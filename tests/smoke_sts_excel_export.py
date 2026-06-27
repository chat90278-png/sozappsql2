import os
import sys
from pathlib import Path
from tempfile import TemporaryDirectory

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from src.models.app_models import ComponentDef, ContractInfo, DeliveryInfo, SystemInfo
from src.services.sts_store import STSStore

try:
    from openpyxl import load_workbook
except Exception:
    print('skip: openpyxl not installed')
    raise SystemExit(0)


def rgb_endswith(cell, suffix):
    value = getattr(cell.fill.fgColor, 'rgb', None) or ''
    return str(value).upper().endswith(suffix)


def rows_by_delivery(ws, delivery_name):
    return [row for row in ws.iter_rows(min_row=2, values_only=True) if row[3] == delivery_name]


with TemporaryDirectory() as td:
    db = Path(td) / 'e.sts'
    s = STSStore(db)
    s.create_platform('AKINCI')
    s.create_platform('TB2')
    s.write_users([{'name':'U1','yi_yd':'Yİ','active':True,'note':''}])
    s.write_components([
        ComponentDef(name='C1', platforms={'AKINCI': True, 'TB2': True}),
        ComponentDef(name='C2', platforms={'AKINCI': True, 'TB2': False}),
    ])

    ci_no_system = ContractInfo(no='K0', platform='AKINCI', user='U1', yi_yd='Yİ', contract_type='Ana Sözleşme', signature_date='', t0_date='', t0_months=0, completion_date='')
    ci_with_delivery = ContractInfo(no='K1', platform='AKINCI', user='U1', yi_yd='Yİ', contract_type='Ana Sözleşme', signature_date='2026-01-01', t0_date='2026-01-02', t0_months=3, completion_date='2026-04-02')
    ci_other_platform = ContractInfo(no='K2', platform='TB2', user='U1', yi_yd='Yİ', contract_type='Ana Sözleşme', signature_date='', t0_date='', t0_months=0, completion_date='')
    ci_system_only = ContractInfo(no='K3', platform='AKINCI', user='U1', yi_yd='Yİ', contract_type='Ana Sözleşme', signature_date='', t0_date='', t0_months=0, completion_date='')

    systems = [SystemInfo(name='S1', components={'C1':3,'C2':2}), SystemInfo(name='S2', components={'C1':1})]
    deliveries = {
        'S1':[DeliveryInfo(name='Teslimat 1', status='Devam Ediyor', acceptance_date='', note='N1', planned={'C1':2,'C2':2}, delivered={'C1':1,'C2':2})],
        'S2':[DeliveryInfo(name='Teslimat 2', status='Tamamlandı', acceptance_date='2026-03-01', note='', planned={'C1':1}, delivered={'C1':1})],
    }

    s.write_contract(ci_no_system, [], {})
    s.write_contract(ci_with_delivery, systems, deliveries)
    s.write_contract(ci_other_platform, systems, deliveries)
    s.write_contract(ci_system_only, [SystemInfo(name='Sistem Teslimat Yok', components={'C1':5,'C2':4})], {})

    progress = []
    out1 = Path(td)/'full.xlsx'
    s.export_to_excel(out1, options={'scope':'all'}, progress_cb=lambda p,m: progress.append((p,m)))
    assert out1.exists() and len(progress) >= 3
    wb1 = load_workbook(out1)
    assert 'Özet' in wb1.sheetnames
    assert 'AKINCI' in wb1.sheetnames and 'TB2' in wb1.sheetnames

    ws = wb1['AKINCI']
    expected_base_headers = [
        'Sözleşme No', 'Sözleşme Türü', 'Sistem Adı', 'Teslimat Adı', 'Kullanıcı', 'Yİ/YD',
        'Durum', 'İmza Tarihi', 'T0 Tarihi', 'T0 Ay', 'Termin Tarihi', 'Gerçek Teslimat',
        'Etiketler', 'Not',
    ]
    header_row = [cell.value for cell in ws[1]]
    assert header_row[:14] == expected_base_headers
    assert header_row[14:] == [
        'C1 Teslim Edilecek', 'C1 Teslim Edilen', 'C1 Kalan',
        'C2 Teslim Edilecek', 'C2 Teslim Edilen', 'C2 Kalan',
    ]

    all_rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert any(row[0] == 'K0' and row[2] == 'GENEL' and row[3] == 'Sözleşme Toplamı' for row in all_rows)
    assert any(row[2] == 'Sistem Teslimat Yok' and row[3] == 'Sistem Toplamı' for row in all_rows)
    assert rows_by_delivery(ws, 'Teslimat 1') and rows_by_delivery(ws, 'Teslimat 2')

    k1_total = next(row for row in all_rows if row[0] == 'K1' and row[3] == 'Sözleşme Toplamı')
    assert k1_total[14:20] == (4, 2, 2, 2, 2, 0)
    k3_system = next(row for row in all_rows if row[2] == 'Sistem Teslimat Yok' and row[3] == 'Sistem Toplamı')
    assert k3_system[14:20] == (5, 0, 5, 4, 0, 4)

    teslimat1_row_index, teslimat1 = next((idx, row) for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2) if row[3] == 'Teslimat 1')
    teslimat2_row_index, teslimat2 = next((idx, row) for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2) if row[3] == 'Teslimat 2')
    assert teslimat1[16] == 1  # C1 kalan = 2 - 1, formülsüz sayı
    assert teslimat1[19] == 0  # C2 kalan = 2 - 2

    assert rgb_endswith(ws['A1'], '0D2B55')
    assert ws['A1'].font.bold is True
    assert str(ws['A1'].font.color.rgb).upper().endswith('FFFFFF')
    assert ws.row_dimensions[1].height == 22
    assert ws.freeze_panes == 'A2'
    assert ws.auto_filter.ref == f'A1:T{ws.max_row}'
    assert rgb_endswith(ws.cell(row=teslimat1_row_index, column=17), 'FFF2CC')
    assert rgb_endswith(ws.cell(row=teslimat1_row_index, column=20), 'FFFFFF')
    assert rgb_endswith(ws.cell(row=teslimat1_row_index, column=7), 'DDEEFF')
    assert rgb_endswith(ws.cell(row=teslimat2_row_index, column=7), 'C6EFCE')
    assert any(str(rng).startswith('A') for rng in ws.merged_cells.ranges)
    assert any(str(rng).startswith('B') for rng in ws.merged_cells.ranges)

    ws_tb2 = wb1['TB2']
    tb2_headers = [cell.value for cell in ws_tb2[1]]
    assert tb2_headers[14:] == ['C1 Teslim Edilecek', 'C1 Teslim Edilen', 'C1 Kalan']

    out2 = Path(td)/'selected.xlsx'
    s.export_to_excel(out2, options={'scope':'selected','platforms':['AKINCI']})
    wb2 = load_workbook(out2, read_only=True)
    assert 'AKINCI' in wb2.sheetnames
    assert 'TB2' not in wb2.sheetnames

    out3 = Path(td)/'summary.xlsx'
    s.export_to_excel(out3, options={'scope':'summary_only'})
    wb3 = load_workbook(out3, read_only=True)
    assert wb3.sheetnames == ['Özet']

    try:
        from PySide6.QtWidgets import QApplication
        from src.ui.dialogs.excel_export_options import ExcelExportDialog
    except Exception:
        print('skip: PySide6 not installed')
    else:
        app = QApplication.instance() or QApplication([])
        dlg_all = ExcelExportDialog(s, active_platform=None, contract_index=[])
        assert {row.platform_name for row in dlg_all.platform_rows if row.is_checked()} == {'AKINCI', 'TB2'}
        dlg_all.accept_options()
        assert dlg_all.result_options == {
            'scope': 'selected',
            'platforms': ['AKINCI', 'TB2'],
            'include_summary': True,
            'include_contract_rows': True,
            'include_system_rows': True,
            'include_delivery_rows': True,
            'include_component_columns': True,
            'include_tags': True,
        }

        dlg_active = ExcelExportDialog(s, active_platform='AKINCI', contract_index=[])
        assert [row.platform_name for row in dlg_active.platform_rows if row.is_checked()] == ['AKINCI']
        dlg_active.clear_platform_selection()
        assert not dlg_active._selected_platforms()
        dlg_active.select_all_platforms()
        assert {row.platform_name for row in dlg_active.platform_rows if row.is_checked()} == {'AKINCI', 'TB2'}

    logs = s.list_logs(limit=200)
    ex = [x for x in logs if x.get('action') == 'excel_exported']
    assert ex
    assert 'options' in (ex[0].get('payload_json') or '')

print('ok')
