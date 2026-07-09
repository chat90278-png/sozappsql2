from __future__ import annotations

import gc
import tempfile
import time
import tracemalloc
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from analysis_center.analysis_custom_dashboard import CustomAnalysisDashboardController
from analysis_center.analysis_dashboard_workspace import DashboardWorkspace
from analysis_center.analysis_excel_export import export_dashboard_excel
from analysis_center.analysis_models import AnalysisCard, AnalysisEntity, CardType, DashboardItem
from analysis_center.analysis_repository import MemoryAnalysisRepository
from analysis_center.analysis_service import AnalysisService


@dataclass(frozen=True)
class Scenario:
    name: str
    card_count: int
    rows_per_card: int
    columns: int


SCENARIOS = (
    Scenario("10_cards_small_tables", 10, 20, 6),
    Scenario("20_cards_small_tables", 20, 20, 6),
    Scenario("1_table_1000_rows", 1, 1_000, 8),
    Scenario("1_table_10000_rows", 1, 10_000, 8),
)


def _card(index: int, rows: int, columns: int) -> AnalysisCard:
    column_ids = [f"field_{col}" for col in range(columns)]
    data = []
    for row in range(rows):
        data.append(
            {
                column: (
                    row * (col + 1)
                    if col % 3 == 0
                    else f"Kart {index} / satır {row} / alan {col}"
                )
                for col, column in enumerate(column_ids)
            }
        )
    return AnalysisCard(
        f"perf-card-{index}",
        f"Performans Kartı {index + 1}",
        AnalysisEntity.CONTRACT,
        CardType.TABLE,
        columns=column_ids,
        data=data,
        screen_id="perf_prepared",
    )


def _measure(root: Path, scenario: Scenario) -> tuple[float, float, int, int]:
    service = AnalysisService(use_sample=True, repository=MemoryAnalysisRepository())
    service.refresh_data()
    controller = CustomAnalysisDashboardController(service)
    cards = [_card(index, scenario.rows_per_card, scenario.columns) for index in range(scenario.card_count)]
    workspace = DashboardWorkspace("perf")
    for card in cards:
        assert workspace.pin(card)
    item = DashboardItem("perf_prepared", "Performance", cards=cards)
    output = root / f"{scenario.name}.xlsx"

    gc.collect()
    tracemalloc.start()
    started = time.perf_counter()
    result = export_dashboard_excel(
        output,
        workspace=workspace,
        dashboard_items=[item],
        custom_controller=controller,
    )
    elapsed = time.perf_counter() - started
    _current, peak = tracemalloc.get_traced_memory()
    tracemalloc.stop()

    workbook = load_workbook(output, read_only=False)
    assert len(workbook.sheetnames) == scenario.card_count + 1
    return elapsed, peak / (1024 * 1024), output.stat().st_size, result.sheet_count


def main() -> None:
    with tempfile.TemporaryDirectory(prefix="tur21-excel-performance-") as temp_dir:
        root = Path(temp_dir)
        print("measurement=wall_clock_export_only")
        print("memory=tracemalloc_python_allocation_peak")
        for scenario in SCENARIOS:
            elapsed, peak_mb, file_size, sheet_count = _measure(root, scenario)
            print(
                f"scenario={scenario.name},cards={scenario.card_count},"
                f"rows_per_card={scenario.rows_per_card},columns={scenario.columns},"
                f"elapsed_seconds={elapsed:.4f},python_peak_mb={peak_mb:.2f},"
                f"file_size_bytes={file_size},sheet_count={sheet_count}"
            )
    print("TUR 21 DASHBOARD EXCEL PERFORMANCE: PASS")


if __name__ == "__main__":
    main()
