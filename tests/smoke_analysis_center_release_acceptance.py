from __future__ import annotations

import ast
import importlib
import json
import os
import subprocess
import sys
import tempfile
from datetime import datetime, timedelta, timezone
from pathlib import Path

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

from openpyxl import load_workbook

from analysis_center.analysis_custom_dashboard import (
    CUSTOM_ANALYSIS_DASHBOARD_SOURCE_ID,
    CustomAnalysisDashboardController,
)
from analysis_center.analysis_dashboard_workspace import (
    DashboardWorkspace,
    DashboardWorkspaceStore,
    source_workspace_key,
)
from analysis_center.analysis_definitions import AnalysisDefinition, MeasureDefinition
from analysis_center.analysis_excel_export import export_dashboard_excel
from analysis_center.analysis_models import AnalysisCard, AnalysisEntity, CardType, DashboardItem
from analysis_center.analysis_repository import (
    ANALYSIS_REPOSITORY_SCHEMA_VERSION,
    AnalysisRepositoryCorruptError,
    FileAnalysisRepository,
)
from analysis_center.analysis_service import AnalysisService


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "STS-S-VR-S-NEK---TBD---1__share-edit__2026-07-07_14-04.sts"
SMOKES = (
    "smoke_dashboard_hardening_tur12.py",
    "smoke_custom_analysis_builder_tur13.py",
    "smoke_custom_analysis_persistence_tur14.py",
    "smoke_custom_dashboard_integration_tur15.py",
    "smoke_visual_settings_tur16.py",
    "smoke_builder_ux_cleanup_tur17.py",
    "smoke_prepared_analysis_template_tur18.py",
    "smoke_dashboard_quick_edit_tur19.py",
    "smoke_dashboard_excel_export_tur20.py",
)


def _run_smoke(script_name: str) -> None:
    script = ROOT / "tests" / script_name
    env = dict(os.environ)
    env["PYTHONPATH"] = str(ROOT)
    env["QT_QPA_PLATFORM"] = "offscreen"
    completed = subprocess.run(
        [sys.executable, str(script)],
        cwd=ROOT,
        env=env,
        text=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        check=False,
    )
    print(f"--- {script_name} ---")
    print(completed.stdout.rstrip())
    if completed.returncode != 0:
        raise RuntimeError(f"Acceptance smoke failed: {script_name} rc={completed.returncode}")


def _analysis_definition() -> AnalysisDefinition:
    return AnalysisDefinition(
        analysis_id="preview-release",
        title="Release Acceptance Custom KPI",
        dataset="acceptances",
        visualization="kpi",
        measures=[MeasureDefinition("", "count_rows")],
    )


def _persistence_acceptance(root: Path) -> tuple[str, Path, Path, int]:
    source_a = root / "source-a.sts"
    source_b = root / "source-b.sts"
    source_a.write_bytes(b"A")
    source_b.write_bytes(b"B")
    analysis_root = root / "analysis-repository"
    workspace_root = root / "workspace-repository"

    repository_a = FileAnalysisRepository(source_a, analysis_root)
    repository_b = FileAnalysisRepository(source_b, analysis_root)
    service_a = AnalysisService(use_sample=True, repository=repository_a)
    service_a.refresh_data()
    saved = service_a.create_saved_analysis(_analysis_definition())
    assert saved.analysis_id.startswith("custom-")
    assert repository_b.list_analyses() == []
    assert repository_a.repository_path() != repository_b.repository_path()

    custom_controller = CustomAnalysisDashboardController(service_a)
    workspace_a = DashboardWorkspace(source_key=source_workspace_key(source_a))
    assert custom_controller.pin(workspace_a, saved.analysis_id)
    workspace_b = DashboardWorkspace(source_key=source_workspace_key(source_b))
    store = DashboardWorkspaceStore(workspace_root)
    store.save(source_a, workspace_a)
    store.save(source_b, workspace_b)
    assert store.workspace_path(source_a) != store.workspace_path(source_b)

    repository_text = repository_a.repository_path().read_text(encoding="utf-8")
    workspace_text = store.workspace_path(source_a).read_text(encoding="utf-8")
    assert "preview-" not in repository_text
    assert "preview-" not in workspace_text
    assert CUSTOM_ANALYSIS_DASHBOARD_SOURCE_ID in workspace_text
    assert saved.analysis_id in workspace_text

    payload = json.loads(repository_text)
    payload["analyses"].append({"analysis_id": "broken-entry"})
    repository_a.repository_path().write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    isolated = FileAnalysisRepository(source_a, analysis_root)
    assert [item.analysis_id for item in isolated.list_analyses()] == [saved.analysis_id]
    assert len(isolated.load_issues) == 1

    corrupt_source = root / "source-corrupt.sts"
    corrupt_source.write_bytes(b"C")
    corrupt_repository = FileAnalysisRepository(corrupt_source, analysis_root)
    corrupt_repository.repository_path().parent.mkdir(parents=True, exist_ok=True)
    corrupt_repository.repository_path().write_text("{", encoding="utf-8")
    corrupt_repository = FileAnalysisRepository(corrupt_source, analysis_root)
    assert isinstance(corrupt_repository.load_error, AnalysisRepositoryCorruptError)
    before = corrupt_repository.repository_path().read_text(encoding="utf-8")
    try:
        corrupt_repository.save_analysis(_analysis_definition())
    except AnalysisRepositoryCorruptError:
        pass
    else:
        raise AssertionError("Corrupt repository save unexpectedly succeeded")
    assert corrupt_repository.repository_path().read_text(encoding="utf-8") == before

    print(f"source_a_repository={repository_a.repository_path()}")
    print(f"source_b_repository={repository_b.repository_path()}")
    print(f"saved_custom_id={saved.analysis_id}")
    print(f"workspace_custom_placement_count={len(workspace_a.placements)}")
    print(f"repository_schema_version={ANALYSIS_REPOSITORY_SCHEMA_VERSION}")
    print(f"invalid_entry_issue_count={len(isolated.load_issues)}")
    print("corrupt_repository_overwrite_protection=PASS")
    return saved.analysis_id, store.workspace_path(source_a), repository_a.repository_path(), len(workspace_a.placements)


def _excel_edge_acceptance(root: Path) -> tuple[Path, int]:
    service = AnalysisService(use_sample=True)
    service.refresh_data()
    controller = CustomAnalysisDashboardController(service)
    aware = datetime(2026, 7, 9, 10, 30, tzinfo=timezone(timedelta(hours=3)))
    card = AnalysisCard(
        "release-edge-table",
        "Release\x0b Edge",
        AnalysisEntity.CONTRACT,
        CardType.TABLE,
        columns=["name", "metric", "when", "note"],
        data=[
            {
                "name": "bad\x0btext",
                "metric": float("nan"),
                "when": aware,
                "note": "X" * 40_000,
            }
        ],
        screen_id="release_prepared",
    )
    workspace = DashboardWorkspace("release")
    assert workspace.pin(card)
    item = DashboardItem("release_prepared", "Release Prepared", cards=[card])
    output_path = root / "release-edge.xlsx"
    result = export_dashboard_excel(
        output_path,
        workspace=workspace,
        dashboard_items=[item],
        custom_controller=controller,
        source=SOURCE,
        exported_at=datetime(2026, 7, 9, 9, 0),
    )
    workbook = load_workbook(output_path)
    assert workbook.sheetnames[0] == "Dashboard Özeti"
    detail = workbook[workbook.sheetnames[1]]
    assert detail["A4"].value == "badtext"
    assert detail["B4"].value is None
    assert detail["C4"].value == datetime(2026, 7, 9, 7, 30)
    assert len(detail["D4"].value) == 32_767
    print(f"release_edge_export_path={output_path}")
    print(f"release_edge_workbook_sheet_count={len(workbook.sheetnames)}")
    print("excel_illegal_xml_nonfinite_timezone_longtext=PASS")
    return output_path, len(workbook.sheetnames)


def _runtime_acceptance() -> None:
    for module_name in (
        "app",
        "analysis_center",
        "analysis_center.analysis_window",
        "analysis_center.analysis_excel_export",
    ):
        importlib.import_module(module_name)
        print(f"import_{module_name}=PASS")
    for spec_name in ("STS.spec", "SozAppSQL.spec"):
        text = (ROOT / spec_name).read_text(encoding="utf-8")
        ast.parse(text)
        assert 'collect_submodules("openpyxl")' in text
        print(f"spec_{spec_name}=openpyxl_hiddenimports:PASS")


def main() -> None:
    assert SOURCE.exists(), SOURCE
    print(f"release_source={SOURCE}")
    for script_name in SMOKES:
        _run_smoke(script_name)

    with tempfile.TemporaryDirectory(prefix="tur21-release-acceptance-") as temp_dir:
        root = Path(temp_dir)
        saved_id, workspace_path, repository_path, placement_count = _persistence_acceptance(root)
        export_path, sheet_count = _excel_edge_acceptance(root)
        _runtime_acceptance()
        print(f"acceptance_saved_custom_id={saved_id}")
        print(f"acceptance_workspace_path={workspace_path}")
        print(f"acceptance_repository_path={repository_path}")
        print(f"acceptance_placement_count={placement_count}")
        print(f"acceptance_export_path={export_path}")
        print(f"acceptance_workbook_sheet_count={sheet_count}")

    print(f"acceptance_smoke_count={len(SMOKES)}")
    print("TUR 21 ANALYSIS CENTER RELEASE ACCEPTANCE: PASS")


if __name__ == "__main__":
    main()
