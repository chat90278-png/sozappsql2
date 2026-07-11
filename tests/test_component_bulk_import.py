from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from src.domain.component_bulk_import import (
    BulkComponentRow,
    WorkbookMapping,
    component_key,
    detect_workbook_mapping,
    list_workbook_sheets,
    load_workbook_matrix,
    merge_new_components,
    normalize_active,
    normalize_space,
    normalize_unit,
    rows_from_matrix,
    validate_bulk_rows,
)


def test_normalizers_cover_turkish_values():
    assert normalize_space("  Hava   Aracı  ") == "Hava Aracı"
    assert component_key(" YKİ ") == component_key("yki̇")
    assert normalize_unit("takim") == "Takım"
    assert normalize_unit("custom unit") == "custom unit"
    assert normalize_active("Aktif") is True
    assert normalize_active("hayır") is False


def test_detects_detailed_header_and_parses_rows():
    matrix = [
        ["Bileşen Adı", "Birim", "Not", "Durum"],
        ["Hava Aracı", "Adet", "Ana araç", "Aktif"],
        ["Yer Destek", "Set", None, "Pasif"],
    ]
    mapping = detect_workbook_mapping(matrix)
    assert mapping == WorkbookMapping(0, 1, 2, 3, 0)

    rows = rows_from_matrix(matrix, mapping=mapping)
    assert rows == [
        BulkComponentRow("Hava Aracı", "Adet", "Ana araç", True, 2),
        BulkComponentRow("Yer Destek", "Set", "", False, 3),
    ]


def test_headerless_single_column_is_supported():
    matrix = [["Hava Aracı"], ["YKİ"], ["YVT"]]
    mapping = detect_workbook_mapping(matrix)
    assert mapping.name_column == 0
    assert mapping.header_row is None
    assert [row.name for row in rows_from_matrix(matrix, mapping=mapping)] == [
        "Hava Aracı",
        "YKİ",
        "YVT",
    ]


def test_duplicates_and_existing_rows_are_classified():
    rows = [
        BulkComponentRow("Motor"),
        BulkComponentRow(" motor "),
        BulkComponentRow("YKİ"),
        BulkComponentRow(""),
        BulkComponentRow("Yeni"),
    ]
    validated = validate_bulk_rows(rows, [{"name": "YKİ"}])
    assert [item.state for item in validated] == [
        "duplicate",
        "duplicate",
        "existing",
        "blank",
        "ready",
    ]


def test_merge_appends_ready_rows_and_preserves_existing_payload():
    existing = [
        {
            "id": 7,
            "name": "Hava Aracı",
            "unit": "Adet",
            "active": True,
            "note": "Eski not",
            "display_order": 4,
            "platforms": {"AKINCI": True},
        }
    ]
    rows = [
        BulkComponentRow("Hava Aracı"),
        BulkComponentRow("Motor", "Takım", "Yeni", False),
    ]
    result = merge_new_components(existing, rows)

    assert result.added == 1
    assert result.skipped_existing == 1
    assert len(result.components) == 2
    assert result.components[0]["platforms"] == {"AKINCI": True}
    assert result.components[1] == {
        "id": None,
        "name": "Motor",
        "version": "",
        "unit": "Takım",
        "active": False,
        "usage": 1,
        "note": "Yeni",
        "display_order": 5,
        "platforms": {},
    }


def test_workbook_helpers_read_selected_sheet(tmp_path: Path):
    path = tmp_path / "components.xlsx"
    workbook = Workbook()
    first = workbook.active
    first.title = "Liste"
    first.append(["Bileşen Adı"])
    first.append(["Motor"])
    second = workbook.create_sheet("Detay")
    second.append(["Bileşen Adı", "Birim", "Durum"])
    second.append(["YKİ", "Set", "Pasif"])
    workbook.save(path)
    workbook.close()

    assert list_workbook_sheets(path) == ["Liste", "Detay"]
    matrix = load_workbook_matrix(path, "Detay")
    rows = rows_from_matrix(matrix)
    assert rows == [BulkComponentRow("YKİ", "Set", "", False, 2)]
