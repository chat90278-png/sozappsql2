from __future__ import annotations

from copy import deepcopy
from datetime import date, datetime
from pathlib import Path

import pytest
from openpyxl import load_workbook
from openpyxl.chart import BarChart, DoughnutChart, LineChart

from analysis_center.analysis_custom_dashboard import (
    CUSTOM_ANALYSIS_DASHBOARD_SOURCE_ID,
    CustomAnalysisDashboardController,
)
from analysis_center.analysis_dashboard_layout import DashboardCardPlacement
from analysis_center.analysis_dashboard_workspace import DashboardWorkspace
from analysis_center.analysis_definitions import AnalysisDefinition, MeasureDefinition
from analysis_center.analysis_excel_export import (
    DashboardExcelExportError,
    DashboardExportCollector,
    export_dashboard_excel,
    safe_unique_sheet_name,
)
from analysis_center.analysis_models import (
    AnalysisCard,
    AnalysisEntity,
    CardType,
    ChartType,
    DashboardItem,
)
from analysis_center.analysis_repository import MemoryAnalysisRepository
from analysis_center.analysis_service import AnalysisService
from analysis_center.analysis_visual_settings import AnalysisVisualSettings


def _service() -> AnalysisService:
    service = AnalysisService(use_sample=True, repository=MemoryAnalysisRepository())
    service.refresh_data()
    return service


def _custom_definition(analysis_id: str, visualization: str = "horizontal_bar", title: str = "Özel"):
    if visualization == "kpi":
        return AnalysisDefinition(
            analysis_id=analysis_id,
            title=title,
            dataset="acceptances",
            visualization="kpi",
            measures=[MeasureDefinition("", "count_rows")],
        )
    if visualization == "table":
        return AnalysisDefinition(
            analysis_id=analysis_id,
            title=title,
            dataset="acceptances",
            visualization="table",
            select_fields=["platform", "name", "planned_acceptance_date", "planned_total"],
            limit=20,
        )
    return AnalysisDefinition(
        analysis_id=analysis_id,
        title=title,
        dataset="acceptances",
        visualization=visualization,
        dimensions=["platform"],
        measures=[MeasureDefinition("", "count_rows")],
        limit=20,
    )


def _prepared_cards() -> list[AnalysisCard]:
    visual = AnalysisVisualSettings.defaults()
    visual.replace_kpi(prefix="#", suffix=" adet", decimal_places=0, subtitle="Alt Başlık")
    return [
        AnalysisCard(
            "prepared-kpi",
            "Toplam",
            AnalysisEntity.CONTRACT,
            CardType.KPI,
            value=3,
            subtitle="Legacy subtitle",
            screen_id="prepared",
            meta={"visual_settings_enabled": True, "visual_settings": visual, "dataset": "contracts"},
        ),
        AnalysisCard(
            "prepared-chart",
            "Dağılım",
            AnalysisEntity.CONTRACT,
            CardType.CHART,
            chart_type=ChartType.DONUT,
            data=[{"label": "A", "value": 2}, {"label": "B", "value": 1}],
            screen_id="prepared",
        ),
        AnalysisCard(
            "prepared-table",
            "Kayıtlar",
            AnalysisEntity.ACCEPTANCE,
            CardType.TABLE,
            columns=["platform", "planned_acceptance_date", "planned_total", "name"],
            data=[
                {
                    "platform": "AKINCI",
                    "planned_acceptance_date": "2026-07-09",
                    "planned_total": 5.5,
                    "name": None,
                }
            ],
            screen_id="prepared",
            meta={"dataset": "acceptances"},
        ),
    ]


def _workspace_with_prepared(cards: list[AnalysisCard]) -> tuple[DashboardWorkspace, DashboardItem]:
    workspace = DashboardWorkspace("test")
    item = DashboardItem("prepared", "Hazır", cards=cards)
    for card in cards:
        assert workspace.pin(card)
    return workspace, item


def test_collector_exports_only_pinned_prepared_and_custom_in_placement_order():
    service = _service()
    first = service.create_saved_analysis(_custom_definition("preview-a", title="A"))
    second = service.create_saved_analysis(_custom_definition("preview-b", title="B"))
    controller = CustomAnalysisDashboardController(service)
    cards = _prepared_cards()[:1]
    workspace, item = _workspace_with_prepared(cards)
    assert controller.pin(workspace, second.analysis_id)

    executed: list[str] = []
    real_execute = service.execute_analysis

    def spy(definition):
        executed.append(definition.analysis_id)
        return real_execute(definition)

    service.execute_analysis = spy
    collection = DashboardExportCollector(controller).collect(workspace, [item])

    assert [export_item.card.card_id for export_item in collection.items] == [
        "prepared-kpi",
        second.analysis_id,
    ]
    assert [export_item.source_kind for export_item in collection.items] == ["prepared", "custom"]
    assert executed == [second.analysis_id]
    assert first.analysis_id not in executed


def test_collector_unresolved_custom_becomes_issue_without_dropping_prepared():
    service = _service()
    controller = CustomAnalysisDashboardController(service)
    prepared = _prepared_cards()[0]
    workspace, item = _workspace_with_prepared([prepared])
    workspace.add_placement(
        DashboardCardPlacement(
            "missing-custom",
            CUSTOM_ANALYSIS_DASHBOARD_SOURCE_ID,
            "custom-missing",
            3,
            0,
            3,
            2,
        )
    )

    collection = DashboardExportCollector(controller).collect(workspace, [item])

    assert len(collection.items) == 1
    assert collection.items[0].card.card_id == prepared.card_id
    assert len(collection.issues) == 1
    assert collection.issues[0].source_kind == "custom"
    assert "bulunamadı" in collection.issues[0].message
    assert workspace.contains(CUSTOM_ANALYSIS_DASHBOARD_SOURCE_ID, "custom-missing")


def test_empty_dashboard_raises_controlled_error(tmp_path):
    service = _service()
    with pytest.raises(DashboardExcelExportError, match="aktarılabilir kart"):
        export_dashboard_excel(
            tmp_path / "empty.xlsx",
            workspace=DashboardWorkspace("test"),
            dashboard_items=[],
            custom_controller=CustomAnalysisDashboardController(service),
        )


def test_workbook_summary_detail_order_hyperlinks_and_unique_sheet_names(tmp_path):
    service = _service()
    controller = CustomAnalysisDashboardController(service)
    cards = _prepared_cards()
    cards[0].title = "Platform/Dağılımı"
    cards[1].title = "Platform:Dağılımı"
    workspace, item = _workspace_with_prepared(cards)

    result = export_dashboard_excel(
        tmp_path / "dashboard",
        workspace=workspace,
        dashboard_items=[item],
        custom_controller=controller,
        source=tmp_path / "Kaynak.sts",
        exported_at=datetime(2026, 7, 9, 8, 30),
    )
    workbook = load_workbook(result.output_path)

    assert result.output_path.suffix == ".xlsx"
    assert workbook.sheetnames[0] == "Dashboard Özeti"
    assert workbook.sheetnames[1:] == ["Platform Dağılımı", "Platform Dağılımı (2)", "Kayıtlar"]
    summary = workbook["Dashboard Özeti"]
    assert summary["B3"].value == "Kaynak.sts"
    assert summary["B4"].value == datetime(2026, 7, 9, 8, 30)
    assert summary["F10"].hyperlink.target == "#'Platform Dağılımı'!A1"
    assert summary["F11"].hyperlink.target == "#'Platform Dağılımı (2)'!A1"
    assert result.exported_card_count == 3
    assert result.sheet_count == 4


def test_safe_sheet_name_is_case_insensitive_unique_and_31_chars():
    used: set[str] = set()
    first = safe_unique_sheet_name("A/B:C?D*E[F]", used)
    second = safe_unique_sheet_name(first.lower(), used)
    long_name = safe_unique_sheet_name("X" * 100, used)
    assert first == "A B C D E F"
    assert second.endswith("(2)")
    assert len(long_name) == 31


def test_kpi_export_preserves_numeric_value_and_formatted_display(tmp_path):
    service = _service()
    cards = _prepared_cards()[:1]
    workspace, item = _workspace_with_prepared(cards)
    result = export_dashboard_excel(
        tmp_path / "kpi.xlsx",
        workspace=workspace,
        dashboard_items=[item],
        custom_controller=CustomAnalysisDashboardController(service),
    )
    sheet = load_workbook(result.output_path)["Toplam"]
    assert sheet["A5"].value == 3
    assert sheet["A5"].number_format == '"#"#,##0" adet"'
    assert sheet["B10"].value == "#3 adet"
    assert sheet["A3"].value == "Legacy subtitle"


@pytest.mark.parametrize(
    ("chart_type", "chart_class"),
    [
        (ChartType.BAR, BarChart),
        (ChartType.HORIZONTAL_BAR, BarChart),
        (ChartType.DONUT, DoughnutChart),
        (ChartType.LINE, LineChart),
    ],
)
def test_chart_export_creates_native_excel_chart_and_preserves_source_data(tmp_path, chart_type, chart_class):
    service = _service()
    settings = AnalysisVisualSettings.defaults()
    settings.replace_chart(show_legend=False, legend_position="bottom", show_values=True, palette="pastel")
    card = AnalysisCard(
        "chart",
        f"Chart {chart_type.value}",
        AnalysisEntity.CONTRACT,
        CardType.CHART,
        chart_type=chart_type,
        data=[{"label": "A", "value": 5}, {"label": "B", "value": 3}],
        screen_id="prepared",
        meta={"visual_settings_enabled": True, "visual_settings": settings},
    )
    before = deepcopy(card.data)
    workspace, item = _workspace_with_prepared([card])
    result = export_dashboard_excel(
        tmp_path / f"{chart_type.value}.xlsx",
        workspace=workspace,
        dashboard_items=[item],
        custom_controller=CustomAnalysisDashboardController(service),
    )
    sheet = load_workbook(result.output_path)[card.title]
    assert len(sheet._charts) == 1
    native_chart = sheet._charts[0]
    assert isinstance(native_chart, chart_class)
    assert native_chart.legend is None
    assert native_chart.dLbls is not None and native_chart.dLbls.showVal is True
    assert sheet["A5"].value == "A"
    assert sheet["B5"].value == 5
    assert card.data == before


def test_chart_legend_bottom_and_palette_state_are_written(tmp_path):
    service = _service()
    settings = AnalysisVisualSettings.defaults()
    settings.replace_chart(show_legend=True, legend_position="bottom", palette="pastel")
    card = AnalysisCard(
        "donut",
        "Donut",
        AnalysisEntity.CONTRACT,
        CardType.CHART,
        chart_type=ChartType.DONUT,
        data=[{"label": "A", "value": 2}, {"label": "B", "value": 1}],
        screen_id="prepared",
        meta={"visual_settings_enabled": True, "visual_settings": settings},
    )
    workspace, item = _workspace_with_prepared([card])
    result = export_dashboard_excel(
        tmp_path / "donut.xlsx",
        workspace=workspace,
        dashboard_items=[item],
        custom_controller=CustomAnalysisDashboardController(service),
    )
    chart = load_workbook(result.output_path)["Donut"]._charts[0]
    assert chart.legend.position == "b"
    assert len(chart.series[0].data_points) == 2
    assert chart.series[0].data_points[0].graphicalProperties.solidFill.srgbClr.casefold().lstrip("#") == "93c5fd"


def test_table_export_preserves_column_order_types_dates_blanks_and_excel_table(tmp_path):
    service = _service()
    card = _prepared_cards()[2]
    workspace, item = _workspace_with_prepared([card])
    result = export_dashboard_excel(
        tmp_path / "table.xlsx",
        workspace=workspace,
        dashboard_items=[item],
        custom_controller=CustomAnalysisDashboardController(service),
    )
    sheet = load_workbook(result.output_path)["Kayıtlar"]
    assert [sheet.cell(3, index).value for index in range(1, 5)] == [
        "Platform",
        "Planlanan Kabul Tarihi",
        "Planlanan Toplam",
        "Teslimat Adı",
    ]
    assert sheet["A4"].value == "AKINCI"
    assert sheet["B4"].value.date() == date(2026, 7, 9)
    assert sheet["C4"].value == 5.5
    assert sheet["D4"].value is None
    assert len(sheet.tables) == 1
    assert sheet.freeze_panes == "A4"


def test_repository_load_error_is_export_warning_and_custom_placement_preserved(tmp_path):
    service = _service()
    controller = CustomAnalysisDashboardController(service)
    prepared = _prepared_cards()[0]
    workspace, item = _workspace_with_prepared([prepared])
    workspace.add_placement(
        DashboardCardPlacement(
            "custom-placement",
            CUSTOM_ANALYSIS_DASHBOARD_SOURCE_ID,
            "custom-safe",
            3,
            0,
            3,
            2,
        )
    )
    service.repository_load_error = lambda: RuntimeError("corrupt")
    result = export_dashboard_excel(
        tmp_path / "warning.xlsx",
        workspace=workspace,
        dashboard_items=[item],
        custom_controller=controller,
    )
    workbook = load_workbook(result.output_path)
    assert result.exported_card_count == 1
    assert result.warning_count == 1
    assert any(cell.value == "Aktarım Uyarıları" for row in workbook["Dashboard Özeti"] for cell in row)
    assert workspace.contains(CUSTOM_ANALYSIS_DASHBOARD_SOURCE_ID, "custom-safe")


def test_output_save_error_is_controlled(monkeypatch, tmp_path):
    service = _service()
    workspace, item = _workspace_with_prepared(_prepared_cards()[:1])

    def fail_save(self, filename):
        raise OSError("disk")

    monkeypatch.setattr("analysis_center.analysis_excel_export.Workbook.save", fail_save)
    with pytest.raises(DashboardExcelExportError, match="kaydedilemedi"):
        export_dashboard_excel(
            tmp_path / "fail.xlsx",
            workspace=workspace,
            dashboard_items=[item],
            custom_controller=CustomAnalysisDashboardController(service),
        )

@pytest.mark.parametrize(
    ("visualization", "expected_kind"),
    [("kpi", "kpi"), ("horizontal_bar", "chart"), ("table", "table")],
)
def test_saved_custom_kpi_chart_table_export_through_real_engine(tmp_path, visualization, expected_kind):
    service = _service()
    saved = service.create_saved_analysis(_custom_definition("preview", visualization=visualization, title=f"Custom {visualization}"))
    controller = CustomAnalysisDashboardController(service)
    workspace = DashboardWorkspace("test")
    assert controller.pin(workspace, saved.analysis_id)

    result = export_dashboard_excel(
        tmp_path / f"custom-{visualization}.xlsx",
        workspace=workspace,
        dashboard_items=[],
        custom_controller=controller,
    )
    workbook = load_workbook(result.output_path)
    sheet = workbook[workbook.sheetnames[1]]
    assert result.exported_card_count == 1
    if expected_kind == "kpi":
        assert sheet["A5"].value is not None
    elif expected_kind == "chart":
        assert len(sheet._charts) == 1
    else:
        assert len(sheet.tables) == 1


def test_legacy_donut_uses_right_legend_and_long_table_width_is_bounded(tmp_path):
    service = _service()
    cards = _prepared_cards()[1:]
    cards[1].data[0]["name"] = "X" * 300
    workspace, item = _workspace_with_prepared(cards)
    result = export_dashboard_excel(
        tmp_path / "legacy.xlsx",
        workspace=workspace,
        dashboard_items=[item],
        custom_controller=CustomAnalysisDashboardController(service),
    )
    workbook = load_workbook(result.output_path)
    donut = workbook["Dağılım"]._charts[0]
    assert donut.legend.position == "r"
    assert workbook["Kayıtlar"].column_dimensions["D"].width <= 40


def test_legacy_kpi_unit_is_in_summary_and_native_number_format(tmp_path):
    service = _service()
    card = AnalysisCard(
        "unit-kpi",
        "Bütçe",
        AnalysisEntity.CONTRACT,
        CardType.KPI,
        value=42,
        unit="adet",
        screen_id="prepared",
    )
    workspace, item = _workspace_with_prepared([card])
    result = export_dashboard_excel(
        tmp_path / "unit.xlsx",
        workspace=workspace,
        dashboard_items=[item],
        custom_controller=CustomAnalysisDashboardController(service),
    )
    workbook = load_workbook(result.output_path)
    assert workbook["Dashboard Özeti"]["E10"].value == "42 adet"
    sheet = workbook["Bütçe"]
    assert sheet["A5"].value == 42
    assert sheet["A5"].number_format == '#,##0" adet"'
    assert sheet["B10"].value == "42 adet"


def test_export_sanitizes_illegal_xml_text_and_truncates_long_cells(tmp_path):
    service = _service()
    card = AnalysisCard(
        "unsafe-text",
        "Bad\x0bTitle",
        AnalysisEntity.CONTRACT,
        CardType.TABLE,
        columns=["name", "note"],
        data=[{"name": "bad\x0btext", "note": "X" * 40_000}],
        screen_id="prepared",
    )
    workspace, item = _workspace_with_prepared([card])

    result = export_dashboard_excel(
        tmp_path / "unsafe.xlsx",
        workspace=workspace,
        dashboard_items=[item],
        custom_controller=CustomAnalysisDashboardController(service),
    )
    workbook = load_workbook(result.output_path)
    assert workbook.sheetnames[1] == "BadTitle"
    sheet = workbook["BadTitle"]
    assert sheet["A4"].value == "badtext"
    assert len(sheet["B4"].value) == 32_767
    assert "\x0b" not in workbook["Dashboard Özeti"]["B10"].value


def test_export_normalizes_nonfinite_numbers_and_timezone_aware_datetimes(tmp_path):
    from datetime import timezone, timedelta

    service = _service()
    aware = datetime(2026, 7, 9, 10, 30, tzinfo=timezone(timedelta(hours=3)))
    card = AnalysisCard(
        "edge-values",
        "Edge Values",
        AnalysisEntity.CONTRACT,
        CardType.TABLE,
        columns=["metric", "when"],
        data=[
            {"metric": float("nan"), "when": aware},
            {"metric": float("inf"), "when": aware},
            {"metric": float("-inf"), "when": aware},
        ],
        screen_id="prepared",
    )
    workspace, item = _workspace_with_prepared([card])

    result = export_dashboard_excel(
        tmp_path / "edge-values.xlsx",
        workspace=workspace,
        dashboard_items=[item],
        custom_controller=CustomAnalysisDashboardController(service),
    )
    sheet = load_workbook(result.output_path)["Edge Values"]
    assert [sheet.cell(row, 1).value for row in range(4, 7)] == [None, None, None]
    assert sheet["B4"].value == datetime(2026, 7, 9, 7, 30)
    assert sheet["B4"].value.tzinfo is None


def test_empty_chart_data_exports_without_invalid_chart_reference(tmp_path):
    service = _service()
    card = AnalysisCard(
        "empty-chart",
        "Boş Grafik",
        AnalysisEntity.CONTRACT,
        CardType.CHART,
        chart_type=ChartType.BAR,
        data=[],
        screen_id="prepared",
    )
    workspace, item = _workspace_with_prepared([card])

    result = export_dashboard_excel(
        tmp_path / "empty-chart.xlsx",
        workspace=workspace,
        dashboard_items=[item],
        custom_controller=CustomAnalysisDashboardController(service),
    )
    sheet = load_workbook(result.output_path)["Boş Grafik"]
    assert len(sheet._charts) == 0
    assert sheet["A4"].value == "Kategori"
    assert sheet["B4"].value == "Değer"
