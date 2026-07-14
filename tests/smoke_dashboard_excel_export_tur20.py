from __future__ import annotations

import os
import tempfile
from datetime import date, datetime
from pathlib import Path
from unittest.mock import patch

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

from openpyxl import load_workbook
from PySide6.QtWidgets import QApplication

from analysis_center.analysis_dashboard_layout import placement_order
from analysis_center.analysis_dashboard_workspace import DashboardWorkspaceStore
from analysis_center.analysis_definitions import AnalysisDefinition, MeasureDefinition
from analysis_center.analysis_excel_export import DashboardExportCollector, export_dashboard_excel
from analysis_center.analysis_models import CardType, VisualSettings
from analysis_center.analysis_qt_window import AnalysisCenterWindow
from analysis_center.analysis_repository import FileAnalysisRepository
from analysis_center.analysis_visual_settings import AnalysisVisualSettings


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "STS-S-VR-S-NEK---TBD---1__share-edit__2026-07-07_14-04.sts"


class FrozenSmokeDate(date):
    @classmethod
    def today(cls):
        return cls(2026, 7, 1)


def _settings() -> VisualSettings:
    return VisualSettings(show_disabled_sections=False, empty_state_uses_sample=False)


def _custom_definitions():
    kpi_visual = AnalysisVisualSettings.defaults()
    kpi_visual.replace_kpi(subtitle="Aktif teslimat kapsamı", prefix="Toplam: ", suffix=" kayıt", decimal_places=0)
    chart_visual = AnalysisVisualSettings.defaults()
    chart_visual.replace_chart(
        show_legend=True,
        legend_position="bottom",
        show_values=True,
        palette="pastel",
        max_categories=1,
        group_others=True,
    )
    table_fields = ["platform", "name", "planned_total", "planned_acceptance_date"]
    table_visual = AnalysisVisualSettings.defaults(selected_table_fields=table_fields)
    table_visual.replace_table(column_order=("name", "platform", "planned_acceptance_date", "planned_total"))
    return [
        AnalysisDefinition(
            analysis_id="preview-kpi",
            title="Özel Teslimat Sayısı",
            dataset="acceptances",
            visualization="kpi",
            measures=[MeasureDefinition("", "count_rows")],
            options=kpi_visual.to_options(),
        ),
        AnalysisDefinition(
            analysis_id="preview-chart",
            title="Özel Teslimat Donut",
            dataset="acceptances",
            visualization="donut",
            dimensions=["name"],
            measures=[MeasureDefinition("", "count_rows")],
            limit=20,
            options=chart_visual.to_options(),
        ),
        AnalysisDefinition(
            analysis_id="preview-table",
            title="Özel Teslimat Tablosu",
            dataset="acceptances",
            visualization="table",
            select_fields=table_fields,
            limit=20,
            options=table_visual.to_options(),
        ),
    ]


def _pin_prepared_types(window: AnalysisCenterWindow) -> list[str]:
    selected: list[str] = []
    for target in (CardType.KPI, CardType.CHART, CardType.TABLE):
        card = next(
            card
            for item in window._dashboard_items
            for card in item.cards
            if card.enabled and card.card_type == target
        )
        window._toggle_dashboard_card(card)
        selected.append(card.card_id)
    return selected


@patch("analysis_center.analysis_metrics.date", FrozenSmokeDate)
def main() -> None:
    assert SOURCE.exists(), SOURCE
    app = QApplication.instance() or QApplication([])
    with tempfile.TemporaryDirectory(prefix="tur20-dashboard-export-") as temp_dir:
        root = Path(temp_dir)
        analysis_root = root / "analyses"
        workspace_root = root / "dashboards"
        output_path = root / "dashboard-export.xlsx"
        repository = FileAnalysisRepository(SOURCE, analysis_root)
        store = DashboardWorkspaceStore(workspace_root)
        window = AnalysisCenterWindow(
            source=SOURCE,
            settings=_settings(),
            analysis_repository=repository,
            workspace_store=store,
        )
        window.show()
        app.processEvents()
        try:
            assert window._payload["metrics"]["generated_at"] == "2026-07-01"
            prepared_ids = _pin_prepared_types(window)
            service = window.controller.analysis_service
            saved = [service.create_saved_analysis(definition) for definition in _custom_definitions()]
            unpinned = service.create_saved_analysis(
                AnalysisDefinition(
                    analysis_id="preview-unpinned",
                    title="Pinlenmemiş Özel Analiz",
                    dataset="acceptances",
                    visualization="kpi",
                    measures=[MeasureDefinition("", "count_rows")],
                )
            )
            for definition in saved:
                assert window._set_custom_analysis_dashboard_pinned(definition.analysis_id, True)
            assert len(window.workspace.placements) == 6
            expected_collection = DashboardExportCollector(window._custom_dashboard_controller).collect(
                window.workspace,
                window._dashboard_items,
            )
            expected_titles = [item.card.title for item in expected_collection.items]
            assert len(expected_titles) == 6

            executed: list[str] = []
            refresh_calls = 0
            real_execute = service.execute_analysis
            real_refresh = window.controller.refresh_payload

            def execute_spy(definition):
                executed.append(definition.analysis_id)
                return real_execute(definition)

            def refresh_spy():
                nonlocal refresh_calls
                refresh_calls += 1
                return real_refresh()

            service.execute_analysis = execute_spy
            window.controller.refresh_payload = refresh_spy
            result = export_dashboard_excel(
                output_path,
                workspace=window.workspace,
                dashboard_items=window._dashboard_items,
                custom_controller=window._custom_dashboard_controller,
                source=SOURCE,
                exported_at=datetime(2026, 7, 9, 8, 0),
            )
            assert executed == [definition.analysis_id for definition in saved]
            assert unpinned.analysis_id not in executed
            assert refresh_calls == 0
            assert result.exported_card_count == 6
            assert result.warning_count == 0
            assert result.sheet_count == 7

            workbook = load_workbook(output_path)
            assert workbook.sheetnames[0] == "Dashboard Özeti"
            summary = workbook["Dashboard Özeti"]
            assert summary["B3"].value == SOURCE.name
            assert summary["B5"].value == 6
            assert summary["B6"].value == 6
            detail_names = [summary.cell(row, 6).value for row in range(10, 16)]
            summary_titles = [summary.cell(row, 2).value for row in range(10, 16)]
            assert summary_titles == expected_titles
            assert workbook.sheetnames[1:] == detail_names

            collection_by_title = {item.card.title: item for item in expected_collection.items}
            for title, detail_name in zip(summary_titles, detail_names):
                item = collection_by_title[title]
                sheet = workbook[detail_name]
                if item.card.card_type == CardType.KPI:
                    assert sheet["A5"].value is not None
                elif item.card.card_type == CardType.CHART:
                    assert len(sheet._charts) == 1
                elif item.card.card_type == CardType.TABLE:
                    assert len(sheet.tables) == 1

            custom_kpi_sheet = workbook[detail_names[summary_titles.index("Özel Teslimat Sayısı")]]
            assert isinstance(custom_kpi_sheet["A5"].value, (int, float))
            assert custom_kpi_sheet["B10"].value.startswith("Toplam: ")
            custom_chart_sheet = workbook[detail_names[summary_titles.index("Özel Teslimat Donut")]]
            custom_chart = custom_chart_sheet._charts[0]
            assert custom_chart.legend.position == "b"
            assert custom_chart.dLbls.showVal is True
            assert custom_chart_sheet["A6"].value == "Diğer"
            assert custom_chart.series[0].data_points[0].graphicalProperties.solidFill.srgbClr.casefold().lstrip("#") == "93c5fd"
            custom_table_sheet = workbook[detail_names[summary_titles.index("Özel Teslimat Tablosu")]]
            assert [custom_table_sheet.cell(3, column).value for column in range(1, 5)] == [
                "Teslimat Adı",
                "Platform",
                "Planlanan Kabul Tarihi",
                "Planlanan Toplam",
            ]
            assert len(custom_table_sheet.tables) == 1

            placement_signature = [
                (p.placement_id, p.source_screen_id, p.card_id, p.x, p.y, p.w, p.h)
                for p in sorted(window.workspace.placements, key=placement_order)
            ]
            repository_path = repository.repository_path()
            window.close()
            app.processEvents()

            reloaded_repository = FileAnalysisRepository(SOURCE, analysis_root)
            reloaded = AnalysisCenterWindow(
                source=SOURCE,
                settings=_settings(),
                analysis_repository=reloaded_repository,
                workspace_store=DashboardWorkspaceStore(workspace_root),
            )
            reloaded.show()
            app.processEvents()
            try:
                reloaded_signature = [
                    (p.placement_id, p.source_screen_id, p.card_id, p.x, p.y, p.w, p.h)
                    for p in sorted(reloaded.workspace.placements, key=placement_order)
                ]
                assert reloaded_signature == placement_signature
                restart_output = root / "dashboard-export-restart.xlsx"
                restart_result = export_dashboard_excel(
                    restart_output,
                    workspace=reloaded.workspace,
                    dashboard_items=reloaded._dashboard_items,
                    custom_controller=reloaded._custom_dashboard_controller,
                    source=SOURCE,
                )
                assert restart_result.exported_card_count == 6
                assert load_workbook(restart_output).sheetnames[0] == "Dashboard Özeti"
            finally:
                reloaded.close()

            print("TUR 20 DASHBOARD EXCEL EXPORT SMOKE: PASS")
            print(f"source={SOURCE}")
            print(f"repository_path={repository_path}")
            print(f"workspace_path={store.workspace_path(SOURCE)}")
            print(f"output_path={output_path}")
            print(f"prepared_card_ids={prepared_ids}")
            print(f"pinned_custom_ids={[definition.analysis_id for definition in saved]}")
            print(f"unpinned_custom_id={unpinned.analysis_id}")
            print(f"executed_custom_ids={executed}")
            print(f"refresh_payload_calls={refresh_calls}")
            print(f"exported_card_count={result.exported_card_count}")
            print(f"warning_count={result.warning_count}")
            print(f"sheet_count={result.sheet_count}")
            print(f"sheet_order={workbook.sheetnames}")
            print(f"custom_chart_legend={custom_chart.legend.position}")
            print(f"custom_chart_show_values={custom_chart.dLbls.showVal}")
            print(f"custom_chart_categories={[custom_chart_sheet.cell(row, 1).value for row in range(5, 7)]}")
            print("verification=openpyxl structural verification; no human Excel visual QA")
        finally:
            try:
                window.close()
            except Exception:
                pass


if __name__ == "__main__":
    main()
