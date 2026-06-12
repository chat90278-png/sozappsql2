from __future__ import annotations

import sqlite3
import traceback
import unicodedata
from pathlib import Path
from types import SimpleNamespace
from urllib.parse import quote

from PySide6.QtCore import QObject, Signal
from openpyxl import load_workbook

from src.services.sts_excel_exporter import export_sts_to_excel


class ExcelExportWorker(QObject):
    """Excel dışa aktarımını arka plan thread'inde çalıştırır.

    Notlar:
    - Worker, run() içinde kendi salt-okunur SQLite bağlantısını açar.
    - Canlı store/ana thread bağlantısına dokunmaz.
    - Eski teşhis amaçlı .export_debug.log üretimi kaldırıldı.
    - Seçili platform exportunda Özet üst sayaçları seçili kapsamla uyumlu hale getirilir.
    """

    progress = Signal(int, str)
    finished = Signal(dict)
    failed = Signal(str)

    def __init__(self, db_path, output_path, options):
        super().__init__()
        self.db_path = Path(db_path)
        self.output_path = output_path
        self.options = options or {}

    def _open_readonly_conn(self) -> sqlite3.Connection:
        raw = str(self.db_path.resolve()).replace("\\", "/")
        uri = "file:" + quote(raw, safe="/:") + "?mode=ro"
        conn = sqlite3.connect(uri, uri=True)
        conn.execute("PRAGMA busy_timeout=5000")
        return conn

    @staticmethod
    def _norm(value: object) -> str:
        text = str(value or "").strip().casefold()
        text = text.replace("ı", "i").replace("İ", "i")
        text = unicodedata.normalize("NFKD", text)
        text = "".join(ch for ch in text if not unicodedata.combining(ch))
        return " ".join(text.split())

    @staticmethod
    def _as_int(value: object, default: int = 0) -> int:
        try:
            if value is None or value == "":
                return default
            return int(float(value))
        except Exception:
            return default

    def _selected_platform_names(self) -> list[str]:
        opts = self.options or {}
        platforms = opts.get("platforms") or []
        out: list[str] = []
        seen: set[str] = set()
        for item in platforms:
            name = str(item or "").strip()
            key = self._norm(name)
            if name and key not in seen:
                seen.add(key)
                out.append(name)
        return out

    def _count_contracts_from_platform_sheets(self, wb, selected_platforms: list[str]) -> int:
        selected_keys = {self._norm(p) for p in selected_platforms}
        count = 0
        for ws in wb.worksheets:
            if self._norm(ws.title) in {"ozet", "summary"}:
                continue
            if selected_keys and self._norm(ws.title) not in selected_keys:
                continue
            for row_idx in range(2, ws.max_row + 1):
                value = ws.cell(row_idx, 1).value
                if str(value or "").strip():
                    count += 1
        return count

    def _fix_selected_scope_summary(self, result: dict | None) -> dict:
        """Seçili platform exportunda Özet sayfasındaki global sayaçları düzeltir.

        Exporter bazı sürümlerde üstteki toplam sözleşme sayısını tüm STS dosyasından,
        alttaki platform tablosunu ise seçili platformdan hesaplıyor. Bu helper,
        seçili platform kapsamı varsa üst sayaçları alttaki platform toplamlarıyla
        eşitler ve returned payload içindeki contract_count değerini de düzeltir.
        """
        payload = dict(result or {})
        opts = self.options or {}
        selected_platforms = self._selected_platform_names()
        if self._norm(opts.get("scope")) != "selected" or not selected_platforms:
            return payload

        output_path = Path(str(self.output_path))
        if not output_path.exists():
            return payload

        try:
            wb = load_workbook(output_path)
        except Exception:
            return payload

        selected_keys = {self._norm(p) for p in selected_platforms}
        platform_count = len(selected_platforms)
        contract_count = None
        system_count = None
        delivery_count = None
        changed = False

        ws = wb["Özet"] if "Özet" in wb.sheetnames else None
        if ws is not None:
            platform_header_row = None
            for r in range(1, ws.max_row + 1):
                a = self._norm(ws.cell(r, 1).value)
                b = self._norm(ws.cell(r, 2).value)
                if a == "platform" and "sozlesme" in b:
                    platform_header_row = r
                    break

            if platform_header_row:
                total_contracts = 0
                total_systems = 0
                total_deliveries = 0
                matched_platforms = 0
                for r in range(platform_header_row + 1, ws.max_row + 1):
                    platform_name = str(ws.cell(r, 1).value or "").strip()
                    if not platform_name:
                        break
                    if self._norm(platform_name) not in selected_keys:
                        continue
                    matched_platforms += 1
                    total_contracts += self._as_int(ws.cell(r, 2).value)
                    total_systems += self._as_int(ws.cell(r, 3).value)
                    total_deliveries += self._as_int(ws.cell(r, 4).value)

                if matched_platforms:
                    platform_count = matched_platforms
                    contract_count = total_contracts
                    system_count = total_systems
                    delivery_count = total_deliveries

            # Platform tablosu bulunamadıysa, en azından platform sheetlerinden sözleşme sayısını çıkar.
            if contract_count is None:
                contract_count = self._count_contracts_from_platform_sheets(wb, selected_platforms)

            label_to_value = {
                "platform sayisi": platform_count,
                "sozlesme sayisi": contract_count,
            }
            if system_count is not None:
                label_to_value["sistem sayisi"] = system_count
            if delivery_count is not None:
                label_to_value["kabul/teslimat sayisi"] = delivery_count

            for r in range(1, ws.max_row + 1):
                key = self._norm(ws.cell(r, 1).value)
                if key in label_to_value and label_to_value[key] is not None:
                    new_value = int(label_to_value[key])
                    if ws.cell(r, 2).value != new_value:
                        ws.cell(r, 2).value = new_value
                        changed = True
        else:
            contract_count = self._count_contracts_from_platform_sheets(wb, selected_platforms)

        if changed:
            try:
                wb.save(output_path)
            except Exception:
                pass

        payload["platform_count"] = int(platform_count)
        if contract_count is not None:
            payload["contract_count"] = int(contract_count)
        if system_count is not None:
            payload["system_count"] = int(system_count)
        if delivery_count is not None:
            payload["delivery_count"] = int(delivery_count)
        return payload

    def run(self):
        conn = None
        try:
            # Eski debug log dosyası aynı çıktı adıyla kaldıysa temizle; yeni export log üretmez.
            try:
                Path(str(self.output_path)).with_suffix(".export_debug.log").unlink(missing_ok=True)
            except Exception:
                pass

            conn = self._open_readonly_conn()

            def _on_progress(p, m):
                try:
                    pct = int(max(0, min(100, int(p or 0))))
                except Exception:
                    pct = 0
                # Exporter 100 verdikten sonra seçili kapsam özet düzeltmesi yapılacağı için
                # son yüzdeyi worker kendisi gönderecek.
                if pct >= 100:
                    pct = 97
                self.progress.emit(pct, str(m or ""))

            db = SimpleNamespace(conn=conn, path=self.db_path)
            res = export_sts_to_excel(db, self.output_path, options=self.options, progress_cb=_on_progress)
            res = self._fix_selected_scope_summary(res or {})
            self.progress.emit(100, "Excel dosyası oluşturuldu.")
            self.finished.emit(res or {})
        except Exception as exc:
            self.failed.emit(str(exc) + "\n\n" + traceback.format_exc())
        finally:
            if conn is not None:
                try:
                    conn.close()
                except Exception:
                    pass
