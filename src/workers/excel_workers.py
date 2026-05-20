# -*- coding: utf-8 -*-
from __future__ import annotations

import re
import time
from pathlib import Path
from typing import Dict, List, Tuple, Optional

from PySide6.QtCore import QObject, Signal
from PySide6.QtWidgets import QApplication, QDialog, QLabel, QProgressBar, QVBoxLayout
from openpyxl import load_workbook

from src.config.app_config import TAG_SHEET, TAG_HEADERS, TAG_KIND_ASSIGN, MAIN_TOTAL_LABEL, BASE_HEADERS, EXTRA_SYSTEM_SHEET_NAMES
from src.domain.constants import DATA_START_ROW, CORE_SHEETS
from src.models.app_models import ComponentDef
from src.services.excel_store import ExcelStore, as_number
from src.ui.theme import STYLE


def normalize_sheet_name(name: str) -> str:
    txt = str(name or "").strip().lower()
    repl = {"ı": "i", "İ": "i", "ş": "s", "ğ": "g", "ü": "u", "ö": "o", "ç": "c"}
    for a, b in repl.items():
        txt = txt.replace(a, b)
    return txt


def is_system_sheet_name(name: str) -> bool:
    if not name:
        return True
    n = normalize_sheet_name(name)
    core_norm = {normalize_sheet_name(x) for x in CORE_SHEETS}
    if n in core_norm or n in EXTRA_SYSTEM_SHEET_NAMES or n.startswith("_"):
        return True
    if str(name).startswith("_"):
        return True
    return False


def safe_sheet_name(name: str) -> str:
    import re as _re
    n = _re.sub(r"[\/*?:\[\]]", "_", name.strip().upper())
    return n[:31] or "PLATFORM"

class ExcelLoadWorker(QObject):
    store_ready = Signal(object)
    batch_ready = Signal(str, object, int, str)
    index_ready = Signal(object, object, object)
    finished = Signal(object, object)
    failed = Signal(str)
    progress = Signal(int, str)

    def __init__(self, path: Path):
        super().__init__()
        self.path = Path(path)

    def _norm_label(self, text: str) -> str:
        s = str(text or "").strip().lower()
        s = s.translate(str.maketrans({
            "\u0131": "i", "\u0130": "i", "\u015f": "s", "\u015e": "s",
            "\u011f": "g", "\u011e": "g", "\u00fc": "u", "\u00dc": "u",
            "\u00f6": "o", "\u00d6": "o", "\u00e7": "c", "\u00c7": "c",
        }))
        s = re.sub(r"[^\w\s]", " ", s)
        s = s.replace("_", " ")
        return re.sub(r"\s+", " ", s).strip()

    def _is_main_total_row_text(self, delivery_name: str) -> bool:
        n = self._norm_label(delivery_name)
        return n in {self._norm_label(MAIN_TOTAL_LABEL), "ana sozlesme toplam", "ana sozlesme"}

    def _is_tag_assign_kind_text(self, kind: str) -> bool:
        return str(kind or "").strip().upper() in {"ASSIGN", TAG_KIND_ASSIGN}

    def _read_only_excluded_platforms(self, wb) -> set[str]:
        config_sheet = "_PlatformConfig"
        if config_sheet not in wb.sheetnames:
            return set()
        ws = wb[config_sheet]
        excluded = set()
        for row in ws.iter_rows(min_row=1, max_col=1, values_only=True):
            val = str(row[0] or "").strip() if row else ""
            if val:
                excluded.add(val)
        return excluded

    def _read_only_platform_names(self, wb) -> List[str]:
        excluded = self._read_only_excluded_platforms(wb)
        items: List[str] = []
        for ws in wb.worksheets:
            if getattr(ws, "sheet_state", "visible") != "visible":
                continue
            if is_system_sheet_name(ws.title):
                continue
            if ws.title in excluded:
                continue
            items.append(ws.title)
        return items

    def _read_only_tags_map(self, wb) -> Dict[Tuple[str, str, str], List[str]]:
        if TAG_SHEET not in wb.sheetnames:
            return {}
        ws = wb[TAG_SHEET]
        out: Dict[Tuple[str, str, str], List[str]] = {}
        for row_data in ws.iter_rows(min_row=2, max_col=len(TAG_HEADERS), values_only=True):
            def val(col: int, default=None):
                return row_data[col - 1] if col - 1 < len(row_data) else default

            if not self._is_tag_assign_kind_text(str(val(1) or "").strip().upper()):
                continue
            nm = str(val(2) or "").strip()
            p = str(val(6) or "").strip()
            no = str(val(7) or "").strip()
            ctype = str(val(8) or "").strip()
            if not nm or not p or not no:
                continue
            out.setdefault((p, no, ctype), []).append(nm)
        return out

    def _read_only_contract_rows(self, wb, platform: str, tags_map: Dict[Tuple[str, str, str], List[str]]) -> List[dict]:
        if platform not in wb.sheetnames:
            return []
        ws = wb[platform]
        rows: List[dict] = []
        p = safe_sheet_name(platform)
        for r, row_data in enumerate(
            ws.iter_rows(min_row=DATA_START_ROW, max_col=len(BASE_HEADERS), values_only=True),
            start=DATA_START_ROW,
        ):
            def val(col: int, default=None):
                return row_data[col - 1] if col - 1 < len(row_data) else default

            activity = str(val(5) or "").strip()
            delivery = str(val(6) or "").strip()
            if activity.upper() != "GENEL" or not self._is_main_total_row_text(delivery):
                continue

            ctype = str(val(4) or "").strip()
            is_main = self._norm_label(ctype) == "ana sozlesme"
            no = str(val(1) or "")
            if not no.strip():
                continue
            tags = tags_map.get((p, no.strip(), ctype.strip()), [])
            item = {
                "row": r,
                "platform": p,
                "no": no,
                "user": val(2),
                "type": ctype,
                "type_display": ctype if is_main else f"\u21b3 {ctype}",
                "link": "Ana S\u00f6zle\u015fme" if is_main else "Ana s\u00f6zle\u015fmeye ba\u011fl\u0131 SD",
                "status": val(12),
                "completion_date": val(11),
                "content": val(7),
                "is_main": is_main,
                "tags": tags,
            }
            item["search"] = " ".join(
                str(item.get(k, "") or "")
                for k in ["platform", "no", "user", "type", "link", "status", "completion_date", "content"]
            ).lower()
            if tags:
                item["search"] = (item["search"] + " " + " ".join(tags).lower()).strip()
            rows.append(item)
        return rows

    def _build_read_only_index(self) -> Tuple[List[str], List[dict], Dict[str, float]]:
        timings: Dict[str, float] = {}
        t0 = time.perf_counter()
        wb = load_workbook(self.path, read_only=True, data_only=True, keep_links=False)
        timings["read_only_open"] = time.perf_counter() - t0
        try:
            t_platforms = time.perf_counter()
            platforms = self._read_only_platform_names(wb)
            timings["platforms"] = time.perf_counter() - t_platforms

            t_tags = time.perf_counter()
            tags_map = self._read_only_tags_map(wb)
            timings["tags"] = time.perf_counter() - t_tags

            index: List[dict] = []
            total = len(platforms)
            t_index = time.perf_counter()
            if total == 0:
                self.progress.emit(80, "İndeks hazır")
                timings["index"] = 0.0
                return platforms, index, timings
            for i, platform in enumerate(platforms, start=1):
                base = int(((i - 1) / total) * 100)
                mapped_base = 10 + int((max(0, min(100, base)) * 72) / 100)
                self.progress.emit(mapped_base, f"Hızlı indeksleniyor: {platform} ({i}/{total})")
                index.extend(self._read_only_contract_rows(wb, platform, tags_map))
                done = int((i / total) * 100)
                mapped_done = 10 + int((max(0, min(100, done)) * 72) / 100)
                self.progress.emit(mapped_done, f"Hızlı indeksleniyor: {platform} ({i}/{total})")
            timings["index"] = time.perf_counter() - t_index
            return platforms, index, timings
        finally:
            try:
                wb.close()
            except Exception:
                pass

    def run(self):
        try:
            if str(self.path).lower().endswith(".sts"):
                raise RuntimeError("STS dosyası Excel worker ile açılamaz; STSStore kullanılmalı.")
            total_start = time.perf_counter()
            self.progress.emit(5, "Excel hızlı okunuyor...")
            platforms, index, timings = self._build_read_only_index()
            timings["read_only_total"] = sum(timings.values())
            self.index_ready.emit(platforms, index, dict(timings))

            self.progress.emit(86, "Excel düzenleme modu hazırlanıyor...")
            t_store = time.perf_counter()
            store = ExcelStore(self.path)
            # Eski platformlara CF kuralları ekle (sadece CF'siz sayfalara, bir kez)
            try:
                self.progress.emit(92, "Koşullu biçimlendirme kontrol ediliyor...")
                store.migrate_platform_cf_rules()
            except Exception:
                pass
            timings["full_store_open"] = time.perf_counter() - t_store
            timings["total"] = time.perf_counter() - total_start
            self.progress.emit(98, "Arayüz hazırlanıyor...")
            self.progress.emit(
                99,
                f"Ölçüm: indeks {timings.get('index', 0):.1f}s, Excel açılış {timings.get('full_store_open', 0):.1f}s",
            )
            # Perf kaydı
            try:
                from src.services.perf_tracker import record as _perf_record, OP_EXCEL_LOAD, file_size_mb
                _perf_record(
                    OP_EXCEL_LOAD, self.path,
                    timings["total"] * 1000,
                    meta={
                        "ro_open_ms":   round(timings.get("read_only_open", 0) * 1000, 1),
                        "full_open_ms": round(timings.get("full_store_open", 0) * 1000, 1),
                        "platforms":    len(platforms),
                        "contracts":    len(index),
                        "file_mb":      file_size_mb(self.path),
                    },
                )
            except Exception:
                pass
            self.finished.emit(store, index)
        except Exception as exc:
            self.failed.emit(str(exc))


class ComponentSaveWorker(QObject):
    progress = Signal(int, str)
    finished = Signal(object)
    failed = Signal(str)

    def __init__(self, path: Path, components_payload: List[dict], actor: str):
        super().__init__()
        self.path = Path(path)
        self.components_payload = [dict(x or {}) for x in list(components_payload or [])]
        self.actor = str(actor or "Sistem")

    def _to_component_defs(self) -> List[ComponentDef]:
        out: List[ComponentDef] = []
        for raw in self.components_payload:
            out.append(
                ComponentDef(
                    name=str(raw.get("name", "") or "").strip(),
                    version=str(raw.get("version", "") or ""),
                    unit=str(raw.get("unit", "Adet") or "Adet"),
                    active=bool(raw.get("active", True)),
                    usage=int(as_number(raw.get("usage", 1) or 1)),
                    platforms={str(k): bool(v) for k, v in dict(raw.get("platforms", {}) or {}).items()},
                )
            )
        return out

    def run(self):
        try:
            if str(self.path).lower().endswith(".sts"):
                raise RuntimeError("STS dosyası Excel worker ile açılamaz; STSStore kullanılmalı.")
            self.progress.emit(8, "Excel açılıyor...")
            store = ExcelStore(self.path)
            with store.batch_save():
                self.progress.emit(16, "Eski bileşen eşleşmeleri hazırlanıyor...")
                old_components = {str(c.name or "").strip().lower(): c for c in store.load_components()}
                result = self._to_component_defs()
                new_components = {str(c.name or "").strip().lower(): c for c in result}
                platforms = store.platform_names()
                # Sadece EKLEME yapilan platformlari guncelle (silme durumunda platform sayfalarina dokunma)
                # Silinen bilesenlerin eski verileri platform sayfasinda kalmali - veri kaybi olmamali.
                added_platforms = set()
                all_names = set(old_components.keys()) | set(new_components.keys())
                for nm in all_names:
                    oc = old_components.get(nm)
                    nc = new_components.get(nm)
                    for p in platforms:
                        old_assigned = bool(oc and oc.active and oc.platforms.get(p, False))
                        new_assigned = bool(nc and nc.active and nc.platforms.get(p, False))
                        # Sadece yeni atama (ekleme) varsa baslik guncelle
                        if not old_assigned and new_assigned:
                            added_platforms.add(p)
                        # Silme/cikarma durumunda (old_assigned=True, new_assigned=False) -> DOKUNMA

                self.progress.emit(34, "Bileşenler kaydediliyor...")
                store.write_components(result, actor=self.actor)

                targets = sorted(added_platforms)
                total = max(1, len(targets))
                for i, p in enumerate(targets, start=1):
                    pct = 34 + int((i * 56) / total)
                    self.progress.emit(pct, f"{p} yeni bileşen başlığı ekleniyor... ({i}/{total})")
                    # Sadece yeni bilesen eklenmisse platform basligini guncelle
                    store.rebuild_platform_headers(p, style_rows=False)
                    try:
                        _ws = store.wb[safe_sheet_name(p)]
                        _max_col = len(BASE_HEADERS) + 3 * len(store.assigned_components(p))
                        store._apply_platform_cf_rules(_ws, _max_col)
                    except Exception:
                        pass

                self.progress.emit(96, "Excel kaydediliyor...")
                store.save()
            self.progress.emit(100, "Tamamlandı")
            self.finished.emit({"changed_platforms": targets})
        except Exception as exc:
            self.failed.emit(str(exc))


class UserSaveWorker(QObject):
    progress = Signal(int, str)
    finished = Signal()
    failed = Signal(str)

    def __init__(self, path: Path, users_payload: List[dict], actor: str):
        super().__init__()
        self.path = Path(path)
        self.users_payload = [dict(x or {}) for x in list(users_payload or [])]
        self.actor = str(actor or "Sistem")

    def run(self):
        try:
            if str(self.path).lower().endswith(".sts"):
                raise RuntimeError("STS dosyası Excel worker ile açılamaz; STSStore kullanılmalı.")
            self.progress.emit(10, "Excel açılıyor...")
            store = ExcelStore(self.path)
            with store.batch_save():
                self.progress.emit(42, "Kullanıcılar kaydediliyor...")
                store.write_users(self.users_payload, actor=self.actor)
                self.progress.emit(94, "Excel kaydediliyor...")
                store.save()
            self.progress.emit(100, "Tamamlandı")
            self.finished.emit()
        except Exception as exc:
            self.failed.emit(str(exc))



class ContractSaveWorker(QObject):
    progress = Signal(int, str)
    finished = Signal(dict)
    failed = Signal(str)

    def __init__(
        self,
        path: Path,
        action: str,
        platform: str,
        contract_no: str,
        ci=None,
        systems=None,
        deliveries=None,
        old_contract_no: str = "",
        old_start_row: int = 0,
        start_row: int = 0,
        actor: str = "",
        store=None,   # Bellekteki store — load_workbook'u atlatır
    ):
        super().__init__()
        self.path = Path(path)
        self.action = action
        self.platform = platform
        self.contract_no = contract_no
        self.ci = ci
        self.systems = systems or []
        self.deliveries = deliveries or {}
        self.old_contract_no = old_contract_no
        self.old_start_row = old_start_row
        self.start_row = start_row
        self.actor = actor or "Sistem"
        self._store = store  # Varsa wb yeniden açılmaz

    def _open_store(self):
        """
        Mevcut bellekteki store'da wb yüklüyse doğrudan kullan.
        Yoksa yeni ExcelStore aç (yavaş yol — sadece ilk kayıtta).
        """
        s = self._store
        if str(self.path).lower().endswith(".sts"):
            if s is not None:
                return s, False
            raise RuntimeError("STS dosyası Excel worker ile açılamaz; STSStore kullanılmalı.")
        if s is not None and getattr(s, 'wb', None) is not None:
            return s, False  # (store, opened_new=False)
        return ExcelStore(self.path), True

    def run(self):
        _t0 = time.perf_counter()
        try:
            self.progress.emit(10, "Hazırlanıyor...")
            store, opened_new = self._open_store()
            if opened_new:
                self.progress.emit(20, "Excel açılıyor...")
            else:
                self.progress.emit(20, "Bellekteki Excel kullanılıyor...")

            payload = None
            with store.batch_save():
                if self.action == "write":
                    self.progress.emit(30, "Sözleşme yazılıyor...")
                    new_row = store.write_contract(
                        self.ci,
                        self.systems,
                        self.deliveries,
                        old_contract_no=self.old_contract_no or None,
                        old_start_row=self.old_start_row or None,
                    )
                    self.progress.emit(70, "Stiller uygulanıyor...")
                    store.flush_pending_styles()
                    self.progress.emit(95, "Excel kaydediliyor...")
                    store.save()
                    payload = {
                        "action": "write",
                        "platform": self.platform,
                        "contract_no": self.contract_no,
                        "start_row": int(new_row or 0),
                    }
                elif self.action == "delete":
                    self.progress.emit(40, "Sözleşme siliniyor...")
                    result = store.delete_contract(
                        self.platform,
                        self.contract_no,
                        start_row=self.start_row or None,
                        actor=self.actor or None,
                        progress_cb=lambda p, m: self.progress.emit(
                            40 + int(p * 0.5), m
                        ),
                    )
                    self.progress.emit(95, "Excel kaydediliyor...")
                    store.save()
                    payload = {
                        "action": "delete",
                        "platform": self.platform,
                        "contract_no": self.contract_no,
                        "start_row": self.start_row,
                        "result": result or {},
                    }
                elif self.action == "migrate_cf":
                    self.progress.emit(10, "CF kuralları kontrol ediliyor...")
                    migrated = store.migrate_platform_cf_rules()
                    self.progress.emit(90, f"{len(migrated)} platform güncellendi...")
                    payload = {"action": "migrate_cf", "migrated": migrated}
            total_ms = (time.perf_counter() - _t0) * 1000
            try:
                from src.services.perf_tracker import record as _pr, OP_CONTRACT_SAVE, OP_CONTRACT_DELETE
                if self.action == "write":
                    op = OP_CONTRACT_SAVE
                elif self.action == "delete":
                    op = OP_CONTRACT_DELETE
                else:
                    op = self.action or "excel_action"
                _pr(op, self.path, total_ms,
                    meta={"platform": self.platform, "contract_no": self.contract_no,
                          "reused_wb": not opened_new})
            except Exception:
                pass
            if payload is not None:
                self.finished.emit(payload)
        except Exception as exc:
            try:
                from src.services.perf_tracker import record as _pr, OP_CONTRACT_SAVE, OP_CONTRACT_DELETE
                if self.action == "write":
                    op = OP_CONTRACT_SAVE
                elif self.action == "delete":
                    op = OP_CONTRACT_DELETE
                else:
                    op = self.action or "excel_action"
                _pr(op, self.path, (time.perf_counter() - _t0) * 1000,
                    success=False, meta={"error": str(exc)})
            except Exception:
                pass
            self.failed.emit(str(exc))


class AnalyzeDialog(QDialog):
    """Excel okunurken kullanıcıya kısa analiz ekranı gösterir."""
    def __init__(self, path: Path, parent=None):
        super().__init__(parent)
        self.path = Path(path)
        self.store: Optional[ExcelStore] = None
        self.index: List[dict] = []
        self.setWindowTitle("Excel analiz ediliyor")
        self.setModal(True)
        self.resize(520, 190)
        self.setStyleSheet(STYLE)
        self.build()

    def build(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(24, 24, 24, 24)
        root.setSpacing(12)
        title = QLabel("Excel analiz ediliyor")
        title.setObjectName("mainTitle")
        root.addWidget(title)
        self.msg = QLabel("Platformlar, kullanıcılar, bileşenler ve sözleşme indeksi hazırlanıyor...")
        self.msg.setWordWrap(True)
        self.msg.setObjectName("muted")
        root.addWidget(self.msg)
        self.progress = QProgressBar()
        self.progress.setRange(0, 0)
        root.addWidget(self.progress)

    def run(self):
        self.show()
        QApplication.processEvents()
        if str(self.path).lower().endswith(".sts"):
            raise RuntimeError("STS dosyası Excel worker ile açılamaz; STSStore kullanılmalı.")
        self.store = ExcelStore(self.path)
        self.index = self.store.build_contract_index()
        QApplication.processEvents()
        self.accept()
        return self.store, self.index


