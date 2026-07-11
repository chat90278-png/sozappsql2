from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

SUPPORTED_UNITS: tuple[str, ...] = ("Adet", "Takım", "Set", "Metre", "Kg", "Litre")

_NAME_HEADERS = {
    "bileşen adı",
    "bilesen adi",
    "bileşen",
    "bilesen",
    "component name",
    "component",
    "name",
}
_UNIT_HEADERS = {"birim", "unit", "ölçü birimi", "olcu birimi"}
_NOTE_HEADERS = {"not", "açıklama", "aciklama", "description", "note"}
_STATUS_HEADERS = {"durum", "aktif", "active", "status", "aktif mi"}

_TRUE_VALUES = {
    "1",
    "true",
    "yes",
    "evet",
    "aktif",
    "active",
    "açık",
    "acik",
}
_FALSE_VALUES = {
    "0",
    "false",
    "no",
    "hayır",
    "hayir",
    "pasif",
    "inactive",
    "kapalı",
    "kapali",
}


@dataclass(frozen=True, slots=True)
class BulkComponentRow:
    name: str
    unit: str = "Adet"
    note: str = ""
    active: bool = True
    source_row: int = 0


@dataclass(frozen=True, slots=True)
class ValidatedBulkRow:
    row: BulkComponentRow
    state: str
    message: str

    @property
    def is_ready(self) -> bool:
        return self.state == "ready"


@dataclass(frozen=True, slots=True)
class WorkbookMapping:
    name_column: int | None
    unit_column: int | None = None
    note_column: int | None = None
    status_column: int | None = None
    header_row: int | None = None


@dataclass(frozen=True, slots=True)
class BulkMergeResult:
    components: list[dict[str, Any]]
    added: int
    skipped_existing: int
    skipped_duplicate: int
    skipped_blank: int



def normalize_space(value: Any) -> str:
    """Trim text and collapse consecutive whitespace without changing letter case."""

    return " ".join(str(value or "").strip().split())



def component_key(value: Any) -> str:
    """Return the canonical key used for duplicate and existing-name checks."""

    return normalize_space(value).casefold()



def normalize_unit(value: Any, default: str = "Adet") -> str:
    text = normalize_space(value)
    if not text:
        return normalize_space(default) or "Adet"

    aliases = {
        "adet": "Adet",
        "takım": "Takım",
        "takim": "Takım",
        "set": "Set",
        "metre": "Metre",
        "meter": "Metre",
        "m": "Metre",
        "kg": "Kg",
        "kilogram": "Kg",
        "litre": "Litre",
        "liter": "Litre",
        "lt": "Litre",
        "l": "Litre",
    }
    return aliases.get(text.casefold(), text)



def normalize_active(value: Any, default: bool = True) -> bool:
    if value is None or normalize_space(value) == "":
        return bool(default)
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)

    key = normalize_space(value).casefold()
    if key in _TRUE_VALUES:
        return True
    if key in _FALSE_VALUES:
        return False
    return bool(default)



def _cell(row: Sequence[Any], index: int | None) -> Any:
    if index is None or index < 0 or index >= len(row):
        return None
    return row[index]



def _header_key(value: Any) -> str:
    return normalize_space(value).casefold()



def _first_nonempty_row(matrix: Sequence[Sequence[Any]]) -> int | None:
    for index, row in enumerate(matrix):
        if any(normalize_space(cell) for cell in row):
            return index
    return None



def detect_workbook_mapping(matrix: Sequence[Sequence[Any]]) -> WorkbookMapping:
    """Detect simple-list or detailed worksheet columns.

    A recognized header row wins. Otherwise the first non-empty column is treated
    as the component name column and the sheet is parsed as a headerless list.
    """

    first_row_index = _first_nonempty_row(matrix)
    if first_row_index is None:
        return WorkbookMapping(name_column=None)

    first_row = matrix[first_row_index]
    mapping: dict[str, int] = {}
    for column, raw in enumerate(first_row):
        key = _header_key(raw)
        if not key:
            continue
        if key in _NAME_HEADERS and "name" not in mapping:
            mapping["name"] = column
        elif key in _UNIT_HEADERS and "unit" not in mapping:
            mapping["unit"] = column
        elif key in _NOTE_HEADERS and "note" not in mapping:
            mapping["note"] = column
        elif key in _STATUS_HEADERS and "status" not in mapping:
            mapping["status"] = column

    if "name" in mapping:
        return WorkbookMapping(
            name_column=mapping["name"],
            unit_column=mapping.get("unit"),
            note_column=mapping.get("note"),
            status_column=mapping.get("status"),
            header_row=first_row_index,
        )

    max_columns = max((len(row) for row in matrix), default=0)
    for column in range(max_columns):
        if any(normalize_space(_cell(row, column)) for row in matrix[first_row_index:]):
            return WorkbookMapping(name_column=column, header_row=None)
    return WorkbookMapping(name_column=None)



def rows_from_matrix(
    matrix: Sequence[Sequence[Any]],
    *,
    mapping: WorkbookMapping | None = None,
    default_unit: str = "Adet",
    default_active: bool = True,
) -> list[BulkComponentRow]:
    """Convert worksheet-like rows to normalized component rows."""

    selected = mapping or detect_workbook_mapping(matrix)
    if selected.name_column is None:
        return []

    start = (selected.header_row + 1) if selected.header_row is not None else 0
    output: list[BulkComponentRow] = []
    for source_index, raw_row in enumerate(matrix[start:], start=start + 1):
        name = normalize_space(_cell(raw_row, selected.name_column))
        if not name:
            continue
        output.append(
            BulkComponentRow(
                name=name,
                unit=normalize_unit(_cell(raw_row, selected.unit_column), default_unit),
                note=normalize_space(_cell(raw_row, selected.note_column)),
                active=normalize_active(_cell(raw_row, selected.status_column), default_active),
                source_row=source_index,
            )
        )
    return output



def validate_bulk_rows(
    rows: Sequence[BulkComponentRow],
    existing_components: Iterable[Any] = (),
) -> list[ValidatedBulkRow]:
    """Classify every row as ready, blank, duplicate, or existing.

    All occurrences of an in-list duplicate are marked duplicate, matching the
    live table behavior in the approved design.
    """

    existing_keys: set[str] = set()
    for component in existing_components:
        if isinstance(component, Mapping):
            raw_name = component.get("name")
        else:
            raw_name = getattr(component, "name", "")
        key = component_key(raw_name)
        if key:
            existing_keys.add(key)

    counts: dict[str, int] = {}
    for row in rows:
        key = component_key(row.name)
        if key:
            counts[key] = counts.get(key, 0) + 1

    validated: list[ValidatedBulkRow] = []
    for row in rows:
        key = component_key(row.name)
        if not key:
            validated.append(ValidatedBulkRow(row, "blank", "Bileşen adı eksik"))
        elif counts.get(key, 0) > 1:
            validated.append(ValidatedBulkRow(row, "duplicate", "Toplu listede tekrar ediyor"))
        elif key in existing_keys:
            validated.append(ValidatedBulkRow(row, "existing", "STS içinde zaten mevcut; atlanacak"))
        else:
            validated.append(ValidatedBulkRow(row, "ready", "Eklemeye hazır"))
    return validated



def _component_to_dict(component: Any) -> dict[str, Any]:
    if isinstance(component, Mapping):
        payload = dict(component)
    else:
        payload = {
            "id": getattr(component, "id", None),
            "name": getattr(component, "name", ""),
            "version": getattr(component, "version", ""),
            "unit": getattr(component, "unit", "Adet"),
            "active": getattr(component, "active", True),
            "usage": getattr(component, "usage", 1),
            "note": getattr(component, "note", ""),
            "display_order": getattr(component, "display_order", None),
            "platforms": dict(getattr(component, "platforms", {}) or {}),
        }
    payload["name"] = str(payload.get("name") or "").strip()
    payload["unit"] = str(payload.get("unit") or "Adet")
    payload["active"] = bool(payload.get("active", True))
    payload["note"] = str(payload.get("note") or "")
    payload["platforms"] = dict(payload.get("platforms") or {})
    return payload



def merge_new_components(
    existing_components: Sequence[Any],
    rows: Sequence[BulkComponentRow],
) -> BulkMergeResult:
    """Append ready rows to the complete component payload for one atomic write."""

    existing = [_component_to_dict(component) for component in existing_components]
    validated = validate_bulk_rows(rows, existing)

    highest_order = -1
    for index, component in enumerate(existing):
        raw_order = component.get("display_order")
        try:
            order = int(raw_order) if raw_order is not None else index
        except (TypeError, ValueError):
            order = index
        component["display_order"] = order
        highest_order = max(highest_order, order)

    added = 0
    skipped_existing = 0
    skipped_duplicate = 0
    skipped_blank = 0
    for result in validated:
        if result.state == "existing":
            skipped_existing += 1
            continue
        if result.state == "duplicate":
            skipped_duplicate += 1
            continue
        if result.state == "blank":
            skipped_blank += 1
            continue

        highest_order += 1
        existing.append(
            {
                "id": None,
                "name": normalize_space(result.row.name),
                "version": "",
                "unit": normalize_unit(result.row.unit, "Adet"),
                "active": bool(result.row.active),
                "usage": 1,
                "note": normalize_space(result.row.note),
                "display_order": highest_order,
                "platforms": {},
            }
        )
        added += 1

    return BulkMergeResult(
        components=existing,
        added=added,
        skipped_existing=skipped_existing,
        skipped_duplicate=skipped_duplicate,
        skipped_blank=skipped_blank,
    )



def list_workbook_sheets(path: str | Path) -> list[str]:
    from openpyxl import load_workbook

    workbook = load_workbook(filename=Path(path), read_only=True, data_only=True)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()



def load_workbook_matrix(path: str | Path, sheet_name: str | None = None) -> list[list[Any]]:
    from openpyxl import load_workbook

    workbook = load_workbook(filename=Path(path), read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
        rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
    finally:
        workbook.close()

    while rows and not any(normalize_space(cell) for cell in rows[-1]):
        rows.pop()
    return rows
