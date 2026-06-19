from __future__ import annotations

from collections import OrderedDict
from dataclasses import dataclass, field
from math import floor
from pathlib import Path
from typing import Any
import unicodedata

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import quote_sheetname

from src.services.sts_database import now_iso

NAVY = "082F6F"
COL_HEAD = "103C80"
SUMMARY_ROW = "DBE8F6"
BORDER = "000000"
TEXT = "001F54"
COLORS = ["EAD58C", "1BA9DA", "CBCBCB", "E8D5C8", "E7BF9D", "B7CFA5"]
STATUSES = ["", "Hazırlıkta", "Süreci Devam Ediyor", "Kontrol Bekliyor", "Teslim Edildi", "Gecikmede"]

@dataclass
class ReportLine:
    user_id: int; user: str; contract_id: int; contract: str; delivery_date: str
    component_id: int; component: str; quantity: int; serial_no: str; serial_key: str
    internal_location: str = ""; note: str = ""; delivery_location: str = ""
    component_group_key: str = ""

@dataclass
class SummaryRow:
    user_id: int; user: str; contract_id: int; contract: str; delivery_date: str
    status: str = ""; description: str = ""

@dataclass
class ReportData:
    platform_id: int; platform: str; locations: list[str]
    summary: list[SummaryRow] = field(default_factory=list)
    details: "OrderedDict[tuple[int,int], list[ReportLine]]" = field(default_factory=OrderedDict)


def _to_int_qty(value: Any) -> int:
    """Safely coerce a raw quantity (possibly float/None/negative) to a
    non-negative integer. Returns 0 for anything invalid or negative."""
    try:
        number = float(value)
    except (TypeError, ValueError):
        return 0
    if number != number or number < 0:  # NaN guard + negative guard
        return 0
    return int(floor(number + 1e-6))


def _resolve_component_quantity(planned: Any, delivered: Any, unit_count: int) -> int:
    """Determine the real, displayable row count for one delivery component.

    The report must create exactly as many visible detail rows as the
    component quantity says. Unit/serial rows are only labels for the
    Kuyruk No / Seri No cells; they must never inflate the row count. This
    fixes the old case where planned=5 but five/ten default unit rows caused
    the UI to ask for more serial rows than the quantity shown.

    Priority:
    1. planned quantity if positive
    2. delivered quantity if planned is missing/zero and delivered is positive
    3. unit_count only when no positive planned/delivered quantity exists
    4. explicit recorded zero => no row
    5. completely missing quantity info => one manual placeholder row
    """
    has_planned_value = planned is not None
    has_delivered_value = delivered is not None

    base = _to_int_qty(planned) if has_planned_value else 0
    if base <= 0 and has_delivered_value:
        base = _to_int_qty(delivered)
    if base > 0:
        return base

    if has_planned_value or has_delivered_value:
        return 0

    if unit_count > 0:
        return unit_count

    return 1


def _normalize_user_filter(user_id: int | list[int] | tuple[int, ...] | set[int] | None) -> list[int]:
    """Normalize optional user filter. Empty list means no user filter."""
    if user_id is None:
        return []
    raw_values = user_id if isinstance(user_id, (list, tuple, set)) else [user_id]
    result: list[int] = []
    seen: set[int] = set()
    for value in raw_values:
        try:
            uid = int(value)
        except (TypeError, ValueError):
            continue
        if uid > 0 and uid not in seen:
            seen.add(uid)
            result.append(uid)
    return result




def _status_key(value: Any) -> str:
    """Normalize a Turkish status string for safe comparisons."""
    text = str(value or "").strip()
    if not text:
        return ""
    replacements = {
        "ı": "i", "İ": "i", "ş": "s", "Ş": "s",
        "ğ": "g", "Ğ": "g", "ü": "u", "Ü": "u",
        "ö": "o", "Ö": "o", "ç": "c", "Ç": "c",
    }
    for src, dst in replacements.items():
        text = text.replace(src, dst)
    text = unicodedata.normalize("NFKD", text.casefold())
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return " ".join(text.replace("-", " ").replace("_", " ").split())


def _is_completed_contract_status(value: Any) -> bool:
    """Only fully completed/closed contracts are excluded from this report.

    Intermediate statuses such as Başlanmadı, Hazırlıkta/Hazırlanıyor,
    Süreci Devam Ediyor and Parçalı Teslim are intentionally kept.
    """
    key = _status_key(value)
    return key in {
        "tamamlandi",
        "tamamlanmis",
        "bitmis",
        "bitti",
        "kapandi",
        "kapatildi",
        "completed",
        "closed",
    }

def load_report_data(store, platform_name: str, user_id: int | list[int] | tuple[int, ...] | set[int] | None = None, contract_id: int | None = None) -> ReportData:
    """Load Platform Teslimat Özeti data.

    The report's "Kullanıcı / Ülke" filter is intentionally based on the
    contract owner users in contract_users. Delivery users are a different
    concept and can legitimately be empty in acceptance/delivery rows; using
    delivery_user_id here made valid contracts disappear whenever no delivery
    user had been assigned yet.
    """
    conn = store.db.conn
    prow = conn.execute("SELECT id,name FROM platforms WHERE name=?", (platform_name,)).fetchone()
    if not prow:
        return ReportData(0, platform_name or "Platform", [])
    pid = int(prow["id"])
    locations = [str(r[0]) for r in conn.execute("SELECT name FROM internal_locations WHERE COALESCE(is_active,1)=1 ORDER BY sort_order,name").fetchall()]

    # LEFT JOIN contract_platforms with the platform predicate in the JOIN
    # avoids duplicates and still supports legacy rows where only
    # contracts.platform_id was written.
    params: list[Any] = [pid, pid, pid]
    clauses = ["(cp.platform_id IS NOT NULL OR c.platform_id=?)"]

    user_ids = _normalize_user_filter(user_id)
    if user_ids:
        placeholders = ",".join("?" for _ in user_ids)
        clauses.append(f"ru.user_id IN ({placeholders})")
        params.extend(user_ids)
    if contract_id:
        clauses.append("c.id=?")
        params.append(int(contract_id))

    # One row per delivery_component here (NOT per unit). Unit identifiers
    # are fetched separately and then used only as labels for rows created
    # from the real planned/delivered component quantity.
    component_rows = conn.execute(f"""
        WITH report_users AS (
            -- Preferred source: contract owner users selected on the contract.
            SELECT c0.id AS contract_id, u0.id AS user_id, u0.name AS user_name
            FROM contracts c0
            JOIN contract_users cu0 ON cu0.contract_id=c0.id
            JOIN users u0 ON u0.id=cu0.user_id
            WHERE COALESCE(u0.active,1)=1

            UNION

            -- Fallback for legacy/partial data with no contract owner rows:
            -- use assigned delivery users if present.
            SELECT c0.id AS contract_id, du0.id AS user_id, du0.name AS user_name
            FROM contracts c0
            JOIN deliveries d0 ON d0.contract_id=c0.id
            JOIN users du0 ON du0.id=d0.delivery_user_id
            WHERE NOT EXISTS (SELECT 1 FROM contract_users cu0 WHERE cu0.contract_id=c0.id)

            UNION

            -- Final visibility fallback: keep orphaned contracts visible as
            -- Tanımsız rather than hiding the entire report page.
            SELECT c0.id AS contract_id, 0 AS user_id, 'Tanımsız' AS user_name
            FROM contracts c0
            WHERE NOT EXISTS (SELECT 1 FROM contract_users cu0 WHERE cu0.contract_id=c0.id)
              AND NOT EXISTS (SELECT 1 FROM deliveries d0 WHERE d0.contract_id=c0.id AND d0.delivery_user_id IS NOT NULL)
        )
        SELECT dc.id AS dc_id, d.id AS delivery_id, c.id AS contract_id, c.contract_no,
               COALESCE(c.status,'') AS contract_status,
               COALESCE(d.planned_acceptance_date,d.acceptance_date,c.acceptance_date,c.completion_date,'') AS delivery_date,
               ru.user_id AS user_id, COALESCE(ru.user_name,'Tanımsız') AS user_name,
               comp.id AS component_id, comp.name AS component_name, dc.planned, dc.delivered,
               COALESCE(pds.status,'') AS saved_status, COALESCE(pds.description,'') AS saved_description
        FROM deliveries d
        JOIN contracts c ON c.id=d.contract_id
        LEFT JOIN contract_platforms cp ON cp.contract_id=c.id AND cp.platform_id=?
        JOIN report_users ru ON ru.contract_id=c.id
        JOIN delivery_components dc ON dc.delivery_id=d.id
        JOIN components comp ON comp.id=dc.component_id
        LEFT JOIN platform_delivery_report_summary pds
            ON pds.platform_id=? AND pds.user_id=ru.user_id AND pds.contract_id=c.id
        WHERE {' AND '.join(clauses)}
        ORDER BY ru.user_name COLLATE NOCASE, c.contract_no, COALESCE(d.sort_order,0), d.id, COALESCE(comp.display_order,9999), comp.name COLLATE NOCASE, dc.id
    """, params).fetchall()

    # Completed contracts are out of scope for Platform Teslimat Özeti.
    # Keep partial/in-progress/not-started contracts visible; only the final
    # completed/closed contract statuses are removed here.
    component_rows = [r for r in component_rows if not _is_completed_contract_status(r["contract_status"])]

    if not component_rows:
        return ReportData(pid, str(prow["name"]), locations)

    dc_ids = [int(r["dc_id"]) for r in component_rows]
    placeholders = ",".join("?" for _ in dc_ids)
    unit_rows = conn.execute(
        f"""
        SELECT delivery_component_id, identifier, slot_no
        FROM delivery_component_units
        WHERE delivery_component_id IN ({placeholders})
        ORDER BY delivery_component_id, slot_no
        """,
        dc_ids,
    ).fetchall()
    units_by_dc: dict[int, list[tuple[str, int]]] = OrderedDict()
    for ur in unit_rows:
        dc_id = int(ur["delivery_component_id"])
        identifier = str(ur["identifier"] or "").strip()
        slot_no = int(ur["slot_no"] or 0)
        units_by_dc.setdefault(dc_id, []).append((identifier, slot_no))

    saved_rows = conn.execute(
        """
        SELECT user_id, contract_id, component_id, serial_key, internal_location, note, delivery_location
        FROM platform_delivery_report_lines
        WHERE platform_id=?
        """,
        (pid,),
    ).fetchall()
    saved_by_key: dict[tuple[int, int, int, str], tuple[str, str, str]] = {
        (int(sr["user_id"]), int(sr["contract_id"]), int(sr["component_id"]), str(sr["serial_key"])): (
            str(sr["internal_location"] or ""), str(sr["note"] or ""), str(sr["delivery_location"] or ""),
        )
        for sr in saved_rows
    }

    data = ReportData(pid, str(prow["name"]), locations)
    seen = set()
    for r in component_rows:
        uid = int(r["user_id"] or 0)
        cid = int(r["contract_id"] or 0)
        key = (uid, cid)
        if key not in seen:
            seen.add(key)
            data.summary.append(SummaryRow(
                uid,
                str(r["user_name"] or "Tanımsız"),
                cid,
                str(r["contract_no"] or "-"),
                str(r["delivery_date"] or ""),
                str(r["saved_status"] or ""),
                str(r["saved_description"] or ""),
            ))
            data.details[key] = []

        dc_id = int(r["dc_id"])
        component_id = int(r["component_id"] or 0)
        component_name = str(r["component_name"] or "-")
        identifiers = units_by_dc.get(dc_id, [])
        quantity = _resolve_component_quantity(r["planned"], r["delivered"], len(identifiers))
        if quantity <= 0:
            continue

        component_group_key = f"COMP-{component_id}"
        for slot in range(quantity):
            legacy_serial_key = ""
            if slot < len(identifiers) and identifiers[slot][0]:
                identifier, slot_no = identifiers[slot]
                serial_no = identifier
                legacy_serial_key = identifier
                serial_key = f"DC-{dc_id}:{identifier}"
            else:
                serial_no = "TBD"
                serial_key = f"DC-{dc_id}-{slot + 1}"

            saved = saved_by_key.get((uid, cid, component_id, serial_key))
            if saved is None and legacy_serial_key:
                # Backward compatibility for older saved rows that used the
                # raw identifier as serial_key before dc_id was added.
                saved = saved_by_key.get((uid, cid, component_id, legacy_serial_key))
            saved_location, saved_note, saved_delivery_location = saved or ("", "", "")

            data.details[key].append(ReportLine(
                uid,
                str(r["user_name"] or "Tanımsız"),
                cid,
                str(r["contract_no"] or "-"),
                str(r["delivery_date"] or ""),
                component_id,
                component_name,
                quantity,
                serial_no,
                serial_key,
                saved_location,
                saved_note,
                saved_delivery_location,
                component_group_key,
            ))

    # Keep same components together in the detail page, even when the same
    # component appears under multiple delivery rows. This makes the report
    # show one merged block with the total quantity for that system/component
    # instead of repeating the same system several times.
    def _detail_group_key(line: ReportLine) -> tuple:
        explicit = getattr(line, "component_group_key", "") or ""
        if explicit:
            return (line.user_id, line.contract_id, explicit)
        return (line.user_id, line.contract_id, line.component_id, line.component)

    for detail_lines in data.details.values():
        first_order: dict[tuple, int] = {}
        for idx, line in enumerate(detail_lines):
            first_order.setdefault(_detail_group_key(line), idx)
        detail_lines.sort(key=lambda line: first_order.get(_detail_group_key(line), 0))

        totals: dict[tuple, int] = {}
        for line in detail_lines:
            totals[_detail_group_key(line)] = totals.get(_detail_group_key(line), 0) + 1
        for line in detail_lines:
            line.quantity = totals.get(_detail_group_key(line), line.quantity)
    return data

def save_report_data(store, data: ReportData, summary_rows: list[dict], line_rows: list[dict]) -> None:
    """Persist user edits.

    Rows with no real FK-valid report user are shown under "Tanımsız" purely
    for visibility. user_id=0 is a display sentinel only:
    platform_delivery_report_summary/lines both reference users(id), so
    those rows are skipped here until the underlying contract/delivery has a
    real user assignment.
    """
    conn = store.db.conn; ts = now_iso()
    with store.db.tx():
        for r in summary_rows:
            uid = int(r['user_id'] or 0)
            if uid <= 0:
                continue
            conn.execute("""INSERT INTO platform_delivery_report_summary(platform_id,user_id,contract_id,status,description,created_at,updated_at)
            VALUES(?,?,?,?,?,?,?) ON CONFLICT(platform_id,user_id,contract_id) DO UPDATE SET status=excluded.status,description=excluded.description,updated_at=excluded.updated_at""",
            (data.platform_id, uid, int(r['contract_id']), r.get('status',''), r.get('description',''), ts, ts))
        for r in line_rows:
            uid = int(r['user_id'] or 0)
            if uid <= 0:
                continue
            serial_key = str(r.get('serial_key') or r.get('serial_no') or 'TBD').strip() or 'TBD'
            serial_no = str(r.get('serial_no') or serial_key).strip() or serial_key
            conn.execute("""INSERT INTO platform_delivery_report_lines(platform_id,user_id,contract_id,component_id,serial_no,serial_key,internal_location,note,delivery_location,created_at,updated_at)
            VALUES(?,?,?,?,?,?,?,?,?,?,?) ON CONFLICT(platform_id,user_id,contract_id,component_id,serial_key) DO UPDATE SET serial_no=excluded.serial_no,internal_location=excluded.internal_location,note=excluded.note,delivery_location=excluded.delivery_location,updated_at=excluded.updated_at""",
            (data.platform_id, uid, int(r['contract_id']), int(r['component_id']), serial_no, serial_key, r.get('internal_location',''), r.get('note',''), r.get('delivery_location',''), ts, ts))


def _group_consecutive_by_component(lines: list[ReportLine]) -> list[list[ReportLine]]:
    """Group consecutive ReportLine entries that share the same component.

    Rows are already produced contiguously per component by load_report_data,
    so this only needs to look at runs, not do a full re-grouping.
    """
    groups: list[list[ReportLine]] = []
    def group_key(line: ReportLine) -> tuple:
        explicit = getattr(line, "component_group_key", "") or ""
        if explicit:
            return (line.user_id, line.contract_id, explicit)
        return (line.user_id, line.contract_id, line.component_id, line.component)

    for line in lines:
        if groups and group_key(groups[-1][-1]) == group_key(line):
            groups[-1].append(line)
        else:
            groups.append([line])
    return groups


def safe_sheet_title(name: str, used: set[str] | None = None) -> str:
    bad = '[]:*?/\\'
    base = ''.join('_' if c in bad else c for c in str(name or 'Sayfa')).strip() or 'Sayfa'
    used = used if used is not None else set()
    title = base[:31]
    if title not in used:
        used.add(title)
        return title
    counter = 2
    while True:
        suffix = f" ({counter})"
        title = f"{base[:31-len(suffix)]}{suffix}"
        if title not in used:
            used.add(title)
            return title
        counter += 1


def export_report_to_excel(data: ReportData, path: str | Path) -> Path:
    wb = Workbook(); used_sheet_names: set[str] = set(); ws = wb.active; ws.title = safe_sheet_title(f"{data.platform} Teslimat Özeti", used_sheet_names); ws.sheet_properties.tabColor = "E11D48"
    thin = Side(style="thin", color=BORDER); border = Border(left=thin,right=thin,top=thin,bottom=thin)
    main_head = PatternFill("solid", fgColor=NAVY)       # main title band, e.g. "Örnek Platform TESLİMAT ÖZETİ"
    col_head = PatternFill("solid", fgColor=COL_HEAD)     # column header row
    summary_fill = PatternFill("solid", fgColor=SUMMARY_ROW)
    header_font = Font(color="FFFFFF", bold=True)
    body_font = Font(color=TEXT)
    body_font_bold = Font(color=TEXT, bold=True)

    ws.merge_cells("A1:E1")
    ws["A1"] = f"{data.platform} TESLİMAT ÖZETİ"; ws["A1"].fill=main_head; ws["A1"].font=header_font; ws["A1"].alignment=Alignment(horizontal="center", vertical="center")
    headers = ["Kullanıcı", "Sözleşme Adı veya Numarası", "Teslimat Tarihi", "Durum", "Açıklama"]
    for c,h in enumerate(headers,1):
        cell=ws.cell(2,c,h); cell.fill=col_head; cell.font=header_font; cell.alignment=Alignment(horizontal="center", vertical="center"); cell.border=border
    sheet_names = {key: safe_sheet_title(f"{rows[0].user} Teslimat Özeti", used_sheet_names) for key, rows in data.details.items() if rows}
    for r_idx, row in enumerate(data.summary,3):
        vals=[f"{row.user} ↗", row.contract, row.delivery_date, row.status, row.description]
        for c,v in enumerate(vals,1):
            cell=ws.cell(r_idx,c,v); cell.fill=summary_fill; cell.font=body_font; cell.border=border; cell.alignment=Alignment(horizontal="center" if c<5 else "left", vertical="center")
        if (row.user_id,row.contract_id) in sheet_names:
            ws.cell(r_idx,1).hyperlink = f"#{quote_sheetname(sheet_names[(row.user_id,row.contract_id)])}!A1"
            ws.cell(r_idx,1).style = "Hyperlink"
    for w, width in zip("ABCDE", [24,30,18,24,50]): ws.column_dimensions[w].width = width

    for key, lines in data.details.items():
        if not lines: continue
        ws = wb.create_sheet(sheet_names[key]); ws.sheet_properties.tabColor = "22C55E"
        user = lines[0].user; date = lines[0].delivery_date
        ws.merge_cells("A1:E1"); ws["A1"] = f"{user}\nTESLİMAT ÖZETİ"; ws["A1"].fill=main_head; ws["A1"].font=header_font; ws["A1"].alignment=Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws["F1"] = f"TESLİMAT TARİHİ\n({date or 'TBD'})"; ws["F1"].fill=main_head; ws["F1"].font=header_font; ws["F1"].alignment=Alignment(horizontal="center", vertical="center", wrap_text=True)
        for c,h in enumerate(["ANA SİSTEM","MİKTAR","KUYRUK NO / SERİ NO","LOKASYON","NOT","TESLİM EDİLECEK LOKASYON"],1):
            cell=ws.cell(2,c,h); cell.fill=col_head; cell.font=header_font; cell.border=border; cell.alignment=Alignment(horizontal="center", vertical="center")

        comp_color = {}
        row_idx = 3
        for line_group in _group_consecutive_by_component(lines):
            component_name = line_group[0].component
            comp_color.setdefault(component_name, COLORS[len(comp_color) % len(COLORS)])
            fill = PatternFill("solid", fgColor=comp_color[component_name])
            group_start = row_idx
            for row_offset, line in enumerate(line_group):
                r = group_start + row_offset
                # Ana Sistem / Miktar are only written once (top-left cell of
                # the merge); openpyxl errors if you write into a cell that
                # will later be covered by a merge range.
                if row_offset == 0:
                    ws.cell(r, 1, line.component)
                    ws.cell(r, 2, len(line_group))
                for c, v in [(3, line.serial_no), (4, line.internal_location), (5, line.note), (6, line.delivery_location)]:
                    cell = ws.cell(r, c, v)
                    cell.fill = fill; cell.border = border; cell.font = body_font
                    cell.alignment = Alignment(horizontal="center" if c in (3, 4) else "left", vertical="center")
                # Ana Sistem / Miktar cells still need fill+border even on
                # continuation rows so the merge band reads as one solid
                # color block end-to-end.
                for c in (1, 2):
                    cell = ws.cell(r, c)
                    cell.fill = fill; cell.border = border
            group_end = row_idx + len(line_group) - 1
            row_idx = group_end + 1
            if group_end > group_start:
                ws.merge_cells(start_row=group_start, start_column=1, end_row=group_end, end_column=1)
                ws.merge_cells(start_row=group_start, start_column=2, end_row=group_end, end_column=2)
            ws.cell(group_start, 1).font = body_font_bold
            ws.cell(group_start, 1).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(group_start, 2).font = body_font_bold
            ws.cell(group_start, 2).alignment = Alignment(horizontal="center", vertical="center")
        for w,width in zip("ABCDEF", [22,12,28,22,44,36]): ws.column_dimensions[w].width=width
    wb.save(path)
    return Path(path)
