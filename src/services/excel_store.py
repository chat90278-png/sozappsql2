# -*- coding: utf-8 -*-
from __future__ import annotations

import re
import getpass
import base64
import threading
from contextlib import contextmanager
from pathlib import Path
from datetime import date, datetime
from typing import Callable, Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.domain.constants import (
    CORE_SHEETS,
    HEADER_ROW,
    SUBHEADER_ROW,
    DATA_START_ROW,
)
from src.config.app_config import (
    COMP_SHEET,
    USERS_SHEET,
    PLATFORM_LOGO_SHEET,
    TAG_SHEET,
    TAG_KIND_DEF,
    TAG_KIND_ASSIGN,
    LOG_FOLDER_NAME,
    NAVY,
    GREEN,
    GRID,
    BASE_HEADERS,
    MAIN_TOTAL_LABEL,
    SYSTEM_TOTAL_SUFFIX,
    TR_MONTHS,
    LOG_HEADERS,
    TAG_HEADERS,
    EXTRA_SYSTEM_SHEET_NAMES,
)
from src.models.app_models import ComponentDef, ContractInfo, SystemInfo, DeliveryInfo, TagDef
def normalize_sheet_name(name: str) -> str:
    txt = str(name or "").strip().lower()
    repl = {
        "\u0131": "i",
        "\u0130": "i",
        "\u015f": "s",
        "\u011f": "g",
        "\u00fc": "u",
        "\u00f6": "o",
        "\u00e7": "c",
    }
    for a, b in repl.items():
        txt = txt.replace(a, b)
    return txt


def is_system_sheet_name(name: str) -> bool:
    if not name:
        return True
    n = normalize_sheet_name(name)
    core_norm = {normalize_sheet_name(x) for x in CORE_SHEETS}
    if n in core_norm or n in EXTRA_SYSTEM_SHEET_NAMES:
        return True
    if str(name).startswith("_") or n.startswith("_"):
        return True
    return False


def safe_sheet_name(name: str) -> str:
    n = re.sub(r"[\\/*?:\[\]]", "_", name.strip().upper())
    return n[:31] or "PLATFORM"


def to_iso(qdate: QDate) -> str:
    return f"{qdate.year():04d}-{qdate.month():02d}-{qdate.day():02d}"


def parse_iso_date(text: str) -> Optional[date]:
    text = (text or "").strip()
    if not text:
        return None
    try:
        return datetime.strptime(text, "%Y-%m-%d").date()
    except ValueError:
        return None


def add_months(d: date, months: int) -> date:
    month = d.month - 1 + int(months or 0)
    year = d.year + month // 12
    month = month % 12 + 1
    days = [31, 29 if year % 4 == 0 and (year % 100 != 0 or year % 400 == 0) else 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
    day = min(d.day, days[month - 1])
    return date(year, month, day)


def iso_or_blank(text: str) -> str:
    d = parse_iso_date(text)
    return d.isoformat() if d else ""


def as_number(v):
    try:
        return float(v or 0)
    except Exception:
        return 0.0


def fmt_num(v) -> str:
    try:
        f = float(v or 0)
        return str(int(f)) if f == int(f) else str(round(f, 2))
    except Exception:
        return str(v or "")

class ExcelStore:
    def __init__(self, path: Path):
        self.path = Path(path)
        self.wb = None
        self._platform_next_row_hint: Dict[str, int] = {}
        self._sheet_cache: Dict[str, List[List]] = {}
        self._merge_map_cache: Dict[str, Dict[Tuple[int, int], Tuple[int, int]]] = {}
        self._platform_names_cache: Optional[List[str]] = None
        self._pending_style_ranges: Dict[str, List[Tuple[int, int]]] = {}
        self._save_batch_depth = 0
        self._save_requested = False
        self._full_wb_event = threading.Event()  # wb hazır olduğunda set edilir
        self.open_or_create()

    def open_or_create(self):
        needs_initial_save = not self.path.exists()
        had_home_sheet = False
        if self.path.exists():
            keep_vba = self.path.suffix.lower() in {".xlsm", ".xltm"}
            self.wb = load_workbook(
                self.path, keep_vba=keep_vba, data_only=True,
                keep_links=False, rich_text=False
            )
            had_home_sheet = "ANASAYFA" in self.wb.sheetnames
        else:
            self.wb = Workbook()
            self.wb.active.title = "ANASAYFA"
            had_home_sheet = True
        self.ensure_core()
        if needs_initial_save or not had_home_sheet:
            self.save()
        self._full_wb_event.set()  # wb hazır

    def wait_for_full_wb(self, timeout: float = 0.01) -> bool:
        """wb yüklenene kadar bekler. Non-blocking için timeout=0.01 kullan."""
        return self._full_wb_event.wait(timeout=timeout)

    def reload_from_disk(self):
        self._sheet_cache.clear()
        self._merge_map_cache.clear()
        if self.path.exists():
            keep_vba = self.path.suffix.lower() in {".xlsm", ".xltm"}
            self.wb = load_workbook(
                self.path, keep_vba=keep_vba, data_only=True,
                keep_links=False, rich_text=False
            )
        else:
            self.wb = Workbook()
            self.wb.active.title = "ANASAYFA"
        self.ensure_core()
        self._platform_next_row_hint.clear()
        self._full_wb_event.set()

    def _invalidate_runtime_caches(self):
        self._sheet_cache.clear()
        self._merge_map_cache.clear()
        self._platform_names_cache = None

    def _queue_platform_style(self, platform: str, row_start: int, row_end: int):
        p = safe_sheet_name(platform)
        if not p:
            return
        a = max(DATA_START_ROW, int(row_start or DATA_START_ROW))
        b = max(a, int(row_end or a))
        self._pending_style_ranges.setdefault(p, []).append((a, b))

    @contextmanager
    def batch_save(self):
        self._save_batch_depth += 1
        try:
            yield self
        except Exception:
            self._save_batch_depth -= 1
            if self._save_batch_depth == 0:
                self._save_requested = False
            raise
        else:
            self._save_batch_depth -= 1
            if self._save_batch_depth == 0 and self._save_requested:
                self._save_requested = False
                self.save()

    def save(self):
        self._invalidate_runtime_caches()
        if self._save_batch_depth > 0:
            self._save_requested = True
            return
        try:
            self.wb.save(self.path)
        except ValueError as exc:
            # Bazi dosyalarda gomulu gorsellerin kaynak stream'i kapali olabiliyor.
            # Kayit kaybi yasamamak icin gorselleri kaldirip bir kez daha deneriz.
            if "closed file" not in str(exc).lower():
                raise
            for ws in self.wb.worksheets:
                if hasattr(ws, "_images") and ws._images:
                    ws._images = []
            self.wb.save(self.path)

    def _log_folder(self) -> Path:
        return self.path.parent / LOG_FOLDER_NAME

    def current_actor(self) -> str:
        try:
            u = getpass.getuser()
            if u:
                return str(u)
        except Exception:
            pass
        try:
            u = os.getlogin()
            if u:
                return str(u)
        except Exception:
            pass
        return "Sistem"

    def _log_file_for_year(self, year: int) -> Path:
        return self._log_folder() / f"sozlesme_takip_log_{year}.xlsx"

    def _log_sheet_name(self, month: int) -> str:
        m = max(1, min(12, int(month or 1)))
        return f"{m:02d}-{TR_MONTHS[m - 1]}"

    def _ensure_log_sheet(self, wb, sheet_name: str):
        already_exists = sheet_name in wb.sheetnames
        if already_exists:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(sheet_name)

        # Başlıklar zaten varsa her açılışta aynı stili yeniden yazma.
        if already_exists and ws.cell(1, 1).value:
            return ws

        for c, h in enumerate(LOG_HEADERS, 1):
            ws.cell(1, c, h)
        fill = PatternFill("solid", fgColor=NAVY)
        border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
        widths = [22, 16, 14, 16, 18, 14, 20, 24, 18, 18, 30]
        header_font = Font(color="FFFFFF", bold=True)
        header_align = Alignment(horizontal="center", vertical="center")
        for c, w in enumerate(widths, 1):
            cell = ws.cell(1, c)
            cell.fill = fill
            cell.font = header_font
            cell.alignment = header_align
            cell.border = border
            ws.column_dimensions[get_column_letter(c)].width = w
        ws.freeze_panes = "A2"
        return ws

    def _textify(self, v) -> str:
        if v is None:
            return ""
        if isinstance(v, float) or isinstance(v, int):
            return fmt_num(v)
        return str(v)

    def _append_log_rows(self, rows: List[dict]):
        if not rows:
            return
        folder = self._log_folder()
        folder.mkdir(parents=True, exist_ok=True)
        by_year: Dict[int, dict] = {}
        for row in rows:
            ts = row.get("ts")
            if not isinstance(ts, datetime):
                ts = datetime.now()
            year = int(ts.year)
            month = int(ts.month)
            if year not in by_year:
                f = self._log_file_for_year(year)
                if f.exists():
                    wb = load_workbook(f)
                else:
                    wb = Workbook()
                    wb.active.title = self._log_sheet_name(month)
                by_year[year] = {"wb": wb, "file": f}
            wb = by_year[year]["wb"]
            sheet_name = self._log_sheet_name(month)
            ws = self._ensure_log_sheet(wb, sheet_name)
            ws.append([
                ts.strftime("%Y-%m-%d %H:%M:%S"),
                self._textify(row.get("user")),
                self._textify(row.get("platform")),
                self._textify(row.get("contract")),
                self._textify(row.get("delivery")),
                self._textify(row.get("kind")),
                self._textify(row.get("component")),
                self._textify(row.get("field")),
                self._textify(row.get("old")),
                self._textify(row.get("new")),
                self._textify(row.get("reason")),
            ])
        for item in by_year.values():
            item["wb"].save(item["file"])

    def _contract_snapshot(
        self,
        ci: Optional[ContractInfo],
        systems: List[SystemInfo],
        deliveries_by_system: Dict[str, List[DeliveryInfo]],
    ) -> Optional[dict]:
        if ci is None:
            return None
        out = {
            "contract": {
                "no": str(ci.no or ""),
                "platform": str(ci.platform or ""),
                "user": str(ci.user or ""),
                "yi_yd": str(ci.yi_yd or ""),
                "contract_type": str(ci.contract_type or ""),
                "signature_date": str(ci.signature_date or ""),
                "t0_date": str(ci.t0_date or ""),
                "t0_months": int(as_number(ci.t0_months)),
                "completion_date": str(ci.completion_date or ""),
                "status": str(ci.status or ""),
                "note": str(ci.note or ""),
            },
            "systems": {},
            "deliveries": {},
        }
        for s in systems:
            out["systems"][str(s.name)] = {str(k): as_number(v) for k, v in (s.components or {}).items()}
        for sys_name, items in (deliveries_by_system or {}).items():
            for d in items:
                key = f"{sys_name}::{d.name}"
                out["deliveries"][key] = {
                    "system": str(sys_name),
                    "name": str(d.name or ""),
                    "status": str(d.status or ""),
                    "acceptance_date": str(d.acceptance_date or ""),
                    "note": str(d.note or ""),
                    "planned": {str(k): as_number(v) for k, v in (d.planned or {}).items()},
                    "delivered": {str(k): as_number(v) for k, v in (d.delivered or {}).items()},
                }
        return out

    def _contract_diff_logs(self, old_snapshot: Optional[dict], new_snapshot: Optional[dict], reason: str) -> List[dict]:
        rows: List[dict] = []
        if not old_snapshot and not new_snapshot:
            return rows
        ts = datetime.now()
        base = new_snapshot if new_snapshot else old_snapshot
        actor = self.current_actor()
        platform = str(base.get("contract", {}).get("platform") or "")
        contract_no = str(base.get("contract", {}).get("no") or "")

        rows.append({
            "ts": ts,
            "user": actor,
            "platform": platform,
            "contract": contract_no,
            "delivery": "",
            "kind": "Sözleşme",
            "component": "",
            "field": "Kayıt İşlemi",
            "old": "Yok" if not old_snapshot else "Var",
            "new": "Yok" if not new_snapshot else "Var",
            "reason": reason,
        })

        # Yeni sözleşme eklemede sadece tek satır özet log tut.
        if not old_snapshot and new_snapshot:
            return rows

        if not new_snapshot:
            return rows

        old_contract = (old_snapshot or {}).get("contract", {})
        new_contract = new_snapshot.get("contract", {})
        contract_fields = [
            ("user", "Kullanıcı"),
            ("yi_yd", "Yİ/YD"),
            ("contract_type", "Sözleşme Tipi"),
            ("signature_date", "İmza Tarihi"),
            ("t0_date", "T0 Tarihi"),
            ("t0_months", "T0+Ay"),
            ("completion_date", "Termin Tarihi"),
            ("status", "Durum"),
            ("note", "Not"),
        ]
        for key, label in contract_fields:
            ov = self._textify(old_contract.get(key, ""))
            nv = self._textify(new_contract.get(key, ""))
            if ov != nv:
                rows.append({
                    "ts": ts, "user": actor, "platform": platform, "contract": contract_no,
                    "delivery": "", "kind": "Sözleşme", "component": "",
                    "field": label, "old": ov, "new": nv, "reason": reason,
                })

        old_systems = (old_snapshot or {}).get("systems", {})
        new_systems = new_snapshot.get("systems", {})
        all_systems = sorted(set(old_systems.keys()) | set(new_systems.keys()))
        for sname in all_systems:
            old_map = old_systems.get(sname, {})
            new_map = new_systems.get(sname, {})
            for comp in sorted(set(old_map.keys()) | set(new_map.keys())):
                ov = as_number(old_map.get(comp, 0))
                nv = as_number(new_map.get(comp, 0))
                if ov == nv:
                    continue
                rows.append({
                    "ts": ts, "user": actor, "platform": platform, "contract": contract_no,
                    "delivery": f"{sname} Toplamı", "kind": "Sistem", "component": comp,
                    "field": "Sözleşme Adedi", "old": fmt_num(ov), "new": fmt_num(nv), "reason": reason,
                })

        old_delivs = (old_snapshot or {}).get("deliveries", {})
        new_delivs = new_snapshot.get("deliveries", {})
        all_keys = sorted(set(old_delivs.keys()) | set(new_delivs.keys()))
        for key in all_keys:
            od = old_delivs.get(key, {})
            nd = new_delivs.get(key, {})
            dname = str(nd.get("name") or od.get("name") or "")
            for field_key, label in [("status", "Durum"), ("acceptance_date", "Kabul Tarihi"), ("note", "Not")]:
                ov = self._textify(od.get(field_key, ""))
                nv = self._textify(nd.get(field_key, ""))
                if ov != nv:
                    rows.append({
                        "ts": ts, "user": actor, "platform": platform, "contract": contract_no,
                        "delivery": dname, "kind": "Teslimat", "component": "",
                        "field": label, "old": ov, "new": nv, "reason": reason,
                    })
            old_p = od.get("planned", {})
            new_p = nd.get("planned", {})
            old_d = od.get("delivered", {})
            new_d = nd.get("delivered", {})
            for comp in sorted(set(old_p.keys()) | set(new_p.keys()) | set(old_d.keys()) | set(new_d.keys())):
                op = as_number(old_p.get(comp, 0))
                np = as_number(new_p.get(comp, 0))
                if op != np:
                    rows.append({
                        "ts": ts, "user": actor, "platform": platform, "contract": contract_no,
                        "delivery": dname, "kind": "Teslimat", "component": comp,
                        "field": "Teslim Edilecek", "old": fmt_num(op), "new": fmt_num(np), "reason": reason,
                    })
                odv = as_number(old_d.get(comp, 0))
                ndv = as_number(new_d.get(comp, 0))
                if odv != ndv:
                    rows.append({
                        "ts": ts, "user": actor, "platform": platform, "contract": contract_no,
                        "delivery": dname, "kind": "Teslimat", "component": comp,
                        "field": "Teslim Edilen", "old": fmt_num(odv), "new": fmt_num(ndv), "reason": reason,
                    })
        return rows

    def delete_contract(
        self,
        platform: str,
        contract_no: str,
        start_row: Optional[int] = None,
        actor: Optional[str] = None,
        progress_cb: Optional[Callable[[int, str], None]] = None,
    ) -> Optional[dict]:
        def _progress(pct: int, msg: str):
            if progress_cb:
                try:
                    progress_cb(int(max(0, min(100, pct))), str(msg or ""))
                except Exception:
                    pass

        _progress(46, "Sözleşme bloğu aranıyor...")
        p = safe_sheet_name(platform)
        if p not in self.wb.sheetnames:
            return None
        ws = self.wb[p]
        target_no = str(contract_no or "").strip()
        # Avoid normalizing the whole platform sheet; target merges are cleaned after block lookup.
        comp_cols = self.component_col_map(p)
        blocks = self._contract_entry_blocks(ws, comp_cols=comp_cols)
        if not blocks:
            return None

        target_block = None
        if start_row is not None:
            sr = int(start_row)
            target_block = next((b for b in blocks if int(b["start"]) == sr), None)
            if target_block is None:
                target_block = next(
                    (b for b in blocks if int(b["start"]) <= sr <= int(b["end"])),
                    None,
                )
            if target_block is None:
                same_no = [b for b in blocks if str(b.get("no", "")).strip() == target_no]
                if same_no:
                    target_block = min(same_no, key=lambda b: abs(int(b["start"]) - sr))
        if target_block is None:
            same_no = [b for b in blocks if str(b.get("no", "")).strip() == target_no]
            if same_no:
                target_block = same_no[0]
        if target_block is None:
            return None

        start = int(target_block["start"])
        end = int(target_block["end"])
        if end < start:
            return None
        block_no = str(target_block.get("no", "") or target_no)

        prev_ci, prev_systems, prev_deliveries = self.load_contract_structure(
            p,
            block_no,
            start_row=start,
        )
        old_snapshot = self._contract_snapshot(prev_ci, prev_systems, prev_deliveries)

        _progress(58, "Birleştirilmiş hücreler temizleniyor...")
        for rng in list(ws.merged_cells.ranges):
            if rng.max_row < start or rng.min_row > end:
                continue
            if rng.min_row >= DATA_START_ROW:
                ws.unmerge_cells(str(rng))
        _progress(70, "Satırlar siliniyor...")
        ws.delete_rows(start, end - start + 1)
        self._sheet_cache.pop(ws.title, None)
        self._merge_map_cache.pop(ws.title, None)

        _progress(76, "Etiketler temizleniyor...")
        # _repair_contract_merges_near_row kaldırıldı — performans için merge onarımı yapılmaz
        self.delete_contract_tags(p, block_no, str(target_block.get("type", "") or ""))

        # delete_rows, alttaki satırları biçimleriyle birlikte yukarı taşır.
        # Tüm sayfayı baştan stillendirmek (binlerce satırda) çok maliyetli olduğu için
        # burada tam stil turu yapılmaz; veri bütünlüğü korunur ve işlem hızlanır.
        self._queue_platform_style(p, max(DATA_START_ROW, start - 3), min(ws.max_row, start + 8))
        _progress(82, "Excel dosyasına kaydediliyor...")
        self._set_next_row_hint(p, ws.max_row + 1)
        self.save()

        _progress(92, "Loglar yazılıyor...")
        logs = self._contract_diff_logs(old_snapshot, None, "Sözleşme silindi")
        if logs:
            if not actor:
                actor = self.current_actor()
            if actor:
                for row in logs:
                    row["user"] = actor
            self._append_log_rows(logs)
        _progress(100, "Silme tamamlandı")
        return {
            "platform": p,
            "contract_no": block_no,
            "start_row": int(start),
            "end_row": int(end),
            "deleted_rows": int(end - start + 1),
        }

    def _is_effectively_empty_sheet(self, ws) -> bool:
        if ws.max_row > 1 or ws.max_column > 1:
            return False
        return ws.cell(1, 1).value in (None, "")

    def ensure_core(self):
        if "ANASAYFA" not in self.wb.sheetnames:
            visible = [ws for ws in self.wb.worksheets if ws.sheet_state == "visible"]
            if (len(visible) == 1 and self._is_effectively_empty_sheet(visible[0]) and
                    normalize_sheet_name(visible[0].title) in {"sheet", "sayfa1"}):
                visible[0].title = "ANASAYFA"
            else:
                self.wb.create_sheet("ANASAYFA")
        # Yardımcı sayfalar artık yalnızca ilgili özellik ilk kez veri yazdığında
        # oluşturulur. Böylece sayfasız/boş Excel bağlandığında kullanılmayan
        # VeriÇekme, kullanıcı, bileşen, etiket vb. sayfalar gereksiz oluşmaz.
        # Helper sheet styles are validated lazily on write; doing it here makes
        # large workbooks feel slower before the user can interact.

    def platform_names(self) -> List[str]:
        if self._platform_names_cache is not None:
            return list(self._platform_names_cache)
        excluded = set(self.load_excluded_platforms())
        items = []
        for ws in self.wb.worksheets:
            if ws.sheet_state != "visible":
                continue
            if is_system_sheet_name(ws.title):
                continue
            if ws.title in excluded:
                continue
            items.append(ws.title)
        self._platform_names_cache = list(items)
        return list(items)

    def all_sheet_names(self) -> List[str]:
        """Sistem sayfaları dışındaki tüm görünür sayfalar (excluded dahil)."""
        items = []
        for ws in self.wb.worksheets:
            if ws.sheet_state != "visible":
                continue
            if is_system_sheet_name(ws.title):
                continue
            items.append(ws.title)
        return items

    def load_excluded_platforms(self) -> List[str]:
        """Config sayfasından 'platform değil' olarak işaretlenmiş sayfa adlarını okur."""
        config_sheet = "_PlatformConfig"
        if config_sheet not in self.wb.sheetnames:
            return []
        ws = self.wb[config_sheet]
        excluded = []
        for row in ws.iter_rows(min_row=1, values_only=True):
            val = str(row[0] or "").strip() if row else ""
            if val:
                excluded.append(val)
        return excluded

    def save_excluded_platforms(self, excluded_names: List[str]):
        """Config sayfasına 'platform değil' olarak işaretlenen sayfa adlarını yazar."""
        config_sheet = "_PlatformConfig"
        if config_sheet in self.wb.sheetnames:
            del self.wb[config_sheet]
        ws = self.wb.create_sheet(config_sheet)
        ws.sheet_state = "hidden"
        for i, name in enumerate(excluded_names, start=1):
            ws.cell(i, 1, name)
        self.save()

    def ensure_component_sheet(self):
        ws = self.wb[COMP_SHEET] if COMP_SHEET in self.wb.sheetnames else self.wb.create_sheet(COMP_SHEET)
        headers = ["Bileşen Adı", "Birim", "Aktif", "Kullanım"]
        for c, h in enumerate(headers, 1):
            ws.cell(1, c, h)
        self.style_component_sheet()


    def ensure_user_sheet(self):
        ws = self.wb[USERS_SHEET] if USERS_SHEET in self.wb.sheetnames else self.wb.create_sheet(USERS_SHEET)
        headers = ["Kullanıcı Adı", "Yİ/YD", "Aktif", "Not"]
        for c, h in enumerate(headers, 1):
            ws.cell(1, c, h)
        fill = PatternFill("solid", fgColor=NAVY)
        border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
        widths = [28, 10, 10, 40]
        for c, w in enumerate(widths, 1):
            cell = ws.cell(1, c)
            cell.fill = fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
            ws.column_dimensions[get_column_letter(c)].width = w
        ws.freeze_panes = "A2"

    def ensure_platform_logo_sheet(self):
        ws = self.wb[PLATFORM_LOGO_SHEET] if PLATFORM_LOGO_SHEET in self.wb.sheetnames else self.wb.create_sheet(PLATFORM_LOGO_SHEET)
        headers = ["Platform", "LogoB64", "Format", "Guncelleme"]
        for c, h in enumerate(headers, 1):
            ws.cell(1, c, h)
        fill = PatternFill("solid", fgColor=NAVY)
        border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
        widths = [24, 80, 12, 22]
        for c, w in enumerate(widths, 1):
            cell = ws.cell(1, c)
            cell.fill = fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
            ws.column_dimensions[get_column_letter(c)].width = w
        ws.freeze_panes = "A2"

    def _logo_sheet(self):
        self.ensure_platform_logo_sheet()
        return self.wb[PLATFORM_LOGO_SHEET]

    def _logo_row(self, platform: str) -> int:
        ws = self._logo_sheet()
        p = safe_sheet_name(platform)
        for r in range(2, ws.max_row + 1):
            if str(ws.cell(r, 1).value or "").strip() == p:
                return r
        return ws.max_row + 1

    def set_platform_logo(self, platform: str, source_path: str) -> bool:
        src = Path(str(source_path or "").strip())
        if not src.exists() or not src.is_file():
            return False
        ext = src.suffix.lower()
        if ext not in {".png", ".jpg", ".jpeg", ".bmp", ".webp"}:
            return False
        try:
            raw = src.read_bytes()
        except Exception:
            return False
        if not raw:
            return False
        b64 = base64.b64encode(raw).decode("ascii")
        ws = self._logo_sheet()
        r = self._logo_row(platform)
        ws.cell(r, 1, safe_sheet_name(platform))
        ws.cell(r, 2, b64)
        ws.cell(r, 3, ext.lstrip("."))
        ws.cell(r, 4, datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        self.save()
        return True

    def get_platform_logo_bytes(self, platform: str) -> Optional[bytes]:
        if PLATFORM_LOGO_SHEET not in self.wb.sheetnames:
            return None
        ws = self.wb[PLATFORM_LOGO_SHEET]
        p = safe_sheet_name(platform)
        for r in range(2, ws.max_row + 1):
            if str(ws.cell(r, 1).value or "").strip() != p:
                continue
            b64 = str(ws.cell(r, 2).value or "").strip()
            if not b64:
                return None
            try:
                return base64.b64decode(b64)
            except Exception:
                return None
        return None

    def ensure_tag_sheet(self):
        ws = self.wb[TAG_SHEET] if TAG_SHEET in self.wb.sheetnames else self.wb.create_sheet(TAG_SHEET)
        for c, h in enumerate(TAG_HEADERS, 1):
            ws.cell(1, c, h)
        fill = PatternFill("solid", fgColor=NAVY)
        border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
        widths = [12, 26, 12, 36, 10, 16, 16, 18, 16, 18]
        for c, w in enumerate(widths, 1):
            cell = ws.cell(1, c)
            cell.fill = fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
            ws.column_dimensions[get_column_letter(c)].width = w
        ws.freeze_panes = "A2"

    def _is_tag_def_kind(self, kind: str) -> bool:
        k = str(kind or "").strip().upper()
        return k in {"TAG", TAG_KIND_DEF}

    def _is_tag_assign_kind(self, kind: str) -> bool:
        k = str(kind or "").strip().upper()
        return k in {"ASSIGN", TAG_KIND_ASSIGN}

    def _normalize_tag_sheet_rows(self):
        self.ensure_tag_sheet()
        ws = self.wb[TAG_SHEET]
        defs: List[List[object]] = []
        assigns: List[List[object]] = []
        for r in range(2, ws.max_row + 1):
            vals = [ws.cell(r, c).value for c in range(1, len(TAG_HEADERS) + 1)]
            kind = str(vals[0] or "").strip().upper()
            if self._is_tag_def_kind(kind):
                vals[0] = TAG_KIND_DEF
                defs.append(vals)
            elif self._is_tag_assign_kind(kind):
                vals[0] = TAG_KIND_ASSIGN
                assigns.append(vals)
        defs.sort(key=lambda a: self._normalize_label(str(a[1] or "")))
        assigns.sort(key=lambda a: (
            self._normalize_label(str(a[5] or "")),
            self._contract_no_sort_key(str(a[6] or "")),
            self._normalize_label(str(a[7] or "")),
            self._normalize_label(str(a[1] or "")),
            str(a[8] or ""),
        ))
        ws.delete_rows(2, max(0, ws.max_row - 1))
        row = 2
        for vals in defs + assigns:
            for c, v in enumerate(vals, 1):
                ws.cell(row, c, v)
            row += 1

    def load_tag_defs(self, active_only: bool = False) -> List[TagDef]:
        if TAG_SHEET not in self.wb.sheetnames:
            return []
        ws = self.wb[TAG_SHEET]
        out: List[TagDef] = []
        data = self._get_sheet_data(ws)

        def val(row_data, col: int, default=None):
            return row_data[col - 1] if col - 1 < len(row_data) else default

        for _r_idx, row_data in enumerate(data[1:], start=2):
            kind = str(val(row_data, 1) or "").strip().upper()
            if not self._is_tag_def_kind(kind):
                continue
            name = str(val(row_data, 2) or "").strip()
            if not name:
                continue
            color = str(val(row_data, 3) or "#3B82F6").strip() or "#3B82F6"
            note = str(val(row_data, 4) or "").strip()
            active_txt = str(val(row_data, 5) or "Evet").strip().lower()
            active = active_txt in {"evet", "true", "1", "aktif", "yes"}
            if active_only and not active:
                continue
            out.append(TagDef(name=name, color=color, note=note, active=active))
        out.sort(key=lambda t: self._normalize_label(t.name))
        return out

    def write_tag_defs(self, tags: List[TagDef], actor: Optional[str] = None):
        self.ensure_tag_sheet()
        ws = self.wb[TAG_SHEET]
        assigns = []
        for r in range(2, ws.max_row + 1):
            kind = str(ws.cell(r, 1).value or "").strip().upper()
            if self._is_tag_assign_kind(kind):
                assigns.append([ws.cell(r, c).value for c in range(1, len(TAG_HEADERS) + 1)])
        ws.delete_rows(2, max(0, ws.max_row - 1))
        row = 2
        ordered_tags = sorted(list(tags or []), key=lambda t: self._normalize_label(str(getattr(t, "name", "") or "")))
        for t in ordered_tags:
            ws.cell(row, 1, TAG_KIND_DEF)
            ws.cell(row, 2, str(t.name or "").strip())
            ws.cell(row, 3, str(t.color or "#3B82F6").strip() or "#3B82F6")
            ws.cell(row, 4, str(t.note or "").strip())
            ws.cell(row, 5, "Evet" if t.active else "Hayır")
            row += 1
        valid_names = {str(t.name or "").strip() for t in tags}
        for data in assigns:
            tag_name = str(data[1] or "").strip()
            if tag_name and tag_name not in valid_names:
                continue
            data[0] = TAG_KIND_ASSIGN
            for c, v in enumerate(data, 1):
                ws.cell(row, c, v)
            row += 1
        self._normalize_tag_sheet_rows()
        self.save()

    def upsert_tag_def(self, tag: TagDef):
        self.ensure_tag_sheet()
        ws = self.wb[TAG_SHEET]
        target = self._normalize_label(tag.name)
        row_idx = None
        for r in range(2, ws.max_row + 1):
            kind = str(ws.cell(r, 1).value or "").strip().upper()
            if not self._is_tag_def_kind(kind):
                continue
            name = str(ws.cell(r, 2).value or "").strip()
            if self._normalize_label(name) == target:
                row_idx = r
                break
        if row_idx is None:
            row_idx = ws.max_row + 1
        ws.cell(row_idx, 1, TAG_KIND_DEF)
        ws.cell(row_idx, 2, str(tag.name or "").strip())
        ws.cell(row_idx, 3, str(tag.color or "#3B82F6").strip() or "#3B82F6")
        ws.cell(row_idx, 4, str(tag.note or "").strip())
        ws.cell(row_idx, 5, "Evet" if tag.active else "Hayır")
        self._normalize_tag_sheet_rows()
        self.save()

    def delete_tag_def(self, tag_name: str):
        self.ensure_tag_sheet()
        ws = self.wb[TAG_SHEET]
        target = self._normalize_label(tag_name)
        keep_rows: List[List[object]] = []
        for r in range(2, ws.max_row + 1):
            vals = [ws.cell(r, c).value for c in range(1, len(TAG_HEADERS) + 1)]
            name = str(vals[1] or "").strip()
            if self._normalize_label(name) == target:
                continue
            keep_rows.append(vals)
        ws.delete_rows(2, max(0, ws.max_row - 1))
        row = 2
        for vals in keep_rows:
            for c, v in enumerate(vals, 1):
                ws.cell(row, c, v)
            row += 1
        self._normalize_tag_sheet_rows()
        self.save()

    def load_contract_tags(self, platform: str, contract_no: str, contract_type: str) -> List[dict]:
        if TAG_SHEET not in self.wb.sheetnames:
            return []
        ws = self.wb[TAG_SHEET]
        p = safe_sheet_name(platform)
        no = str(contract_no or "").strip()
        ctype = str(contract_type or "").strip()
        out: List[dict] = []
        data = self._get_sheet_data(ws)

        def val(row_data, col: int, default=None):
            return row_data[col - 1] if col - 1 < len(row_data) else default

        for _r_idx, row_data in enumerate(data[1:], start=2):
            kind = str(val(row_data, 1) or "").strip().upper()
            if not self._is_tag_assign_kind(kind):
                continue
            rp = str(val(row_data, 6) or "").strip()
            rno = str(val(row_data, 7) or "").strip()
            rt = str(val(row_data, 8) or "").strip()
            if rp != p or rno != no or rt != ctype:
                continue
            name = str(val(row_data, 2) or "").strip()
            if not name:
                continue
            out.append({
                "name": name,
                "color": str(val(row_data, 3) or "#3B82F6").strip() or "#3B82F6",
                "note": str(val(row_data, 4) or "").strip(),
                "assigned_at": str(val(row_data, 9) or "").strip(),
                "assigned_by": str(val(row_data, 10) or "").strip(),
            })
        return out

    def save_contract_tags(
        self,
        platform: str,
        contract_no: str,
        contract_type: str,
        tags: List[dict],
        actor: Optional[str] = None,
    ):
        if TAG_SHEET not in self.wb.sheetnames and not tags:
            return
        self.ensure_tag_sheet()
        ws = self.wb[TAG_SHEET]
        p = safe_sheet_name(platform)
        no = str(contract_no or "").strip()
        ctype = str(contract_type or "").strip()
        keep_rows: List[List[object]] = []
        for r in range(2, ws.max_row + 1):
            row_vals = [ws.cell(r, c).value for c in range(1, len(TAG_HEADERS) + 1)]
            kind = str(row_vals[0] or "").strip().upper()
            if not self._is_tag_assign_kind(kind):
                keep_rows.append(row_vals)
                continue
            rp = str(row_vals[5] or "").strip()
            rno = str(row_vals[6] or "").strip()
            rt = str(row_vals[7] or "").strip()
            if rp == p and rno == no and rt == ctype:
                continue
            keep_rows.append(row_vals)
        ws.delete_rows(2, max(0, ws.max_row - 1))
        row = 2
        for vals in keep_rows:
            for c, v in enumerate(vals, 1):
                ws.cell(row, c, v)
            row += 1
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        who = str(actor or self.current_actor())
        ordered_tags = sorted(
            [dict(t or {}) for t in list(tags or [])],
            key=lambda t: self._normalize_label(str(t.get("name", "") or "")),
        )
        for t in ordered_tags:
            nm = str((t or {}).get("name") or "").strip()
            if not nm:
                continue
            ws.cell(row, 1, TAG_KIND_ASSIGN)
            ws.cell(row, 2, nm)
            ws.cell(row, 3, str((t or {}).get("color") or "#3B82F6"))
            ws.cell(row, 4, str((t or {}).get("note") or "").strip())
            ws.cell(row, 5, "Evet")
            ws.cell(row, 6, p)
            ws.cell(row, 7, no)
            ws.cell(row, 8, ctype)
            ws.cell(row, 9, ts)
            ws.cell(row, 10, who)
            row += 1
        self._normalize_tag_sheet_rows()
        self.save()

    def delete_contract_tags(self, platform: str, contract_no: str, contract_type: str):
        self.ensure_tag_sheet()
        ws = self.wb[TAG_SHEET]
        p = safe_sheet_name(platform)
        no = str(contract_no or "").strip()
        ctype = str(contract_type or "").strip()
        keep_rows: List[List[object]] = []
        for r in range(2, ws.max_row + 1):
            vals = [ws.cell(r, c).value for c in range(1, len(TAG_HEADERS) + 1)]
            kind = str(vals[0] or "").strip().upper()
            if not self._is_tag_assign_kind(kind):
                keep_rows.append(vals)
                continue
            rp = str(vals[5] or "").strip()
            rno = str(vals[6] or "").strip()
            rt = str(vals[7] or "").strip()
            if rp == p and rno == no and rt == ctype:
                continue
            keep_rows.append(vals)
        ws.delete_rows(2, max(0, ws.max_row - 1))
        row = 2
        for vals in keep_rows:
            for c, v in enumerate(vals, 1):
                ws.cell(row, c, v)
            row += 1
        self._normalize_tag_sheet_rows()

    def tag_usage_counts(self) -> Dict[str, int]:
        self.ensure_tag_sheet()
        ws = self.wb[TAG_SHEET]
        counts: Dict[str, int] = {}
        for r in range(2, ws.max_row + 1):
            if not self._is_tag_assign_kind(str(ws.cell(r, 1).value or "").strip().upper()):
                continue
            name = str(ws.cell(r, 2).value or "").strip()
            if not name:
                continue
            counts[name] = counts.get(name, 0) + 1
        return counts

    def list_tag_assignments(self, tag_name: str) -> List[dict]:
        self.ensure_tag_sheet()
        ws = self.wb[TAG_SHEET]
        target = self._normalize_label(tag_name)
        out: List[dict] = []
        for r in range(2, ws.max_row + 1):
            kind = str(ws.cell(r, 1).value or "").strip().upper()
            if not self._is_tag_assign_kind(kind):
                continue
            name = str(ws.cell(r, 2).value or "").strip()
            if self._normalize_label(name) != target:
                continue
            out.append({
                "name": name,
                "color": str(ws.cell(r, 3).value or "#3B82F6").strip() or "#3B82F6",
                "note": str(ws.cell(r, 4).value or "").strip(),
                "platform": str(ws.cell(r, 6).value or "").strip(),
                "no": str(ws.cell(r, 7).value or "").strip(),
                "type": str(ws.cell(r, 8).value or "").strip(),
                "assigned_at": str(ws.cell(r, 9).value or "").strip(),
                "assigned_by": str(ws.cell(r, 10).value or "").strip(),
            })
        return out

    def load_tag_snapshot(self) -> Tuple[List[TagDef], Dict[str, List[dict]]]:
        if TAG_SHEET not in self.wb.sheetnames:
            return [], {}
        ws = self.wb[TAG_SHEET]
        defs: List[TagDef] = []
        assigns_by_key: Dict[str, List[dict]] = {}
        data = self._get_sheet_data(ws)

        def val(row_data, col: int, default=None):
            return row_data[col - 1] if col - 1 < len(row_data) else default

        for _r_idx, row_data in enumerate(data[1:], start=2):
            kind = str(val(row_data, 1) or "").strip().upper()
            if self._is_tag_def_kind(kind):
                name = str(val(row_data, 2) or "").strip()
                if not name:
                    continue
                color = str(val(row_data, 3) or "#3B82F6").strip() or "#3B82F6"
                note = str(val(row_data, 4) or "").strip()
                active_txt = str(val(row_data, 5) or "Evet").strip().lower()
                active = active_txt in {"evet", "true", "1", "aktif", "yes"}
                defs.append(TagDef(name=name, color=color, note=note, active=active))
                continue
            if not self._is_tag_assign_kind(kind):
                continue
            name = str(val(row_data, 2) or "").strip()
            if not name:
                continue
            key = self._normalize_label(name)
            assigns_by_key.setdefault(key, []).append({
                "name": name,
                "color": str(val(row_data, 3) or "#3B82F6").strip() or "#3B82F6",
                "note": str(val(row_data, 4) or "").strip(),
                "platform": str(val(row_data, 6) or "").strip(),
                "no": str(val(row_data, 7) or "").strip(),
                "type": str(val(row_data, 8) or "").strip(),
                "assigned_at": str(val(row_data, 9) or "").strip(),
                "assigned_by": str(val(row_data, 10) or "").strip(),
            })
        defs.sort(key=lambda t: self._normalize_label(t.name))
        for k in list(assigns_by_key.keys()):
            assigns_by_key[k].sort(
                key=lambda a: (
                    self._normalize_label(str(a.get("platform", "") or "")),
                    self._contract_no_sort_key(str(a.get("no", "") or "")),
                    self._normalize_label(str(a.get("type", "") or "")),
                    str(a.get("assigned_at", "") or ""),
                )
            )
        return defs, assigns_by_key

    def rename_tag_assignments(self, old_name: str, new_name: str, new_color: Optional[str] = None):
        self.ensure_tag_sheet()
        ws = self.wb[TAG_SHEET]
        old_norm = self._normalize_label(old_name)
        for r in range(2, ws.max_row + 1):
            kind = str(ws.cell(r, 1).value or "").strip().upper()
            if not self._is_tag_assign_kind(kind):
                continue
            name = str(ws.cell(r, 2).value or "").strip()
            if self._normalize_label(name) != old_norm:
                continue
            ws.cell(r, 2, str(new_name or "").strip())
            if new_color:
                ws.cell(r, 3, str(new_color))
        self._normalize_tag_sheet_rows()
        self.save()

    def _user_col_indices(self, ws) -> Tuple[int, int, int, int]:
        headers = {str(ws.cell(1, c).value or "").strip().lower().replace("ı", "i"): c for c in range(1, ws.max_column + 1)}
        name_col = headers.get("kullanıcı adı") or headers.get("kullanici adi") or 1
        yi_yd_col = headers.get("yi/yd") or headers.get("yi yd") or headers.get("yı/yd") or 2
        active_col = headers.get("aktif") or 3
        note_col = headers.get("not") or 4
        return int(name_col), int(yi_yd_col), int(active_col), int(note_col)

    def load_users(self, active_only: bool = True) -> List[dict]:
        if USERS_SHEET not in self.wb.sheetnames:
            return []
        ws = self.wb[USERS_SHEET]
        name_col, yi_yd_col, active_col, note_col = self._user_col_indices(ws)
        users = []
        data = self._get_sheet_data(ws)

        def val(row_data, col: int, default=None):
            return row_data[col - 1] if col - 1 < len(row_data) else default

        for _r_idx, row_data in enumerate(data[1:], start=2):
            name = str(val(row_data, name_col) or "").strip()
            if not name:
                continue
            yi_yd_raw = str(val(row_data, yi_yd_col) or "").strip().upper()
            yi_yd = "YD" if yi_yd_raw == "YD" else "Yİ"
            active_txt = str(val(row_data, active_col) or "Evet").strip().lower()
            active = active_txt in ["evet", "true", "1", "aktif", "yes"]
            if active_only and not active:
                continue
            users.append({
                "name": name,
                "yi_yd": yi_yd,
                "active": active,
                "note": str(val(row_data, note_col) or ""),
            })
        return users

    def write_users(self, users: List[dict], actor: str = "Sistem"):
        old_users = {u.get("name", ""): u for u in self.load_users(active_only=False)}
        ws = self.wb[USERS_SHEET] if USERS_SHEET in self.wb.sheetnames else self.wb.create_sheet(USERS_SHEET)
        ws.delete_rows(1, ws.max_row)
        headers = ["Kullanıcı Adı", "Yİ/YD", "Aktif", "Not"]
        for c, h in enumerate(headers, 1):
            ws.cell(1, c, h)
        for r, u in enumerate(users, 2):
            ws.cell(r, 1, u.get("name", ""))
            ws.cell(r, 2, "YD" if str(u.get("yi_yd", "Yİ")).strip().upper() == "YD" else "Yİ")
            ws.cell(r, 3, "Evet" if u.get("active", True) else "Hayır")
            ws.cell(r, 4, u.get("note", ""))
        self.ensure_user_sheet()
        self.save()
        new_users = {str(u.get("name", "")).strip(): u for u in users if str(u.get("name", "")).strip()}
        logs: List[dict] = []
        ts = datetime.now()
        all_names = sorted(set(old_users.keys()) | set(new_users.keys()))
        for name in all_names:
            old = old_users.get(name, {})
            new = new_users.get(name, {})
            for key, label in [("yi_yd", "Yİ/YD"), ("active", "Aktif"), ("note", "Not")]:
                ov = self._textify(old.get(key, "" if key != "active" else False))
                nv = self._textify(new.get(key, "" if key != "active" else False))
                if key == "active":
                    ov = "Evet" if str(ov).lower() in {"true", "1", "evet"} else "Hayır"
                    nv = "Evet" if str(nv).lower() in {"true", "1", "evet"} else "Hayır"
                if ov == nv and old and new:
                    continue
                if not old and not new:
                    continue
                logs.append({
                    "ts": ts,
                    "user": actor,
                    "platform": "",
                    "contract": "",
                    "delivery": "",
                    "kind": "Kullanıcı",
                    "component": "",
                    "field": f"{name} / {label}",
                    "old": ov if old else "",
                    "new": nv if new else "",
                    "reason": "Kullanıcı listesi güncellendi",
                })
        self._append_log_rows(logs)

    def style_component_sheet(self):
        ws = self.wb[COMP_SHEET] if COMP_SHEET in self.wb.sheetnames else self.wb.create_sheet(COMP_SHEET)
        dark_fill = PatternFill("solid", fgColor=NAVY)
        border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
        for cell in ws[1]:
            cell.fill = dark_fill if cell.column <= 4 else PatternFill("solid", fgColor="38761D")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        widths = {1: 32, 2: 12, 3: 10, 4: 10}
        for col, width in widths.items():
            ws.column_dimensions[get_column_letter(col)].width = width
        for col in range(5, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 14
        ws.freeze_panes = "A2"

    def _normalize_label(self, text: str) -> str:
        s = str(text or "").strip().lower()
        mojibake = {
            "Ã¶": "\u00f6",
            "Ã¼": "\u00fc",
            "Ã§": "\u00e7",
            "Ã–": "\u00d6",
            "ÃŸ": "\u015f",
            "ÅŸ": "\u015f",
            "Ä±": "\u0131",
            "Ä°": "\u0130",
            "ÄŸ": "\u011f",
        }
        for bad, good in mojibake.items():
            s = s.replace(str(bad).lower(), str(good).lower())
        s = s.translate(str.maketrans({
            "\u0131": "i",
            "\u0130": "i",
            "\u015f": "s",
            "\u015e": "s",
            "\u011f": "g",
            "\u011e": "g",
            "\u00fc": "u",
            "\u00dc": "u",
            "\u00f6": "o",
            "\u00d6": "o",
            "\u00e7": "c",
            "\u00c7": "c",
        }))
        s = re.sub(r"[^\w\s]", " ", s)
        s = s.replace("_", " ")
        s = re.sub(r"\s+", " ", s).strip()
        return s

    def _contract_no_sort_key(self, no_text: str):
        txt = str(no_text or "").strip()
        if txt.isdigit():
            return (0, int(txt), txt.lower())
        parts = re.split(r"(\d+)", txt.lower())
        key = []
        for p in parts:
            if p.isdigit():
                key.append((0, int(p)))
            else:
                key.append((1, p))
        return (1, key)

    def _is_main_total_row(self, delivery_name: str) -> bool:
        n = self._normalize_label(delivery_name)
        return n in {
            self._normalize_label(MAIN_TOTAL_LABEL),
            "ana sozlesme toplam",
            "ana sozlesme",
        }

    def _is_system_total_row(self, delivery_name: str) -> bool:
        n = self._normalize_label(delivery_name)
        if not n or n == self._normalize_label(MAIN_TOTAL_LABEL):
            return False
        return n.endswith(" toplami") or n.endswith(" toplam")

    def _extract_system_name_from_total(self, delivery_name: str, activity: str, fallback_index: int) -> str:
        raw = str(delivery_name or "").strip()
        act = str(activity or "").strip()
        cleaned = re.sub(r"[?]+$", "", raw).strip()
        if not cleaned:
            cleaned = raw
        candidate = re.sub(
            r"\s*(toplam\u0131|toplami|toplam)\s*[?]?\s*$",
            "",
            cleaned,
            flags=re.IGNORECASE,
        ).strip()
        if candidate:
            return candidate
        if act:
            return act
        return f"Sistem {fallback_index}"

    def ensure_platform_column_in_components(self, platform: str):
        ws = self.wb[COMP_SHEET] if COMP_SHEET in self.wb.sheetnames else self.wb.create_sheet(COMP_SHEET)
        platform = safe_sheet_name(platform)
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        if platform not in headers:
            col = ws.max_column + 1
            ws.cell(1, col, platform)
            self.style_component_sheet()

    def load_components(self) -> List[ComponentDef]:
        if COMP_SHEET not in self.wb.sheetnames:
            return []
        ws = self.wb[COMP_SHEET]

        # Güncel Bileşen sayfası şeması:
        # A: Bileşen Adı | B: Birim | C: Aktif | D: Kullanım | E...: Platformlar
        # Versiyon kolonu artık yoktur. Eski/kaymış şema tolere edilmez; Excel bu düzene göre olmalıdır.
        data = self._get_sheet_data(ws)

        def val(row_data, col: int, default=None):
            return row_data[col - 1] if col - 1 < len(row_data) else default

        header = data[0] if data else []
        platforms = [header[c - 1] for c in range(5, len(header) + 1) if header[c - 1]]
        yes_values = {"✓", "x", "X", "1", "True", "true", "EVET", "Evet", "evet", "aktif", "Aktif"}
        active_values = {"evet", "true", "1", "aktif", "✓", "x"}

        items = []
        for _r_idx, row_data in enumerate(data[1:], start=2):
            name = str(val(row_data, 1) or "").strip()
            if not name:
                continue

            unit = str(val(row_data, 2) or "Adet").strip() or "Adet"
            active_txt = str(val(row_data, 3) or "Evet").strip()
            usage_val = val(row_data, 4) or 1

            comp = ComponentDef(
                name=name,
                version="",
                unit=unit,
                active=str(active_txt).strip().lower() in active_values,
                usage=int(as_number(usage_val or 1)),
                platforms={}
            )
            for i, p in enumerate(platforms):
                col = 5 + i
                cell_val = val(row_data, col)
                comp.platforms[str(p)] = str(cell_val or "").strip() in yes_values
            items.append(comp)
        return items

    def write_components(self, components: List[ComponentDef], actor: str = "Sistem"):
        old_components = {c.name: c for c in self.load_components()}
        ws = self.wb[COMP_SHEET] if COMP_SHEET in self.wb.sheetnames else self.wb.create_sheet(COMP_SHEET)
        platforms = self.platform_names()
        ws.delete_rows(1, ws.max_row)
        headers = ["Bileşen Adı", "Birim", "Aktif", "Kullanım"] + platforms
        for c, h in enumerate(headers, 1):
            ws.cell(1, c, h)
        for r, comp in enumerate(components, 2):
            ws.cell(r, 1, comp.name)
            ws.cell(r, 2, comp.unit or "Adet")
            ws.cell(r, 3, "Evet" if comp.active else "Hayır")
            ws.cell(r, 4, comp.usage or 1)
            for c, p in enumerate(platforms, 5):
                cell = ws.cell(r, c, "✓" if comp.platforms.get(p, False) else "")
                cell.alignment = Alignment(horizontal="center")
                if cell.value:
                    cell.fill = PatternFill("solid", fgColor=GREEN)
        self.style_component_sheet()
        self.save()
        new_components = {c.name: c for c in components}
        logs: List[dict] = []
        ts = datetime.now()
        all_names = sorted(set(old_components.keys()) | set(new_components.keys()))
        for name in all_names:
            old = old_components.get(name)
            new = new_components.get(name)
            if old and not new:
                logs.append({
                    "ts": ts, "user": actor, "platform": "", "contract": "", "delivery": "",
                    "kind": "Bileşen", "component": name, "field": "Kayıt", "old": "Var", "new": "Silindi",
                    "reason": "Bileşen listesi güncellendi",
                })
                continue
            if new and not old:
                logs.append({
                    "ts": ts, "user": actor, "platform": "", "contract": "", "delivery": "",
                    "kind": "Bileşen", "component": name, "field": "Kayıt", "old": "Yok", "new": "Eklendi",
                    "reason": "Bileşen listesi güncellendi",
                })
                continue
            if not old or not new:
                continue
            pairs = [
                ("unit", "Birim", self._textify(old.unit), self._textify(new.unit)),
                ("active", "Aktif", "Evet" if old.active else "Hayır", "Evet" if new.active else "Hayır"),
            ]
            for _k, label, ov, nv in pairs:
                if ov == nv:
                    continue
                logs.append({
                    "ts": ts, "user": actor, "platform": "", "contract": "", "delivery": "",
                    "kind": "Bileşen", "component": name, "field": label, "old": ov, "new": nv,
                    "reason": "Bileşen listesi güncellendi",
                })
        self._append_log_rows(logs)

    def assigned_components(self, platform: str) -> List[str]:
        platform = safe_sheet_name(platform)
        comps = []
        for c in self.load_components():
            if c.active and c.platforms.get(platform, False):
                comps.append((c.name, int(c.usage or 0)))
        comps.sort(key=lambda x: (-x[1], str(x[0]).lower()))
        return [name for name, _u in comps]
    
    def increment_component_usage(self, used_component_names: List[str]):
        """Kullanımı arka planda artır: UI'da görünmez, sıralama için kullanılır."""
        if not used_component_names:
            return
        ws = self.wb[COMP_SHEET] if COMP_SHEET in self.wb.sheetnames else self.wb.create_sheet(COMP_SHEET)
        used = {str(x or "").strip().lower() for x in used_component_names if str(x or "").strip()}
        if not used:
            return
        for r in range(2, ws.max_row + 1):
            name = str(ws.cell(r, 1).value or "").strip().lower()
            if name in used:
                cur = int(as_number(ws.cell(r, 4).value or 0))
                ws.cell(r, 4, cur + 1)

    def create_platform(self, platform: str, actor: str = "Sistem", logo_source: Optional[str] = None):
        p = safe_sheet_name(platform)
        created = False
        if p not in self.wb.sheetnames:
            ws = self.wb.create_sheet(p)
            self.setup_platform_sheet(p)
            created = True
        self.ensure_platform_column_in_components(p)
        if logo_source:
            self.set_platform_logo(p, logo_source)
        self.save()
        if created:
            self._append_log_rows([{
                "ts": datetime.now(),
                "user": actor,
                "platform": p,
                "contract": "",
                "delivery": "",
                "kind": "Platform",
                "component": "",
                "field": "Platform Adı",
                "old": "",
                "new": p,
                "reason": "Platform oluşturuldu",
            }])

    def setup_platform_sheet(self, platform: str):
        ws = self.wb[platform]
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = f"A{DATA_START_ROW}"

        # Ust 3 satir sabit platform sablonu
        top_fill = PatternFill("solid", fgColor="D7DEE8")
        for r in range(1, 4):
            for c in range(1, 15):
                ws.cell(r, c).fill = top_fill

        ws.merge_cells("B1:D2")
        ws["B1"] = safe_sheet_name(platform)
        ws["B1"].fill = PatternFill("solid", fgColor=NAVY)
        ws["B1"].font = Font(color="FFFFFF", bold=True, size=16)
        ws["B1"].alignment = Alignment(horizontal="center", vertical="center")

        # Referans g?rseldeki gibi sat?r 3 ?erit g?r?n?m? (ayr? koyu buton yok).
        ws["B3"] = ""
        ws["B3"].fill = top_fill
        ws["B3"].font = Font(color="4F6280", bold=True, size=9)
        ws["B3"].alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells("F1:N3")
        ws["F1"] = "Platform g\u00f6rseli alan\u0131  -  Sa\u011f t\u0131kla -> Resim Ekle"
        ws["F1"].font = Font(color="8CA3C8", size=10)
        ws["F1"].alignment = Alignment(horizontal="center", vertical="center")

        self.rebuild_platform_headers(platform)

    def rebuild_platform_headers(self, platform: str, style_rows: bool = True):
        ws = self.wb[platform]
        comps = self.assigned_components(platform)
        # ?st bantta B3 h?cresini n?tr tut (setup sonras? da korunur).
        top_fill = PatternFill("solid", fgColor="D7DEE8")
        ws["B3"] = ""
        ws["B3"].fill = top_fill
        ws["B3"].font = Font(color="4F6280", bold=True, size=9)
        ws["B3"].alignment = Alignment(horizontal="center", vertical="center")
        # Header alanında eski merge varsa önce kaldır. Aksi halde MergedCell read-only hatası verir.
        for rng in list(ws.merged_cells.ranges):
            if rng.min_row <= SUBHEADER_ROW and rng.max_row >= HEADER_ROW:
                ws.unmerge_cells(str(rng))
        # clear header area
        _empty_fill = PatternFill(fill_type=None)
        _empty_font = Font()
        _empty_align = Alignment()
        for row in range(HEADER_ROW, SUBHEADER_ROW + 1):
            for col in range(1, 80):
                cell = ws.cell(row, col)
                cell.value = None
                cell.fill = _empty_fill
                cell.font = _empty_font
                cell.alignment = _empty_align
        for c, h in enumerate(BASE_HEADERS, 1):
            ws.cell(HEADER_ROW, c, h)
            ws.cell(HEADER_ROW, c).fill = PatternFill("solid", fgColor=NAVY)
            ws.cell(HEADER_ROW, c).font = Font(color="FFFFFF", bold=True, size=8)
            ws.cell(HEADER_ROW, c).alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
            ws.cell(SUBHEADER_ROW, c).fill = PatternFill("solid", fgColor=NAVY)
            ws.cell(SUBHEADER_ROW, c).font = Font(color="FFFFFF", bold=True, size=8)
        start_col = len(BASE_HEADERS) + 1
        for comp in comps:
            c1 = start_col
            c3 = start_col + 2
            try:
                ws.merge_cells(start_row=HEADER_ROW, start_column=c1, end_row=HEADER_ROW, end_column=c3)
            except Exception:
                pass
            ws.cell(HEADER_ROW, c1, comp)
            ws.cell(HEADER_ROW, c1).fill = PatternFill("solid", fgColor=NAVY)
            ws.cell(HEADER_ROW, c1).font = Font(color="FFFFFF", bold=True, size=8)
            ws.cell(HEADER_ROW, c1).alignment = Alignment(horizontal="center", vertical="center")
            for i, sub in enumerate(["Teslim Edilecek", "Teslim Edilen", "Kalan"]):
                cell = ws.cell(SUBHEADER_ROW, c1 + i, sub)
                cell.fill = PatternFill("solid", fgColor=NAVY)
                cell.font = Font(color="FFFFFF", bold=True, size=8)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
            start_col += 3
        widths = [24, 13, 9, 16, 18, 20, 38, 18, 15, 8, 22, 12, 14, 20]
        for c, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(c)].width = w
        for c in range(len(BASE_HEADERS) + 1, len(BASE_HEADERS) + 1 + 3 * len(comps)):
            ws.column_dimensions[get_column_letter(c)].width = 14

        max_col = len(BASE_HEADERS) + 3 * len(comps)
        self._apply_platform_cf_rules(ws, max_col)

        if style_rows:
            self.style_platform_rows(platform)

    def _apply_platform_cf_rules(self, ws, max_col: int) -> None:
        """Platform veri satırları için kalıcı ConditionalFormatting kuralları yazar."""
        from openpyxl.formatting.formatting import ConditionalFormattingList
        from openpyxl.formatting.rule import FormulaRule
        from openpyxl.styles import Font as _Font, PatternFill as _PF

        ws.conditional_formatting = ConditionalFormattingList()

        last_col_letter = get_column_letter(max(max_col, len(BASE_HEADERS)))
        cf_range = f"A{DATA_START_ROW}:{last_col_letter}1048576"

        ws.conditional_formatting.add(
            cf_range,
            FormulaRule(
                formula=[f'$F{DATA_START_ROW}="Ana Sözleşme Toplamı"'],
                fill=_PF("solid", fgColor="DCEBFF"),
                font=_Font(bold=True),
                stopIfTrue=True,
            ),
        )
        ws.conditional_formatting.add(
            cf_range,
            FormulaRule(
                formula=[f'ISNUMBER(SEARCH("Toplamı",$F{DATA_START_ROW}))'],
                fill=_PF("solid", fgColor="EEF2F6"),
                font=_Font(bold=True),
                stopIfTrue=True,
            ),
        )
        ws.conditional_formatting.add(
            cf_range,
            FormulaRule(
                formula=[f'$F{DATA_START_ROW}<>""'],
                fill=_PF("solid", fgColor="FFFFFF"),
                font=_Font(bold=False),
            ),
        )

    def migrate_platform_cf_rules(self, platform: Optional[str] = None) -> List[str]:
        """
        Mevcut platform sayfalarına ConditionalFormatting kuralları ekler.

        - platform=None  → tüm platformları kontrol eder
        - platform="X"   → yalnızca o platformu kontrol eder
        - CF kuralları zaten olan sayfaları ATLAR (her açılışta yeniden yazmaz)

        Döndürür: CF kuralı yazılan platform adları listesi
        """
        targets = [platform] if platform else self.platform_names()
        migrated = []
        for p in targets:
            pname = safe_sheet_name(p)
            if pname not in self.wb.sheetnames:
                continue
            ws = self.wb[pname]
            # CF kuralları zaten varsa atla — performans için kritik
            try:
                has_cf = len(list(ws.conditional_formatting)) > 0
            except Exception:
                has_cf = False
            if has_cf:
                continue
            comps = self.assigned_components(pname)
            max_col = len(BASE_HEADERS) + 3 * len(comps)
            self._apply_platform_cf_rules(ws, max_col)
            migrated.append(pname)
        if migrated:
            self.save()
        return migrated

    def component_col_map(self, platform: str) -> Dict[str, Tuple[int, int, int]]:
        comps = self.assigned_components(platform)
        start = len(BASE_HEADERS) + 1
        return {comp: (start + i*3, start + i*3 + 1, start + i*3 + 2) for i, comp in enumerate(comps)}

    def style_platform_rows(self, platform: str):
        ws = self.wb[platform]
        thin = Side(style="thin", color=GRID)
        _border = Border(left=thin, right=thin, top=thin, bottom=thin)
        _align = Alignment(vertical="center", wrap_text=True)
        _main_fill = PatternFill("solid", fgColor="DCEBFF")
        _system_fill = PatternFill("solid", fgColor="EEF2F6")
        _default_fill = PatternFill("solid", fgColor="FFFFFF")
        _bold_font = Font(bold=True)
        _normal_font = Font(bold=False)
        max_col = max(
            len(BASE_HEADERS) + 3 * len(self.assigned_components(platform)),
            len(BASE_HEADERS),
        )
        for r in range(HEADER_ROW, max(ws.max_row, SUBHEADER_ROW) + 1):
            if r >= DATA_START_ROW:
                delivery = str(self.cell_value(ws, r, 6) or "")
                if self._is_main_total_row(delivery):
                    fill, font = _main_fill, _bold_font
                elif self._is_system_total_row(delivery):
                    fill, font = _system_fill, _bold_font
                else:
                    fill, font = _default_fill, _normal_font
            else:
                fill = font = None
            for c in range(1, max_col + 1):
                cell = ws.cell(r, c)
                cell.border = _border
                cell.alignment = _align
                if fill is not None:
                    cell.fill = fill
                    cell.font = font

    def style_platform_rows_range(self, platform: str, row_start: int, row_end: int):
        p = safe_sheet_name(platform)
        if p not in self.wb.sheetnames:
            return
        ws = self.wb[p]
        a = max(DATA_START_ROW, int(row_start or DATA_START_ROW))
        b = min(ws.max_row, int(row_end or ws.max_row))
        if b < a:
            return
        thin = Side(style="thin", color=GRID)
        _border = Border(left=thin, right=thin, top=thin, bottom=thin)
        _align = Alignment(vertical="center", wrap_text=True)
        _main_fill = PatternFill("solid", fgColor="DCEBFF")
        _system_fill = PatternFill("solid", fgColor="EEF2F6")
        _default_fill = PatternFill("solid", fgColor="FFFFFF")
        _bold_font = Font(bold=True)
        _normal_font = Font(bold=False)
        max_col = max(
            len(BASE_HEADERS) + 3 * len(self.assigned_components(p)),
            len(BASE_HEADERS),
        )
        for r in range(a, b + 1):
            delivery = str(self.cell_value(ws, r, 6) or "")
            if self._is_main_total_row(delivery):
                fill, font = _main_fill, _bold_font
            elif self._is_system_total_row(delivery):
                fill, font = _system_fill, _bold_font
            else:
                fill, font = _default_fill, _normal_font
            for c in range(1, max_col + 1):
                cell = ws.cell(r, c)
                cell.border = _border
                cell.alignment = _align
                cell.fill = fill
                cell.font = font


    def flush_pending_styles(self) -> int:
        """Bekleyen platform stil güncellemelerini uygular. Arka plan thread'inden çağrılmalıdır."""
        pending = dict(self._pending_style_ranges)
        self._pending_style_ranges.clear()
        styled = 0
        for p, ranges in pending.items():
            if p in self.wb.sheetnames:
                for start, end in ranges:
                    self.style_platform_rows_range(p, start, end)
                    styled += 1
        if pending:
            self.save()
        return styled

    def _unmerge_data_cells_preserve_values(
        self,
        ws,
        row_start: int = DATA_START_ROW,
        row_end: Optional[int] = None,
        columns: Optional[set[int]] = None,
    ) -> None:
        """Data satırlarındaki merge'leri kaldırır ve üst-sol değeri tüm hücrelere yayar.

        Platform sayfalarında A-D ve E sütunlarındaki birleşimler insert/delete sırasında
        openpyxl tarafından güvenilir taşınmadığı için veri bölgesinde merge tutmuyoruz.
        Bu yardımcı eski dosyalardaki birleşimleri de veri kaybı olmadan normalize eder.
        """
        end_row = int(row_end or ws.max_row or row_start)
        col_filter = set(columns or [])
        for rng in list(ws.merged_cells.ranges):
            if rng.min_row < row_start or rng.max_row > end_row:
                continue
            if col_filter and (rng.max_col < min(col_filter) or rng.min_col > max(col_filter)):
                continue
            if col_filter and not any(rng.min_col <= c <= rng.max_col for c in col_filter):
                continue
            value = ws.cell(rng.min_row, rng.min_col).value
            ws.unmerge_cells(str(rng))
            for r in range(rng.min_row, rng.max_row + 1):
                for c in range(rng.min_col, rng.max_col + 1):
                    if col_filter and c not in col_filter:
                        continue
                    ws.cell(r, c, value)
        if end_row >= row_start:
            self.style_platform_rows_range(ws.title, row_start, end_row)


    def _apply_visual_merges_for_block(self, ws, start: int, end: int) -> None:
        """A sütununu merge etmeden görünür sözleşme/sistem alanlarını birleştirir.

        A sütunu satır-bazlı teknik anahtar olarak kalır; insert/delete öncesi tüm
        data merge'leri yine kaldırıldığı için bu görsel merge'ler veri güvenliğini
        etkilemez.
        """
        start = int(start or 0)
        end = int(end or 0)
        if start < DATA_START_ROW or end < start:
            return
        # A sütunu bilinçli olarak merge edilmez. B-D sözleşme üst bilgisini görsel toparlar.
        for c in (2, 3, 4):
            if end > start:
                ws.merge_cells(start_row=start, start_column=c, end_row=end, end_column=c)
            ws.cell(start, c).alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)

        sys_starts: List[Tuple[int, str]] = []
        for r in range(start, end + 1):
            delivery = str(self.cell_value(ws, r, 6) or "").strip()
            if not self._is_system_total_row(delivery):
                continue
            activity = str(self.cell_value(ws, r, 5) or "").strip()
            sys_name = self._extract_system_name_from_total(delivery, activity, len(sys_starts) + 1)
            sys_starts.append((r, sys_name))
        for i, (sr, sys_name) in enumerate(sys_starts):
            er = (sys_starts[i + 1][0] - 1) if i + 1 < len(sys_starts) else end
            ws.cell(sr, 5, sys_name)
            ws.cell(sr, 5).alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
            if er > sr:
                ws.merge_cells(start_row=sr, start_column=5, end_row=er, end_column=5)

    def _repair_single_contract_block_merges(self, ws, block: dict):
        start = int(block.get("start", 0) or 0)
        end = int(block.get("end", 0) or 0)
        if start < DATA_START_ROW or end < start:
            return
        self._unmerge_data_cells_preserve_values(ws, start, end, {1, 2, 3, 4, 5})
        self._apply_visual_merges_for_block(ws, start, end)

    def _repair_contract_merges_near_row(self, platform: str, row_hint: int):
        p = safe_sheet_name(platform)
        if p not in self.wb.sheetnames:
            return
        ws = self.wb[p]
        comp_cols = self.component_col_map(p)
        blocks = self._contract_entry_blocks(ws, comp_cols=comp_cols)
        if not blocks:
            return

        idx = 0
        for i, b in enumerate(blocks):
            s = int(b["start"])
            e = int(b["end"])
            if s <= row_hint <= e:
                idx = i
                break
            if row_hint < s:
                idx = i
                break
            idx = i

        lo = max(0, idx - 2)
        hi = min(len(blocks) - 1, idx + 2)
        for i in range(lo, hi + 1):
            self._repair_single_contract_block_merges(ws, blocks[i])

    def _get_sheet_data(self, ws) -> List[List]:
        """Sayfa verisini tek seferde önbelleğe alır. Sonraki çağrılar bellekten döner."""
        title = ws.title
        if title not in self._sheet_cache:
            self._sheet_cache[title] = [list(row) for row in ws.iter_rows(values_only=True)]
        return self._sheet_cache[title]

    def _get_cell(self, ws, row: int, col: int):
        """Cache üzerinden 1-indexed hücre okuma (cell_value'nın hızlı versiyonu)."""
        data = self._get_sheet_data(ws)
        r = row - 1
        c = col - 1
        if r < 0 or r >= len(data):
            return None
        row_data = data[r]
        if c < 0 or c >= len(row_data):
            return None
        return row_data[c]

    def _build_merge_map(self, ws) -> Dict[Tuple[int, int], Tuple[int, int]]:
        """Merge haritası: (r,c) → (top_row, top_col). Sadece slave hücreler haritada."""
        merge_map: Dict[Tuple[int, int], Tuple[int, int]] = {}
        for rng in ws.merged_cells.ranges:
            mr, mc = rng.min_row, rng.min_col
            for r in range(rng.min_row, rng.max_row + 1):
                for c in range(rng.min_col, rng.max_col + 1):
                    if r == mr and c == mc:
                        continue
                    merge_map[(r, c)] = (mr, mc)
        return merge_map

    def _get_merge_map(self, ws) -> Dict[Tuple[int, int], Tuple[int, int]]:
        title = ws.title
        if title not in self._merge_map_cache:
            self._merge_map_cache[title] = self._build_merge_map(ws)
        return self._merge_map_cache[title]

    def cell_value(self, ws, row: int, col: int):
        """Read cell value safely; if inside a merged range, return top-left value."""
        merge_map = self._get_merge_map(ws)
        actual = merge_map.get((row, col))
        if actual:
            row, col = actual
        return self._get_cell(ws, row, col)

    def _contract_block_rows(self, ws, contract_nos: set[str], comp_cols: Optional[Dict[str, Tuple[int, int, int]]] = None) -> List[int]:
        """Find all contiguous rows belonging to the target contract, even if merge in col A is broken."""
        targets = {str(x or "").strip() for x in contract_nos if str(x or "").strip()}
        if not targets:
            return []
        rows: List[int] = []
        in_block = False
        for r in range(DATA_START_ROW, ws.max_row + 1):
            c1 = str(self.cell_value(ws, r, 1) or "").strip()
            if not in_block:
                if c1 in targets:
                    in_block = True
                    rows.append(r)
                continue

            if c1 and c1 not in targets:
                break

            activity = str(self.cell_value(ws, r, 5) or "").strip()
            delivery = str(self.cell_value(ws, r, 6) or "").strip()
            has_data = bool(activity or delivery)
            if not has_data and comp_cols:
                for _comp, (pc, dc, rc) in comp_cols.items():
                    if as_number(self.cell_value(ws, r, pc)) or as_number(self.cell_value(ws, r, dc)) or as_number(self.cell_value(ws, r, rc)):
                        has_data = True
                        break

            if c1 in targets or has_data:
                rows.append(r)
            else:
                break
        return rows

    def _set_next_row_hint(self, platform: str, next_row: int):
        p = safe_sheet_name(platform)
        self._platform_next_row_hint[p] = max(DATA_START_ROW, int(next_row or DATA_START_ROW))

    def next_row(self, platform: str) -> int:
        p = safe_sheet_name(platform)
        ws = self.wb[p]
        fallback = max(ws.max_row + 1, DATA_START_ROW)
        hinted = int(self._platform_next_row_hint.get(p, 0) or 0)
        if hinted < DATA_START_ROW or hinted > fallback:
            return fallback
        # Hint satırına veri yazılmışsa riske girmeden fallback kullan.
        has_val = False
        for c in (1, 5, 6):
            if str(self.cell_value(ws, hinted, c) or "").strip():
                has_val = True
                break
        if has_val:
            return fallback
        return hinted

    def _platform_headers_match_components(self, platform: str, comps: List[str]) -> bool:
        p = safe_sheet_name(platform)
        if p not in self.wb.sheetnames:
            return False
        ws = self.wb[p]
        start_col = len(BASE_HEADERS) + 1
        subs = ["Teslim Edilecek", "Teslim Edilen", "Kalan"]
        for i, comp in enumerate(comps):
            c1 = start_col + i * 3
            if self._normalize_label(str(ws.cell(HEADER_ROW, c1).value or "")) != self._normalize_label(comp):
                return False
            for j, sub in enumerate(subs):
                if self._normalize_label(str(ws.cell(SUBHEADER_ROW, c1 + j).value or "")) != self._normalize_label(sub):
                    return False
        tail_start = start_col + len(comps) * 3
        for c in range(tail_start, ws.max_column + 1):
            if str(ws.cell(HEADER_ROW, c).value or "").strip():
                return False
            if str(ws.cell(SUBHEADER_ROW, c).value or "").strip():
                return False
        return True

    def _contract_entry_blocks(self, ws, comp_cols: Optional[Dict[str, Tuple[int, int, int]]] = None) -> List[dict]:
        starts: List[Tuple[int, str, str]] = []
        for r in range(DATA_START_ROW, ws.max_row + 1):
            activity = str(self.cell_value(ws, r, 5) or "").strip()
            delivery = str(self.cell_value(ws, r, 6) or "").strip()
            if not self._is_main_total_row(delivery):
                continue
            if activity.upper() != "GENEL":
                continue
            no = str(self.cell_value(ws, r, 1) or "").strip()
            ctype = str(self.cell_value(ws, r, 4) or "").strip()
            if not no:
                continue
            starts.append((r, no, ctype))

        def row_has_contract_data(row: int) -> bool:
            for c in range(1, len(BASE_HEADERS) + 1):
                if str(self.cell_value(ws, row, c) or "").strip():
                    return True
            if comp_cols:
                for _comp, (pc, dc, rc) in comp_cols.items():
                    for c in (pc, dc, rc):
                        v = self.cell_value(ws, row, c)
                        if v not in (None, ""):
                            return True
            return False

        blocks: List[dict] = []
        for i, (start, no, ctype) in enumerate(starts):
            next_start = starts[i + 1][0] if i + 1 < len(starts) else (ws.max_row + 1)
            end = start
            for r in range(start, next_start):
                if row_has_contract_data(r):
                    end = r
            blocks.append({
                "start": start,
                "end": max(start, end),
                "no": no,
                "type": ctype,
                "is_sd": bool(re.match(r"^SD-\d+$", str(ctype or "").strip().upper())),
            })
        return blocks

    def find_main_contract_info(self, platform: str, contract_no: str) -> Optional[dict]:
        p = safe_sheet_name(platform)
        if p not in self.wb.sheetnames:
            return None
        ws = self.wb[p]
        comp_cols = self.component_col_map(p)
        target = str(contract_no or "").strip()
        for b in self._contract_entry_blocks(ws, comp_cols=comp_cols):
            if b["no"] != target:
                continue
            if self._normalize_label(str(b["type"])) != self._normalize_label("Ana Sözleşme"):
                continue
            r = int(b["start"])
            return {
                "row": r,
                "block_start": int(b["start"]),
                "block_end": int(b["end"]),
                "no": str(self.cell_value(ws, r, 1) or ""),
                "user": str(self.cell_value(ws, r, 2) or ""),
                "yi_yd": str(self.cell_value(ws, r, 3) or ""),
                "type": str(self.cell_value(ws, r, 4) or ""),
                "status": str(self.cell_value(ws, r, 12) or "PLAN"),
                "platform": str(platform or ""),
            }
        return None

    def next_sd_code(self, platform: str, contract_no: str) -> str:
        p = safe_sheet_name(platform)
        if p not in self.wb.sheetnames:
            return "SD-1"
        ws = self.wb[p]
        comp_cols = self.component_col_map(p)
        target = str(contract_no or "").strip()
        max_n = 0
        for b in self._contract_entry_blocks(ws, comp_cols=comp_cols):
            if b["no"] != target:
                continue
            m = re.match(r"^SD-(\d+)$", str(b.get("type", "")).strip().upper())
            if m:
                max_n = max(max_n, int(m.group(1)))
        return f"SD-{max_n + 1}"

    # ─────────────────────────────────────────────────────────────────────────
    # BULLETPROOF INSERT / MERGE HELPERS
    # ─────────────────────────────────────────────────────────────────────────

    def _physical_family_last_row(self, ws, contract_no: str) -> int:
        """
        Gerçek (fiziksel) aile-son-satırı tespiti.

        ALGORITHM
        ---------
        1. Sütun 1'deki tüm merge aralıkları taranır; başlık hücresi == contract_no
           olan aralıkların tüm satırları 'aile' sayılır.
        2. Sütun 1'de doğrudan (merge dışı) contract_no içeren satırlar eklenir.
        3. Sonuç: max(aile_satırları)  →  ekleme noktası bu satır + 1 olmalıdır.

        INVARIANT
        ---------
        Dönen satır, söz konusu sözleşme ailesine ait veri içeren son satırdır.
        Bir sonraki satır güvenli ekleme noktasıdır ve başka bir sözleşmeye ait
        hiçbir veri içermez.
        """
        target = str(contract_no or "").strip()
        if not target:
            return DATA_START_ROW - 1

        family_rows: set = set()

        # 1) Merge aralıklarından tespit
        for rng in ws.merged_cells.ranges:
            if rng.min_col <= 1 <= rng.max_col and rng.min_row >= DATA_START_ROW:
                top_val = str(ws.cell(rng.min_row, rng.min_col).value or "").strip()
                if top_val == target:
                    for r in range(rng.min_row, rng.max_row + 1):
                        family_rows.add(r)

        # 2) Doğrudan hücre değerinden tespit (merge dışı ya da çözülmüş merge)
        for r in range(DATA_START_ROW, ws.max_row + 1):
            if str(ws.cell(r, 1).value or "").strip() == target:
                family_rows.add(r)

        return max(family_rows) if family_rows else DATA_START_ROW - 1

    def _safe_insert_rows(self, ws, insert_row: int, count: int) -> None:
        """
        openpyxl.insert_rows'u güvenli şekilde çağırır; merge'leri bozmaz.

        SORUN
        -----
        openpyxl.insert_rows(X, N) şu kurala göre merge aralıklarını günceller:
          - min_row >= X  →  tüm aralık N satır aşağı kaydırılır  ✓
          - min_row < X AND max_row < X  →  değişmez              ✓
          - min_row < X AND max_row >= X  →  max_row += N ile GENİŞLETİLİR ✗
            (Bu, ekleme noktasının tam sınırında olan merge'leri bozar.)

        ÇÖZÜM
        -----
        1. Ekleme noktasına yaslanan (min_row < X ≤ max_row) merge'ler
           önceden UNMERGE edilir.
        2. ws.insert_rows(X, N) çağrılır.
        3. Unmerge edilen aralıklar İKİYE BÖLÜNEREK yeniden merge edilir:
             • Üst parça : [min_row, X-1]        (ekleme öncesi pozisyon)
             • Alt parça : [X+N, max_row+N]      (kaydırılmış pozisyon)
           Böylece ne üstteki blok bozulur ne de alttaki blok.

        INVARIANT
        ---------
        Bu fonksiyon çağrıldıktan sonra:
          • insert_row ve üzerindeki TÜM satırlar +count kaydırılmıştır.
          • Hiçbir merge aralığı yeni eklenen boş satırları kapsamamaktadır.
          • Üst blokların merge'leri [min_row, X-1] aralığında korunur.
          • Alt blokların merge'leri [X+N, ...] aralığında doğru konumdadır.
        """
        if count <= 0:
            return

        # Ekleme noktasına yaslanan (spanning) merge'leri tespit et
        # KOŞUL: min_row < insert_row  VE  max_row >= insert_row
        # Bu aralıklar ekleme yapılırsa yanlış genişletilir.
        spanning: List[Tuple[int, int, int, int]] = []
        for rng in list(ws.merged_cells.ranges):
            if (rng.min_row >= DATA_START_ROW          # header satırlarına dokunma
                    and rng.min_row < insert_row        # ekleme noktasının üstünde başlıyor
                    and rng.max_row >= insert_row):     # ekleme noktasında bitiyor/aşıyor
                spanning.append((rng.min_row, rng.min_col,
                                 rng.max_row, rng.max_col))

        # 1) Ekleme öncesi: spanning merge'leri çöz
        for min_r, min_c, max_r, max_c in spanning:
            ws.unmerge_cells(start_row=min_r, start_column=min_c,
                             end_row=max_r, end_column=max_c)

        # 2) Satır ekle (artık spanning merge yok, openpyxl hatasız çalışır)
        ws.insert_rows(insert_row, count)

        # 3) Spanning merge'leri ikiye bölerek yeniden uygula
        for min_r, min_c, max_r, max_c in spanning:
            upper_end = insert_row - 1   # ekleme noktasının hemen üstü
            lower_start = insert_row + count          # kaydırılmış alt başlangıç
            lower_end = max_r + count                 # kaydırılmış alt bitiş

            # Üst parça: blok ekleme noktasından önce veri içeriyorsa merge et
            if upper_end >= min_r and upper_end > min_r:
                ws.merge_cells(start_row=min_r, start_column=min_c,
                               end_row=upper_end, end_column=max_c)

            # Alt parça: ekleme noktasından sonra veri içeriyorsa merge et
            if lower_end >= lower_start and lower_end > lower_start:
                ws.merge_cells(start_row=lower_start, start_column=min_c,
                               end_row=lower_end, end_column=max_c)

    def _sd_insert_row_from_anchor(self, ci: ContractInfo, ws, entry_blocks: List[dict]) -> Optional[int]:
        """
        SD sözleşmesinin ekleneceği satırı belirler.

        ALGORITHM
        ---------
        1. anchor_start ile doğrulanmış Ana Sözleşme bloğu bulunur.
        2. Aynı contract_no'ya ait tüm blokların mantıksal sonu (entry_blocks
           üzerinden) hesaplanır  →  logical_end
        3. Fiziksel son satır (_physical_family_last_row) hesaplanır  → physical_end
        4. insert_row = max(logical_end, physical_end) + 1
           Her iki yöntem de tutarlı sonuç veriyorsa aynı değer; biri hatalıysa
           diğeri onu düzeltir.

        INVARIANT
        ---------
        Dönen satır, söz konusu aile bloğunun DIŞINDA ve HEMEN ALTINDADIR.
        Bu satıra _safe_insert_rows ile eklendiğinde hiçbir mevcut blok zarar
        görmez.
        """
        anchor_start = int(getattr(ci, "sd_anchor_start_row", 0) or 0)
        anchor_no = str(getattr(ci, "sd_anchor_no", "") or "").strip()
        anchor_platform = safe_sheet_name(str(getattr(ci, "sd_anchor_platform", "") or ""))
        target_no = str(ci.no or "").strip()
        target_platform = safe_sheet_name(str(ci.platform or ""))
        if not anchor_start or not target_no:
            return None
        if anchor_no != target_no or anchor_platform != target_platform:
            return None

        # --- Mantıksal son satır (entry_blocks üzerinden) ---
        main_idx = -1
        for i, b in enumerate(entry_blocks):
            if int(b.get("start", 0) or 0) != anchor_start:
                continue
            if str(b.get("no", "")).strip() != target_no:
                return None
            if self._normalize_label(str(b.get("type", ""))) != self._normalize_label("Ana Sözleşme"):
                return None
            main_idx = i
            break
        if main_idx < 0:
            return None

        logical_end = int(entry_blocks[main_idx].get("end", 0) or 0)
        for j in range(main_idx + 1, len(entry_blocks)):
            bj = entry_blocks[j]
            if str(bj.get("no", "")).strip() != target_no:
                break
            logical_end = max(logical_end, int(bj.get("end", 0) or 0))

        # --- Fiziksel son satır (merge + raw cell taraması) ---
        physical_end = self._physical_family_last_row(ws, target_no)

        # --- Güvenli ekleme noktası: her iki yöntemin maksimumu ---
        # NEDEN MAX: eğer logical_end < physical_end ise, yani block detection
        # bazı satırları "göremediyse", physical_end doğru olanı verir.
        # Eğer physical_end < logical_end ise, yani merge yanlış okunuyorsa,
        # logical_end doğru olanı verir. İkisi de tutarlıysa eşit değer döner.
        safe_end = max(logical_end, physical_end)
        if safe_end <= 0:
            return None
        return safe_end + 1

    def write_contract(
        self,
        ci: ContractInfo,
        systems: List[SystemInfo],
        deliveries_by_system: Dict[str, List[DeliveryInfo]],
        old_contract_no: Optional[str] = None,
        old_start_row: Optional[int] = None,
    ) -> int:
        p = safe_sheet_name(ci.platform)
        self.create_platform(p)
        comps = self.assigned_components(p)
        if not self._platform_headers_match_components(p, comps):
            self.rebuild_platform_headers(p)
        ws = self.wb[p]
        comp_cols = self.component_col_map(p)
        # Avoid normalizing the whole platform sheet; target/insert merges are cleaned locally below.
        entry_blocks = self._contract_entry_blocks(ws, comp_cols=comp_cols)

        rows_data = []

        selected_components = {str(c) for s in systems for c in (s.components or {}).keys()}
        totals = {c: sum(as_number(s.components.get(c, 0)) for s in systems) for c in comps if c in selected_components}
        delivered_totals = {
            c: sum(
                sum(as_number(d.delivered.get(c, 0)) for d in deliveries_by_system.get(s.name, []))
                for s in systems
            )
            for c in totals.keys()
        }
        content = ", ".join(f"{int(v) if v == int(v) else v} {k}" for k, v in totals.items() if v)

        rows_data.append((
            [ci.no, ci.user, ci.yi_yd, ci.contract_type, "GENEL", "Ana S\u00f6zle\u015fme Toplam\u0131", content,
             ci.signature_date, ci.t0_date, ci.t0_months, ci.completion_date, ci.status, ci.acceptance_date or "", ci.note],
            {c: (totals[c], delivered_totals.get(c, 0), totals[c] - delivered_totals.get(c, 0)) for c in totals.keys()},
            None,
        ))

        system_spans_idx: List[Tuple[int, int]] = []
        for s in systems:
            sys_start_idx = len(rows_data)
            s_delivs = deliveries_by_system.get(s.name, [])
            s_component_keys = {str(c) for c in (s.components or {}).keys()}
            for d in s_delivs:
                s_component_keys.update(str(c) for c in (d.planned or {}).keys())
                s_component_keys.update(str(c) for c in (d.delivered or {}).keys())
            s_component_keys = {c for c in s_component_keys if c in comps}
            s_delivered = {c: sum(as_number(d.delivered.get(c, 0)) for d in s_delivs) for c in s_component_keys}
            s_content = ", ".join(f"{int(v) if v == int(v) else v} {k}" for k, v in s.components.items() if v)

            sys_t0 = s.t0_date or ci.t0_date
            sys_months = int(s.t0_months if s.t0_date else (ci.t0_months or 0))
            sys_completion = s.completion_date or (add_months(parse_iso_date(sys_t0), sys_months).isoformat() if parse_iso_date(sys_t0) else ci.completion_date)
            rows_data.append((
                [ci.no, ci.user, ci.yi_yd, ci.contract_type, s.name, f"{s.name} {SYSTEM_TOTAL_SUFFIX}", s_content,
                 ci.signature_date, sys_t0, sys_months, sys_completion, s.status or ci.status, s.acceptance_date or "", ""],
                {
                    c: (
                        as_number((s.components or {}).get(c, 0)),
                        s_delivered.get(c, 0),
                        as_number((s.components or {}).get(c, 0)) - s_delivered.get(c, 0),
                    )
                    for c in s_component_keys
                },
                s.name,
            ))

            for d in s_delivs:
                d_content = ", ".join(
                    f"{int(d.planned.get(c,0)) if d.planned.get(c,0)==int(d.planned.get(c,0)) else d.planned.get(c,0)} {c}"
                    for c in comps if d.planned.get(c, 0)
                )
                rows_data.append((
                    [ci.no, ci.user, ci.yi_yd, ci.contract_type, s.name, d.name, d_content,
                     ci.signature_date, sys_t0, sys_months, sys_completion, d.status, d.acceptance_date, d.note],
                    {
                        c: (
                            as_number((d.planned or {}).get(c, 0)),
                            as_number((d.delivered or {}).get(c, 0)),
                            as_number((d.planned or {}).get(c, 0)) - as_number((d.delivered or {}).get(c, 0)),
                        )
                        for c in comps
                        if c in (d.planned or {}) or c in (d.delivered or {})
                    },
                    s.name,
                ))
            system_spans_idx.append((sys_start_idx, len(rows_data) - 1))

        target_count = len(rows_data)
        existing_rows: List[int] = []
        start_row = 0
        end_row = 0
        existing_count = 0
        old_snapshot: Optional[dict] = None

        existing_block = None
        if old_start_row:
            existing_block = next((b for b in entry_blocks if int(b["start"]) == int(old_start_row)), None)

        if existing_block:
            start_row = int(existing_block["start"])
            end_row = int(existing_block["end"])
            existing_rows = list(range(start_row, end_row + 1))
            existing_count = len(existing_rows)
        else:
            norm_ci_type = self._normalize_label(str(ci.contract_type or ""))
            is_sd_entry = bool(re.match(r"^SD-\d+$", str(ci.contract_type or "").strip().upper()))
            if is_sd_entry:
                # ws parametresi ile güncellenmiş (fiziksel + mantıksal) hesaplama
                start_row = int(self._sd_insert_row_from_anchor(ci, ws, entry_blocks) or 0)
                if not start_row:
                    family = [b for b in entry_blocks if str(b["no"]) == str(ci.no).strip()]
                    if family:
                        logical_end = max(int(b["end"]) for b in family)
                        physical_end = self._physical_family_last_row(ws, str(ci.no).strip())
                        start_row = max(logical_end, physical_end) + 1
                    else:
                        start_row = self.next_row(p)
            else:
                same_no_blocks = [
                    b for b in entry_blocks
                    if str(b.get("no", "")).strip() in {str(old_contract_no or "").strip(), str(ci.no).strip()}
                ]
                same_type_blocks = [
                    b for b in same_no_blocks
                    if self._normalize_label(str(b.get("type", ""))) == norm_ci_type
                ]
                candidate = same_type_blocks[0] if same_type_blocks else None
                if candidate:
                    start_row = int(candidate["start"])
                    end_row = int(candidate["end"])
                    existing_rows = list(range(start_row, end_row + 1))
                    existing_count = len(existing_rows)
                else:
                    start_row = self.next_row(p)
            if not start_row:
                start_row = self.next_row(p)
            if existing_count == 0:
                end_row = start_row - 1

        if existing_count:
            prev_ci, prev_systems, prev_deliveries = self.load_contract_structure(
                p,
                str(old_contract_no or ci.no),
                start_row=start_row,
            )
            old_snapshot = self._contract_snapshot(prev_ci, prev_systems, prev_deliveries)

        if existing_count:
            for rng in list(ws.merged_cells.ranges):
                if rng.max_row < start_row or rng.min_row > end_row:
                    continue
                if rng.min_row >= DATA_START_ROW:
                    ws.unmerge_cells(str(rng))

        # SATIR EKLEME / SİLME — güvenli yardımcı ile yapılır.
        # _safe_insert_rows: ekleme noktasına yaslanan merge'leri ikiye böler,
        # böylece ne üst bloğun ne de alt bloğun merge'i bozulmaz.
        if existing_count < target_count:
            self._safe_insert_rows(ws, start_row + existing_count,
                                   target_count - existing_count)
        elif existing_count > target_count:
            # Silme öncesi: silinecek satır aralığındaki merge'leri temizle
            del_start = start_row + target_count
            del_end = start_row + existing_count - 1
            for rng in list(ws.merged_cells.ranges):
                if rng.min_row <= del_end and rng.max_row >= del_start:
                    if rng.min_row >= DATA_START_ROW:
                        ws.unmerge_cells(str(rng))
            ws.delete_rows(del_start, existing_count - target_count)

        # Yazma öncesi: hedef satır aralığındaki artık merge'leri temizle.
        # (insert_rows/delete_rows sonrası kalmış olabilir; çakışmayı önler.)
        write_end = start_row + target_count - 1
        for rng in list(ws.merged_cells.ranges):
            if rng.min_row >= DATA_START_ROW:
                if rng.min_row <= write_end and rng.max_row >= start_row:
                    ws.unmerge_cells(str(rng))

        for i, (base_vals, comp_vals, _sys_name) in enumerate(rows_data):
            r = start_row + i
            for c, v in enumerate(base_vals, 1):
                ws.cell(r, c, v)
            for c in range(len(BASE_HEADERS) + 1, ws.max_column + 1):
                ws.cell(r, c, None)
            for comp in comps:
                pc, dc, rc = comp_cols[comp]
                pv, dv, rv = comp_vals.get(comp, (None, None, None))
                ws.cell(r, pc, pv)
                ws.cell(r, dc, dv)
                ws.cell(r, rc, rv)

        end_row = start_row + target_count - 1
        if end_row >= start_row:
            _center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            for r in range(start_row, end_row + 1):
                for c in (1, 2, 3, 4, 5):
                    ws.cell(r, c).alignment = _center_align
            # _apply_visual_merges_for_block kaldırıldı — performans için merge yapılmaz

        used_components = [c for c, v in totals.items() if as_number(v) > 0]
        self.increment_component_usage(used_components)
        self._queue_platform_style(p, start_row, end_row)
        self._set_next_row_hint(p, ws.max_row + 1)
        self.save()
        new_snapshot = self._contract_snapshot(ci, systems, deliveries_by_system)
        reason = "Sözleşme güncellendi" if old_snapshot else "Sözleşme eklendi"
        self._append_log_rows(self._contract_diff_logs(old_snapshot, new_snapshot, reason))
        return int(start_row)


    def update_linked_sd_contract_numbers(
        self,
        platform: str,
        old_contract_no: str,
        new_contract_no: str,
        actor: Optional[str] = None,
    ) -> int:
        """Ana sözleşme no değiştiğinde aynı platformdaki bağlı SD bloklarının no alanını taşır."""
        p = safe_sheet_name(platform)
        old_no = str(old_contract_no or "").strip()
        new_no = str(new_contract_no or "").strip()
        if not p or not old_no or not new_no or old_no == new_no or p not in self.wb.sheetnames:
            return 0

        ws = self.wb[p]
        comp_cols = self.component_col_map(p)
        blocks = self._contract_entry_blocks(ws, comp_cols=comp_cols)

        def is_sd_type(text: str) -> bool:
            raw = str(text or "").strip()
            return bool(
                re.match(r"^SD-\d+$", raw.upper()) or
                self._normalize_label(raw) == self._normalize_label("Sözleşme Değişikliği")
            )

        linked_blocks = [
            b for b in blocks
            if str(b.get("no", "") or "").strip() == old_no
            and is_sd_type(str(b.get("type", "") or ""))
        ]
        if not linked_blocks:
            return 0

        linked_rows = {int(b.get("start") or 0) for b in linked_blocks}
        linked_types = {str(b.get("type", "") or "").strip() for b in linked_blocks}
        linked_type_keys = {self._normalize_label(t) for t in linked_types}
        for b in blocks:
            if int(b.get("start") or 0) in linked_rows:
                continue
            b_no = str(b.get("no", "") or "").strip()
            b_type = str(b.get("type", "") or "").strip()
            if b_no == new_no and self._normalize_label(b_type) in linked_type_keys:
                raise ValueError(
                    f"'{p}' platformunda '{new_no}' no ve '{b_type}' tipi için zaten kayıt var."
                )

        for b in linked_blocks:
            start = int(b["start"])
            ws.cell(start, 1, new_no)

        self.ensure_tag_sheet()
        tag_ws = self.wb[TAG_SHEET]
        changed_tags = False
        for r in range(2, tag_ws.max_row + 1):
            kind = str(tag_ws.cell(r, 1).value or "").strip().upper()
            if not self._is_tag_assign_kind(kind):
                continue
            rp = str(tag_ws.cell(r, 6).value or "").strip()
            rno = str(tag_ws.cell(r, 7).value or "").strip()
            rt = str(tag_ws.cell(r, 8).value or "").strip()
            if rp == p and rno == old_no and self._normalize_label(rt) in linked_type_keys:
                tag_ws.cell(r, 7, new_no)
                tag_ws.cell(r, 9, datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
                tag_ws.cell(r, 10, str(actor or self.current_actor()))
                changed_tags = True

        first = min(int(b["start"]) for b in linked_blocks)
        last = max(int(b["end"]) for b in linked_blocks)
        self.style_platform_rows_range(p, first, last)
        if changed_tags:
            self._normalize_tag_sheet_rows()
        self.save()
        return len(linked_blocks)

    def list_main_contracts(self, platform: str, tags_map: Optional[Dict[Tuple[str, str, str], List[str]]] = None) -> List[dict]:
        p = safe_sheet_name(platform)
        if p not in self.wb.sheetnames:
            return []
        ws = self.wb[p]
        if tags_map is None:
            tags_map = self.all_contract_tags_map()
        rows = []
        for r, row_data in enumerate(
            ws.iter_rows(min_row=DATA_START_ROW, max_col=len(BASE_HEADERS), values_only=True),
            start=DATA_START_ROW,
        ):
            def val(col: int, default=None):
                return row_data[col - 1] if col - 1 < len(row_data) else default

            activity = str(val(5) or "").strip()
            delivery = str(val(6) or "").strip()
            if activity.upper() != "GENEL" or not self._is_main_total_row(delivery):
                continue

            ctype = str(val(4) or "").strip()
            is_main = self._normalize_label(ctype) == "ana sozlesme"
            no = str(val(1) or "")
            if not no.strip():
                continue
            tags = tags_map.get((p, no.strip(), ctype.strip()), [])
            item = {
                "row": r,
                "platform": p,
                "no": no,
                "user": val(2),
                "type": ctype,
                "type_display": ctype if is_main else f"\u21b3 {ctype}",
                "link": "Ana S\u00f6zle\u015fme" if is_main else "Ana s\u00f6zle\u015fmeye ba\u011fl\u0131 SD",
                "status": val(12),
                "completion_date": val(11),
                "acceptance_date": val(13),
                "content": val(7),
                "is_main": is_main,
                "tags": tags,
            }
            item["search"] = " ".join(
                str(item.get(k, "") or "")
                for k in ["platform", "no", "user", "type", "link", "status", "completion_date", "content"]
            ).lower()
            if tags:
                item["search"] = (item["search"] + " " + " ".join(tags).lower()).strip()
            rows.append(item)
        return rows

    def all_contract_tags_map(self) -> Dict[Tuple[str, str, str], List[str]]:
        if TAG_SHEET not in self.wb.sheetnames:
            return {}
        ws = self.wb[TAG_SHEET]
        out: Dict[Tuple[str, str, str], List[str]] = {}
        data = self._get_sheet_data(ws)

        def val(row_data, col: int, default=None):
            return row_data[col - 1] if col - 1 < len(row_data) else default

        for _r_idx, row_data in enumerate(data[1:], start=2):
            if not self._is_tag_assign_kind(str(val(row_data, 1) or "").strip().upper()):
                continue
            nm = str(val(row_data, 2) or "").strip()
            p = str(val(row_data, 6) or "").strip()
            no = str(val(row_data, 7) or "").strip()
            ctype = str(val(row_data, 8) or "").strip()
            if not nm or not p or not no:
                continue
            key = (p, no, ctype)
            out.setdefault(key, []).append(nm)
        return out

    def build_contract_index(self, progress_cb=None) -> List[dict]:
        """Tüm platformlardaki sözleşme bloklarını (Ana + SD) tek seferde indeksler.
        Arama ekranı bu hazır listeyi filtreler; her tuşta Excel tekrar okunmaz.
        """
        def emit(p: int, msg: str):
            if progress_cb:
                try:
                    progress_cb(int(max(0, min(100, p))), str(msg))
                except Exception:
                    pass

        index = []
        platforms = self.platform_names()
        tags_map = self.all_contract_tags_map()
        total = len(platforms)
        emit(0, "Platform listesi hazırlanıyor...")
        if total == 0:
            emit(100, "İndeks hazır")
            return index
        for i, platform in enumerate(platforms, start=1):
            base = int(((i - 1) / total) * 100)
            emit(base, f"İndeksleniyor: {platform} ({i}/{total})")
            index.extend(self.list_main_contracts(platform, tags_map=tags_map))
            done = int((i / total) * 100)
            emit(done, f"İndeksleniyor: {platform} ({i}/{total})")
        return index

    def load_contract_structure(
        self,
        platform: str,
        contract_no: str,
        start_row: Optional[int] = None,
    ) -> Tuple[Optional[ContractInfo], List[SystemInfo], Dict[str, List[DeliveryInfo]]]:
        """Excel platform sayfasından tek bir sözleşme bloğunu (Ana/SD) sistem-kabul yapısıyla okur."""
        p = safe_sheet_name(platform)
        if p not in self.wb.sheetnames:
            return None, [], {}
        ws = self.wb[p]
        comps = self.assigned_components(p)
        comp_cols = self.component_col_map(p)
        ci = None
        systems: List[SystemInfo] = []
        deliveries: Dict[str, List[DeliveryInfo]] = {}
        current_system = None

        blocks = self._contract_entry_blocks(ws, comp_cols=comp_cols)
        target_block = None
        if start_row is not None:
            target_block = next((b for b in blocks if int(b["start"]) == int(start_row)), None)
        if target_block is None:
            target_no = str(contract_no or "").strip()
            matched = [b for b in blocks if str(b["no"]) == target_no]
            if matched:
                main_first = [b for b in matched if self._normalize_label(str(b.get("type", ""))) == self._normalize_label("Ana Sözleşme")]
                target_block = (main_first[0] if main_first else matched[0])

        if not target_block:
            return None, [], {}
        block_rows = list(range(int(target_block["start"]), int(target_block["end"]) + 1))

        for r in block_rows:
            delivery_name = str(self.cell_value(ws, r, 6) or "").strip()
            activity = str(self.cell_value(ws, r, 5) or "").strip()

            if ci is None:
                ci = ContractInfo(
                    no=str(self.cell_value(ws, r, 1) or ""),
                    platform=p,
                    user=str(self.cell_value(ws, r, 2) or ""),
                    yi_yd=str(self.cell_value(ws, r, 3) or ""),
                    contract_type=str(self.cell_value(ws, r, 4) or ""),
                    signature_date=str(self.cell_value(ws, r, 8) or ""),
                    t0_date=str(self.cell_value(ws, r, 9) or ""),
                    t0_months=int(as_number(self.cell_value(ws, r, 10))),
                    completion_date=str(self.cell_value(ws, r, 11) or ""),
                    status=str(self.cell_value(ws, r, 12) or "PLAN"),
                    note=str(self.cell_value(ws, r, 14) or ""),
                    acceptance_date=str(self.cell_value(ws, r, 13) or ""),
                    entry_start_row=int(target_block["start"]),
                )

            if self._is_main_total_row(delivery_name):
                continue

            if self._is_system_total_row(delivery_name):
                sys_name = self._extract_system_name_from_total(
                    delivery_name, activity, len(systems) + 1
                )
                sys_comps = {}
                for c in comps:
                    pc, dc, rc = comp_cols[c]
                    val = as_number(self.cell_value(ws, r, pc))
                    if val:  # Sadece deger > 0 olanlari ekle
                        sys_comps[c] = val
                si = SystemInfo(
                    sys_name,
                    sys_comps,
                    t0_date=str(self.cell_value(ws, r, 9) or ""),
                    t0_months=int(as_number(self.cell_value(ws, r, 10))),
                    completion_date=str(self.cell_value(ws, r, 11) or ""),
                    status=str(self.cell_value(ws, r, 12) or "Başlanmadı"),
                    acceptance_date=str(self.cell_value(ws, r, 13) or ""),
                )
                systems.append(si)
                deliveries[si.name] = []
                current_system = si.name
                continue

            sys_name = activity or current_system
            if not sys_name:
                continue
            if sys_name not in deliveries:
                si = SystemInfo(sys_name, {})
                systems.append(si)
                deliveries[sys_name] = []
            sys_obj = next((s for s in systems if s.name == sys_name), None)
            selected_keys = set((sys_obj.components or {}).keys()) if sys_obj else set()
            planned = {}
            delivered = {}
            for c in comps:
                pc, dc, rc = comp_cols[c]
                raw_planned = self.cell_value(ws, r, pc)
                raw_delivered = self.cell_value(ws, r, dc)
                planned_val = as_number(raw_planned)
                delivered_val = as_number(raw_delivered)
                if c in selected_keys or planned_val or delivered_val:
                    planned[c] = planned_val
                    delivered[c] = delivered_val
            deliveries[sys_name].append(DeliveryInfo(
                name=delivery_name,
                status=str(self.cell_value(ws, r, 12) or "PLAN"),
                acceptance_date=str(self.cell_value(ws, r, 13) or ""),
                note=str(self.cell_value(ws, r, 14) or ""),
                planned=planned,
                delivered=delivered,
                t0_date=str(self.cell_value(ws, r, 9) or ""),
                t0_months=int(as_number(self.cell_value(ws, r, 10))),
                completion_date=str(self.cell_value(ws, r, 11) or ""),
            ))
        return ci, systems, deliveries


# ---------- Dialog helpers ----------


# ---------- Sistem Tipi / Bileşen Paketi helpers ----------
# Bu bölüm Sistem Ekle penceresindeki "Sistem Tipi" hızlı seçimlerini Excel'de saklar.
# Veri sayfası: SistemTipleri

SYSTEM_TYPE_SHEET = "SistemTipleri"
SYSTEM_TYPE_HEADERS = ["Tip Adı", "Platform", "Bileşen Adı", "Aktif", "Güncelleme", "Adet"]
_SYSTEM_TYPE_INACTIVE_VALUES = {"hayır", "hayir", "no", "false", "0", "pasif"}


def _ensure_system_type_sheet(self, persist: bool = False):
    """SistemTipleri sayfasını oluşturur ve başlıklarını hazırlar."""
    if SYSTEM_TYPE_SHEET in self.wb.sheetnames:
        ws = self.wb[SYSTEM_TYPE_SHEET]
    else:
        ws = self.wb.create_sheet(SYSTEM_TYPE_SHEET)
        persist = True

    changed = False
    for c, h in enumerate(SYSTEM_TYPE_HEADERS, 1):
        if ws.cell(1, c).value != h:
            ws.cell(1, c, h)
            changed = True

    fill = PatternFill("solid", fgColor=NAVY)
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    widths = [28, 18, 38, 10, 20, 12]
    for c, w in enumerate(widths, 1):
        cell = ws.cell(1, c)
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"

    if persist or changed:
        self.save()
    return ws


def _system_type_platform_key(platform: str) -> str:
    return safe_sheet_name(str(platform or "")) if str(platform or "").strip() else ""


def list_system_type_names(self, platform: str = "") -> List[str]:
    """Platforma ait sistem tipi adlarını döndürür. Platformu boş olan tipler geneldir."""
    if SYSTEM_TYPE_SHEET not in self.wb.sheetnames:
        return []
    ws = _ensure_system_type_sheet(self)
    p = _system_type_platform_key(platform)
    names: List[str] = []
    seen = set()
    for r in range(2, ws.max_row + 1):
        name = str(ws.cell(r, 1).value or "").strip()
        row_platform = str(ws.cell(r, 2).value or "").strip()
        active = str(ws.cell(r, 4).value or "Evet").strip().lower()
        if not name or active in _SYSTEM_TYPE_INACTIVE_VALUES:
            continue
        if row_platform and p and row_platform != p:
            continue
        key = name.lower()
        if key not in seen:
            seen.add(key)
            names.append(name)
    return sorted(names, key=lambda x: x.lower())


def get_system_type_components(self, type_name: str, platform: str = "") -> List[str]:
    """Seçili tipteki aktif bileşenleri döndürür."""
    if SYSTEM_TYPE_SHEET not in self.wb.sheetnames:
        return []
    ws = _ensure_system_type_sheet(self)
    target = str(type_name or "").strip().lower()
    p = _system_type_platform_key(platform)
    comps: List[str] = []
    seen = set()
    if not target:
        return comps
    for r in range(2, ws.max_row + 1):
        name = str(ws.cell(r, 1).value or "").strip()
        row_platform = str(ws.cell(r, 2).value or "").strip()
        comp = str(ws.cell(r, 3).value or "").strip()
        active = str(ws.cell(r, 4).value or "Evet").strip().lower()
        if name.lower() != target or not comp:
            continue
        if active in _SYSTEM_TYPE_INACTIVE_VALUES:
            continue
        if row_platform and p and row_platform != p:
            continue
        key = comp.lower()
        if key not in seen:
            seen.add(key)
            comps.append(comp)
    return comps


def get_system_type_component_quantities(self, type_name: str, platform: str = "") -> Dict[str, float]:
    """Seçili tipteki aktif bileşenleri adetleriyle döndürür. Eski kayıtlarda adet yoksa 1 kabul edilir."""
    if SYSTEM_TYPE_SHEET not in self.wb.sheetnames:
        return {}
    ws = _ensure_system_type_sheet(self)
    target = str(type_name or "").strip().lower()
    p = _system_type_platform_key(platform)
    out: Dict[str, float] = {}
    if not target:
        return out
    for r in range(2, ws.max_row + 1):
        name = str(ws.cell(r, 1).value or "").strip()
        row_platform = str(ws.cell(r, 2).value or "").strip()
        comp = str(ws.cell(r, 3).value or "").strip()
        active = str(ws.cell(r, 4).value or "Evet").strip().lower()
        if name.lower() != target or not comp:
            continue
        if active in _SYSTEM_TYPE_INACTIVE_VALUES:
            continue
        if row_platform and p and row_platform != p:
            continue
        qty = as_number(ws.cell(r, 6).value)
        out[comp] = qty if qty > 0 else 1
    return out


def save_system_type(self, type_name: str, platform: str, components) -> int:
    """
    Seçili bileşenleri SistemTipleri sayfasına yazar.
    Aynı Tip Adı + Platform varsa eski satırları tamamen yeniler.
    Dönüş: kaydedilen bileşen sayısı.
    """
    type_name = str(type_name or "").strip()
    if not type_name:
        raise ValueError("Tip adı boş olamaz.")

    comps: List[Tuple[str, float]] = []
    seen = set()
    source = (components or {}).items() if isinstance(components, dict) else [(c, 1) for c in (components or [])]
    for c, qty in source:
        comp = str(c or "").strip()
        key = comp.lower()
        amount = as_number(qty)
        if comp and key not in seen and amount > 0:
            seen.add(key)
            comps.append((comp, amount))
    if not comps:
        raise ValueError("Kaydedilecek bileşen yok.")

    ws = _ensure_system_type_sheet(self)
    p = _system_type_platform_key(platform)
    target = type_name.lower()

    # Aynı tip + platform için eski satırları sil, sonra yeni seçimi yaz.
    for r in range(ws.max_row, 1, -1):
        name = str(ws.cell(r, 1).value or "").strip().lower()
        row_platform = str(ws.cell(r, 2).value or "").strip()
        if name == target and row_platform == p:
            ws.delete_rows(r, 1)

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    start_row = max(ws.max_row + 1, 2)
    for i, (comp, qty) in enumerate(comps):
        row = start_row + i
        ws.cell(row, 1, type_name)
        ws.cell(row, 2, p)
        ws.cell(row, 3, comp)
        ws.cell(row, 4, "Evet")
        ws.cell(row, 5, now)
        ws.cell(row, 6, qty)

    self.save()
    return len(comps)


ExcelStore.ensure_system_type_sheet = _ensure_system_type_sheet
ExcelStore.list_system_type_names = list_system_type_names
ExcelStore.get_system_type_components = get_system_type_components
ExcelStore.get_system_type_component_quantities = get_system_type_component_quantities
ExcelStore.save_system_type = save_system_type
