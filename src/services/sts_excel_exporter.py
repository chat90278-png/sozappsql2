from __future__ import annotations

import time
from pathlib import Path


BASE_HEADERS = [
    "Sözleşme No",
    "Sözleşme Türü",
    "Sistem Adı",
    "Kabul Adı",
    "Kullanıcı",
    "Yİ/YD",
    "Durum",
    "İmza Tarihi",
    "T0 Tarihi",
    "T0 Ay",
    "Termin Tarihi",
    "Kabul Tarihi",
    "Etiketler",
    "Not",
]

STATUS_STYLES = {
    "tamamlandı": ("C6EFCE", "276221"),
    "tamamlandi": ("C6EFCE", "276221"),
    "devam ediyor": ("DDEEFF", "1F4E79"),
    "gecikti": ("FFD7D7", "9C0006"),
    "başlanmadı": ("F2F2F2", "595959"),
    "baslanmadi": ("F2F2F2", "595959"),
    "plan": ("F2F2F2", "595959"),
}


def _safe_sheet_name(name: str, used: set[str]) -> str:
    bad = ':\\/?*[]'
    n = ''.join('_' if ch in bad else ch for ch in str(name or '').strip())[:31] or 'Platform'
    base = n
    i = 2
    while n in used:
        suf = f'_{i}'
        n = (base[:31-len(suf)] + suf)
        i += 1
    used.add(n)
    return n


def _norm_status(value: str) -> str:
    text = str(value or "").strip().lower()
    return (
        text.replace("ı", "i")
        .replace("İ", "i")
        .replace("ş", "s")
        .replace("ğ", "g")
        .replace("ü", "u")
        .replace("ö", "o")
        .replace("ç", "c")
    )


def _display_status(value: str) -> str:
    text = str(value or "").strip()
    if not text or text.upper() == "PLAN":
        return "Başlanmadı"
    return text


def _number(value) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def _component_names_for_platform(conn, platform_id: int) -> list[str]:
    assignment_count = conn.execute(
        "SELECT COUNT(*) FROM component_platforms WHERE platform_id=?",
        (platform_id,),
    ).fetchone()[0]
    if assignment_count:
        rows = conn.execute(
            """
            SELECT c.name
            FROM component_platforms cp
            JOIN components c ON c.id=cp.component_id
            WHERE cp.platform_id=? AND cp.enabled=1 AND c.active=1
            ORDER BY c.name COLLATE NOCASE, c.name
            """,
            (platform_id,),
        ).fetchall()
    else:
        rows = conn.execute(
            "SELECT name FROM components WHERE active=1 ORDER BY name COLLATE NOCASE, name"
        ).fetchall()
    return [str(row[0]) for row in rows]


def _tag_text(conn, contract_id: int) -> str:
    rows = conn.execute(
        """
        SELECT t.name
        FROM contract_tags ct
        JOIN tags t ON t.id=ct.tag_id
        WHERE ct.contract_id=?
        ORDER BY t.name COLLATE NOCASE, t.name
        """,
        (contract_id,),
    ).fetchall()
    return ", ".join(str(row[0]) for row in rows)


def _contract_user_text(conn, contract_id: int) -> str:
    rows = conn.execute(
        """
        SELECT u.name
        FROM contract_users cu
        JOIN users u ON u.id=cu.user_id
        WHERE cu.contract_id=?
        ORDER BY u.name COLLATE NOCASE, u.name
        """,
        (contract_id,),
    ).fetchall()
    return ", ".join(str(row[0]) for row in rows)


def _write_summary_sheet(wb, conn, db, platforms: list[str], all_platforms: list[str], comp_count: int):
    summary = wb.create_sheet('Özet')
    c_count = conn.execute('SELECT COUNT(*) FROM contracts').fetchone()[0]
    s_count = conn.execute('SELECT COUNT(*) FROM systems').fetchone()[0]
    d_count = conn.execute('SELECT COUNT(*) FROM deliveries').fetchone()[0]
    l_count = conn.execute('SELECT COUNT(*) FROM activity_logs').fetchone()[0]
    summary.append(['Öğe', 'Değer'])
    summary.append(['Oluşturma zamanı', time.strftime('%Y-%m-%d %H:%M:%S')])
    summary.append(['Kaynak STS', str(getattr(db, 'path', ''))])
    summary.append(['Platform sayısı', len(platforms) if platforms else len(all_platforms)])
    summary.append(['Sözleşme sayısı', c_count])
    summary.append(['Sistem sayısı', s_count])
    summary.append(['Kabul/Teslimat sayısı', d_count])
    summary.append(['Bileşen sayısı', comp_count])
    summary.append(['Log sayısı', l_count])
    summary.append([])
    summary.append(['Platform', 'Sözleşme Sayısı', 'Sistem Sayısı', 'Teslimat Sayısı'])
    for platform in (platforms if platforms else all_platforms):
        p_contracts = conn.execute('SELECT COUNT(*) FROM contracts c JOIN platforms p ON p.id=c.platform_id WHERE p.name=?', (platform,)).fetchone()[0]
        p_systems = conn.execute('SELECT COUNT(*) FROM systems s JOIN contracts c ON c.id=s.contract_id JOIN platforms p ON p.id=c.platform_id WHERE p.name=?', (platform,)).fetchone()[0]
        p_delivs = conn.execute('SELECT COUNT(*) FROM deliveries d JOIN contracts c ON c.id=d.contract_id JOIN platforms p ON p.id=c.platform_id WHERE p.name=?', (platform,)).fetchone()[0]
        summary.append([platform, p_contracts, p_systems, p_delivs])


def _apply_platform_formatting(ws, max_row: int, max_col: int, component_start_col: int, row_groups: list[dict]):
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor="0D2B55")
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    zebra_fill = PatternFill("solid", fgColor="F7FAFF")
    remaining_fill = PatternFill("solid", fgColor="FFF2CC")
    thin = Side(style="thin", color="D9E2EF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.row_dimensions[1].height = 22
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row_idx in range(2, max_row + 1):
        fill = white_fill if row_idx % 2 else zebra_fill
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row_idx, col_idx)
            cell.fill = fill
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = border

        status_cell = ws.cell(row_idx, 7)
        key = _norm_status(status_cell.value)
        status_colors = STATUS_STYLES.get(key)
        if status_colors:
            status_cell.fill = PatternFill("solid", fgColor=status_colors[0])
            status_cell.font = Font(color=status_colors[1], bold=True)

        for col_idx in range(component_start_col + 2, max_col + 1, 3):
            cell = ws.cell(row_idx, col_idx)
            if _number(cell.value) > 0:
                cell.fill = remaining_fill
            else:
                cell.fill = white_fill

    def merge_ranges(column: int, key_name: str):
        start = 2
        last_key = None
        for offset, group in enumerate(row_groups, start=2):
            key = group[key_name]
            if last_key is None:
                last_key = key
                start = offset
            elif key != last_key:
                if offset - 1 > start:
                    ws.merge_cells(start_row=start, start_column=column, end_row=offset - 1, end_column=column)
                    ws.cell(start, column).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                start = offset
                last_key = key
        if last_key is not None and max_row > start:
            ws.merge_cells(start_row=start, start_column=column, end_row=max_row, end_column=column)
            ws.cell(start, column).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    merge_ranges(1, "contract_id")
    merge_ranges(2, "contract_id")
    merge_ranges(3, "system_id")

    ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"
    ws.freeze_panes = "A2"

    for col_idx in range(1, max_col + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[letter]:
            value = cell.value
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 45)


def _component_totals(rows) -> dict[str, tuple[float, float]]:
    totals: dict[str, list[float]] = {}
    for name, planned, delivered in rows:
        key = str(name)
        current = totals.setdefault(key, [0.0, 0.0])
        current[0] += _number(planned)
        current[1] += _number(delivered)
    return {key: (value[0], value[1]) for key, value in totals.items()}


def _system_component_totals(conn, system_id: int) -> dict[str, tuple[float, float]]:
    planned_rows = conn.execute(
        """
        SELECT c.name, sc.qty
        FROM system_components sc
        JOIN components c ON c.id=sc.component_id
        WHERE sc.system_id=?
        """,
        (system_id,),
    ).fetchall()
    delivered_rows = conn.execute(
        """
        SELECT c.name, SUM(dc.delivered)
        FROM deliveries d
        JOIN delivery_components dc ON dc.delivery_id=d.id
        JOIN components c ON c.id=dc.component_id
        WHERE d.system_id=?
        GROUP BY c.name
        """,
        (system_id,),
    ).fetchall()
    totals: dict[str, list[float]] = {}
    for name, planned in planned_rows:
        totals.setdefault(str(name), [0.0, 0.0])[0] += _number(planned)
    for name, delivered in delivered_rows:
        totals.setdefault(str(name), [0.0, 0.0])[1] += _number(delivered)
    return {key: (value[0], value[1]) for key, value in totals.items()}


def _contract_component_totals(conn, contract_id: int) -> dict[str, tuple[float, float]]:
    planned_rows = conn.execute(
        """
        SELECT c.name, SUM(sc.qty)
        FROM systems s
        JOIN system_components sc ON sc.system_id=s.id
        JOIN components c ON c.id=sc.component_id
        WHERE s.contract_id=?
        GROUP BY c.name
        """,
        (contract_id,),
    ).fetchall()
    delivered_rows = conn.execute(
        """
        SELECT c.name, SUM(dc.delivered)
        FROM deliveries d
        JOIN delivery_components dc ON dc.delivery_id=d.id
        JOIN components c ON c.id=dc.component_id
        WHERE d.contract_id=?
        GROUP BY c.name
        """,
        (contract_id,),
    ).fetchall()
    totals: dict[str, list[float]] = {}
    for name, planned in planned_rows:
        totals.setdefault(str(name), [0.0, 0.0])[0] += _number(planned)
    for name, delivered in delivered_rows:
        totals.setdefault(str(name), [0.0, 0.0])[1] += _number(delivered)
    return {key: (value[0], value[1]) for key, value in totals.items()}


def _delivery_component_totals(conn, delivery_id: int) -> dict[str, tuple[float, float]]:
    return _component_totals(
        conn.execute(
            """
            SELECT c.name, dc.planned, dc.delivered
            FROM delivery_components dc
            JOIN components c ON c.id=dc.component_id
            WHERE dc.delivery_id=?
            """,
            (delivery_id,),
        ).fetchall()
    )


def _append_export_row(ws, base_values: list, components: list[str], component_values: dict[str, tuple[float, float]]):
    row = list(base_values)
    for component in components:
        planned, delivered = component_values.get(component, (0.0, 0.0))
        row.extend([planned, delivered, planned - delivered])
    ws.append(row)


def export_sts_to_excel(db, output_path, options=None, progress_cb=None):
    opts = {
        "scope": "all", "platforms": None, "include_summary": True,
        "include_contract_rows": True, "include_system_rows": True,
        "include_delivery_rows": True, "include_component_columns": True,
        "include_tags": True,
    }
    if isinstance(options, dict):
        opts.update(options)
    if progress_cb:
        progress_cb(0, "Excel oluşturma başlatılıyor...")
    try:
        from openpyxl import Workbook
    except Exception as exc:
        raise RuntimeError('Excel aktarımı için openpyxl kurulu olmalıdır.') from exc

    t0 = time.time()
    conn = db.conn
    out = Path(output_path)
    wb = Workbook(write_only=False)
    wb.remove(wb.active)

    platform_rows = conn.execute(
        "SELECT id,name FROM platforms WHERE is_active=1 ORDER BY sort_order,name"
    ).fetchall()
    all_platforms = [row[1] for row in platform_rows]
    platform_ids = {row[1]: int(row[0]) for row in platform_rows}
    scope = str(opts.get('scope') or 'all')
    if scope == 'selected':
        platforms = [p for p in (opts.get('platforms') or []) if p in all_platforms]
    elif scope == 'active':
        platforms = [p for p in (opts.get('platforms') or []) if p in all_platforms][:1]
    elif scope == 'summary_only':
        platforms = []
        opts['include_summary'] = True
    else:
        platforms = list(all_platforms)

    active_component_count = conn.execute("SELECT COUNT(*) FROM components WHERE active=1").fetchone()[0]
    if opts.get('include_summary', True):
        if progress_cb:
            progress_cb(5, "Özet hazırlanıyor...")
        _write_summary_sheet(wb, conn, db, platforms, all_platforms, active_component_count)

    if progress_cb:
        progress_cb(10, "Platformlar hazırlanıyor...")
    used = {'Özet'} if 'Özet' in wb.sheetnames else set()

    for platform_index, platform in enumerate(platforms):
        if progress_cb:
            pct = 10 + int((platform_index / max(1, len(platforms))) * 80)
            progress_cb(pct, f"{platform} platformu yazılıyor...")

        platform_id = platform_ids[platform]
        components = _component_names_for_platform(conn, platform_id) if opts.get('include_component_columns', True) else []
        headers = list(BASE_HEADERS)
        if not opts.get('include_tags', True):
            headers[12] = "Etiketler"
        for component in components:
            headers.extend([
                f"{component} Teslim Edilecek",
                f"{component} Teslim Edilen",
                f"{component} Kalan",
            ])

        ws = wb.create_sheet(_safe_sheet_name(platform, used))
        ws.append(headers)
        row_groups = []

        contracts = conn.execute(
            """
            SELECT
                c.id,
                c.contract_no,
                c.contract_type,
                c.yi_yd,
                c.status,
                c.signed_date,
                c.t0_date,
                c.t0_months,
                c.completion_date,
                c.acceptance_date,
                c.note
            FROM contracts c
            JOIN platforms p ON p.id=c.platform_id
            WHERE p.name=?
            ORDER BY c.contract_no COLLATE NOCASE, c.contract_type COLLATE NOCASE, c.id
            """,
            (platform,),
        ).fetchall()

        for contract in contracts:
            contract_id = int(contract[0])
            contract_users = _contract_user_text(conn, contract_id)
            tag_txt = _tag_text(conn, contract_id) if opts.get('include_tags', True) else ""
            contract_base = [
                contract[1] or "",
                contract[2] or "",
                "GENEL",
                "Sözleşme Toplamı",
                contract_users,
                contract[3] or "",
                _display_status(contract[4]),
                contract[5] or "",
                contract[6] or "",
                contract[7] or 0,
                contract[8] or "",
                contract[9] or "",
                tag_txt,
                contract[10] or "",
            ]
            if opts.get('include_contract_rows', True):
                _append_export_row(
                    ws,
                    contract_base,
                    components,
                    _contract_component_totals(conn, contract_id),
                )
                row_groups.append({"contract_id": contract_id, "system_id": f"contract:{contract_id}:summary"})

            systems = conn.execute(
                """
                SELECT id, name, status, completion_date, acceptance_date, note
                FROM systems
                WHERE contract_id=?
                ORDER BY sort_order, id
                """,
                (contract_id,),
            ).fetchall()
            for system in systems:
                system_id = int(system[0])
                deliveries = conn.execute(
                    """
                    SELECT
                        d.id,
                        d.name,
                        d.status,
                        d.acceptance_date,
                        d.note,
                        u.name AS delivery_user
                    FROM deliveries d
                    LEFT JOIN users u ON u.id=d.delivery_user_id
                    WHERE d.contract_id=? AND d.system_id=?
                    ORDER BY d.sort_order, d.id
                    """,
                    (contract_id, system_id),
                ).fetchall()

                if deliveries:
                    if not opts.get('include_delivery_rows', True):
                        continue
                    for delivery in deliveries:
                        user_txt = str(delivery[5] or "").strip() or contract_users
                        row_base = [
                            contract[1] or "",
                            contract[2] or "",
                            system[1] or "",
                            delivery[1] or "",
                            user_txt,
                            contract[3] or "",
                            _display_status(delivery[2] or contract[4]),
                            contract[5] or "",
                            contract[6] or "",
                            contract[7] or 0,
                            contract[8] or "",
                            delivery[3] or contract[9] or "",
                            tag_txt,
                            delivery[4] or contract[10] or "",
                        ]
                        _append_export_row(
                            ws,
                            row_base,
                            components,
                            _delivery_component_totals(conn, int(delivery[0])),
                        )
                        row_groups.append({"contract_id": contract_id, "system_id": system_id})
                elif opts.get('include_system_rows', True):
                    row_base = [
                        contract[1] or "",
                        contract[2] or "",
                        system[1] or "",
                        "Sistem Toplamı",
                        contract_users,
                        contract[3] or "",
                        _display_status(system[2] or contract[4]),
                        contract[5] or "",
                        contract[6] or "",
                        contract[7] or 0,
                        system[3] or contract[8] or "",
                        system[4] or contract[9] or "",
                        tag_txt,
                        system[5] or contract[10] or "",
                    ]
                    _append_export_row(
                        ws,
                        row_base,
                        components,
                        _system_component_totals(conn, system_id),
                    )
                    row_groups.append({"contract_id": contract_id, "system_id": system_id})

        max_row = max(ws.max_row, 1)
        max_col = max(ws.max_column, len(headers))
        _apply_platform_formatting(ws, max_row, max_col, len(BASE_HEADERS) + 1, row_groups)

    if not wb.sheetnames:
        wb.create_sheet('Özet')

    if progress_cb:
        progress_cb(95, "Excel dosyası kaydediliyor...")
    wb.save(str(out))
    duration = round(time.time() - t0, 3)
    if progress_cb:
        progress_cb(100, "Excel dosyası oluşturuldu.")
    c_count = conn.execute('SELECT COUNT(*) FROM contracts').fetchone()[0]
    return {'output_path': str(out), 'platform_count': len(platforms), 'contract_count': c_count, 'duration_sec': duration}
