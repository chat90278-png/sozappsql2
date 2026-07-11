from __future__ import annotations

from pathlib import Path
from typing import Any, Sequence

from PySide6.QtCore import QEvent, QPoint, QRectF, Qt, QTimer, Signal
from PySide6.QtGui import QColor, QKeySequence, QPainter, QPen
from PySide6.QtWidgets import (
    QAbstractItemDelegate,
    QAbstractItemView,
    QApplication,
    QButtonGroup,
    QCheckBox,
    QComboBox,
    QDialog,
    QFileDialog,
    QFrame,
    QGraphicsDropShadowEffect,
    QGridLayout,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QLineEdit,
    QPushButton,
    QSizePolicy,
    QStackedWidget,
    QStyledItemDelegate,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)

from src.domain.component_bulk_import import (
    SUPPORTED_UNITS,
    BulkComponentRow,
    WorkbookMapping,
    component_key,
    detect_workbook_mapping,
    list_workbook_sheets,
    load_workbook_matrix,
    merge_new_components,
    normalize_space,
    rows_from_matrix,
    validate_bulk_rows,
)
from src.ui.message_boxes import ask_yes_no, show_information, show_warning


COL_NUMBER = 0
COL_NAME = 1
COL_UNIT = 2
COL_NOTE = 3
COL_STATUS = 4
COL_CONTROL = 5
COL_REMOVE = 6


class SlideToggle(QWidget):
    toggled = Signal(bool)

    def __init__(self, checked: bool = True, parent=None):
        super().__init__(parent)
        self._checked = bool(checked)
        self.setFixedSize(40, 23)
        self.setCursor(Qt.PointingHandCursor)

    def isChecked(self) -> bool:
        return self._checked

    def setChecked(self, checked: bool) -> None:
        checked = bool(checked)
        if self._checked == checked:
            return
        self._checked = checked
        self.update()
        self.toggled.emit(checked)

    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton:
            self.setChecked(not self._checked)
            event.accept()
            return
        super().mousePressEvent(event)

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing)
        painter.setPen(Qt.NoPen)
        painter.setBrush(QColor("#3B6FE8" if self._checked else "#CBD5E1"))
        painter.drawRoundedRect(QRectF(0, 1.5, 40, 20), 10, 10)
        thumb_x = 20.0 if self._checked else 2.0
        painter.setBrush(QColor("#FFFFFF"))
        painter.drawEllipse(QRectF(thumb_x, 3.5, 16, 16))
        painter.end()


class DragHeader(QFrame):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._drag_offset: QPoint | None = None

    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton:
            self._drag_offset = event.globalPosition().toPoint() - self.window().frameGeometry().topLeft()
            event.accept()
            return
        super().mousePressEvent(event)

    def mouseMoveEvent(self, event):
        if self._drag_offset is not None and event.buttons() & Qt.LeftButton:
            self.window().move(event.globalPosition().toPoint() - self._drag_offset)
            event.accept()
            return
        super().mouseMoveEvent(event)

    def mouseReleaseEvent(self, event):
        self._drag_offset = None
        super().mouseReleaseEvent(event)


class DropZone(QFrame):
    fileDropped = Signal(str)

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setObjectName("bulkDropZone")
        self.setAcceptDrops(True)
        self.setMinimumHeight(220)

    @staticmethod
    def _path_from_event(event) -> str:
        urls = event.mimeData().urls() if event.mimeData().hasUrls() else []
        for url in urls:
            path = url.toLocalFile()
            if path and Path(path).suffix.casefold() == ".xlsx":
                return path
        return ""

    def dragEnterEvent(self, event):
        if self._path_from_event(event):
            self.setProperty("dragActive", True)
            self.style().unpolish(self)
            self.style().polish(self)
            event.acceptProposedAction()
            return
        event.ignore()

    def dragMoveEvent(self, event):
        if self._path_from_event(event):
            event.acceptProposedAction()
            return
        event.ignore()

    def dragLeaveEvent(self, event):
        self.setProperty("dragActive", False)
        self.style().unpolish(self)
        self.style().polish(self)
        super().dragLeaveEvent(event)

    def dropEvent(self, event):
        path = self._path_from_event(event)
        self.setProperty("dragActive", False)
        self.style().unpolish(self)
        self.style().polish(self)
        if path:
            self.fileDropped.emit(path)
            event.acceptProposedAction()
            return
        event.ignore()


class BulkTextDelegate(QStyledItemDelegate):
    """Spreadsheet-like text editor behavior for bulk name/note cells."""

    advanceRequested = Signal(int, int)

    def createEditor(self, parent, option, index):
        editor = super().createEditor(parent, option, index)
        if editor is not None:
            editor.setProperty("bulkEditorRow", index.row())
            editor.setProperty("bulkEditorColumn", index.column())
        return editor

    def eventFilter(self, editor, event):
        if event.type() == QEvent.KeyPress and event.key() in (Qt.Key_Return, Qt.Key_Enter):
            row = int(editor.property("bulkEditorRow") or 0)
            column = int(editor.property("bulkEditorColumn") or 0)
            self.commitData.emit(editor)
            self.closeEditor.emit(editor, QAbstractItemDelegate.NoHint)
            QTimer.singleShot(0, lambda r=row, c=column: self.advanceRequested.emit(r, c))
            event.accept()
            return True
        return super().eventFilter(editor, event)


class BulkComponentTable(QTableWidget):
    pasteRequested = Signal(str)

    def keyPressEvent(self, event):
        if event.matches(QKeySequence.Paste):
            self.pasteRequested.emit(QApplication.clipboard().text())
            event.accept()
            return

        # QTableWidget's AnyKeyPressed trigger can open the editor while losing
        # the first typed character. Start the editor ourselves and inject that
        # character so spreadsheet-style direct typing is lossless.
        blocked_modifiers = Qt.ControlModifier | Qt.AltModifier | Qt.MetaModifier
        item = self.currentItem()
        if (
            self.state() != QAbstractItemView.EditingState
            and self.currentColumn() in (COL_NAME, COL_NOTE)
            and item is not None
            and bool(item.flags() & Qt.ItemIsEditable)
            and bool(event.text())
            and not bool(event.modifiers() & blocked_modifiers)
        ):
            self.editItem(item)
            editor = QApplication.focusWidget()
            if isinstance(editor, QLineEdit):
                editor.selectAll()
                editor.insert(event.text())
                event.accept()
                return

        super().keyPressEvent(event)


class ComponentEntryDialog(QDialog):
    """Single and bulk component entry dialog used by PlatformComponentManagerDialog.

    The dialog keeps the approved two-level navigation:
    - Tekli Bileşen / Toplu Bileşen
    - Toplu mode: Elle Giriş / Excel'den Aktar

    Bulk writes pass the complete merged component list to ``write_components``
    once, preserving STSStore's transaction boundary and audit behavior.
    """

    componentsSaved = Signal(int, int)

    def __init__(self, store, parent=None, initial_mode: str = "single"):
        super().__init__(parent)
        self.store = store
        self.result_summary = ""
        self._saved = False
        self._dirty = False
        self._building = True
        self._primary_mode = "single"
        self._bulk_mode = "manual"
        self._validated_rows = []
        self._excel_path = ""
        self._excel_matrix: list[list[Any]] = []
        self._excel_detected_mapping = WorkbookMapping(name_column=None)

        self.setWindowTitle("Bileşen Ekle")
        self.setModal(True)
        self.setWindowFlags(Qt.Dialog | Qt.FramelessWindowHint)
        self.setAttribute(Qt.WA_TranslucentBackground, True)
        self._existing_components = self._load_existing_components()
        self._build()
        self._apply_style()
        self._add_initial_rows()
        self._building = False
        self._set_primary_mode("bulk" if initial_mode == "bulk" else "single")
        self._validate_bulk()

    # ------------------------------------------------------------------
    # Build
    # ------------------------------------------------------------------
    def _build(self) -> None:
        outer = QVBoxLayout(self)
        outer.setContentsMargins(14, 14, 14, 14)
        outer.setSpacing(0)

        self.card = QFrame(objectName="componentEntryCard")
        shadow = QGraphicsDropShadowEffect(self.card)
        shadow.setBlurRadius(42)
        shadow.setOffset(0, 12)
        shadow.setColor(QColor(15, 31, 61, 88))
        self.card.setGraphicsEffect(shadow)
        outer.addWidget(self.card)

        card_layout = QVBoxLayout(self.card)
        card_layout.setContentsMargins(0, 0, 0, 0)
        card_layout.setSpacing(0)

        self.header = DragHeader()
        self.header.setObjectName("componentEntryHeader")
        head = QHBoxLayout(self.header)
        head.setContentsMargins(22, 17, 18, 15)
        head.setSpacing(13)

        icon = QLabel("＋", objectName="componentEntryIcon")
        icon.setFixedSize(42, 42)
        icon.setAlignment(Qt.AlignCenter)

        head_copy = QVBoxLayout()
        head_copy.setSpacing(2)
        title = QLabel("Bileşen Ekle", objectName="componentEntryTitle")
        self.header_subtitle = QLabel(
            "Bileşen bilgilerini girin.",
            objectName="componentEntrySubtitle",
        )
        head_copy.addWidget(title)
        head_copy.addWidget(self.header_subtitle)

        close_button = QPushButton("×", objectName="componentEntryClose")
        close_button.setFixedSize(34, 34)
        close_button.clicked.connect(self.close)

        head.addWidget(icon)
        head.addLayout(head_copy, 1)
        head.addWidget(close_button)
        card_layout.addWidget(self.header)

        self.primary_tabs = QFrame(objectName="componentPrimaryTabs")
        tabs_layout = QHBoxLayout(self.primary_tabs)
        tabs_layout.setContentsMargins(22, 10, 22, 0)
        tabs_layout.setSpacing(8)
        self.single_tab = self._tab_button("Tekli Bileşen")
        self.bulk_tab = self._tab_button("Toplu Bileşen")
        self.primary_group = QButtonGroup(self)
        self.primary_group.setExclusive(True)
        self.primary_group.addButton(self.single_tab)
        self.primary_group.addButton(self.bulk_tab)
        self.single_tab.clicked.connect(lambda: self._set_primary_mode("single"))
        self.bulk_tab.clicked.connect(lambda: self._set_primary_mode("bulk"))
        tabs_layout.addWidget(self.single_tab)
        tabs_layout.addWidget(self.bulk_tab)
        tabs_layout.addStretch(1)
        card_layout.addWidget(self.primary_tabs)

        self.pages = QStackedWidget(objectName="componentEntryPages")
        self.single_page = self._build_single_page()
        self.bulk_page = self._build_bulk_page()
        self.pages.addWidget(self.single_page)
        self.pages.addWidget(self.bulk_page)
        card_layout.addWidget(self.pages, 1)

        self.summary_bar = self._build_summary_bar()
        card_layout.addWidget(self.summary_bar)
        card_layout.addWidget(self._build_footer())

    def _tab_button(self, text: str) -> QPushButton:
        button = QPushButton(text, objectName="componentTabButton")
        button.setCheckable(True)
        button.setCursor(Qt.PointingHandCursor)
        button.setFixedHeight(36)
        return button

    def _build_single_page(self) -> QWidget:
        page = QWidget(objectName="singleComponentPage")
        root = QVBoxLayout(page)
        root.setContentsMargins(22, 14, 22, 18)
        root.setSpacing(12)

        panel = QFrame(objectName="singleFormPanel")
        grid = QGridLayout(panel)
        grid.setContentsMargins(16, 15, 16, 15)
        grid.setHorizontalSpacing(9)
        grid.setVerticalSpacing(7)

        grid.addWidget(self._field_label("BİLEŞEN ADI"), 0, 0)
        grid.addWidget(self._field_label("BİRİM"), 0, 1)
        self.single_name = QLineEdit(objectName="componentField")
        self.single_name.setPlaceholderText("Bileşen adı")
        self.single_unit = QComboBox(objectName="componentCombo")
        self.single_unit.addItems(SUPPORTED_UNITS)
        grid.addWidget(self.single_name, 1, 0)
        grid.addWidget(self.single_unit, 1, 1)

        grid.addWidget(self._field_label("NOT"), 2, 0, 1, 2)
        self.single_note = QLineEdit(objectName="componentField")
        self.single_note.setPlaceholderText("İsteğe bağlı kısa not...")
        grid.addWidget(self.single_note, 3, 0, 1, 2)

        grid.addWidget(self._field_label("DURUM"), 4, 0, 1, 2)
        status_host = QWidget()
        status_layout = QHBoxLayout(status_host)
        status_layout.setContentsMargins(0, 2, 0, 1)
        status_layout.setSpacing(9)
        self.single_active = SlideToggle(True)
        self.single_active_label = QLabel("Aktif", objectName="componentStatusText")
        self.single_active.toggled.connect(
            lambda checked: self.single_active_label.setText("Aktif" if checked else "Pasif")
        )
        status_layout.addWidget(self.single_active)
        status_layout.addWidget(self.single_active_label)
        status_layout.addStretch(1)
        grid.addWidget(status_host, 5, 0, 1, 2)
        grid.setColumnStretch(0, 1)
        grid.setColumnMinimumWidth(1, 118)

        self.single_name.textChanged.connect(self._mark_dirty)
        self.single_note.textChanged.connect(self._mark_dirty)
        self.single_unit.currentTextChanged.connect(self._mark_dirty)
        self.single_active.toggled.connect(self._mark_dirty)

        root.addWidget(panel)
        root.addStretch(1)
        return page

    def _build_bulk_page(self) -> QWidget:
        page = QWidget(objectName="bulkComponentPage")
        root = QVBoxLayout(page)
        root.setContentsMargins(22, 12, 22, 0)
        root.setSpacing(10)

        mode_row = QHBoxLayout()
        mode_row.setSpacing(8)
        self.manual_mode_button = self._tab_button("⌨  Elle Giriş")
        self.excel_mode_button = self._tab_button("▦  Excel'den Aktar")
        self.bulk_mode_group = QButtonGroup(self)
        self.bulk_mode_group.setExclusive(True)
        self.bulk_mode_group.addButton(self.manual_mode_button)
        self.bulk_mode_group.addButton(self.excel_mode_button)
        self.manual_mode_button.clicked.connect(lambda: self._set_bulk_mode("manual"))
        self.excel_mode_button.clicked.connect(lambda: self._set_bulk_mode("excel"))
        mode_row.addWidget(self.manual_mode_button)
        mode_row.addWidget(self.excel_mode_button)
        mode_row.addStretch(1)
        root.addLayout(mode_row)

        self.bulk_pages = QStackedWidget()
        self.manual_page = self._build_manual_page()
        self.excel_page = self._build_excel_page()
        self.bulk_pages.addWidget(self.manual_page)
        self.bulk_pages.addWidget(self.excel_page)
        root.addWidget(self.bulk_pages, 1)
        return page

    def _build_manual_page(self) -> QWidget:
        page = QWidget()
        root = QVBoxLayout(page)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(10)

        toolbar = QFrame(objectName="bulkToolbar")
        toolbar_layout = QHBoxLayout(toolbar)
        toolbar_layout.setContentsMargins(11, 10, 11, 10)
        toolbar_layout.setSpacing(8)

        add_row = self._tool_button("＋ Satır Ekle", "bulkToolPrimary")
        paste = self._tool_button("▤ Panodan Yapıştır")
        clear = self._tool_button("⌫ Listeyi Temizle")
        add_row.clicked.connect(lambda: self._add_bulk_row(focus=True))
        paste.clicked.connect(lambda: self._paste_text(QApplication.clipboard().text()))
        clear.clicked.connect(self._clear_bulk_rows)
        toolbar_layout.addWidget(add_row)
        toolbar_layout.addWidget(paste)
        toolbar_layout.addWidget(clear)
        toolbar_layout.addStretch(1)

        default_label = QLabel("Varsayılan birim", objectName="bulkToolbarLabel")
        self.default_unit = QComboBox(objectName="bulkCompactCombo")
        self.default_unit.addItems(SUPPORTED_UNITS)
        self.default_unit.currentTextChanged.connect(self._apply_default_unit_to_blank_rows)
        toolbar_layout.addWidget(default_label)
        toolbar_layout.addWidget(self.default_unit)

        details_host = QWidget()
        details_layout = QHBoxLayout(details_host)
        details_layout.setContentsMargins(3, 0, 0, 0)
        details_layout.setSpacing(7)
        self.details_toggle = SlideToggle(False)
        details_label = QLabel("Detaylı alanlar", objectName="bulkToolbarLabel")
        details_label.setCursor(Qt.PointingHandCursor)
        self.details_toggle.toggled.connect(self._set_details_visible)
        details_label.mousePressEvent = lambda event: self.details_toggle.setChecked(
            not self.details_toggle.isChecked()
        )
        details_layout.addWidget(self.details_toggle)
        details_layout.addWidget(details_label)
        toolbar_layout.addWidget(details_host)
        root.addWidget(toolbar)

        hint = QFrame(objectName="bulkPasteHint")
        hint_layout = QHBoxLayout(hint)
        hint_layout.setContentsMargins(12, 9, 12, 9)
        hint_text = QLabel(
            "<b>Hızlı giriş:</b> Excel'deki bir sütunu kopyalayıp ilk Bileşen Adı hücresine yapıştırabilirsiniz.",
            objectName="bulkHintText",
        )
        shortcut = QLabel("Ctrl + V", objectName="bulkShortcut")
        hint_layout.addWidget(hint_text, 1)
        hint_layout.addWidget(shortcut)
        root.addWidget(hint)

        self.bulk_table = BulkComponentTable(objectName="bulkComponentTable")
        self.bulk_table.setColumnCount(7)
        self.bulk_table.setHorizontalHeaderLabels(
            ["#", "BİLEŞEN ADI *", "BİRİM", "NOT", "DURUM", "KONTROL", ""]
        )
        self.bulk_table.verticalHeader().setVisible(False)
        self.bulk_table.setAlternatingRowColors(True)
        self.bulk_table.setSelectionBehavior(QAbstractItemView.SelectItems)
        self.bulk_table.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.bulk_table.setEditTriggers(
            QAbstractItemView.DoubleClicked
            | QAbstractItemView.EditKeyPressed
            | QAbstractItemView.AnyKeyPressed
            | QAbstractItemView.SelectedClicked
        )
        # Tek tıkla hücre editörünü aç; ayrıca seçili hücrede doğrudan yazmaya
        # başlandığında ilk karakteri kaybetmeden düzenlemeyi başlat.
        self.bulk_table.setFocusPolicy(Qt.StrongFocus)
        self.bulk_table.cellClicked.connect(self._begin_bulk_cell_edit)
        self.bulk_table.horizontalHeader().setSectionResizeMode(COL_NUMBER, QHeaderView.Fixed)
        self.bulk_table.horizontalHeader().setSectionResizeMode(COL_NAME, QHeaderView.Stretch)
        self.bulk_table.horizontalHeader().setSectionResizeMode(COL_UNIT, QHeaderView.Fixed)
        self.bulk_table.horizontalHeader().setSectionResizeMode(COL_NOTE, QHeaderView.Stretch)
        self.bulk_table.horizontalHeader().setSectionResizeMode(COL_STATUS, QHeaderView.Fixed)
        self.bulk_table.horizontalHeader().setSectionResizeMode(COL_CONTROL, QHeaderView.Fixed)
        self.bulk_table.horizontalHeader().setSectionResizeMode(COL_REMOVE, QHeaderView.Fixed)
        self.bulk_table.setColumnWidth(COL_NUMBER, 46)
        self.bulk_table.setColumnWidth(COL_UNIT, 126)
        self.bulk_table.setColumnWidth(COL_STATUS, 116)
        self.bulk_table.setColumnWidth(COL_CONTROL, 145)
        self.bulk_table.setColumnWidth(COL_REMOVE, 46)
        self.bulk_text_delegate = BulkTextDelegate(self.bulk_table)
        self.bulk_text_delegate.advanceRequested.connect(self._advance_bulk_cell)
        self.bulk_table.setItemDelegateForColumn(COL_NAME, self.bulk_text_delegate)
        self.bulk_table.setItemDelegateForColumn(COL_NOTE, self.bulk_text_delegate)
        self.bulk_table.itemChanged.connect(self._on_bulk_item_changed)
        self.bulk_table.pasteRequested.connect(self._paste_text)
        root.addWidget(self.bulk_table, 1)
        return page

    def _build_excel_page(self) -> QWidget:
        page = QWidget()
        root = QVBoxLayout(page)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(10)

        self.drop_zone = DropZone()
        drop_layout = QVBoxLayout(self.drop_zone)
        drop_layout.setContentsMargins(24, 22, 24, 18)
        drop_layout.setSpacing(7)
        drop_layout.setAlignment(Qt.AlignCenter)
        icon = QLabel("▦", objectName="bulkDropIcon")
        icon.setFixedSize(56, 56)
        icon.setAlignment(Qt.AlignCenter)
        title = QLabel("Excel dosyasını buraya bırakın", objectName="bulkDropTitle")
        description = QLabel(
            ".xlsx dosyanız tek sütunlu bir liste veya Bileşen Adı / Birim / Not / Durum başlıklarını içeren bir tablo olabilir.",
            objectName="bulkDropDescription",
        )
        description.setWordWrap(True)
        description.setAlignment(Qt.AlignCenter)
        choose = QPushButton("Dosya Seç", objectName="bulkChooseFile")
        choose.setFixedHeight(36)
        choose.clicked.connect(self._choose_excel_file)
        self.file_meta = QLabel("", objectName="bulkFileMeta")
        self.file_meta.setWordWrap(True)
        self.file_meta.hide()
        drop_layout.addWidget(icon, 0, Qt.AlignCenter)
        drop_layout.addWidget(title, 0, Qt.AlignCenter)
        drop_layout.addWidget(description)
        drop_layout.addWidget(choose, 0, Qt.AlignCenter)
        drop_layout.addWidget(self.file_meta)
        self.drop_zone.fileDropped.connect(self._load_excel_path)
        root.addWidget(self.drop_zone)

        options = QFrame(objectName="bulkExcelOptions")
        options_layout = QVBoxLayout(options)
        options_layout.setContentsMargins(11, 10, 11, 10)
        options_layout.setSpacing(9)

        file_options = QHBoxLayout()
        file_options.setSpacing(8)
        self.sheet_label = QLabel("Sayfa", objectName="bulkExcelFieldLabel")
        self.sheet_combo = QComboBox(objectName="bulkMappingCombo")
        self.sheet_combo.currentTextChanged.connect(self._reload_excel_sheet)
        self.header_checkbox = QCheckBox("İlk dolu satır başlık", objectName="bulkHeaderCheck")
        self.header_checkbox.toggled.connect(self._refresh_excel_mapping)
        template = QPushButton("Örnek Şablon Oluştur", objectName="bulkTemplateButton")
        template.clicked.connect(self._export_template)
        file_options.addWidget(self.sheet_label)
        file_options.addWidget(self.sheet_combo, 1)
        file_options.addWidget(self.header_checkbox)
        file_options.addStretch(1)
        file_options.addWidget(template)
        options_layout.addLayout(file_options)

        mapping_grid = QGridLayout()
        mapping_grid.setHorizontalSpacing(9)
        mapping_grid.setVerticalSpacing(5)
        self.name_mapping = self._mapping_combo()
        self.unit_mapping = self._mapping_combo(include_default=True, default_text="Varsayılan: Adet")
        self.note_mapping = self._mapping_combo(include_default=True, default_text="Aktarma")
        self.status_mapping = self._mapping_combo(include_default=True, default_text="Varsayılan: Aktif")
        for column, (label_text, combo) in enumerate(
            [
                ("BİLEŞEN ADI", self.name_mapping),
                ("BİRİM", self.unit_mapping),
                ("NOT", self.note_mapping),
                ("DURUM", self.status_mapping),
            ]
        ):
            mapping_grid.addWidget(self._field_label(label_text), 0, column)
            mapping_grid.addWidget(combo, 1, column)
        options_layout.addLayout(mapping_grid)
        root.addWidget(options)
        root.addStretch(1)
        return page

    def _build_summary_bar(self) -> QFrame:
        bar = QFrame(objectName="bulkSummaryBar")
        layout = QHBoxLayout(bar)
        layout.setContentsMargins(22, 9, 22, 9)
        layout.setSpacing(8)
        self.total_badge = QLabel("0 satır", objectName="summaryNeutral")
        self.ready_badge = QLabel("0 hazır", objectName="summaryGood")
        self.warning_badge = QLabel("0 kontrol gerekli", objectName="summaryWarn")
        note = QLabel("Mevcut bileşenler varsayılan olarak atlanır.", objectName="summaryNote")
        layout.addWidget(self.total_badge)
        layout.addWidget(self.ready_badge)
        layout.addWidget(self.warning_badge)
        layout.addStretch(1)
        layout.addWidget(note)
        return bar

    def _build_footer(self) -> QFrame:
        footer = QFrame(objectName="componentEntryFooter")
        layout = QHBoxLayout(footer)
        layout.setContentsMargins(22, 12, 22, 16)
        layout.setSpacing(10)
        layout.addStretch(1)
        cancel = QPushButton("Vazgeç", objectName="componentCancel")
        cancel.setFixedSize(106, 38)
        cancel.clicked.connect(self.close)
        self.save_button = QPushButton("Kaydet", objectName="componentPrimary")
        self.save_button.setMinimumSize(118, 38)
        self.save_button.clicked.connect(self._save_current)
        layout.addWidget(cancel)
        layout.addWidget(self.save_button)
        return footer

    def _field_label(self, text: str) -> QLabel:
        return QLabel(text, objectName="componentFieldLabel")

    def _tool_button(self, text: str, object_name: str = "bulkToolButton") -> QPushButton:
        button = QPushButton(text, objectName=object_name)
        button.setFixedHeight(34)
        button.setCursor(Qt.PointingHandCursor)
        return button

    def _mapping_combo(self, include_default: bool = False, default_text: str = "") -> QComboBox:
        combo = QComboBox(objectName="bulkMappingCombo")
        if include_default:
            combo.addItem(default_text, None)
        combo.currentIndexChanged.connect(self._mark_dirty)
        return combo

    # ------------------------------------------------------------------
    # State / navigation
    # ------------------------------------------------------------------
    def _set_primary_mode(self, mode: str) -> None:
        self._primary_mode = "bulk" if mode == "bulk" else "single"
        bulk = self._primary_mode == "bulk"
        self.bulk_tab.setChecked(bulk)
        self.single_tab.setChecked(not bulk)
        self.pages.setCurrentWidget(self.bulk_page if bulk else self.single_page)
        self.summary_bar.setVisible(bulk)
        if bulk:
            self.header_subtitle.setText(
                "Bileşenleri elle girin, Excel'den yapıştırın veya dosyadan aktarın."
            )
            self.setMinimumSize(900, 620)
            self.resize(1040, 720)
            self._set_bulk_mode(self._bulk_mode)
        else:
            self.header_subtitle.setText("Bileşen bilgilerini girin.")
            self.setMinimumSize(430, 470)
            self.resize(450, 520)
            self.save_button.setText("Kaydet")
            self.save_button.setEnabled(bool(normalize_space(self.single_name.text())))
            self.single_name.setFocus()

    def _set_bulk_mode(self, mode: str) -> None:
        self._bulk_mode = "excel" if mode == "excel" else "manual"
        excel = self._bulk_mode == "excel"
        self.excel_mode_button.setChecked(excel)
        self.manual_mode_button.setChecked(not excel)
        self.bulk_pages.setCurrentWidget(self.excel_page if excel else self.manual_page)
        self._update_save_button()

    def _mark_dirty(self, *args) -> None:
        if not self._building:
            self._dirty = True
        if self._primary_mode == "single":
            self.save_button.setEnabled(bool(normalize_space(self.single_name.text())))

    def closeEvent(self, event):
        if self._dirty and not self._saved:
            if not ask_yes_no(
                self,
                "Kaydedilmemiş Değişiklik",
                "Girdiğiniz bilgiler kaydedilmedi. Pencere kapatılsın mı?",
                default_yes=False,
            ):
                event.ignore()
                return
        event.accept()

    # ------------------------------------------------------------------
    # Existing/store
    # ------------------------------------------------------------------
    def _load_existing_components(self) -> list[dict[str, Any]]:
        if hasattr(self.store, "load_components_full"):
            return [dict(item) for item in self.store.load_components_full()]
        output = []
        for index, component in enumerate(
            self.store.load_components() if hasattr(self.store, "load_components") else []
        ):
            output.append(
                {
                    "id": getattr(component, "id", None),
                    "name": str(getattr(component, "name", "") or ""),
                    "version": str(getattr(component, "version", "") or ""),
                    "unit": str(getattr(component, "unit", "Adet") or "Adet"),
                    "active": bool(getattr(component, "active", True)),
                    "usage": getattr(component, "usage", 1),
                    "note": str(getattr(component, "note", "") or ""),
                    "display_order": getattr(component, "display_order", index),
                    "platforms": dict(getattr(component, "platforms", {}) or {}),
                }
            )
        return output

    def _write_component_list(self, components: Sequence[dict[str, Any]]) -> None:
        if not hasattr(self.store, "write_components"):
            raise RuntimeError("Bu veri kaynağı toplu bileşen kaydını desteklemiyor.")
        actor = self.store.current_actor() if hasattr(self.store, "current_actor") else "Sistem"
        self.store.write_components(list(components), actor=actor)

    # ------------------------------------------------------------------
    # Single save
    # ------------------------------------------------------------------
    def _save_single(self) -> None:
        name = normalize_space(self.single_name.text())
        if not name:
            show_warning(self, "Eksik", "Bileşen adı girin.")
            self.single_name.setFocus()
            return
        if any(component_key(item.get("name")) == component_key(name) for item in self._existing_components):
            show_warning(self, "Mevcut Bileşen", f"{name} STS içinde zaten bulunuyor.")
            return

        row = BulkComponentRow(
            name=name,
            unit=self.single_unit.currentText(),
            note=self.single_note.text(),
            active=self.single_active.isChecked(),
            source_row=1,
        )
        merged = merge_new_components(self._existing_components, [row])
        try:
            self._write_component_list(merged.components)
        except Exception as exc:
            show_warning(self, "Bileşen Kaydedilemedi", str(exc))
            return

        self.result_summary = "Bileşen kaydedildi"
        self._saved = True
        self._dirty = False
        self.componentsSaved.emit(1, 0)
        self.accept()

    # ------------------------------------------------------------------
    # Manual bulk table
    # ------------------------------------------------------------------
    def _add_initial_rows(self) -> None:
        for _ in range(4):
            self._add_bulk_row()
        self._set_details_visible(False)

    def _add_bulk_row(self, row: BulkComponentRow | None = None, focus: bool = False) -> None:
        row = row or BulkComponentRow(name="", unit=self.default_unit.currentText())
        self.bulk_table.blockSignals(True)
        table_row = self.bulk_table.rowCount()
        self.bulk_table.insertRow(table_row)
        self.bulk_table.setRowHeight(table_row, 44)

        number = QTableWidgetItem(str(table_row + 1))
        number.setFlags(Qt.ItemIsEnabled)
        number.setTextAlignment(Qt.AlignCenter)
        self.bulk_table.setItem(table_row, COL_NUMBER, number)

        name_item = QTableWidgetItem(row.name)
        name_item.setData(Qt.UserRole, int(row.source_row or 0))
        self.bulk_table.setItem(table_row, COL_NAME, name_item)

        unit_combo = QComboBox(objectName="bulkCellCombo")
        unit_combo.addItems(SUPPORTED_UNITS)
        if row.unit not in [unit_combo.itemText(i) for i in range(unit_combo.count())]:
            unit_combo.addItem(row.unit)
        unit_combo.setCurrentText(row.unit or self.default_unit.currentText())
        unit_combo.currentTextChanged.connect(self._on_bulk_widget_changed)
        self.bulk_table.setCellWidget(table_row, COL_UNIT, unit_combo)

        self.bulk_table.setItem(table_row, COL_NOTE, QTableWidgetItem(row.note))

        status_combo = QComboBox(objectName="bulkCellCombo")
        status_combo.addItems(["Aktif", "Pasif"])
        status_combo.setCurrentText("Aktif" if row.active else "Pasif")
        status_combo.currentTextChanged.connect(self._on_bulk_widget_changed)
        self.bulk_table.setCellWidget(table_row, COL_STATUS, status_combo)

        state = QLabel("Eksik", objectName="bulkStateEmpty")
        state.setAlignment(Qt.AlignCenter)
        state.setToolTip("Bileşen adı eksik")
        self.bulk_table.setCellWidget(table_row, COL_CONTROL, state)

        remove = QPushButton("×", objectName="bulkRemoveRow")
        remove.setFixedSize(28, 28)
        remove.clicked.connect(lambda _checked=False, button=remove: self._remove_bulk_row(button))
        remove_host = QWidget()
        remove_layout = QHBoxLayout(remove_host)
        remove_layout.setContentsMargins(0, 0, 0, 0)
        remove_layout.setAlignment(Qt.AlignCenter)
        remove_layout.addWidget(remove)
        self.bulk_table.setCellWidget(table_row, COL_REMOVE, remove_host)
        self.bulk_table.blockSignals(False)
        self._renumber_rows()
        if not self._building:
            self._mark_dirty()
        self._validate_bulk()
        if focus:
            self.bulk_table.setCurrentCell(table_row, COL_NAME)
            self.bulk_table.editItem(name_item)

    def _remove_bulk_row(self, button: QPushButton) -> None:
        for row in range(self.bulk_table.rowCount()):
            host = self.bulk_table.cellWidget(row, COL_REMOVE)
            if host and button in host.findChildren(QPushButton):
                self.bulk_table.removeRow(row)
                break
        if self.bulk_table.rowCount() == 0:
            self._add_bulk_row()
        self._renumber_rows()
        self._mark_dirty()
        self._validate_bulk()

    def _renumber_rows(self) -> None:
        for row in range(self.bulk_table.rowCount()):
            item = self.bulk_table.item(row, COL_NUMBER)
            if item is None:
                item = QTableWidgetItem()
                item.setFlags(Qt.ItemIsEnabled)
                item.setTextAlignment(Qt.AlignCenter)
                self.bulk_table.setItem(row, COL_NUMBER, item)
            item.setText(str(row + 1))

    def _clear_bulk_rows(self) -> None:
        self.bulk_table.setRowCount(0)
        for _ in range(4):
            self._add_bulk_row()
        self._mark_dirty()
        self._validate_bulk()

    def _begin_bulk_cell_edit(self, row: int, column: int) -> None:
        """Open text cells on the first click instead of requiring a double click."""
        if column not in (COL_NAME, COL_NOTE):
            return
        item = self.bulk_table.item(row, column)
        if item is None or not bool(item.flags() & Qt.ItemIsEditable):
            return
        self.bulk_table.setCurrentItem(item)
        self.bulk_table.editItem(item)

    def _advance_bulk_cell(self, row: int, column: int) -> None:
        """Commit Enter, move one row down and keep the next text cell ready."""
        target_row = row + 1
        if target_row >= self.bulk_table.rowCount():
            self._add_bulk_row()
        item = self.bulk_table.item(target_row, column)
        if item is None:
            return
        self.bulk_table.setCurrentItem(item)
        self.bulk_table.scrollToItem(item, QAbstractItemView.PositionAtCenter)
        self.bulk_table.setFocus()

    def _on_bulk_item_changed(self, item: QTableWidgetItem) -> None:
        if item.column() in (COL_NAME, COL_NOTE):
            self._mark_dirty()
            self._validate_bulk()

    def _on_bulk_widget_changed(self, *args) -> None:
        self._mark_dirty()
        self._validate_bulk()

    def _table_rows(self) -> list[BulkComponentRow]:
        output: list[BulkComponentRow] = []
        for row in range(self.bulk_table.rowCount()):
            name_item = self.bulk_table.item(row, COL_NAME)
            note_item = self.bulk_table.item(row, COL_NOTE)
            unit_combo = self.bulk_table.cellWidget(row, COL_UNIT)
            status_combo = self.bulk_table.cellWidget(row, COL_STATUS)
            source_row = int(name_item.data(Qt.UserRole) or row + 1) if name_item else row + 1
            output.append(
                BulkComponentRow(
                    name=normalize_space(name_item.text() if name_item else ""),
                    unit=unit_combo.currentText() if isinstance(unit_combo, QComboBox) else self.default_unit.currentText(),
                    note=normalize_space(note_item.text() if note_item else ""),
                    active=(status_combo.currentText() != "Pasif") if isinstance(status_combo, QComboBox) else True,
                    source_row=source_row,
                )
            )
        return output

    def _validate_bulk(self) -> None:
        if not hasattr(self, "bulk_table"):
            return
        self._validated_rows = validate_bulk_rows(self._table_rows(), self._existing_components)
        ready = 0
        warnings = 0
        for row, validated in enumerate(self._validated_rows):
            label = self.bulk_table.cellWidget(row, COL_CONTROL)
            if not isinstance(label, QLabel):
                continue
            if validated.state == "ready":
                label.setObjectName("bulkStateReady")
                label.setText("●  Hazır")
                ready += 1
            elif validated.state == "existing":
                label.setObjectName("bulkStateExisting")
                label.setText("●  Mevcut")
                warnings += 1
            elif validated.state == "duplicate":
                label.setObjectName("bulkStateDuplicate")
                label.setText("●  Tekrar")
                warnings += 1
            else:
                label.setObjectName("bulkStateEmpty")
                label.setText("●  Eksik")
                warnings += 1
            label.setToolTip(validated.message)
            label.style().unpolish(label)
            label.style().polish(label)

        total = self.bulk_table.rowCount()
        self.total_badge.setText(f"{total} satır")
        self.ready_badge.setText(f"{ready} hazır")
        self.warning_badge.setText(f"{warnings} kontrol gerekli")
        self._update_save_button()

    def _update_save_button(self) -> None:
        if self._primary_mode == "single":
            self.save_button.setText("Kaydet")
            self.save_button.setEnabled(bool(normalize_space(self.single_name.text())))
            return
        if self._bulk_mode == "excel":
            self.save_button.setText("Dosyayı Önizle")
            self.save_button.setEnabled(bool(self._excel_matrix and self.name_mapping.currentData() is not None))
            return
        ready = sum(1 for row in self._validated_rows if row.is_ready)
        self.save_button.setText(f"{ready} Bileşeni Ekle" if ready else "Bileşenleri Ekle")
        self.save_button.setEnabled(ready > 0)

    def _set_details_visible(self, visible: bool) -> None:
        for column in (COL_UNIT, COL_NOTE, COL_STATUS):
            self.bulk_table.setColumnHidden(column, not visible)

    def _apply_default_unit_to_blank_rows(self, unit: str) -> None:
        for row in range(self.bulk_table.rowCount()):
            name_item = self.bulk_table.item(row, COL_NAME)
            combo = self.bulk_table.cellWidget(row, COL_UNIT)
            if (not name_item or not normalize_space(name_item.text())) and isinstance(combo, QComboBox):
                combo.blockSignals(True)
                combo.setCurrentText(unit)
                combo.blockSignals(False)
        self._mark_dirty()
        self._validate_bulk()

    def _paste_text(self, text: str) -> None:
        raw = str(text or "")
        lines = [line for line in raw.splitlines() if line.strip()]
        if not lines:
            show_warning(self, "Pano Boş", "Panoda aktarılabilecek satır bulunamadı.")
            return
        start_row = self.bulk_table.currentRow()
        if start_row < 0:
            start_row = 0
        for offset, line in enumerate(lines):
            columns = line.split("\t")
            target = start_row + offset
            while target >= self.bulk_table.rowCount():
                self._add_bulk_row()
            self.bulk_table.blockSignals(True)
            name_item = self.bulk_table.item(target, COL_NAME)
            note_item = self.bulk_table.item(target, COL_NOTE)
            if name_item:
                name_item.setText(columns[0].strip() if columns else "")
            if len(columns) > 1:
                combo = self.bulk_table.cellWidget(target, COL_UNIT)
                if isinstance(combo, QComboBox):
                    value = normalize_space(columns[1])
                    if value and value not in [combo.itemText(i) for i in range(combo.count())]:
                        combo.addItem(value)
                    if value:
                        combo.setCurrentText(value)
            if len(columns) > 2 and note_item:
                note_item.setText(columns[2].strip())
            if len(columns) > 3:
                status = self.bulk_table.cellWidget(target, COL_STATUS)
                if isinstance(status, QComboBox):
                    status.setCurrentText("Pasif" if columns[3].strip().casefold() in {"pasif", "0", "false", "hayır", "hayir"} else "Aktif")
            self.bulk_table.blockSignals(False)
        self._mark_dirty()
        self._validate_bulk()
        self.bulk_table.setCurrentCell(start_row, COL_NAME)

    def _save_bulk(self) -> None:
        ready_rows = [validated.row for validated in self._validated_rows if validated.is_ready]
        if not ready_rows:
            show_warning(self, "Eklenecek Bileşen Yok", "Eklemeye hazır en az bir satır bulunmalıdır.")
            return
        merged = merge_new_components(self._existing_components, self._table_rows())
        try:
            self._write_component_list(merged.components)
        except Exception as exc:
            show_warning(self, "Bileşenler Kaydedilemedi", str(exc))
            return

        skipped = merged.skipped_existing + merged.skipped_duplicate + merged.skipped_blank
        lines = [f"{merged.added} bileşen eklendi."]
        if merged.skipped_existing:
            lines.append(f"{merged.skipped_existing} mevcut bileşen atlandı.")
        if merged.skipped_duplicate:
            lines.append(f"{merged.skipped_duplicate} tekrar eden satır atlandı.")
        self.result_summary = f"{merged.added} bileşen eklendi"
        show_information(self, "Toplu Ekleme Tamamlandı", "\n".join(lines))
        self._saved = True
        self._dirty = False
        self.componentsSaved.emit(merged.added, skipped)
        self.accept()

    # ------------------------------------------------------------------
    # Excel
    # ------------------------------------------------------------------
    def _choose_excel_file(self) -> None:
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Bileşen Listesi Seç",
            str(Path.home()),
            "Excel Dosyaları (*.xlsx)",
        )
        if path:
            self._load_excel_path(path)

    def _load_excel_path(self, path: str) -> None:
        selected = Path(path)
        if selected.suffix.casefold() != ".xlsx":
            show_warning(self, "Geçersiz Dosya", "Yalnızca .xlsx dosyaları desteklenir.")
            return
        try:
            sheets = list_workbook_sheets(selected)
        except Exception as exc:
            show_warning(self, "Excel Açılamadı", str(exc))
            return
        if not sheets:
            show_warning(self, "Excel Açılamadı", "Dosyada okunabilir çalışma sayfası bulunamadı.")
            return

        self._excel_path = str(selected)
        self.sheet_combo.blockSignals(True)
        self.sheet_combo.clear()
        self.sheet_combo.addItems(sheets)
        self.sheet_combo.blockSignals(False)
        self.file_meta.setText(f"<b>{selected.name}</b><br>{len(sheets)} sayfa bulundu · Eşleştirme için hazır")
        self.file_meta.show()
        self._load_selected_sheet(sheets[0])
        self._mark_dirty()

    def _reload_excel_sheet(self, sheet_name: str) -> None:
        if self._excel_path and sheet_name:
            self._load_selected_sheet(sheet_name)

    def _load_selected_sheet(self, sheet_name: str) -> None:
        try:
            self._excel_matrix = load_workbook_matrix(self._excel_path, sheet_name)
        except Exception as exc:
            show_warning(self, "Sayfa Okunamadı", str(exc))
            self._excel_matrix = []
            self._update_save_button()
            return
        self._excel_detected_mapping = detect_workbook_mapping(self._excel_matrix)
        self.header_checkbox.blockSignals(True)
        self.header_checkbox.setChecked(self._excel_detected_mapping.header_row is not None)
        self.header_checkbox.blockSignals(False)
        self._refresh_excel_mapping()

    @staticmethod
    def _column_letter(index: int) -> str:
        value = index + 1
        result = ""
        while value:
            value, remainder = divmod(value - 1, 26)
            result = chr(65 + remainder) + result
        return result

    def _first_nonempty_row_index(self) -> int | None:
        for index, row in enumerate(self._excel_matrix):
            if any(normalize_space(cell) for cell in row):
                return index
        return None

    def _refresh_excel_mapping(self, *args) -> None:
        if not self._excel_matrix:
            for combo in (self.name_mapping, self.unit_mapping, self.note_mapping, self.status_mapping):
                combo.clear()
            self._update_save_button()
            return

        first_index = self._first_nonempty_row_index()
        if first_index is None:
            self._update_save_button()
            return
        first_row = self._excel_matrix[first_index]
        max_columns = max((len(row) for row in self._excel_matrix), default=0)
        options = []
        for column in range(max_columns):
            header = normalize_space(first_row[column] if column < len(first_row) else "")
            label = f"Sütun {self._column_letter(column)}"
            if header:
                label += f" — {header}"
            options.append((label, column))

        combos = [
            (self.name_mapping, False, ""),
            (self.unit_mapping, True, "Varsayılan: Adet"),
            (self.note_mapping, True, "Aktarma"),
            (self.status_mapping, True, "Varsayılan: Aktif"),
        ]
        previous = [combo.currentData() for combo, _has_default, _text in combos]
        for combo, has_default, default_text in combos:
            combo.blockSignals(True)
            combo.clear()
            if has_default:
                combo.addItem(default_text, None)
            for label, column in options:
                combo.addItem(label, column)
            combo.blockSignals(False)

        detected = self._excel_detected_mapping
        defaults = [detected.name_column, detected.unit_column, detected.note_column, detected.status_column]
        for (combo, has_default, _text), old_value, detected_value in zip(combos, previous, defaults):
            target = detected_value if detected_value is not None else old_value
            target_index = combo.findData(target)
            if target_index >= 0:
                combo.setCurrentIndex(target_index)
            elif has_default:
                combo.setCurrentIndex(0)
        self._update_save_button()

    def _preview_excel(self) -> None:
        name_column = self.name_mapping.currentData()
        if name_column is None:
            show_warning(self, "Eksik Eşleştirme", "Bileşen Adı sütununu seçin.")
            return
        header_row = self._first_nonempty_row_index() if self.header_checkbox.isChecked() else None
        mapping = WorkbookMapping(
            name_column=int(name_column),
            unit_column=self.unit_mapping.currentData(),
            note_column=self.note_mapping.currentData(),
            status_column=self.status_mapping.currentData(),
            header_row=header_row,
        )
        imported = rows_from_matrix(
            self._excel_matrix,
            mapping=mapping,
            default_unit=self.default_unit.currentText(),
            default_active=True,
        )
        if not imported:
            show_warning(self, "Veri Bulunamadı", "Seçilen eşleştirmeyle aktarılacak bileşen bulunamadı.")
            return

        self.bulk_table.setRowCount(0)
        for row in imported:
            self._add_bulk_row(row)
        self._add_bulk_row()
        self.details_toggle.setChecked(True)
        self._set_bulk_mode("manual")
        self._mark_dirty()
        self._validate_bulk()
        show_information(self, "Excel Önizlemesi", f"{len(imported)} satır tabloya aktarıldı. Kontrol edip kaydedebilirsiniz.")

    def _export_template(self) -> None:
        default_path = str(Path.home() / "STS_Bilesen_Aktarim_Sablonu.xlsx")
        path, _ = QFileDialog.getSaveFileName(
            self,
            "Bileşen Şablonu Oluştur",
            default_path,
            "Excel Dosyaları (*.xlsx)",
        )
        if not path:
            return
        target = Path(path)
        if target.suffix.casefold() != ".xlsx":
            target = target.with_suffix(".xlsx")
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill

            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Bileşenler"
            sheet.append(["Bileşen Adı", "Birim", "Not", "Durum"])
            sheet.append(["Hava Aracı", "Adet", "Örnek satır", "Aktif"])
            sheet.append(["Yer Destek Sistemi", "Set", "", "Aktif"])
            for cell in sheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="0F1F3D")
            sheet.column_dimensions["A"].width = 30
            sheet.column_dimensions["B"].width = 14
            sheet.column_dimensions["C"].width = 34
            sheet.column_dimensions["D"].width = 14
            workbook.save(target)
            workbook.close()
        except Exception as exc:
            show_warning(self, "Şablon Oluşturulamadı", str(exc))
            return
        show_information(self, "Şablon Oluşturuldu", str(target))

    # ------------------------------------------------------------------
    # Save dispatcher
    # ------------------------------------------------------------------
    def _save_current(self) -> None:
        if self._primary_mode == "single":
            self._save_single()
        elif self._bulk_mode == "excel":
            self._preview_excel()
        else:
            self._save_bulk()

    # ------------------------------------------------------------------
    # Styling
    # ------------------------------------------------------------------
    def _apply_style(self) -> None:
        self.setStyleSheet(
            """
QDialog { background:transparent; }
QFrame#componentEntryCard {
    background:#FFFFFF;
    border:1px solid #C9D4E2;
    border-radius:17px;
}
QFrame#componentEntryHeader { background:#FFFFFF; border-top-left-radius:17px; border-top-right-radius:17px; }
QLabel#componentEntryIcon {
    background:#F0FDFA; color:#0D9488; border:1px solid #BFF1E8;
    border-radius:12px; font-size:24px; font-weight:500;
}
QLabel#componentEntryTitle { color:#152238; font-size:18px; font-weight:800; background:transparent; }
QLabel#componentEntrySubtitle { color:#718096; font-size:12px; background:transparent; }
QPushButton#componentEntryClose {
    background:transparent; color:#58708D; border:0; border-radius:9px; font-size:22px;
}
QPushButton#componentEntryClose:hover { background:#EDF3FA; color:#0F1F3D; }
QFrame#componentPrimaryTabs { background:#FFFFFF; border-top:1px solid #EDF1F5; }
QPushButton#componentTabButton {
    background:#F7F9FC; color:#56708E; border:1px solid #DCE5EF;
    border-radius:9px; padding:0 15px; font-size:12px; font-weight:700;
}
QPushButton#componentTabButton:checked {
    background:#EDF4FF; color:#214EAC; border:1px solid #BED0F6;
}
QWidget#componentEntryPages, QWidget#singleComponentPage, QWidget#bulkComponentPage { background:#FFFFFF; }
QFrame#singleFormPanel {
    background:#FFFFFF; border:1px solid #DCE5EF; border-radius:12px;
}
QLabel#componentFieldLabel {
    color:#526985; background:transparent; font-size:10px; font-weight:800; letter-spacing:.5px;
}
QLineEdit#componentField, QComboBox#componentCombo, QComboBox#bulkCompactCombo,
QComboBox#bulkMappingCombo, QComboBox#bulkCellCombo {
    min-height:32px; background:#FFFFFF; color:#24364C;
    border:1px solid #CBD7E6; border-radius:8px; padding:0 9px; font-size:12px;
}
QLineEdit#componentField:focus, QComboBox#componentCombo:focus,
QComboBox#bulkMappingCombo:focus, QComboBox#bulkCellCombo:focus { border:1.5px solid #3B6FE8; }
QLabel#componentStatusText { color:#334155; font-size:13px; background:transparent; }
QFrame#bulkToolbar {
    background:#F8FAFC; border:1px solid #DCE5EF; border-radius:12px;
}
QPushButton#bulkToolButton, QPushButton#bulkToolPrimary, QPushButton#bulkTemplateButton {
    background:#FFFFFF; color:#354C69; border:1px solid #CBD7E6;
    border-radius:8px; padding:0 12px; font-size:12px; font-weight:700;
}
QPushButton#bulkToolButton:hover, QPushButton#bulkTemplateButton:hover { background:#F7FBFF; border-color:#9FB7D4; }
QPushButton#bulkToolPrimary { background:#EFF6FF; color:#2356B2; border-color:#C8D9FB; }
QLabel#bulkToolbarLabel { color:#526985; background:transparent; font-size:11px; font-weight:700; }
QFrame#bulkPasteHint {
    background:#F9FBFE; border:1px dashed #B8C8DC; border-radius:10px;
}
QLabel#bulkHintText { color:#617792; background:transparent; font-size:11px; }
QLabel#bulkShortcut {
    background:#FFFFFF; color:#344B66; border:1px solid #CBD7E6;
    border-radius:5px; padding:3px 8px; font-size:10px; font-weight:800;
}
QTableWidget#bulkComponentTable {
    background:#FFFFFF; alternate-background-color:#F8FAFC; color:#24364C;
    border:1px solid #DCE5EF; border-radius:11px; gridline-color:#E2E8F0;
    selection-background-color:#EDF4FF; selection-color:#152238;
}
QTableWidget#bulkComponentTable QHeaderView::section {
    background:#EDF2F7; color:#47617F; border:0; border-right:1px solid #D7E0EA;
    border-bottom:1px solid #CED9E6; padding:0 9px; height:38px;
    font-size:10px; font-weight:800;
}
QTableWidget#bulkComponentTable::item { padding:5px 8px; border:0; }
QPushButton#bulkRemoveRow { background:transparent; color:#94A3B8; border:0; border-radius:7px; font-size:18px; }
QPushButton#bulkRemoveRow:hover { background:#FFF1F2; color:#DC2626; }
QLabel#bulkStateReady, QLabel#bulkStateExisting, QLabel#bulkStateDuplicate, QLabel#bulkStateEmpty {
    border-radius:13px; padding:4px 8px; font-size:10px; font-weight:800;
}
QLabel#bulkStateReady { background:#DCFCE7; color:#12813C; }
QLabel#bulkStateExisting { background:#EFF6FF; color:#2356B2; }
QLabel#bulkStateDuplicate { background:#FFF7ED; color:#B45309; }
QLabel#bulkStateEmpty { background:#EEF2F7; color:#7A8AA3; }
QFrame#bulkDropZone {
    background:#FFFFFF; border:2px dashed #BDCCE0; border-radius:14px;
}
QFrame#bulkDropZone[dragActive="true"] { background:#F4F8FF; border-color:#6D93EB; }
QLabel#bulkDropIcon {
    background:#EDF4FF; color:#2E5FC8; border-radius:16px; font-size:27px; font-weight:800;
}
QLabel#bulkDropTitle { color:#152238; background:transparent; font-size:16px; font-weight:800; }
QLabel#bulkDropDescription { color:#718096; background:transparent; font-size:12px; }
QPushButton#bulkChooseFile {
    background:#3B6FE8; color:#FFFFFF; border:0; border-radius:9px;
    padding:0 16px; font-size:12px; font-weight:800;
}
QPushButton#bulkChooseFile:hover { background:#2F5FD1; }
QLabel#bulkFileMeta {
    background:#F2F7FF; color:#38577E; border:1px solid #CBDAF4;
    border-radius:10px; padding:10px 12px; font-size:11px;
}
QFrame#bulkExcelOptions { background:#FBFCFE; border:1px solid #DCE5EF; border-radius:12px; }
QLabel#bulkExcelFieldLabel { color:#526985; background:transparent; font-size:11px; font-weight:800; }
QCheckBox#bulkHeaderCheck { color:#526985; font-size:11px; font-weight:700; }
QFrame#bulkSummaryBar { background:#FBFCFE; border-top:1px solid #DCE5EF; }
QLabel#summaryNeutral, QLabel#summaryGood, QLabel#summaryWarn {
    border-radius:8px; padding:5px 10px; font-size:11px; font-weight:800;
}
QLabel#summaryNeutral { background:#EEF2F7; color:#536B86; }
QLabel#summaryGood { background:#DCFCE7; color:#16803C; }
QLabel#summaryWarn { background:#FFF7ED; color:#B45309; }
QLabel#summaryNote { color:#7B8BA1; background:transparent; font-size:11px; }
QFrame#componentEntryFooter {
    background:#FFFFFF; border-top:1px solid #DCE5EF;
    border-bottom-left-radius:17px; border-bottom-right-radius:17px;
}
QPushButton#componentCancel {
    background:#FFFFFF; color:#40536C; border:1px solid #CBD7E6;
    border-radius:9px; font-size:12px; font-weight:800;
}
QPushButton#componentCancel:hover { background:#F8FAFC; }
QPushButton#componentPrimary {
    background:#3B6FE8; color:#FFFFFF; border:0; border-radius:9px;
    padding:0 16px; font-size:12px; font-weight:800;
}
QPushButton#componentPrimary:hover { background:#2F5FD1; }
QPushButton#componentPrimary:disabled { background:#AFC0E9; color:#F5F7FB; }
QScrollBar:vertical { background:#F1F5F9; width:10px; }
QScrollBar::handle:vertical { background:#CBD5E1; border-radius:5px; min-height:24px; }
QScrollBar:horizontal { background:#F1F5F9; height:10px; }
QScrollBar::handle:horizontal { background:#CBD5E1; border-radius:5px; min-width:24px; }
"""
        )
