from __future__ import annotations

import json
import math
import re
import unicodedata
from copy import deepcopy
from dataclasses import dataclass
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.marker import DataPoint
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

from .analysis_custom_dashboard import (
    CUSTOM_ANALYSIS_DASHBOARD_SOURCE_ID,
    CustomAnalysisDashboardController,
)
from .analysis_dashboard_layout import DashboardCardPlacement, placement_order
from .analysis_dashboard_workspace import DashboardWorkspace
from .analysis_models import AnalysisCard, CardType, ChartType, DashboardItem
from .analysis_registry import AnalysisRegistry
from .analysis_utils import parse_date, parse_datetime
from .analysis_visual_settings import (
    AnalysisVisualSettings,
    format_kpi_value,
    palette_colors,
)
from .analysis_widgets import chart_series, table_rows


_INVALID_SHEET_CHARS_RE = re.compile(r"[\\/*?:\[\]]+")
_INVALID_FILENAME_CHARS_RE = re.compile(r'[<>:"/\\|?*]+')
_ILLEGAL_XML_CHARS_RE = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")
_EXCEL_CELL_TEXT_LIMIT = 32_767

_DARK = "263341"
_BLUE = "1F5BE3"
_LIGHT_BLUE = "E8F0FE"
_HEADER = "DDE7F2"
_BORDER = "D8E2ED"
_TEXT = "0F172A"
_MUTED = "64748B"
_WHITE = "FFFFFF"

_SOURCE_TITLES = {"prepared": "Hazır Analiz", "custom": "Özel Analiz"}
_VISUALIZATION_TITLES = {
    (CardType.KPI, ChartType.NONE): "KPI",
    (CardType.CHART, ChartType.BAR): "Dikey Çubuk",
    (CardType.CHART, ChartType.HORIZONTAL_BAR): "Yatay Çubuk",
    (CardType.CHART, ChartType.DONUT): "Donut",
    (CardType.CHART, ChartType.LINE): "Çizgi",
    (CardType.TABLE, ChartType.NONE): "Tablo",
}
_FALLBACK_COLUMN_TITLES = {
    "platform": "Platform",
    "contract_no": "Sözleşme No",
    "contract_type": "Sözleşme Tipi",
    "entity": "Varlık",
    "name": "Ad",
    "system_name": "Sistem",
    "due_date": "Termin",
    "days": "Gün",
    "status": "Durum",
    "planned_acceptance_date": "Planlanan Kabul",
    "acceptance_date": "Teslimat Tarihi",
    "planned_total": "Planlanan",
    "delivered_total": "Teslim Edilen",
    "completed": "Tamamlandı",
    "date_field": "Termin Alanı",
    "raw_date_value": "Ham Tarih",
    "date_status": "Tarih Durumu",
    "source_type": "Kaynak Türü",
}


class DashboardExcelExportError(RuntimeError):
    pass


@dataclass(frozen=True, slots=True)
class DashboardExportItem:
    placement: DashboardCardPlacement
    card: AnalysisCard
    source_kind: str


@dataclass(frozen=True, slots=True)
class DashboardExportIssue:
    placement_id: str
    card_key: str
    message: str
    source_kind: str


@dataclass(frozen=True, slots=True)
class DashboardExportCollection:
    items: tuple[DashboardExportItem, ...]
    issues: tuple[DashboardExportIssue, ...]


@dataclass(frozen=True, slots=True)
class DashboardExcelExportResult:
    output_path: Path
    exported_card_count: int
    warning_count: int
    sheet_count: int


class DashboardExportCollector:
    """Resolve Dashboard cards through the same prepared/custom model pipeline as Qt."""

    def __init__(self, custom_controller: CustomAnalysisDashboardController):
        self.custom_controller = custom_controller

    def collect(
        self,
        workspace: DashboardWorkspace,
        dashboard_items: Iterable[DashboardItem],
    ) -> DashboardExportCollection:
        working = workspace.working_copy()
        custom_cards, custom_issues = self.custom_controller.resolve_pinned_cards(working)
        cards, missing = working.resolve_cards(dashboard_items, additional_cards=custom_cards)

        placement_by_id = {placement.placement_id: placement for placement in working.placements}
        items: list[DashboardExportItem] = []
        for card in cards:
            placement_id = str(card.meta.get("dashboard_placement_id") or "")
            placement = placement_by_id.get(placement_id)
            if placement is None:
                continue
            source_kind = (
                "custom"
                if placement.source_screen_id == CUSTOM_ANALYSIS_DASHBOARD_SOURCE_ID
                else "prepared"
            )
            items.append(DashboardExportItem(placement, deepcopy(card), source_kind))
        items.sort(key=lambda item: placement_order(item.placement))

        issues: list[DashboardExportIssue] = []
        custom_issue_keys: set[str] = set()
        for issue in custom_issues:
            key = f"{CUSTOM_ANALYSIS_DASHBOARD_SOURCE_ID}:{issue.analysis_id}"
            custom_issue_keys.add(key)
            placements = [
                placement
                for placement in working.placements
                if placement.source_screen_id == CUSTOM_ANALYSIS_DASHBOARD_SOURCE_ID
                and placement.card_id == issue.analysis_id
            ]
            for placement in placements:
                issues.append(
                    DashboardExportIssue(
                        placement_id=placement.placement_id,
                        card_key=key,
                        message=issue.message,
                        source_kind="custom",
                    )
                )
        for key in missing:
            if key in custom_issue_keys:
                continue
            source_screen_id, _, _card_id = key.partition(":")
            placement = next(
                (
                    candidate
                    for candidate in working.placements
                    if f"{candidate.source_screen_id}:{candidate.card_id}" == key
                ),
                None,
            )
            issues.append(
                DashboardExportIssue(
                    placement_id=placement.placement_id if placement is not None else "",
                    card_key=key,
                    message="Dashboard kartı mevcut analiz setinde çözümlenemedi.",
                    source_kind=(
                        "custom"
                        if source_screen_id == CUSTOM_ANALYSIS_DASHBOARD_SOURCE_ID
                        else "prepared"
                    ),
                )
            )
        issues.sort(key=lambda issue: _issue_order(issue, placement_by_id))
        return DashboardExportCollection(tuple(items), tuple(issues))


def export_dashboard_excel(
    output_path: Path | str,
    *,
    workspace: DashboardWorkspace,
    dashboard_items: Iterable[DashboardItem],
    custom_controller: CustomAnalysisDashboardController,
    source: Any = None,
    exported_at: datetime | None = None,
) -> DashboardExcelExportResult:
    collection = DashboardExportCollector(custom_controller).collect(workspace, dashboard_items)
    return export_dashboard_collection_excel(
        output_path,
        collection=collection,
        registry=custom_controller.service.registry,
        source=source,
        exported_at=exported_at,
        workspace_card_count=len(workspace.placements),
    )


def export_dashboard_collection_excel(
    output_path: Path | str,
    *,
    collection: DashboardExportCollection,
    registry: AnalysisRegistry,
    source: Any = None,
    exported_at: datetime | None = None,
    workspace_card_count: int | None = None,
) -> DashboardExcelExportResult:
    """Write a pre-resolved Dashboard snapshot without touching Qt or source data."""

    if not collection.items:
        raise DashboardExcelExportError("Dashboard'da aktarılabilir kart bulunamadı.")

    path = ensure_xlsx_path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    now = exported_at or datetime.now()
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Dashboard Özeti"

    used_sheet_names = {summary.title.casefold()}
    detail_names: list[str] = []
    for item in collection.items:
        detail_name = safe_unique_sheet_name(item.card.title, used_sheet_names)
        detail_names.append(detail_name)
        sheet = workbook.create_sheet(detail_name)
        if item.card.card_type == CardType.KPI:
            _write_kpi_sheet(sheet, item, registry)
        elif item.card.card_type == CardType.CHART:
            _write_chart_sheet(sheet, item)
        elif item.card.card_type == CardType.TABLE:
            _write_table_sheet(sheet, item, registry)
        else:
            _write_generic_sheet(sheet, item)

    _write_summary_sheet(
        summary,
        collection=collection,
        detail_names=detail_names,
        source=source,
        exported_at=now,
        workspace_card_count=(
            int(workspace_card_count)
            if workspace_card_count is not None
            else len(collection.items) + len(collection.issues)
        ),
    )
    try:
        workbook.save(path)
    except (OSError, PermissionError) as exc:
        raise DashboardExcelExportError(f"Excel dosyası kaydedilemedi: {path}") from exc
    return DashboardExcelExportResult(
        output_path=path,
        exported_card_count=len(collection.items),
        warning_count=len(collection.issues),
        sheet_count=len(workbook.sheetnames),
    )

def ensure_xlsx_path(value: Path | str) -> Path:
    path = Path(value)
    if path.suffix.casefold() != ".xlsx":
        path = path.with_suffix(".xlsx")
    return path


def suggested_dashboard_excel_path(source: Any, *, when: datetime | None = None) -> Path:
    source_path = _source_path(source)
    base_name = safe_filename_part(source_path.stem if source_path is not None else "STS", "STS")
    stamp = (when or datetime.now()).strftime("%Y-%m-%d_%H-%M")
    filename = f"{base_name}__ANALIZ_DASHBOARD__{stamp}.xlsx"
    folder = source_path.parent if source_path is not None else Path.cwd()
    return folder / filename


def safe_filename_part(value: object, fallback: str = "DOSYA") -> str:
    text = str(value or "").strip() or str(fallback or "DOSYA")
    replacements = {
        "ı": "i", "İ": "I", "ş": "s", "Ş": "S", "ğ": "g", "Ğ": "G",
        "ü": "u", "Ü": "U", "ö": "o", "Ö": "O", "ç": "c", "Ç": "C",
    }
    for src, dst in replacements.items():
        text = text.replace(src, dst)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(char for char in text if not unicodedata.combining(char))
    text = _INVALID_FILENAME_CHARS_RE.sub("_", text)
    text = re.sub(r"\s+", "_", text)
    text = re.sub(r"_+", "_", text).strip("._ ")
    return text or str(fallback or "DOSYA")


def safe_unique_sheet_name(title: object, used_names: set[str]) -> str:
    raw = _INVALID_SHEET_CHARS_RE.sub(" ", _excel_safe_text(title).strip())
    base = " ".join(raw.split()).strip(" '") or "Analiz"
    base = base[:31]
    candidate = base
    index = 2
    while candidate.casefold() in used_names:
        suffix = f" ({index})"
        candidate = f"{base[: max(1, 31 - len(suffix))]}{suffix}"
        index += 1
    used_names.add(candidate.casefold())
    return candidate


def card_display_value(card: AnalysisCard) -> str:
    visual = card.meta.get("visual_settings")
    if card.meta.get("visual_settings_enabled") and isinstance(visual, AnalysisVisualSettings):
        rendered = format_kpi_value(card.value, visual.kpi)
    else:
        rendered = _display_value(card.value)
    unit = str(card.unit or "").strip()
    return _excel_safe_text(f"{rendered} {unit}" if unit else rendered)


def _write_summary_sheet(
    sheet,
    *,
    collection: DashboardExportCollection,
    detail_names: Sequence[str],
    source: Any,
    exported_at: datetime,
    workspace_card_count: int,
) -> None:
    sheet.sheet_view.showGridLines = False
    sheet.merge_cells("A1:F1")
    sheet["A1"] = "Analiz Merkezi — Dashboard Excel Aktarımı"
    _style_title(sheet["A1"])
    sheet.row_dimensions[1].height = 28

    metadata = [
        ("Kaynak STS", _source_name(source)),
        ("Aktarım Tarihi", exported_at),
        ("Dashboard Kart Sayısı", workspace_card_count),
        ("Aktarılan Kart Sayısı", len(collection.items)),
    ]
    for row_index, (label, value) in enumerate(metadata, start=3):
        sheet.cell(row_index, 1, label)
        sheet.cell(row_index, 2, value)
        sheet.cell(row_index, 1).font = Font(bold=True, color=_TEXT)
        sheet.cell(row_index, 1).fill = PatternFill("solid", fgColor=_LIGHT_BLUE)
        sheet.cell(row_index, 1).border = _thin_border()
        sheet.cell(row_index, 2).border = _thin_border()
    sheet["B4"].number_format = "dd.mm.yyyy hh:mm"

    start_row = 9
    headers = ["Sıra", "Analiz", "Kaynak", "Görünüm", "Özet", "Detay Sayfası"]
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(start_row, column, header)
        _style_table_header(cell)
    for index, (item, detail_name) in enumerate(zip(collection.items, detail_names), start=1):
        row = start_row + index
        values = [
            index,
            item.card.title,
            _SOURCE_TITLES[item.source_kind],
            _visualization_title(item.card),
            _summary_text(item.card),
            detail_name,
        ]
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row, column, _excel_cell_value(value))
            cell.border = _thin_border()
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        link = sheet.cell(row, 6)
        escaped = detail_name.replace("'", "''")
        link.hyperlink = f"#'{escaped}'!A1"
        link.style = "Hyperlink"

    if collection.issues:
        issue_title_row = start_row + len(collection.items) + 3
        sheet.merge_cells(start_row=issue_title_row, start_column=1, end_row=issue_title_row, end_column=6)
        cell = sheet.cell(issue_title_row, 1, "Aktarım Uyarıları")
        cell.fill = PatternFill("solid", fgColor="FFF7ED")
        cell.font = Font(bold=True, color="9A3412")
        for offset, issue in enumerate(collection.issues, start=1):
            row = issue_title_row + offset
            sheet.cell(row, 1, offset)
            sheet.cell(row, 2, _excel_safe_text(_SOURCE_TITLES.get(issue.source_kind, issue.source_kind)))
            sheet.cell(row, 3, _excel_safe_text(issue.card_key))
            sheet.cell(row, 4, "Aktarılamadı")
            sheet.cell(row, 5, _excel_safe_text(issue.message))
            sheet.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
            for column in range(1, 7):
                issue_cell = sheet.cell(row, column)
                issue_cell.border = _thin_border()
                issue_cell.alignment = Alignment(vertical="top", wrap_text=True)

    sheet.freeze_panes = "A10"
    widths = {"A": 8, "B": 34, "C": 18, "D": 18, "E": 26, "F": 30}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width


def _write_kpi_sheet(
    sheet,
    item: DashboardExportItem,
    registry: AnalysisRegistry,
) -> None:
    card = item.card
    _prepare_detail_sheet(sheet, card.title)
    if card.subtitle:
        sheet.merge_cells("A3:H3")
        sheet["A3"] = _excel_safe_text(card.subtitle)
        sheet["A3"].font = Font(italic=True, color=_MUTED, size=11)
    sheet.merge_cells("A5:H8")
    raw_numeric = _numeric_value(card.value)
    sheet["A5"] = raw_numeric if raw_numeric is not None else card_display_value(card)
    sheet["A5"].font = Font(bold=True, size=28, color=_BLUE)
    sheet["A5"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    visual = card.meta.get("visual_settings")
    if raw_numeric is not None:
        unit_suffix = f" {str(card.unit).strip()}" if str(card.unit or "").strip() else ""
        if card.meta.get("visual_settings_enabled") and isinstance(visual, AnalysisVisualSettings):
            settings = visual.kpi
            sheet["A5"].number_format = _excel_kpi_number_format(
                settings.prefix,
                f"{settings.suffix}{unit_suffix}",
                settings.decimal_places,
            )
        else:
            places = 0 if isinstance(raw_numeric, int) or float(raw_numeric).is_integer() else 2
            sheet["A5"].number_format = _excel_kpi_number_format("", unit_suffix, places)
    sheet["A10"] = "Görünen Değer"
    sheet["B10"] = card_display_value(card)
    _style_metadata_pair(sheet, 10)
    metadata = [
        ("Veri Kaynağı", _dataset_title(registry, card)),
        ("Görünüm", _visualization_title(card)),
        ("Analiz Türü", _SOURCE_TITLES[item.source_kind]),
    ]
    for row, (label, value) in enumerate(metadata, start=12):
        sheet.cell(row, 1, label)
        sheet.cell(row, 2, _excel_cell_value(value))
        _style_metadata_pair(sheet, row)
    sheet.column_dimensions["A"].width = 22
    sheet.column_dimensions["B"].width = 32


def _write_chart_sheet(sheet, item: DashboardExportItem) -> None:
    card = item.card
    _prepare_detail_sheet(sheet, card.title)
    rows = chart_series(card.data)
    sheet["A4"] = "Kategori"
    sheet["B4"] = "Değer"
    _style_table_header(sheet["A4"])
    _style_table_header(sheet["B4"])
    for index, row in enumerate(rows, start=5):
        sheet.cell(index, 1, _excel_safe_text(row.get("label") or "Eksik"))
        value = _numeric_value(row.get("value"))
        sheet.cell(index, 2, value if value is not None else _excel_cell_value(row.get("value")))
        sheet.cell(index, 1).border = _thin_border()
        sheet.cell(index, 2).border = _thin_border()

    if rows:
        chart = _create_native_chart(card)
        data = Reference(sheet, min_col=2, min_row=4, max_row=4 + len(rows))
        categories = Reference(sheet, min_col=1, min_row=5, max_row=4 + len(rows))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.title = _excel_safe_text(card.title)
        chart.height = 10
        chart.width = 18
        _apply_chart_visual_settings(chart, card, len(rows))
        sheet.add_chart(chart, "D4")
    sheet.freeze_panes = "A5"
    sheet.column_dimensions["A"].width = 30
    sheet.column_dimensions["B"].width = 16


def _write_table_sheet(
    sheet,
    item: DashboardExportItem,
    registry: AnalysisRegistry,
) -> None:
    card = item.card
    _prepare_detail_sheet(sheet, card.title)
    columns = list(card.columns or [])
    rows = table_rows(card.data, columns)
    dataset_id = str(card.meta.get("dataset") or "")
    field_definitions = [_field_definition(registry, dataset_id, column) for column in columns]
    headers = [
        field.title if field is not None else _FALLBACK_COLUMN_TITLES.get(column, column)
        for field, column in zip(field_definitions, columns)
    ]
    header_row = 3
    for column_index, header in enumerate(headers, start=1):
        _style_table_header(sheet.cell(header_row, column_index, _excel_safe_text(header)))
    for row_index, row in enumerate(rows, start=header_row + 1):
        for column_index, column in enumerate(columns, start=1):
            field = field_definitions[column_index - 1]
            value = _excel_cell_value(
                row.get(column),
                field_type=field.field_type if field is not None else "",
            )
            cell = sheet.cell(row_index, column_index, value)
            cell.border = _thin_border()
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if isinstance(value, datetime):
                cell.number_format = "dd.mm.yyyy hh:mm"
            elif isinstance(value, date):
                cell.number_format = "dd.mm.yyyy"
    if columns and rows:
        table_name = _unique_table_name(f"DashboardTable_{item.placement.placement_id}")
        table = Table(displayName=table_name, ref=f"A{header_row}:{_column_letter(len(columns))}{header_row + len(rows)}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)
    sheet.freeze_panes = f"A{header_row + 1}"
    for index, column in enumerate(columns, start=1):
        max_length = len(str(headers[index - 1]))
        for row in rows[:200]:
            max_length = max(max_length, len(str(row.get(column) or "")))
        sheet.column_dimensions[_column_letter(index)].width = min(40, max(10, max_length + 2))


def _write_generic_sheet(sheet, item: DashboardExportItem) -> None:
    _prepare_detail_sheet(sheet, item.card.title)
    sheet["A3"] = "Bu kart türü için yapılandırılmış KPI, grafik veya tablo verisi bulunmuyor."


def _prepare_detail_sheet(sheet, title: str) -> None:
    sheet.sheet_view.showGridLines = False
    sheet.merge_cells("A1:H1")
    sheet["A1"] = _excel_safe_text(title)
    _style_title(sheet["A1"])
    sheet.row_dimensions[1].height = 28


def _create_native_chart(card: AnalysisCard):
    if card.chart_type == ChartType.HORIZONTAL_BAR:
        chart = BarChart()
        chart.type = "bar"
        chart.style = 10
        return chart
    if card.chart_type == ChartType.DONUT:
        chart = DoughnutChart()
        chart.holeSize = 55
        return chart
    if card.chart_type == ChartType.LINE:
        return LineChart()
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    return chart


def _apply_chart_visual_settings(chart, card: AnalysisCard, row_count: int) -> None:
    visual = card.meta.get("visual_settings")
    enabled = bool(card.meta.get("visual_settings_enabled")) and isinstance(visual, AnalysisVisualSettings)
    if enabled:
        settings = visual.chart
        show_legend = settings.show_legend and card.chart_type != ChartType.LINE
        legend_position = settings.legend_position
        show_values = settings.show_values
        colors = palette_colors(settings.palette)
    else:
        show_legend = card.chart_type == ChartType.DONUT
        legend_position = "right"
        show_values = card.chart_type in {ChartType.BAR, ChartType.HORIZONTAL_BAR, ChartType.DONUT}
        colors = palette_colors("corporate")

    if not show_legend:
        chart.legend = None
    elif chart.legend is not None:
        chart.legend.position = "b" if legend_position == "bottom" else "r"
    if show_values:
        chart.dLbls = DataLabelList()
        chart.dLbls.showVal = True

    if not chart.series:
        return
    series = chart.series[0]
    if card.chart_type == ChartType.DONUT:
        series.data_points = [
            DataPoint(idx=index, spPr=GraphicalProperties(solidFill=_excel_color(colors[index % len(colors)])))
            for index in range(row_count)
        ]
    else:
        series.graphicalProperties.solidFill = _excel_color(colors[0])
        series.graphicalProperties.line.solidFill = _excel_color(colors[0])


def _dataset_title(
    registry: AnalysisRegistry,
    card: AnalysisCard,
) -> str:
    dataset_id = str(card.meta.get("dataset") or "")
    if dataset_id:
        try:
            return registry.get_dataset(dataset_id).title
        except Exception:
            return dataset_id
    return "-"


def _field_definition(
    registry: AnalysisRegistry,
    dataset_id: str,
    column: str,
):
    if dataset_id:
        try:
            return registry.get_field(dataset_id, column)
        except Exception:
            return None
    return None


def _summary_text(card: AnalysisCard) -> str:
    if card.card_type == CardType.KPI:
        return card_display_value(card)
    if card.card_type == CardType.CHART:
        return f"{len(chart_series(card.data))} kategori"
    if card.card_type == CardType.TABLE:
        return f"{len(table_rows(card.data, card.columns))} satır"
    return "Kart"


def _visualization_title(card: AnalysisCard) -> str:
    return _VISUALIZATION_TITLES.get((card.card_type, card.chart_type), card.card_type.value)


def _display_value(value: Any) -> str:
    if value is None or value == "":
        return "-"
    if isinstance(value, bool):
        return "Evet" if value else "Hayır"
    if isinstance(value, float):
        if not math.isfinite(value):
            return "-"
        if value.is_integer():
            return f"{int(value):,}".replace(",", ".")
        return f"{value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    if isinstance(value, int):
        return f"{value:,}".replace(",", ".")
    return str(value)


def _numeric_value(value: Any) -> int | float | None:
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        return None
    if isinstance(value, float) and not math.isfinite(value):
        return None
    return value


def _excel_kpi_number_format(prefix: str, suffix: str, decimal_places: int) -> str:
    decimals = "" if decimal_places <= 0 else "." + ("0" * decimal_places)
    base = f"#,##0{decimals}"
    prefix_escaped = _excel_safe_text(prefix).replace('"', '""')
    suffix_escaped = _excel_safe_text(suffix).replace('"', '""')
    if prefix_escaped:
        base = f'"{prefix_escaped}"{base}'
    if suffix_escaped:
        base = f'{base}"{suffix_escaped}"'
    return base


def _excel_safe_text(value: object) -> str:
    text = str(value or "")
    text = _ILLEGAL_XML_CHARS_RE.sub("", text)
    return text[:_EXCEL_CELL_TEXT_LIMIT]


def _excel_datetime_value(value: datetime) -> datetime:
    if value.utcoffset() is None:
        return value
    return value.astimezone(timezone.utc).replace(tzinfo=None)


def _excel_cell_value(value: Any, *, field_type: str = "") -> Any:
    if value is None:
        return None
    if field_type == "date":
        parsed = parse_date(value)
        if parsed is not None:
            return parsed
    elif field_type == "datetime":
        parsed = parse_datetime(value)
        if parsed is not None:
            return _excel_datetime_value(parsed)
    if isinstance(value, float) and not math.isfinite(value):
        return None
    if isinstance(value, str):
        return _excel_safe_text(value)
    if isinstance(value, datetime):
        return _excel_datetime_value(value)
    if isinstance(value, (int, float, bool, date)):
        return value
    if isinstance(value, (Mapping, list, tuple, set)):
        try:
            if isinstance(value, set):
                value = sorted(value, key=str)
            return _excel_safe_text(
                json.dumps(value, ensure_ascii=False, sort_keys=True, default=str)
            )
        except Exception:
            return _excel_safe_text(value)
    return _excel_safe_text(value)


def _style_title(cell) -> None:
    cell.fill = PatternFill("solid", fgColor=_DARK)
    cell.font = Font(bold=True, size=16, color=_WHITE)
    cell.alignment = Alignment(vertical="center")


def _style_table_header(cell) -> None:
    cell.fill = PatternFill("solid", fgColor=_HEADER)
    cell.font = Font(bold=True, color=_TEXT)
    cell.border = _thin_border()
    cell.alignment = Alignment(vertical="center", wrap_text=True)


def _style_metadata_pair(sheet, row: int) -> None:
    sheet.cell(row, 1).font = Font(bold=True, color=_TEXT)
    sheet.cell(row, 1).fill = PatternFill("solid", fgColor=_LIGHT_BLUE)
    sheet.cell(row, 1).border = _thin_border()
    sheet.cell(row, 2).border = _thin_border()


def _thin_border() -> Border:
    side = Side(style="thin", color=_BORDER)
    return Border(left=side, right=side, top=side, bottom=side)


def _source_path(source: Any) -> Path | None:
    if isinstance(source, (str, Path)):
        try:
            return Path(source)
        except Exception:
            return None
    for attr in ("path", "db_path", "database_path"):
        value = getattr(source, attr, None)
        if value:
            try:
                return Path(value)
            except Exception:
                return None
    return None


def _source_name(source: Any) -> str:
    source_path = _source_path(source)
    if source_path is not None:
        return source_path.name
    return "STS veri kaynağı"


def _issue_order(
    issue: DashboardExportIssue,
    placement_by_id: Mapping[str, DashboardCardPlacement],
) -> tuple[int, int, str]:
    placement = placement_by_id.get(issue.placement_id)
    if placement is None:
        return (10**9, 10**9, issue.card_key)
    return placement_order(placement)


def _excel_color(value: str) -> str:
    return str(value or "").strip().lstrip("#").upper() or "1F5BE3"


def _unique_table_name(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9_]", "_", value)
    if not cleaned or cleaned[0].isdigit():
        cleaned = f"T_{cleaned}"
    return cleaned[:240]


def _column_letter(index: int) -> str:
    value = index
    letters = ""
    while value:
        value, remainder = divmod(value - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters or "A"


__all__ = [
    "DashboardExcelExportError",
    "DashboardExcelExportResult",
    "DashboardExportCollection",
    "DashboardExportCollector",
    "DashboardExportIssue",
    "DashboardExportItem",
    "card_display_value",
    "ensure_xlsx_path",
    "export_dashboard_collection_excel",
    "export_dashboard_excel",
    "safe_filename_part",
    "safe_unique_sheet_name",
    "suggested_dashboard_excel_path",
]
