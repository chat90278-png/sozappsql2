# -*- coding: utf-8 -*-
from __future__ import annotations

import sys
import re
import ctypes
import getpass
import os
import base64
import copy
import time
import traceback
import tempfile
import zipfile
import sqlite3
import unicodedata
import logging
from pathlib import Path
from datetime import date, datetime, timedelta
from typing import Callable, Dict, List, Optional, Protocol, Tuple
from auto_accept import open_auto_accept_dialog

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.domain.constants import (
    CORE_SHEETS,
    HEADER_ROW,
    SUBHEADER_ROW,
    DATA_START_ROW,
    STATUS_VALUES,
)
from src.config.app_config import (
    APP_TITLE,
    APP_ICON_PATH,
    APP_ICON_ICO_PATH,
    APP_ID,
    DEFAULT_FILE,
    COMP_SHEET,
    USERS_SHEET,
    PLATFORM_LOGO_SHEET,
    TAG_SHEET,
    TAG_KIND_DEF,
    TAG_KIND_ASSIGN,
    LOG_FOLDER_NAME,
    NAVY,
    LIGHT,
    CARD,
    HEAD,
    BLUE,
    GREEN,
    GRID,
    TEXT_MUTED,
    BASE_HEADERS,
    MAIN_TOTAL_LABEL,
    SYSTEM_TOTAL_SUFFIX,
    TR_MONTHS,
    TR_WEEKDAYS,
    LOG_HEADERS,
    TAG_HEADERS,
    EXTRA_SYSTEM_SHEET_NAMES,
)
from src.models.app_models import ComponentDef, ContractInfo, SystemInfo, DeliveryInfo, TagDef
from src.domain.contract_timing import contract_timing, is_completed_status
from src.domain.delivery_coverage import acceptance_coverage_issues
from src.domain.flexible_date import flexible_or_blank, is_exact_date, parse_flexible_date, validate_flexible_date, format_flexible_date
from src.ui.widgets import stat_card, set_card_value
from src.ui.theme import STYLE
from src.ui.tarih import ContractCalendarWindow
from src.ui.ozet import ContractSummaryDialog
from src.ui.date_picker import build_date_input as _build_date_input
from src.ui.kullanim_kilavuzu import UsageGuideDialog
from src.ui.dialogs.platform_component_manager import PlatformComponentManagerDialog
from src.ui.dialogs.delivery_schedule_report_dialog import DeliveryScheduleReportDialog
from src.ui.dialogs.platform_delivery_report_dialog import PlatformTeslimatDurumuReportDialog
from src.ui.message_boxes import ask_yes_no

from PySide6.QtCore import Qt, QDate, QObject, QThread, Signal, QTimer, QPoint, QSize, QRect, QEvent, QPropertyAnimation, QEasingCurve, QUrl
from PySide6.QtGui import QFont, QFontMetrics, QColor, QPixmap, QIcon, QPainter, QAction, QCursor, QCloseEvent, QDesktopServices, QKeySequence, QShortcut
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget,QGraphicsOpacityEffect, QGraphicsDropShadowEffect, QVBoxLayout, QHBoxLayout, QGridLayout,
    QLabel, QPushButton, QListWidget, QListWidgetItem, QTableWidget, QTableWidgetItem,
    QTreeWidget, QTreeWidgetItem, QDialog, QLineEdit, QComboBox, QDateEdit, QSpinBox, QDoubleSpinBox,
    QMessageBox, QFileDialog, QFrame, QScrollArea, QCheckBox, QHeaderView,
    QSizePolicy, QProgressBar, QProgressDialog, QStyledItemDelegate, QTextEdit,
    QToolButton, QMenu, QInputDialog, QWidgetAction, QStackedWidget, QAbstractItemView, QStyle, QRadioButton, QButtonGroup
)
from shiboken6 import isValid as _qt_is_valid


def qt_obj_alive(obj) -> bool:
    """PySide/Shiboken objeleri C++ tarafında silindiyse False döner.

    Qt'nin C++ tarafından çağırdığı eventFilter/sizeHint gibi override'larda
    silinmiş objeye dokunmak RuntimeError üretir; PySide bunu yakalayamazsa
    uygulama `Fatal Python error: Aborted` ile kapanabilir.
    """
    if obj is None:
        return False
    try:
        return bool(_qt_is_valid(obj))
    except Exception:
        return True


def normalized_tag_key(value: str) -> str:
    """Return a stable comparison key for tag names, including Turkish case variants."""
    text = str(value or "").strip().replace("ı", "i").replace("İ", "i")
    text = unicodedata.normalize("NFKD", text.casefold())
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return " ".join(text.split())


def _share_metadata_from_path(path: Path | str) -> dict:
    """Return share metadata for STS share packages; empty dict for normal files."""
    try:
        p = Path(path)
        if not p.exists() or p.suffix.lower() != ".sts":
            return {}
        conn = sqlite3.connect(str(p))
        conn.row_factory = sqlite3.Row
        try:
            row = conn.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='share_metadata'").fetchone()
            if not row:
                return {}
            rows = conn.execute("SELECT key,value FROM share_metadata").fetchall()
            meta = {str(r["key"]): str(r["value"] or "") for r in rows}
            return meta if str(meta.get("share_mode", "")).lower() == "true" else {}
        finally:
            conn.close()
    except Exception:
        return {}


def _write_share_metadata(path: Path | str, metadata: dict) -> None:
    conn = sqlite3.connect(str(path))
    try:
        conn.execute("CREATE TABLE IF NOT EXISTS share_metadata(key TEXT PRIMARY KEY, value TEXT)")
        for key, value in dict(metadata or {}).items():
            conn.execute(
                "INSERT INTO share_metadata(key,value) VALUES(?,?) ON CONFLICT(key) DO UPDATE SET value=excluded.value",
                (str(key), str(value)),
            )
        conn.commit()
    finally:
        conn.close()


class ContractFileDropButton(QPushButton):
    """Clickable upload target that also accepts local file drops."""

    filesDropped = Signal(list)
    invalidDrop = Signal(str)

    def __init__(self, text: str, parent=None):
        super().__init__(text, parent)
        self._default_text = text
        self.setAcceptDrops(True)

    def dragEnterEvent(self, event):
        mime = event.mimeData()
        if mime.hasUrls():
            event.acceptProposedAction()
            self.setText("  ↓    Dosyaları buraya bırakın")
            return
        event.ignore()

    def dragMoveEvent(self, event):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
        else:
            event.ignore()

    def dragLeaveEvent(self, event):
        self.setText(self._default_text)
        super().dragLeaveEvent(event)

    def dropEvent(self, event):
        self.setText(self._default_text)
        mime = event.mimeData()
        if not mime.hasUrls():
            self.invalidDrop.emit("Yalnızca yerel dosyalar yüklenebilir.")
            event.ignore()
            return
        file_paths = []
        folder_paths = []
        unsupported_url = False
        for url in mime.urls():
            if not url.isLocalFile():
                unsupported_url = True
                continue
            path = Path(url.toLocalFile())
            if path.is_dir():
                folder_paths.append(str(path))
            else:
                file_paths.append(str(path))
        if unsupported_url:
            self.invalidDrop.emit("Web bağlantısı yüklenemez, lütfen yerel dosya seçin.")
        if folder_paths:
            # Klasör sürükle-bırak: parent widget'a ilet
            parent = self.parent()
            while parent is not None:
                if hasattr(parent, "_import_contract_folders"):
                    parent._import_contract_folders(folder_paths, parent_folder_id=None)
                    break
                parent = parent.parent() if hasattr(parent, "parent") else None
        if file_paths:
            self.filesDropped.emit(file_paths)
        if file_paths or folder_paths:
            event.acceptProposedAction()
        else:
            event.ignore()


class ContractFileTreeWidget(QTreeWidget):
    """Folder-aware document tree with external file drop support."""

    filesDropped = Signal(list, object)
    invalidDrop = Signal(str)

    # Signal: (item_kind, item_id, target_folder_id_or_None)
    itemMoved = Signal(str, int, object)

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setAcceptDrops(True)
        self.setDragEnabled(True)
        self.setDragDropMode(QAbstractItemView.DragDrop)
        self.setDropIndicatorShown(True)
        self.setHeaderHidden(True)
        self.setRootIsDecorated(True)
        self.setAlternatingRowColors(False)
        self.setSelectionMode(QAbstractItemView.SingleSelection)
        self.setEditTriggers(
            QAbstractItemView.EditKeyPressed
            | QAbstractItemView.SelectedClicked
        )
        self.setContextMenuPolicy(Qt.CustomContextMenu)
        self._drag_item_kind = None
        self._drag_item_id = None

    def _drop_folder_id(self, pos):
        item = self.itemAt(pos)
        if item is None:
            return None
        kind = item.data(0, Qt.UserRole)
        if kind == "folder":
            return item.data(0, Qt.UserRole + 1)
        if kind == "file":
            return item.data(0, Qt.UserRole + 2)
        return None

    def _local_file_paths_from_event(self, event):
        """Dosya ve klasör yollarını ayrı listeler halinde döner: (file_paths, folder_paths, error)"""
        mime = event.mimeData()
        if not mime.hasUrls():
            return [], [], "Yalnızca yerel dosyalar yüklenebilir."
        file_paths = []
        folder_paths = []
        unsupported_url = False
        for url in mime.urls():
            if not url.isLocalFile():
                unsupported_url = True
                continue
            path = Path(url.toLocalFile())
            if path.is_dir():
                folder_paths.append(str(path))
            else:
                file_paths.append(str(path))
        if unsupported_url:
            self.invalidDrop.emit("Web bağlantısı yüklenemez, lütfen yerel dosya seçin.")
        if not file_paths and not folder_paths:
            return [], [], "Yüklenecek yerel dosya veya klasör bulunamadı."
        return file_paths, folder_paths, ""

    def _get_dragged_item_info(self):
        """Sürüklenen item'ın kind ve id'sini döner."""
        item = self.currentItem()
        if not item:
            return None, None
        kind = item.data(0, Qt.UserRole)
        if kind in ("file", "folder"):
            return kind, int(item.data(0, Qt.UserRole + 1))
        return None, None

    def dragEnterEvent(self, event):
        if event.source() is self:
            # Internal move – kendi tree'sinden sürükleme
            kind, item_id = self._get_dragged_item_info()
            if kind in ("file", "folder"):
                self._drag_item_kind = kind
                self._drag_item_id = item_id
                event.acceptProposedAction()
                return
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
        else:
            event.ignore()

    def dragMoveEvent(self, event):
        if event.source() is self:
            event.acceptProposedAction()
        elif event.mimeData().hasUrls():
            event.acceptProposedAction()
        else:
            event.ignore()

    def dropEvent(self, event):
        pos = event.position().toPoint() if hasattr(event, "position") else event.pos()
        drop_folder_id = self._drop_folder_id(pos)

        if event.source() is self:
            # Internal move
            kind = self._drag_item_kind
            item_id = self._drag_item_id
            self._drag_item_kind = None
            self._drag_item_id = None
            if kind and item_id is not None:
                self.itemMoved.emit(kind, item_id, drop_folder_id)
            event.acceptProposedAction()
            return

        file_paths, folder_paths, error = self._local_file_paths_from_event(event)
        if not file_paths and not folder_paths:
            if error:
                self.invalidDrop.emit(error)
            event.ignore()
            return
        if folder_paths:
            parent = self.parent()
            while parent is not None:
                if hasattr(parent, "_import_contract_folders"):
                    parent._import_contract_folders(folder_paths, parent_folder_id=drop_folder_id)
                    break
                parent = parent.parent() if hasattr(parent, "parent") else None
        if file_paths:
            self.filesDropped.emit(file_paths, drop_folder_id)
        event.acceptProposedAction()

def app_icon_path() -> Path:
    """Return the native Windows icon when available, otherwise the SVG logo."""
    if APP_ICON_ICO_PATH.exists():
        return APP_ICON_ICO_PATH
    return APP_ICON_PATH


def configure_windows_app_identity() -> None:
    """Set a stable Windows AppUserModelID so taskbar icons use the STS icon."""
    if sys.platform != "win32":
        return
    try:
        ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(APP_ID)
    except Exception:
        pass

def normalize_sheet_name(name: str) -> str:
    txt = str(name or "").strip().lower()
    repl = {
        "\u0131": "i",
        "\u0130": "i",
        "\u015f": "s",
        "\u011f": "g",
        "\u00fc": "u",
        "\u00f6": "o",
        "\u00e7": "c",
    }
    for a, b in repl.items():
        txt = txt.replace(a, b)
    return txt


def is_system_sheet_name(name: str) -> bool:
    if not name:
        return True
    n = normalize_sheet_name(name)
    core_norm = {normalize_sheet_name(x) for x in CORE_SHEETS}
    if n in core_norm or n in EXTRA_SYSTEM_SHEET_NAMES:
        return True
    if str(name).startswith("_") or n.startswith("_"):
        return True
    return False


def safe_sheet_name(name: str) -> str:
    n = re.sub(r"[\\/*?:\[\]]", "_", name.strip().upper())
    return n[:31] or "PLATFORM"


def safe_filename_part(value: object, fallback: str = "DOSYA") -> str:
    """Windows dosya adına güvenli, okunabilir parça üretir."""
    text = str(value or "").strip() or str(fallback or "DOSYA")
    replacements = {
        "ı": "i", "İ": "I", "ş": "s", "Ş": "S", "ğ": "g", "Ğ": "G",
        "ü": "u", "Ü": "U", "ö": "o", "Ö": "O", "ç": "c", "Ç": "C",
    }
    for src, dst in replacements.items():
        text = text.replace(src, dst)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r'[<>:"/\\|?*]+', "_", text)
    text = re.sub(r"\s+", "_", text)
    text = re.sub(r"_+", "_", text).strip("._ ")
    return text or str(fallback or "DOSYA")


def suggested_export_excel_path(sts_path: object, options: Optional[dict] = None) -> Path:
    """Export için varsayılan dosya adı: STSADI__PLATFORM__export__yyyy-MM-dd_HH-mm.xlsx"""
    try:
        source = Path(sts_path) if sts_path else Path.cwd() / "sts_export.sts"
    except Exception:
        source = Path.cwd() / "sts_export.sts"

    base_name = safe_filename_part(source.stem, "STS")
    opts = dict(options or {})
    platforms_raw = opts.get("platforms") or []
    platforms: List[str] = []
    seen: set[str] = set()
    for item in platforms_raw:
        name = str(item or "").strip()
        key = name.casefold()
        if name and key not in seen:
            seen.add(key)
            platforms.append(name)

    scope = str(opts.get("scope") or "").strip().lower()
    if scope == "selected" and len(platforms) == 1:
        scope_name = safe_filename_part(platforms[0], "PLATFORM")
    elif scope == "selected" and len(platforms) > 1:
        scope_name = "SECILI_PLATFORMLAR"
    else:
        scope_name = "TUM_PLATFORMLAR"

    stamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
    filename = f"{base_name}__{scope_name}__export__{stamp}.xlsx"
    folder = source.parent if str(source.parent) not in {"", "."} else Path.cwd()
    return folder / filename


_STS_VERSIONED_RE = re.compile(
    r"^(?P<code>STS-[A-Z]\d+)__v(?P<ver>\d+)__(?P<stamp>\d{4}-\d{2}-\d{2}_\d{2}-\d{2})$",
    re.IGNORECASE,
)


def parse_sts_versioned_stem(stem: str) -> Tuple[str, int]:
    """STS-A1__v004__2026-06-15_10-30 -> (STS-A1, 4)."""
    raw = str(stem or "").strip()
    m = _STS_VERSIONED_RE.match(raw)
    if m:
        code = str(m.group("code") or "STS-A1").upper()
        try:
            ver = int(m.group("ver") or 0)
        except Exception:
            ver = 0
        return code, max(0, ver)
    # Eski/serbest adla açılmış dosyalar ilk kapanışta standart isme taşınır.
    return "STS-A1", 0


def next_sts_versioned_path(current_path: object) -> Path:
    """Tek aktif .sts dosyası için sonraki ad: STS-A1__v004__yyyy-MM-dd_HH-mm.sts"""
    current = Path(current_path)
    folder = current.parent if str(current.parent) not in {"", "."} else Path.cwd()
    code, current_ver = parse_sts_versioned_stem(current.stem)

    max_ver = current_ver
    try:
        for item in folder.glob(f"{code}__v*__*.sts"):
            parsed_code, parsed_ver = parse_sts_versioned_stem(item.stem)
            if parsed_code == code:
                max_ver = max(max_ver, parsed_ver)
    except Exception:
        pass

    stamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
    next_ver = max_ver + 1
    while True:
        candidate = folder / f"{code}__v{next_ver:03d}__{stamp}.sts"
        if not candidate.exists() or candidate.resolve() == current.resolve():
            return candidate
        next_ver += 1


def close_store_for_file_rename(store) -> None:
    """Windows'ta SQLite dosyasını yeniden adlandırabilmek için bağlantıyı kapatır."""
    if store is None:
        return
    try:
        if hasattr(store, "save"):
            store.save()
    except Exception:
        pass
    for close_name in ("close", "disconnect"):
        try:
            closer = getattr(store, close_name, None)
            if callable(closer):
                closer()
                return
        except Exception:
            pass
    try:
        db = getattr(store, "db", None)
        conn = getattr(db, "conn", None)
        if conn is not None:
            try:
                conn.commit()
            except Exception:
                pass
            conn.close()
    except Exception:
        pass


def rename_sts_file_to_next_version(store, current_path: object) -> Optional[Path]:
    """Değişiklik varsa kapanışta tek aktif STS dosyasını bir üst versiyon adına taşır."""
    try:
        old_path = Path(current_path)
        if old_path.suffix.lower() != ".sts" or not old_path.exists():
            return None
        new_path = next_sts_versioned_path(old_path)
        if new_path.resolve() == old_path.resolve():
            return old_path
        close_store_for_file_rename(store)
        old_path.replace(new_path)
        try:
            setattr(store, "path", new_path)
            db = getattr(store, "db", None)
            if db is not None:
                setattr(db, "path", new_path)
        except Exception:
            pass
        return new_path
    except Exception:
        return None


def to_iso(qdate: QDate) -> str:
    return f"{qdate.year():04d}-{qdate.month():02d}-{qdate.day():02d}"


def parse_iso_date(text: str) -> Optional[date]:
    text = (text or "").strip()
    if not text:
        return None
    try:
        return datetime.strptime(text, "%Y-%m-%d").date()
    except ValueError:
        return None


def add_months(d: date, months: int) -> date:
    month = d.month - 1 + int(months or 0)
    year = d.year + month // 12
    month = month % 12 + 1
    days = [31, 29 if year % 4 == 0 and (year % 100 != 0 or year % 400 == 0) else 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
    day = min(d.day, days[month - 1])
    return date(year, month, day)


def iso_or_blank(text: str) -> str:
    d = parse_iso_date(text)
    return d.isoformat() if d else ""


def contract_date_picker_events(ci: Optional[ContractInfo], systems: Optional[List[SystemInfo]] = None) -> List[dict]:
    events: List[dict] = []
    if ci:
        contract_deadline = parse_iso_date(str(getattr(ci, "completion_date", "") or ""))
        if contract_deadline:
            no = str(getattr(ci, "no", "") or "").strip() or "Sözleşme"
            ctype = str(getattr(ci, "contract_type", "") or "").strip()
            title = f"{no} {ctype}".strip()
            events.append({
                "date": contract_deadline,
                "title": title,
                "lines": [
                    f"Termin tarihi: {contract_deadline.isoformat()}",
                    f"Durum: {str(getattr(ci, 'status', '') or '-')}",
                ],
                "tag": "Sözleşme termini",
                "color": "#f97316",
            })
    for sys_info in list(systems or []):
        system_deadline = parse_iso_date(str(getattr(sys_info, "completion_date", "") or ""))
        if not system_deadline:
            continue
        no = str(getattr(ci, "no", "") or "").strip() if ci else ""
        name = str(getattr(sys_info, "name", "") or "").strip() or "Sistem"
        events.append({
            "date": system_deadline,
            "title": f"{no} {name}".strip(),
            "lines": [
                f"Termin tarihi: {system_deadline.isoformat()}",
                f"Durum: {str(getattr(sys_info, 'status', '') or '-')}",
            ],
            "tag": "Sistem termini",
            "color": "#2563eb",
        })
    return events


def as_number(v):
    try:
        if callable(v):
            return 0.0
        if isinstance(v, (dict, list, tuple, set)):
            return 0.0
        return float(v or 0)
    except RecursionError:
        try:
            sys.__stderr__.write("RecursionError in as_number; returning 0.0\n")
        except Exception:
            pass
        return 0.0
    except Exception:
        return 0.0


def _global_exc_handler(exc_type, exc, tb):
    if issubclass(exc_type, RecursionError):
        try:
            sys.__stderr__.write("".join(traceback.format_exception_only(exc_type, exc)))
        except Exception:
            pass
        return
    try:
        traceback.print_exception(exc_type, exc, tb)
    except Exception:
        pass
    app_instance = QApplication.instance() if "QApplication" in globals() else None
    if app_instance is not None:
        try:
            QMessageBox.critical(None, "Beklenmeyen Hata", f"Uygulamada beklenmeyen bir hata oluştu.\n\n{exc}")
        except Exception:
            pass


def fmt_num(v) -> str:
    try:
        f = float(v or 0)
        return str(int(f)) if f == int(f) else str(round(f, 2))
    except Exception:
        return str(v or "")


def _hex_to_rgb(color: str) -> Tuple[int, int, int]:
    txt = str(color or "").strip().lstrip("#")
    if len(txt) == 3:
        txt = "".join(ch * 2 for ch in txt)
    if len(txt) != 6:
        return (59, 130, 246)
    try:
        return (int(txt[0:2], 16), int(txt[2:4], 16), int(txt[4:6], 16))
    except Exception:
        return (59, 130, 246)


def _mix_rgb(a: Tuple[int, int, int], b: Tuple[int, int, int], ratio: float) -> Tuple[int, int, int]:
    r = max(0.0, min(1.0, float(ratio)))
    return (
        int(a[0] * (1 - r) + b[0] * r),
        int(a[1] * (1 - r) + b[1] * r),
        int(a[2] * (1 - r) + b[2] * r),
    )


def _rgb_to_hex(rgb: Tuple[int, int, int]) -> str:
    return "#{:02X}{:02X}{:02X}".format(
        max(0, min(255, int(rgb[0]))),
        max(0, min(255, int(rgb[1]))),
        max(0, min(255, int(rgb[2]))),
    )


def tag_chip_style(color: str, selected: bool = False) -> str:
    base = _hex_to_rgb(color)
    bg = _mix_rgb(base, (255, 255, 255), 0.78 if not selected else 0.68)
    border = _mix_rgb(base, (255, 255, 255), 0.28 if not selected else 0.12)
    lum = (0.299 * base[0] + 0.587 * base[1] + 0.114 * base[2]) / 255.0
    txt = "#0F172A" if lum > 0.58 else "#FFFFFF"
    if not selected:
        txt = _rgb_to_hex(_mix_rgb(base, (15, 23, 42), 0.22))
    return (
        f"QPushButton {{ background:{_rgb_to_hex(bg)}; color:{txt}; border:1px solid {_rgb_to_hex(border)}; "
        "border-radius:13px; padding:4px 10px; font-weight:800; } "
        f"QPushButton:hover {{ border-color:{_rgb_to_hex(_mix_rgb(base, (15, 23, 42), 0.18))}; }}"
    )


class ElidedLabel(QLabel):
    def __init__(self, text: str = "", parent=None):
        super().__init__("", parent)
        self._full_text = ""
        self.setText(text)

    def setText(self, text: str):  # type: ignore[override]
        self._full_text = str(text or "")
        self.setToolTip(self._full_text)
        self._refresh_elide()

    def resizeEvent(self, event):
        super().resizeEvent(event)
        self._refresh_elide()

    def _refresh_elide(self):
        width = max(12, self.width())
        super().setText(self.fontMetrics().elidedText(self._full_text, Qt.ElideRight, width))



from src.services.excel_store import ExcelStore
from src.services.sts_store import STSStore
from src import auth
from src.workers import ExcelLoadWorker, UserSaveWorker, ContractSaveWorker, AnalyzeDialog

_log = logging.getLogger("STS")


class STSLoadWorker(QObject):
    """STS dosyasını ana thread'i bloklamadan önce doğrular.

    KURAL: Bu worker hiçbir zaman STSStore, STSDatabase veya sqlite3.connect
    nesnesi ANA THREAD'E AKTARMAZ. SQLite connection thread'e bağlıdır;
    worker thread'de oluşturulan bir connection ana thread'de kullanılırsa
    ProgrammingError (check_same_thread=True default) veya veri bozulması olur.

    Worker yalnızca dosya varlığı ve magic-bytes doğrulaması yapar, ardından
    parametresiz finished() sinyali gönderir. Asıl STSStore ve contract index
    ana thread'de _on_sts_load_finished() içinde oluşturulur.
    """

    progress = Signal(int, str)
    # Sinyalde STSStore / bağlantı nesnesi YOK — sadece kontrol sonucu
    finished = Signal()
    failed = Signal(str)

    def __init__(self, path: Path):
        super().__init__()
        self.path = Path(path)

    def run(self):
        try:
            self.progress.emit(15, "STS dosyası doğrulanıyor...")
            if not self.path.exists():
                raise FileNotFoundError(f"Dosya bulunamadı: {self.path}")
            if not self.path.is_file():
                raise ValueError(f"Geçerli bir dosya değil: {self.path}")
            # Hafif ön-kontrol: SQLite magic bytes (connection açmadan)
            with open(self.path, "rb") as fh:
                header = fh.read(16)
            if not header.startswith(b"SQLite format 3"):
                raise ValueError("Dosya geçerli bir STS/SQLite veritabanı değil.")
            self.progress.emit(80, "Doğrulama tamamlandı, yükleniyor...")
            self.finished.emit()
        except Exception as exc:
            _log.exception("STSLoadWorker doğrulama hatası")
            self.failed.emit(str(exc))


class SystemTypeStore(Protocol):
    """System dialogs depend on this store API in both ExcelStore and STSStore modes."""

    def assigned_components(self, platform: str) -> List[str]: ...
    def list_system_type_names(self, platform: str = "") -> List[str]: ...
    def get_system_type_components(self, type_name: str, platform: str = "") -> List[str]: ...
    def get_system_type_component_quantities(self, type_name: str, platform: str = "") -> Dict[str, float]: ...
    def save_system_type(self, type_name: str, platform: str, components) -> int: ...


COL_PLATFORM = 0
COL_TYPE = 1
COL_CONTRACT_NO = 2
COL_USER = 3
COL_STATUS = 4
COL_T_DATE = 5
COL_REMAINING = 6
COL_TAGS = 7
COL_SUMMARY = 8
PLATFORM_SELECTED_ROLE = Qt.UserRole + 100


class PlatformListDelegate(QStyledItemDelegate):
    """Platform listesi: renkli kısaltma kutusu + isim + sayı."""

    _PALETTES = [
        ("#dbeafe", "#1e40af"),
        ("#fce7f3", "#9d174d"),
        ("#d1fae5", "#065f46"),
        ("#fef3c7", "#92400e"),
        ("#ede9fe", "#5b21b6"),
        ("#fee2e2", "#991b1b"),
        ("#e0f2fe", "#075985"),
        ("#fef9c3", "#713f12"),
    ]

    def __init__(self, parent=None):
        super().__init__(parent)
        self._counts: dict = {}   # row -> count string

    def set_count(self, row: int, count: int):
        self._counts[row] = str(count) if count else ""

    def paint(self, painter, option, index):
        from PySide6.QtGui import QColor
        from PySide6.QtCore import QRect
        from PySide6.QtWidgets import QStyle

        painter.save()

        # PySide6: option.state is StateFlag enum — use QStyle.StateFlag members
        state = option.state
        # QListWidget'in native selection state'i tekli seçim/current item ile sınırlı
        # kalabildiği için platform seçim mantığını ayrı bir item role'ünden okuyoruz.
        # Böylece çoklu seçimde selected_platforms set'indeki her satır aynı aktif
        # stili alır; focus/current satır tek başına selected gibi boyanmaz.
        is_selected = bool(index.data(PLATFORM_SELECTED_ROLE))
        is_hover    = bool(state & QStyle.State_MouseOver)

        # Arka plan
        if is_selected:
            painter.fillRect(option.rect, QColor("#eff6ff"))
            # Sol mavi çizgi
            painter.fillRect(
                QRect(option.rect.left(), option.rect.top(), 3, option.rect.height()),
                QColor("#2563eb")
            )
        elif is_hover:
            painter.fillRect(option.rect, QColor("#f0f7ff"))
        else:
            painter.fillRect(option.rect, QColor("#ffffff"))

        row = index.row()
        pal_bg, pal_fg = self._PALETTES[row % len(self._PALETTES)]

        # UserRole'den platform adını al (text'te sayaç olabilir)
        platform_name = str(index.data(Qt.UserRole) or index.data(Qt.DisplayRole) or "").strip()
        abbr = platform_name[:3].upper() if platform_name else "?"

        rect = option.rect
        abbr_rect = QRect(
            rect.left() + 10,
            rect.top() + (rect.height() - 26) // 2,
            34, 26
        )

        # Kısaltma kutusu
        painter.setPen(Qt.NoPen)
        painter.setBrush(QColor(pal_bg))
        painter.drawRoundedRect(abbr_rect, 5, 5)

        af = painter.font()
        af.setPointSize(8)
        af.setBold(True)
        painter.setFont(af)
        painter.setPen(QColor(pal_fg))
        painter.drawText(abbr_rect, Qt.AlignCenter, abbr)

        # Platform adı
        name_x = abbr_rect.right() + 10
        count_str = self._counts.get(row, "")
        count_w = 30 if count_str else 0
        name_rect = QRect(name_x, rect.top(), rect.width() - name_x - count_w - 8, rect.height())

        nf = painter.font()
        nf.setPointSize(10)
        nf.setBold(is_selected)
        painter.setFont(nf)
        painter.setPen(QColor("#1e40af") if is_selected else QColor("#374151"))
        painter.drawText(name_rect, Qt.AlignVCenter | Qt.AlignLeft, platform_name)

        # Sağda sözleşme sayısı
        if count_str:
            cnt_rect = QRect(rect.right() - count_w - 6, rect.top(), count_w, rect.height())
            cf = painter.font()
            cf.setPointSize(9)
            cf.setBold(False)
            painter.setFont(cf)
            painter.setPen(QColor("#94a3b8"))
            painter.drawText(cnt_rect, Qt.AlignVCenter | Qt.AlignRight, count_str)

        painter.restore()

    def sizeHint(self, option, index):
        try:
            w = int(option.rect.width()) if option is not None else 200
            return QSize(w if w > 0 else 200, 46)
        except Exception:
            return QSize(200, 46)




class FilterableHeaderView(QHeaderView):
    """Excel gibi sutun filtresi destekleyen header.
    Her sutun basliginda kucuk bir filtre ikonu gosterir.
    Tiklayinca o sutunun benzersiz degerlerini checkbox listesi olarak acar.
    """
    filterChanged = Signal()

    def __init__(self, orientation, parent=None):
        super().__init__(orientation, parent)
        self.setSectionsClickable(True)
        self.setHighlightSections(True)
        # col_index -> set of selected values (None = tumu secili)
        self._col_filters: Dict[int, Optional[set]] = {}
        self._date_ranges: Dict[int, Tuple[Optional[date], Optional[date]]] = {}
        self._day_ranges: Dict[int, Tuple[Optional[int], Optional[int]]] = {}
        self.sectionClicked.connect(self._on_section_clicked)

    def has_active_filter(self, col: int) -> bool:
        return (
            (col in self._col_filters and self._col_filters[col] is not None)
            or col in self._date_ranges
            or col in self._day_ranges
        )

    def get_filter(self, col: int) -> Optional[set]:
        return self._col_filters.get(col)

    def clear_filter(self, col: int):
        self._col_filters.pop(col, None)
        self._date_ranges.pop(col, None)
        self._day_ranges.pop(col, None)
        self.filterChanged.emit()
        self.viewport().update()

    def clear_all_filters(self):
        self._col_filters.clear()
        self._date_ranges.clear()
        self._day_ranges.clear()
        self.filterChanged.emit()
        self.viewport().update()

    def _get_column_values(self, col: int) -> List[str]:
        table = self.parent()
        if not isinstance(table, QTableWidget):
            return []
        vals = set()
        # Tum satirlari tara (visible_rows varsa onu kullan)
        source = getattr(table, '_all_rows_for_filter', None)
        if source is not None:
            col_keys = getattr(table, "_filter_col_keys", ["platform", "type", "no", "user", "status", "date", "days", "tags", "summary"])
            if col < len(col_keys):
                window = table.window()
                for row_index, it in enumerate(source, start=1):
                    key = col_keys[col]
                    if key == "type":
                        v = str(it.get("type_display", it.get("type", "")) or "").strip()
                    elif key in {"status", "date", "days"} and hasattr(window, "_contract_health"):
                        _cls, st, days, dt = window._contract_health(it)
                        v = {"status": st, "date": dt, "days": days}.get(key, "")
                    elif key == "tags":
                        for tg in list(it.get("tags", []) or []):
                            tg = str(tg or "").strip()
                            if tg:
                                vals.add(tg)
                        continue
                    elif key == "summary":
                        v = "Özet"
                    else:
                        v = str(it.get(key, "") or "").strip()
                    if v:
                        vals.add(v)
        else:
            for r in range(table.rowCount()):
                item = table.item(r, col)
                v = str(item.text() if item else "").strip()
                if v:
                    vals.add(v)
        return sorted(vals, key=lambda x: x.lower())

    def _on_section_clicked(self, col: int):
        values = self._get_column_values(col)
        if not values and col not in (COL_T_DATE, COL_REMAINING):
            return
        current_filter = self._col_filters.get(col)  # None = tumu

        popup = QMenu(self.viewport())
        popup.setObjectName("filterPopup")
        popup.setStyleSheet(
            "QMenu { background:#fff; border:1px solid #d8e2ed; border-radius:6px; padding:4px; }"
            "QMenu::item { padding:4px 14px; border-radius:4px; }"
            "QMenu::item:selected { background:#EEF2F6; }"
        )

        # Tumunu sec / temizle
        select_all_action = popup.addAction("✔ Tümünü Seç")
        clear_action = popup.addAction("✕ Filtreyi Temizle")
        clear_action.setEnabled(current_filter is not None or col in self._date_ranges or col in self._day_ranges)
        popup.addSeparator()
        # Kalan Gun sutunu - siralama secenekleri ekle
        sort_asc_action = None
        sort_desc_action = None
        if col == COL_REMAINING:  # Kalan Gun sutunu
            sort_asc_action = popup.addAction("↑ Artan Sırala (Az → Çok)")
            sort_desc_action = popup.addAction("↓ Azalan Sırala (Çok → Az)")
            popup.addSeparator()


        if col == COL_T_DATE:
            self._add_date_range_controls(popup, col)
            popup.addSeparator()
        elif col == COL_REMAINING:
            self._add_day_range_controls(popup, col)
            popup.addSeparator()
        # Her deger icin checkbox action
        check_actions: List[Tuple[QAction, str]] = []
        if col not in (COL_T_DATE, COL_REMAINING):
            for val in values:
                icon_txt = "✔" if (current_filter is None or val in current_filter) else "□"
                a = popup.addAction(f"{icon_txt}  {val}")
                a.setCheckable(False)
                check_actions.append((a, val))

        # Sutun basliginin ekran koordinati
        x = self.sectionViewportPosition(col)
        y = self.height()
        global_pos = self.viewport().mapToGlobal(QPoint(x, y))

        chosen = popup.exec(global_pos)
        if not chosen:
            return
        if chosen is select_all_action:
            self._col_filters.pop(col, None)
            self._date_ranges.pop(col, None)
            self._day_ranges.pop(col, None)
        elif chosen is clear_action:
            self._col_filters.pop(col, None)
            self._date_ranges.pop(col, None)
            self._day_ranges.pop(col, None)
        elif sort_asc_action is not None and chosen is sort_asc_action:
            # Kalan gun artan siralama - table parent'ina sort_mode set et
            table = self.parent()
            if hasattr(table, '_sort_mode'):
                table._sort_mode = 'days_asc'
            self.filterChanged.emit()
            return
        elif sort_desc_action is not None and chosen is sort_desc_action:
            table = self.parent()
            if hasattr(table, '_sort_mode'):
                table._sort_mode = 'days_desc'
            self.filterChanged.emit()
            return
        else:
            # Tiklanana gore toggle
            clicked_val = next((v for a, v in check_actions if a is chosen), None)
            if clicked_val is not None:
                if current_filter is None:
                    # Ilk tiklamada sadece o degeri sec
                    self._col_filters[col] = {clicked_val}
                elif clicked_val in current_filter:
                    current_filter.discard(clicked_val)
                    if not current_filter:
                        self._col_filters.pop(col, None)
                    else:
                        self._col_filters[col] = current_filter
                else:
                    current_filter.add(clicked_val)
                    self._col_filters[col] = current_filter
        self.filterChanged.emit()
        self.viewport().update()

    def _add_date_range_controls(self, popup: QMenu, col: int):
        box = QWidget()
        lay = QVBoxLayout(box)
        lay.setContentsMargins(8, 6, 8, 6)
        lay.setSpacing(5)
        title = QLabel("Tarih aralığı")
        title.setStyleSheet("font-weight:800;color:#1b3150;")
        lay.addWidget(title)
        current_from, current_to = self._date_ranges.get(col, (None, None))
        start = QDateEdit()
        start.setCalendarPopup(True)
        start.setDisplayFormat("dd.MM.yyyy")
        start.setSpecialValueText("Başlangıç")
        start.setMinimumDate(QDate(2000, 1, 1))
        start.setMaximumDate(QDate(2100, 12, 31))
        start.setDate(QDate(current_from.year, current_from.month, current_from.day) if current_from else QDate.currentDate())
        if not current_from:
            start.lineEdit().clear()
        end = QDateEdit()
        end.setCalendarPopup(True)
        end.setDisplayFormat("dd.MM.yyyy")
        end.setSpecialValueText("Bitiş")
        end.setMinimumDate(QDate(2000, 1, 1))
        end.setMaximumDate(QDate(2100, 12, 31))
        end.setDate(QDate(current_to.year, current_to.month, current_to.day) if current_to else QDate.currentDate())
        if not current_to:
            end.lineEdit().clear()
        lay.addWidget(start)
        lay.addWidget(end)
        buttons = QHBoxLayout()
        apply_btn = QPushButton("Uygula")
        clear_btn = QPushButton("Aralığı Temizle")
        buttons.addWidget(apply_btn)
        buttons.addWidget(clear_btn)
        lay.addLayout(buttons)

        def apply_range():
            start_date = start.date().toPython() if start.lineEdit().text().strip() else None
            end_date = end.date().toPython() if end.lineEdit().text().strip() else None
            if start_date or end_date:
                self._date_ranges[col] = (start_date, end_date)
            else:
                self._date_ranges.pop(col, None)
            popup.close()
            self.filterChanged.emit()
            self.viewport().update()

        def clear_range():
            self._date_ranges.pop(col, None)
            popup.close()
            self.filterChanged.emit()
            self.viewport().update()

        apply_btn.clicked.connect(apply_range)
        clear_btn.clicked.connect(clear_range)
        action = QWidgetAction(popup)
        action.setDefaultWidget(box)
        popup.addAction(action)

    def _add_day_range_controls(self, popup: QMenu, col: int):
        box = QWidget()
        lay = QVBoxLayout(box)
        lay.setContentsMargins(8, 6, 8, 6)
        lay.setSpacing(5)
        title = QLabel("Kalan gün aralığı")
        title.setStyleSheet("font-weight:800;color:#1b3150;")
        lay.addWidget(title)
        current_min, current_max = self._day_ranges.get(col, (None, None))
        min_edit = QLineEdit()
        min_edit.setPlaceholderText("Min / tek gün")
        min_edit.setText("" if current_min is None else str(current_min))
        max_edit = QLineEdit()
        max_edit.setPlaceholderText("Maks")
        max_edit.setText("" if current_max is None else str(current_max))
        lay.addWidget(min_edit)
        lay.addWidget(max_edit)
        buttons = QHBoxLayout()
        apply_btn = QPushButton("Uygula")
        clear_btn = QPushButton("Aralığı Temizle")
        buttons.addWidget(apply_btn)
        buttons.addWidget(clear_btn)
        lay.addLayout(buttons)

        def parse_int(widget: QLineEdit):
            txt = widget.text().strip()
            if not txt:
                return None
            try:
                return int(txt)
            except ValueError:
                return None

        def apply_range():
            min_val = parse_int(min_edit)
            max_val = parse_int(max_edit)
            if min_val is not None or max_val is not None:
                if min_val is not None and max_val is None:
                    max_val = min_val
                self._day_ranges[col] = (min_val, max_val)
            else:
                self._day_ranges.pop(col, None)
            popup.close()
            self.filterChanged.emit()
            self.viewport().update()

        def clear_range():
            self._day_ranges.pop(col, None)
            popup.close()
            self.filterChanged.emit()
            self.viewport().update()

        apply_btn.clicked.connect(apply_range)
        clear_btn.clicked.connect(clear_range)
        action = QWidgetAction(popup)
        action.setDefaultWidget(box)
        popup.addAction(action)

    def paintSection(self, painter, rect, logical_index):
        super().paintSection(painter, rect, logical_index)
        # Filtre aktifse ikonu farkli goster
        active = self.has_active_filter(logical_index)
        icon = "▼" if active else "▾"  # solid down vs outline down
        painter.save()
        painter.setPen(QColor("#1F5BE3" if active else "#94A3B8"))
        f = painter.font()
        f.setPointSize(9)  # daha buyuk
        f.setBold(active)
        painter.setFont(f)
        painter.drawText(rect.adjusted(0, 0, -6, 0), Qt.AlignRight | Qt.AlignVCenter, icon)
        painter.restore()


class StyledDialog(QDialog):
    def __init__(self, title, parent=None):
        super().__init__(parent)
        self.setWindowTitle(title)
        self.setModal(True)
        self.setStyleSheet(STYLE)

    def make_footer_status_label(self) -> QLabel:
        label = QLabel("")
        label.setObjectName("footerStatus")
        label.setMinimumWidth(0)
        label.setMinimumHeight(34)
        label.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        label.hide()
        return label

    def show_footer_status(self, message: str, kind: str = "success", duration: int = 2600):
        label = getattr(self, "footer_status", None)
        if label is None:
            ToastNotification.show_in(self, message, kind=kind, duration=duration)
            return

        colors = {
            "success": ("#166534", "#dcfce7", "#16a34a"),
            "error": ("#991b1b", "#fee2e2", "#dc2626"),
            "info": ("#1e3a8a", "#dbeafe", "#2563eb"),
        }
        icons = {"success": "\u2713", "error": "\u2715", "info": "i"}
        fg, bg, border = colors.get(kind, colors["success"])
        text = str(message or "")
        visible_text = text if len(text) <= 96 else f"{text[:93]}..."
        label.setText(f"{icons.get(kind, 'i')}  {visible_text}")
        label.setToolTip(text)
        label.setStyleSheet(
            f"color:{fg};background-color:{bg};border:1px solid {border};"
            "border-radius:7px;padding:6px 10px;font-size:12px;font-weight:700;"
        )
        label.show()

        token = getattr(self, "_footer_status_token", 0) + 1
        self._footer_status_token = token
        QTimer.singleShot(duration, lambda: self.clear_footer_status(token))

    def clear_footer_status(self, token: Optional[int] = None):
        if token is not None and token != getattr(self, "_footer_status_token", None):
            return
        label = getattr(self, "footer_status", None)
        if label is not None:
            label.hide()


def form_label(txt):
    l = QLabel(txt)
    l.setObjectName("formLabel")
    return l


def _legacy_build_date_input_unused(
    parent: QWidget,
    placeholder: str = "yyyy-aa-gg",
    max_date: Optional[date] = None,
) -> Tuple[QLineEdit, QWidget]:
    edit = QLineEdit()
    edit.setPlaceholderText(placeholder)

    btn = QPushButton("📅")
    btn.setObjectName("dateBtn")
    btn.setFixedSize(34, 34)
    btn.setToolTip("Takvimden tarih seç")

    wrap = QWidget(parent)
    lay = QHBoxLayout(wrap)
    lay.setContentsMargins(0, 0, 0, 0)
    lay.setSpacing(6)
    lay.addWidget(edit, 1)
    lay.addWidget(btn, 0)

    def choose_date():
        popup = QDialog(parent, Qt.Popup | Qt.FramelessWindowHint)
        popup.setObjectName("calendarPopup")
        pop_lay = QVBoxLayout(popup)
        pop_lay.setContentsMargins(6, 6, 6, 6)
        pop_lay.setSpacing(0)
        cal = QCalendarWidget(popup)
        cal.setVerticalHeaderFormat(QCalendarWidget.NoVerticalHeader)
        cal.setGridVisible(True)
        if max_date:
            cal.setMaximumDate(QDate(max_date.year, max_date.month, max_date.day))
            disabled_fmt = QTextCharFormat()
            disabled_fmt.setBackground(QColor("#EEF2F7"))
            disabled_fmt.setForeground(QColor("#94A3B8"))
            month_days = calendar.monthrange(max_date.year, max_date.month)[1]
            for day_num in range(max_date.day + 1, month_days + 1):
                cal.setDateTextFormat(QDate(max_date.year, max_date.month, day_num), disabled_fmt)
        current = parse_iso_date(edit.text())
        if current:
            if max_date and current > max_date:
                current = max_date
            cal.setSelectedDate(QDate(current.year, current.month, current.day))

        def on_pick(qd: QDate):
            edit.setText(to_iso(qd))
            popup.accept()

        cal.clicked.connect(on_pick)
        pop_lay.addWidget(cal)
        popup.adjustSize()
        popup.move(btn.mapToGlobal(QPoint(0, btn.height() + 2)))
        popup.exec()

    btn.clicked.connect(choose_date)
    return edit, wrap


def build_date_input(
    parent: QWidget,
    placeholder: str = "yyyy-aa-gg",
    max_date: Optional[date] = None,
    events_provider: Optional[Callable[[], List[dict]]] = None,
) -> Tuple[QLineEdit, QWidget]:
    return _build_date_input(parent, placeholder=placeholder, max_date=max_date, events_provider=events_provider)


from src.ui.delegates import CompactNumberDelegate, CenterTableDelegate, DropdownDelegate

from src.ui.toast import ToastNotification

class UserManagerDialog(StyledDialog):
    def __init__(self, store: ExcelStore, parent=None):
        super().__init__("Kullanıcı Yönetimi", parent)
        self.store = store
        self.users = store.load_users(active_only=False)
        self.changed = False
        self._save_thread: Optional[QThread] = None
        self._save_worker: Optional[UserSaveWorker] = None
        self._save_payload: List[dict] = []
        self._saving = False
        self._busy_cursor_on = False
        self.resize(760, 500)
        self.build()
        self.load_table()

    def build(self):
        root = QVBoxLayout(self)
        title = QLabel("Kullanıcı Yönetimi")
        title.setObjectName("dialogTitle")
        root.addWidget(title)
        desc = QLabel("Sözleşme girişinde seçilecek kullanıcıları burada tanımlayın. Aktif olmayanlar yeni sözleşme ekranında görünmez.")
        desc.setObjectName("muted")
        root.addWidget(desc)

        btns = QHBoxLayout()
        add = QPushButton("+ Kullanıcı Ekle")
        add.clicked.connect(self.add_user)
        delete = QPushButton("Seçili Kullanıcıyı Sil")
        delete.setObjectName("danger")
        delete.clicked.connect(self.delete_selected)
        btns.addWidget(add)
        btns.addWidget(delete)
        btns.addStretch()
        root.addLayout(btns)

        self.table = QTableWidget()
        self.table.setAlternatingRowColors(True)
        self.table.setColumnCount(4)
        self.table.setHorizontalHeaderLabels(["Kullanıcı Adı", "Yİ/YD", "Aktif", "Not"])
        root.addWidget(self.table, 1)

        foot = QHBoxLayout()
        self.footer_status = self.make_footer_status_label()
        foot.addWidget(self.footer_status, 1, Qt.AlignVCenter)
        foot.addStretch()
        self.save_btn = QPushButton("Kaydet")
        self.save_btn.clicked.connect(self.save)
        self.close_btn = QPushButton("Kapat")
        self.close_btn.setObjectName("secondary")
        self.close_btn.clicked.connect(self.reject)
        foot.addWidget(self.save_btn)
        foot.addWidget(self.close_btn)
        root.addLayout(foot)

        self.busy_overlay = QFrame(self)
        self.busy_overlay.setStyleSheet("QFrame { background: rgba(248, 251, 255, 0.86); }")
        self.busy_overlay.hide()
        self.busy_card = QFrame(self.busy_overlay)
        self.busy_card.setStyleSheet(
            "QFrame { background: rgba(255,255,255,0.97); border: 1px solid #d8e2ed; border-radius: 10px; }"
        )
        bl = QVBoxLayout(self.busy_card)
        bl.setContentsMargins(18, 14, 18, 14)
        bl.setSpacing(8)
        self.busy_label = QLabel("İşlem yapılıyor...")
        self.busy_label.setObjectName("mainTitle")
        self.busy_label.setAlignment(Qt.AlignCenter)
        self.busy_progress = QProgressBar()
        self.busy_progress.setRange(0, 100)
        self.busy_progress.setValue(0)
        self.busy_progress.setTextVisible(True)
        self.busy_progress.setFormat("%p%")
        bl.addWidget(self.busy_label)
        bl.addWidget(self.busy_progress)
        self.position_busy_overlay()

    def load_table(self):
        self.table.setRowCount(len(self.users))
        for r, u in enumerate(self.users):
            vals = [u.get("name", ""), u.get("yi_yd", "Yİ"), "Evet" if u.get("active", True) else "Hayır", u.get("note", "")]
            for c, val in enumerate(vals):
                self.table.setItem(r, c, QTableWidgetItem(str(val)))
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        # Sütun 1: Yİ / YD dropdown
        self.table.setItemDelegateForColumn(1, DropdownDelegate(["Yİ", "YD"], self.table))
        # Sütun 2: Evet / Hayır dropdown
        self.table.setItemDelegateForColumn(2, DropdownDelegate(["Evet", "Hayır"], self.table))

    def position_busy_overlay(self):
        if not hasattr(self, "busy_overlay"):
            return
        self.busy_overlay.setGeometry(self.rect())
        w, h = 420, 130
        x = max((self.busy_overlay.width() - w) // 2, 0)
        y = max((self.busy_overlay.height() - h) // 2, 0)
        self.busy_card.setGeometry(x, y, w, h)
        self.busy_overlay.raise_()

    def set_busy(self, visible: bool, message: str = "İşlem yapılıyor...", percent: int = 0):
        if not hasattr(self, "busy_overlay"):
            return
        self._saving = bool(visible)
        if visible:
            self.busy_label.setText(str(message or "İşlem yapılıyor..."))
            self.busy_progress.setValue(int(max(0, min(100, percent))))
            self.position_busy_overlay()
            self.save_btn.setEnabled(False)
            self.close_btn.setEnabled(False)
            self.table.setEnabled(False)
            if hasattr(self, "frozen_table"): self.frozen_table.setEnabled(False)
            self.busy_overlay.show()
            if not self._busy_cursor_on:
                QApplication.setOverrideCursor(Qt.WaitCursor)
                self._busy_cursor_on = True
            QApplication.processEvents()
        else:
            self.busy_overlay.hide()
            self.save_btn.setEnabled(True)
            self.close_btn.setEnabled(True)
            self.table.setEnabled(True)
            if hasattr(self, "frozen_table"): self.frozen_table.setEnabled(True)
            if self._busy_cursor_on:
                QApplication.restoreOverrideCursor()
                self._busy_cursor_on = False
            QApplication.processEvents()

    def resizeEvent(self, event):
        super().resizeEvent(event)
        self.position_busy_overlay()

    def closeEvent(self, event):
        if self._saving:
            event.ignore()
            return
        super().closeEvent(event)

    def _sync_table_to_users(self):
        """Tablodaki hücre değerlerini self.users listesine yansıt."""
        for r in range(min(self.table.rowCount(), len(self.users))):
            u = self.users[r]
            if self.table.item(r, 0):
                u["name"] = self.table.item(r, 0).text().strip() or u.get("name", "")
            if self.table.item(r, 1):
                u["yi_yd"] = self.table.item(r, 1).text().strip() or "Yİ"
            if self.table.item(r, 2):
                u["active"] = self.table.item(r, 2).text().strip().lower() in ["evet", "true", "1", "aktif"]
            if self.table.item(r, 3):
                u["note"] = self.table.item(r, 3).text().strip()

    def add_user(self):
        self._sync_table_to_users()   # ← önce mevcut değerleri koru
        self.users.append({"name": "Yeni Kullanıcı", "yi_yd": "Yİ", "active": True, "note": ""})
        self.load_table()
        last = self.table.rowCount() - 1
        self.table.setCurrentCell(last, 0)
        self.table.editItem(self.table.item(last, 0))

    def delete_selected(self):
        r = self.table.currentRow()
        if r >= 0:
            self._sync_table_to_users()   # ← önce mevcut değerleri koru
            self.users.pop(r)
            self.load_table()

    def save(self):
        result = []
        seen = set()
        for r in range(self.table.rowCount()):
            name = (self.table.item(r, 0).text() if self.table.item(r, 0) else "").strip()
            if not name:
                continue
            if name.lower() in seen:
                QMessageBox.warning(self, "Uyarı", f"Tekrarlanan kullanıcı: {name}")
                return
            seen.add(name.lower())
            yi_yd_txt = (self.table.item(r, 1).text() if self.table.item(r, 1) else "Yİ").strip().upper()
            yi_yd = "YD" if yi_yd_txt == "YD" else "Yİ"
            active_txt = (self.table.item(r, 2).text() if self.table.item(r, 2) else "Evet").strip().lower()
            result.append({
                "name": name,
                "yi_yd": yi_yd,
                "active": active_txt in ["evet", "true", "1", "aktif", "yes"],
                "note": (self.table.item(r, 3).text() if self.table.item(r, 3) else ""),
            })
        self._save_payload = list(result)
        if _is_sts_store(self.store):
            try:
                self.set_busy(True, "Kullanıcılar kaydediliyor...", 25)
                self.store.write_users(self._save_payload, actor=self.store.current_actor())
                self.store.save()
                self.on_save_finished()
            except Exception as exc:
                self.on_save_failed(str(exc))
            return
        self._start_async_save()

    def _start_async_save(self):
        if self._save_thread and self._save_thread.isRunning():
            return
        self.set_busy(True, "Kullanıcı güncellemesi başlatılıyor...", 6)
        self._save_thread = QThread(self)
        self._save_worker = UserSaveWorker(self.store.path, self._save_payload, self.store.current_actor())
        self._save_worker.moveToThread(self._save_thread)
        self._save_thread.started.connect(self._save_worker.run)
        self._save_worker.progress.connect(self.on_save_progress)
        self._save_worker.finished.connect(self.on_save_finished)
        self._save_worker.failed.connect(self.on_save_failed)
        self._save_worker.finished.connect(self._save_thread.quit)
        self._save_worker.failed.connect(self._save_thread.quit)
        self._save_thread.finished.connect(self._save_worker.deleteLater)
        self._save_thread.finished.connect(self._save_thread.deleteLater)
        self._save_thread.finished.connect(self._clear_save_refs)
        self._save_thread.start()

    def _clear_save_refs(self):
        self._save_worker = None
        self._save_thread = None

    def on_save_progress(self, percent: int, message: str):
        self.set_busy(True, str(message or "İşlem yapılıyor..."), int(max(0, min(100, int(percent or 0)))))

    def on_save_finished(self):
        try:
            self.set_busy(True, "Yerel önbellek yenileniyor...", 98)
            self.store.reload_from_disk()
            self.users = self.store.load_users(active_only=False)
            self.changed = True
            self.set_busy(False)
            self.show_footer_status("Kullanıcılar kaydedildi", kind="success")
        except Exception as exc:
            self.set_busy(False)
            self.show_footer_status(f"Yenileme hatası: {exc}", kind="error", duration=4000)

    def on_save_failed(self, error_text: str):
        self.set_busy(False)
        self.show_footer_status("Kaydetme hatası! Detay için loga bakın.", kind="error", duration=4000)
        QMessageBox.critical(self, "Kullanıcı kaydetme hatası", f"Kaydetme sırasında hata oluştu:\n\n{error_text}")




class _PlatformRowWidget(QWidget):
    """Platform dropdown satırı: renkli kısaltma + isim."""

    clicked = Signal(str)

    _PALETTES = [
        ("#e8f0fe", "#1e40af"),
        ("#fce7f3", "#9d174d"),
        ("#ecfdf5", "#065f46"),
        ("#fef3c7", "#92400e"),
        ("#ede9fe", "#5b21b6"),
        ("#fee2e2", "#991b1b"),
        ("#d1fae5", "#065f46"),
        ("#e0f2fe", "#075985"),
    ]

    def __init__(self, name: str, index: int, selected: bool = False, parent=None):
        super().__init__(parent)
        self._name = name
        self._selected = selected
        bg, fg = self._PALETTES[index % len(self._PALETTES)]
        self._bg, self._fg = bg, fg
        self.setCursor(Qt.PointingHandCursor)
        self.setFixedHeight(40)

        lay = QHBoxLayout(self)
        lay.setContentsMargins(10, 0, 12, 0)
        lay.setSpacing(10)

        abbr = QLabel(name[:3].upper())
        abbr.setFixedSize(30, 26)
        abbr.setAlignment(Qt.AlignCenter)
        abbr.setStyleSheet(
            f"background:{bg};color:{fg};border-radius:5px;"
            "font-size:9px;font-weight:800;letter-spacing:0.5px;"
        )
        lay.addWidget(abbr)

        lbl = QLabel(name)
        lbl.setStyleSheet("font-size:13px;color:#0f172a;background:transparent;")
        lay.addWidget(lbl, 1)

        if selected:
            tick = QLabel("✓")
            tick.setStyleSheet("color:#2563eb;font-size:13px;font-weight:700;background:transparent;")
            lay.addWidget(tick)

        self._update_bg()

    def _update_bg(self):
        if self._selected:
            self.setStyleSheet("QWidget{background:#f0f7ff;}QWidget:hover{background:#e8f3ff;}")
        else:
            self.setStyleSheet("QWidget{background:white;}QWidget:hover{background:#f8fafc;}")

    def mousePressEvent(self, event):
        self.clicked.emit(self._name)


class PlatformSelectWidget(QWidget):
    """
    Tek seçimli platform dropdown.
    platform_names listesinden seçim yapar, renkli kısaltma + isim gösterir.

    Public API:
        set_platforms(names: List[str])
        set_current(name: str)
        current_text() -> str
        currentTextChanged Signal(str)
    """

    currentTextChanged = Signal(str)

    def __init__(self, parent=None):
        super().__init__(parent)
        self._platforms: List[str] = []
        self._current: str = ""
        self._dropdown: Optional[QFrame] = None

        lay = QVBoxLayout(self)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(0)

        self._display = QFrame()
        self._display.setObjectName("platformSelectDisplay")
        self._display.setCursor(Qt.PointingHandCursor)
        self._display.setMinimumHeight(34)
        self._display.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Minimum)
        self._display.setStyleSheet(
            "QFrame#platformSelectDisplay{"
            "background:white;border:1.5px solid #d8e2ed;border-radius:6px;"
            "}"
            "QFrame#platformSelectDisplay:hover{border-color:#93c5fd;}"
        )
        dl = QHBoxLayout(self._display)
        dl.setContentsMargins(8, 4, 8, 4)
        dl.setSpacing(8)

        self._abbr_lbl = QLabel("")
        self._abbr_lbl.setFixedSize(30, 24)
        self._abbr_lbl.setAlignment(Qt.AlignCenter)
        self._abbr_lbl.setStyleSheet(
            "background:#e8f0fe;color:#1e40af;border-radius:4px;"
            "font-size:9px;font-weight:800;letter-spacing:0.5px;"
        )
        self._abbr_lbl.hide()
        dl.addWidget(self._abbr_lbl)

        self._name_lbl = QLabel("Platform seçiniz...")
        self._name_lbl.setStyleSheet("font-size:13px;color:#94a3b8;background:transparent;")
        dl.addWidget(self._name_lbl, 1)

        self._chev = QLabel("▾")
        self._chev.setStyleSheet("color:#94a3b8;font-size:13px;background:transparent;")
        dl.addWidget(self._chev)

        lay.addWidget(self._display)
        self._display.mousePressEvent = self._toggle_dropdown

        # currentIndex ve currentText compat shims for QComboBox drop-in replacement
        self._current_index: int = -1

    # ── QComboBox compat ────────────────────────────────────────────────

    def addItems(self, names):
        self.set_platforms(list(names))

    def addItem(self, name):
        self._platforms.append(str(name))
        self._rebuild_dropdown()

    def currentText(self) -> str:
        return self._current

    def currentIndex(self) -> int:
        try:
            return self._platforms.index(self._current)
        except ValueError:
            return -1

    def setCurrentIndex(self, idx: int):
        if 0 <= idx < len(self._platforms):
            self._set_current(self._platforms[idx])

    def setCurrentText(self, text: str):
        self._set_current(str(text or ""))

    # currentIndexChanged shim — bağlamalar için
    @property
    def currentIndexChanged(self):
        return self.currentTextChanged

    # ── Public API ──────────────────────────────────────────────────────

    def set_platforms(self, names: List[str]):
        self._platforms = [str(n) for n in names if n]
        if self._current not in self._platforms:
            self._current = self._platforms[0] if self._platforms else ""
        self._update_display()

    def set_current(self, name: str):
        self._set_current(str(name or ""))

    # ── İç metodlar ────────────────────────────────────────────────────

    _PALETTES = [
        ("#e8f0fe", "#1e40af"),
        ("#fce7f3", "#9d174d"),
        ("#ecfdf5", "#065f46"),
        ("#fef3c7", "#92400e"),
        ("#ede9fe", "#5b21b6"),
        ("#fee2e2", "#991b1b"),
        ("#d1fae5", "#065f46"),
        ("#e0f2fe", "#075985"),
    ]

    def _palette(self, name: str):
        idx = self._platforms.index(name) if name in self._platforms else 0
        return self._PALETTES[idx % len(self._PALETTES)]

    def _set_current(self, name: str):
        if name == self._current:
            return
        self._current = name
        self._update_display()
        self.currentTextChanged.emit(name)

    def _update_display(self):
        if self._current and self._current in self._platforms:
            bg, fg = self._palette(self._current)
            self._abbr_lbl.setText(self._current[:3].upper())
            self._abbr_lbl.setStyleSheet(
                f"background:{bg};color:{fg};border-radius:4px;"
                "font-size:9px;font-weight:800;letter-spacing:0.5px;"
            )
            self._abbr_lbl.show()
            self._name_lbl.setText(self._current)
            self._name_lbl.setStyleSheet(
                "font-size:13px;color:#0f172a;background:transparent;"
            )
        else:
            self._abbr_lbl.hide()
            self._name_lbl.setText("Platform seçiniz...")
            self._name_lbl.setStyleSheet(
                "font-size:13px;color:#94a3b8;background:transparent;"
            )

    def _toggle_dropdown(self, event=None):
        if self._dropdown and self._dropdown.isVisible():
            self._dropdown.hide()
            return
        self._open_dropdown()

    def _open_dropdown(self):
        if self._dropdown is None:
            self._dropdown = QFrame(None)
            self._dropdown.setObjectName("platformDropdown")
            self._dropdown.setStyleSheet(
                "QFrame#platformDropdown{"
                "background:white;border:1.5px solid #e2e8f0;border-radius:10px;"
                "}"
            )
            self._dropdown.setWindowFlags(Qt.Popup | Qt.FramelessWindowHint)

        # Mevcut satırları temizle
        if self._dropdown.layout():
            while self._dropdown.layout().count():
                item = self._dropdown.layout().takeAt(0)
                if item.widget():
                    item.widget().deleteLater()
            QFrame().setLayout(self._dropdown.layout())

        root = QVBoxLayout(self._dropdown)
        root.setContentsMargins(0, 4, 0, 4)
        root.setSpacing(0)

        for i, name in enumerate(self._platforms):
            row = _PlatformRowWidget(name, i, name == self._current)
            row.clicked.connect(self._on_platform_clicked)
            root.addWidget(row)

        self._dropdown.setFixedWidth(max(self.width(), 200))
        self._dropdown.adjustSize()

        pos = self.mapToGlobal(self._display.rect().bottomLeft())
        self._dropdown.move(pos.x(), pos.y() + 2)
        self._dropdown.show()
        self._dropdown.raise_()

    def _on_platform_clicked(self, name: str):
        self._dropdown.hide()
        self._set_current(name)

    def resizeEvent(self, event):
        super().resizeEvent(event)
        if self._dropdown and self._dropdown.isVisible():
            self._dropdown.setFixedWidth(max(self.width(), 200))

class _UserRowWidget(QWidget):
    """Dropdown içindeki tek kullanıcı satırı: avatar + isim + checkbox."""

    toggled = Signal(str, bool)  # (name, is_checked)

    _AVATAR_PALETTES = [
        ("#e8f0fe", "#1e40af"),
        ("#fce7f3", "#9d174d"),
        ("#ecfdf5", "#065f46"),
        ("#fef3c7", "#92400e"),
        ("#ede9fe", "#5b21b6"),
        ("#fee2e2", "#991b1b"),
        ("#e0f2fe", "#075985"),
        ("#d1fae5", "#065f46"),
    ]

    def __init__(self, name: str, palette_index: int, checked: bool = False, parent=None):
        super().__init__(parent)
        self._name = name
        self._checked = checked
        bg, fg = self._AVATAR_PALETTES[palette_index % len(self._AVATAR_PALETTES)]
        self._bg = bg
        self._fg = fg
        self.setCursor(Qt.PointingHandCursor)
        self.setFixedHeight(40)

        lay = QHBoxLayout(self)
        lay.setContentsMargins(10, 0, 12, 0)
        lay.setSpacing(10)

        # Avatar dairesi
        self._avatar = QLabel(self._initials(name))
        self._avatar.setFixedSize(26, 26)
        self._avatar.setAlignment(Qt.AlignCenter)
        self._avatar.setStyleSheet(
            f"background:{bg};color:{fg};border-radius:13px;"
            "font-size:10px;font-weight:700;"
        )
        lay.addWidget(self._avatar)

        # İsim
        self._label = QLabel(name)
        self._label.setStyleSheet("font-size:13px;color:#0f172a;background:transparent;")
        lay.addWidget(self._label, 1)

        # Checkbox kutusu
        self._check = QLabel()
        self._check.setFixedSize(18, 18)
        self._update_check_style()
        lay.addWidget(self._check)

        self._update_bg()

    @staticmethod
    def _initials(name: str) -> str:
        parts = str(name or "").strip().split()
        if len(parts) >= 2:
            return (parts[0][0] + parts[-1][0]).upper()
        return name[:2].upper() if name else "?"

    def _update_check_style(self):
        if self._checked:
            self._check.setStyleSheet(
                "background:#2563eb;border-radius:4px;border:1.5px solid #2563eb;"
            )
            self._check.setText("✓")
            self._check.setAlignment(Qt.AlignCenter)
            self._check.setStyleSheet(
                "background:#2563eb;border-radius:4px;border:1.5px solid #2563eb;"
                "color:white;font-size:11px;font-weight:700;"
            )
        else:
            self._check.setText("")
            self._check.setStyleSheet(
                "background:white;border-radius:4px;border:1.5px solid #cbd5e1;"
            )

    def _update_bg(self):
        if self._checked:
            self.setStyleSheet("QWidget{background:#f0f7ff;}QWidget:hover{background:#e8f3ff;}")
        else:
            self.setStyleSheet("QWidget{background:white;}QWidget:hover{background:#f8fafc;}")

    def set_checked(self, checked: bool):
        self._checked = checked
        self._update_check_style()
        self._update_bg()

    def is_checked(self) -> bool:
        return self._checked

    def name(self) -> str:
        return self._name

    def mousePressEvent(self, event):
        self._checked = not self._checked
        self._update_check_style()
        self._update_bg()
        self.toggled.emit(self._name, self._checked)

    def matches_filter(self, query: str) -> bool:
        return query.lower() in self._name.lower()


class _MultiUserDropdown(QFrame):
    """Açılır panel: arama + kullanıcı satırları + alt bilgi."""

    selection_changed = Signal(list)  # List[str]

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setObjectName("multiUserDropdown")
        self.setStyleSheet(
            "QFrame#multiUserDropdown{"
            "background:white;border:1.5px solid #e2e8f0;"
            "border-radius:10px;"
            "}"
        )
        self.setWindowFlags(Qt.Popup | Qt.FramelessWindowHint)
        self.setAttribute(Qt.WA_TranslucentBackground, False)

        self._rows: List[_UserRowWidget] = []
        self._selected: List[str] = []

        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        # Arama satırı
        search_row = QWidget()
        search_row.setStyleSheet("background:white;border-bottom:1px solid #f1f5f9;")
        sr = QHBoxLayout(search_row)
        sr.setContentsMargins(10, 6, 10, 6)
        sr.setSpacing(6)
        lupe = QLabel("⌕")
        lupe.setStyleSheet("color:#94a3b8;font-size:16px;background:transparent;")
        sr.addWidget(lupe)
        self._search = QLineEdit()
        self._search.setPlaceholderText("Ara...")
        self._search.setStyleSheet(
            "border:none;outline:none;font-size:13px;color:#0f172a;"
            "background:transparent;"
        )
        self._search.textChanged.connect(self._apply_filter)
        sr.addWidget(self._search, 1)
        root.addWidget(search_row)

        # Scroll alan
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        scroll.setFixedHeight(210)
        scroll.setStyleSheet("QScrollArea{background:white;border:none;}")
        self._list_container = QWidget()
        self._list_container.setStyleSheet("background:white;")
        self._list_layout = QVBoxLayout(self._list_container)
        self._list_layout.setContentsMargins(0, 4, 0, 4)
        self._list_layout.setSpacing(0)
        scroll.setWidget(self._list_container)
        root.addWidget(scroll)

        # Alt satır
        footer = QWidget()
        footer.setStyleSheet(
            "background:white;border-top:1px solid #f1f5f9;"
        )
        fr = QHBoxLayout(footer)
        fr.setContentsMargins(12, 6, 12, 6)
        self._count_lbl = QLabel("0 seçili")
        self._count_lbl.setStyleSheet("font-size:11px;color:#64748b;background:transparent;")
        fr.addWidget(self._count_lbl)
        fr.addStretch()
        clear_btn = QPushButton("Temizle")
        clear_btn.setStyleSheet(
            "QPushButton{border:none;background:transparent;color:#3b82f6;"
            "font-size:11px;font-weight:700;padding:0;}"
            "QPushButton:hover{color:#1d4ed8;}"
        )
        clear_btn.setCursor(Qt.PointingHandCursor)
        clear_btn.clicked.connect(self._clear_all)
        fr.addWidget(clear_btn)
        root.addWidget(footer)

    def populate(self, available: List[str], selected: List[str]):
        # Mevcut satırları temizle
        while self._list_layout.count():
            item = self._list_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()
        self._rows.clear()
        self._selected = list(selected)

        for i, name in enumerate(available):
            row = _UserRowWidget(name, i, name in self._selected)
            row.toggled.connect(self._on_row_toggled)
            self._list_layout.addWidget(row)
            self._rows.append(row)

        self._list_layout.addStretch()
        self._update_count()
        self._search.clear()

    def _on_row_toggled(self, name: str, checked: bool):
        if checked and name not in self._selected:
            self._selected.append(name)
        elif not checked:
            self._selected = [x for x in self._selected if x != name]
        self._update_count()
        self.selection_changed.emit(list(self._selected))

    def _clear_all(self):
        self._selected.clear()
        for row in self._rows:
            row.set_checked(False)
        self._update_count()
        self.selection_changed.emit([])

    def _update_count(self):
        n = len(self._selected)
        self._count_lbl.setText(f"{n} seçili" if n else "Seçim yok")

    def _apply_filter(self, query: str):
        for row in self._rows:
            row.setVisible(row.matches_filter(query))

    def selected_names(self) -> List[str]:
        return list(self._selected)

    def keyPressEvent(self, event):
        if event.key() in (Qt.Key_Escape, Qt.Key_Return, Qt.Key_Enter):
            self.hide()
        else:
            super().keyPressEvent(event)


class MultiUserSelectWidget(QWidget):
    """
    Kullanıcı seçim widget'ı — pill + flow wrap + avatar dropdown.

    Public API (değişmez):
        set_available_users(names: List[str])
        set_users(names: List[str])
        selected_users() -> List[str]
        changed  Signal
    """

    changed = Signal()

    _PILL_COLORS = [
        ("#e8f0fe", "#1e40af"),
        ("#fce7f3", "#9d174d"),
        ("#ecfdf5", "#065f46"),
        ("#fef3c7", "#92400e"),
        ("#ede9fe", "#5b21b6"),
        ("#fee2e2", "#991b1b"),
        ("#e0f2fe", "#075985"),
    ]

    def __init__(self, parent=None):
        super().__init__(parent)
        self._available: List[str] = []
        self._selected: List[str] = []
        self._placeholder = "Kullanıcı seçiniz..."
        self._dropdown: Optional[_MultiUserDropdown] = None
        self._rendered_rows = 1
        self._display_height = 40
        self._max_visible_rows = 4
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)

        outer = QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.setSpacing(0)

        self._display = QFrame()
        self._display.setObjectName("multiUserDisplay")
        self._display.setCursor(Qt.PointingHandCursor)
        self._display.setMinimumHeight(40)
        self._display.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        self._display.setStyleSheet(
            "QFrame#multiUserDisplay{"
            "background:white;border:1.5px solid #d8e2ed;border-radius:8px;"
            "}"
            "QFrame#multiUserDisplay:hover{border-color:#93c5fd;}"
        )
        self._vlay = QVBoxLayout(self._display)
        self._vlay.setContentsMargins(8, 6, 8, 6)
        self._vlay.setSpacing(5)

        self._scroll = QScrollArea()
        self._scroll.setObjectName("multiUserScroll")
        self._scroll.setWidgetResizable(True)
        self._scroll.setFrameShape(QFrame.NoFrame)
        self._scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self._scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self._scroll.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        self._scroll.setStyleSheet("QScrollArea#multiUserScroll{background:transparent;border:none;}")
        self._scroll.setWidget(self._display)

        outer.addWidget(self._scroll)
        self._display.mousePressEvent = self._toggle_dropdown
        self._render_pills()

    # ── Public API ────────────────────────────────────────────────────────

    def set_available_users(self, user_names: List[str]):
        seen: set = set()
        vals: List[str] = []
        for u in list(user_names or []):
            n = str(u or "").strip()
            if not n:
                continue
            k = n.casefold()
            if k in seen:
                continue
            seen.add(k)
            vals.append(n)
        self._available = vals
        self._selected = [x for x in self._selected if x in self._available]
        self._render_pills()

    def set_users(self, user_names: List[str]):
        seen: set = set()
        vals: List[str] = []
        for u in list(user_names or []):
            n = str(u or "").strip()
            if not n:
                continue
            k = n.casefold()
            if k in seen:
                continue
            seen.add(k)
            vals.append(n)
        if vals == self._selected:
            self._render_pills()
            return
        self._selected = vals
        self._render_pills()
        self.changed.emit()

    def selected_users(self) -> List[str]:
        return list(self._selected)

    # ── İç metodlar ──────────────────────────────────────────────────────

    def _pill_colors(self, name: str):
        idx = sum(ord(c) for c in name) % len(self._PILL_COLORS)
        return self._PILL_COLORS[idx]

    def _make_pill(self, name: str) -> QWidget:
        bg, fg = self._pill_colors(name)
        pill = QWidget()
        pill.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        pill.setMinimumHeight(self._row_height())
        pill.setStyleSheet(f"QWidget{{background:{bg};border-radius:11px;border:none;}}")
        pl = QHBoxLayout(pill)
        pl.setContentsMargins(10, 4, 8, 4)
        pl.setSpacing(5)
        max_label_width = max(90, self._available_pill_width() - 44)
        display_name = QFontMetrics(self.font()).elidedText(str(name), Qt.ElideRight, max_label_width)
        lbl = QLabel(display_name)
        lbl.setToolTip(str(name))
        lbl.setMaximumWidth(max_label_width)
        lbl.setStyleSheet(
            f"color:{fg};font-size:12px;font-weight:600;background:transparent;border:none;"
        )
        pl.addWidget(lbl)
        if self._pill_removable(name):
            x = QLabel("×")
            x.setStyleSheet(
                f"color:{fg};font-size:15px;background:transparent;border:none;padding:0 1px;"
            )
            x.setCursor(Qt.PointingHandCursor)
            x.mousePressEvent = lambda e, n=name: self._remove_user(e, n)
            pl.addWidget(x)
        return pill

    def _pill_removable(self, name: str) -> bool:
        return True

    def _pill_width(self, name: str) -> int:
        metrics = QFontMetrics(self.font())
        return metrics.horizontalAdvance(str(name or "")) + 44

    def _row_height(self) -> int:
        return max(QFontMetrics(self.font()).height() + 12, 28)

    def _available_pill_width(self, width: Optional[int] = None) -> int:
        source_width = int(width or self._display.width() or self.width() or 0)
        # İç marginler + son satırdaki açılır liste oku için güvenli alan bırak.
        # Ok genişliği hesaba katılmazsa özellikle 4. chip son satırda kırpılabiliyor.
        reserved = 58
        return max(source_width - reserved, 160) if source_width > 20 else 300

    def _pill_rows(self, width: Optional[int] = None) -> List[List[str]]:
        if not self._selected:
            return [[]]
        avail = self._available_pill_width(width)
        gap = 5
        rows: List[List[str]] = []
        cur_row: List[str] = []
        cur_w = 0
        for name in self._selected:
            pw = min(self._pill_width(name), max(avail, 160))
            if cur_row and cur_w + gap + pw > avail:
                rows.append(cur_row)
                cur_row = [name]
                cur_w = pw
            else:
                cur_row.append(name)
                cur_w += (gap if len(cur_row) > 1 else 0) + pw
        if cur_row:
            rows.append(cur_row)
        return rows or [[]]

    def _height_for_rows(self, row_count: int) -> int:
        margins = self._vlay.contentsMargins()
        spacing = self._vlay.spacing()
        rows = max(1, int(row_count or 1))
        visible_rows = min(rows, self._max_visible_rows)
        return max(40, margins.top() + margins.bottom() + visible_rows * self._row_height() + max(0, visible_rows - 1) * spacing)

    def _content_height_for_rows(self, row_count: int) -> int:
        margins = self._vlay.contentsMargins()
        spacing = self._vlay.spacing()
        rows = max(1, int(row_count or 1))
        return max(40, margins.top() + margins.bottom() + rows * self._row_height() + max(0, rows - 1) * spacing)

    def _apply_display_height(self, row_count: int):
        height = self._height_for_rows(row_count)
        content_height = self._content_height_for_rows(row_count)
        if height != self._display_height:
            self._display_height = height
        self.setMinimumHeight(height)
        self._scroll.setMinimumHeight(height)
        self._scroll.setMaximumHeight(height)
        self._display.setMinimumHeight(content_height)
        self._display.setMaximumHeight(16777215)
        self._rendered_rows = max(1, int(row_count or 1))
        self._display.updateGeometry()
        self._scroll.updateGeometry()
        self.updateGeometry()
        parent = self.parentWidget()
        if parent is not None:
            layout = parent.layout()
            if layout is not None:
                layout.invalidate()
            parent.updateGeometry()

    def hasHeightForWidth(self) -> bool:
        return True

    def heightForWidth(self, width: int) -> int:
        return self._height_for_rows(len(self._pill_rows(width)))

    def sizeHint(self) -> QSize:
        hint = super().sizeHint()
        width = max(hint.width(), self.width(), 240)
        return QSize(width, self.heightForWidth(width))

    def minimumSizeHint(self) -> QSize:
        return QSize(160, self.heightForWidth(max(self.width(), 160)))

    def _clear_rows(self):
        """_vlay içindeki tüm satır widget'larını sil (her seferinde yeni QLabel yaratılıyor)."""
        while self._vlay.count():
            item = self._vlay.takeAt(0)
            w = item.widget()
            if w is not None:
                w.deleteLater()

    def _render_pills(self):
        """Her çağrıda tüm satırları sıfırdan yeni widget'larla yeniden çizer.
        _placeholder / _chevron artık kalıcı widget değil — her seferinde yeni QLabel."""
        self._clear_rows()

        if not self._selected:
            row = QWidget()
            row.setStyleSheet("background:transparent;border:none;")
            row.setMinimumHeight(self._row_height())
            rl = QHBoxLayout(row)
            rl.setContentsMargins(0, 0, 0, 0)
            rl.setSpacing(4)
            ph = QLabel(getattr(self, "_placeholder", "Kullanıcı seçiniz..."))
            ph.setStyleSheet(
                "color:#94a3b8;font-size:13px;background:transparent;border:none;"
            )
            rl.addWidget(ph)
            rl.addStretch()
            ch = QLabel("▾")
            ch.setStyleSheet(
                "color:#94a3b8;font-size:13px;background:transparent;border:none;"
            )
            rl.addWidget(ch)
            self._vlay.addWidget(row)
        else:
            GAP = 5
            all_rows = self._pill_rows()

            for ri, row_names in enumerate(all_rows):
                row = QWidget()
                row.setStyleSheet("background:transparent;border:none;")
                row.setMinimumHeight(self._row_height())
                rl = QHBoxLayout(row)
                rl.setContentsMargins(0, 0, 0, 0)
                rl.setSpacing(GAP)
                for nm in row_names:
                    rl.addWidget(self._make_pill(nm))
                rl.addStretch()
                if ri == len(all_rows) - 1:
                    ch = QLabel("▾")
                    ch.setStyleSheet(
                        "color:#94a3b8;font-size:13px;background:transparent;border:none;"
                    )
                    rl.addWidget(ch)
                self._vlay.addWidget(row)

        self._apply_display_height(1 if not self._selected else len(all_rows))
        self._display.setToolTip(", ".join(self._selected))

    def _remove_user(self, event, name: str):
        event.accept()
        self._selected = [x for x in self._selected if x != name]
        self._render_pills()
        if self._dropdown and self._dropdown.isVisible():
            self._dropdown.populate(self._available, self._selected)
        self.changed.emit()

    def _toggle_dropdown(self, event=None):
        if self._dropdown is None:
            self._dropdown = _MultiUserDropdown(self)
            self._dropdown.setFixedWidth(max(self.width(), 240))
            self._dropdown.selection_changed.connect(self._on_dropdown_changed)

        if self._dropdown.isVisible():
            self._dropdown.hide()
            return

        self._dropdown.setFixedWidth(max(self.width(), 240))
        self._dropdown.populate(self._available, self._selected)
        pos = self.mapToGlobal(self._scroll.rect().bottomLeft())
        self._dropdown.move(pos.x(), pos.y() + 2)
        self._dropdown.show()
        self._dropdown.raise_()

    def _on_dropdown_changed(self, names: List[str]):
        self._selected = list(names)
        self._render_pills()
        self.changed.emit()

    def resizeEvent(self, event):
        super().resizeEvent(event)
        self._render_pills()
        if self._dropdown:
            self._dropdown.setFixedWidth(max(self.width(), 240))


class MultiStaffSelectWidget(MultiUserSelectWidget):
    """Personel/staff seçim widget'ı — isim chip'i tutar, kayıt için staff id döndürür."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self._placeholder = "Sorumlu mühendis seçiniz..."
        self._staff_id_by_name: Dict[str, int] = {}
        self._staff_name_by_id: Dict[int, str] = {}

    def set_staff_options(self, staff_rows: List[dict]):
        names: List[str] = []
        self._staff_id_by_name = {}
        self._staff_name_by_id = {}
        for row in staff_rows or []:
            try:
                sid = int(row.get("id") or row.get("staff_id") or 0)
            except Exception:
                sid = 0
            name = str(row.get("full_name") or row.get("name") or "").strip()
            if not sid or not name:
                continue
            names.append(name)
            self._staff_id_by_name[name] = sid
            self._staff_name_by_id[sid] = name
        self.set_available_users(names)

    def set_users(self, users: List[str]):
        super().set_users(list(users or [])[:1])

    def set_selected_staff_ids(self, staff_ids: List[int]):
        names: List[str] = []
        for sid in staff_ids or []:
            name = self._staff_name_by_id.get(int(sid or 0))
            if name:
                names.append(name)
                break
        self.set_users(names)

    def _on_dropdown_changed(self, names: List[str]):
        self._selected = list(names or [])[-1:]
        self._render_pills()
        self.changed.emit()

    def selected_staff_ids(self) -> List[int]:
        ids: List[int] = []
        seen = set()
        for name in self.selected_users():
            sid = int(self._staff_id_by_name.get(name) or 0)
            if sid and sid not in seen:
                seen.add(sid)
                ids.append(sid)
                break
        return ids

    def selected_staff_id(self) -> int:
        ids = self.selected_staff_ids()
        return int(ids[0]) if ids else 0


class MultiPlatformSelectWidget(MultiUserSelectWidget):
    """Faz 1 çoklu platform seçim prototipi.

    MultiUserSelectWidget'in denenmiş chip/dropdown altyapısını generic API ile
    kullanır; state UI tarafında çoklu platform adlarını tutar, backend'e sadece
    ilk seçili platform currentText uyumluluğu ile verilir.
    """

    currentTextChanged = Signal(str)

    def __init__(self, parent=None):
        super().__init__(parent)
        self._placeholder = "Platform seçiniz..."
        self._platform_id_by_name: Dict[str, int] = {}
        self._locked_platform_keys: set[str] = set()
        self._display.setObjectName("multiPlatformDisplay")
        self._display.setStyleSheet(
            "QFrame#multiPlatformDisplay{background:white;border:1.5px solid #d8e2ed;border-radius:8px;}"
            "QFrame#multiPlatformDisplay:hover{border-color:#93c5fd;}"
        )

    def set_platforms(self, platforms: List[object]):
        names: List[str] = []
        self._platform_id_by_name = {}
        for item in platforms or []:
            if isinstance(item, dict):
                pid = int(item.get("id") or item.get("platform_id") or 0)
                name = str(item.get("name") or item.get("platform_name") or "").strip()
            else:
                pid = 0
                name = str(item or "").strip()
            if name:
                names.append(name)
                if pid:
                    self._platform_id_by_name[name] = pid
        self.set_available_users(names)

    def selected_platform_names(self) -> List[str]:
        return self.selected_users()

    def selected_platform_ids(self) -> List[int]:
        ids: List[int] = []
        seen = set()
        for name in self.selected_platform_names():
            pid = int(self._platform_id_by_name.get(name) or 0)
            if pid and pid not in seen:
                seen.add(pid); ids.append(pid)
        return ids

    def selected_platforms(self) -> List[str]:
        return self.selected_platform_names()

    def selected_platform_records(self) -> List[dict]:
        ids = self.selected_platform_ids()
        names = self.selected_platform_names()
        return [{"id": pid, "name": name, "platform_id": pid, "platform_name": name} for pid, name in zip(ids, names)]

    def set_selected_platforms(self, names: List[str]):
        self.set_users(names)

    def set_locked_platforms(self, names: List[str]):
        self._locked_platform_keys = {str(name or "").strip().casefold() for name in names or [] if str(name or "").strip()}
        self._render_pills()

    def _pill_removable(self, name: str) -> bool:
        return str(name or "").strip().casefold() not in self._locked_platform_keys

    def currentText(self) -> str:
        vals = self.selected_platforms()
        return vals[0] if vals else ""

    def currentIndex(self) -> int:
        cur = self.currentText()
        try:
            return self._available.index(cur)
        except ValueError:
            return -1

    def setCurrentIndex(self, idx: int):
        if 0 <= idx < len(self._available):
            self.set_selected_platforms([self._available[idx]])

    def setCurrentText(self, text: str):
        t = str(text or "").strip()
        self.set_selected_platforms([t] if t else [])

    @property
    def currentIndexChanged(self):
        return self.currentTextChanged

    def _on_dropdown_changed(self, names: List[str]):
        old = self.currentText()
        locked = [name for name in self._selected if str(name or "").strip().casefold() in self._locked_platform_keys]
        merged: List[str] = []
        seen = set()
        for name in list(locked) + list(names or []):
            clean = str(name or "").strip()
            key = clean.casefold()
            if clean and key not in seen:
                seen.add(key)
                merged.append(clean)
        super()._on_dropdown_changed(merged)
        new = self.currentText()
        if old != new:
            self.currentTextChanged.emit(new)

    def _remove_user(self, event, name: str):
        old = self.currentText()
        super()._remove_user(event, name)
        new = self.currentText()
        if old != new:
            self.currentTextChanged.emit(new)


class ElidedValueLabel(QLabel):
    """Tek satır ellipsis + tam değer tooltip gösteren kompakt header etiketi."""

    def __init__(self, text: str = "", parent=None):
        super().__init__(parent)
        self._full_text = ""
        self.setWordWrap(False)
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.setMinimumWidth(40)
        super().setText("")
        self.setText(text)

    def setText(self, text: str):
        self._full_text = str(text or "-")
        self.setToolTip(self._full_text)
        self._apply_elide()

    def resizeEvent(self, event):
        super().resizeEvent(event)
        self._apply_elide()

    def _apply_elide(self):
        metrics = QFontMetrics(self.font())
        super().setText(metrics.elidedText(self._full_text, Qt.ElideRight, max(20, self.width() - 2)))


class PlatformTabsWidget(QWidget):
    """Header içinde gömülü premium/neon platform sekme rayı."""

    activePlatformChanged = Signal(int)

    DEFAULT_RAIL_HEIGHT = 38

    def __init__(self, parent=None):
        super().__init__(parent)
        self._platforms: List[dict] = []
        self._active = 0
        self._content_width = 76
        self._min_scroll_width = 300
        self._max_width = 420
        self._buttons: Dict[int, QPushButton] = {}
        self._rail_height = self.DEFAULT_RAIL_HEIGHT
        self.setFixedHeight(self._rail_height)
        self.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)

        outer = QHBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.setSpacing(0)

        self._rail = QFrame(self)
        self._rail.setObjectName("PlatformTabRail")
        self._rail.setFixedHeight(self._rail_height)
        self._rail.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        rail_lay = QHBoxLayout(self._rail)
        rail_lay.setContentsMargins(2, 2, 2, 2)
        rail_lay.setSpacing(0)

        self._scroll = QScrollArea(self._rail)
        self._scroll.setObjectName("PlatformTabScroll")
        self._scroll.setWidgetResizable(False)
        self._scroll.setFrameShape(QFrame.NoFrame)
        self._scroll.setFixedHeight(self._rail_height)
        self._scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self._scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self._scroll.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        self._scroll.viewport().setStyleSheet("background:transparent;border:0;")
        rail_lay.addWidget(self._scroll, 1, Qt.AlignVCenter)
        outer.addWidget(self._rail, 1, Qt.AlignVCenter)

        self._host = QWidget()
        self._host.setObjectName("PlatformTabScrollContent")
        self._host.setStyleSheet("QWidget#PlatformTabScrollContent{background:transparent;border:0;}")
        self._host.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        self._host.setFixedHeight(self._rail_height)
        self._lay = QHBoxLayout(self._host)
        self._lay.setContentsMargins(0, 0, 0, 0)
        self._lay.setSpacing(2)
        self._lay.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        self._scroll.setWidget(self._host)
        self._apply_rail_style()

    def _apply_rail_style(self):
        self._rail.setStyleSheet("""
            QFrame#PlatformTabRail {
                background: rgba(5, 18, 43, 0.62);
                border: 1px solid rgba(96, 165, 250, 0.34);
                border-radius: 17px;
                padding: 0px;
                margin: 0px;
            }
        """)
        self._scroll.setStyleSheet("""
            QScrollArea#PlatformTabScroll {
                background: transparent;
                border: none;
                padding: 0px;
                margin: 0px;
            }
            QScrollArea#PlatformTabScroll > QWidget > QWidget { background:transparent; }
            QScrollBar:horizontal { height:0px; background:transparent; }
        """)

    def set_platforms(self, platforms: List[object], active_platform_id: int = 0):
        vals=[]; seen=set()
        for p in platforms or []:
            if isinstance(p, dict):
                pid = int(p.get("platform_id") or p.get("id") or 0)
                name = str(p.get("platform_name") or p.get("name") or "").strip()
                is_primary = bool(p.get("is_primary"))
            else:
                name = str(p or "").strip()
                pid = 0
                is_primary = False
            key = pid if pid else name.casefold()
            if name and key not in seen:
                seen.add(key)
                vals.append({"platform_id": pid, "platform_name": name, "is_primary": is_primary})
        self._platforms = vals
        active_id = int(active_platform_id or 0)
        valid_ids = {int(p.get("platform_id") or 0) for p in self._platforms}
        if active_id not in valid_ids:
            primary = next((p for p in self._platforms if p.get("is_primary")), None)
            active_id = int((primary or (self._platforms[0] if self._platforms else {})).get("platform_id") or 0)
        self._active = active_id
        self._render()

    def _render(self):
        self._buttons.clear()
        while self._lay.count():
            item = self._lay.takeAt(0)
            widget = item.widget()
            if widget:
                widget.setParent(None)
                widget.deleteLater()
        metrics = QFontMetrics(self.font())
        total_width = 0
        single = len(self._platforms) <= 1
        margins = (0, 0, 0, 0)
        self._lay.setContentsMargins(*margins)
        self._lay.setSpacing(2)
        for platform in self._platforms:
            name = str(platform.get("platform_name") or "")
            pid = int(platform.get("platform_id") or 0)
            btn = QPushButton(name)
            btn.setObjectName("PlatformTabButton")
            btn.setProperty("platform_id", pid)
            btn.setCheckable(True)
            btn.setCursor(Qt.PointingHandCursor)
            # Height is derived from the polished button sizeHint; do not force it here.
            chip_width = min(168, max(92 if single else 86, metrics.horizontalAdvance(name) + (44 if single else 40)))
            btn.setFixedWidth(chip_width)
            btn.setFixedHeight(30)
            btn.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
            btn.setToolTip(name)
            btn.setGraphicsEffect(None)
            btn.setStyleSheet("""
                QPushButton#PlatformTabButton {
                    background: rgba(9, 31, 68, 0.72);
                    border: 1px solid rgba(96, 165, 250, 0.30);
                    color: rgba(226, 239, 255, 0.88);
                    font-weight: 900;
                    font-size: 11px;
                    letter-spacing: 0.45px;
                    padding: 4px 13px;
                    border-radius: 14px;
                    text-align: center;
                }
                QPushButton#PlatformTabButton[active="true"] {
                    background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                        stop:0 rgba(37, 99, 235, 235),
                        stop:0.55 rgba(59, 130, 246, 245),
                        stop:1 rgba(14, 74, 160, 238));
                    border: 1px solid rgba(191, 226, 255, 0.98);
                    color: #FFFFFF;
                }
                QPushButton#PlatformTabButton[active="false"] {
                    background: rgba(8, 30, 64, 0.62);
                    border: 1px solid rgba(96, 165, 250, 0.26);
                    color: rgba(219, 234, 254, 0.88);
                }
                QPushButton#PlatformTabButton[active="false"]:hover {
                    background: rgba(30, 83, 170, 0.42);
                    border-color: rgba(147, 197, 253, 0.60);
                    color: #FFFFFF;
                }
                QPushButton#PlatformTabButton[active="true"]:hover {
                    border-color: rgba(224, 242, 254, 1.0);
                    background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                        stop:0 rgba(37, 99, 235, 245),
                        stop:0.55 rgba(59, 130, 246, 255),
                        stop:1 rgba(14, 74, 160, 245));
                }
            """)
            btn.clicked.connect(lambda _=False, platform_id=pid: self._set_active(platform_id))
            btn.ensurePolished()
            self._lay.addWidget(btn, 0, Qt.AlignVCenter)
            self._buttons[pid] = btn
            total_width += chip_width
        self._sync_measured_heights()
        self._refresh_button_states()
        if self._platforms:
            total_width += self._lay.spacing() * max(0, len(self._platforms) - 1)
        left, _top, right, _bottom = margins
        total_width += left + right
        self._host.setFixedSize(max(1, total_width), self._host_height())
        self._content_width = max(70, total_width)
        fixed_width = min(self._content_width, self._max_width)
        if len(self._platforms) >= 4:
            fixed_width = min(self._max_width, max(self._min_scroll_width, fixed_width))
        self.setMinimumWidth(min(fixed_width, self._min_scroll_width if len(self._platforms) >= 4 else fixed_width))
        self.setMaximumWidth(self._max_width)
        self.setFixedWidth(fixed_width)
        self._rail.setFixedWidth(fixed_width)
        self._scroll.setFixedWidth(max(1, fixed_width - 4))
        self.updateGeometry()
        self._ensure_active_visible()
        self.setToolTip(", ".join(str(p.get("platform_name") or "") for p in self._platforms))

    def _ensure_active_visible(self):
        active_btn = self._buttons.get(int(self._active or 0))
        if active_btn is not None:
            self._scroll.ensureWidgetVisible(active_btn, 24, 0)

    def _refresh_button_states(self):
        active_id = int(self._active or 0)
        for pid, btn in self._buttons.items():
            active = bool(pid) and int(pid) == active_id
            btn.setProperty("active", "true" if active else "false")
            btn.setChecked(active)
            if active:
                glow = QGraphicsDropShadowEffect(btn)
                glow.setBlurRadius(18)
                glow.setOffset(0, 0)
                glow.setColor(QColor(56, 189, 248, 120))
                btn.setGraphicsEffect(glow)
            else:
                btn.setGraphicsEffect(None)
            btn.style().unpolish(btn)
            btn.style().polish(btn)
            btn.update()
        self._ensure_active_visible()

    def wheelEvent(self, event):
        delta = event.angleDelta().x() or event.angleDelta().y()
        if delta:
            bar = self._scroll.horizontalScrollBar()
            bar.setValue(bar.value() - delta)
            event.accept()
            return
        super().wheelEvent(event)

    def sizeHint(self) -> QSize:
        width = min(max(70, self._content_width), self._max_width)
        if len(self._platforms) >= 4:
            width = max(self._min_scroll_width, width)
        return QSize(width, self._rail_height)

    def minimumSizeHint(self) -> QSize:
        width = min(max(70, self._content_width), self._max_width)
        if len(self._platforms) >= 4:
            width = min(width, self._min_scroll_width)
        return QSize(width, self._rail_height)

    def _host_height(self) -> int:
        left, top, right, bottom = self._lay.getContentsMargins()
        button_height = max((btn.sizeHint().height() for btn in self._buttons.values()), default=0)
        return max(1, button_height + top + bottom)

    def _sync_measured_heights(self) -> None:
        host_height = self._host_height()
        _left, rail_top, _right, rail_bottom = self._rail.layout().getContentsMargins()
        rail_frame = self._rail.frameWidth() * 2
        rail_height = max(self.DEFAULT_RAIL_HEIGHT, host_height + rail_top + rail_bottom + rail_frame)
        self._rail_height = rail_height
        self._host.setFixedHeight(host_height)
        self._scroll.setFixedHeight(host_height)
        self._rail.setFixedHeight(rail_height)
        self.setFixedHeight(rail_height)

    def _set_active(self, platform_id: int):
        platform_id = int(platform_id or 0)
        if platform_id == int(self._active or 0):
            return
        self._active = platform_id
        self._refresh_button_states()
        self.activePlatformChanged.emit(platform_id)


class HeaderUserPopup(QFrame):
    """Modern owner list popup for the contract detail header."""

    def __init__(self, parent=None):
        super().__init__(parent, Qt.Popup | Qt.FramelessWindowHint)
        self.setObjectName("HeaderUserPopup")
        self.setAttribute(Qt.WA_ShowWithoutActivating, True)
        self.setMouseTracking(True)
        self.setMinimumWidth(220)
        self.setMinimumHeight(80)
        self._users: List[str] = []
        self._lay = QVBoxLayout(self)
        self._lay.setContentsMargins(12, 10, 12, 12)
        self._lay.setSpacing(8)
        self.setStyleSheet("""
            QFrame#HeaderUserPopup {
                background-color: #071C3A;
                border: 1px solid #2D6EB8;
                border-radius: 10px;
            }
            QLabel#HeaderUserPopupTitle {
                color: #8fc8ff;
                background: transparent;
                border: 0;
                font-size: 10px;
                font-weight: 900;
            }
            QLabel#HeaderUserPopupName {
                color: #ffffff;
                background: transparent;
                border: 0;
                font-size: 12px;
                font-weight: 800;
            }
            QLabel#HeaderUserPopupAvatar {
                color: #eaf5ff;
                background-color: #2F7DFF;
                border: 1px solid #96D3FF;
                border-radius: 11px;
                font-size: 9px;
                font-weight: 900;
            }
        """)

    def set_users(self, users: List[str]) -> None:
        clean_users = [str(u).strip() for u in users if str(u).strip()]
        if clean_users == self._users and self._lay.count():
            return
        self._users = clean_users
        while self._lay.count():
            item = self._lay.takeAt(0)
            widget = item.widget()
            if widget:
                widget.setParent(None)
                widget.deleteLater()
        title = QLabel("SÖZLEŞME SAHİPLERİ")
        title.setObjectName("HeaderUserPopupTitle")
        self._lay.addWidget(title)
        for name in self._users:
            row = QWidget(self)
            row.setMouseTracking(True)
            lay = QHBoxLayout(row)
            lay.setContentsMargins(0, 0, 0, 0)
            lay.setSpacing(8)
            avatar = QLabel(self._initials(name))
            avatar.setObjectName("HeaderUserPopupAvatar")
            avatar.setAlignment(Qt.AlignCenter)
            avatar.setFixedSize(22, 22)
            label = QLabel(name)
            label.setObjectName("HeaderUserPopupName")
            label.setMinimumWidth(130)
            lay.addWidget(avatar, 0, Qt.AlignVCenter)
            lay.addWidget(label, 1, Qt.AlignVCenter)
            self._lay.addWidget(row)
        self.adjustSize()

    @staticmethod
    def _initials(name: str) -> str:
        parts = [p for p in re.split(r"\s+", str(name or "").strip()) if p]
        if not parts:
            return "?"
        letters = "".join(part[0] for part in parts[:2]).upper()
        return letters[:2]


class ContractDialog(StyledDialog):
    def __init__(self, store: ExcelStore, parent=None):
        super().__init__("Yeni Sözleşme", parent)
        self.store = store
        self.user_records = store.load_users()
        self.user_to_yi_yd = {u.get("name", ""): u.get("yi_yd", "Yİ") for u in self.user_records}
        self.current_staff = getattr(parent, "current_staff", None) if parent is not None else auth.current_staff
        self.staff_records = store.list_staff_for_engineer_selection() if hasattr(store, "list_staff_for_engineer_selection") else []
        self.result: Optional[ContractInfo] = None
        self._sd_verified_info: Optional[dict] = None
        self._sd_anchor_start_row: int = 0
        self._sd_anchor_end_row: int = 0
        self._sd_anchor_platform: str = ""
        self._sd_anchor_no: str = ""
        self._default_size = QSize(820, 600)
        self.build()
        self._resize_to_safe_default()

    def build(self):
        root = QVBoxLayout(self)
        title = QLabel("Yeni Sözleşme")
        title.setObjectName("dialogTitle")
        root.addWidget(title)
        desc = QLabel("Ana sözleşme bilgilerini girin. SD kayıtları ana sözleşme detay ekranından eklenecek.")
        desc.setObjectName("muted")
        root.addWidget(desc)

        grid = QGridLayout()
        self.no = QLineEdit(); self.no.setPlaceholderText("Örn: SZL-2026-001")

        # Sözleşme No satırı: input + Doğrula butonu + hata etiketi alt alta
        no_container = QWidget(self)
        no_container_lay = QVBoxLayout(no_container)
        no_container_lay.setContentsMargins(0, 0, 0, 0)
        no_container_lay.setSpacing(2)
        no_row = QWidget()
        no_lay = QHBoxLayout(no_row)
        no_lay.setContentsMargins(0, 0, 0, 0)
        no_lay.setSpacing(6)
        no_lay.addWidget(self.no, 1)
        self.verify_btn = QPushButton("Doğrula")
        self.verify_btn.setObjectName("secondary")
        self.verify_btn.setMinimumHeight(34)
        no_lay.addWidget(self.verify_btn, 0)
        self.no_dup_warn = QLabel("")
        self.no_dup_warn.setStyleSheet(
            "color:#dc2626; font-size:11px; font-weight:700; padding:0px;"
        )
        self.no_dup_warn.setVisible(False)
        no_container_lay.addWidget(no_row)
        no_container_lay.addWidget(self.no_dup_warn)

        self.platform = MultiPlatformSelectWidget(self); self.platform.set_platforms(self.store.load_platforms() if hasattr(self.store, "load_platforms") else self.store.platform_names())
        self.user = MultiUserSelectWidget(self)
        self.user.set_available_users([u.get("name", "") for u in self.user_records])
        self.yi_yd = QLineEdit(); self.yi_yd.setReadOnly(True); self.yi_yd.setText("Yİ")
        self.responsible_engineers = MultiStaffSelectWidget(self)
        self.responsible_engineers.set_staff_options(self.staff_records)
        # Sorumlu mühendis sözleşme seviyesinde ayrı ve opsiyonel bir alandır; varsayılan seçim yapılmaz.
        self.ctype = QComboBox(); self.ctype.addItems(["-", "Ana Sözleşme"])
        self.sd_code = QLineEdit(); self.sd_code.setPlaceholderText("SD-1"); self.sd_code.setEnabled(False)
        self.note = QLineEdit(); self.note.setPlaceholderText("Not")

        self.user.changed.connect(self._on_user_selection_changed)
        self.ctype.currentTextChanged.connect(self.on_contract_type_changed)
        self.verify_btn.clicked.connect(lambda: self.verify_sd_reference(show_message=False))
        self.platform.currentTextChanged.connect(self.on_sd_ref_changed)
        self.no.textChanged.connect(self.on_sd_ref_changed)
        # Anlık duplikasyon kontrolü: no veya platform veya tip değişince tekrar kontrol
        self.no.editingFinished.connect(self._check_no_duplicate)
        self.platform.currentIndexChanged.connect(lambda _: self._check_no_duplicate())
        self.ctype.currentIndexChanged.connect(lambda _: self._check_no_duplicate())

        def add_field(label: str, widget, row: int, col: int):
            grid.addWidget(form_label(label), row * 2, col)
            grid.addWidget(widget, row * 2 + 1, col)

        self.unknown_no_btn = QPushButton("No bilinmiyor")
        self.unknown_no_btn.setObjectName("secondary")
        no_lay.addWidget(self.unknown_no_btn, 0)
        self.unknown_no_btn.clicked.connect(self.fill_unknown_contract_no)
        add_field("Platform", self.platform, 0, 0)
        add_field("Sözleşme No", no_container, 0, 1)
        add_field("Sözleşmenin Sahibi Kullanıcı", self.user, 1, 0)
        add_field("Sorumlu Mühendis", self.responsible_engineers, 1, 1)
        add_field("Sözleşme Tipi", self.ctype, 2, 0)
        grid.setColumnStretch(0, 1)
        grid.setColumnStretch(1, 1)
        root.addLayout(grid)

        root.addWidget(form_label("Not"))
        root.addWidget(self.note)

        self.sd_verify_hint = QLabel("")
        self.sd_verify_hint.setObjectName("muted")
        self.sd_verify_hint.setVisible(False)
        root.addWidget(self.sd_verify_hint)

        self.update_user_yi_yd()
        self.on_contract_type_changed()

        row = QHBoxLayout(); row.addStretch()
        save = QPushButton("Devam Et")
        save.clicked.connect(self.save)
        row.addWidget(save)
        root.addLayout(row)

    def fill_unknown_contract_no(self):
        platform = self.platform.currentText().strip()
        if not platform:
            self.no_dup_warn.setText("Geçici sözleşme numarası oluşturmak için önce platform seçin.")
            self.no_dup_warn.setVisible(True)
            return
        pattern = re.compile(rf"^\s*{re.escape(platform)}\s*-\s*TBD\s*-\s*(\d+)\s*$", re.IGNORECASE)
        max_n = 0
        try:
            for ex in self.store.list_main_contracts(platform):
                m = pattern.match(str(ex.get("no", "") or ""))
                if m:
                    max_n = max(max_n, int(m.group(1)))
        except Exception:
            pass
        self.no.setText(f"{platform} - TBD - {max_n + 1}")
        self.no_dup_warn.setVisible(False)

    def _resize_to_safe_default(self):
        layout = self.layout()
        if layout is not None:
            layout.invalidate()
            layout.activate()
        hint = self.minimumSizeHint()
        size_hint = self.sizeHint()
        target = QSize(
            max(self._default_size.width(), hint.width(), size_hint.width()),
            max(self._default_size.height(), hint.height(), size_hint.height()),
        )
        screen = QApplication.screenAt(self.pos()) or QApplication.primaryScreen()
        if screen:
            available = screen.availableGeometry().size()
            target.setWidth(min(target.width(), max(720, available.width() - 80)))
            target.setHeight(min(target.height(), max(560, available.height() - 100)))
        self.resize(target)

    def _on_user_selection_changed(self):
        self.update_user_yi_yd()
        QTimer.singleShot(0, self._resize_for_user_selection)

    def _resize_for_user_selection(self):
        layout = self.layout()
        if layout is not None:
            layout.invalidate()
            layout.activate()
        self.user.updateGeometry()
        hint = self.minimumSizeHint()
        preferred_height = max(self._default_size.height(), self.sizeHint().height(), hint.height())
        screen = QApplication.primaryScreen()
        if screen is not None:
            available_height = screen.availableGeometry().height() - 80
            if available_height >= hint.height():
                preferred_height = min(preferred_height, available_height)
        self.resize(max(self.width(), self._default_size.width(), hint.width()), preferred_height)

    def is_sd_mode(self) -> bool:
        return self.ctype.currentText().strip() == "Sözleşme Değişikliği"

    def on_sd_ref_changed(self):
        self._sd_verified_info = None
        self._sd_anchor_start_row = 0
        self._sd_anchor_end_row = 0
        self._sd_anchor_platform = ""
        self._sd_anchor_no = ""
        if self.is_sd_mode():
            suggested = self.store.next_sd_code(self.platform.currentText(), self.no.text().strip())
            if not self.sd_code.text().strip():
                self.sd_code.setText(suggested)
            self.sd_verify_hint.setText("Doğrulama bekleniyor.")
            self.sd_verify_hint.setStyleSheet("color:#64748b; font-weight:700;")

    def on_contract_type_changed(self):
        sd = self.is_sd_mode()
        self.sd_code.setEnabled(sd)
        self.verify_btn.setVisible(sd)
        if hasattr(self, "sd_label"):
            self.sd_label.setVisible(sd)
        self.sd_code.setVisible(sd)
        self.sd_verify_hint.setVisible(sd)
        if sd:
            self.no.setPlaceholderText("Mevcut kontrat no")
            if not self.sd_code.text().strip():
                self.sd_code.setText(self.store.next_sd_code(self.platform.currentText(), self.no.text().strip()))
            self.sd_verify_hint.setText("Doğrulama bekleniyor.")
            self.sd_verify_hint.setStyleSheet("color:#64748b; font-weight:700;")
            self.user.setEnabled(False)
            self.yi_yd.setReadOnly(True)
        else:
            self.no.setPlaceholderText("Örn: SZL-2026-001")
            self.sd_code.clear()
            self.sd_verify_hint.clear()
            self.sd_verify_hint.setStyleSheet("")
            self._sd_verified_info = None
            self._sd_anchor_start_row = 0
            self._sd_anchor_end_row = 0
            self._sd_anchor_platform = ""
            self._sd_anchor_no = ""
            self.user.setEnabled(True)
            self.update_user_yi_yd()

    def _set_user_from_main_contract(self, info: dict):
        target_user = str(info.get("user", "") or "").strip()
        if target_user:
            cur = self.user.selected_users()
            if target_user not in cur:
                cur = [target_user]
            self.user.set_users(cur)
        yi_yd = str(info.get("yi_yd", "Yİ") or "Yİ").strip().upper()
        self.yi_yd.setText("YD" if yi_yd == "YD" else "Yİ")

    def verify_sd_reference(self, show_message: bool = True) -> bool:
        if not self.is_sd_mode():
            return True
        no = self.no.text().strip()
        platform = self.platform.currentText().strip()
        if not no or not platform:
            if show_message:
                QMessageBox.warning(self, "Eksik", "Önce platform ve kontrat no girin.")
            self.sd_verify_hint.setText("Önce platform ve kontrat no girin.")
            self.sd_verify_hint.setStyleSheet("color:#b91c1c; font-weight:700;")
            return False
        info = self.store.find_main_contract_info(platform, no)
        if not info:
            if show_message:
                QMessageBox.warning(self, "Bulunamadı", "Bu platformda girilen kontrat no için Ana Sözleşme bulunamadı.")
            self._sd_verified_info = None
            self._sd_anchor_start_row = 0
            self._sd_anchor_end_row = 0
            self._sd_anchor_platform = ""
            self._sd_anchor_no = ""
            self.sd_verify_hint.setText("✗ Ana sözleşme bulunamadı.")
            self.sd_verify_hint.setStyleSheet("color:#b91c1c; font-weight:700;")
            return False
        self._sd_verified_info = info
        self._sd_anchor_start_row = int(info.get("block_start") or info.get("row") or 0)
        self._sd_anchor_end_row = int(info.get("block_end") or self._sd_anchor_start_row or 0)
        self._sd_anchor_platform = str(platform or "")
        self._sd_anchor_no = str(no or "")
        self._set_user_from_main_contract(info)
        # Her doğrulamada ilgili kontrat için bir sonraki SD kodunu otomatik getir.
        self.sd_code.setText(self.store.next_sd_code(platform, no))
        self.sd_verify_hint.setText(f"✓ Ana sözleşme bulundu: {no}")
        self.sd_verify_hint.setStyleSheet("color:#047857; font-weight:800;")
        if show_message:
            QMessageBox.information(self, "Doğrulandı", f"Ana sözleşme bulundu. SD kaydı {no} kontrat no altında eklenecek.")
        return True

    def update_completion_date(self):
        return

    def date_picker_events(self) -> List[dict]:
        return []

    def update_user_yi_yd(self):
        if self.is_sd_mode() and self._sd_verified_info:
            self._set_user_from_main_contract(self._sd_verified_info)
            return
        selected = (self.user.selected_users() or [""])[0].strip()
        yi_yd = self.user_to_yi_yd.get(selected, "Yİ")
        self.yi_yd.setText("YD" if str(yi_yd).upper() == "YD" else "Yİ")

    def _normalized_sd_code(self) -> str:
        raw = str(self.sd_code.text() or "").strip().upper().replace(" ", "")
        if not raw:
            return ""
        m = re.match(r"^SD[-_]?(\d+)$", raw)
        if m:
            return f"SD-{int(m.group(1))}"
        return ""

    def _check_no_duplicate(self):
        """Sözleşme no + platform + tip kombinasyonu zaten varsa kırmızı uyarı göster."""
        if not hasattr(self, 'no_dup_warn'):
            return
        no = self.no.text().strip()
        platform = self.platform.currentText().strip()
        if not no or not platform:
            self.no_dup_warn.setVisible(False)
            self.no.setStyleSheet("")
            return
        contract_type = self.ctype.currentText().strip() or "-"
        if self.is_sd_mode():
            sd = self._normalized_sd_code()
            contract_type = sd if sd else self.sd_code.text().strip()
        try:
            existing = self.store.list_main_contracts(platform)
            for ex in existing:
                ex_no = self.store._normalize_label(str(ex.get("no", "") or "").strip())
                ex_type = self.store._normalize_label(str(ex.get("type", "") or "").strip())
                if (ex_no == self.store._normalize_label(no) and
                        ex_type == self.store._normalize_label(contract_type)):
                    self.no_dup_warn.setText(
                        f"⚠  '{platform}' platformunda bu sözleşme no zaten mevcut!"
                    )
                    self.no_dup_warn.setVisible(True)
                    self.no.setStyleSheet(
                        "QLineEdit{border:1.5px solid #dc2626; background:#fff5f5;}"
                    )
                    return
        except Exception:
            pass
        self.no_dup_warn.setVisible(False)
        self.no.setStyleSheet("")

    def _highlight_required(self, widget, error: bool):
        """Zorunlu alan boşsa kırmızı çerçeve, doluysa normal."""
        if error:
            widget.setStyleSheet("QLineEdit{border:1.5px solid #dc2626; background:#fff5f5;}"
                                 "QSpinBox{border:1.5px solid #dc2626; background:#fff5f5;}")
        else:
            widget.setStyleSheet("")

    def selected_platform_ids(self) -> List[int]:
        return self.platform.selected_platform_ids() if hasattr(self.platform, "selected_platform_ids") else []

    def _confirm_empty_responsible_engineer(self) -> bool:
        return True

    def save(self):
        if not self.no.text().strip():
            QMessageBox.warning(self, "Eksik", "Sözleşme no girin.")
            return
        if not self.platform.selected_platforms():
            QMessageBox.warning(self, "Eksik", "Lütfen en az bir platform seçiniz.")
            return
        if self.is_sd_mode() and not self.verify_sd_reference(show_message=False):
            QMessageBox.warning(self, "Doğrulama", "Sözleşme Değişikliği için önce geçerli kontrat no doğrulaması gerekir.")
            return
        sel_users = self.user.selected_users()
        if not sel_users:
            QMessageBox.warning(self, "Eksik", "Önce Kullanıcı Yönetimi ekranından kullanıcı tanımlayın.")
            return
        if not self._confirm_empty_responsible_engineer():
            return
        contract_type = self.ctype.currentText().strip() or "-"
        if self.is_sd_mode():
            sd_code = self._normalized_sd_code()
            if not sd_code:
                QMessageBox.warning(self, "Format", "SD kodu SD-1, SD-2 gibi sayısal formatta olmalı.")
                return
            self.sd_code.setText(sd_code)
            contract_type = sd_code

        # Aynı platform + sözleşme no + tip kombinasyonuna izin verme
        platform_check = self.platform.currentText()
        no_check = self.no.text().strip()
        try:
            existing_contracts = self.store.list_main_contracts(platform_check)
            for ex in existing_contracts:
                ex_no = self.store._normalize_label(str(ex.get("no", "") or "").strip())
                ex_type = self.store._normalize_label(str(ex.get("type", "") or "").strip())
                if (ex_no == self.store._normalize_label(no_check) and
                        ex_type == self.store._normalize_label(contract_type)):
                    QMessageBox.warning(
                        self, "Tekrar Eden Kayıt",
                        f"'{platform_check}' platformunda '{no_check}' sözleşme numarası ve "
                        f"'{contract_type}' tipi için zaten bir kayıt mevcut.\n\n"
                        "Aynı platform + no + tip kombinasyonu kullanılamaz."
                    )
                    return
        except Exception:
            pass  # Kontrol başarısız olursa devam et

        users = self.user.selected_users()
        user_display = ", ".join(users)
        self.result = ContractInfo(
            no=self.no.text().strip(),
            platform=self.platform.currentText(),
            user=user_display,
            yi_yd=self.yi_yd.text().strip() or "Yİ",
            contract_type=contract_type,
            signature_date="",
            t0_date="",
            t0_months=0,
            completion_date="",
            status="Başlanmadı",
            note=self.note.text().strip(),
            acceptance_date="",
            sd_anchor_start_row=self._sd_anchor_start_row if self.is_sd_mode() else 0,
            sd_anchor_end_row=self._sd_anchor_end_row if self.is_sd_mode() else 0,
            sd_anchor_platform=self._sd_anchor_platform if self.is_sd_mode() else "",
            sd_anchor_no=self._sd_anchor_no if self.is_sd_mode() else "",
            users=users,
            platforms=self.platform.selected_platform_records(),
            platform_names=self.platform.selected_platform_names(),
            platform_ids=self.platform.selected_platform_ids(),
        )
        responsible_id = self.responsible_engineers.selected_staff_id()
        responsible_name = self.responsible_engineers._staff_name_by_id.get(responsible_id, "") if responsible_id else ""
        self.result.responsible_engineer_id = responsible_id
        self.result.responsible_engineer_name = responsible_name
        setattr(self.result, "responsible_engineer_ids", [responsible_id] if responsible_id else [])
        setattr(self.result, "responsible_engineers", [
            {"staff_id": responsible_id, "full_name": responsible_name}
        ] if responsible_id else [])
        self.accept()


class ContractEditDialog(StyledDialog):
    """Mevcut sözleşmenin ANA BİLGİLERİNİ güncelleme ekranı.

    - Sözleşme No düzenlenebilir; Platform ve Sözleşme Tipi salt okunur.
    - Diğer temel alanlar düzenlenebilir; sistemler ve teslimatlar değişmez.
    - Güncellemede platform + sözleşme tipi + sözleşme no kombinasyonu tekil kalır.
    """

    def __init__(
        self,
        store: ExcelStore,
        ci: ContractInfo,
        parent=None,
        title_text: str = "Ana Bilgileri Düzenle",
        save_text: str = "Güncelle",
        info_text: Optional[str] = None,
    ):
        super().__init__(title_text, parent)
        self.store = store
        self.ci = ci
        self.title_text = title_text
        self.save_text = save_text
        self.info_text = info_text
        self.external_events_provider = getattr(parent, "date_picker_events", None)
        self.user_records = self.store.load_users()
        self.user_to_yi_yd = {u.get("name", ""): u.get("yi_yd", "Yİ") for u in self.user_records}
        self.staff_records = self.store.list_staff_for_engineer_selection() if hasattr(self.store, "list_staff_for_engineer_selection") else []
        self.result: Optional[ContractInfo] = None
        self._default_size = QSize(820, 660)
        self.build()
        self._resize_to_safe_default()

    def build(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(20, 14, 20, 14)
        root.setSpacing(10)

        title = QLabel(self.title_text)
        title.setObjectName("dialogTitle")
        root.addWidget(title)

        info = QLabel(
            self.info_text or
            "Yalnızca sözleşmenin temel bilgileri güncellenir. "
            "Sistemler ve teslimatlar değişmez."
        )
        info.setObjectName("muted")
        info.setWordWrap(True)
        root.addWidget(info)

        self._form_scroll = QScrollArea(self)
        self._form_scroll.setFrameShape(QFrame.NoFrame)
        self._form_scroll.setWidgetResizable(True)
        self._form_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self._form_scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self._form_scroll.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self._form_scroll.setStyleSheet("QScrollArea{background:transparent;border:0;} QScrollArea > QWidget > QWidget{background:transparent;}")
        self._form_container = QWidget(self._form_scroll)
        self._form_container.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        form = QVBoxLayout(self._form_container)
        form.setContentsMargins(0, 0, 0, 0)
        form.setSpacing(12)
        self._form_scroll.setWidget(self._form_container)
        root.addWidget(self._form_scroll, 1)

        grid = QGridLayout()
        grid.setHorizontalSpacing(16)
        grid.setVerticalSpacing(10)
        grid.setColumnStretch(0, 1)
        grid.setColumnStretch(1, 1)

        def readonly(text: str) -> QLineEdit:
            w = QLineEdit(str(text or ""))
            w.setReadOnly(True)
            w.setStyleSheet("background:#f1f5f9; color:#64748B; border:1px solid #e2e8f0;")
            return w

        # ── Salt okunur alanlar ──────────────────────────────────────
        ctype_text = str(self.ci.contract_type or "").strip()
        self._is_sd_contract = bool(
            re.match(r"^SD-\d+$", ctype_text.upper()) or
            self.store._normalize_label(ctype_text) == self.store._normalize_label("Sözleşme Değişikliği")
        )
        self._no_lbl       = QLineEdit(str(self.ci.no or ""))
        self._locked_platforms = self._initial_platforms()
        self._platform_select = MultiPlatformSelectWidget(self)
        self._platform_select.set_platforms(self.store.load_platforms() if hasattr(self.store, "load_platforms") else self.store.platform_names())
        self._platform_select.set_selected_platforms(self._locked_platforms)
        self._platform_select.set_locked_platforms(self._locked_platforms)
        self._platform_help = QLabel("Kayıtlı platformlar çıkarılamaz; yalnızca yeni platform eklenebilir.")
        self._platform_help.setObjectName("muted")
        self._platform_help.setWordWrap(True)
        self._platform_box = QWidget(self)
        self._platform_box.setStyleSheet("QWidget{background:transparent;border:0;}")
        platform_box_lay = QVBoxLayout(self._platform_box)
        platform_box_lay.setContentsMargins(0, 0, 0, 0)
        platform_box_lay.setSpacing(3)
        platform_box_lay.addWidget(self._platform_select)
        platform_box_lay.addWidget(self._platform_help)
        self._type_lbl     = readonly(self.ci.contract_type)
        self._type_lbl.setPlaceholderText("Örn: SD-1")
        if self._is_sd_contract:
            self._no_lbl.setReadOnly(True)
            self._no_lbl.setStyleSheet("background:#f1f5f9; color:#64748B; border:1px solid #e2e8f0;")
            self._type_lbl.setReadOnly(False)
            self._type_lbl.setEnabled(True)
            self._type_lbl.setStyleSheet("")
            self._type_lbl.editingFinished.connect(self._normalize_sd_code_field)
            self._type_lbl.textChanged.connect(self._check_duplicate_contract_key)
            no_warn_text = ""
        else:
            no_warn_text = "Aynı platform + sözleşme tipi + sözleşme no kombinasyonu kullanılamaz."
            self._no_lbl.textChanged.connect(self._check_duplicate_contract_key)
        self._no_dup_warn = QLabel(no_warn_text)
        self._no_dup_warn.setObjectName("warning")
        self._no_dup_warn.setWordWrap(True)
        self._no_dup_warn.setVisible(False)

        # ── Düzenlenebilir alanlar ────────────────────────────────────
        self.user = MultiUserSelectWidget(self)
        self.user.set_available_users([u.get("name", "") for u in self.user_records])
        init_users = list(getattr(self.ci, "users", []) or [])
        if not init_users and str(self.ci.user or "").strip():
            init_users = [x.strip() for x in str(self.ci.user or "").split(",") if x.strip()]
        self.user.set_users(init_users)

        self.yi_yd = QLineEdit()
        self.yi_yd.setReadOnly(True)
        self.yi_yd.setText(str(self.ci.yi_yd or "Yİ"))

        self.responsible_engineers = MultiStaffSelectWidget(self)
        self.responsible_engineers.set_staff_options(self.staff_records)
        responsible_ids = [int(getattr(self.ci, "responsible_engineer_id", 0) or 0)]
        if not responsible_ids[0]:
            responsible_ids = [int(x.get("staff_id") or x.get("id") or 0) for x in list(getattr(self.ci, "responsible_engineers", []) or []) if int(x.get("staff_id") or x.get("id") or 0)]
        if not responsible_ids:
            responsible_ids = [int(x or 0) for x in list(getattr(self.ci, "responsible_engineer_ids", []) or []) if int(x or 0)]
        self.responsible_engineers.set_selected_staff_ids(responsible_ids[:1])

        self.note = QLineEdit()
        self.note.setPlaceholderText("Not")
        self.note.setText(str(self.ci.note or ""))

        self.user.changed.connect(self.update_user_yi_yd)
        self.user.changed.connect(self._on_dynamic_field_changed)
        self._platform_select.currentTextChanged.connect(lambda _text: self._on_dynamic_field_changed())
        self.responsible_engineers.changed.connect(self._on_dynamic_field_changed)
        self.update_user_yi_yd()

        def add_field(label: str, widget, row: int, col: int):
            grid.addWidget(form_label(label), row * 2, col)
            grid.addWidget(widget, row * 2 + 1, col)

        add_field("Sözleşme No", self._no_lbl, 0, 0)
        add_field("Platform", self._platform_box, 0, 1)
        add_field("Sözleşmenin Sahibi Kullanıcı", self.user, 1, 0)
        add_field("Sorumlu Mühendis", self.responsible_engineers, 1, 1)
        add_field("Sözleşme Tipi", self._type_lbl, 2, 0)
        form.addLayout(grid)

        form.addWidget(form_label("Not"))
        form.addWidget(self.note)
        form.addWidget(self._no_dup_warn)
        form.addStretch(1)

        btn_row = QHBoxLayout()
        btn_row.addStretch()
        cancel = QPushButton("İptal")
        cancel.setObjectName("secondary")
        cancel.clicked.connect(self.reject)
        save_btn = QPushButton(self.save_text)
        save_btn.setDefault(True)
        save_btn.setAutoDefault(True)
        save_btn.clicked.connect(self.save)
        self._save_btn = save_btn
        btn_row.addWidget(cancel)
        btn_row.addWidget(save_btn)
        root.addLayout(btn_row)

    def _on_dynamic_field_changed(self):
        for widget_name in ("_form_container", "_form_scroll", "user", "_platform_select", "responsible_engineers"):
            widget = getattr(self, widget_name, None)
            if isinstance(widget, QWidget):
                widget.updateGeometry()
        layout = getattr(self, "_form_container", None).layout() if hasattr(self, "_form_container") else None
        if layout is not None:
            layout.invalidate()
        QTimer.singleShot(0, self._resize_to_safe_default)

    def keyPressEvent(self, event):
        if event.key() in (Qt.Key_Return, Qt.Key_Enter):
            if isinstance(self.focusWidget(), QTextEdit):
                super().keyPressEvent(event)
                return
            self.save()
            return
        if event.key() == Qt.Key_Escape:
            self.reject()
            return
        super().keyPressEvent(event)

    def fill_unknown_contract_no(self):
        platform = self.platform.currentText().strip()
        if not platform:
            self.no_dup_warn.setText("Geçici sözleşme numarası oluşturmak için önce platform seçin.")
            self.no_dup_warn.setVisible(True)
            return
        pattern = re.compile(rf"^\s*{re.escape(platform)}\s*-\s*TBD\s*-\s*(\d+)\s*$", re.IGNORECASE)
        max_n = 0
        try:
            for ex in self.store.list_main_contracts(platform):
                m = pattern.match(str(ex.get("no", "") or ""))
                if m:
                    max_n = max(max_n, int(m.group(1)))
        except Exception:
            pass
        self.no.setText(f"{platform} - TBD - {max_n + 1}")
        self.no_dup_warn.setVisible(False)

    def _resize_to_safe_default(self):
        layout = self.layout()
        if layout is not None:
            layout.invalidate()
            layout.activate()
        hint = self.minimumSizeHint()
        size_hint = self.sizeHint()
        target = QSize(
            max(self._default_size.width(), hint.width(), size_hint.width()),
            max(self._default_size.height(), hint.height(), size_hint.height()),
        )
        screen = QApplication.screenAt(self.pos()) or QApplication.primaryScreen()
        if screen:
            available = screen.availableGeometry().size()
            target.setWidth(min(target.width(), max(720, available.width() - 80)))
            target.setHeight(min(target.height(), max(560, available.height() - 100)))
        self.resize(target)

    def _initial_platforms(self) -> List[str]:
        raw = list(getattr(self.ci, "platforms", []) or [])
        if not raw and str(self.ci.platform or "").strip():
            raw = [str(self.ci.platform or "").strip()]
        out: List[str] = []
        seen = set()
        for item in raw:
            n = str(item.get("platform_name") or item.get("name") or "").strip() if isinstance(item, dict) else str(item or "").strip()
            key = n.casefold()
            if n and key not in seen:
                seen.add(key)
                out.append(n)
        return out

    def _selected_platforms(self) -> List[str]:
        vals = self._platform_select.selected_platforms() if hasattr(self, "_platform_select") else []
        out: List[str] = []
        seen = set()
        for name in vals:
            n = str(name or "").strip()
            key = n.casefold()
            if n and key not in seen:
                seen.add(key)
                out.append(n)
        return out

    def _removed_locked_platforms(self) -> List[str]:
        selected_keys = {p.casefold() for p in self._selected_platforms()}
        return [p for p in self._locked_platforms if p.casefold() not in selected_keys]

    def update_user_yi_yd(self):
        selected = (self.user.selected_users() or [""])[0].strip()
        yi_yd = self.user_to_yi_yd.get(selected, "Yİ")
        self.yi_yd.setText("YD" if str(yi_yd).upper() == "YD" else "Yİ")


    def _recalc(self):
        return

    def date_picker_events(self) -> List[dict]:
        if callable(self.external_events_provider):
            try:
                return list(self.external_events_provider() or [])
            except Exception:
                return []
        return []

    def _normalized_sd_code(self) -> str:
        raw = str(self._type_lbl.text() or "").strip().upper().replace(" ", "")
        if not raw:
            return ""
        m = re.match(r"^SD[-_]?(\d+)$", raw)
        if m:
            return f"SD-{int(m.group(1))}"
        return ""

    def _normalize_sd_code_field(self):
        if not self._is_sd_contract:
            return
        sd_code = self._normalized_sd_code()
        if sd_code:
            self._type_lbl.setText(sd_code)

    def _current_contract_type_text(self) -> str:
        if self._is_sd_contract:
            return self._normalized_sd_code() or self._type_lbl.text().strip()
        return str(self.ci.contract_type or "").strip()

    def _check_duplicate_contract_key(self) -> bool:
        """Başka bir kayıtta aynı platform + tip + no varsa uyarı gösterir."""
        no_text = self._no_lbl.text().strip()
        platform = str(self.ci.platform or "").strip()
        contract_type = self._current_contract_type_text()
        if not no_text or not platform or not contract_type:
            self._no_dup_warn.setVisible(self._is_sd_contract)
            if not self._is_sd_contract:
                self._no_lbl.setStyleSheet("")
            return False

        norm_no = self.store._normalize_label(no_text)
        norm_type = self.store._normalize_label(contract_type)
        current_row = int(getattr(self.ci, "entry_start_row", 0) or 0)
        try:
            existing_contracts = self.store.list_main_contracts(platform)
        except Exception:
            existing_contracts = []

        def mark_duplicate(message: str) -> bool:
            self._no_dup_warn.setText(message)
            self._no_dup_warn.setVisible(True)
            if not self._is_sd_contract:
                self._no_lbl.setStyleSheet(
                    "QLineEdit{border:1.5px solid #dc2626; background:#fff5f5;}"
                )
            return True

        for ex in existing_contracts:
            ex_row = int(ex.get("row") or 0)
            if current_row > 0 and ex_row == current_row:
                continue
            ex_no = self.store._normalize_label(str(ex.get("no", "") or "").strip())
            ex_type = self.store._normalize_label(str(ex.get("type", "") or "").strip())
            if ex_no == norm_no and ex_type == norm_type:
                return mark_duplicate(
                    f"⚠ '{platform}' platformunda '{contract_type}' tipi için "
                    f"'{no_text}' sözleşme numarası zaten mevcut. "
                    "Aynı platform + tip + no kombinasyonu kullanılamaz."
                )

        # Ana sözleşme no değişirken aynı no'ya bağlı SD kayıtları da taşınacak.
        # Bu yüzden taşınacak her SD tipi için hedef no altında çakışma var mı önceden kontrol edilir.
        old_no = str(self.ci.no or "").strip()
        if (not self._is_sd_contract and
                self.store._normalize_label(contract_type) == self.store._normalize_label("Ana Sözleşme") and
                self.store._normalize_label(old_no) != norm_no):
            linked_sd_rows = set()
            linked_sd_types = []
            for ex in existing_contracts:
                ex_type_raw = str(ex.get("type", "") or "").strip()
                is_sd_type = bool(
                    re.match(r"^SD-\d+$", ex_type_raw.upper()) or
                    self.store._normalize_label(ex_type_raw) == self.store._normalize_label("Sözleşme Değişikliği")
                )
                if not is_sd_type:
                    continue
                if self.store._normalize_label(str(ex.get("no", "") or "").strip()) != self.store._normalize_label(old_no):
                    continue
                linked_sd_rows.add(int(ex.get("row") or 0))
                linked_sd_types.append(ex_type_raw)
            for sd_type in linked_sd_types:
                norm_sd_type = self.store._normalize_label(sd_type)
                for ex in existing_contracts:
                    if int(ex.get("row") or 0) in linked_sd_rows:
                        continue
                    ex_no = self.store._normalize_label(str(ex.get("no", "") or "").strip())
                    ex_type = self.store._normalize_label(str(ex.get("type", "") or "").strip())
                    if ex_no == norm_no and ex_type == norm_sd_type:
                        return mark_duplicate(
                            f"⚠ Ana sözleşme no güncellenirse bağlı '{sd_type}' kaydı da "
                            f"'{no_text}' no'ya taşınacak; ancak bu platformda aynı no ve SD tipi zaten var."
                        )

        self._no_dup_warn.setVisible(False)
        if not self._is_sd_contract:
            self._no_lbl.setStyleSheet("")
        return False

    def _confirm_empty_responsible_engineer(self) -> bool:
        return True

    def save(self):
        new_no_text = self._no_lbl.text().strip()
        if not new_no_text:
            QMessageBox.warning(self, "Zorunlu Alan", "Sözleşme No girilmelidir.")
            return
        removed_platforms = self._removed_locked_platforms()
        if removed_platforms:
            QMessageBox.warning(
                self,
                "Platform çıkarılamaz",
                "Kayıt yazıldıktan sonra mevcut platform çıkarılamaz. "
                "Sadece yeni platform ekleyebilirsiniz. Çıkarılan: " + ", ".join(removed_platforms)
            )
            self._platform_select.set_selected_platforms(self._locked_platforms + [p for p in self._selected_platforms() if p.casefold() not in {x.casefold() for x in self._locked_platforms}])
            return
        selected_platforms = self._selected_platforms()
        if not selected_platforms:
            QMessageBox.warning(self, "Zorunlu Alan", "Lütfen en az bir platform seçiniz.")
            return
        if self._check_duplicate_contract_key():
            QMessageBox.warning(
                self,
                "Tekrar Eden Kayıt",
                "Aynı platform, sözleşme tipi ve sözleşme no ile başka bir kayıt bulundu. "
                "Lütfen farklı bir sözleşme no girin."
            )
            return
        norm_no = self.store._normalize_label(new_no_text)
        norm_type = self.store._normalize_label(self._current_contract_type_text())
        locked_keys = {p.casefold() for p in self._locked_platforms}
        for platform_name in selected_platforms:
            if platform_name.casefold() in locked_keys:
                continue
            try:
                candidates = self.store.list_main_contracts(platform_name)
            except Exception:
                candidates = []
            for ex in candidates:
                if (self.store._normalize_label(str(ex.get("no", "") or "")) == norm_no and
                        self.store._normalize_label(str(ex.get("type", "") or "")) == norm_type):
                    QMessageBox.warning(
                        self,
                        "Tekrar Eden Kayıt",
                        f"'{platform_name}' platformunda aynı sözleşme no ve tip zaten var. "
                        "Bu platform eklenemez."
                    )
                    return
        contract_type = str(self.ci.contract_type or "").strip()
        if self._is_sd_contract:
            sd_code = self._normalized_sd_code()
            if not sd_code:
                QMessageBox.warning(self, "Format", "SD kodu SD-1, SD-2 gibi sayısal formatta olmalı.")
                return
            self._type_lbl.setText(sd_code)
            contract_type = sd_code
        new_ci = copy.copy(self.ci)
        new_ci.no              = new_no_text
        new_ci.contract_type   = contract_type
        selected_users = self.user.selected_users()
        if not selected_users:
            QMessageBox.warning(self, "Zorunlu Alan", "En az bir kullanıcı seçmelisiniz.")
            return
        if not self._confirm_empty_responsible_engineer():
            return
        new_ci.users           = selected_users
        new_ci.user            = ", ".join(selected_users)
        new_ci.yi_yd           = self.yi_yd.text().strip() or "Yİ"
        new_ci.signature_date  = str(getattr(self.ci, "signature_date", "") or "")
        new_ci.t0_date         = str(getattr(self.ci, "t0_date", "") or "")
        new_ci.t0_months       = int(getattr(self.ci, "t0_months", 0) or 0)
        new_ci.completion_date = str(getattr(self.ci, "completion_date", "") or "")
        new_ci.status          = str(self.ci.status or "Başlanmadı")
        new_ci.note            = self.note.text().strip()
        selected_platform_ids = self._platform_select.selected_platform_ids() if hasattr(self._platform_select, "selected_platform_ids") else []
        if not selected_platform_ids:
            for name in selected_platforms:
                pid = self.store.get_platform_id(name, create=False)
                if pid is not None:
                    selected_platform_ids.append(int(pid))
        setattr(new_ci, "platforms", [{"platform_id": pid, "platform_name": name} for pid, name in zip(selected_platform_ids, selected_platforms)])
        setattr(new_ci, "platform_names", selected_platforms)
        setattr(new_ci, "platform_ids", selected_platform_ids)
        responsible_id = self.responsible_engineers.selected_staff_id()
        responsible_name = self.responsible_engineers._staff_name_by_id.get(responsible_id, "") if responsible_id else ""
        new_ci.responsible_engineer_id = responsible_id
        new_ci.responsible_engineer_name = responsible_name
        setattr(new_ci, "responsible_engineer_ids", [responsible_id] if responsible_id else [])
        setattr(new_ci, "responsible_engineers", [
            {"staff_id": responsible_id, "full_name": responsible_name}
        ] if responsible_id else [])
        self.result = new_ci
        self.accept()


class TagAssignDialog(StyledDialog):
    def __init__(self, store: ExcelStore, already_assigned: Optional[List[dict]] = None, parent=None):
        super().__init__("Etiket Ekle", parent)
        self.store = store
        self.all_tags = list(store.load_tag_defs(active_only=True))
        self.already_keys = {
            self._tag_key(str((t or {}).get("name") or ""))
            for t in list(already_assigned or [])
            if str((t or {}).get("name") or "").strip()
        }
        self.available_tags = [tag for tag in self.all_tags if self._tag_key(tag.name) not in self.already_keys]
        self.selected: Dict[str, TagDef] = {}
        self.result: List[dict] = []
        self.save_btn: Optional[QPushButton] = None
        self.resize(520, 380)
        self.build()

    def _tag_key(self, name: str) -> str:
        return normalized_tag_key(name)

    def build(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(18, 14, 18, 14)
        root.setSpacing(10)
        title = QLabel("Etiket Ekle")
        title.setObjectName("dialogTitle")
        root.addWidget(title)

        root.addWidget(form_label("Etiket Seç"))
        self.tags_wrap = QFrame()
        self.tags_wrap.setObjectName("tagPanel")
        tags_lay = QGridLayout(self.tags_wrap)
        tags_lay.setContentsMargins(10, 10, 10, 10)
        tags_lay.setHorizontalSpacing(8)
        tags_lay.setVerticalSpacing(8)
        if not self.all_tags:
            warn = QLabel("Aktif etiket yok. Önce Etiket Yönetimi ekranından etiket oluşturun.")
            warn.setObjectName("warning")
            warn.setWordWrap(True)
            tags_lay.addWidget(warn, 0, 0, 1, 3)
        elif not self.available_tags:
            empty = QLabel("Atanabilecek etiket bulunmuyor.")
            empty.setObjectName("warning")
            empty.setWordWrap(True)
            tags_lay.addWidget(empty, 0, 0, 1, 3)
        else:
            for i, t in enumerate(self.available_tags):
                b = QPushButton(f"● {t.name}")
                b.setCheckable(True)
                b.setObjectName("tagChipBtn")
                b.setStyleSheet(tag_chip_style(t.color, selected=False))
                b.clicked.connect(lambda checked, tag=t, btn=b: self.toggle_tag(tag, btn, checked))
                tags_lay.addWidget(b, i // 3, i % 3)
        root.addWidget(self.tags_wrap)

        root.addWidget(form_label("Sözleşmeye Özel Not (Opsiyonel)"))
        self.note = QTextEdit()
        self.note.setPlaceholderText("Bu atama için özel bir not ekleyin...")
        self.note.setMinimumHeight(72)
        root.addWidget(self.note)

        row = QHBoxLayout()
        row.addStretch()
        cancel = QPushButton("İptal")
        cancel.setObjectName("secondary")
        cancel.clicked.connect(self.reject)
        save = QPushButton("Ekle")
        save.setEnabled(False)
        save.clicked.connect(self.save)
        self.save_btn = save
        row.addWidget(cancel)
        row.addWidget(save)
        root.addLayout(row)

    def toggle_tag(self, tag: TagDef, btn: QPushButton, checked: bool):
        key = self._tag_key(tag.name)
        if checked:
            self.selected[key] = tag
        else:
            self.selected.pop(key, None)
        btn.setStyleSheet(tag_chip_style(tag.color, selected=bool(checked)))
        if self.save_btn is not None:
            self.save_btn.setEnabled(bool(self.selected))

    def save(self):
        if not self.available_tags:
            return
        if not self.selected:
            QMessageBox.warning(self, "Seçim", "En az bir etiket seçin.")
            return
        if self.save_btn is not None:
            self.save_btn.setEnabled(False)
        note = self.note.toPlainText().strip()
        out: List[dict] = []
        for tag in self.selected.values():
            out.append({
                "name": str(tag.name or "").strip(),
                "color": str(tag.color or "#3B82F6"),
                "note": note or str(tag.note or "").strip(),
            })
        self.result = out
        self.accept()


class TagManagerDialog(StyledDialog):
    def __init__(self, store: ExcelStore, contract_index: Optional[List[dict]] = None, parent=None):
        super().__init__("Etiket Yönetimi", parent)
        self.store = store
        self.contract_index = list(contract_index or [])
        self.changed = False
        self.tags: List[TagDef] = []
        self.usage: Dict[str, int] = {}
        self.assignments_by_key: Dict[str, List[dict]] = {}
        self._contract_map: Dict[Tuple[str, str, str], dict] = {}
        self.selected_tag_key: Optional[str] = None
        self.selected_color = "#3B82F6"
        self._color_buttons: List[QPushButton] = []
        self._draft_tags: Dict[str, TagDef] = {}
        self._draft_order: List[str] = []
        self._draft_seq = 1
        self.resize(1240, 700)
        self.build()
        self._rebuild_contract_map()
        self.reload_data(keep_selection=False)

    def build(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(12, 10, 12, 10)
        root.setSpacing(10)

        top = QHBoxLayout()
        title = QLabel("Etiket Yönetimi")
        title.setObjectName("dialogTitle")
        top.addWidget(title, 1)
        new_btn = QPushButton("+ Yeni Etiket")
        new_btn.clicked.connect(self.new_tag)
        top.addWidget(new_btn, 0)
        root.addLayout(top)

        body = QHBoxLayout()
        body.setSpacing(10)
        root.addLayout(body, 1)

        left = QFrame()
        left.setObjectName("panel")
        left.setFixedWidth(310)
        ll = QVBoxLayout(left)
        ll.setContentsMargins(10, 10, 10, 10)
        ll.setSpacing(8)
        ll.addWidget(form_label("Etiketler"))
        self.search = QLineEdit()
        self.search.setPlaceholderText("Etiket ara...")
        self.search.textChanged.connect(self.refresh_tag_list)
        ll.addWidget(self.search)
        self.tag_list = QListWidget()
        self.tag_list.setObjectName("tagList")
        self.tag_list.currentRowChanged.connect(self.on_tag_selected)
        self.tag_list.setAlternatingRowColors(False)
        ll.addWidget(self.tag_list, 1)
        body.addWidget(left, 0)

        right = QFrame()
        right.setObjectName("panel")
        rl = QVBoxLayout(right)
        rl.setContentsMargins(14, 12, 14, 12)
        rl.setSpacing(10)
        body.addWidget(right, 1)

        detail_row = QHBoxLayout()
        detail_row.addWidget(section_label("ETİKET DETAYI"), 1)
        self.contract_count = QLabel("0 bağlı sözleşme")
        self.contract_count.setObjectName("ctxPill")
        detail_row.addWidget(self.contract_count, 0)
        rl.addLayout(detail_row)

        form = QGridLayout()
        form.setHorizontalSpacing(12)
        form.setVerticalSpacing(8)
        self.name_edit = QLineEdit()
        self.note_edit = QTextEdit()
        self.note_edit.setMinimumHeight(86)
        form.addWidget(form_label("Etiket Adı"), 0, 0)
        form.addWidget(self.name_edit, 1, 0)
        form.addWidget(form_label("Açıklama / Not"), 2, 0)
        form.addWidget(self.note_edit, 3, 0)

        color_box = QVBoxLayout()
        color_box.setContentsMargins(0, 0, 0, 0)
        color_box.setSpacing(6)
        color_box.addWidget(form_label("Renk Seçimi"))
        colors = ["#EF4444", "#F59E0B", "#22C55E", "#3B82F6", "#8B5CF6", "#EC4899", "#14B8A6", "#94A3B8"]
        color_row = QHBoxLayout()
        color_row.setSpacing(8)
        for c in colors:
            b = QPushButton("")
            b.setObjectName("colorDotBtn")
            b.setCheckable(True)
            b.setFixedSize(28, 28)
            b.setProperty("tag_color", c)
            b.setStyleSheet(
                "QPushButton { border-radius:14px; border:2px solid #e2e8f0; background:%s; } "
                "QPushButton:checked { border:3px solid #0f172a; }" % c
            )
            b.clicked.connect(lambda _=False, color=c: self.select_color(color))
            self._color_buttons.append(b)
            color_row.addWidget(b)
        color_row.addStretch()
        color_box.addLayout(color_row)
        self.active_check = QCheckBox("Etiket Aktif")
        self.active_check.setChecked(True)
        color_box.addWidget(self.active_check)
        form.addLayout(color_box, 0, 1, 4, 1)

        rl.addLayout(form)

        btn_row = QHBoxLayout()
        self.op_hint = QLabel("")
        self.op_hint.setObjectName("muted")
        btn_row.addWidget(self.op_hint, 1)
        btn_row.addStretch()
        self.save_btn = QPushButton("Kaydet")
        self.save_btn.clicked.connect(self.save_tag)
        self.cancel_btn = QPushButton("İptal")
        self.cancel_btn.setObjectName("secondary")
        self.cancel_btn.clicked.connect(self.reject)
        self.del_btn = QPushButton("Etiketi Sil")
        self.del_btn.setObjectName("danger")
        self.del_btn.clicked.connect(self.delete_tag)
        btn_row.addWidget(self.save_btn)
        btn_row.addWidget(self.cancel_btn)
        btn_row.addWidget(self.del_btn)
        rl.addLayout(btn_row)

        rl.addWidget(section_label("Bu Etikete Bağlı Sözleşmeler"))
        self.contracts_table = QTableWidget(0, 6)
        configure_table(self.contracts_table, compact=True)
        self.contracts_table.setHorizontalHeaderLabels(["Platform", "Sözleşme No", "Kullanıcı", "Tür", "Durum", "Atama Tarihi"])
        self.contracts_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        rl.addWidget(self.contracts_table, 1)

    def _tag_key(self, name: str) -> str:
        return self.store._normalize_label(name)

    def _rebuild_contract_map(self):
        out: Dict[Tuple[str, str, str], dict] = {}
        for it in self.contract_index:
            key = (
                str(it.get("platform", "") or "").strip(),
                str(it.get("no", "") or "").strip(),
                str(it.get("type", "") or "").strip(),
            )
            out[key] = it
        self._contract_map = out

    def select_color(self, color: str):
        self.selected_color = str(color or "#3B82F6")
        for b in self._color_buttons:
            is_this = str(b.property("tag_color") or "").upper() == self.selected_color.upper()
            b.blockSignals(True)
            b.setChecked(is_this)
            b.blockSignals(False)

    def reload_data(self, keep_selection: bool = True):
        prev = self.selected_tag_key if keep_selection else None
        self.tags, self.assignments_by_key = self.store.load_tag_snapshot()
        self.usage = {}
        for k, vals in self.assignments_by_key.items():
            if vals:
                self.usage[k] = len(vals)
        self.refresh_tag_list()
        if prev:
            for i in range(self.tag_list.count()):
                it = self.tag_list.item(i)
                if str(it.data(Qt.UserRole) or "") == prev:
                    self.tag_list.setCurrentRow(i)
                    return
        if self.tag_list.count():
            self.tag_list.setCurrentRow(0)
        else:
            self.clear_detail_form()

    def _build_tag_row(self, color: str, name: str, count: int, active: bool) -> QFrame:
        row = QFrame()
        row.setObjectName("tagListRow")
        lay = QHBoxLayout(row)
        lay.setContentsMargins(10, 8, 10, 8)
        lay.setSpacing(8)

        dot = QLabel("●")
        dot.setObjectName("tagDot")
        dot.setStyleSheet(f"color:{color};")
        lay.addWidget(dot, 0)

        txt_col = QVBoxLayout()
        txt_col.setContentsMargins(0, 0, 0, 0)
        txt_col.setSpacing(1)
        name_lbl = QLabel(name)
        name_lbl.setObjectName("tagName")
        count_lbl = QLabel(f"{count} sözleşme")
        count_lbl.setObjectName("tagCount")
        txt_col.addWidget(name_lbl)
        txt_col.addWidget(count_lbl)
        lay.addLayout(txt_col, 1)

        st_lbl = QLabel("Aktif" if active else "Pasif")
        st_lbl.setObjectName("tagStateOn" if active else "tagStateOff")
        lay.addWidget(st_lbl, 0, Qt.AlignVCenter)
        return row

    def _apply_tag_row_state(self, row: Optional[QFrame], selected: bool):
        if row is None:
            return
        if selected:
            row.setStyleSheet(
                "QFrame#tagListRow { background:#dcecff; border-left:4px solid #1f5be3; border-radius:8px; }"
            )
        else:
            row.setStyleSheet(
                "QFrame#tagListRow { background:transparent; border-left:4px solid transparent; border-radius:8px; }"
            )

    def _refresh_tag_row_visuals(self):
        current = self.tag_list.currentRow()
        for i in range(self.tag_list.count()):
            item = self.tag_list.item(i)
            roww = self.tag_list.itemWidget(item)
            self._apply_tag_row_state(roww, i == current)

    def _all_ui_tags(self) -> List[Tuple[str, TagDef, bool]]:
        out: List[Tuple[str, TagDef, bool]] = []
        for dk in self._draft_order:
            dt = self._draft_tags.get(dk)
            if dt is not None:
                out.append((dk, dt, True))
        for t in self.tags:
            out.append((self._tag_key(t.name), t, False))
        return out

    def _is_draft_key(self, key: str) -> bool:
        return str(key or "").startswith("__draft__:")

    def _make_unique_draft_name(self, base: str = "Yeni Etiket") -> str:
        used = {self._tag_key(t.name) for t in self.tags}
        used.update(self._tag_key(v.name) for v in self._draft_tags.values())
        name = base
        if self._tag_key(name) not in used:
            return name
        i = 2
        while True:
            candidate = f"{base} {i}"
            if self._tag_key(candidate) not in used:
                return candidate
            i += 1

    def refresh_tag_list(self):
        q = self._tag_key(self.search.text())
        current = self.selected_tag_key
        self.tag_list.clear()
        for key, t, _is_draft in self._all_ui_tags():
            if q and q not in self._tag_key(t.name):
                continue
            cnt = int(self.usage.get(key, 0))
            item = QListWidgetItem("")
            item.setData(Qt.UserRole, key)
            item.setData(Qt.UserRole + 1, t.name)
            item.setSizeHint(QSize(0, 60))
            self.tag_list.addItem(item)
            roww = self._build_tag_row(t.color, t.name, cnt, bool(t.active))
            self._apply_tag_row_state(roww, key == current)
            self.tag_list.setItemWidget(item, roww)
        if current:
            for i in range(self.tag_list.count()):
                it = self.tag_list.item(i)
                if str(it.data(Qt.UserRole) or "") == current:
                    self.tag_list.setCurrentRow(i)
                    break
        if self.tag_list.count() and self.tag_list.currentRow() < 0:
            self.tag_list.setCurrentRow(0)
        if not self.tag_list.count():
            self.clear_detail_form()

    def clear_detail_form(self):
        self.selected_tag_key = None
        self.name_edit.clear()
        self.note_edit.clear()
        self.active_check.setChecked(True)
        self.select_color("#3B82F6")
        self.contract_count.setText("0 bağlı sözleşme")
        self.contracts_table.setRowCount(0)

    def on_tag_selected(self, row: int):
        if row < 0:
            self.clear_detail_form()
            return
        item = self.tag_list.item(row)
        if not item:
            self.clear_detail_form()
            return
        name = str(item.data(Qt.UserRole + 1) or "").strip()
        key = str(item.data(Qt.UserRole) or "")
        if self._is_draft_key(key):
            tag = self._draft_tags.get(key)
        else:
            tag = next((t for t in self.tags if self._tag_key(t.name) == key), None)
        if not tag:
            self.clear_detail_form()
            return
        self.selected_tag_key = key
        self.name_edit.setText(tag.name)
        self.note_edit.setPlainText(tag.note)
        self.active_check.setChecked(bool(tag.active))
        self.select_color(tag.color)
        self.refresh_assignments(tag.name if not self._is_draft_key(key) else "")
        self._refresh_tag_row_visuals()

    def refresh_assignments(self, tag_name: str):
        key = self._tag_key(tag_name)
        assigns = list(self.assignments_by_key.get(key, []))
        self.contract_count.setText(f"{len(assigns)} bağlı sözleşme")
        self.contracts_table.setUpdatesEnabled(False)
        self.contracts_table.setRowCount(len(assigns))
        for r, a in enumerate(assigns):
            key = (str(a.get("platform", "")), str(a.get("no", "")), str(a.get("type", "")))
            it = self._contract_map.get(key, {})
            vals = [
                key[0],
                key[1],
                str(it.get("user", "") or "-"),
                key[2],
                str(it.get("status", "") or "-"),
                str(a.get("assigned_at", "") or "-"),
            ]
            for c, v in enumerate(vals):
                cell = QTableWidgetItem(str(v))
                cell.setFlags(cell.flags() & ~Qt.ItemIsEditable)
                self.contracts_table.setItem(r, c, cell)
        self.contracts_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.contracts_table.setUpdatesEnabled(True)

    def new_tag(self):
        key = f"__draft__:{self._draft_seq}"
        self._draft_seq += 1
        draft = TagDef(name=self._make_unique_draft_name(), color="#3B82F6", note="", active=True)
        self._draft_tags[key] = draft
        self._draft_order.insert(0, key)
        self.selected_tag_key = key
        self.refresh_tag_list()
        for i in range(self.tag_list.count()):
            it = self.tag_list.item(i)
            if str(it.data(Qt.UserRole) or "") == key:
                self.tag_list.setCurrentRow(i)
                break
        self.name_edit.selectAll()
        self.name_edit.setFocus()

    def save_tag(self):
        name = self.name_edit.text().strip()
        if not name:
            QMessageBox.warning(self, "Eksik", "Etiket adı boş olamaz.")
            return
        key = self._tag_key(name)
        same = next((t for t in self.tags if self._tag_key(t.name) == key), None)
        is_draft = self._is_draft_key(self.selected_tag_key or "")
        if self.selected_tag_key and (not is_draft) and same and self._tag_key(same.name) != self.selected_tag_key:
            QMessageBox.warning(self, "Çakışma", "Aynı isimde başka etiket var.")
            return
        if (not self.selected_tag_key or is_draft) and same:
            QMessageBox.warning(self, "Çakışma", "Bu etiket zaten var.")
            return

        old_name = ""
        if self.selected_tag_key and not is_draft:
            old = next((t for t in self.tags if self._tag_key(t.name) == self.selected_tag_key), None)
            old_name = str(old.name if old else "")

        tag = TagDef(
            name=name,
            color=self.selected_color,
            note=self.note_edit.toPlainText().strip(),
            active=bool(self.active_check.isChecked()),
        )
        # processEvents öncesi tüm işlem butonlarını kapat — reentrancy önleme
        self.save_btn.setEnabled(False)
        self.del_btn.setEnabled(False)
        self.cancel_btn.setEnabled(False)
        try:
            self.op_hint.setText("Etiket kaydediliyor...")
            QApplication.processEvents()
            self.store.upsert_tag_def(tag)
            if old_name and self._tag_key(old_name) != self._tag_key(name):
                self.store.rename_tag_assignments(old_name, name, tag.color)
                self.store.delete_tag_def(old_name)
            if is_draft and self.selected_tag_key:
                self._draft_tags.pop(self.selected_tag_key, None)
                self._draft_order = [k for k in self._draft_order if k != self.selected_tag_key]
            self.changed = True
            self.reload_data(keep_selection=False)
            self.op_hint.setText("")
            for i in range(self.tag_list.count()):
                it = self.tag_list.item(i)
                if str(it.data(Qt.UserRole) or "") == self._tag_key(name):
                    self.tag_list.setCurrentRow(i)
                    break
        finally:
            self.save_btn.setEnabled(True)
            self.del_btn.setEnabled(True)
            self.cancel_btn.setEnabled(True)

    def delete_tag(self):
        if not self.selected_tag_key:
            QMessageBox.warning(self, "Seçim", "Silmek için bir etiket seçin.")
            return
        if self._is_draft_key(self.selected_tag_key):
            self._draft_tags.pop(self.selected_tag_key, None)
            self._draft_order = [k for k in self._draft_order if k != self.selected_tag_key]
            self.refresh_tag_list()
            if self.tag_list.count():
                self.tag_list.setCurrentRow(0)
            else:
                self.clear_detail_form()
            return
        tag = next((t for t in self.tags if self._tag_key(t.name) == self.selected_tag_key), None)
        if not tag:
            return
        if not ask_yes_no(
            self,
            "Etiketi Sil",
            f"'{tag.name}' etiketi silinecek.\nBu etikete ait tüm atamalar da kaldırılır.\n\nDevam edilsin mi?",
        ):
            return
        # processEvents öncesi tüm işlem butonlarını kapat — reentrancy önleme
        self.save_btn.setEnabled(False)
        self.del_btn.setEnabled(False)
        self.cancel_btn.setEnabled(False)
        try:
            self.op_hint.setText("Etiket siliniyor...")
            QApplication.processEvents()
            self.store.delete_tag_def(tag.name)
            self.changed = True
            self.reload_data(keep_selection=False)
            self.op_hint.setText("")
        finally:
            self.save_btn.setEnabled(True)
            self.del_btn.setEnabled(True)
            self.cancel_btn.setEnabled(True)


class SystemDialog(StyledDialog):
    def __init__(
        self,
        store: SystemTypeStore,
        platform: str,
        default_name: str = "Sistem 1",
        parent=None,
        existing_system: Optional[SystemInfo] = None,
        edit_mode: bool = False,
        pre_selected: Optional[List[str]] = None,
        default_t0_date: str = "",
        events_provider: Optional[Callable[[], List[dict]]] = None,
    ):
        super().__init__("Sistemi Düzenle" if edit_mode else "Sistem Ekle", parent)
        self.store = store
        self.platform = platform
        self.default_name = default_name
        self.existing_system = existing_system
        self.edit_mode = edit_mode
        self.default_t0_date = str(default_t0_date or "")
        self.external_events_provider = events_provider
        # pre_selected: edit modunda hangi bilesenlerin secili gosterilecegi
        self.pre_selected: Optional[set] = set(pre_selected) if pre_selected is not None else None
        initial_keys = pre_selected if pre_selected is not None else (getattr(existing_system, "components", {}) or {}).keys()
        self.initial_component_keys = set(initial_keys or [])
        self.result: Optional[SystemInfo] = None
        try:
            self.system_types = list(self.store.list_system_type_names(self.platform))
        except Exception:
            self.system_types = []
        self.resize(560, 720)
        self.inputs: Dict[str, QCheckBox] = {}
        self.build()

    def build(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(22, 10, 22, 16)
        root.setSpacing(10)

        root.addWidget(form_label("Sistem Adı"))
        self.name = QLineEdit()
        self.name.setPlaceholderText("Örn: Sistem 1")
        self.name.setText(self.default_name)
        self.name.selectAll()
        root.addWidget(self.name)

        # Sistem seviyesinde tarih widget oluşturulmaz; tarih bilgisi teslimatlarda tutulur.

        # ── Sistem Tipi / Bileşen Paketi: sadece hızlı seçim sağlar ──
        type_card = QFrame()
        type_card.setObjectName("systemFormCard")
        type_lay = QGridLayout(type_card)
        type_lay.setContentsMargins(10, 8, 10, 8)
        type_lay.setHorizontalSpacing(8)
        type_lay.setVerticalSpacing(6)
        type_lay.setColumnStretch(0, 1)

        type_lay.addWidget(form_label("Sistem Tipi"), 0, 0)
        self.system_type_combo = QComboBox()
        self.system_type_combo.setMinimumHeight(34)
        self.system_type_combo.addItem("Tip seçiniz...")
        for t in self.system_types:
            self.system_type_combo.addItem(t)
        type_lay.addWidget(self.system_type_combo, 1, 0)

        self.apply_type_btn = QPushButton("Tipi Uygula")
        self.apply_type_btn.setObjectName("secondary")
        self.apply_type_btn.setMinimumHeight(34)
        self.apply_type_btn.clicked.connect(self.apply_selected_system_type)
        type_lay.addWidget(self.apply_type_btn, 1, 1)
        root.addWidget(type_card)

        comp_head = QHBoxLayout()
        comp_head.addWidget(form_label("Bileşenler"), 0)
        self.selected_count_lbl = QLabel("0 seçili")
        self.selected_count_lbl.setObjectName("selectionPill")
        comp_head.addWidget(self.selected_count_lbl, 0)
        comp_head.addStretch(1)
        self.select_all_btn = QPushButton("Tümünü Seç")
        self.select_all_btn.setObjectName("secondary")
        self.select_all_btn.setMinimumHeight(32)
        self.clear_all_btn = QPushButton("Hiçbiri")
        self.clear_all_btn.setObjectName("secondary")
        self.clear_all_btn.setMinimumHeight(32)
        comp_head.addWidget(self.select_all_btn)
        comp_head.addWidget(self.clear_all_btn)
        root.addLayout(comp_head)

        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Bileşen ara...")
        root.addWidget(self.search_input)

        assigned = self.store.assigned_components(self.platform)
        existing_keys = list((self.existing_system.components if self.existing_system else {}).keys())
        extras = [c for c in existing_keys if c not in assigned]
        comps = assigned + extras
        if not comps:
            warn = QLabel("Bu platforma atanmış bileşen yok. Önce Bileşen Yönetimi ekranından platforma bileşen atayın.")
            warn.setObjectName("warning")
            warn.setWordWrap(True)
            root.addWidget(warn)

        self.comp_table = QTableWidget(len(comps), 2)
        self.comp_table.setObjectName("systemCompTable")
        configure_table(self.comp_table, compact=True)
        self.comp_table.setHorizontalHeaderLabels(["", "Bileşen"])
        self.comp_table.setSelectionMode(QTableWidget.NoSelection)
        self.comp_table.verticalHeader().setVisible(False)
        self.comp_table.horizontalHeader().setVisible(False)
        self.comp_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Fixed)
        self.comp_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.comp_table.setColumnWidth(0, 38)
        self.comp_table.setMinimumHeight(230)

        for r, comp in enumerate(comps):
            cell_wrap = QWidget()
            cell_layout = QHBoxLayout(cell_wrap)
            cell_layout.setContentsMargins(0, 0, 0, 0)
            cell_layout.setAlignment(Qt.AlignCenter)
            cb = QCheckBox()
            if self.edit_mode and self.existing_system:
                if self.pre_selected is not None:
                    is_checked = comp in self.pre_selected
                else:
                    # Sadece qty > 0 olan bilesenleri secili goster
                    is_checked = comp in existing_keys and self.existing_system.components.get(comp, 0) > 0
            else:
                is_checked = False
            cb.setChecked(is_checked)
            cb.stateChanged.connect(lambda _state, row=r: self._sync_component_row_style(row))
            cb.stateChanged.connect(lambda _state: self.update_selected_count())
            cell_layout.addWidget(cb)
            self.comp_table.setCellWidget(r, 0, cell_wrap)

            name_item = QTableWidgetItem(comp)
            name_item.setFlags(name_item.flags() & ~Qt.ItemIsEditable)
            self.comp_table.setItem(r, 1, name_item)
            self.comp_table.setRowHeight(r, 31)
            self.inputs[comp] = cb
            self._sync_component_row_style(r)

        root.addWidget(self.comp_table, 1)
        self.update_selected_count()
        self.search_input.textChanged.connect(self.filter_components)
        self.comp_table.cellClicked.connect(self.on_component_cell_clicked)
        self.select_all_btn.clicked.connect(self.select_all_components)
        self.clear_all_btn.clicked.connect(self.clear_all_components)

        row = QHBoxLayout()
        self.save_type_btn = QPushButton("Seçimi Tip Olarak Kaydet")
        self.save_type_btn.setObjectName("secondary")
        self.save_type_btn.clicked.connect(self.save_selection_as_system_type)
        row.addWidget(self.save_type_btn)
        row.addStretch()
        save = QPushButton("Güncelle" if self.edit_mode else "Sistemi Ekle")
        save.clicked.connect(self.save)
        row.addWidget(save)
        root.addLayout(row)

    def update_selected_count(self):
        if hasattr(self, "selected_count_lbl"):
            self.selected_count_lbl.setText(f"{len(self.selected_components())} seçili")

    def _sync_component_row_style(self, row: int):
        cb = self._row_checkbox(row) if hasattr(self, "comp_table") else None
        checked = bool(cb and cb.isChecked())

        bg = QColor("#dcecff") if checked else QColor("#ffffff")
        fg = QColor("#1f5be3") if checked else QColor("#0f172a")

        cell_wrap = self.comp_table.cellWidget(row, 0)
        if cell_wrap:
            cell_wrap.setStyleSheet(f"background:{bg.name()};")

        for c in range(self.comp_table.columnCount()):
            item = self.comp_table.item(row, c)
            if item:
                item.setBackground(bg)
                item.setForeground(fg)

        name_item = self.comp_table.item(row, 1)
        if name_item:
            # Eski tik varsa temizle, ismi saf haliyle tut.
            raw = str(
                name_item.data(Qt.UserRole)
                or name_item.text().replace("✓", "").strip()
            )

            name_item.setData(Qt.UserRole, raw)
            name_item.setText(raw)
            name_item.setForeground(fg)

    def selected_components(self) -> List[str]:
        return [comp for comp, cb in self.inputs.items() if cb.isChecked()]

    def apply_selected_system_type(self):
        type_name = self.system_type_combo.currentText().strip()
        if not type_name or type_name == "Tip seçiniz...":
            QMessageBox.information(self, "Sistem Tipi", "Uygulamak için bir sistem tipi seçin.")
            return
        try:
            comps = self.store.get_system_type_components(type_name, self.platform)
        except Exception as exc:
            QMessageBox.warning(self, "Sistem Tipi", f"Sistem tipi okunamadı:\n{exc}")
            return
        if not comps:
            QMessageBox.information(self, "Sistem Tipi", "Bu sistem tipinde kayıtlı bileşen bulunamadı.")
            return
        comp_set = set(comps)
        for comp, cb in self.inputs.items():
            cb.setChecked(comp in comp_set)

    def save_selection_as_system_type(self):
        comps = self.selected_components()
        if not comps:
            QMessageBox.warning(self, "Eksik", "Tip olarak kaydetmek için en az bir bileşen seçin.")
            return
        default_name = ""
        current = self.system_type_combo.currentText().strip()
        if current and current != "Tip seçiniz...":
            default_name = current
        type_name, ok = QInputDialog.getText(
            self,
            "Sistem Tipi Kaydet",
            "Tip adı:",
            QLineEdit.Normal,
            default_name,
        )
        if not ok:
            return
        type_name = type_name.strip()
        if not type_name:
            QMessageBox.warning(self, "Eksik", "Tip adı girin.")
            return
        try:
            saved_count = self.store.save_system_type(type_name, self.platform, comps)
        except Exception as exc:
            QMessageBox.warning(self, "Sistem Tipi", f"Tip kaydedilemedi:\n{exc}")
            return

        # Kaydedilen tip sözleşme kaydı beklemeden hemen dropdown'a gelsin.
        current = self.system_type_combo.currentText().strip()
        self.system_type_combo.blockSignals(True)
        self.system_type_combo.clear()
        self.system_type_combo.addItem("Tip seçiniz...")
        try:
            for t in self.store.list_system_type_names(self.platform):
                self.system_type_combo.addItem(t)
        except Exception:
            if self.system_type_combo.findText(type_name) < 0:
                self.system_type_combo.addItem(type_name)
        self.system_type_combo.blockSignals(False)
        idx = self.system_type_combo.findText(type_name)
        if idx >= 0:
            self.system_type_combo.setCurrentIndex(idx)
        elif current:
            self.system_type_combo.setCurrentText(current)

        QMessageBox.information(self, "Sistem Tipi", f"'{type_name}' sistem tipi kaydedildi. ({saved_count} bileşen)")

    def _row_checkbox(self, row: int) -> Optional[QCheckBox]:
        cell = self.comp_table.cellWidget(row, 0)
        if not cell:
            return None
        cb = cell.findChild(QCheckBox)
        return cb

    def on_component_cell_clicked(self, row: int, col: int):
        if col not in (0, 1):
            return
        cb = self._row_checkbox(row)
        if cb:
            cb.setChecked(not cb.isChecked())

    def select_all_components(self):
        for r in range(self.comp_table.rowCount()):
            cb = self._row_checkbox(r)
            if cb:
                cb.setChecked(True)

    def clear_all_components(self):
        for r in range(self.comp_table.rowCount()):
            cb = self._row_checkbox(r)
            if cb:
                cb.setChecked(False)

    def filter_components(self, text: str):
        def norm(s: str) -> str:
            return str(s or "").strip().lower().replace("ı", "i").replace("İ", "i")
        q = norm(text)
        for r in range(self.comp_table.rowCount()):
            item = self.comp_table.item(r, 1)
            name = norm(item.text() if item else "")
            self.comp_table.setRowHidden(r, bool(q and q not in name))

    def _recalc_completion(self):
        return

    def date_picker_events(self) -> List[dict]:
        if callable(self.external_events_provider):
            try:
                return list(self.external_events_provider() or [])
            except Exception:
                return []
        return []

    def save(self):
        name = self.name.text().strip()
        if not name:
            QMessageBox.warning(self, "Eksik", "Sistem adı girin.")
            return
        t0_text = ""
        old = self.existing_system.components if (self.edit_mode and self.existing_system) else {}
        selected = set(self.selected_components())
        removed = []
        if self.edit_mode and self.existing_system:
            removed = sorted(self.initial_component_keys - selected, key=lambda x: str(x).lower())
            if removed:
                shown = "\n".join(f"• {name}" for name in removed[:12])
                if len(removed) > 12:
                    shown += f"\n• ... ve {len(removed) - 12} bileşen daha"
                if not ask_yes_no(
                    self,
                    "Bileşenler Silinecek",
                    "Aşağıdaki bileşenlerin onay kutusunu kaldırdınız. Güncelleme sonrası bu bileşenler "
                    "sistemden ve bu sisteme ait teslimatlardan silinecek; Excel'deki ilgili değer hücreleri boşaltılacak.\n\n"
                    f"{shown}\n\nOnaylıyor musunuz?",
                ):
                    return
        comps = {comp: old.get(comp, 0.0) for comp in self.inputs.keys() if comp in selected}
        if not comps:
            QMessageBox.warning(self, "Eksik", "En az bir bileşen seçin.")
            return
        self.result = SystemInfo(
            name=name,
            components=comps,
            t0_date="",
            t0_months=0,
            completion_date="",
            status=getattr(self.existing_system, "status", "Başlanmadı") or "Başlanmadı",
            acceptance_date=getattr(self.existing_system, "acceptance_date", "") or "",
        )
        self.result.removed_components = set(removed)
        self.accept()


class MultiSystemDialog(StyledDialog):
    def __init__(
        self,
        store: SystemTypeStore,
        platform: str,
        contract_t0_date: str = "",
        existing_names: Optional[List[str]] = None,
        parent=None,
        events_provider: Optional[Callable[[], List[dict]]] = None,
    ):
        super().__init__("Çoklu Sistem Ekle", parent)
        self.store = store
        self.platform = str(platform or "")
        self.contract_t0_date = str(contract_t0_date or "")
        self.existing_names = set(existing_names or [])
        self.external_events_provider = events_provider
        self.result: List[SystemInfo] = []
        self.drafts: List[dict] = []
        self.current_index = 0
        self._loading = False
        self.component_rows: Dict[str, QCheckBox] = {}
        self.components = list(self.store.assigned_components(self.platform))
        try:
            self.system_types = list(self.store.list_system_type_names(self.platform))
        except Exception:
            self.system_types = []
        self.resize(1120, 760)
        self.build()
        self.add_blank_system(select=True)

    def build(self):
        self.setStyleSheet(self.styleSheet() + """
        QFrame#multiShell { background:#eef3f8; border:1px solid #cbd8e8; border-radius:14px; }
        QFrame#multiLeft, QFrame#multiMiddle { background:#ffffff; border:1px solid #d8e4f0; border-radius:14px; }
        QLabel#multiTitle { background:transparent; color:#102033; font-weight:950; font-size:14px; }
        QLabel#miniPill { background:#dbeafe; color:#1f5be3; border-radius:10px; padding:3px 8px; font-size:11px; font-weight:950; }
        QLabel#miniPillGreen { background:#dcfce7; color:#16a34a; border-radius:10px; padding:3px 8px; font-size:11px; font-weight:950; }
        QLabel#miniPillOrange { background:#fff7ed; color:#ea580c; border-radius:10px; padding:3px 8px; font-size:11px; font-weight:950; }
        QLabel#typeHint { background:transparent; color:#64748b; font-size:11px; font-weight:750; }
        QFrame#dateStrip { background:#f8fbff; border:1px solid #d8e4f0; border-radius:12px; }
        QFrame#saveTypeStrip { background:#f8fbff; border:1px dashed #93c5fd; border-radius:10px; }
        QListWidget#multiSystemList { background:transparent; border:0; outline:0; }
        QListWidget#multiSystemList::item { border:0; padding:0; margin:0; }
        QTableWidget#multiComponentTable { background:#ffffff; border:1px solid #d8e4f0; border-radius:10px; gridline-color:#edf2f7; }
        QTableWidget#multiComponentTable::item { border-bottom:1px solid #edf2f7; padding:6px 8px; }
        QLineEdit#qtyCellInput { padding:3px; border-radius:7px; font-weight:950; qproperty-alignment: AlignCenter; }
        """)
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        body = QHBoxLayout()
        body.setContentsMargins(16, 12, 16, 12)
        body.setSpacing(14)
        root.addLayout(body, 1)

        left = QFrame()
        left.setObjectName("multiLeft")
        left.setFixedWidth(300)
        left_lay = QVBoxLayout(left)
        left_lay.setContentsMargins(12, 12, 12, 12)
        left_lay.setSpacing(10)

        left_head = QHBoxLayout()
        systems_title = QLabel("Sistemler")
        systems_title.setObjectName("multiTitle")
        left_head.addWidget(systems_title, 1)
        add_btn = QPushButton("+ Sistem")
        add_btn.clicked.connect(lambda: self.add_blank_system(select=True))
        left_head.addWidget(add_btn, 0)
        left_lay.addLayout(left_head)

        self.system_list = QListWidget()
        self.system_list.setObjectName("multiSystemList")
        self.system_list.currentRowChanged.connect(self.select_draft)
        left_lay.addWidget(self.system_list, 1)

        duplicate_btn = QPushButton("Bu sistemi çoğalt")
        duplicate_btn.setObjectName("secondary")
        duplicate_btn.clicked.connect(self.duplicate_current)
        left_lay.addWidget(duplicate_btn)
        delete_btn = QPushButton("Seçili sistemi sil")
        delete_btn.setObjectName("danger")
        delete_btn.clicked.connect(self.delete_current)
        left_lay.addWidget(delete_btn)
        body.addWidget(left, 0)

        middle = QFrame()
        middle.setObjectName("multiMiddle")
        mid = QVBoxLayout(middle)
        mid.setContentsMargins(12, 12, 12, 12)
        mid.setSpacing(10)
        body.addWidget(middle, 1)

        mid.addWidget(form_label("Sistem Adı"))
        self.name_edit = QLineEdit()
        self.name_edit.textChanged.connect(self.on_name_changed)
        mid.addWidget(self.name_edit)

        # Çoklu sistemde tarih widget oluşturulmaz; tarih bilgisi teslimatlarda tutulur.

        type_row = QGridLayout()
        type_row.setHorizontalSpacing(10)
        type_row.setVerticalSpacing(6)
        type_row.addWidget(form_label("Sistem Tipi / opsiyonel"), 0, 0)
        self.type_combo = QComboBox()
        self.type_combo.addItem("Tip seçiniz...")
        for t in self.system_types:
            self.type_combo.addItem(t)
        type_row.addWidget(self.type_combo, 1, 0)
        apply_btn = QPushButton("Tipi Uygula")
        apply_btn.setObjectName("secondary")
        apply_btn.clicked.connect(self.apply_selected_type)
        type_row.addWidget(apply_btn, 1, 1)
        hint = QLabel("Tip seçmek zorunlu değil. Tip seçilmezse bileşenleri aşağıdan manuel seçip adet girebilirsin.")
        hint.setObjectName("typeHint")
        type_row.addWidget(hint, 2, 0, 1, 2)
        type_row.setColumnStretch(0, 1)
        mid.addLayout(type_row)

        comp_head = QHBoxLayout()
        comp_title = QLabel("Bileşenler")
        comp_title.setObjectName("multiTitle")
        comp_head.addWidget(comp_title, 0)
        self.selected_count_lbl = QLabel("0 seçili")
        self.selected_count_lbl.setObjectName("miniPill")
        comp_head.addWidget(self.selected_count_lbl, 0)
        comp_head.addStretch()
        select_all = QPushButton("Tümünü Seç")
        select_all.setObjectName("secondary")
        select_all.clicked.connect(self.select_all_components)
        clear_all = QPushButton("Hiçbiri")
        clear_all.setObjectName("secondary")
        clear_all.clicked.connect(self.clear_components)
        comp_head.addWidget(select_all)
        comp_head.addWidget(clear_all)
        mid.addLayout(comp_head)

        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Bileşen ara...")
        self.search_input.textChanged.connect(self.filter_components)
        mid.addWidget(self.search_input)

        self.comp_table = QTableWidget(0, 3)
        self.comp_table.setObjectName("multiComponentTable")
        configure_table(self.comp_table, compact=True)
        self.comp_table.setHorizontalHeaderLabels(["", "Bileşen", "Adet"])
        self.comp_table.verticalHeader().setVisible(False)
        self.comp_table.setSelectionMode(QTableWidget.NoSelection)
        self.comp_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Fixed)
        self.comp_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.comp_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.Fixed)
        self.comp_table.setColumnWidth(0, 34)
        self.comp_table.setColumnWidth(2, 96)
        self.comp_table.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.comp_table.itemChanged.connect(self.on_component_table_item_changed)
        self.comp_table.setMinimumHeight(330)
        mid.addWidget(self.comp_table, 1)

        save_type = QFrame()
        save_type.setObjectName("saveTypeStrip")
        st_lay = QHBoxLayout(save_type)
        st_lay.setContentsMargins(10, 8, 10, 8)
        st_lay.setSpacing(10)
        st_hint = QLabel("Bu sistemde seçtiğin bileşen/adet kombinasyonunu daha sonra tekrar kullanmak için tip olarak kaydet.")
        st_hint.setObjectName("typeHint")
        st_hint.setWordWrap(True)
        st_lay.addWidget(st_hint, 1)
        save_type_btn = QPushButton("Bunu Tip Olarak Kaydet")
        save_type_btn.setObjectName("secondary")
        save_type_btn.clicked.connect(self.save_current_as_type)
        st_lay.addWidget(save_type_btn, 0)
        mid.addWidget(save_type)

        footer = QHBoxLayout()
        footer.setContentsMargins(16, 10, 16, 12)
        self.system_count_badge = QLabel("")
        self.system_count_badge.setObjectName("miniPill")
        self.total_qty_badge = QLabel("")
        self.total_qty_badge.setObjectName("miniPill")
        self.warning_badge = QLabel("")
        self.warning_badge.setObjectName("miniPillOrange")
        footer.addWidget(self.system_count_badge, 0)
        footer.addWidget(self.total_qty_badge, 0)
        footer.addWidget(self.warning_badge, 0)
        footer.addStretch()
        cancel = QPushButton("İptal")
        cancel.setObjectName("secondary")
        cancel.clicked.connect(self.reject)
        self.submit_btn = QPushButton("Sistemleri Ekle")
        self.submit_btn.clicked.connect(self.accept_drafts)
        footer.addWidget(cancel)
        footer.addWidget(self.submit_btn)
        root.addLayout(footer)

    def date_picker_events(self) -> List[dict]:
        if callable(self.external_events_provider):
            try:
                return list(self.external_events_provider() or [])
            except Exception:
                return []
        return []

    def _draft_components(self, draft: dict) -> Dict[str, int]:
        return {str(k): int(v) for k, v in (draft.get("components") or {}).items() if int(v or 0) > 0}

    def make_unique_system_name(self) -> str:
        used = {normalize_sheet_name(n) for n in self.existing_names}
        used.update(normalize_sheet_name(str(d.get("name", ""))) for d in self.drafts)
        i = 1
        while True:
            name = f"Sistem {i}"
            if normalize_sheet_name(name) not in used:
                return name
            i += 1

    def make_blank_draft(self) -> dict:
        return {
            "name": self.make_unique_system_name(),
            "t0_date": "",
            "t0_months": 0,
            "completion_date": "",
            "system_type": "",
            "components": {},
        }

    def add_blank_system(self, select: bool = True):
        self.drafts.append(self.make_blank_draft())
        self.refresh_system_list(keep_row=len(self.drafts) - 1 if select else self.current_index)

    def duplicate_current(self):
        if not self.drafts:
            return
        src = copy.deepcopy(self.drafts[self.current_index])
        src["name"] = self.make_unique_system_name()
        self.drafts.append(src)
        self.refresh_system_list(keep_row=len(self.drafts) - 1)

    def delete_current(self):
        if len(self.drafts) <= 1:
            QMessageBox.information(self, "Sistem silinemez", "En az 1 sistem kalmalı.")
            return
        self.drafts.pop(self.current_index)
        self.refresh_system_list(keep_row=min(self.current_index, len(self.drafts) - 1))

    def refresh_system_list(self, keep_row: Optional[int] = None, reload_form: bool = True):
        target = self.current_index if keep_row is None else int(keep_row)
        self.system_list.blockSignals(True)
        self.system_list.clear()
        for idx, draft in enumerate(self.drafts):
            item = QListWidgetItem()
            item.setSizeHint(QSize(0, 72))
            self.system_list.addItem(item)
            self.system_list.setItemWidget(item, self.build_system_card(draft, idx == target))
        if self.drafts:
            target = max(0, min(target, len(self.drafts) - 1))
            self.current_index = target
            self.system_list.setCurrentRow(target)
        self.system_list.blockSignals(False)
        if self.drafts and reload_form:
            self.load_current_draft()
        self.update_footer()

    def build_system_card(self, draft: dict, selected: bool) -> QFrame:
        card = QFrame()
        card.setStyleSheet(
            "QFrame { background:%s; border:1px solid %s; border-radius:13px; } QLabel { background:transparent; color:%s; }"
            % ("#0b2f6b" if selected else "#ffffff", "#061f49" if selected else "#cbdff4", "#ffffff" if selected else "#0f172a")
        )
        lay = QVBoxLayout(card)
        lay.setContentsMargins(10, 8, 10, 8)
        lay.setSpacing(6)
        top = QHBoxLayout()
        name = QLabel(str(draft.get("name") or "Sistem"))
        name.setStyleSheet("font-weight:950; font-size:13px;")
        top.addWidget(name, 1)
        count = QLabel(f"{len(self._draft_components(draft))} bileşen")
        count.setObjectName("miniPillGreen")
        top.addWidget(count, 0)
        lay.addLayout(top)
        meta = QHBoxLayout()
        meta.setSpacing(6)
        typ = str(draft.get("system_type") or "").strip() or "Özel seçim"
        typ_lbl = QLabel(typ)
        typ_lbl.setObjectName("miniPill")
        meta.addWidget(typ_lbl, 0)
        meta.addStretch()
        lay.addLayout(meta)
        return card

    def select_draft(self, row: int):
        if self._loading or row < 0 or row >= len(self.drafts):
            return
        self.current_index = row
        self.refresh_system_list(keep_row=row)

    def current_draft(self) -> dict:
        return self.drafts[self.current_index]

    def load_current_draft(self):
        if not self.drafts:
            return
        draft = self.current_draft()
        self._loading = True
        try:
            self.name_edit.setText(str(draft.get("name") or ""))
            typ = str(draft.get("system_type") or "")
            idx = self.type_combo.findText(typ) if typ else 0
            self.type_combo.setCurrentIndex(idx if idx >= 0 else 0)
            self.refresh_component_table()
        finally:
            self._loading = False
        self.update_selected_count()

    def recalc_current_completion(self):
        draft = self.current_draft()
        draft["t0_date"] = ""
        draft["t0_months"] = 0
        draft["completion_date"] = ""

    def on_name_changed(self, text: str):
        if self._loading:
            return
        self.current_draft()["name"] = text.strip()
        self.refresh_system_list(keep_row=self.current_index, reload_form=False)

    def on_t0_changed(self, text: str):
        if self._loading:
            return
        self.current_draft()["t0_date"] = text.strip()
        self.recalc_current_completion()
        self.refresh_system_list(keep_row=self.current_index, reload_form=False)

    def on_months_changed(self, value: int):
        if self._loading:
            return
        self.current_draft()["t0_months"] = int(value)
        self.recalc_current_completion()
        self.refresh_system_list(keep_row=self.current_index, reload_form=False)

    def set_custom_type(self):
        draft = self.current_draft()
        draft["system_type"] = ""
        self.type_combo.blockSignals(True)
        self.type_combo.setCurrentIndex(0)
        self.type_combo.blockSignals(False)

    def refresh_component_table(self):
        self.component_rows.clear()
        self.comp_table.blockSignals(True)
        self.comp_table.setRowCount(len(self.components))
        draft = self.current_draft()
        components = self._draft_components(draft)
        self.comp_table.setUpdatesEnabled(False)
        for r, comp in enumerate(self.components):
            qty = int(components.get(comp, 0))
            cb_wrap = QWidget()
            cb_lay = QHBoxLayout(cb_wrap)
            cb_lay.setContentsMargins(0, 0, 0, 0)
            cb_lay.setAlignment(Qt.AlignCenter)
            cb = QCheckBox()
            cb.setChecked(qty > 0)
            cb.toggled.connect(lambda checked, c=comp: self.on_component_checked(c, checked))
            cb_lay.addWidget(cb)
            self.comp_table.setCellWidget(r, 0, cb_wrap)

            name_item = QTableWidgetItem(comp)
            name_item.setFlags(Qt.ItemIsEnabled)
            self.comp_table.setItem(r, 1, name_item)

            qty_item = QTableWidgetItem(str(qty))
            qty_item.setData(Qt.UserRole, comp)
            qty_item.setTextAlignment(Qt.AlignCenter)
            qty_item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable | Qt.ItemIsEditable)
            self.comp_table.setItem(r, 2, qty_item)
            self.component_rows[comp] = cb
            self.comp_table.setRowHeight(r, 36)
            self.apply_component_row_style(r, qty > 0)
        self.comp_table.setUpdatesEnabled(True)
        self.comp_table.blockSignals(False)
        self.filter_components(self.search_input.text())

    def apply_component_row_style(self, row: int, selected: bool):
        bg = QColor("#f8fbff") if selected else QColor("#ffffff")
        fg = QColor("#1f5be3") if selected else QColor("#0f172a")
        for col in range(self.comp_table.columnCount()):
            item = self.comp_table.item(row, col)
            if item:
                item.setBackground(bg)
                item.setForeground(fg)
        for col in (0, 2):
            widget = self.comp_table.cellWidget(row, col)
            if widget:
                widget.setStyleSheet(f"background:{bg.name()};")

    def component_row_index(self, comp: str) -> int:
        try:
            return self.components.index(comp)
        except ValueError:
            return -1

    def update_component_qty(self, comp: str, qty: int, manual: bool = True):
        qty = max(0, int(qty or 0))
        draft = self.current_draft()
        if qty > 0:
            draft.setdefault("components", {})[comp] = qty
        else:
            draft.setdefault("components", {}).pop(comp, None)
        cb = self.component_rows.get(comp)
        if cb:
            cb.blockSignals(True)
            cb.setChecked(qty > 0)
            cb.blockSignals(False)
        row = self.component_row_index(comp)
        if row >= 0:
            item = self.comp_table.item(row, 2)
            if item:
                self.comp_table.blockSignals(True)
                item.setText(str(qty))
                self.comp_table.blockSignals(False)
        row = self.component_row_index(comp)
        if row >= 0:
            self.apply_component_row_style(row, qty > 0)
        if manual and not self._loading:
            self.set_custom_type()
        self.update_selected_count()
        self.refresh_system_list(keep_row=self.current_index, reload_form=False)

    def on_component_checked(self, comp: str, checked: bool):
        if self._loading:
            return
        current = int(self.current_draft().setdefault("components", {}).get(comp, 0) or 0)
        self.update_component_qty(comp, 1 if checked and current <= 0 else (current if checked else 0))

    def on_component_table_item_changed(self, item: QTableWidgetItem):
        if self._loading or not item or item.column() != 2:
            return
        comp_item = self.comp_table.item(item.row(), 1)
        comp = str((comp_item.text() if comp_item else item.data(Qt.UserRole)) or "").strip()
        if not comp:
            return
        text = str(item.text() or "").strip()
        qty = int(text) if text.isdigit() else 0
        if text != str(qty):
            self.comp_table.blockSignals(True)
            item.setText(str(qty))
            self.comp_table.blockSignals(False)
        self.update_component_qty(comp, qty)

    def on_qty_changed(self, comp: str, text: str):
        if self._loading:
            return
        qty = int(text) if str(text or "").isdigit() else 0
        self.update_component_qty(comp, qty)

    def normalize_qty_input(self, comp: str):
        row = self.component_row_index(comp)
        item = self.comp_table.item(row, 2) if row >= 0 else None
        text = item.text().strip() if item else ""
        qty = int(text) if text.isdigit() else 0
        self.update_component_qty(comp, qty)

    def update_selected_count(self):
        selected = len(self._draft_components(self.current_draft())) if self.drafts else 0
        self.selected_count_lbl.setText(f"{selected} seçili")
        self.update_footer()

    def filter_components(self, text: str):
        q = normalize_sheet_name(text)
        for r, comp in enumerate(self.components):
            self.comp_table.setRowHidden(r, bool(q and q not in normalize_sheet_name(comp)))

    def select_all_components(self):
        for comp in self.components:
            if int(self.current_draft().setdefault("components", {}).get(comp, 0) or 0) <= 0:
                self.update_component_qty(comp, 1, manual=False)
        self.set_custom_type()

    def clear_components(self):
        for comp in list(self.components):
            self.update_component_qty(comp, 0, manual=False)
        self.set_custom_type()

    def apply_selected_type(self):
        type_name = self.type_combo.currentText().strip()
        if not type_name or type_name == "Tip seçiniz...":
            QMessageBox.information(self, "Sistem Tipi", "Uygulamak için bir sistem tipi seçin.")
            return
        try:
            qty_map = self.store.get_system_type_component_quantities(type_name, self.platform)
        except Exception:
            qty_map = {}
        if not qty_map:
            try:
                qty_map = {comp: 1 for comp in self.store.get_system_type_components(type_name, self.platform)}
            except Exception as exc:
                QMessageBox.warning(self, "Sistem Tipi", f"Sistem tipi okunamadı:\n{exc}")
                return
        if not qty_map:
            QMessageBox.information(self, "Sistem Tipi", "Bu sistem tipinde kayıtlı bileşen bulunamadı.")
            return
        self.current_draft()["system_type"] = type_name
        self.current_draft()["components"] = {c: int(max(as_number(v), 0)) for c, v in qty_map.items() if c in self.components and as_number(v) > 0}
        self.refresh_component_table()
        self.update_selected_count()
        self.refresh_system_list(keep_row=self.current_index, reload_form=False)

    def save_current_as_type(self):
        components = self._draft_components(self.current_draft())
        if not components:
            QMessageBox.warning(self, "Eksik", "Tip olarak kaydetmek için en az bir bileşen adedi girin.")
            return
        type_name, ok = QInputDialog.getText(self, "Sistem Tipi Kaydet", "Tip adı:", QLineEdit.Normal, "")
        if not ok:
            return
        type_name = type_name.strip()
        if not type_name:
            QMessageBox.warning(self, "Eksik", "Tip adı boş olamaz.")
            return
        existing = {normalize_sheet_name(n) for n in self.store.list_system_type_names(self.platform)}
        if normalize_sheet_name(type_name) in existing:
            QMessageBox.warning(self, "Çakışma", "Aynı isimde bir sistem tipi zaten var.")
            return
        try:
            saved_count = self.store.save_system_type(type_name, self.platform, components)
        except Exception as exc:
            QMessageBox.warning(self, "Sistem Tipi", f"Tip kaydedilemedi:\n{exc}")
            return
        self.system_types = list(self.store.list_system_type_names(self.platform))
        self.type_combo.blockSignals(True)
        self.type_combo.clear()
        self.type_combo.addItem("Tip seçiniz...")
        for t in self.system_types:
            self.type_combo.addItem(t)
        idx = self.type_combo.findText(type_name)
        self.type_combo.setCurrentIndex(idx if idx >= 0 else 0)
        self.type_combo.blockSignals(False)
        self.current_draft()["system_type"] = type_name
        self.refresh_system_list(keep_row=self.current_index, reload_form=False)
        QMessageBox.information(self, "Sistem Tipi", f"'{type_name}' sistem tipi kaydedildi. ({saved_count} bileşen)")

    def warning_count(self) -> int:
        count = 0
        names = []
        existing_norm = {normalize_sheet_name(n) for n in self.existing_names}
        for draft in self.drafts:
            name_norm = normalize_sheet_name(str(draft.get("name", "")))
            if not name_norm or name_norm in existing_norm or name_norm in names:
                count += 1
            names.append(name_norm)
            if not self._draft_components(draft):
                count += 1
        return count

    def update_footer(self):
        total_qty = sum(sum(self._draft_components(d).values()) for d in self.drafts)
        self.system_count_badge.setText(f"{len(self.drafts)} sistem hazır")
        self.total_qty_badge.setText(f"{total_qty} toplam bileşen adedi")
        warnings = self.warning_count()
        self.warning_badge.setText(f"{warnings} uyarı")
        self.warning_badge.setVisible(warnings > 0)
        self.submit_btn.setText(f"{len(self.drafts)} Sistemi Ekle")

    def accept_drafts(self):
        existing_norm = {normalize_sheet_name(n) for n in self.existing_names}
        seen = set()
        out: List[SystemInfo] = []
        for draft in self.drafts:
            name = str(draft.get("name", "") or "").strip()
            norm = normalize_sheet_name(name)
            if not name:
                QMessageBox.warning(self, "Eksik", "Sistem adı boş olamaz.")
                return
            if norm in existing_norm or norm in seen:
                QMessageBox.warning(self, "Çakışma", f"'{name}' sistem adı zaten kullanılıyor.")
                return
            seen.add(norm)
            t0 = ""
            comps = self._draft_components(draft)
            if not comps:
                QMessageBox.warning(self, "Bileşen yok", f"{name}: bileşen adedi toplamı 0 olamaz.")
                return
            months = 0
            completion = ""
            out.append(SystemInfo(
                name=name,
                components={k: float(v) for k, v in comps.items()},
                t0_date="",
                t0_months=0,
                completion_date="",
                status="Başlanmadı",
                acceptance_date="",
            ))
        self.result = out
        self.accept()



# ─────────────────────────────────────────────────────────────────────────────
# UNIT TRACKING — Kuyruk no / seri no takibi için sol panel yardımcı sınıfları
# ─────────────────────────────────────────────────────────────────────────────

class UnitTrackingSlotCard(QFrame):
    """Sol panelde tek bir kuyruk no / seri no slotunu gösteren kompakt kart."""

    changed = Signal()

    def __init__(self, slot_no: int, comp_name: str, label: str = "Kuyruk No / Seri No", identifier: str = "", parent=None):
        super().__init__(parent)
        self._slot_no = int(slot_no or 0)
        self._comp_name = str(comp_name or "")
        self._label = str(label or "Kuyruk No / Seri No")
        self._is_duplicate = False
        self.setObjectName("unitSlotCard")
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self._build(identifier)
        self._apply_state_style()

    def _build(self, identifier: str):
        lay = QVBoxLayout(self)
        lay.setContentsMargins(10, 8, 10, 8)
        lay.setSpacing(6)

        top = QHBoxLayout()
        top.setSpacing(8)
        self._slot_lbl = QLabel(f"#{self._slot_no:03d}")
        self._slot_lbl.setObjectName("unitSlotNo")
        self._slot_lbl.setFixedWidth(44)
        self._slot_lbl.setAlignment(Qt.AlignCenter)
        top.addWidget(self._slot_lbl, 0)

        self._comp_lbl = QLabel(self._comp_name.upper())
        self._comp_lbl.setObjectName("unitSlotComp")
        top.addWidget(self._comp_lbl, 1)

        self._status_lbl = QLabel("Eksik")
        self._status_lbl.setObjectName("unitSlotBadge")
        self._status_lbl.setAlignment(Qt.AlignCenter)
        top.addWidget(self._status_lbl, 0)
        lay.addLayout(top)

        self._edit = QLineEdit()
        self._edit.setObjectName("unitSlotInput")
        self._edit.setPlaceholderText(self._label)
        self._edit.setText(str(identifier or ""))
        self._edit.textChanged.connect(self._on_changed)
        lay.addWidget(self._edit)

    def _on_changed(self, _text: str = ""):
        self._apply_state_style()
        self.changed.emit()

    def _status(self) -> str:
        if self.identifier() and self._is_duplicate:
            return "duplicate"
        if self.identifier():
            return "defined"
        return "missing"

    def _apply_state_style(self):
        status = self._status()
        if status == "duplicate":
            text = "Tekrar Var"; border = "#FCA5A5"; bg = "#FFF1F2"; badge_bg = "#FEE2E2"; badge_fg = "#B91C1C"
        elif status == "defined":
            text = "Tanımlandı"; border = "#86EFAC"; bg = "#F0FDF4"; badge_bg = "#DCFCE7"; badge_fg = "#15803D"
        else:
            text = "Eksik"; border = "#D8E2EE"; bg = "#FFFFFF"; badge_bg = "#F1F5F9"; badge_fg = "#64748B"
        self._status_lbl.setText(text)
        self.setStyleSheet(f"""
            QFrame#unitSlotCard {{ background:{bg}; border:1px solid {border}; border-radius:10px; }}
            QLabel#unitSlotNo {{ background:#EFF6FF; border:1px solid #BFDBFE; border-radius:7px; color:#0F3B82; font-weight:900; font-size:12px; padding:5px 0; }}
            QLabel#unitSlotComp {{ color:#64748B; font-size:10px; font-weight:900; background:transparent; }}
            QLabel#unitSlotBadge {{ background:{badge_bg}; color:{badge_fg}; border:1px solid {border}; border-radius:8px; font-size:10px; font-weight:900; padding:2px 8px; }}
            QLineEdit#unitSlotInput {{ background:#FFFFFF; border:1px solid #CBD5E1; border-radius:7px; padding:5px 8px; color:#0F172A; font-weight:700; }}
            QLineEdit#unitSlotInput:focus {{ border-color:#2563EB; }}
        """)

    def slot_no(self) -> int:
        return self._slot_no

    def identifier(self) -> str:
        return self._edit.text().strip()

    def set_identifier(self, value: str):
        self._edit.blockSignals(True)
        self._edit.setText(str(value or ""))
        self._edit.blockSignals(False)
        self._apply_state_style()

    def set_duplicate(self, duplicate: bool):
        if self._is_duplicate == bool(duplicate):
            return
        self._is_duplicate = bool(duplicate)
        self._apply_state_style()

    def matches(self, query: str, filter_key: str) -> bool:
        ident = self.identifier()
        status = self._status()
        if filter_key == "missing" and status != "missing":
            return False
        if filter_key == "defined" and status != "defined":
            return False
        if filter_key == "duplicate" and status != "duplicate":
            return False
        q = normalize_sheet_name(query)
        if not q:
            return True
        slot_text = f"{self._slot_no} {self._slot_no:03d} #{self._slot_no:03d}"
        return q in normalize_sheet_name(slot_text) or q in normalize_sheet_name(ident)

    def to_dict(self) -> dict:
        return {"slot_no": self._slot_no, "identifier": self.identifier(), "is_delivered": 0, "note": ""}


class UnitTrackingSidePanel(QFrame):
    """DeliveryDialog sol panelinde kuyruk no / seri no listesini yönetir."""

    changed = Signal()
    backRequested = Signal()
    clearRequested = Signal()

    def __init__(self, parent=None):
        super().__init__(parent)
        self._comp_name = ""
        self._label = "Kuyruk No / Seri No"
        self._filter = "all"
        self._cards: List[UnitTrackingSlotCard] = []
        self._build()

    def _build(self):
        outer = QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.setSpacing(8)

        top = QHBoxLayout()
        top.setSpacing(8)
        self._back_btn = QPushButton("← Bileşen Atama Durumu")
        self._back_btn.setObjectName("secondary")
        self._back_btn.setFlat(True)
        self._back_btn.setStyleSheet("text-align:left; color:#1D4ED8; font-weight:900; border:0; background:transparent;")
        self._back_btn.clicked.connect(self.backRequested.emit)
        top.addWidget(self._back_btn, 1)
        self._clear_btn = QPushButton("Temizle")
        self._clear_btn.setObjectName("danger")
        self._clear_btn.setMinimumHeight(32)
        self._clear_btn.clicked.connect(self._on_clear)
        top.addWidget(self._clear_btn, 0)
        outer.addLayout(top)

        self._title = QLabel("")
        self._title.setStyleSheet("font-weight:900; font-size:14px; color:#0F172A; background:transparent;")
        outer.addWidget(self._title)

        stats = QHBoxLayout(); stats.setSpacing(8)
        self._defined_card = self._stat_card("0/0", "Tanımlı", "#ECFDF5", "#86EFAC", "#047857")
        self._missing_card = self._stat_card("0", "Eksik", "#FFF7ED", "#FDBA74", "#C2410C")
        stats.addWidget(self._defined_card[0], 1)
        stats.addWidget(self._missing_card[0], 1)
        outer.addLayout(stats)

        self._progress = QProgressBar()
        self._progress.setRange(0, 100)
        self._progress.setTextVisible(False)
        self._progress.setFixedHeight(8)
        self._progress.setStyleSheet("QProgressBar{border:0;background:#E2E8F0;border-radius:4px;} QProgressBar::chunk{background:#1D9A8A;border-radius:4px;}")
        outer.addWidget(self._progress)

        self._search = QLineEdit()
        self._search.setPlaceholderText("Kuyruk No / Seri No veya slot ara...")
        self._search.textChanged.connect(self._apply_visibility)
        outer.addWidget(self._search)

        chips = QHBoxLayout(); chips.setSpacing(5)
        self._filter_buttons: Dict[str, QPushButton] = {}
        for key, label in (("all", "Tümü"), ("missing", "Eksik"), ("defined", "Tanımlı"), ("duplicate", "Tekrar")):
            btn = QPushButton(label)
            btn.setCheckable(True)
            btn.setMinimumHeight(28)
            btn.clicked.connect(lambda _=False, k=key: self._set_filter(k))
            self._filter_buttons[key] = btn
            chips.addWidget(btn)
        chips.addStretch()
        outer.addLayout(chips)

        self._scroll = QScrollArea()
        self._scroll.setFrameShape(QFrame.NoFrame)
        self._scroll.setWidgetResizable(True)
        self._scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self._scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self._scroll.setStyleSheet(self._modern_scrollbar_qss("QScrollArea"))
        self._host = QWidget()
        self._list = QVBoxLayout(self._host)
        self._list.setContentsMargins(0, 4, 4, 4)
        self._list.setSpacing(8)
        self._list.addStretch()
        self._scroll.setWidget(self._host)
        outer.addWidget(self._scroll, 1)
        self._set_filter("all")


    @staticmethod
    def _modern_scrollbar_qss(root_selector: str = "QScrollArea") -> str:
        return f"""
{root_selector}{{background:transparent;border:0;}}
{root_selector} > QWidget > QWidget{{background:transparent;}}
{root_selector} QScrollBar:vertical{{
    width:11px;
    background:#F1F5F9;
    margin:4px 2px 4px 2px;
    border-radius:5px;
}}
{root_selector} QScrollBar::handle:vertical{{
    background:#B8C7D9;
    border-radius:5px;
    min-height:28px;
}}
{root_selector} QScrollBar::handle:vertical:hover{{background:#8FA8C6;}}
{root_selector} QScrollBar::add-line:vertical,
{root_selector} QScrollBar::sub-line:vertical{{height:0px; background:transparent; border:0;}}
{root_selector} QScrollBar::add-page:vertical,
{root_selector} QScrollBar::sub-page:vertical{{background:transparent;}}
{root_selector} QScrollBar:horizontal{{
    height:11px;
    background:#F1F5F9;
    margin:2px 4px 2px 4px;
    border-radius:5px;
}}
{root_selector} QScrollBar::handle:horizontal{{
    background:#B8C7D9;
    border-radius:5px;
    min-width:28px;
}}
{root_selector} QScrollBar::handle:horizontal:hover{{background:#8FA8C6;}}
{root_selector} QScrollBar::add-line:horizontal,
{root_selector} QScrollBar::sub-line:horizontal{{width:0px; background:transparent; border:0;}}
{root_selector} QScrollBar::add-page:horizontal,
{root_selector} QScrollBar::sub-page:horizontal{{background:transparent;}}
"""

    def _stat_card(self, value: str, caption: str, bg: str, border: str, fg: str):
        frame = QFrame(); frame.setObjectName("unitStatCard")
        frame.setStyleSheet(f"QFrame#unitStatCard{{background:{bg}; border:1px solid {border}; border-radius:9px;}}")
        lay = QVBoxLayout(frame); lay.setContentsMargins(8, 7, 8, 7); lay.setSpacing(1)
        val = QLabel(value); val.setAlignment(Qt.AlignCenter); val.setStyleSheet(f"font-weight:900; color:{fg}; font-size:13px; background:transparent;")
        cap = QLabel(caption); cap.setAlignment(Qt.AlignCenter); cap.setStyleSheet("font-weight:800; color:#64748B; font-size:10px; background:transparent;")
        lay.addWidget(val); lay.addWidget(cap)
        return frame, val, cap

    def _on_clear(self):
        for card in self._cards:
            card.set_identifier("")
        self._refresh_all()
        self.clearRequested.emit()
        self.changed.emit()

    def _set_filter(self, key: str):
        self._filter = key
        for k, btn in self._filter_buttons.items():
            btn.setChecked(k == key)
        self._style_filter_buttons()
        self._apply_visibility()

    def _style_filter_buttons(self):
        for key, btn in self._filter_buttons.items():
            if btn.isChecked():
                btn.setStyleSheet("QPushButton{background:#0F3B82;color:white;border:1px solid #0F3B82;border-radius:12px;padding:3px 9px;font-weight:900;font-size:11px;}")
            else:
                btn.setStyleSheet("QPushButton{background:white;color:#334155;border:1px solid #CBD5E1;border-radius:12px;padding:3px 9px;font-weight:800;font-size:11px;}")

    def set_component(self, comp_name: str, label: str, units: list):
        self._comp_name = str(comp_name or "")
        self._label = str(label or "Kuyruk No / Seri No")
        self._title.setText(f"{self._comp_name} {self._label} Listesi")
        self._search.setPlaceholderText(f"{self._label} veya slot ara...")
        self._rebuild_cards(units or [])

    def _rebuild_cards(self, units: list):
        while self._cards:
            card = self._cards.pop()
            self._list.removeWidget(card)
            card.setParent(None)
            card.deleteLater()
        for unit in sorted(units or [], key=lambda u: int(u.get("slot_no", 0) or 0)):
            card = UnitTrackingSlotCard(
                int(unit.get("slot_no", 0) or 0),
                self._comp_name,
                self._label,
                str(unit.get("identifier") or ""),
            )
            card.changed.connect(self._on_card_changed)
            self._list.insertWidget(self._list.count() - 1, card)
            self._cards.append(card)
        self._refresh_all()

    def _on_card_changed(self):
        self._refresh_all()
        self.changed.emit()

    def _refresh_all(self):
        self._check_duplicates()
        self._update_stats()
        self._apply_visibility()

    def _check_duplicates(self):
        counts: Dict[str, int] = {}
        for card in self._cards:
            ident = card.identifier()
            if ident:
                counts[normalize_sheet_name(ident)] = counts.get(normalize_sheet_name(ident), 0) + 1
        for card in self._cards:
            ident = card.identifier()
            card.set_duplicate(bool(ident and counts.get(normalize_sheet_name(ident), 0) > 1))

    def _counts(self) -> Tuple[int, int, int, int]:
        total = len(self._cards)
        defined = sum(1 for c in self._cards if c.identifier())
        duplicate = sum(1 for c in self._cards if c.identifier() and c._status() == "duplicate")
        missing = total - defined
        return total, defined, missing, duplicate

    def _update_stats(self):
        total, defined, missing, duplicate = self._counts()
        self._defined_card[1].setText(f"{defined}/{total}")
        self._missing_card[1].setText(str(missing))
        self._progress.setValue(int((defined / total) * 100) if total else 0)
        labels = {"all": f"Tümü {total}", "missing": f"Eksik {missing}", "defined": f"Tanımlı {defined}", "duplicate": f"Tekrar {duplicate}"}
        for key, text in labels.items():
            self._filter_buttons[key].setText(text)

    def _apply_visibility(self, *_args):
        query = self._search.text().strip()
        for card in self._cards:
            card.setVisible(card.matches(query, self._filter))

    def get_units(self) -> list:
        return [card.to_dict() for card in self._cards]

    def has_duplicates(self) -> bool:
        return any(card.identifier() and card._status() == "duplicate" for card in self._cards)

    def all_defined(self) -> bool:
        return all(card.identifier() for card in self._cards)


class DeliveryDialog(StyledDialog):
    """Teslimat ekleme / düzenleme dialog'u.

    Tüm bileşenlerde kuyruk no / seri no takibi kullanılabilir:
    - Bileşen hücresinde ▶/◀ ok ikonu görünür.
    - Satır/ok seçilince sol panel kuyruk no / seri no listesine dönüşür.
    - Ana tablo 4 sütun kalır; inline detail row oluşturulmaz.
    """

    def __init__(
        self,
        system: SystemInfo,
        default_name: str = "Teslimat 1",
        parent=None,
        component_keys: Optional[List[str]] = None,
        planned_assigned: Optional[Dict[str, float]] = None,
        contract_t0_date: str = "",
        events_provider: Optional[Callable[[], List[dict]]] = None,
        allow_delete: bool = False,
        existing_delivery: Optional["DeliveryInfo"] = None,
    ):
        super().__init__("Teslimatı Düzenle" if existing_delivery else "Teslimat Ekle", parent)
        self.system = system
        self.store = getattr(parent, "store", None)
        self.default_name = default_name
        raw_components = getattr(self.system, "components", {}) or {}
        try:
            component_names = list(raw_components.keys()) if hasattr(raw_components, "keys") else list(dict(raw_components).keys())
        except RecursionError:
            try:
                sys.__stderr__.write("RecursionError while reading system components; using an empty component list.\n")
            except Exception:
                pass
            component_names = []
        except Exception:
            component_names = []
        self.component_keys = list(component_keys or component_names)
        self.planned_assigned = dict(planned_assigned or {})
        self.contract_t0_date = contract_t0_date
        self.events_provider = events_provider
        self.allow_delete = bool(allow_delete)
        self.delete_requested = False
        self.result: Optional[DeliveryInfo] = None
        self._existing_delivery = existing_delivery
        self.resize(1280, 700)
        self.inputs: Dict[str, Tuple[QTableWidgetItem, QTableWidgetItem, QTableWidgetItem]] = {}
        self._updating_qty = False
        self._updating_qty_table = False
        self._status_auto_filling = False
        # Unit tracking state
        self._unit_tracking_map: Dict[str, str] = {}  # {comp_name: label}
        self.left_panel_mode = "assignment"
        self.active_unit_component: Optional[str] = None
        self.unit_filter = "all"
        self.unit_search_text = ""
        self._component_units_state: Dict[str, list] = {}
        # comp_name -> table row index
        self._comp_row: Dict[str, int] = {}
        # Actual qty_table row -> comp_name
        self._row_comp: Dict[int, Optional[str]] = {}
        self._load_unit_tracking_map()
        self.build()

    def _load_unit_tracking_map(self):
        self._unit_tracking_map = {comp: "Kuyruk No / Seri No" for comp in self.component_keys}
        if self.store is not None:
            try:
                stored_labels = self.store.get_unit_tracking_components()
            except Exception:
                stored_labels = {}
            for comp, label in (stored_labels or {}).items():
                if comp in self._unit_tracking_map and str(label or "").strip():
                    self._unit_tracking_map[comp] = str(label).strip()

    def _is_unit_tracking(self, comp: str) -> bool:
        return bool(comp)

    def _safe_system_components(self) -> Dict[str, float]:
        raw = getattr(self.system, "components", {}) or {}
        if isinstance(raw, dict):
            return raw
        try:
            return dict(raw)
        except RecursionError:
            try:
                sys.__stderr__.write("RecursionError while normalizing system components; using empty quantities.\n")
            except Exception:
                pass
            return {}
        except Exception:
            return {}

    def _system_component_qty(self, comp: str) -> float:
        try:
            components = self._safe_system_components()
            return max(as_number(components.get(comp, 0)), 0)
        except RecursionError:
            try:
                sys.__stderr__.write(f"RecursionError while reading component quantity for {comp}; using 0.\n")
            except Exception:
                pass
            return 0.0
        except Exception:
            return 0.0

    # ------------------------------------------------------------------ build
    def build(self):
        outer = QHBoxLayout(self)
        outer.setContentsMargins(18, 18, 18, 18)
        outer.setSpacing(14)

        left_card = QFrame()
        left_card.setObjectName("contentPanel")
        left_card.setFixedWidth(380)
        left_card.setStyleSheet(
            "QFrame#contentPanel{background:#F8FBFF; border:1px solid #D8E2EE; border-radius:12px;}"
        )
        left_lay = QVBoxLayout(left_card)
        left_lay.setContentsMargins(12, 12, 12, 12)
        left_lay.setSpacing(8)
        alloc_title = QLabel("Bileşen Atama Durumu")
        alloc_title.setAlignment(Qt.AlignCenter)
        alloc_title.setStyleSheet("font-weight:900; font-size:14px;")
        left_lay.addWidget(alloc_title)
        alloc_hint = QLabel("Tanımlanabilir değeri 0 olan bileşenler listeden gizlenir.")
        alloc_hint.setObjectName("muted")
        alloc_hint.setWordWrap(True)
        left_lay.addWidget(alloc_hint)
        self.assignment_table = QTableWidget(0, 3)
        self.assignment_table.setObjectName("qtyTable")
        configure_table(self.assignment_table, compact=True)
        self.assignment_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.assignment_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.assignment_table.setHorizontalHeaderLabels(["Bileşen", "Tanımlanmış", "Tanımlanabilir"])
        self.assignment_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.assignment_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Fixed)
        self.assignment_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.Fixed)
        self.assignment_table.setColumnWidth(1, 120)
        self.assignment_table.setColumnWidth(2, 132)
        self.assignment_table.setStyleSheet(UnitTrackingSidePanel._modern_scrollbar_qss("QTableWidget"))
        left_lay.addWidget(self.assignment_table, 1)

        self.assignment_panel = QWidget()
        assignment_lay = left_lay
        self.unit_side_panel = UnitTrackingSidePanel()
        self.unit_side_panel.changed.connect(self._on_unit_side_panel_changed)
        self.unit_side_panel.backRequested.connect(self._show_assignment_panel)
        self.unit_side_panel.clearRequested.connect(self._on_unit_side_panel_changed)

        # left_card layout is reused as the assignment page; wrap pages in a stack by moving widgets.
        self.left_stack = QStackedWidget()
        self.left_stack.setStyleSheet("background:transparent;")
        while left_lay.count():
            item = left_lay.takeAt(0)
            widget = item.widget()
            if widget is not None:
                widget.setParent(self.assignment_panel)
        assignment_page_lay = QVBoxLayout(self.assignment_panel)
        assignment_page_lay.setContentsMargins(0, 0, 0, 0)
        assignment_page_lay.setSpacing(8)
        assignment_page_lay.addWidget(alloc_title)
        assignment_page_lay.addWidget(alloc_hint)
        assignment_page_lay.addWidget(self.assignment_table, 1)
        left_lay.addWidget(self.left_stack, 1)
        self.left_stack.addWidget(self.assignment_panel)
        self.left_stack.addWidget(self.unit_side_panel)
        outer.addWidget(left_card, 0)

        right = QWidget()
        root = QVBoxLayout(right)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(12)
        outer.addWidget(right, 1)

        title = QLabel(f"{self.system.name} için Teslimat")
        title.setObjectName("dialogTitle")
        root.addWidget(title)

        grid = QGridLayout()
        grid.setHorizontalSpacing(14)
        grid.setVerticalSpacing(8)
        self.name = QLineEdit()
        self.name.setPlaceholderText("Örn: Teslimat 1")
        self.name.setText(self.default_name)
        self.name.selectAll()
        self.status = QComboBox()
        self.status.addItems(STATUS_VALUES)
        self.status.currentTextChanged.connect(self.on_status_changed)
        self._delivery_t0_date = str(getattr(self.system, "t0_date", "") or self.contract_t0_date or "")
        self._delivery_t0_months = int(getattr(self.system, "t0_months", 0) or 0)
        self._delivery_completion_date = str(getattr(self.system, "completion_date", "") or "")
        self.note = QLineEdit()
        self.note.setPlaceholderText("Not")
        self.delivery_user_combo = QComboBox()
        self.delivery_user_combo.addItem("Seçiniz...")
        if self.store is not None:
            for user in self.store.load_users(active_only=True):
                uname = str(user.get("name", "") or "").strip()
                if uname:
                    self.delivery_user_combo.addItem(uname)
        self.planned_acceptance_date, self.planned_acceptance_date_wrap = build_date_input(
            self, events_provider=self.events_provider
        )
        self.acceptance_date, self.acceptance_date_wrap = build_date_input(
            self, max_date=date.today(), events_provider=self.events_provider
        )
        grid.addWidget(form_label("Teslimat Adı"), 0, 0)
        grid.addWidget(self.name, 1, 0)
        grid.addWidget(form_label("Durum"), 0, 1)
        grid.addWidget(self.status, 1, 1)
        grid.addWidget(form_label("Planlanan Teslimat Tarihi"), 2, 0)
        grid.addWidget(self.planned_acceptance_date_wrap, 3, 0)
        grid.addWidget(form_label("Gerçek Teslimat Tarihi"), 4, 0)
        grid.addWidget(self.acceptance_date_wrap, 5, 0)
        grid.addWidget(form_label("Not"), 2, 1)
        grid.addWidget(self.note, 3, 1)
        grid.addWidget(form_label("Teslim Edilecek Kullanıcı"), 4, 1)
        grid.addWidget(self.delivery_user_combo, 5, 1)
        root.addLayout(grid)

        info_row = QHBoxLayout()
        info = QLabel("Bileşen miktarlarını aşağıdaki tabloda girin. Kalan değeri otomatik hesaplanır.")
        info.setObjectName("muted")
        info_row.addWidget(info, 1)
        self.fill_all_btn = QPushButton("Tüm Sistemi Ekle")
        self.fill_all_btn.setObjectName("secondary")
        self.fill_all_btn.setMinimumHeight(32)
        self.fill_all_btn.clicked.connect(self.fill_all_system_planned)
        info_row.addWidget(self.fill_all_btn, 0)
        self.fill_remaining_btn = QPushButton("Kalan Sistemi Ekle")
        self.fill_remaining_btn.setObjectName("secondary")
        self.fill_remaining_btn.setMinimumHeight(32)
        self.fill_remaining_btn.clicked.connect(self.fill_remaining_system_planned)
        info_row.addWidget(self.fill_remaining_btn, 0)
        root.addLayout(info_row)

        # Main quantity table (4 columns, fixed structure)
        self.qty_table = QTableWidget(0, 4)
        self.qty_table.setObjectName("qtyTable")
        configure_table(self.qty_table, compact=True)
        self.qty_table.setHorizontalHeaderLabels(["Bileşen", "Teslim Edilecek", "Teslim Edilen", "Kalan"])
        self.qty_table.verticalHeader().setVisible(False)
        self.qty_table.setAlternatingRowColors(False)
        self.qty_table.setShowGrid(True)
        self.qty_table.setVerticalScrollMode(QTableWidget.ScrollPerPixel)
        self.qty_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.qty_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.qty_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.Stretch)
        self.qty_table.horizontalHeader().setSectionResizeMode(3, QHeaderView.Fixed)
        self.qty_table.setColumnWidth(3, 110)
        self.qty_table.setMinimumHeight(200)
        self.qty_table.setStyleSheet(
            UnitTrackingSidePanel._modern_scrollbar_qss("QTableWidget")
            + """
            QTableWidget#qtyTable {
                background:#ffffff;
                alternate-background-color:#ffffff;
                gridline-color:#d8e2ed;
                selection-background-color:#dbeafe;
                selection-color:#0f172a;
            }
            QTableWidget#qtyTable::item {
                background:#ffffff;
                color:#0f172a;
                padding:4px 6px;
            }
            QTableWidget#qtyTable::item:hover {
                background:#f8fbff;
            }
            QTableWidget#qtyTable::item:selected {
                background:#dbeafe;
                color:#0f172a;
            }
            QTableWidget#qtyTable QWidget {
                background:#ffffff;
            }
            """
        )
        self.qty_table.viewport().setStyleSheet("background:#ffffff;")
        self.qty_table.setItemDelegateForColumn(1, CompactNumberDelegate(self.qty_table))
        self.qty_table.setItemDelegateForColumn(2, CompactNumberDelegate(self.qty_table))
        self.component_search = QLineEdit()
        self.component_search.setPlaceholderText("Bileşen ara...")
        self.component_search.textChanged.connect(self.filter_qty_components)

        self._populate_qty_table()

        root.addWidget(self.component_search, 0)
        root.addWidget(self.qty_table, 1)

        row = QHBoxLayout()
        if self.allow_delete:
            delete_btn = QPushButton("Teslimat Sil")
            delete_btn.setObjectName("danger")
            delete_btn.clicked.connect(self.request_delete)
            row.addWidget(delete_btn)
        row.addStretch()
        cancel = QPushButton("İptal")
        cancel.setObjectName("secondary")
        cancel.clicked.connect(self.reject)
        save = QPushButton("Kaydet")
        save.clicked.connect(self.save)
        row.addWidget(cancel)
        row.addWidget(save)
        root.addLayout(row)

    # ------------------------------------------------------------------ table population
    def _populate_qty_table(self):
        """Tüm bileşenler için satırları oluşturur."""
        existing = self._existing_delivery
        self._updating_qty = True
        was_blocked = self.qty_table.blockSignals(True)
        try:
            self.qty_table.setRowCount(0)
            self._comp_row.clear()
            self._row_comp.clear()
            self.inputs.clear()

            current_row = 0
            for comp in self.component_keys:
                self._comp_row[comp] = current_row

                # Component name cell with optional arrow for unit tracking
                if self._is_unit_tracking(comp):
                    comp_widget = self._make_arrow_cell(comp)
                    comp_item = QTableWidgetItem("")
                    comp_item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
                else:
                    comp_widget = None
                    comp_item = QTableWidgetItem(comp)
                    comp_item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)

                planned_val = "0"
                delivered_val = "0"
                if existing:
                    planned_val = fmt_num(float((existing.planned or {}).get(comp, 0) or 0))
                    delivered_val = fmt_num(float((existing.delivered or {}).get(comp, 0) or 0))

                planned = QTableWidgetItem(planned_val)
                delivered = QTableWidgetItem(delivered_val)
                remaining = QTableWidgetItem("0")

                planned.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable | Qt.ItemIsEditable)
                delivered.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable | Qt.ItemIsEditable)
                remaining.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
                for it in (planned, delivered, remaining):
                    it.setTextAlignment(Qt.AlignCenter)

                self.qty_table.insertRow(current_row)
                if comp_widget:
                    self.qty_table.setCellWidget(current_row, 0, comp_widget)
                else:
                    self.qty_table.setItem(current_row, 0, comp_item)
                self.qty_table.setItem(current_row, 1, planned)
                self.qty_table.setItem(current_row, 2, delivered)
                self.qty_table.setItem(current_row, 3, remaining)
                self.qty_table.setRowHeight(current_row, 30)
                self._row_comp[current_row] = comp
                self.inputs[comp] = (planned, delivered, remaining)
                self._update_remaining_row(current_row)
                if self._is_unit_tracking(comp):
                    self._ensure_component_units(comp, int(as_number(planned.text())))
                current_row += 1
        finally:
            self.qty_table.blockSignals(was_blocked)
            self._updating_qty = False

        self.qty_table.itemChanged.connect(self.on_qty_item_changed)
        self.qty_table.cellClicked.connect(self._on_cell_clicked)
        self.refresh_assignment_card()

    def _make_arrow_cell(self, comp: str) -> QWidget:
        """Bileşen adının önünde sol paneli açan ok ikonu olan widget döner."""
        widget = QWidget()
        widget.setStyleSheet("background: transparent;")
        lay = QHBoxLayout(widget)
        lay.setContentsMargins(6, 0, 6, 0)
        lay.setSpacing(0)
        arrow = QPushButton("▶")
        arrow.setObjectName("unitTrackingArrow")
        arrow.setFixedSize(20, 20)
        arrow.clicked.connect(lambda _=False, c=comp: self._toggle_unit_component(c))
        arrow.setProperty("comp", comp)
        lbl = QLabel(comp)
        lbl.setAlignment(Qt.AlignCenter)
        lbl.setStyleSheet("background: transparent; color:#0F172A; font-weight: 500;")
        spacer = QWidget()
        spacer.setFixedWidth(20)
        lay.addWidget(arrow, 0, Qt.AlignLeft | Qt.AlignVCenter)
        lay.addWidget(lbl, 1)
        lay.addWidget(spacer, 0)
        widget.setProperty("arrow_btn", arrow)
        widget.setProperty("label_widget", lbl)
        return widget

    def _get_arrow_btn(self, comp: str) -> Optional[QPushButton]:
        data_row = self._comp_row.get(comp)
        if data_row is None:
            return None
        w = self.qty_table.cellWidget(data_row, 0)
        if w:
            return w.property("arrow_btn")
        return None

    def _get_component_label(self, comp: str) -> Optional[QLabel]:
        data_row = self._comp_row.get(comp)
        if data_row is None:
            return None
        w = self.qty_table.cellWidget(data_row, 0)
        if w:
            return w.property("label_widget")
        return None

    def _existing_units_for(self, comp: str) -> list:
        if not self._existing_delivery:
            return []
        return list((getattr(self._existing_delivery, "component_units", None) or {}).get(comp, []) or [])

    def _ensure_component_units(self, comp: str, planned_qty: int) -> list:
        """Component slot state'ini planned adede göre normalize eder; mevcut değerleri korur."""
        planned_qty = max(0, int(planned_qty or 0))
        current = self._component_units_state.get(comp)
        source = current if current is not None else self._existing_units_for(comp)
        by_slot = {}
        for unit in source or []:
            try:
                slot_no = int(unit.get("slot_no", 0) or 0)
            except Exception:
                slot_no = 0
            if slot_no > 0:
                by_slot[slot_no] = {
                    "slot_no": slot_no,
                    "identifier": str(unit.get("identifier") or "").strip(),
                    "is_delivered": int(unit.get("is_delivered", 0) or 0),
                    "note": str(unit.get("note") or ""),
                }
        normalized = []
        keep_slots = set(range(1, planned_qty + 1))
        for slot_no, unit in by_slot.items():
            has_data = bool(
                str(unit.get("identifier") or "").strip()
                or str(unit.get("note") or "").strip()
                or int(unit.get("is_delivered", 0) or 0)
            )
            if slot_no > planned_qty and has_data:
                keep_slots.add(slot_no)
        for slot_no in sorted(keep_slots):
            unit = dict(by_slot.get(slot_no, {}))
            unit["slot_no"] = slot_no
            unit["identifier"] = str(unit.get("identifier") or "").strip()
            unit["is_delivered"] = int(unit.get("is_delivered", 0) or 0)
            unit["note"] = str(unit.get("note") or "")
            normalized.append(unit)
        self._component_units_state[comp] = normalized
        return normalized

    def _current_units_from_panel_if_active(self):
        comp = self.active_unit_component
        if comp and self.left_panel_mode == "unit_tracking" and hasattr(self, "unit_side_panel"):
            self._component_units_state[comp] = self.unit_side_panel.get_units()

    def _activate_unit_component(self, comp: str):
        if not self._is_unit_tracking(comp):
            return
        self._current_units_from_panel_if_active()
        planned_item, delivered_item, _ = self.inputs.get(comp, (None, None, None))
        planned_qty = int(max(as_number(planned_item.text()) if planned_item else 0, as_number(delivered_item.text()) if delivered_item else 0))
        units = self._ensure_component_units(comp, planned_qty)
        self.left_panel_mode = "unit_tracking"
        self.active_unit_component = comp
        self.unit_side_panel.set_component(comp, self._unit_tracking_map.get(comp, "Kuyruk No / Seri No"), units)
        self.left_stack.setCurrentWidget(self.unit_side_panel)
        self._refresh_unit_row_selection()

    def _toggle_unit_component(self, comp: str):
        if not self._is_unit_tracking(comp):
            return
        if self.left_panel_mode == "unit_tracking" and self.active_unit_component == comp:
            self._show_assignment_panel()
            return
        self._activate_unit_component(comp)

    def _show_assignment_panel(self):
        self._current_units_from_panel_if_active()
        self.left_panel_mode = "assignment"
        self.active_unit_component = None
        self.left_stack.setCurrentWidget(self.assignment_panel)
        self._refresh_unit_row_selection()
        self.refresh_assignment_card()

    def _on_unit_side_panel_changed(self):
        comp = self.active_unit_component
        if not comp:
            return
        self._component_units_state[comp] = self.unit_side_panel.get_units()

    def _validate_unit_tracking_qty(self, comp: str, qty: float) -> bool:
        """Unit tracking bileşende ondalıklı adet uyarısı."""
        if not self._is_unit_tracking(comp):
            return True
        if qty != int(qty):
            QMessageBox.warning(
                self, "Ondalıklı Adet",
                f"Bu bileşende teslim edilecek adet tam sayı olmalıdır.\n({comp})"
            )
            return False
        return True

    def _on_cell_clicked(self, row: int, col: int):
        comp = self._row_comp.get(row)
        if comp and col == 0 and self._is_unit_tracking(comp):
            self._toggle_unit_component(comp)

    def _update_panel_slot_count(self, comp: str, new_qty: int):
        if not self._is_unit_tracking(comp):
            return
        if comp == self.active_unit_component and self.left_panel_mode == "unit_tracking":
            self._component_units_state[comp] = self.unit_side_panel.get_units()
        units = self._ensure_component_units(comp, int(new_qty or 0))
        if comp == self.active_unit_component and self.left_panel_mode == "unit_tracking":
            self.unit_side_panel.set_component(comp, self._unit_tracking_map.get(comp, "Kuyruk No / Seri No"), units)
        self._refresh_unit_row_selection()

    def _refresh_unit_row_selection(self):
        selected_bg = QColor("#EAF3FF")
        normal_bg = QColor("#FFFFFF")
        selected_fg = QColor("#0F3B82")
        normal_fg = QColor("#0F172A")
        was_blocked = self.qty_table.blockSignals(True)
        try:
            for comp, row in self._comp_row.items():
                active = comp == self.active_unit_component and self.left_panel_mode == "unit_tracking"
                cell_widget = self.qty_table.cellWidget(row, 0)
                if cell_widget:
                    cell_widget.setStyleSheet(f"background:{'#EAF3FF' if active else 'transparent'};")
                label = self._get_component_label(comp)
                if label:
                    label.setStyleSheet(
                        "background: transparent; "
                        f"color:{'#0F3B82' if active else '#0F172A'}; "
                        f"font-weight:{'900' if active else '500'};"
                    )
                btn = self._get_arrow_btn(comp)
                if btn:
                    btn.setText("◀" if active else "▶")
                    if active:
                        btn.setStyleSheet("QPushButton#unitTrackingArrow{background:#0F3B82;color:white;border:1px solid #0F3B82;border-radius:5px;font-size:10px;font-weight:900;padding:0;}")
                    else:
                        btn.setStyleSheet("QPushButton#unitTrackingArrow{background:#DBEAFE;color:#1D4ED8;border:1px solid #93C5FD;border-radius:5px;font-size:10px;font-weight:900;padding:0;} QPushButton#unitTrackingArrow:hover{background:#BFDBFE;}")
                for c in range(self.qty_table.columnCount()):
                    item = self.qty_table.item(row, c)
                    if item:
                        item.setBackground(selected_bg if active else normal_bg)
                        item.setForeground(selected_fg if active else normal_fg)
        finally:
            self.qty_table.blockSignals(was_blocked)

    # ------------------------------------------------------------------ existing methods
    def request_delete(self):
        confirm = QMessageBox(self)
        confirm.setIcon(QMessageBox.Warning)
        confirm.setWindowTitle("Teslimat Sil")
        confirm.setText(
            "Bu teslimat silinecek. Bu teslimata ait teslim miktarları artık teslim edilmiş sayılmayacak. "
            "Sistem ana bileşen adetleri değişmeyecek. Devam etmek istiyor musunuz?"
        )
        delete_btn = confirm.addButton("Evet, Sil", QMessageBox.DestructiveRole)
        confirm.addButton("Vazgeç", QMessageBox.RejectRole)
        confirm.exec()
        if confirm.clickedButton() != delete_btn:
            return
        self.delete_requested = True
        self.accept()

    def _is_delivered_status(self) -> bool:
        status = self.status.currentText().strip() if hasattr(self, "status") else ""
        norm = status.lower().replace("ı", "i").replace("İ", "i")
        return norm in {"teslim edildi", "tamamlandi", "tamamlandı"}

    def on_status_changed(self, _text: str = ""):
        if self._status_auto_filling or not self._is_delivered_status():
            return
        self.fill_delivered_to_planned()

    def fill_delivered_to_planned(self):
        self._status_auto_filling = True
        self._updating_qty = True
        was_blocked = self.qty_table.blockSignals(True)
        try:
            for comp in self.component_keys:
                data_row = self._comp_row.get(comp)
                if data_row is None:
                    continue
                items = self.inputs.get(comp)
                if not items:
                    continue
                planned_item, delivered_item, _ = items
                if not planned_item or not delivered_item:
                    continue
                delivered_item.setText(fmt_num(as_number(planned_item.text())))
                self._update_remaining_row(data_row)
        finally:
            self.qty_table.blockSignals(was_blocked)
            self._updating_qty = False
            self._status_auto_filling = False
        self.refresh_assignment_card()

    def _planned_remaining_state(
        self, planned: Dict[str, float], delivered: Dict[str, float]
    ) -> Tuple[bool, List[str]]:
        active_components = [comp for comp, qty in planned.items() if max(as_number(qty), 0) > 0.0001]
        remaining = [
            comp for comp in active_components
            if max(as_number(planned.get(comp, 0)) - as_number(delivered.get(comp, 0)), 0) > 0.0001
        ]
        return bool(active_components) and not remaining, remaining

    def _recalc_completion(self):
        return

    def _current_planned_for(self, comp: str) -> float:
        items = self.inputs.get(comp)
        if not items:
            return 0.0
        return max(as_number(items[0].text()), 0)

    def assignment_rows(self) -> List[Tuple[str, float, float]]:
        rows = []
        for comp in self.component_keys:
            total = self._system_component_qty(comp)
            assigned = max(as_number(self.planned_assigned.get(comp, 0)), 0) + self._current_planned_for(comp)
            available = total - assigned
            if abs(available) > 0.0001:
                rows.append((comp, assigned, available))
        return rows

    def over_assigned_components(self) -> set:
        return {comp for comp, _assigned, available in self.assignment_rows() if available < -0.0001}

    def filter_qty_components(self, text: str):
        query = normalize_sheet_name(text)
        for comp, data_row in self._comp_row.items():
            hidden = bool(query and query not in normalize_sheet_name(comp))
            self.qty_table.setRowHidden(data_row, hidden)

    def refresh_qty_issue_highlights(self):
        over = self.over_assigned_components()
        issue_bg = QColor("#FEE2E2")
        issue_fg = QColor("#991B1B")
        normal_bg = QColor("#FFFFFF")
        normal_fg = QColor("#0F172A")
        was_blocked = self.qty_table.blockSignals(True)
        try:
            for comp, data_row in self._comp_row.items():
                has_issue = comp in over
                for c in range(self.qty_table.columnCount()):
                    item = self.qty_table.item(data_row, c)
                    if not item:
                        continue
                    item.setBackground(issue_bg if has_issue else normal_bg)
                    item.setForeground(issue_fg if has_issue else normal_fg)
        finally:
            self.qty_table.blockSignals(was_blocked)

    def refresh_assignment_card(self):
        rows = self.assignment_rows()
        was_blocked = self.assignment_table.blockSignals(True)
        try:
            self.assignment_table.setRowCount(len(rows))
            for r, (comp, assigned, available) in enumerate(rows):
                has_issue = available < -0.0001
                bg = QColor("#FEE2E2") if has_issue else QColor("#FFFFFF")
                fg = QColor("#991B1B") if has_issue else QColor("#0F172A")
                values = [comp, assigned, available]
                for c, v in enumerate(values):
                    item = QTableWidgetItem(fmt_num(v) if c else str(v))
                    item.setTextAlignment(Qt.AlignCenter if c else Qt.AlignLeft | Qt.AlignVCenter)
                    item.setBackground(bg)
                    item.setForeground(fg)
                    self.assignment_table.setItem(r, c, item)
                self.assignment_table.setRowHeight(r, 30)
        finally:
            self.assignment_table.blockSignals(was_blocked)
        self.refresh_qty_issue_highlights()
        self._refresh_unit_row_selection()

    def fill_all_system_planned(self):
        self._updating_qty = True
        was_blocked = self.qty_table.blockSignals(True)
        try:
            for comp in self.component_keys:
                data_row = self._comp_row.get(comp)
                if data_row is None:
                    continue
                items = self.inputs.get(comp)
                if not items:
                    continue
                planned_item, delivered_item, _ = items
                if not planned_item:
                    continue
                system_qty = self._system_component_qty(comp)
                assigned_qty = max(as_number(self.planned_assigned.get(comp, 0)), 0)
                allowed_qty = max(system_qty - assigned_qty, 0)
                planned_item.setText(fmt_num(allowed_qty))
                if delivered_item and as_number(delivered_item.text()) > allowed_qty:
                    delivered_item.setText(fmt_num(allowed_qty))
                self._update_remaining_row(data_row)
                if self._is_unit_tracking(comp):
                    self._update_panel_slot_count(comp, int(allowed_qty))
        finally:
            self.qty_table.blockSignals(was_blocked)
            self._updating_qty = False
        self.refresh_assignment_card()

    def fill_remaining_system_planned(self):
        self._updating_qty = True
        was_blocked = self.qty_table.blockSignals(True)
        try:
            for comp in self.component_keys:
                data_row = self._comp_row.get(comp)
                if data_row is None:
                    continue
                items = self.inputs.get(comp)
                if not items:
                    continue
                planned_item, delivered_item, _ = items
                if not planned_item:
                    continue
                system_qty = self._system_component_qty(comp)
                assigned_qty = max(as_number(self.planned_assigned.get(comp, 0)), 0)
                remaining_qty = max(system_qty - assigned_qty, 0)
                planned_item.setText(fmt_num(remaining_qty))
                if delivered_item and as_number(delivered_item.text()) > remaining_qty:
                    delivered_item.setText(fmt_num(remaining_qty))
                self._update_remaining_row(data_row)
                if self._is_unit_tracking(comp):
                    self._update_panel_slot_count(comp, int(remaining_qty))
        finally:
            self.qty_table.blockSignals(was_blocked)
            self._updating_qty = False
        self.refresh_assignment_card()

    def _update_remaining_row(self, row: int):
        p = self.qty_table.item(row, 1)
        d = self.qty_table.item(row, 2)
        r = self.qty_table.item(row, 3)
        if not p or not d or not r:
            return
        pv = as_number(p.text())
        dv = as_number(d.text())
        was_blocked = self.qty_table.blockSignals(True)
        try:
            r.setText(fmt_num(max(pv - dv, 0)))
        finally:
            self.qty_table.blockSignals(was_blocked)

    def on_qty_item_changed(self, item: QTableWidgetItem):
        if self._updating_qty or self._updating_qty_table or not item:
            return
        col = item.column()
        if col not in (1, 2):
            return
        row = item.row()
        comp = self._row_comp.get(row)
        if not comp:
            return
        self._updating_qty_table = True
        self._updating_qty = True
        was_blocked = self.qty_table.blockSignals(True)
        try:
            item.setText(fmt_num(as_number(item.text())))
            if col == 1:
                # Teslim edilecek değişti
                new_qty = as_number(item.text())
                if self._is_unit_tracking(comp):
                    # Ondalıklı değer uyarısı
                    if new_qty != int(new_qty):
                        QMessageBox.warning(
                            self, "Ondalıklı Adet",
                            f"Bu bileşende teslim edilecek adet tam sayı olmalıdır.\n({comp})"
                        )
                        item.setText("0")
                        self._update_remaining_row(row)
                        return
                    delivered_item = self.qty_table.item(row, 2)
                    delivered_qty = as_number(delivered_item.text()) if delivered_item else 0
                    self._update_panel_slot_count(comp, int(max(new_qty, delivered_qty)))
                else:
                    if self._is_delivered_status():
                        delivered_item = self.qty_table.item(row, 2)
                        if delivered_item:
                            delivered_item.setText(fmt_num(new_qty))
            elif col == 2 and self._is_unit_tracking(comp):
                planned_item = self.qty_table.item(row, 1)
                planned_qty = as_number(planned_item.text()) if planned_item else 0
                delivered_qty = as_number(item.text())
                self._update_panel_slot_count(comp, int(max(planned_qty, delivered_qty)))
            self._update_remaining_row(row)
        finally:
            self.qty_table.blockSignals(was_blocked)
            self._updating_qty = False
            self._updating_qty_table = False
        self.refresh_assignment_card()

    # ------------------------------------------------------------------ save
    def save(self):
        if not self.name.text().strip():
            QMessageBox.warning(self, "Eksik", "Teslimat adı girin.")
            return

        self._current_units_from_panel_if_active()
        planned: Dict[str, float] = {}
        delivered: Dict[str, float] = {}
        component_units: Dict[str, list] = {}

        for comp, (p, d, _r) in self.inputs.items():
            pv = as_number(p.text())
            dv = as_number(d.text())
            assigned_other = max(as_number(self.planned_assigned.get(comp, 0)), 0)
            system_qty = self._system_component_qty(comp)
            if pv + assigned_other > system_qty + 0.0001:
                QMessageBox.warning(self, "Hata", f"{comp}: tanımlanan toplam miktar sistem adedini aşamaz.")
                return
            if dv > pv:
                QMessageBox.warning(self, "Hata", f"{comp}: teslim edilen, teslim edilecekten büyük olamaz.")
                return
            planned[comp] = pv
            delivered[comp] = dv

            # Unit tracking validation
            if self._is_unit_tracking(comp) and pv > 0:
                if pv != int(pv):
                    QMessageBox.warning(
                        self, "Ondalıklı Adet",
                        f"Bu bileşende teslim edilecek adet tam sayı olmalıdır.\n({comp})"
                    )
                    return
                if comp == self.active_unit_component and self.left_panel_mode == "unit_tracking":
                    self._component_units_state[comp] = self.unit_side_panel.get_units()
                units = self._ensure_component_units(comp, int(max(pv, dv)))
                counts: Dict[str, int] = {}
                for unit in units:
                    ident = normalize_sheet_name(unit.get("identifier", ""))
                    if ident:
                        counts[ident] = counts.get(ident, 0) + 1
                if any(v > 1 for v in counts.values()):
                    QMessageBox.warning(
                        self, "Tekrar Var",
                        f"{comp}: Aynı kuyruk no / seri no iki kez girilemez. Lütfen düzeltin."
                    )
                    return
                component_units[comp] = units

        t0_text = str(getattr(self.system, "t0_date", "") or self._delivery_t0_date).strip()
        completion = str(getattr(self.system, "completion_date", "") or self._delivery_completion_date).strip()
        plan_acc_text = self.planned_acceptance_date.text().strip()
        ok, message = validate_flexible_date(plan_acc_text, allow_empty=True)
        if not ok:
            QMessageBox.warning(self, "Tarih hatası", f"Planlanan Teslimat Tarihi: {message}")
            return
        acc_text = self.acceptance_date.text().strip()
        ok, message = validate_flexible_date(acc_text, allow_empty=True)
        if not ok:
            QMessageBox.warning(self, "Tarih hatası", f"Gerçek Teslimat Tarihi: {message}")
            return
        acc_date = parse_flexible_date(acc_text)
        if acc_text and not acc_date and self._is_delivered_status():
            QMessageBox.warning(self, "Tarih hatası", "Teslim Edildi durumunda Gerçek Teslimat Tarihi kesin YYYY-MM-DD olmalı.")
            return
        if acc_date and acc_date > date.today():
            QMessageBox.warning(self, "Tarih hatası", "Gerçek Teslimat Tarihi bugünden ileri olamaz.")
            return

        all_delivered, remaining_components = self._planned_remaining_state(planned, delivered)
        if self._is_delivered_status():
            if not acc_text:
                QMessageBox.warning(self, "Gerçek Teslimat Tarihi Gerekli", "Durum 'Teslim Edildi' olduğunda Gerçek Teslimat Tarihi zorunludur.")
                return
            if remaining_components:
                QMessageBox.warning(
                    self, "Teslim Edilen Eksik",
                    "Durum 'Teslim Edildi' olduğunda bu teslimattaki tüm bileşenlerin kalan değeri 0 olmalıdır.\n\n"
                    "Eksik kalan bileşenler:\n• " + "\n• ".join(remaining_components),
                )
                return
        elif all_delivered:
            QMessageBox.warning(
                self, "Durum Uyumsuz",
                "Bu teslimatta tüm bileşenlerin kalanı 0. Kaydetmeden önce Durum alanını 'Teslim Edildi' yapın.",
            )
            return

        self.result = DeliveryInfo(
            name=self.name.text().strip(),
            status=self.status.currentText(),
            acceptance_date=flexible_or_blank(acc_text),
            note=self.note.text().strip(),
            planned_acceptance_date=flexible_or_blank(plan_acc_text),
            planned=planned,
            delivered=delivered,
            t0_date=iso_or_blank(t0_text),
            t0_months=int(getattr(self.system, "t0_months", self._delivery_t0_months) or 0),
            completion_date=completion,
            delivery_user="" if self.delivery_user_combo.currentIndex() <= 0 else self.delivery_user_combo.currentText().strip(),
            component_units=component_units,
        )
        self.accept()




class ContractSharePopover(QFrame):
    """Compact contract-scoped sharing panel used inside ContractActionTabs popover."""

    _CARD_BASE = (
        "QPushButton#sharePermCard{"
        "background:#ffffff; color:#0b2b54; border:1px solid #D8E5F5;"
        "border-radius:12px; padding:8px 12px; font-size:12px; font-weight:700;"
        "text-align:left; min-height:54px;"
        "}"
        "QPushButton#sharePermCard:hover{"
        "background:#F4F8FF; border-color:#93C5FD;"
        "}"
    )
    _CARD_ACTIVE = (
        "QPushButton#sharePermCard{"
        "background:#EFF6FF; color:#1D4ED8; border:1px solid #60A5FA;"
        "border-radius:12px; padding:8px 12px; font-size:12px; font-weight:800;"
        "text-align:left; min-height:54px;"
        "}"
    )

    def __init__(self, owner: "ContractWorkWindow", parent=None):
        super().__init__(parent)
        self.owner = owner
        self._share_mode_value = "goruntule"
        self.setObjectName("contractSharePanel")
        self.setMinimumWidth(336)
        self.setStyleSheet(
            "QFrame#contractSharePanel{background:#F8FBFF;border:1px solid #BFDBFE;border-radius:14px;}"
            "QLabel{background:transparent;border:0;}"
            "QPushButton#shareCloseButton{background:transparent;color:#64748B;border:0;border-radius:8px;font-size:16px;font-weight:900;min-width:28px;min-height:28px;}"
            "QPushButton#shareCloseButton:hover{background:#EAF3FF;color:#1D4ED8;}"
            "QPushButton#shareCreateButton{background:#2563eb;color:#fff;border:0;border-radius:10px;"
            "padding:7px 16px;font-size:12px;font-weight:800;min-height:34px;}"
            "QPushButton#shareCreateButton:hover{background:#1d4ed8;}"
            "QLabel#sharePreview{background:#ffffff;color:#1e3a8a;border:1px solid #bfdbfe;"
            "border-radius:9px;padding:8px 10px;font-family:Consolas,monospace;font-size:11px;}"
        )
        lay = QVBoxLayout(self)
        lay.setContentsMargins(14, 12, 14, 14)
        lay.setSpacing(10)

        header = QHBoxLayout()
        header.setContentsMargins(0, 0, 0, 0)
        title = QLabel("Sözleşme Paylaşımı")
        title.setStyleSheet("color:#10233d;font-size:14px;font-weight:900;")
        close_btn = QPushButton("×")
        close_btn.setObjectName("shareCloseButton")
        close_btn.clicked.connect(self.owner.close_side_meta_popover)
        header.addWidget(title, 1)
        header.addWidget(close_btn, 0, Qt.AlignRight | Qt.AlignVCenter)
        lay.addLayout(header)

        info = QLabel("Sadece bu sözleşmeyi içeren bağımsız STS dosyası oluştur.")
        info.setWordWrap(True)
        info.setStyleSheet("color:#475569;font-size:11px;")
        lay.addWidget(info)

        cards_row = QHBoxLayout()
        cards_row.setSpacing(8)
        cards_row.setContentsMargins(0, 2, 0, 2)

        self._btn_view = QPushButton("👁  Görüntüle\nSadece okuma")
        self._btn_view.setObjectName("sharePermCard")
        self._btn_view.setCursor(Qt.PointingHandCursor)

        self._btn_edit = QPushButton("✏  Düzenle\nDüzenleme izni")
        self._btn_edit.setObjectName("sharePermCard")
        self._btn_edit.setCursor(Qt.PointingHandCursor)

        self._btn_view.clicked.connect(lambda: self._set_mode("goruntule"))
        self._btn_edit.clicked.connect(lambda: self._set_mode("duzenle"))

        cards_row.addWidget(self._btn_view, 1)
        cards_row.addWidget(self._btn_edit, 1)
        lay.addLayout(cards_row)

        self.preview = QLabel("")
        self.preview.setObjectName("sharePreview")
        lay.addWidget(self.preview)

        btn = QPushButton("Paylaşım Dosyası Oluştur")
        btn.setObjectName("shareCreateButton")
        btn.clicked.connect(self.create_share_file)
        lay.addWidget(btn, 0, Qt.AlignRight)

        self._set_mode("goruntule")

    def _set_mode(self, mode: str):
        self._share_mode_value = mode
        self._btn_view.setStyleSheet(
            self._CARD_ACTIVE if mode == "goruntule" else self._CARD_BASE
        )
        self._btn_edit.setStyleSheet(
            self._CARD_ACTIVE if mode == "duzenle" else self._CARD_BASE
        )
        self.update_preview()

    def share_mode(self) -> str:
        return self._share_mode_value

    def filename(self) -> str:
        no = re.sub(r"[^A-Za-z0-9_.-]+", "-", str(getattr(self.owner.ci, "no", "") or "sozlesme")).strip("-._") or "sozlesme"
        stamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
        mode = "edit" if self.share_mode() == "duzenle" else "view"
        return f"STS-{no}__share-{mode}__{stamp}.sts"

    def update_preview(self):
        self.preview.setText(self.filename())

    def create_share_file(self):
        self.owner.create_contract_share_file(self.share_mode(), self.filename())




class BadgeTabButton(QFrame):
    """Contract meta tab with a fixed top-right badge anchor.

    Badge geometry is intentionally independent from the text layout.  Only the
    text changes when the count changes, so active/passive state and 0/1/9/10/99
    counts cannot push the badge or the button content around.
    """

    clicked = Signal(bool)

    BADGE_W = 28
    BADGE_H = 20

    def __init__(self, icon: str, text: str, count: Optional[int] = None, parent=None):
        super().__init__(parent)
        self._checked = False
        self._text = str(text or "")
        self._has_badge = count is not None
        self.setObjectName("badgeTabButton")
        self.setCursor(Qt.PointingHandCursor)
        self.setMinimumHeight(42)
        self.setMinimumWidth(128)
        self.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)

        lay = QHBoxLayout(self)
        # Reserve a stable right-side anchor area for the absolute badge.
        lay.setContentsMargins(14, 7, 38 if self._has_badge else 14, 7)
        lay.setSpacing(8)

        self.icon_lbl = QLabel(str(icon or ""), self)
        self.icon_lbl.setObjectName("badgeTabIcon")
        self.icon_lbl.setAlignment(Qt.AlignCenter)
        self.icon_lbl.setFixedWidth(18)
        lay.addWidget(self.icon_lbl, 0, Qt.AlignVCenter)

        self.text_lbl = QLabel(self._text, self)
        self.text_lbl.setObjectName("badgeTabText")
        self.text_lbl.setAlignment(Qt.AlignVCenter | Qt.AlignLeft)
        self.text_lbl.setSizePolicy(QSizePolicy.MinimumExpanding, QSizePolicy.Preferred)
        lay.addWidget(self.text_lbl, 1, Qt.AlignVCenter)

        self.badge = QLabel("0" if count is None else str(count), self)
        self.badge.setObjectName("badgeTabCount")
        self.badge.setAlignment(Qt.AlignCenter)
        self.badge.setFixedSize(self.BADGE_W, self.BADGE_H)
        self.badge.setAttribute(Qt.WA_TransparentForMouseEvents, True)
        self.badge.setVisible(self._has_badge)
        self._apply_style()
        self._position_badge()

    def setChecked(self, checked: bool):
        checked = bool(checked)
        if self._checked == checked:
            return
        self._checked = checked
        self._apply_style()
        self._position_badge()

    def isChecked(self) -> bool:
        return self._checked

    def setCount(self, count: int):
        self._has_badge = True
        self.badge.setVisible(True)
        self.badge.setText(str(count))
        # Fixed badge width keeps 0/1/9/10/99 visually anchored.
        self._position_badge()

    def resizeEvent(self, event):
        super().resizeEvent(event)
        self._position_badge()

    def _position_badge(self):
        if not hasattr(self, "badge"):
            return
        x = max(4, self.width() - self.BADGE_W - 8)
        y = 5
        self.badge.setGeometry(x, y, self.BADGE_W, self.BADGE_H)
        self.badge.raise_()

    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton:
            self.clicked.emit(False)
            event.accept()
            return
        super().mousePressEvent(event)

    def keyPressEvent(self, event):
        if event.key() in (Qt.Key_Return, Qt.Key_Enter, Qt.Key_Space):
            self.clicked.emit(False)
            event.accept()
            return
        super().keyPressEvent(event)

    def _apply_style(self):
        if self._checked:
            bg = "#EFF6FF"; border = "#60A5FA"; fg = "#1D4ED8"
        else:
            bg = "#FFFFFF"; border = "#D8E5F5"; fg = "#0F2747"
        self.setStyleSheet(f"""
            QFrame#badgeTabButton {{
                background:{bg};
                border:1px solid {border};
                border-top-left-radius:0px;
                border-top-right-radius:0px;
                border-bottom-left-radius:12px;
                border-bottom-right-radius:12px;
            }}
            QFrame#badgeTabButton:hover {{ background:#F4F8FF; border-color:#93C5FD; }}
            QLabel#badgeTabIcon {{ background:transparent; color:{fg}; border:0; font-size:15px; }}
            QLabel#badgeTabText {{ background:transparent; color:{fg}; border:0; font-size:12px; font-weight:700; }}
            QLabel#badgeTabCount {{
                background:#DBEAFE;
                color:#2563EB;
                border:1px solid #93C5FD;
                border-radius:10px;
                padding:0px;
                font-size:10px;
                font-weight:900;
            }}
        """)

class ContractActionTabs(QFrame):
    """Contract-level hanging tabs anchored under the header bar."""

    def __init__(self, owner: "ContractWorkWindow", parent=None):
        super().__init__(parent)
        self.owner = owner
        self.setObjectName("contractActionTabs")

    def open_tags(self):
        self.owner.toggle_side_meta_popover("tags")

    def open_files(self):
        self.owner.toggle_side_meta_popover("files")

    def open_share(self):
        self.owner.toggle_side_meta_popover("share")

class ContractWorkWindow(QDialog):
    def __init__(self, store: ExcelStore, ci: ContractInfo, parent=None, systems: Optional[List[SystemInfo]] = None, deliveries: Optional[Dict[str, List[DeliveryInfo]]] = None):
        super().__init__(parent)
        self.store = store
        parent_staff = getattr(parent, "current_staff", None) if parent is not None else None
        self.current_staff = parent_staff or auth.current_staff
        self._document_lock_state: dict = {}
        # Yeni sozlesme mi (systems/deliveries verilmemis) yoksa mevcut mu
        self.is_new_contract = (systems is None and deliveries is None)
        self.ci = ci
        self.original_platform = str(ci.platform or "")
        self.original_contract_no = str(ci.no or "")
        self.original_contract_type = str(ci.contract_type or "")
        self.original_entry_start_row = int(getattr(ci, "entry_start_row", 0) or 0)
        self.systems: List[SystemInfo] = systems or []
        self.deliveries: Dict[str, List[DeliveryInfo]] = deliveries or {}
        self.contract_tags: List[dict] = self.store.load_contract_tags(
            self.original_platform,
            self.original_contract_no,
            self.original_contract_type,
        )
        dedup: Dict[str, dict] = {}
        for t in self.contract_tags:
            k = normalized_tag_key(str((t or {}).get("name") or ""))
            if not k:
                continue
            dedup[k] = dict(t)
        self.contract_tags = list(dedup.values())
        self.selected_system: Optional[str] = None
        self.expanded_delivery_index: Optional[int] = None
        self._delivery_row_map: Dict[int, int] = {}
        self._updating_summary = False
        self.deleted_contract_info: Optional[dict] = None
        self._context_cache: Dict[Tuple[str, str, str], dict] = {}
        self._deleted_delivery_systems: set[str] = set()
        self._refreshing_contract_context = False
        self._contract_save_thread = None
        self._contract_save_worker = None
        self._pending_contract_save_context = None
        self._file_dialog_open: bool = False
        self._documents_changed: bool = False
        self._is_dirty: bool = False   # Kullanıcı henüz değişiklik yapmadı
        # Yeni sözleşme modunda belgeler için in-memory bekleme yapısı
        self._pending_doc_folders: list = []   # [{id, parent_id, name}]
        self._pending_doc_files: list = []     # [{id, folder_id, filename, content_blob, file_ext, mime_type, size_bytes, note, created_at}]
        self._pending_doc_next_id: int = -1    # Negatif id'ler pending anlamına gelir
        self.share_mode_enabled = False
        self.share_permission_mode = "edit"
        self.setWindowTitle(APP_TITLE)
        # QDialog varsayılan olarak ? butonu gösterir — standart pencere butonları ekle
        self.setWindowFlags(
            Qt.Window |
            Qt.WindowTitleHint |
            Qt.WindowSystemMenuHint |
            Qt.WindowMinimizeButtonHint |
            Qt.WindowMaximizeButtonHint |
            Qt.WindowCloseButtonHint
        )
        self.resize(1450, 860)
        self.setWindowState(Qt.WindowMaximized)  # Varsayılan tam ekran
        self.setStyleSheet(STYLE)
        self.build()
        self.build_busy_overlay()
        # Mevcut sozlesmede 'Secili Sistemi Sil' butonunu gizle
        self.delete_system_btn.setVisible(self.is_new_contract)
        self.refresh()
        # Değişiklik tespiti için başlangıç snapshot'ı al
        if not self.is_new_contract:
            self._apply_derived_statuses(self.ci, self.systems, self.deliveries)
        self._initial_snapshot = self._make_data_snapshot()

    def set_share_mode(self, permission_mode: str = "view"):
        self.share_mode_enabled = True
        self.share_permission_mode = "edit" if str(permission_mode or "").lower() == "edit" else "view"
        self._apply_share_permissions()

    def _share_is_view_only(self) -> bool:
        return bool(getattr(self, "share_mode_enabled", False)) and str(getattr(self, "share_permission_mode", "view")) != "edit"

    def _apply_share_permissions(self):
        if not getattr(self, "share_mode_enabled", False):
            return
        view_only = self._share_is_view_only()
        self.setWindowTitle(f"{APP_TITLE} - Paylaşım ({'Görüntüleme' if view_only else 'Düzenleme'})")
        band = getattr(self, "share_info_band", None)
        if band is not None:
            band.setText("Paylaşım Modu: Görüntüleme — Bu dosyada düzenleme kapalıdır." if view_only else "Paylaşım Modu: Düzenleme — Yalnızca bu sözleşme düzenlenebilir.")
            band.setVisible(True)
        for attr in ("save_btn", "edit_system_btn", "add_system_btn", "add_delivery_btn", "auto_accept_btn", "delete_system_btn"):
            widget = getattr(self, attr, None)
            if widget is not None:
                widget.setEnabled(not view_only)
                if view_only:
                    widget.setToolTip("Paylaşım görüntüleme modunda bu işlem kapalıdır.")
        if hasattr(self, "delete_contract_btn"):
            self.delete_contract_btn.setVisible(not view_only)
            self.delete_contract_btn.setEnabled(not view_only)
            if view_only:
                self.delete_contract_btn.setToolTip("Paylaşım görüntüleme modunda bu işlem kapalıdır.")
        header_edit = getattr(self, "header_edit_btn", None)
        if header_edit is not None:
            header_edit.setEnabled(not view_only)
            if view_only:
                header_edit.setToolTip("Paylaşım görüntüleme modunda bu işlem kapalıdır.")
        for table_name in ("summary", "del_table"):
            table = getattr(self, table_name, None)
            if table is not None:
                table.setEditTriggers(QAbstractItemView.NoEditTriggers if view_only else QAbstractItemView.DoubleClicked | QAbstractItemView.EditKeyPressed | QAbstractItemView.AnyKeyPressed)

    def _ensure_share_can_edit(self, title: str = "Paylaşım") -> bool:
        if self._share_is_view_only():
            QMessageBox.information(self, title, "Bu paylaşım dosyası görüntüleme yetkisiyle açıldı; düzenleme yapılamaz.")
            return False
        return True

    def has_permission(self, permission_code: str) -> bool:
        db_conn = getattr(getattr(self.store, "db", None), "conn", None)
        return auth.has_permission(self.current_staff, permission_code, db_conn)

    def require_permission_ui(self, permission_code: str, title: str = "Yetki gerekli") -> bool:
        if self.has_permission(permission_code):
            return True
        QMessageBox.warning(self, "Yetkisiz İşlem", "Bu işlemi yapmak için gerekli yetkiye sahip değilsiniz.")
        return False

    def _set_dirty(self) -> None:
        """Kullanıcı bir değişiklik yaptığında çağrılır."""
        self._is_dirty = True

    def _mark_documents_changed(self) -> None:
        """Belge/klasör işlemleri STS veritabanına anında yazılır; Kaydet uyarısını bastırmak için izlenir."""
        self._documents_changed = True

    def _finish_persisted_side_meta_only_save(self) -> None:
        self._documents_changed = False
        self._is_dirty = False
        self.accept()

    def _make_data_snapshot(self) -> str:
        """ci + systems + deliveries + tags verilerinin JSON özeti."""
        import json as _json
        ci = self.ci
        data = {
            "no":              str(ci.no or ""),
            "user":            str(ci.user or ""),
            "yi_yd":           str(ci.yi_yd or ""),
            "contract_type":   str(ci.contract_type or ""),
            "signature_date":  str(ci.signature_date or ""),
            "t0_date":         str(ci.t0_date or ""),
            "t0_months":       int(ci.t0_months or 0),
            "completion_date": str(ci.completion_date or ""),
            "status":          str(ci.status or ""),
            "note":            str(ci.note or ""),
            "acceptance_date": str(ci.acceptance_date or ""),
            "systems": sorted([
                {
                    "name":            str(s.name or ""),
                    "components":      {k: float(v) for k, v in sorted((s.components or {}).items())},
                    "component_notes": {k: str(v or "") for k, v in sorted((getattr(s, "component_notes", {}) or {}).items()) if str(v or "")},
                    "t0_date":         str(s.t0_date or ""),
                    "t0_months":       int(s.t0_months or 0),
                    "completion_date": str(s.completion_date or ""),
                    "status":          str(s.status or ""),
                    "acceptance_date": str(s.acceptance_date or ""),
                    "deliveries": sorted([
                        {
                            "name":            str(d.name or ""),
                            "status":          str(d.status or ""),
                            "planned_acceptance_date": str(getattr(d, "planned_acceptance_date", "") or ""),
                            "acceptance_date": str(d.acceptance_date or ""),
                            "note":            str(d.note or ""),
                            "delivery_user":   str(getattr(d, "delivery_user", "") or ""),
                            "planned":  {k: float(v) for k, v in sorted((d.planned or {}).items())},
                            "delivered":{k: float(v) for k, v in sorted((d.delivered or {}).items())},
                            "component_units": {
                                k: sorted([{kk: vv for kk, vv in u.items()} for u in v], key=lambda x: x.get("slot_no", 0))
                                for k, v in sorted((getattr(d, "component_units", {}) or {}).items())
                            },
                        }
                        for d in self.deliveries.get(s.name, [])
                    ], key=lambda x: x["name"]),
                }
                for s in self.systems
            ], key=lambda x: x["name"]),
            "tags": sorted(str((t or {}).get("name", "") or "") for t in self.contract_tags),
        }
        return _json.dumps(data, sort_keys=True, ensure_ascii=False)

    def date_picker_events(self) -> List[dict]:
        return contract_date_picker_events(self.ci, self.systems)

    def build(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(12, 8, 12, 10)
        root.setSpacing(6)

        header = QFrame(); header.setObjectName("contractHeader")
        header.setFixedHeight(68)
        self.contract_header = header  # sekme konumlandırması için referans
        h = QHBoxLayout(header); h.setContentsMargins(18, 7, 16, 7); h.setSpacing(14)

        info_row = QWidget(); info_row.setObjectName("contractHeaderInfoRow")
        info_row.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        info = QHBoxLayout(info_row); info.setContentsMargins(0, 0, 0, 0); info.setSpacing(0)

        actions = QWidget(); actions.setObjectName("contractHeaderActions")
        actions.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        actions_lay = QHBoxLayout(actions); actions_lay.setContentsMargins(0, 0, 0, 0); actions_lay.setSpacing(8)

        self.meta_values: Dict[str, QLabel] = {}
        self.user_tooltip_text = ""
        self.user_popup_users: List[str] = []
        self.user_summary_container: Optional[QWidget] = None
        self._user_tooltip_widgets: set[QWidget] = set()
        self._user_popup: Optional[HeaderUserPopup] = None
        self.platform_tabs_widget: Optional[PlatformTabsWidget] = None

        def meta_cell(key, label_text, value_text, *, min_w=70, max_w=None, value_widget=None, tooltip: str = ""):
            cell = QWidget(); cell.setObjectName("metaCell")
            cell.setMinimumWidth(min_w)
            if max_w:
                cell.setMaximumWidth(max_w)
            cell.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
            if tooltip:
                cell.setToolTip(tooltip)
                cell.setMouseTracking(True)
                if key == "user":
                    self._register_user_tooltip_widget(cell)
            cl = QVBoxLayout(cell); cl.setContentsMargins(10, 0, 10, 0); cl.setSpacing(2)
            lbl = QLabel(label_text.upper()); lbl.setObjectName("metaHeaderLabel")
            if value_widget is None:
                val = ElidedValueLabel(value_text if value_text else "-"); val.setObjectName("metaHeaderValue")
                self.meta_values[key] = val
            else:
                val = value_widget
            if tooltip and hasattr(val, "setToolTip"):
                val.setToolTip(tooltip)
                val.setMouseTracking(True)
                if key == "user" and isinstance(val, QWidget):
                    self._register_user_tooltip_widget(val)
            cl.addWidget(lbl); cl.addWidget(val)
            div = QFrame(); div.setObjectName("metaHeaderDiv")
            div.setFixedSize(1, 32)
            return cell, div

        def compact_users(value: str, user_list: Optional[List[str]] = None) -> Tuple[str, str]:
            users = [str(x).strip() for x in (user_list or []) if str(x).strip()]
            if not users:
                users = [x.strip() for x in re.split(r"[,;]+", str(value or "")) if x.strip()]
            if not users:
                return "-", ""
            shown = users[0] if len(users) == 1 else f"{users[0]} +{len(users) - 1}"
            return shown, "\n".join(users)

        def inline_icon_text(text: str, icon_svg: bytes, object_name: str, tooltip: str = "") -> QWidget:
            wrap = QWidget(); wrap.setObjectName(object_name)
            row = QHBoxLayout(wrap); row.setContentsMargins(0, 0, 0, 0); row.setSpacing(6)
            pix = QPixmap(); pix.loadFromData(icon_svg, "SVG")
            icon = QLabel(); icon.setObjectName("headerUserIcon"); icon.setPixmap(pix); icon.setFixedSize(16, 16)
            icon.setStyleSheet("QLabel#headerUserIcon{background:transparent;border:0;padding:0;margin:0;}")
            val = ElidedValueLabel(text if text else "-"); val.setObjectName("metaHeaderValue")
            if tooltip:
                if object_name == "user":
                    self.user_summary_container = wrap
                for widget in (wrap, icon, val):
                    widget.setToolTip(tooltip)
                    widget.setMouseTracking(True)
                    if object_name == "user":
                        self._register_user_tooltip_widget(widget)
            self.meta_values[object_name] = val
            row.addWidget(icon, 0, Qt.AlignVCenter)
            row.addWidget(val, 1, Qt.AlignVCenter)
            return wrap

        def status_widget(text: str) -> QWidget:
            wrap = QWidget(); wrap.setObjectName("headerStatusWrap")
            row = QHBoxLayout(wrap); row.setContentsMargins(0, 0, 0, 0); row.setSpacing(7)
            dot = QLabel(); dot.setObjectName("headerStatusDot"); dot.setFixedSize(10, 10)
            val = ElidedValueLabel(text if text else "Başlanmadı"); val.setObjectName("metaHeaderValue")
            self.meta_values["status"] = val
            row.addWidget(dot, 0, Qt.AlignVCenter)
            row.addWidget(val, 1, Qt.AlignVCenter)
            return wrap

        user_text, user_tip = compact_users(self.ci.user, list(getattr(self.ci, "users", []) or []))
        responsible_name = str(getattr(self.ci, "responsible_engineer_name", "") or "").strip()
        if not responsible_name:
            responsible_items = list(getattr(self.ci, "responsible_engineers", []) or [])
            if responsible_items:
                responsible_name = str(responsible_items[0].get("full_name") or responsible_items[0].get("name") or "").strip()
        responsible_text = responsible_name or "-"
        self.user_tooltip_text = user_tip
        self.user_popup_users = [u.strip() for u in user_tip.splitlines() if u.strip()]
        user_svg = b"""<svg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 16 16' width='16' height='16'>
          <circle cx='8' cy='5.35' r='2.45' fill='none' stroke='#c8e2ff' stroke-width='1.35'/>
          <path d='M3.15 13.35c.48-2.75 2.18-4.15 4.85-4.15s4.37 1.4 4.85 4.15'
                fill='none' stroke='#9fc7f2' stroke-width='1.35' stroke-linecap='round' stroke-linejoin='round'/>
          <path d='M4.55 12.55c.7-1.45 1.86-2.16 3.45-2.16s2.75.71 3.45 2.16'
                fill='none' stroke='#4e93ff' stroke-opacity='.55' stroke-width='.85' stroke-linecap='round'/>
        </svg>"""
        self.platform_tabs_widget = PlatformTabsWidget(self)
        self.active_platform_id = int(getattr(self.ci, "platform_id", 0) or getattr(self.ci, "primary_platform_id", 0) or 0)
        self.platform_tabs_widget.set_platforms(self._linked_contract_platforms(), self.active_platform_id)
        self.platform_tabs_widget.activePlatformChanged.connect(self.set_active_platform)

        cells = [
            (*meta_cell("no", "Sözleşme No", self.ci.no, min_w=122, max_w=210), 17),
            (*meta_cell("platform", "Platform", "", min_w=260, max_w=340, value_widget=self.platform_tabs_widget), 0),
            (*meta_cell("type", "Tür", self.ci.contract_type, min_w=96, max_w=160), 13),
            (*meta_cell("responsible_engineer", "Sorumlu Mühendis", responsible_text, min_w=138, max_w=220, tooltip=responsible_name), 17),
            (*meta_cell("user", "Kullanıcı", user_text, min_w=140, max_w=230, value_widget=inline_icon_text(user_text, user_svg, "user", user_tip), tooltip=user_tip), 18),
            (*meta_cell("status", "Durum", "", min_w=112, max_w=170, value_widget=status_widget(self.ci.status or "Başlanmadı")), 13),
        ]
        if user_tip and self.meta_values.get("user"):
            self.meta_values["user"].setToolTip(user_tip)
        for i, (cell, div, stretch) in enumerate(cells):
            info.addWidget(cell, stretch)
            if i < len(cells) - 1:
                info.addWidget(div)
                info.addSpacing(2)

        h.addWidget(info_row, 1)
        e = QPushButton("  Ana Bilgileri Düzenle"); e.setObjectName("headerEditBtn")
        e.setFixedHeight(36)
        _svg = b"""<svg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 16 16' width='15' height='15'>
          <path d='M11.5 1.5a1.5 1.5 0 0 1 2.12 2.12l-9 9a1 1 0 0 1-.4.24l-3 1a.5.5 0 0 1-.63-.63l1-3a1 1 0 0 1 .24-.4z'
                fill='none' stroke='#d8eaff' stroke-width='1.4' stroke-linecap='round' stroke-linejoin='round'/>
        </svg>"""
        _pix = QPixmap()
        _pix.loadFromData(_svg, "SVG")
        e.setIcon(QIcon(_pix))
        e.setIconSize(QSize(15, 15))
        e.clicked.connect(self.edit_contract_info)
        self.header_edit_btn = e
        actions_lay.addWidget(e)
        self.delete_contract_btn = QPushButton("Sözleşmeyi Sil")
        self.delete_contract_btn.setObjectName("danger")
        self.delete_contract_btn.setFixedHeight(36)
        trash_svg = b"""<svg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 16 16' width='15' height='15'>
          <path d='M2.8 4.2h10.4M6.2 4.2V2.7h3.6v1.5M4.2 5.8l.45 7.1c.05.75.48 1.1 1.2 1.1h4.3c.72 0 1.15-.35 1.2-1.1l.45-7.1M6.8 7.4v4.1M9.2 7.4v4.1'
                fill='none' stroke='#dc2626' stroke-width='1.25' stroke-linecap='round' stroke-linejoin='round'/>
        </svg>"""
        trash_pix = QPixmap(); trash_pix.loadFromData(trash_svg, "SVG")
        self.delete_contract_btn.setIcon(QIcon(trash_pix))
        self.delete_contract_btn.setIconSize(QSize(15, 15))
        self.delete_contract_btn.clicked.connect(self.delete_contract)
        actions_lay.addWidget(self.delete_contract_btn)
        h.addWidget(actions, 0, Qt.AlignRight | Qt.AlignVCenter)
        root.addWidget(header)

        body = QHBoxLayout(); body.setSpacing(10); root.addLayout(body, 1)

        left_block_width = 404
        left_block = QWidget()
        left_block.setFixedWidth(left_block_width)
        left_block_lay = QVBoxLayout(left_block)
        left_block_lay.setContentsMargins(0, 0, 0, 0)
        left_block_lay.setSpacing(10)

        left_row = QHBoxLayout()
        left_row.setContentsMargins(0, 0, 0, 0)
        left_row.setSpacing(10)

        version_bar = QFrame(); version_bar.setObjectName("contractVersionBar"); version_bar.setFixedWidth(94)
        vb = QVBoxLayout(version_bar); vb.setContentsMargins(8, 10, 8, 10); vb.setSpacing(8)
        self.sd_list = QListWidget()
        self.sd_list.setObjectName("sdList")
        self.sd_list.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.sd_list.currentRowChanged.connect(self.on_sd_nav_changed)
        vb.addWidget(self.sd_list, 1)
        self.add_sd_btn = QPushButton("+")
        self.add_sd_btn.setToolTip("Bu ana sözleşmeye SD ekle")
        self.add_sd_btn.clicked.connect(self.add_sd_contract)
        vb.addWidget(self.add_sd_btn, 0)
        left_row.addWidget(version_bar, 0)

        left_content_host = QWidget()
        left_content_host.setObjectName("sideMetaHost")
        left_content_host.setFixedWidth(300)
        left_content_layout = QVBoxLayout(left_content_host)
        left_content_layout.setContentsMargins(0, 0, 0, 0)
        left_content_layout.setSpacing(10)

        self.systems_panel = QFrame(); self.systems_panel.setObjectName("sidebar")
        lv = QVBoxLayout(self.systems_panel); lv.setContentsMargins(10, 12, 10, 12); lv.setSpacing(10)
        top = QHBoxLayout(); lbl = QLabel("SİSTEMLER"); lbl.setObjectName("sideTitle"); top.addWidget(lbl); top.addStretch()
        add = QPushButton("+"); add.clicked.connect(self.add_system); add.setMinimumHeight(30); add.setMaximumWidth(34); self.add_system_btn = add; top.addWidget(add); lv.addLayout(top)
        self.system_list = QListWidget(); self.system_list.setObjectName("systemList"); self.system_list.currentRowChanged.connect(self.select_system); lv.addWidget(self.system_list, 1)
        delsys = QPushButton("Seçili Sistemi Sil")
        delsys.setObjectName("secondary")
        delsys.clicked.connect(self.delete_system)
        delsys.setMinimumHeight(38)
        self.delete_system_btn = delsys
        lv.addWidget(delsys)
        left_content_layout.addWidget(self.systems_panel, 1)
        left_row.addWidget(left_content_host, 1)
        left_block_lay.addLayout(left_row, 1)
        body.addWidget(left_block, 0)

        right = QFrame(); right.setObjectName("contentPanel"); rv = QVBoxLayout(right); rv.setContentsMargins(16, 10, 16, 12); rv.setSpacing(8); body.addWidget(right, 1)

        self.share_info_band = QLabel("")
        self.share_info_band.setObjectName("shareInfoBand")
        self.share_info_band.setWordWrap(True)
        self.share_info_band.setVisible(False)
        self.share_info_band.setStyleSheet("QLabel#shareInfoBand{background:#eff6ff;color:#1e3a8a;border:1px solid #bfdbfe;border-radius:8px;padding:6px 10px;font-size:12px;font-weight:700;}")
        rv.addWidget(self.share_info_band, 0)

        self.side_meta_host = QWidget(self)
        self.side_meta_host.setObjectName("contractTabsHost")
        self.side_meta_host.setStyleSheet("QWidget#contractTabsHost{background:transparent;border:0;}")
        self.side_meta_host.setAttribute(Qt.WA_NoSystemBackground, True)
        self.side_meta_host.setAttribute(Qt.WA_TranslucentBackground, True)
        self.side_meta_host.setAutoFillBackground(False)
        self.side_meta_host.installEventFilter(self)
        tabs_host_lay = QVBoxLayout(self.side_meta_host)
        tabs_host_lay.setContentsMargins(0, 0, 0, 0)
        tabs_host_lay.setSpacing(0)
        self.contract_action_tabs = ContractActionTabs(self, self.side_meta_host)
        self.build_side_meta_popover_bar(0)
        tabs_host_lay.addWidget(self.side_meta_bar, 0, Qt.AlignRight)
        # side_meta_host layout'a değil, header altına overlay olarak konumlanır
        QTimer.singleShot(80, self._place_tab_bar)
        self.render_contract_tags()

        # ── Üst satır: SİSTEM BİLEŞENLERİ etiketi + Sistemi Düzenle butonu aynı hizada ──
        self.title = QLabel("")  # refresh_right'ta güncellenir ama görünmez
        self.title.setVisible(False)
        top_row = QHBoxLayout()
        top_row.setContentsMargins(0, 0, 0, 0)
        top_row.setSpacing(10)
        self.system_metric_labels: Dict[str, QLabel] = {}
        self.system_metric_cards: Dict[str, QFrame] = {}

        def system_metric_card(key: str, title: str) -> QFrame:
            card = QFrame()
            card.setObjectName("systemMetricCard")
            lay = QVBoxLayout(card)
            lay.setContentsMargins(12, 7, 12, 7)
            lay.setSpacing(2)
            t = QLabel(title.upper())
            t.setObjectName("systemMetricTitle")
            v = QLabel("-")
            v.setObjectName("systemMetricValue")
            v.setWordWrap(True)
            card.setMinimumWidth(130)
            self.system_metric_labels[key] = v
            self.system_metric_cards[key] = card
            lay.addWidget(t)
            lay.addWidget(v)
            return card

        top_row.addWidget(system_metric_card("completion", "Termin Tarihi"), 0)
        top_row.addWidget(system_metric_card("days", "Kalan Gün"), 0)
        top_row.addWidget(system_metric_card("acceptance", "Gerçek Teslimat"), 0)
        top_row.addWidget(system_metric_card("user", "Kullanıcı"), 0)
        top_row.addStretch(1)
        self.edit_system_btn = QPushButton("✎ Sistemi Düzenle")
        self.edit_system_btn.setObjectName("secondary")
        top_row.addWidget(self.edit_system_btn, 0)
        rv.addLayout(top_row)
        self.edit_system_btn.clicked.connect(self.edit_system)

        self.summary = QTableWidget(0, 5)
        configure_table(self.summary)
        self.summary.verticalHeader().setDefaultSectionSize(38)
        self.summary.setHorizontalHeaderLabels(["Bileşen", "Sözleşme Adedi", "Teslim Edilen", "Kalan", "Not"])
        self.configure_summary_columns()
        self.summary.itemChanged.connect(self.on_summary_changed)
        self.summary.setMinimumHeight(340)
        self.summary.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        rv.addWidget(self.summary, 1)  # Kalan dikey alanın ana bölümü sistem bileşenlerine ayrılsın

        # ── Boşluk + TESLİMATLAR ──────────────────────────
        rv.addSpacing(18)
        dh = QHBoxLayout()
        dh.addWidget(section_label("TESLİMATLAR"))
        dh.addStretch()

        ad = QPushButton("+ Teslimat Ekle")
        self.add_delivery_btn = ad
        ad.clicked.connect(self.add_delivery)
        dh.addWidget(ad)

        auto_btn = QPushButton("Otomatik Teslimat Oluştur")
        self.auto_accept_btn = auto_btn
        auto_btn.clicked.connect(lambda: open_auto_accept_dialog(self))
        dh.addWidget(auto_btn)

        rv.addLayout(dh)

        self.del_table = QTableWidget(0, 0)
        configure_table(self.del_table)
        self.del_table.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.del_table.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.del_table.setVerticalScrollMode(QTableWidget.ScrollPerPixel)
        self.del_table.setMinimumHeight(118)
        self.del_table.setMaximumHeight(180)
        self.del_table.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Maximum)
        self.del_table.cellClicked.connect(self.on_delivery_clicked)
        self.del_table.horizontalHeader().setVisible(True)
        self.pinned_delivery = QTableWidget(0, 0)
        self.pinned_delivery.setObjectName("pinnedDelivery")
        self.pinned_delivery.setVisible(False)
        self.pinned_delivery.setMaximumHeight(0)

        rv.addWidget(self.del_table, 0)

        foot = QHBoxLayout(); foot.addStretch()
        save = QPushButton("Kaydet"); self.save_btn = save; save.clicked.connect(self.save_all)
        close = QPushButton("• Kapat"); close.setObjectName("secondary"); close.clicked.connect(self.reject)
        foot.addWidget(save); foot.addWidget(close); root.addLayout(foot)

    def build_busy_overlay(self):
        self.busy_overlay = QFrame(self)
        self.busy_overlay.setStyleSheet("QFrame { background: rgba(248, 251, 255, 0.82); }")
        self.busy_overlay.hide()
        self.busy_overlay.raise_()

        self.busy_card = QFrame(self.busy_overlay)
        self.busy_card.setStyleSheet(
            "QFrame { background: rgba(255,255,255,0.97); border: 1px solid #d8e2ed; border-radius: 12px; }"
            "QLabel { background: transparent; }"
        )
        lay = QVBoxLayout(self.busy_card)
        lay.setContentsMargins(20, 18, 20, 18)
        lay.setSpacing(10)
        self.busy_label = QLabel("İşlem yapılıyor...")
        self.busy_label.setObjectName("mainTitle")
        self.busy_label.setAlignment(Qt.AlignCenter)
        self.busy_progress = QProgressBar()
        self.busy_progress.setRange(0, 100)
        self.busy_progress.setValue(0)
        self.busy_progress.setFormat("%p%")
        self.busy_progress.setTextVisible(True)
        lay.addWidget(self.busy_label)
        lay.addWidget(self.busy_progress)
        self.position_busy_overlay()

    def position_busy_overlay(self):
        if not hasattr(self, "busy_overlay"):
            return
        self.busy_overlay.setGeometry(self.rect())
        w, h = 420, 130
        x = max((self.busy_overlay.width() - w) // 2, 0)
        y = max((self.busy_overlay.height() - h) // 2, 0)
        self.busy_card.setGeometry(x, y, w, h)
        self.busy_overlay.raise_()

    def set_busy_overlay(self, visible: bool, message: str = "İşlem yapılıyor...", percent: int = 0):
        if not hasattr(self, "busy_overlay"):
            return
        if not hasattr(self, "_busy_cursor_on"):
            self._busy_cursor_on = False
        if visible:
            self.busy_label.setText(message)
            self.busy_progress.setRange(0, 100)
            self.busy_progress.setValue(int(max(0, min(100, percent))))
            self.position_busy_overlay()
            self.busy_overlay.show()
            if not self._busy_cursor_on:
                QApplication.setOverrideCursor(Qt.WaitCursor)
                self._busy_cursor_on = True
            QApplication.processEvents()
        else:
            self.busy_overlay.hide()
            if self._busy_cursor_on:
                QApplication.restoreOverrideCursor()
                self._busy_cursor_on = False
            QApplication.processEvents()

    @staticmethod
    def _is_widget_inside(widget, parent_widget) -> bool:
        if widget is None or parent_widget is None:
            return False
        return widget is parent_widget or parent_widget.isAncestorOf(widget)

    def _side_meta_event_widget(self, obj, event):
        if isinstance(obj, QWidget):
            return obj
        global_pos = None
        if hasattr(event, "globalPosition"):
            global_pos = event.globalPosition().toPoint()
        elif hasattr(event, "globalPos"):
            global_pos = event.globalPos()
        return QApplication.widgetAt(global_pos) if global_pos is not None else None

    def _is_side_meta_inside_click(self, obj, event) -> bool:
        widget = self._side_meta_event_widget(obj, event)
        popover = getattr(self, "side_meta_popover", None)
        bar = getattr(self, "side_meta_bar", None)
        return self._is_widget_inside(widget, popover) or self._is_widget_inside(widget, bar)

    def eventFilter(self, obj, event):
        # Bu dialogda da dosya kartları/event filter kullanılıyor. Silinmiş Qt
        # objelerine dokunmak PySide6'da uygulamayı abort ettirebildiği için
        # eventFilter içinde hiçbir hatanın dışarı kaçmasına izin vermiyoruz.
        try:
            if not qt_obj_alive(obj) or event is None:
                return False
            try:
                etype = event.type()
            except Exception:
                return False

            user_tooltip_widgets = getattr(self, "_user_tooltip_widgets", set())
            if obj in user_tooltip_widgets:
                if etype == QEvent.Enter:
                    self._show_user_popup_now(obj)
                    return False
                if etype == QEvent.Leave:
                    self._schedule_user_popup_hide()
                    return False

            popup = getattr(self, "_user_popup", None)
            if obj is popup:
                if etype == QEvent.Enter:
                    return False
                if etype == QEvent.Leave:
                    self._schedule_user_popup_hide()
                    return False

            side_host = getattr(self, "side_meta_host", None)
            if obj is side_host and etype in (QEvent.Resize, QEvent.Show):
                self.position_side_meta_popover()
                self._place_tab_bar()

            _editing = (
                getattr(self, "_tree_editing", False)
                or getattr(self, "_file_dialog_open", False)
                or getattr(self, "_side_meta_modal_open", False)
            )
            panel_open = bool(getattr(self, "_side_meta_open_panel", None))

            if etype in (QEvent.WindowDeactivate, QEvent.ApplicationDeactivate) and panel_open and not _editing:
                self.close_side_meta_popover()

            if etype == QEvent.MouseButtonPress and panel_open:
                if not _editing and not self._is_side_meta_inside_click(obj, event):
                    self.close_side_meta_popover()

            if etype == QEvent.MouseButtonDblClick:
                file_id = None
                try:
                    if hasattr(obj, "property"):
                        file_id = obj.property("contractFileId")
                except Exception:
                    file_id = None
                if file_id:
                    self.open_contract_file(int(file_id))
                    return True

        except Exception:
            try:
                traceback.print_exc()
            except Exception:
                pass
            return False

        return False


    def _register_user_tooltip_widget(self, widget):
        if not isinstance(widget, QWidget):
            return
        anchor = getattr(self, "user_summary_container", None)
        if isinstance(anchor, QWidget) and widget is not anchor:
            return
        widget.installEventFilter(self)
        widget.setMouseTracking(True)
        self._user_tooltip_widgets.add(widget)

    def _show_user_popup_now(self, obj):
        users = [str(u).strip() for u in getattr(self, "user_popup_users", []) if str(u).strip()]
        if not users:
            return
        anchor = getattr(self, "user_summary_container", None)
        if not isinstance(anchor, QWidget) or not qt_obj_alive(anchor):
            anchor = obj if isinstance(obj, QWidget) else None
        if not isinstance(anchor, QWidget) or not qt_obj_alive(anchor):
            return
        popup = getattr(self, "_user_popup", None)
        if not isinstance(popup, HeaderUserPopup) or not qt_obj_alive(popup):
            popup = HeaderUserPopup(None)
            popup.installEventFilter(self)
            popup.setMouseTracking(True)
            self._user_popup = popup
        popup.set_users(users)
        popup.adjustSize()
        hint = popup.sizeHint().expandedTo(QSize(220, 80))
        popup.resize(hint)
        pos = anchor.mapToGlobal(QPoint(0, anchor.height() + 6))
        screen = QApplication.screenAt(pos) or QApplication.primaryScreen()
        if screen:
            available = screen.availableGeometry()
            if pos.x() + popup.width() > available.right():
                pos.setX(max(available.left(), available.right() - popup.width() - 8))
            if pos.y() + popup.height() > available.bottom():
                pos.setY(max(available.top(), anchor.mapToGlobal(QPoint(0, -popup.height() - 6)).y()))
        if popup.isVisible():
            if popup.pos() != pos:
                popup.move(pos)
            return
        popup.move(pos)
        popup.show()
        popup.raise_()

    def _schedule_user_popup_hide(self):
        QTimer.singleShot(120, self._hide_user_popup_if_outside)

    def _hide_user_popup_if_outside(self):
        popup = getattr(self, "_user_popup", None)
        if not isinstance(popup, HeaderUserPopup) or not qt_obj_alive(popup) or not popup.isVisible():
            return
        pos = QCursor.pos()
        if popup.geometry().contains(pos):
            return
        anchor = getattr(self, "user_summary_container", None)
        if isinstance(anchor, QWidget) and qt_obj_alive(anchor):
            rect = anchor.rect()
            if rect.isValid() and rect.contains(anchor.mapFromGlobal(pos)):
                return
        popup.hide()

    def _hide_user_tooltip_now(self):
        popup = getattr(self, "_user_popup", None)
        if isinstance(popup, HeaderUserPopup) and qt_obj_alive(popup):
            popup.hide()

    def configure_summary_columns(self):
        header = self.summary.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.Stretch)
        header.setStretchLastSection(True)
        self.summary.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)

    def update_timeline_bar(self):
        """Compact timeline bar varsa güncelle; yoksa sessizce geç."""
        if hasattr(self, '_tl_prog'):
            from datetime import date as _d
            today = _d.today()
            t0 = parse_iso_date(str(self.ci.t0_date or ""))
            comp = parse_iso_date(str(self.ci.completion_date or ""))
            pct = 0
            if t0 and comp and comp > t0:
                pct = max(0, min(100, int((today - t0).days * 100 / (comp - t0).days)))
            self._tl_prog.setValue(pct)
            self._tl_name_lbl.setText(str(self.ci.no or "—"))

    def _linked_contract_platforms(self) -> List[dict]:
        explicit = list(getattr(self.ci, "platforms", []) or [])
        if not explicit:
            names = list(getattr(self.ci, "platform_names", []) or [])
            ids = list(getattr(self.ci, "platform_ids", []) or [])
            explicit = [{"platform_id": int(ids[i]) if i < len(ids) else 0, "platform_name": name} for i, name in enumerate(names)]
        if getattr(self.ci, "entry_start_row", 0):
            try:
                rows = self.store.get_contract_platforms(int(getattr(self.ci, "entry_start_row", 0) or 0))
                if rows:
                    explicit = rows
            except Exception:
                pass
        base = str(self.ci.platform or "").strip()
        base_id = int(getattr(self.ci, "platform_id", 0) or getattr(self.ci, "primary_platform_id", 0) or 0)
        found: List[dict] = []
        seen = set()
        for item in explicit + ([{"platform_id": base_id, "platform_name": base}] if base else []):
            if isinstance(item, dict):
                pid = int(item.get("platform_id") or item.get("id") or 0)
                name = str(item.get("platform_name") or item.get("name") or "").strip()
                is_primary = bool(item.get("is_primary"))
            else:
                name = str(item or "").strip()
                pid = self.store.get_platform_id(name, create=False) or 0
                is_primary = False
            key = pid if pid else name.casefold()
            if name and key not in seen:
                seen.add(key)
                found.append({"platform_id": int(pid or 0), "platform_name": name, "is_primary": is_primary})
        return found



    def set_active_platform(self, platform_id: int):
        try:
            platform_id = int(platform_id or 0)
        except Exception:
            platform_id = 0
        linked = self._linked_contract_platforms()
        valid = {int(p.get("platform_id") or 0): str(p.get("platform_name") or "") for p in linked if int(p.get("platform_id") or 0)}
        if not platform_id or platform_id not in valid:
            primary = self.store.get_primary_contract_platform(int(getattr(self.ci, "entry_start_row", 0) or 0)) if getattr(self.ci, "entry_start_row", 0) else None
            platform_id = int((primary or {}).get("platform_id") or self.active_platform_id or 0)
            if not platform_id or platform_id not in valid:
                return
        if platform_id == int(getattr(self, "active_platform_id", 0) or 0):
            return
        self._cache_current_context()
        try:
            ci, systems, deliveries = self.store.load_contract_structure(
                self.ci.no,
                contract_no=self.ci.no,
                start_row=self.ci.entry_start_row,
                contract_type=self.ci.contract_type,
                platform_id=platform_id,
            )
        except Exception as exc:
            QMessageBox.warning(self, "Platform yüklenemedi", f"Seçilen platform verisi yüklenemedi:\n{exc}")
            return
        self.active_platform_id = platform_id
        self.ci.platform = valid.get(platform_id, str(getattr(ci, "platform", "") or ""))
        self.ci.platform_id = platform_id
        setattr(self.ci, "platforms", linked)
        setattr(self.ci, "platform_names", [p.get("platform_name") for p in linked])
        setattr(self.ci, "platform_ids", [int(p.get("platform_id") or 0) for p in linked if int(p.get("platform_id") or 0)])
        self.systems = systems or []
        self.deliveries = deliveries or {}
        self.selected_system = self.systems[0].name if self.systems else None
        self.expanded_delivery_index = None
        self.refresh_contract_header()
        self.refresh()

    def refresh_contract_header(self):
        if not hasattr(self, "meta_values"):
            return
        users = [str(x).strip() for x in (list(getattr(self.ci, "users", []) or [])) if str(x).strip()]
        if not users:
            users = [x.strip() for x in re.split(r"[,;]+", str(self.ci.user or "")) if x.strip()]
        user_text = "-" if not users else (users[0] if len(users) == 1 else f"{users[0]} +{len(users) - 1}")
        mapping = {
            "no": self.ci.no,
            "type": self.ci.contract_type,
            "responsible_engineer": str(getattr(self.ci, "responsible_engineer_name", "") or "-").strip() or "-",
            "user": user_text,
            "status": self.ci.status or "Başlanmadı",
        }
        for k, v in mapping.items():
            lab = self.meta_values.get(k)
            if lab:
                lab.setText(str(v or "-"))
                if k == "user":
                    lab.setToolTip("\n".join(users))
        tabs = getattr(self, "platform_tabs_widget", None)
        if tabs:
            tabs.set_platforms(self._linked_contract_platforms(), self.active_platform_id)

    def _notify_parent_contract_updated(self, old_platform: str, new_platform: str) -> None:
        """Ana bilgi güncellemesi sonrası ana listeyi ve açık takvimi tazeler."""
        parent = self.parent()
        if not parent:
            return
        try:
            if hasattr(parent, "request_refresh"):
                old_p = str(old_platform or "").strip()
                new_p = str(new_platform or "").strip()
                if old_p and new_p and old_p != new_p:
                    parent.request_refresh(select_platform=old_p, scope="platform", platform=old_p)
                    parent.request_refresh(select_platform=new_p, scope="platform", platform=new_p)
                else:
                    target = new_p or old_p
                    parent.request_refresh(select_platform=target, scope="platform", platform=target)
            elif hasattr(parent, "refresh_open_calendar"):
                parent.refresh_open_calendar()
        except Exception:
            pass

    def edit_contract_info(self):
        if not self._ensure_share_can_edit("Ana Bilgileri Düzenle"):
            return
        setattr(self.ci, "platforms", self._linked_contract_platforms())
        dlg = ContractEditDialog(self.store, self.ci, self)
        if not dlg.exec() or not dlg.result:
            return

        new_ci = dlg.result
        previous_platforms = self._linked_contract_platforms()
        previous_platform_names = [str(p.get("platform_name") or "").strip() if isinstance(p, dict) else str(p or "").strip() for p in previous_platforms]
        previous_platform_keys = {p.casefold() for p in previous_platform_names if p}
        candidate_platforms = list(getattr(new_ci, "platforms", []) or getattr(new_ci, "platform_names", []) or [])
        added_platforms = []
        for item in candidate_platforms:
            name = str(item.get("platform_name") or item.get("name") or "").strip() if isinstance(item, dict) else str(item or "").strip()
            if name and name.casefold() not in previous_platform_keys:
                added_platforms.append(name)
        old_platform = str(self.ci.platform or "").strip()
        old_no       = str(self.ci.no or "").strip()
        old_type     = str(self.ci.contract_type or "").strip()

        # SD bağlantı bilgilerini koru
        if re.match(r"^SD-\d+$", str(new_ci.contract_type or "").strip().upper()):
            for attr in ("sd_anchor_start_row", "sd_anchor_end_row",
                         "sd_anchor_platform", "sd_anchor_no"):
                if not getattr(new_ci, attr, None):
                    setattr(new_ci, attr, getattr(self.ci, attr, None))

        new_ci.entry_start_row = self.original_entry_start_row
        self.ci = new_ci
        self._set_dirty()

        # Excel'e yaz
        try:
            self.sync_summary_to_system()
            self._apply_derived_statuses(self.ci, self.systems, self.deliveries)
            with self.store.batch_save():
                written_start = self.store.write_contract(
                    self.ci,
                    self.systems,
                    self.deliveries,
                    old_contract_no=old_no,
                    old_start_row=self.original_entry_start_row,
                )
                # Etiket anahtarı değişmişse taşı
                new_platform = str(self.ci.platform or "").strip()
                new_no       = str(self.ci.no or "").strip()
                new_type     = str(self.ci.contract_type or "").strip()
                actor = self.store.current_actor()
                if (self.store._normalize_label(old_type) == self.store._normalize_label("Ana Sözleşme") and
                        old_platform == new_platform and old_no != new_no):
                    self.store.update_linked_sd_contract_numbers(
                        new_platform, old_no, new_no, actor=actor
                    )
                if (old_platform, old_no, old_type) != (new_platform, new_no, new_type):
                    self.store.save_contract_tags(old_platform, old_no, old_type, [], actor=actor)
                self.store.save_contract_tags(
                    new_platform, new_no, new_type, self.contract_tags, actor=actor)
                existing_keys = {p.casefold() for p in previous_platform_names}
                existing_keys.add(new_platform.casefold())
                for platform_name in added_platforms:
                    platform_name = str(platform_name or "").strip()
                    if not platform_name or platform_name.casefold() in existing_keys or platform_name == new_platform:
                        continue
                    extra_ci = copy.copy(self.ci)
                    extra_ci.platform = platform_name
                    extra_ci.entry_start_row = 0
                    extra_ci.sd_anchor_platform = ""
                    extra_ci.sd_anchor_no = ""
                    self.store.write_contract(extra_ci, self.systems, self.deliveries)
                    self.store.save_contract_tags(platform_name, new_no, new_type, self.contract_tags, actor=actor)
                    existing_keys.add(platform_name.casefold())

            self.original_entry_start_row = int(written_start or 0)
            self.original_platform        = new_platform
            self.original_contract_no     = new_no
            self.original_contract_type   = new_type
            self.ci.entry_start_row       = self.original_entry_start_row

            self.refresh_contract_header()
            self.update_timeline_bar()
            self._initial_snapshot = self._make_data_snapshot()
            self._is_dirty = False
            self._notify_parent_contract_updated(old_platform, new_platform)
            QMessageBox.information(self, "Güncellendi", "Ana bilgiler başarıyla güncellendi.")
        except Exception as exc:
            QMessageBox.critical(self, "Hata", f"Güncelleme sırasında hata:\n{exc}")

    def _start_contract_save_worker(self, worker: ContractSaveWorker, message: str):
        if self._contract_save_thread and self._contract_save_thread.isRunning():
            return False
        self.set_busy_overlay(True, message, 0)
        thread = QThread(self)
        self._contract_save_thread = thread
        self._contract_save_worker = worker
        worker.moveToThread(thread)
        thread.started.connect(worker.run)
        worker.progress.connect(self.on_contract_save_progress)
        worker.finished.connect(self.on_contract_save_finished)
        worker.failed.connect(self.on_contract_save_failed)
        worker.finished.connect(thread.quit)
        worker.failed.connect(thread.quit)
        thread.finished.connect(worker.deleteLater)
        thread.finished.connect(thread.deleteLater)
        thread.finished.connect(self._clear_contract_save_refs)
        parent = self.parent()
        if parent and hasattr(parent, "on_contract_save_progress"):
            worker.progress.connect(parent.on_contract_save_progress)
        thread.start()
        return True

    def _clear_contract_save_refs(self):
        self._contract_save_thread = None
        self._contract_save_worker = None

    def on_contract_save_progress(self, percent: int, message: str):
        self.set_busy_overlay(True, message, percent)

    def on_contract_save_finished(self, payload: dict):
        self.set_busy_overlay(False)
        parent = self.parent()
        if parent and hasattr(parent, "set_busy_overlay"):
            parent.set_busy_overlay(False)
        action = str((payload or {}).get("action") or "")
        try:
            self.store.reload_from_disk()
            if action == "delete":
                info = dict((payload or {}).get("result") or {})
                if not info:
                    QMessageBox.warning(self, "Hata", "Sözleşme Excel'de bulunamadı veya silinemedi.")
                    return
                self._drop_deleted_context_cache(info)
                self.deleted_contract_info = info
                QMessageBox.information(self, "Silindi", "Sözleşme Excel'den silindi.")
                self.accept()
                return

            ctx = self._pending_contract_save_context or {}
            old_key = ctx.get("old_key") or ("", "", "")
            new_key = ctx.get("new_key") or ("", "", "")
            actor = ctx.get("actor") or self.store.current_actor()
            with self.store.batch_save():
                if old_key != new_key and all(old_key):
                    self.store.save_contract_tags(old_key[0], old_key[1], old_key[2], [], actor=actor)
                self.store.save_contract_tags(new_key[0], new_key[1], new_key[2], ctx.get("tags") or [], actor=actor)
            written_start = int((payload or {}).get("start_row") or 0)
            self.original_entry_start_row = written_start
            self.original_platform = new_key[0]
            self.ci.entry_start_row = self.original_entry_start_row
            self.original_contract_no = new_key[1]
            self.original_contract_type = new_key[2]
            # Yeni sözleşme kaydedildikten sonra pending belgeleri DB'ye aktar
            if getattr(self, "is_new_contract", False) and (self._pending_doc_folders or self._pending_doc_files):
                try:
                    self._flush_pending_documents_to_db()
                except Exception as flush_exc:
                    QMessageBox.warning(self, "Belge aktarma hatası",
                        f"Sözleşme kaydedildi ancak belgeler aktarılırken hata oluştu:\n{flush_exc}")
            self.is_new_contract = False
            self.refresh_sd_sidebar()
            QMessageBox.information(self, "Kaydedildi", "Sözleşme Excel'e yazıldı.")
            self._is_dirty = False
            self.accept()
        except Exception as exc:
            QMessageBox.critical(self, "Hata", f"İşlem tamamlandıktan sonra arayüz güncellenirken hata:\n{exc}")
        finally:
            self._pending_contract_save_context = None

    def on_contract_save_failed(self, message: str):
        self.set_busy_overlay(False)
        parent = self.parent()
        if parent and hasattr(parent, "set_busy_overlay"):
            parent.set_busy_overlay(False)
        self._pending_contract_save_context = None
        QMessageBox.critical(self, "Hata", f"Excel işlemi sırasında hata:\n{message}")

    def delete_contract(self):
        if not self._ensure_share_can_edit("Sözleşmeyi Sil"):
            return
        if not self.require_permission_ui("delete_contracts", "Sözleşmeyi Sil"):
            return
        no = str(self.ci.no or "").strip()
        platform = str(self.ci.platform or "").strip()
        if not no or not platform:
            QMessageBox.warning(self, "Eksik", "Silinecek sözleşme bilgisi bulunamadı.")
            return
        msg = (
            f"{platform} platformundaki '{no}' sözleşmesi silinecek.\n\n"
            "Bu işlem tüm sistemler ve teslimatlar ile birlikte Excel'den kalıcı olarak kaldırır.\n"
            "Devam etmek istiyor musunuz?"
        )
        if not ask_yes_no(self, "Sözleşmeyi Sil", msg):
            return
        worker = ContractSaveWorker(
            self.store.path,
            "delete",
            platform,
            no,
            start_row=self.original_entry_start_row if self.original_entry_start_row > 0 else 0,
            actor=self.store.current_actor(),
            store=self.store,
        )
        self._start_contract_save_worker(worker, "Sözleşme siliniyor...")

    def _tag_key(self, name: str) -> str:
        return normalized_tag_key(name)

    def build_side_meta_popover_bar(self, parent_width: int):
        """Build the compact meta bar and its layout-independent floating popover."""
        self._side_meta_open_panel = None
        self._side_meta_last_panel = "files"
        self._side_meta_manual_height = None
        self._side_meta_files: List[dict] = []
        self.side_meta_bar = QFrame()
        self.side_meta_bar.setObjectName("sideMetaBar")
        self.side_meta_bar.setAutoFillBackground(False)
        self.side_meta_bar.setAttribute(Qt.WA_TranslucentBackground, True)
        self.side_meta_bar.setStyleSheet("QFrame#sideMetaBar { background:transparent; border:0; }")
        bar_layout = QHBoxLayout(self.side_meta_bar)
        bar_layout.setContentsMargins(0, 0, 0, 0)
        bar_layout.setSpacing(10)

        self.side_btn_tags = BadgeTabButton("🏷", "Etiketler", 0)
        self.side_btn_files = BadgeTabButton("📎", "Belgeler", 0)
        self.side_btn_share = BadgeTabButton("↗", "Paylaşım", None)
        self.side_btn_share.badge.hide()

        tabs_controller = getattr(self, "contract_action_tabs", None)
        for panel, button in (("tags", self.side_btn_tags), ("files", self.side_btn_files), ("share", self.side_btn_share)):
            if panel == "tags" and tabs_controller is not None:
                button.clicked.connect(lambda _checked=False, ctl=tabs_controller: ctl.open_tags())
            elif panel == "files" and tabs_controller is not None:
                button.clicked.connect(lambda _checked=False, ctl=tabs_controller: ctl.open_files())
            elif panel == "share" and tabs_controller is not None:
                button.clicked.connect(lambda _checked=False, ctl=tabs_controller: ctl.open_share())
            else:
                button.clicked.connect(lambda _checked=False, name=panel: self.toggle_side_meta_popover(name))
            bar_layout.addWidget(button, 0, Qt.AlignVCenter)

        bar_layout.addStretch(1)

        # Eski sağdaki küçük chevron hücresi görsel olarak boş/işlevsiz bir kutu gibi
        # duruyordu. Popover aç/kapatma zaten sekme butonlarından yapıldığı için layout'a
        # ekstra buton eklemiyoruz.
        self.side_chevron = None

        self.side_meta_popover = QFrame(self)
        self.side_meta_popover.setObjectName("sideMetaPopover")
        self.side_meta_popover.setStyleSheet(
            "QFrame#sideMetaPopover{background:#ffffff; border:1px solid #c8d9ed; border-radius:12px;}"
            "QPushButton#sidePanelAdd{background:#2563eb; color:#ffffff; border:0; border-radius:8px; font-size:20px; font-weight:900; padding:0;}"
            "QPushButton#sidePanelAdd:hover{background:#1d4ed8;}"
            "QPushButton#documentActionPrimary{background:#2563eb; color:#ffffff; border:1px solid #2563eb; border-radius:9px; padding:0 12px; font-size:11px; font-weight:700; min-height:28px; max-height:32px;}"
            "QPushButton#documentActionPrimary:hover{background:#1d4ed8; border-color:#1d4ed8;}"
            "QPushButton#documentAction{background:#f8fbff; color:#102a43; border:1px solid #cfe0f3; border-radius:9px; padding:0 12px; font-size:11px; font-weight:700; min-height:28px; max-height:32px;}"
            "QPushButton#documentAction:hover{background:#edf5ff; border-color:#9ec5f8;}"
            "QPushButton#documentLockButton{background:#f8fbff; color:#0f172a; border:1px solid #cfe0f3; border-radius:9px; padding:0; font-size:15px; font-weight:700; min-width:30px; max-width:30px; min-height:28px; max-height:30px;}"
            "QPushButton#documentLockButton:hover{background:#edf5ff; border-color:#9ec5f8;}"
            "QPushButton#documentLockButton[locked=\"true\"]{background:#fff7ed; color:#9a3412; border-color:#fed7aa;}"
            "QLabel#documentLockInfo{background:#fff7ed; color:#9a3412; border:1px solid #fed7aa; border-radius:8px; padding:4px 8px; font-size:10px; font-weight:700;}"
            "QLabel#documentHint{background:transparent; color:#94a3b8; border:0; font-size:10px; padding:0;}"
            "QFrame#docDropZone{background:#f8fbff; border:1px dashed #c7d7ea; border-radius:8px;}"
            "QFrame#docDropZone[dragOver=\"true\"]{background:#eef6ff; border-color:#2563eb;}"
            "QLabel#sidePanelEmpty{background:transparent; color:#7b8da5; border:0; font-size:12px;}"
            "QLabel#sidePanelEmptySub{background:transparent; color:#94a3b8; border:0; font-size:10px;}"
            "QFrame#docFooterBar{background:#f8fbff; border-top:1px solid #e6eef8; border-bottom-left-radius:10px; border-bottom-right-radius:10px;}"
            "QLabel#documentsFooterLeft{background:transparent; color:#94a3b8; border:0; font-size:10px;}"
            "QLabel#documentsFooterRight{background:transparent; color:#334155; border:0; font-size:10px; font-weight:700;}"
            "QLabel#sidePanelEmptyTag{background:#f8fbff; color:#64748b; border:1px dashed #c7d6e8; border-radius:8px; padding:13px; font-size:12px;}"
            "QPushButton#sidePanelAddInline{background:#eef4ff; color:#1d4ed8; border:1px solid #bcd1f2; border-radius:8px; padding:2px 10px; font-size:11px; font-weight:700;}"
            "QPushButton#sidePanelAddInline:hover{background:#dbeafe;}"
        )
        shadow = QGraphicsDropShadowEffect(self.side_meta_popover)
        shadow.setBlurRadius(18)
        shadow.setOffset(0, 6)
        shadow.setColor(QColor(15, 45, 74, 45))
        self.side_meta_popover.setGraphicsEffect(shadow)
        popover_layout = QVBoxLayout(self.side_meta_popover)
        popover_layout.setContentsMargins(12, 6, 12, 4)
        popover_layout.setSpacing(6)
        self.side_meta_arrow = QLabel("▲")
        self.side_meta_arrow.setObjectName("sideMetaArrow")
        self.side_meta_arrow.setStyleSheet("QLabel#sideMetaArrow{background:transparent;color:#c8d9ed;border:0;font-size:14px;margin:0;padding:0;}")
        popover_layout.addWidget(self.side_meta_arrow, 0, Qt.AlignLeft)
        self.side_meta_popover_body = QWidget()
        self.side_meta_popover_body.setStyleSheet("background:transparent;")
        self.side_meta_popover_body_layout = QVBoxLayout(self.side_meta_popover_body)
        self.side_meta_popover_body_layout.setContentsMargins(0, 0, 0, 0)
        self.side_meta_popover_body_layout.setSpacing(6)
        from PySide6.QtWidgets import QSizePolicy as _QSP
        self.side_meta_popover_body.setSizePolicy(_QSP.Expanding, _QSP.Minimum)
        popover_layout.addWidget(self.side_meta_popover_body, 0)
        popover_layout.addStretch(1)

        _rh = QFrame()
        _rh.setFixedHeight(7)
        _rh.setCursor(Qt.SizeVerCursor)
        _rh.setObjectName("popoverResizeHandle")
        _rh.setStyleSheet(
            "QFrame#popoverResizeHandle{background:transparent;border:0;border-top:2px solid transparent;}"
            "QFrame#popoverResizeHandle:hover{background:#e8f0fe;border-top:2px solid #93c5fd;border-radius:3px;}"
        )
        _rh.setToolTip("Paneli yeniden boyutlandır")
        _rh_drag = [None, None]

        def _rh_press(event, _s=self):
            if event.button() == Qt.LeftButton:
                try:
                    _rh_drag[0] = event.globalPosition().toPoint().y()
                except AttributeError:
                    _rh_drag[0] = event.globalPos().y()
                _rh_drag[1] = _s.side_meta_popover.height()

        def _rh_move(event, _s=self):
            if _rh_drag[0] is None:
                return
            try:
                cur_y = event.globalPosition().toPoint().y()
            except AttributeError:
                cur_y = event.globalPos().y()
            delta = cur_y - _rh_drag[0]
            new_h = max(190, _rh_drag[1] + delta)
            screen = _s.screen()
            avail_h = screen.availableGeometry().height() if screen else 900
            new_h = min(new_h, avail_h - 120)
            _s._side_meta_manual_height = new_h
            _s.position_side_meta_popover()

        def _rh_release(event):
            _rh_drag[0] = None
            _rh_drag[1] = None

        _rh.mousePressEvent = _rh_press
        _rh.mouseMoveEvent = _rh_move
        _rh.mouseReleaseEvent = _rh_release
        popover_layout.addWidget(_rh, 0)

        self.side_meta_popover.hide()
        QApplication.instance().installEventFilter(self)
        self.position_side_meta_popover()

    def _place_tab_bar(self):
        """Sekme barını Durum hücresinin altına, header altından sarkacak şekilde konumlandır."""
        header = getattr(self, "contract_header", None)
        bar = getattr(self, "side_meta_bar", None)
        host = getattr(self, "side_meta_host", None)
        if not header or not bar or not host:
            return
        try:
            bar_h = 46

            # Header'ın dialog içindeki kesin konumu
            header_global = header.mapTo(self, QPoint(0, 0))
            hy = header_global.y()
            hh = header.height()

            # X: Durum hücresinin (metaCell) tam sol kenarı
            status_val = (getattr(self, "meta_values", {}) or {}).get("status")
            sx = None
            if status_val is not None:
                # status_val → headerStatusWrap → metaCell zincirini çık
                w = status_val
                for _ in range(4):  # en fazla 4 seviye üste çık
                    if w is None:
                        break
                    if w.objectName() == "metaCell":
                        sx = w.mapTo(self, QPoint(0, 0)).x()
                        break
                    w = w.parent()

            if sx is None:
                # Fallback: "Ana Bilgileri Düzenle" butonunun %60 solundan başla
                edit_btn = getattr(self, "header_edit_btn", None)
                sx = int(self.width() * 0.55) if edit_btn is None else max(
                    int(self.width() * 0.48),
                    edit_btn.mapTo(self, QPoint(0, 0)).x() - int(self.width() * 0.32)
                )

            # Sağ sınır: "Ana Bilgileri Düzenle" butonunun tam sol kenarı - 10px boşluk
            edit_btn = getattr(self, "header_edit_btn", None)
            if edit_btn is not None:
                right_edge = edit_btn.mapTo(self, QPoint(0, 0)).x() - 10
            else:
                right_edge = self.width() - 16

            available_w = max(280, self.width() - 24)
            desired_w = min(available_w, max(470, right_edge - sx))
            x = max(12, min(sx, right_edge - desired_w))
            host_w = desired_w

            # Y: header'ın tam altından başlasın — biraz içeri girmez
            y = hy + hh - 2  # 2px header'a yapışık

            host.setGeometry(x, y, host_w, bar_h)
            bar.setFixedWidth(host_w)
            bar.setFixedHeight(bar_h)
            host.raise_()
        except Exception:
            import traceback; traceback.print_exc()

    def resizeEvent(self, event):
        super().resizeEvent(event)
        self.position_busy_overlay()
        QTimer.singleShot(30, self._place_tab_bar)

    def showEvent(self, event):
        super().showEvent(event)
        QTimer.singleShot(100, self._place_tab_bar)

    def position_side_meta_popover(self):
        if not hasattr(self, "side_meta_popover") or not hasattr(self, "side_meta_host"):
            return
        host = self.side_meta_host
        if getattr(self, "_side_meta_open_panel", None) == "share":
            w = 360
        else:
            w = max(320, min(470, host.width() if host.width() > 0 else 380))
        self.side_meta_popover.setFixedWidth(w)
        self.side_meta_popover.adjustSize()
        hint_h = self.side_meta_popover.sizeHint().height()
        # Popover self içinde: host konumundan itibaren bar altına yap
        bar_h = self.side_meta_bar.height() if hasattr(self, "side_meta_bar") else 38
        pop_x = host.x() + host.width() - w
        pop_y = host.y() + bar_h + 8
        pop_x = max(8, min(pop_x, max(8, self.width() - w - 8)))
        # Ekranın kullanılabilir yüksekliğini hesaba kat
        screen = self.screen()
        if screen:
            avail_h = screen.availableGeometry().height()
        else:
            avail_h = 900
        max_h = max(190, min(avail_h - 120, self.height() - pop_y - 8))
        # Manuel yükseklik kullanıcı tarafından ayarlandıysa öncelikli kullan
        manual_h = getattr(self, "_side_meta_manual_height", None)
        if getattr(self, "_side_meta_open_panel", None) == "files":
            files = list(getattr(self, "_side_meta_files", []))
            folders = list(getattr(self, "_side_meta_folders", []))
            tree_h = self.document_tree_height(folders, files)
            auto_min_h = max(190, tree_h + 94)
            if manual_h is not None:
                min_h = max(190, min(manual_h, max_h))
            else:
                min_h = auto_min_h
        elif getattr(self, "_side_meta_open_panel", None) == "share":
            min_h = 250
            manual_h = None
        else:
            min_h = 110
            manual_h = None
        h = max(min_h, min(hint_h if manual_h is None else manual_h, max_h))
        self.side_meta_popover.setGeometry(pop_x, pop_y, w, h)
        self._position_side_meta_arrow()
        if self.side_meta_popover.isVisible():
            self.side_meta_popover.raise_()

    def _position_side_meta_arrow(self):
        arrow = getattr(self, "side_meta_arrow", None)
        if arrow is None:
            return
        panel = getattr(self, "_side_meta_open_panel", None)
        button = {
            "tags": getattr(self, "side_btn_tags", None),
            "files": getattr(self, "side_btn_files", None),
            "share": getattr(self, "side_btn_share", None),
        }.get(panel)
        if button is None:
            arrow.setStyleSheet("QLabel#sideMetaArrow{background:transparent;color:#c8d9ed;border:0;font-size:14px;margin-left:12px;padding:0;}")
            return
        try:
            center = button.mapTo(self, QPoint(button.width() // 2, 0)).x() - self.side_meta_popover.x()
        except Exception:
            center = 18
        left = max(8, min(center - 7, max(8, self.side_meta_popover.width() - 24)))
        arrow.setStyleSheet(f"QLabel#sideMetaArrow{{background:transparent;color:#c8d9ed;border:0;font-size:14px;margin-left:{left}px;padding:0;}}")

    def _toggle_side_meta_chevron(self):
        if self._side_meta_open_panel:
            self.close_side_meta_popover()
        else:
            self.toggle_side_meta_popover(self._side_meta_last_panel or "files")

    def toggle_side_meta_popover(self, panel: str):
        if panel not in {"tags", "files", "share"}:
            return
        if self._side_meta_open_panel == panel and self.side_meta_popover.isVisible():
            self.close_side_meta_popover()
            return
        self._side_meta_open_panel = panel
        self._side_meta_last_panel = panel
        self.update_side_meta_badges()
        self.render_side_meta_popover_content(panel)
        self._sync_side_meta_controls()
        self.side_meta_popover.show()
        self.side_meta_popover.raise_()
        QApplication.processEvents()
        self.position_side_meta_popover()
        # Kısa opacity fade — konum/boyut değişmez
        _eff = QGraphicsOpacityEffect(self.side_meta_popover)
        self.side_meta_popover.setGraphicsEffect(_eff)
        _anim = QPropertyAnimation(_eff, b"opacity", self.side_meta_popover)
        _anim.setDuration(130)
        _anim.setStartValue(0.0)
        _anim.setEndValue(1.0)
        _anim.finished.connect(lambda: self.side_meta_popover.setGraphicsEffect(None))
        _anim.start(QPropertyAnimation.DeleteWhenStopped)

    def close_side_meta_popover(self):
        self._side_meta_open_panel = None
        if hasattr(self, "side_meta_popover"):
            self.side_meta_popover.hide()
        self._sync_side_meta_controls()

    def _sync_side_meta_controls(self):
        panel = self._side_meta_open_panel
        for name, button in (("tags", self.side_btn_tags), ("files", self.side_btn_files), ("share", self.side_btn_share)):
            button.setChecked(name == panel)
        chevron = getattr(self, "side_chevron", None)
        if chevron is not None:
            chevron.setText("∧" if panel else "∨")

    def _load_contract_files(self) -> List[dict]:
        if self.is_new_contract:
            # Pending modda: in-memory listeden oku
            folders_by_id = {f["id"]: f for f in self._pending_doc_folders}
            out = []
            for f in self._pending_doc_files:
                item = dict(f)
                fid = f.get("folder_id")
                folder = folders_by_id.get(fid) if fid else None
                item["folder_path"] = folder.get("name", "") if folder else ""
                out.append(item)
            return out
        try:
            return list(self.store.list_contract_files(self.ci.platform, self.ci.no, self.ci.contract_type))
        except Exception:
            return []

    def _load_contract_file_folders(self) -> List[dict]:
        if self.is_new_contract:
            return list(self._pending_doc_folders)
        try:
            return list(self.store.list_contract_file_folders(self.ci.platform, self.ci.no, self.ci.contract_type))
        except Exception:
            return []

    def _set_side_meta_badge_counts(self, tag_count: int, file_count: int):
        self.side_btn_tags.setCount(tag_count)
        self.side_btn_files.setCount(file_count)

    def update_side_meta_badges(self):
        self._side_meta_files = self._load_contract_files()
        self._side_meta_folders = self._load_contract_file_folders()
        self._set_side_meta_badge_counts(len(self._ordered_contract_tags()), len(self._side_meta_files))

    def _clear_side_meta_popover_body(self):
        layout = self.side_meta_popover_body_layout
        while layout.count():
            item = layout.takeAt(0)
            widget = item.widget()
            child_layout = item.layout()
            if widget is not None:
                widget.deleteLater()
            elif child_layout is not None:
                while child_layout.count():
                    child = child_layout.takeAt(0)
                    if child.widget() is not None:
                        child.widget().deleteLater()

    def _document_db_conn(self):
        return getattr(getattr(self.store, "db", None), "conn", None)

    def _document_db_contract_id(self) -> int | None:
        """Açık sözleşmenin DB contract_id'sini döner. STS dışında None."""
        if not hasattr(self.store, "_resolve_contract_id"):
            return None
        ci = getattr(self, "ci", None)
        if ci is None:
            return None
        return self.store._resolve_contract_id(
            str(ci.platform or ""),
            str(ci.no or ""),
            str(ci.contract_type or "Ana Sözleşme"),
        )

    def _default_document_lock_state(self) -> dict:
        return {
            "contract_id": None,
            "is_locked": 0,
            "locked_by_staff_id": None,
            "locked_by_device_name": None,
            "locked_by_full_name": None,
            "locked_at": None,
            "updated_at": None,
        }

    def _load_document_lock_state(self) -> dict:
        try:
            conn = self._document_db_conn()
            if conn is None:
                self._document_lock_state = self._default_document_lock_state()
                return dict(self._document_lock_state)
            if hasattr(self.store, "document_lock_state"):
                ci = getattr(self, "ci", None)
                if ci is not None:
                    self._document_lock_state = self.store.document_lock_state(
                        str(ci.platform or ""),
                        str(ci.no or ""),
                        str(ci.contract_type or "Ana Sözleşme"),
                    )
                    return dict(self._document_lock_state)
            cid = self._document_db_contract_id()
            if cid is None:
                self._document_lock_state = self._default_document_lock_state()
                return dict(self._document_lock_state)
            getter = getattr(auth, "get_document_lock_state", None)
            if not callable(getter):
                self._document_lock_state = self._default_document_lock_state()
                return dict(self._document_lock_state)
            self._document_lock_state = getter(conn, cid)
            return dict(self._document_lock_state)
        except Exception:
            traceback.print_exc()
            # Belgeler paneli, lock state okunamadığında beyaz kalmasın;
            # güvenli varsayılan olarak kilit açık kabul edilir.
            self._document_lock_state = self._default_document_lock_state()
            return dict(self._document_lock_state)

    def _current_document_lock_state(self) -> dict:
        return dict(getattr(self, "_document_lock_state", {}) or self._load_document_lock_state())

    def _documents_access_allowed(self, lock_state: Optional[dict] = None) -> bool:
        state = lock_state or self._current_document_lock_state()
        checker = getattr(auth, "can_current_staff_access_documents", None)
        if callable(checker):
            return checker(state, self.current_staff)
        if int((state or {}).get("is_locked") or 0) == 0:
            return True
        return bool(
            self.current_staff
            and str((self.current_staff or {}).get("device_name") or "") == str((state or {}).get("locked_by_device_name") or "")
        )

    def _document_lock_same_device(self, lock_state: Optional[dict] = None) -> bool:
        state = lock_state or self._current_document_lock_state()
        return bool(
            int(state.get("is_locked") or 0) == 1
            and self.current_staff
            and str((self.current_staff or {}).get("device_name") or "") == str(state.get("locked_by_device_name") or "")
        )

    def _document_unlock_with_password(self, lock_state: Optional[dict] = None) -> bool:
        conn = self._document_db_conn()
        if conn is None:
            return False
        unlock_dialog = getattr(auth, "require_document_unlock_password", None)
        if not callable(unlock_dialog):
            QMessageBox.warning(
                self,
                "Belgeler Kilitli",
                "Belge kilidi doğrulama fonksiyonu yüklenemedi. Lütfen uygulamayı güncel kodla yeniden başlatın.",
            )
            return False
        if unlock_dialog(self, conn, lock_state or self._current_document_lock_state()):
            self._load_document_lock_state()
            self.render_contract_files()
            return True
        return False

    def _ensure_document_access(self, interactive: bool = True) -> bool:
        state = self._load_document_lock_state()
        if self._documents_access_allowed(state):
            return True
        if interactive:
            return self._document_unlock_with_password(state)
        return False

    def _animate_document_lock_button(self):
        btn = getattr(self, "document_lock_btn", None)
        if not btn:
            return
        rect = btn.geometry()
        grow = rect.adjusted(-2, -2, 2, 2)
        anim = QPropertyAnimation(btn, b"geometry", btn)
        anim.setDuration(160)
        anim.setStartValue(rect)
        anim.setKeyValueAt(0.5, grow)
        anim.setEndValue(rect)
        anim.start()
        self._document_lock_anim = anim

    def _toggle_document_lock(self):
        if not self._ensure_share_can_edit("Belge Kilidi"):
            return
        state = self._load_document_lock_state()
        if int(state.get("is_locked") or 0) == 0:
            if not self.require_permission_ui("lock_documents", "Belge Kilitleme"):
                return
        elif self._document_lock_same_device(state):
            if not self.require_permission_ui("unlock_own_documents", "Belge Kilidi"):
                return
        elif not self.require_permission_ui("unlock_all_documents", "Belge Kilidi"):
            return
        conn = self._document_db_conn()
        if conn is None:
            QMessageBox.information(self, "Belgeler", "Belge kilidi yalnızca STS veri dosyalarında desteklenir.")
            return
        state = self._load_document_lock_state()
        if int(state.get("is_locked") or 0) == 0:
            if not self.current_staff:
                QMessageBox.warning(self, "Personel gerekli", "Belgeleri kilitlemek için personel girişi gereklidir.")
                return
            if hasattr(self.store, "lock_documents"):
                ci = getattr(self, "ci", None)
                self._document_lock_state = self.store.lock_documents(
                    str(ci.platform or "") if ci else "",
                    str(ci.no or "") if ci else "",
                    self.current_staff,
                    str(ci.contract_type or "Ana Sözleşme") if ci else "Ana Sözleşme",
                )
            else:
                lock_fn = getattr(auth, "lock_documents", None)
                if not callable(lock_fn):
                    QMessageBox.warning(self, "Belgeler", "Belge kilidi fonksiyonu yüklenemedi. Lütfen uygulamayı güncel kodla yeniden başlatın.")
                    return
                cid = self._document_db_contract_id()
                if cid is None:
                    QMessageBox.warning(self, "Belgeler", "Sözleşme ID'si bulunamadı.")
                    return
                self._document_lock_state = lock_fn(conn, cid, self.current_staff)
            self._animate_document_lock_button()
            self.render_contract_files()
            return
        if self._document_lock_same_device(state):
            actor = str((self.current_staff or {}).get("full_name") or "Personel")
            if hasattr(self.store, "unlock_documents"):
                ci = getattr(self, "ci", None)
                self._document_lock_state = self.store.unlock_documents(
                    str(ci.platform or "") if ci else "",
                    str(ci.no or "") if ci else "",
                    actor=actor,
                    contract_type=str(ci.contract_type or "Ana Sözleşme") if ci else "Ana Sözleşme",
                )
            else:
                unlock_fn = getattr(auth, "unlock_documents", None)
                if not callable(unlock_fn):
                    QMessageBox.warning(self, "Belgeler", "Belge kilidi açma fonksiyonu yüklenemedi. Lütfen uygulamayı güncel kodla yeniden başlatın.")
                    return
                cid = self._document_db_contract_id()
                if cid is None:
                    return
                self._document_lock_state = unlock_fn(conn, cid)
            self._animate_document_lock_button()
            self.render_contract_files()
            return
        if self._document_unlock_with_password(state):
            self._animate_document_lock_button()

    def _document_lock_summary_text(self, lock_state: Optional[dict] = None) -> str:
        state = lock_state or self._current_document_lock_state()
        if int(state.get("is_locked") or 0) == 0:
            return ""
        if self._document_lock_same_device(state):
            return "🔒 Belgeler bu cihaz tarafından kilitlendi."
        locked_by = str(state.get("locked_by_full_name") or "Personel")
        return f"🔒 Belgeler kilitli · Kilitleyen: {locked_by}"

    def render_side_meta_popover_content(self, panel: str):
        self._clear_side_meta_popover_body()
        body = self.side_meta_popover_body_layout
        if panel == "share":
            body.addWidget(ContractSharePopover(self, self.side_meta_popover_body), 0)
            return
        if panel == "tags":
            # + butonu scroll'dan önce değil, kart listesinin en üstünde kompakt satır
            add_row = QHBoxLayout(); add_row.setContentsMargins(0, 0, 0, 2); add_row.addStretch(1)
            add_btn = QPushButton("+ Etiket Ekle"); add_btn.setObjectName("sidePanelAddInline")
            add_btn.setFixedHeight(26); add_btn.setEnabled(not self._share_is_view_only()); add_btn.clicked.connect(self.open_tag_assign_dialog)
            add_row.addWidget(add_btn); body.addLayout(add_row)
            scroll, cards = self._make_card_scroll(); body.addWidget(scroll, 1)
            ordered = self._ordered_contract_tags()
            if ordered:
                for tag in ordered:
                    cards.insertWidget(cards.count() - 1, self.create_tag_card(tag))
            else:
                empty = QLabel("Henüz etiket atanmadı."); empty.setObjectName("sidePanelEmptyTag"); empty.setAlignment(Qt.AlignCenter); cards.insertWidget(0, empty)
        else:
            lock_state = self._load_document_lock_state()
            documents_accessible = self._documents_access_allowed(lock_state)
            documents_locked = int(lock_state.get("is_locked") or 0) == 1
            # ── Toolbar: + Dosya Ekle | + Klasör Ekle | Kilit ───────────────
            toolbar = QHBoxLayout()
            toolbar.setContentsMargins(0, 0, 0, 0)
            toolbar.setSpacing(6)
            btn_file = QPushButton("+ Dosya Ekle")
            btn_file.setObjectName("documentActionPrimary")
            btn_file.setFixedHeight(30)
            btn_file.setCursor(Qt.PointingHandCursor)
            btn_file.setStyleSheet(
                "QPushButton{background:#2563eb;color:#ffffff;border:1px solid #2563eb;"
                "border-radius:9px;padding:0 12px;font-size:11px;font-weight:800;}"
                "QPushButton:hover{background:#1d4ed8;border-color:#1d4ed8;}"
            )
            btn_file.clicked.connect(self._pick_contract_files)
            btn_folder = QPushButton("+ Klasör Ekle")
            btn_folder.setObjectName("documentAction")
            btn_folder.setFixedHeight(30)
            btn_folder.setMinimumWidth(100)
            btn_folder.setCursor(Qt.PointingHandCursor)
            btn_folder.setStyleSheet(
                "QPushButton{background:#f8fbff;color:#102a43;border:1px solid #cfe0f3;"
                "border-radius:9px;padding:0 14px;font-size:11px;font-weight:800;}"
                "QPushButton:hover{background:#edf5ff;border-color:#9ec5f8;}"
            )
            btn_folder.clicked.connect(self.add_contract_file_folder)
            btn_file.setEnabled(documents_accessible and not self._share_is_view_only())
            btn_folder.setEnabled(documents_accessible and not self._share_is_view_only())

            lock_btn = QPushButton("🔒" if documents_locked else "🔓")
            lock_btn.setObjectName("documentLockButton")
            lock_btn.setProperty("locked", "true" if documents_locked else "false")
            lock_btn.setFixedSize(30, 30)
            lock_btn.setCursor(Qt.PointingHandCursor)
            lock_btn.setToolTip("Belgeler kilitli" if documents_locked else "Belgeleri kilitle")
            lock_btn.setEnabled(not self._share_is_view_only())
            lock_btn.clicked.connect(self._toggle_document_lock)
            self.document_lock_btn = lock_btn

            toolbar.addWidget(btn_file)
            toolbar.addWidget(btn_folder)
            toolbar.addWidget(lock_btn)
            toolbar.addStretch(1)
            body.addLayout(toolbar)

            lock_text = self._document_lock_summary_text(lock_state)
            if lock_text:
                lock_info = QLabel(lock_text)
                lock_info.setObjectName("documentLockInfo")
                lock_info.setWordWrap(True)
                body.addWidget(lock_info, 0)

            files = list(self._side_meta_files)
            folders = list(getattr(self, "_side_meta_folders", []))

            # ── Drop zone frame containing tree (or empty state) ─────────────
            drop_frame = QFrame()
            drop_frame.setObjectName("docDropZone")
            drop_frame.setAcceptDrops(True)

            # forward drag-drop events from frame to tree
            def _frame_drag_enter(event):
                if not documents_accessible:
                    event.ignore()
                    return
                if event.mimeData().hasUrls():
                    drop_frame.setProperty("dragOver", "true")
                    drop_frame.style().unpolish(drop_frame)
                    drop_frame.style().polish(drop_frame)
                    event.acceptProposedAction()
                else:
                    event.ignore()

            def _frame_drag_leave(event):
                drop_frame.setProperty("dragOver", "false")
                drop_frame.style().unpolish(drop_frame)
                drop_frame.style().polish(drop_frame)
                super(QFrame, drop_frame).dragLeaveEvent(event)

            def _frame_drop(event):
                drop_frame.setProperty("dragOver", "false")
                drop_frame.style().unpolish(drop_frame)
                drop_frame.style().polish(drop_frame)
                if not documents_accessible:
                    self._ensure_document_access(interactive=True)
                    event.ignore()
                    return
                tree_w = getattr(self, "contract_files_tree", None)
                if tree_w:
                    tree_w.dropEvent(event)
                else:
                    event.ignore()

            drop_frame.dragEnterEvent = _frame_drag_enter
            drop_frame.dragLeaveEvent = _frame_drag_leave
            drop_frame.dropEvent = _frame_drop
            drop_frame.dragMoveEvent = lambda e: (e.acceptProposedAction() if documents_accessible and e.mimeData().hasUrls() else e.ignore())
            if not documents_accessible:
                drop_frame.mousePressEvent = lambda event: self._ensure_document_access(interactive=True)


            drop_frame_layout = QVBoxLayout(drop_frame)
            drop_frame_layout.setContentsMargins(0, 0, 0, 0)
            drop_frame_layout.setSpacing(0)

            try:
                if not documents_accessible:
                    self.contract_files_tree = None
                    locked_host = QWidget()
                    locked_host.setStyleSheet("background:transparent;")
                    lv = QVBoxLayout(locked_host)
                    lv.setContentsMargins(12, 28, 12, 28)
                    lv.setSpacing(6)
                    lv.setAlignment(Qt.AlignCenter)
                    l1 = QLabel("🔒 Belgeler kilitli")
                    l1.setObjectName("sidePanelEmpty")
                    l1.setAlignment(Qt.AlignCenter)
                    l2 = QLabel(f"Kilitleyen: {str(lock_state.get('locked_by_full_name') or 'Personel')}")
                    l2.setObjectName("sidePanelEmptySub")
                    l2.setAlignment(Qt.AlignCenter)
                    l3 = QLabel("Erişmek için tıklayın ve kilitleyen personelin şifresini girin.")
                    l3.setObjectName("sidePanelEmptySub")
                    l3.setAlignment(Qt.AlignCenter)
                    l3.setWordWrap(True)
                    lv.addWidget(l1)
                    lv.addWidget(l2)
                    lv.addWidget(l3)
                    drop_frame_layout.addWidget(locked_host)
                else:
                    tree = self.create_contract_files_tree(folders, files)
                    self.contract_files_tree = tree
                    tree_h = self.document_tree_height(folders, files)
                    # Tree'yi direkt ekle, yükseklik min+max ile sınırla
                    # Max = aynı değer → sabit kutu, ama QTreeWidget kendi
                    # internal scroll'u zaten yönetiyor (editör de içinde kalır)
                    tree.setMinimumHeight(tree_h)
                    tree.setMaximumHeight(tree_h)
                    drop_frame_layout.addWidget(tree)

                    if not files and not folders:
                        # empty state overlay
                        empty_host = QWidget()
                        empty_host.setStyleSheet("background:transparent;")
                        ev = QVBoxLayout(empty_host)
                        ev.setContentsMargins(12, 20, 12, 20)
                        ev.setSpacing(4)
                        ev.setAlignment(Qt.AlignCenter)
                        e1 = QLabel("Henüz belge eklenmedi.")
                        e1.setObjectName("sidePanelEmpty")
                        e1.setAlignment(Qt.AlignCenter)
                        e2 = QLabel("Dosya ekleyin veya buraya sürükleyin.")
                        e2.setObjectName("sidePanelEmptySub")
                        e2.setAlignment(Qt.AlignCenter)
                        ev.addWidget(e1)
                        ev.addWidget(e2)
                        drop_frame_layout.addWidget(empty_host)

            except Exception as exc:
                self.contract_files_tree = None
                message = QLabel(f"Belgeler yüklenemedi: {exc}")
                message.setObjectName("sidePanelEmpty")
                message.setWordWrap(True)
                message.setAlignment(Qt.AlignCenter)
                drop_frame_layout.addWidget(message)

            body.addWidget(drop_frame, 1)

            # ── Footer ──────────────────────────────────────────────────────
            footer_bar = QFrame()
            footer_bar.setObjectName("docFooterBar")
            footer_vlay = QVBoxLayout(footer_bar)
            footer_vlay.setContentsMargins(8, 6, 8, 6)
            footer_vlay.setSpacing(4)

            # ── Satır 1: Seçim bilgisi + butonlar (seçim varken) ────────────
            sel_row = QWidget()
            sel_row.setStyleSheet("background:transparent;")
            sel_row_lay = QHBoxLayout(sel_row)
            sel_row_lay.setContentsMargins(0, 0, 0, 0)
            sel_row_lay.setSpacing(6)

            lbl_sel = QLabel("")
            lbl_sel.setStyleSheet("background:transparent;color:#1d4ed8;font-size:11px;font-weight:700;border:0;")

            btn_bulk_dl = QPushButton("⬇  İndir")
            btn_bulk_zip = QPushButton("  ZIP İndir")
            btn_bulk_zip.setIcon(self._make_file_icon("zip"))
            btn_bulk_zip.setIconSize(__import__('PySide6.QtCore', fromlist=['QSize']).QSize(16, 16))
            for _b in (btn_bulk_dl, btn_bulk_zip):
                _b.setFixedHeight(26)
                _b.setMinimumWidth(80)
                _b.setStyleSheet(
                    "QPushButton{background:#2563eb;color:#fff;border:0;"
                    "border-radius:7px;padding:0 12px;font-size:11px;font-weight:700;}"
                    "QPushButton:hover{background:#1d4ed8;}"
                )

            sel_row_lay.addWidget(lbl_sel, 1)
            sel_row_lay.addWidget(btn_bulk_dl)
            sel_row_lay.addWidget(btn_bulk_zip)
            sel_row.hide()
            footer_vlay.addWidget(sel_row)

            # ── Satır 2: Sürükle-bırak + Toplam boyut (her zaman) ───────────
            info_row = QWidget()
            info_row.setStyleSheet("background:transparent;")
            info_row_lay = QHBoxLayout(info_row)
            info_row_lay.setContentsMargins(0, 0, 0, 0)
            info_row_lay.setSpacing(0)

            lbl_drag = QLabel("⇅  Sürükle-bırak desteklenir")
            lbl_drag.setObjectName("documentsFooterLeft")

            total_bytes = sum(int(item.get("size_bytes", 0) or 0) for item in files)
            lbl_right = QLabel(f"Toplam {self.format_file_size(total_bytes)}")
            lbl_right.setObjectName("documentsFooterRight")

            info_row_lay.addWidget(lbl_drag)
            info_row_lay.addStretch(1)
            info_row_lay.addWidget(lbl_right)
            footer_vlay.addWidget(info_row)

            body.addWidget(footer_bar, 0)

            # Seçim değişince footer güncelle
            def _on_selection_changed():
                tree_w = getattr(self, "contract_files_tree", None)
                if not tree_w:
                    return
                sel_files = [
                    it for it in tree_w.selectedItems()
                    if it.data(0, Qt.UserRole) == "file"
                ]
                if sel_files:
                    lbl_sel.setText(f"{len(sel_files)} dosya seçildi")
                    sel_row.show()
                else:
                    sel_row.hide()

            if hasattr(self, "contract_files_tree") and self.contract_files_tree:
                self.contract_files_tree.itemSelectionChanged.connect(_on_selection_changed)

            def _get_selected_file_ids():
                tree_w = getattr(self, "contract_files_tree", None)
                if not tree_w:
                    return []
                return [
                    int(it.data(0, Qt.UserRole + 1))
                    for it in tree_w.selectedItems()
                    if it.data(0, Qt.UserRole) == "file"
                ]

            btn_bulk_dl.clicked.connect(lambda: self._bulk_download_files(_get_selected_file_ids()))
            btn_bulk_zip.clicked.connect(lambda: self._bulk_zip_files(_get_selected_file_ids()))

    def _contract_document_share_stats(self) -> tuple[int, int]:
        try:
            files = list(self._load_contract_files())
            return len(files), sum(int(item.get("size_bytes", 0) or 0) for item in files)
        except Exception:
            return 0, 0

    def _copy_contract_documents_to_share(self, share_store, share_ci) -> tuple[int, int]:
        folders = list(self._load_contract_file_folders())
        files = list(self._load_contract_files())
        if not folders and not files:
            return 0, 0
        folder_id_map = {}
        pending = [dict(f) for f in folders]
        while pending:
            progressed = False
            for folder in pending[:]:
                old_parent = folder.get("parent_id")
                if old_parent not in (None, "", 0) and int(old_parent) not in folder_id_map:
                    continue
                created = share_store.create_contract_file_folder(
                    str(share_ci.platform or ""),
                    str(share_ci.no or ""),
                    str(share_ci.contract_type or "Ana Sözleşme"),
                    parent_id=folder_id_map.get(int(old_parent)) if old_parent not in (None, "", 0) else None,
                    name=str(folder.get("name") or "Klasör"),
                )
                folder_id_map[int(folder.get("id"))] = int(created.get("id") or 0)
                pending.remove(folder)
                progressed = True
            if not progressed:
                break
        copied = 0
        total = 0
        with tempfile.TemporaryDirectory() as tmpdir:
            tmpdir_path = Path(tmpdir)
            for item in files:
                file_id = int(item.get("id") or 0)
                if not file_id:
                    continue
                filename, _mime, content = self.store.get_contract_file_bytes(file_id)
                safe_name = Path(str(filename or f"belge-{file_id}")).name
                tmp_file = tmpdir_path / safe_name
                suffix = 1
                while tmp_file.exists():
                    tmp_file = tmpdir_path / f"{tmp_file.stem}-{suffix}{tmp_file.suffix}"
                    suffix += 1
                tmp_file.write_bytes(content)
                old_folder_id = item.get("folder_id")
                new_folder_id = folder_id_map.get(int(old_folder_id)) if old_folder_id not in (None, "", 0) else None
                share_store.add_contract_file(
                    str(share_ci.platform or ""),
                    str(share_ci.no or ""),
                    tmp_file,
                    str(share_ci.contract_type or "Ana Sözleşme"),
                    note=str(item.get("note") or ""),
                    folder_id=new_folder_id,
                )
                copied += 1
                total += len(content)
        return copied, total

    def create_contract_share_file(self, permission: str, default_filename: str):
        """Create a real single-contract STS share file with share metadata."""
        if not self.require_permission_ui("export_data", "Sözleşme Paylaşımı"):
            return
        doc_count, doc_bytes = self._contract_document_share_stats()
        if doc_count > 0:
            QMessageBox.information(
                self,
                "Paylaşım Belgeleri",
                "Bu sözleşmeye bağlı belgeler paylaşım dosyasına dahil edilir. Dosya boyutu artabilir.",
            )
        target, _ = QFileDialog.getSaveFileName(self, "Paylaşım Dosyası Oluştur", default_filename, "STS Dosyası (*.sts)")
        if not target:
            return
        if not str(target).lower().endswith(".sts"):
            target += ".sts"
        target_path = Path(target)
        try:
            if target_path.exists():
                target_path.unlink()
            share_store = STSStore(target_path, actor="Sözleşme Paylaşımı")
            share_ci = copy.deepcopy(self.ci)
            share_ci.entry_start_row = 0
            setattr(share_ci, "id", 0)
            setattr(share_ci, "contract_id", 0)
            # Ana DB platform/staff ID'leri yeni boş DB'de geçersiz — sıfırla
            share_ci.platform_ids = []
            share_ci.platforms = []
            share_ci.platform_id = 0
            share_ci.primary_platform_id = 0
            share_ci.responsible_engineer_id = 0
            share_ci.responsible_engineer_ids = []
            contract_id = int(share_store.write_contract(share_ci, copy.deepcopy(self.systems), copy.deepcopy(self.deliveries)) or 0)
            share_store.save_contract_tags(
                str(share_ci.platform or ""),
                str(share_ci.no or ""),
                str(share_ci.contract_type or "Ana Sözleşme"),
                [dict(t or {}) for t in self.contract_tags],
                actor="Sözleşme Paylaşımı",
            )
            copied_docs, copied_doc_bytes = self._copy_contract_documents_to_share(share_store, share_ci)
            try:
                share_store.db.conn.commit()
                share_store.db.close()
            except Exception:
                pass
            _write_share_metadata(target_path, {
                "share_mode": "true",
                "contract_id": contract_id,
                "permission_mode": "edit" if permission == "duzenle" else "view",
                "source_contract_no": str(getattr(self.ci, "no", "") or ""),
                "created_at": datetime.now().isoformat(timespec="seconds"),
                "document_count": copied_docs,
                "document_bytes": copied_doc_bytes,
            })
            QMessageBox.information(self, "Paylaşım", "Paylaşım STS dosyası oluşturuldu.")
        except Exception as exc:
            QMessageBox.warning(self, "Paylaşım dosyası oluşturulamadı.", str(exc))

    def _make_card_scroll(self):
        scroll = QScrollArea(); scroll.setWidgetResizable(True); scroll.setFrameShape(QFrame.NoFrame)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff); scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        scroll.setStyleSheet("QScrollArea{background:transparent; border:0;} QScrollArea > QWidget > QWidget{background:transparent;} QScrollBar:vertical{width:8px; background:transparent;} QScrollBar::handle:vertical{background:#94a3b8; border-radius:4px; min-height:22px;} QScrollBar::add-line:vertical,QScrollBar::sub-line:vertical{height:0;}")
        host = QWidget(); host.setMinimumWidth(0); host.setSizePolicy(QSizePolicy.Ignored, QSizePolicy.Preferred); host.setStyleSheet("background:transparent;")
        cards = QVBoxLayout(host); cards.setContentsMargins(0, 0, 2, 0); cards.setSpacing(6); cards.addStretch()
        scroll.setWidget(host)
        return scroll, cards

    def _ordered_contract_tags(self) -> List[dict]:
        return sorted(
            [dict(t) for t in self.contract_tags if str((t or {}).get("name") or "").strip()],
            key=lambda x: self._tag_key(str(x.get("name", ""))),
        )

    def create_tag_card(self, tag: dict) -> QFrame:
        name = str((tag or {}).get("name") or "").strip()
        color = str((tag or {}).get("color") or "#3B82F6")
        card = QFrame(); card.setObjectName("sideTagCard"); card.setMinimumWidth(0); card.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed); card.setFixedHeight(52)
        card.setStyleSheet("QFrame#sideTagCard{background:#f8fbff; border:1px solid #dbe7f5; border-radius:5px;} QFrame#sideTagCard:hover{background:#eef6ff; border-color:#b8cef0;} QLabel{background:transparent; border:0;} QPushButton#tagRemoveButton{background:#f1f5fb; color:#64748b; border:1.5px solid #c8d8ee; border-radius:4px; font-size:16px; font-weight:700; padding:0;} QPushButton#tagRemoveButton:hover{background:#fee2e2; color:#b91c1c; border-color:#fca5a5;}")
        row = QHBoxLayout(card); row.setContentsMargins(9, 5, 9, 5); row.setSpacing(8)
        dot = QLabel("●"); dot.setFixedWidth(10); dot.setStyleSheet(f"color:{color}; font-size:12px;")
        middle = QWidget(); middle.setMinimumWidth(0); middle.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed); middle.setStyleSheet("background:transparent;"); column = QVBoxLayout(middle); column.setContentsMargins(0, 0, 0, 0); column.setSpacing(1)
        title = ElidedLabel(name); title.setMinimumWidth(0); title.setSizePolicy(QSizePolicy.Ignored, QSizePolicy.Fixed); title.setToolTip(name); title.setStyleSheet("color:#10233d; font-size:12px; font-weight:900;")
        meta = QLabel("Sözleşmeye atanmış etiket"); meta.setStyleSheet("color:#64748b; font-size:10px;")
        column.addWidget(title); column.addWidget(meta)
        remove = QPushButton("×"); remove.setObjectName("tagRemoveButton"); remove.setFixedSize(29, 29); remove.setToolTip("Etiketi kaldır"); remove.setEnabled(not self._share_is_view_only());
        if self._share_is_view_only():
            remove.setToolTip("Paylaşım görüntüleme modunda bu işlem kapalıdır.")
        remove.clicked.connect(lambda _=False, nm=name: self.remove_contract_tag(nm))
        row.addWidget(dot); row.addWidget(middle, 1); row.addWidget(remove)
        return card


    @staticmethod
    @staticmethod
    def _make_folder_icon() -> "QIcon":
        """Renkli klasör ikonu üretir."""
        from PySide6.QtCore import QRectF
        from PySide6.QtGui import QBrush, QPen
        px = QPixmap(32, 32)
        px.fill(Qt.transparent)
        painter = QPainter(px)
        painter.setRenderHint(QPainter.Antialiasing)
        # Klasör gövdesi
        painter.setBrush(QBrush(QColor("#f59e0b")))
        painter.setPen(Qt.NoPen)
        painter.drawRoundedRect(QRectF(1, 10, 30, 20), 4, 4)
        # Klasör sekmesi (üst sol)
        painter.drawRoundedRect(QRectF(1, 6, 13, 8), 3, 3)
        painter.end()
        return QIcon(px)

    @staticmethod
    def _make_file_icon(ext: str) -> "QIcon":
        """Ext'e göre renkli küçük dosya ikonu üretir."""
        ext_map = {
            "pdf":  ("#ef4444", "PDF"),
            "doc":  ("#2563eb", "DOC"), "docx": ("#2563eb", "DOC"),
            "xls":  ("#16a34a", "XLS"), "xlsx": ("#16a34a", "XLS"), "xlsm": ("#16a34a", "XLS"),
            "ppt":  ("#f97316", "PPT"), "pptx": ("#f97316", "PPT"),
            "png":  ("#7c3aed", "IMG"), "jpg":  ("#7c3aed", "IMG"), "jpeg": ("#7c3aed", "IMG"),
            "txt":  ("#64748b", "TXT"),
            "zip":  ("#b45309", "ZIP"),
        }
        color_hex, label = ext_map.get(str(ext).lower(), ("#64748b", str(ext).upper()[:3] or "???"))
        from PySide6.QtCore import QRectF
        from PySide6.QtGui import QBrush, QPen
        px = QPixmap(32, 32)
        px.fill(Qt.transparent)
        painter = QPainter(px)
        painter.setRenderHint(QPainter.Antialiasing)
        bg = QColor(color_hex)
        bg.setAlpha(220)
        painter.setBrush(QBrush(bg))
        painter.setPen(Qt.NoPen)
        painter.drawRoundedRect(QRectF(1, 1, 30, 30), 6, 6)
        painter.setPen(QPen(QColor("#ffffff")))
        font = painter.font()
        font.setPixelSize(9)
        font.setBold(True)
        painter.setFont(font)
        painter.drawText(px.rect(), Qt.AlignCenter, label)
        painter.end()
        return QIcon(px)

    @staticmethod
    def document_tree_height(folders: List[dict], files: List[dict]) -> int:
        """Dinamik ağaç yüksekliği: satır sayısına göre küçük→orta→büyük+scroll."""
        # Açık klasörlerdeki dosyaları da say (görünür satır tahmini)
        folder_ids = {int(f.get("id") or 0) for f in (folders or [])}
        visible = len(folders or [])
        for fi in (files or []):
            fid = fi.get("folder_id")
            if not fid or int(fid or 0) not in folder_ids:
                visible += 1  # kökteki dosya
            else:
                visible += 1  # klasör içi dosya (açık varsayımı)
        if visible <= 0:
            return 110
        if visible <= 6:
            return max(120, visible * 30 + 24)
        if visible <= 15:
            return max(210, visible * 30 + 24)
        return 420   # çok kayıt → internal scroll


    def create_contract_files_tree(self, folders: List[dict], files: List[dict]) -> ContractFileTreeWidget:
        # Expand/collapse ok ikonlarını temp dosyaya yaz
        import tempfile, os

        def _write_arrow_png(expanded: bool) -> str:
            from PySide6.QtCore import QBuffer, QIODevice
            px = QPixmap(16, 16)
            px.fill(Qt.transparent)
            p = QPainter(px)
            p.setRenderHint(QPainter.Antialiasing)
            p.setPen(Qt.NoPen)
            p.setBrush(QColor("#94a3b8"))
            from PySide6.QtGui import QPolygonF
            from PySide6.QtCore import QPointF
            if expanded:
                pts = [QPointF(3, 5), QPointF(13, 5), QPointF(8, 12)]
            else:
                pts = [QPointF(5, 3), QPointF(12, 8), QPointF(5, 13)]
            p.drawPolygon(QPolygonF(pts))
            p.end()
            tmp = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
            tmp.close()
            px.save(tmp.name, "PNG")
            return tmp.name.replace("\\", "/")

        arrow_closed = _write_arrow_png(False)
        arrow_open   = _write_arrow_png(True)

        tree = ContractFileTreeWidget(self)
        tree.setMinimumHeight(110)
        tree.setColumnCount(1)
        tree.setHeaderHidden(True)
        tree.setIndentation(20)
        tree.setIconSize(__import__('PySide6.QtCore', fromlist=['QSize']).QSize(18, 18))
        tree.setStyleSheet(
            "QTreeWidget{background:#f8fbff; border:0; border-radius:0; padding:4px 2px; color:#0f172a; font-size:11px; outline:0;}"
            "QTreeWidget::item{height:28px; border-radius:7px; padding:0 6px 0 2px;}"
            "QTreeWidget::item:selected{background:#dbeafe; color:#0f3f8f;}"
            "QTreeWidget::item:hover:!selected{background:#edf5ff;}"
            "QTreeWidget::branch{background:transparent; width:16px;}"
            f"QTreeWidget::branch:has-children:!has-siblings:closed,"
            f"QTreeWidget::branch:closed:has-children:has-siblings{{"
            f"  border:0; image:url({arrow_closed}); width:16px;"
            f"}}"
            f"QTreeWidget::branch:open:has-children:!has-siblings,"
            f"QTreeWidget::branch:open:has-children:has-siblings{{"
            f"  border:0; image:url({arrow_open}); width:16px;"
            f"}}"
            "QScrollBar:vertical{width:7px; background:transparent; margin:3px 1px 3px 0;}"
            "QScrollBar::handle:vertical{background:#c7d7ea; border-radius:3px; min-height:20px;}"
            "QScrollBar::handle:vertical:hover{background:#9eb8d7;}"
            "QScrollBar::add-line:vertical,QScrollBar::sub-line:vertical{height:0;}"
        )
        if self._share_is_view_only():
            tree.setAcceptDrops(False)
            tree.setDragEnabled(False)
            tree.setDragDropMode(QAbstractItemView.NoDragDrop)
            tree.setEditTriggers(QAbstractItemView.NoEditTriggers)
        tree.filesDropped.connect(lambda paths, folder_id: self._add_contract_files(paths, folder_id))
        tree.invalidDrop.connect(lambda message: QMessageBox.warning(self, "Dosya yüklenemedi", message))
        tree.itemMoved.connect(self._handle_tree_item_move)
        tree.itemDoubleClicked.connect(self.on_contract_file_tree_double_clicked)
        tree.itemChanged.connect(self.on_contract_file_tree_item_changed)
        tree.customContextMenuRequested.connect(self.show_contract_file_tree_menu)
        tree.setSelectionMode(QAbstractItemView.ExtendedSelection)
        tree.header().setSectionResizeMode(0, QHeaderView.Stretch)
        # Sağ tıklamada item seçilsin (önceki seçimi korumaz, tek item)
        _orig_mpe = tree.mousePressEvent
        def _mpe(event, _orig=_orig_mpe):
            if event.button() == Qt.RightButton:
                try:
                    pos = event.position().toPoint()
                except AttributeError:
                    pos = event.pos()
                item_at = tree.itemAt(pos)
                if item_at and item_at not in tree.selectedItems():
                    tree.clearSelection()
                    tree.setCurrentItem(item_at)
            _orig(event)
        tree.mousePressEvent = _mpe
        _folder_icon = self._make_folder_icon()
        self._building_file_tree = True
        try:
            folder_items = {}
            children_by_parent: Dict[object, List[dict]] = {}
            folder_ids = {int(folder.get("id")) for folder in folders if folder.get("id") not in (None, "")}

            def folder_file_count(fid):
                """Klasörün doğrudan dosya sayısını döner."""
                return sum(1 for f in files if int(f.get("folder_id") or 0) == fid)
            for folder in folders:
                parent_id = folder.get("parent_id")
                try:
                    parent_id = int(parent_id) if parent_id not in (None, "", 0) else None
                except (TypeError, ValueError):
                    parent_id = None
                if parent_id not in folder_ids:
                    parent_id = None
                children_by_parent.setdefault(parent_id, []).append(folder)

            def add_folder_items(parent_item, parent_id):
                for folder in sorted(children_by_parent.get(parent_id, []), key=lambda x: str(x.get("name") or "").casefold()):
                    folder_id = int(folder.get("id"))
                    count = folder_file_count(folder_id)
                    item = QTreeWidgetItem([str(folder.get("name") or "")])
                    item.setIcon(0, _folder_icon)
                    item.setToolTip(0, f"{folder.get('name', '')}  ·  {count} dosya")
                    item.setData(0, Qt.UserRole, "folder")
                    item.setData(0, Qt.UserRole + 1, folder_id)
                    item.setData(0, Qt.UserRole + 2, folder.get("_tree_parent_id"))
                    item.setData(0, Qt.UserRole + 3, str(folder.get("name") or ""))
                    item.setFlags(item.flags() | Qt.ItemIsEditable | Qt.ItemIsDropEnabled)
                    if parent_item is None:
                        tree.addTopLevelItem(item)
                    else:
                        parent_item.addChild(item)
                    folder_items[folder_id] = item
                    add_folder_items(item, folder_id)

            add_folder_items(None, None)
            for metadata in sorted(files, key=lambda x: (str(x.get("filename") or "").casefold(), int(x.get("id") or 0))):
                try:
                    file_folder_id = int(metadata.get("folder_id") or 0)
                except (TypeError, ValueError):
                    file_folder_id = 0
                parent_item = folder_items.get(file_folder_id)
                ext = str(metadata.get("file_ext") or "").lower() or "dosya"
                size_text = self.format_file_size(metadata.get("size_bytes", 0))
                item = QTreeWidgetItem([str(metadata.get("filename") or "")])
                item.setIcon(0, self._make_file_icon(ext))
                item.setToolTip(0, f"{ext.upper()} · {size_text} · {self.format_file_date(metadata.get('created_at', ''))}")
                item.setData(0, Qt.UserRole, "file")
                item.setData(0, Qt.UserRole + 1, int(metadata.get("id")))
                item.setData(0, Qt.UserRole + 2, file_folder_id if parent_item is not None else None)
                item.setFlags((item.flags() | Qt.ItemIsDragEnabled) & ~Qt.ItemIsEditable)
                if parent_item is None:
                    tree.addTopLevelItem(item)
                else:
                    parent_item.addChild(item)
            # Sadece kök klasörler açık, alt klasörler kapalı başlar
            for i in range(tree.topLevelItemCount()):
                top = tree.topLevelItem(i)
                if top and top.data(0, Qt.UserRole) == "folder":
                    top.setExpanded(True)
        finally:
            self._building_file_tree = False
        return tree

    def _selected_document_folder_id(self):
        tree = getattr(self, "contract_files_tree", None)
        item = tree.currentItem() if tree else None
        if item is None:
            return None
        kind = item.data(0, Qt.UserRole)
        if kind == "folder":
            return item.data(0, Qt.UserRole + 1)
        if kind == "file":
            return item.data(0, Qt.UserRole + 2)
        return None

    def _start_rename_item(self, item):
        if not self._ensure_share_can_edit("Belgeler"):
            return
        """Klasör rename editörünü aç ve metni seçili göster."""
        tree = getattr(self, "contract_files_tree", None)
        if not tree or not item:
            return
        self._tree_editing = True
        tree.setCurrentItem(item)
        # Delegate: editör açıldığında arka plan tamamen kapatılsın (overlay bug önlemi)
        # Her tree nesnesi için delegate kur (render sonrası yeni tree oluşur)
        if not getattr(tree, "_rename_delegate_set", False):
            from PySide6.QtWidgets import QStyledItemDelegate
            from PySide6.QtGui import QColor
            class _SolidBgDelegate(QStyledItemDelegate):
                def createEditor(self, parent, option, index):
                    editor = super().createEditor(parent, option, index)
                    if editor:
                        editor.setStyleSheet(
                            "QLineEdit{"
                            "  background:#ffffff;"
                            "  color:#0f172a;"
                            "  border:2px solid #1f5be3;"
                            "  border-radius:5px;"
                            "  padding:1px 4px;"
                            "  font-size:11px;"
                            "  min-width:120px;"
                            "}"
                        )
                    return editor
                def paint(self, painter, option, index):
                    # PySide6'da option.state bir StateFlag enum'u – int'e cast ederek kontrol et
                    from PySide6.QtWidgets import QStyle
                    try:
                        is_editing = bool(int(option.state) & int(QStyle.State_Editing))
                    except Exception:
                        is_editing = False
                    if is_editing:
                        from PySide6.QtGui import QColor
                        painter.fillRect(option.rect, QColor("#ffffff"))
                        return
                    super().paint(painter, option, index)
            tree.setItemDelegate(_SolidBgDelegate(tree))
            tree._rename_delegate_set = True
        tree.scrollToItem(item)
        tree.editItem(item, 0)
        def _select_all():
            editor = tree.focusWidget()
            if editor and hasattr(editor, "selectAll"):
                editor.selectAll()
        QTimer.singleShot(30, _select_all)

    def add_contract_file_folder(self):
        if not self._ensure_share_can_edit("Belgeler"):
            return
        if not self._ensure_document_access(interactive=True):
            return
        try:
            parent_id = self._selected_document_folder_id()
            if self.is_new_contract:
                # Pending modda in-memory oluştur
                base_name = "Yeni Klasör"
                existing = {f["name"] for f in self._pending_doc_folders if f.get("parent_id") == parent_id}
                name = base_name
                idx = 2
                while name in existing:
                    name = f"{base_name} ({idx})"
                    idx += 1
                new_id = self._pending_doc_next_id
                self._pending_doc_next_id -= 1
                folder = {"id": new_id, "parent_id": parent_id, "name": name, "created_at": "", "updated_at": ""}
                self._pending_doc_folders.append(folder)
                created = folder
            else:
                created = self.store.create_contract_file_folder(
                    self.ci.platform, self.ci.no, self.ci.contract_type, parent_id=parent_id
                )
            self._mark_documents_changed()
            # render_contract_files() tüm tree'yi yıkıp yeniden kurar →
            # editör yanlış konuma açılıyor. Bunun yerine tree'ye direkt item ekle.
            tree = getattr(self, "contract_files_tree", None)
            if tree:
                # Yeni klasör item'ını direkt tree'ye ekle
                _folder_icon = self._make_folder_icon()
                new_item = QTreeWidgetItem([str(created.get("name") or "Yeni Klasör")])
                new_item.setIcon(0, _folder_icon)
                new_item.setData(0, Qt.UserRole, "folder")
                new_item.setData(0, Qt.UserRole + 1, int(created.get("id") or 0))
                new_item.setData(0, Qt.UserRole + 2, created.get("parent_id"))
                new_item.setData(0, Qt.UserRole + 3, str(created.get("name") or "Yeni Klasör"))
                new_item.setFlags(new_item.flags() | Qt.ItemIsEditable | Qt.ItemIsDropEnabled)
                # Parent klasör varsa ona, yoksa köke ekle
                _parent_folder_id = created.get("parent_id")
                _parent_item = None
                if _parent_folder_id:
                    def _find_parent(n_top):
                        for _i in range(n_top):
                            _top = tree.topLevelItem(_i)
                            if _top and _top.data(0, Qt.UserRole) == "folder":
                                try:
                                    if int(_top.data(0, Qt.UserRole + 1)) == int(_parent_folder_id):
                                        return _top
                                except Exception:
                                    pass
                                for _j in range(_top.childCount()):
                                    _ch = _top.child(_j)
                                    if _ch and _ch.data(0, Qt.UserRole) == "folder":
                                        try:
                                            if int(_ch.data(0, Qt.UserRole + 1)) == int(_parent_folder_id):
                                                return _ch
                                        except Exception:
                                            pass
                        return None
                    _parent_item = _find_parent(tree.topLevelItemCount())
                if _parent_item:
                    _parent_item.addChild(new_item)
                    _parent_item.setExpanded(True)
                else:
                    tree.addTopLevelItem(new_item)
                tree.scrollToItem(new_item)
                self._start_rename_item(new_item)
                # Yüksekliği güncelle (1 item arttı)
                all_f = list(getattr(self, "_side_meta_folders", []))
                all_files = list(getattr(self, "_side_meta_files", []))
                new_h = self.document_tree_height(all_f, all_files)
                tree.setMinimumHeight(new_h)
                tree.setMaximumHeight(new_h)
            else:
                # Tree henüz yok, tam render yap
                self.render_contract_files()
                _created_id = int(created.get("id") or 0)
                def _open_rename_editor():
                    _tree = getattr(self, "contract_files_tree", None)
                    if not _tree:
                        return
                    def _find(_p, _n):
                        for _i in range(_n):
                            _c = _p.child(_i) if _p else _tree.topLevelItem(_i)
                            if _c and _c.data(0, Qt.UserRole) == "folder":
                                try:
                                    if int(_c.data(0, Qt.UserRole + 1)) == _created_id:
                                        _tree.scrollToItem(_c)
                                        self._start_rename_item(_c)
                                        return True
                                except Exception:
                                    pass
                                if _find(_c, _c.childCount()):
                                    return True
                        return False
                    _find(None, _tree.topLevelItemCount())
                QTimer.singleShot(60, _open_rename_editor)
            self.update_side_meta_badges()
        except Exception as exc:
            QMessageBox.warning(self, "Klasör eklenemedi", str(exc))

    def on_contract_file_tree_double_clicked(self, item, column):
        if not self._ensure_document_access(interactive=True):
            return
        if not item:
            return
        if item.data(0, Qt.UserRole) == "file":
            self.open_contract_file(int(item.data(0, Qt.UserRole + 1)))
        elif item.data(0, Qt.UserRole) == "folder":
            self._start_rename_item(item)

    def on_contract_file_tree_item_changed(self, item, column):
        if not self._ensure_share_can_edit("Belgeler"):
            return
        if not self._ensure_document_access(interactive=True):
            return
        if getattr(self, "_building_file_tree", False) or not item or item.data(0, Qt.UserRole) != "folder":
            return
        self._tree_editing = False
        folder_id = int(item.data(0, Qt.UserRole + 1))
        old_name = str(item.data(0, Qt.UserRole + 3) or "")
        new_name = str(item.text(0) or "").strip()
        if new_name == old_name:
            return
        try:
            if self.is_new_contract:
                # Pending modda: in-memory güncelle
                for f in self._pending_doc_folders:
                    if f["id"] == folder_id:
                        parent_id = f.get("parent_id")
                        existing = {pf["name"] for pf in self._pending_doc_folders if pf.get("parent_id") == parent_id and pf["id"] != folder_id}
                        if new_name in existing:
                            raise ValueError("Aynı seviyede bu klasör adı zaten var.")
                        f["name"] = new_name
                        break
                item.setData(0, Qt.UserRole + 3, new_name)
            else:
                renamed = self.store.rename_contract_file_folder(folder_id, new_name)
                item.setData(0, Qt.UserRole + 3, str(renamed.get("name") or new_name))
            self._mark_documents_changed()
            # Ağacı yeniden render ETME – sadece badge güncelle.
            # render_contract_files() tüm tree'yi silip yeniden kurar,
            # bu da expand state'i sıfırlayıp dosyaları "kaybeder" görüntüsü verir.
            self.update_side_meta_badges()
        except Exception as exc:
            self._building_file_tree = True
            try:
                item.setText(0, old_name)
            finally:
                self._building_file_tree = False
            QMessageBox.warning(self, "Klasör adı değiştirilemedi", str(exc))

    def _begin_side_meta_modal_action(self):
        self._side_meta_modal_open = True
        self._tree_editing = True

    def _end_side_meta_modal_action(self):
        self._side_meta_modal_open = False
        self._tree_editing = False
        if getattr(self, "_side_meta_last_panel", None) == "files":
            QTimer.singleShot(0, self._restore_documents_popover_if_needed)

    def _restore_documents_popover_if_needed(self):
        popover = getattr(self, "side_meta_popover", None)
        if popover is None or getattr(self, "_side_meta_last_panel", None) != "files":
            return
        self._side_meta_open_panel = "files"
        self._sync_side_meta_controls()
        self.position_side_meta_popover()
        popover.show()
        popover.raise_()

    def delete_contract_file_folder(self, folder_id, folder_name):
        if not self._ensure_share_can_edit("Belgeler"):
            return
        if not self._ensure_document_access(interactive=True):
            return
        self._begin_side_meta_modal_action()
        try:
            msg = (
                f'"{folder_name}" klasörü, alt klasörleri ve içindeki tüm belgeler STS dosyasından silinecek.\n'
                "Bu işlem geri alınamaz."
            )
            if not ask_yes_no(self, "Klasörü Sil", msg):
                return
            try:
                if self.is_new_contract:
                    # Pending modda recursive sil
                    def _collect_pending_folder_ids(fid):
                        ids = [fid]
                        for f in self._pending_doc_folders:
                            if f.get("parent_id") == fid:
                                ids.extend(_collect_pending_folder_ids(f["id"]))
                        return ids
                    all_ids = set(_collect_pending_folder_ids(folder_id))
                    self._pending_doc_folders = [f for f in self._pending_doc_folders if f["id"] not in all_ids]
                    self._pending_doc_files = [f for f in self._pending_doc_files if f.get("folder_id") not in all_ids]
                else:
                    self.store.delete_contract_file_folder(folder_id)
                self._mark_documents_changed()
                self.render_contract_files()
            except Exception as exc:
                QMessageBox.warning(self, "Klasör silinemedi", str(exc))
        finally:
            self._end_side_meta_modal_action()

    def show_contract_file_tree_menu(self, pos):
        if not self._ensure_share_can_edit("Belgeler"):
            return
        if not self._ensure_document_access(interactive=True):
            return
        tree = getattr(self, "contract_files_tree", None)
        item = tree.itemAt(pos) if tree else None
        if not item:
            return
        kind = item.data(0, Qt.UserRole)
        # Seçili dosyalar
        sel_file_ids = [
            int(it.data(0, Qt.UserRole + 1))
            for it in tree.selectedItems()
            if it.data(0, Qt.UserRole) == "file"
        ]
        menu = QMenu(self)
        menu.setStyleSheet(
            "QMenu{background:#ffffff;border:1px solid #d5e0ee;border-radius:10px;padding:4px;}"
            "QMenu::item{padding:7px 18px 7px 12px;border-radius:7px;font-size:12px;color:#0f172a;}"
            "QMenu::item:selected{background:#eef5ff;color:#1d4ed8;}"
            "QMenu::separator{height:1px;background:#e8eff8;margin:4px 8px;}"
        )
        # Menü açıkken popup kapanmasın
        self._tree_editing = True
        menu.aboutToHide.connect(lambda: setattr(self, "_tree_editing", False))

        _zip_icon = self._make_file_icon("zip")

        if kind == "file":
            file_id = int(item.data(0, Qt.UserRole + 1))
            if len(sel_file_ids) > 1:
                menu.addAction(f"📄  {len(sel_file_ids)} dosya seçili")
                a = menu.actions()[-1]; a.setEnabled(False)
                menu.addSeparator()
                menu.addAction("⬇  Tümünü İndir", lambda ids=sel_file_ids: self._bulk_download_files(ids))
                act = menu.addAction("  ZIP Olarak İndir", lambda ids=sel_file_ids: self._bulk_zip_files(ids))
                act.setIcon(_zip_icon)
                menu.addSeparator()
                menu.addAction("🗑  Seçilenleri Sil", lambda ids=sel_file_ids: self._bulk_delete_files(ids))
            else:
                menu.addAction("📂  Aç", lambda: self.open_contract_file(file_id))
                menu.addAction("⬇  İndir", lambda fid=file_id: self._bulk_download_files([fid]))
                act = menu.addAction("  ZIP Olarak İndir", lambda fid=file_id: self._bulk_zip_files([fid]))
                act.setIcon(_zip_icon)
                menu.addSeparator()
                menu.addAction("🗑  Sil", lambda: self.delete_contract_file(file_id))

        elif kind == "folder":
            folder_id = int(item.data(0, Qt.UserRole + 1))
            folder_name = str(item.data(0, Qt.UserRole + 3) or item.text(0))
            # Seçili klasörün parent_id'sini al (üst klasör eklemek için)
            _folder_parent_id = item.data(0, Qt.UserRole + 2)
            menu.addAction("➕  Dosya Ekle", lambda: self._add_files_to_folder(folder_id))
            menu.addAction("📁  Alt Klasör Ekle", lambda: self._add_subfolder(folder_id))
            menu.addAction("📂  Üst Klasör Ekle", lambda pid=_folder_parent_id: self._add_subfolder(pid))
            menu.addSeparator()
            menu.addAction("⬇  Klasörü İndir (Klasör Olarak)", lambda fid=folder_id, fname=folder_name: self._download_folder(fid, fname, as_zip=False))
            act = menu.addAction("  Klasörü İndir (ZIP)", lambda fid=folder_id, fname=folder_name: self._download_folder(fid, fname, as_zip=True))
            act.setIcon(_zip_icon)
            menu.addSeparator()
            menu.addAction("✏  Yeniden Adlandır", lambda i=item: self._start_rename_item(i))
            menu.addSeparator()
            menu.addAction("🗑  Sil", lambda: self.delete_contract_file_folder(folder_id, folder_name))

        menu.exec(tree.viewport().mapToGlobal(pos))

    def _add_files_to_folder(self, folder_id):
        if not self._ensure_share_can_edit("Belgeler"):
            return
        """Belirli bir klasöre dosya ekle."""
        if not self._ensure_document_access(interactive=True):
            return
        self._file_dialog_open = True
        try:
            paths, _ = QFileDialog.getOpenFileNames(
                self,
                "Klasöre Dosya Ekle",
                "",
                "Documents (*.pdf *.doc *.docx *.xls *.xlsx *.xlsm *.ppt *.pptx *.txt *.png *.jpg *.jpeg);;All Files (*.*)",
            )
        finally:
            self._file_dialog_open = False
        if paths:
            self._add_contract_files(paths, folder_id)

    def _add_subfolder(self, parent_folder_id):
        if not self._ensure_share_can_edit("Belgeler"):
            return
        """Mevcut klasörün altına alt klasör ekle."""
        if not self._ensure_document_access(interactive=True):
            return
        try:
            if self.is_new_contract:
                base_name = "Yeni Klasör"
                existing = {f["name"] for f in self._pending_doc_folders if f.get("parent_id") == parent_folder_id}
                name = base_name
                idx = 2
                while name in existing:
                    name = f"{base_name} ({idx})"
                    idx += 1
                new_id = self._pending_doc_next_id
                self._pending_doc_next_id -= 1
                folder = {"id": new_id, "parent_id": parent_folder_id, "name": name, "created_at": "", "updated_at": ""}
                self._pending_doc_folders.append(folder)
                created = folder
            else:
                created = self.store.create_contract_file_folder(
                    self.ci.platform, self.ci.no, self.ci.contract_type, parent_id=parent_folder_id
                )
            self._mark_documents_changed()
            tree = getattr(self, "contract_files_tree", None)
            if tree:
                _folder_icon = self._make_folder_icon()
                new_item = QTreeWidgetItem([str(created.get("name") or "Yeni Klasör")])
                new_item.setIcon(0, _folder_icon)
                new_item.setData(0, Qt.UserRole, "folder")
                new_item.setData(0, Qt.UserRole + 1, int(created.get("id") or 0))
                new_item.setData(0, Qt.UserRole + 2, created.get("parent_id"))
                new_item.setData(0, Qt.UserRole + 3, str(created.get("name") or "Yeni Klasör"))
                new_item.setFlags(new_item.flags() | Qt.ItemIsEditable | Qt.ItemIsDropEnabled)
                # parent_folder_id'ye sahip tree item'ı bul
                _pid = parent_folder_id
                _par = None
                if _pid:
                    def _fp(n):
                        for _i in range(n):
                            _t = tree.topLevelItem(_i)
                            if _t and _t.data(0, Qt.UserRole) == "folder":
                                try:
                                    if int(_t.data(0, Qt.UserRole + 1)) == int(_pid):
                                        return _t
                                except Exception:
                                    pass
                                for _j in range(_t.childCount()):
                                    _c = _t.child(_j)
                                    if _c and _c.data(0, Qt.UserRole) == "folder":
                                        try:
                                            if int(_c.data(0, Qt.UserRole + 1)) == int(_pid):
                                                return _c
                                        except Exception:
                                            pass
                        return None
                    _par = _fp(tree.topLevelItemCount())
                if _par:
                    _par.addChild(new_item)
                    _par.setExpanded(True)
                else:
                    tree.addTopLevelItem(new_item)
                tree.scrollToItem(new_item)
                self._start_rename_item(new_item)
                all_f = list(getattr(self, "_side_meta_folders", []))
                all_files = list(getattr(self, "_side_meta_files", []))
                new_h = self.document_tree_height(all_f, all_files)
                tree.setMinimumHeight(new_h)
                tree.setMaximumHeight(new_h)
            else:
                self.render_contract_files()
            self.update_side_meta_badges()
        except Exception as exc:
            QMessageBox.warning(self, "Alt klasör eklenemedi", str(exc))

    def _bulk_download_files(self, file_ids: list):
        """Seçili dosyaları klasöre indir."""
        if not self.require_permission_ui("export_data", "Belge Dışa Aktar"):
            return
        if not self._ensure_document_access(interactive=True):
            return
        if not file_ids:
            return
        folder = QFileDialog.getExistingDirectory(self, "İndirme Klasörü Seç")
        if not folder:
            return
        errors = []
        for fid in file_ids:
            try:
                filename, _mime, content = self.store.get_contract_file_bytes(fid)
                target = Path(folder) / filename
                counter = 1
                while target.exists():
                    target = Path(folder) / f"{Path(filename).stem}_{counter}{Path(filename).suffix}"
                    counter += 1
                target.write_bytes(content)
            except Exception as exc:
                errors.append(str(exc))
        if errors:
            QMessageBox.warning(self, "Bazı dosyalar indirilemedi", "\n".join(errors[:5]))
        else:
            QMessageBox.information(self, "İndirme tamamlandı", f"{len(file_ids)} dosya indirildi.")
            QDesktopServices.openUrl(QUrl.fromLocalFile(folder))

    def _bulk_zip_files(self, file_ids: list):
        """Seçili dosyaları ZIP olarak indir."""
        if not self.require_permission_ui("export_data", "Belge Dışa Aktar"):
            return
        if not self._ensure_document_access(interactive=True):
            return
        if not file_ids:
            return
        default_name = f"{self.ci.no}_belgeler.zip" if hasattr(self, "ci") else "belgeler.zip"
        target, _ = QFileDialog.getSaveFileName(self, "ZIP Kaydet", default_name, "ZIP (*.zip)")
        if not target:
            return
        try:
            added_count = 0
            errors = []
            with zipfile.ZipFile(target, "w", zipfile.ZIP_DEFLATED) as zf:
                seen: dict = {}
                for fid in file_ids:
                    try:
                        filename, _mime, file_content = self.store.get_contract_file_bytes(fid)
                        arc_name = filename
                        if arc_name in seen:
                            seen[arc_name] += 1
                            stem = Path(arc_name).stem
                            ext = Path(arc_name).suffix
                            arc_name = f"{stem}_{seen[filename]}{ext}"
                        else:
                            seen[arc_name] = 0
                        zf.writestr(arc_name, file_content)
                        added_count += 1
                    except Exception as exc:
                        errors.append(str(exc))
            if added_count == 0:
                import os
                try:
                    os.unlink(target)
                except Exception:
                    pass
                msg = "Hiçbir dosya ZIP'e eklenemedi."
                if errors:
                    msg += "\n\nHatalar:\n" + "\n".join(errors[:5])
                QMessageBox.warning(self, "ZIP oluşturulamadı", msg)
                return
            if errors:
                msg = f"{added_count} dosya ZIP'e eklendi, {len(errors)} dosya eklenemedi.\n\nİlk hatalar:\n" + "\n".join(errors[:5])
                QMessageBox.warning(self, "ZIP kısmen oluşturuldu", msg)
            else:
                QMessageBox.information(self, "ZIP oluşturuldu", f"{added_count} dosya ZIP'e eklendi.")
            QDesktopServices.openUrl(QUrl.fromLocalFile(str(Path(target).parent)))
        except Exception as exc:
            QMessageBox.warning(self, "ZIP oluşturulamadı", str(exc))

    def _bulk_delete_files(self, file_ids: list):
        if not self._ensure_share_can_edit("Belgeler"):
            return
        """Seçili dosyaları sil."""
        if not self._ensure_document_access(interactive=True):
            return
        if not file_ids:
            return
        self._begin_side_meta_modal_action()
        try:
            if not ask_yes_no(
                self, "Dosyaları Sil",
                f"{len(file_ids)} dosyayı silmek istediğinize emin misiniz?\nBu işlem geri alınamaz.",
            ):
                return
            for fid in file_ids:
                try:
                    self.store.delete_contract_file(fid)
                except Exception:
                    pass
            self._mark_documents_changed()
            self.render_contract_files()
        finally:
            self._end_side_meta_modal_action()

    def _download_folder(self, folder_id: int, folder_name: str, as_zip: bool = True):
        """Klasör ve tüm alt klasörlerini/dosyalarını recursive indir."""
        if not self.require_permission_ui("export_data", "Belge Dışa Aktar"):
            return
        if not self._ensure_document_access(interactive=True):
            return
        # Tüm dosya ve klasör verisini yükle
        all_files = list(self._side_meta_files)
        all_folders = list(getattr(self, "_side_meta_folders", []))

        # Klasör hiyerarşisini kur: folder_id → children
        folders_by_id = {int(f["id"]): f for f in all_folders if f.get("id")}
        children_by_parent: dict = {}
        for f in all_folders:
            pid = f.get("parent_id")
            try:
                pid = int(pid) if pid not in (None, "", 0) else None
            except (TypeError, ValueError):
                pid = None
            children_by_parent.setdefault(pid, []).append(f)

        def collect_folder_file_ids(fid, path_prefix=""):
            """Klasör altındaki tüm dosyaları (recursive) (file_id, arc_path) olarak döner."""
            fname = folders_by_id.get(fid, {}).get("name", str(fid))
            cur_path = f"{path_prefix}{fname}/" if path_prefix else f"{fname}/"
            result = []
            # Bu klasördeki dosyalar
            for f in all_files:
                try:
                    ffid = int(f.get("folder_id") or 0)
                except (TypeError, ValueError):
                    ffid = 0
                if ffid == fid:
                    result.append((int(f["id"]), f.get("filename", "dosya"), cur_path))
            # Alt klasörler recursive
            for child in children_by_parent.get(fid, []):
                result.extend(collect_folder_file_ids(int(child["id"]), cur_path))
            return result

        entries = collect_folder_file_ids(folder_id)

        if not entries:
            QMessageBox.information(self, "Boş Klasör", f'"{folder_name}" klasöründe indirilecek dosya yok.')
            return

        if as_zip:
            # ZIP olarak indir
            default_name = f"{folder_name}.zip"
            target, _ = QFileDialog.getSaveFileName(self, "ZIP Kaydet", default_name, "ZIP (*.zip)")
            if not target:
                return
            try:
                dl_added_count = 0
                dl_errors = []
                seen: dict = {}
                with zipfile.ZipFile(target, "w", zipfile.ZIP_DEFLATED) as zf:
                    for file_id, filename, arc_prefix in entries:
                        arc_name = arc_prefix + filename
                        # Çakışma önle
                        if arc_name in seen:
                            seen[arc_name] += 1
                            stem = Path(filename).stem
                            ext = Path(filename).suffix
                            arc_name = arc_prefix + f"{stem}_{seen[arc_name]}{ext}"
                        else:
                            seen[arc_name] = 0
                        try:
                            _, _mime, dl_content = self.store.get_contract_file_bytes(file_id)
                            zf.writestr(arc_name, dl_content)
                            dl_added_count += 1
                        except Exception as exc:
                            dl_errors.append(f"{filename}: {exc}")
                if dl_added_count == 0:
                    import os as _os
                    try:
                        _os.unlink(target)
                    except Exception:
                        pass
                    msg = "Hiçbir dosya ZIP'e eklenemedi."
                    if dl_errors:
                        msg += "\n\nHatalar:\n" + "\n".join(dl_errors[:5])
                    QMessageBox.warning(self, "ZIP oluşturulamadı", msg)
                    return
                if dl_errors:
                    msg = f"{dl_added_count} dosya ZIP'e eklendi, {len(dl_errors)} dosya eklenemedi.\n\nİlk hatalar:\n" + "\n".join(dl_errors[:5])
                    QMessageBox.warning(self, "ZIP kısmen oluşturuldu", msg)
                else:
                    QMessageBox.information(self, "ZIP oluşturuldu", f"{dl_added_count} dosya ZIP'e eklendi.\n{target}")
                QDesktopServices.openUrl(QUrl.fromLocalFile(str(Path(target).parent)))
            except Exception as exc:
                QMessageBox.warning(self, "ZIP oluşturulamadı", str(exc))
        else:
            # Klasör yapısıyla indir
            base_dir = QFileDialog.getExistingDirectory(self, f'"{folder_name}" için hedef klasör seç')
            if not base_dir:
                return
            errors = []
            for file_id, filename, arc_prefix in entries:
                try:
                    _, _mime, content = self.store.get_contract_file_bytes(file_id)
                    dest_dir = Path(base_dir) / arc_prefix
                    dest_dir.mkdir(parents=True, exist_ok=True)
                    dest_file = dest_dir / filename
                    counter = 1
                    while dest_file.exists():
                        dest_file = dest_dir / f"{Path(filename).stem}_{counter}{Path(filename).suffix}"
                        counter += 1
                    dest_file.write_bytes(content)
                except Exception as exc:
                    errors.append(f"{filename}: {exc}")
            if errors:
                QMessageBox.warning(self, "Bazı dosyalar indirilemedi", "\n".join(errors[:5]))
            else:
                QMessageBox.information(self, "İndirme tamamlandı",
                    f"{len(entries)} dosya indirildi.\nHedef: {base_dir}/{folder_name}/")
                QDesktopServices.openUrl(QUrl.fromLocalFile(base_dir))

    def refresh_contract_tags_panel(self):
        self.update_side_meta_badges()
        if self._side_meta_open_panel == "tags":
            self.render_side_meta_popover_content("tags")

    def render_contract_tags(self):
        self.refresh_contract_tags_panel()

    def render_contract_files(self):
        self.refresh_contract_files_panel()

    @staticmethod
    def format_file_size(size_bytes: int) -> str:
        size = float(size_bytes or 0)
        if size < 1024: return f"{int(size)} B"
        if size < 1024 * 1024: return f"{size / 1024:.0f} KB"
        return f"{size / (1024 * 1024):.1f} MB"

    @staticmethod
    def format_file_date(created_at: str) -> str:
        raw = str(created_at or "").strip()
        if not raw: return "-"
        try:
            parsed = datetime.fromisoformat(raw.replace("Z", "+00:00"))
            return "Bugün" if parsed.date() == date.today() else parsed.strftime("%d.%m.%Y")
        except ValueError:
            return raw[:10]

    @staticmethod
    def file_type_style(ext: str) -> Tuple[str, str]:
        extension = str(ext or "").strip().lower()
        if extension == "pdf": return "PDF", "#ef4444"
        if extension in {"doc", "docx"}: return "DOC", "#2563eb"
        if extension in {"xls", "xlsx", "xlsm"}: return "XLS", "#16a34a"
        if extension in {"ppt", "pptx"}: return "PPT", "#f97316"
        if extension in {"png", "jpg", "jpeg"}: return "IMG", "#7c3aed"
        return "TXT", "#64748b"

    def create_file_card(self, metadata: dict) -> QFrame:
        file_id = int(metadata["id"]); filename = str(metadata.get("filename") or "")
        ext = str(metadata.get("file_ext") or "").upper() or "DOSYA"
        icon_text, icon_color = self.file_type_style(ext)
        card = QFrame(); card.setObjectName("sideFileCard"); card.setMinimumWidth(0); card.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed); card.setFixedHeight(56); card.setProperty("contractFileId", file_id); card.installEventFilter(self)
        card.setStyleSheet("QFrame#sideFileCard{background:#f8fbff; border:1px solid #dbe7f5; border-radius:12px;} QFrame#sideFileCard:hover{background:#eef6ff; border-color:#b8cef0;} QLabel{background:transparent; border:0;} QToolButton{background:#ffffff; color:#334155; border:1px solid #d8e4f2; border-radius:8px; font-size:16px; font-weight:900;} QToolButton:hover{background:#eff6ff; border-color:#b8cef0;}")
        row = QHBoxLayout(card); row.setContentsMargins(9, 7, 9, 7); row.setSpacing(8)
        icon = QLabel(icon_text); icon.setFixedSize(36, 36); icon.setAlignment(Qt.AlignCenter); icon.setStyleSheet(f"background:{icon_color}; color:#ffffff; border-radius:11px; font-size:13px; font-weight:900;")
        middle = QWidget(); middle.setMinimumWidth(0); middle.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed); middle.setStyleSheet("background:transparent;"); column = QVBoxLayout(middle); column.setContentsMargins(0, 0, 0, 0); column.setSpacing(2)
        title = ElidedLabel(filename); title.setMinimumWidth(0); title.setSizePolicy(QSizePolicy.Ignored, QSizePolicy.Fixed); title.setToolTip(filename); title.setStyleSheet("color:#10233d; font-size:12px; font-weight:900;")
        meta = QLabel(f"{ext}  ·  {self.format_file_size(metadata.get('size_bytes', 0))}  ·  {self.format_file_date(metadata.get('created_at', ''))}"); meta.setMinimumWidth(0); meta.setSizePolicy(QSizePolicy.Ignored, QSizePolicy.Fixed); meta.setStyleSheet("color:#64748b; font-size:10px;")
        column.addWidget(title); column.addWidget(meta)
        menu_btn = QToolButton(); menu_btn.setText("⋯"); menu_btn.setFixedSize(30, 30); menu_btn.setToolTip("Menü"); menu_btn.clicked.connect(lambda _=False, fid=file_id, btn=menu_btn: self.show_contract_file_button_menu(fid, btn))
        row.addWidget(icon); row.addWidget(middle, 1); row.addWidget(menu_btn)
        return card

    def refresh_contract_files_panel(self):
        self.update_side_meta_badges()
        if self._side_meta_open_panel == "files":
            self.render_side_meta_popover_content("files")

    def _pick_contract_files(self):
        if not self._ensure_share_can_edit("Belgeler"):
            return
        if not self._ensure_document_access(interactive=True):
            return
        self._file_dialog_open = True
        try:
            paths, _ = QFileDialog.getOpenFileNames(
                self,
                "Sözleşmeye Dosya Ekle",
                "",
                "Documents (*.pdf *.doc *.docx *.xls *.xlsx *.xlsm *.ppt *.pptx *.txt *.png *.jpg *.jpeg);;All Files (*.*)",
            )
        finally:
            self._file_dialog_open = False
        if not paths:
            return
        self._add_contract_files(paths, self._selected_document_folder_id())

    def _handle_tree_item_move(self, kind: str, item_id: int, target_folder_id):
        if not self._ensure_share_can_edit("Belgeler"):
            return
        """Tree içinde sürükle-bırak taşıma işlemini yönet."""
        if not self._ensure_document_access(interactive=True):
            return
        try:
            if self.is_new_contract:
                # Pending modda in-memory taşıma
                if kind == "file":
                    for f in self._pending_doc_files:
                        if f["id"] == item_id:
                            f["folder_id"] = target_folder_id
                            break
                elif kind == "folder":
                    # Kendi altına taşınamaz
                    def is_descendant_pending(fid, anc_id):
                        for ff in self._pending_doc_folders:
                            if ff["id"] == fid:
                                pid = ff.get("parent_id")
                                if pid is None:
                                    return False
                                if pid == anc_id:
                                    return True
                                return is_descendant_pending(pid, anc_id)
                        return False
                    if target_folder_id == item_id:
                        QMessageBox.warning(self, "Taşınamaz", "Klasör kendi içine taşınamaz.")
                        return
                    if target_folder_id is not None and is_descendant_pending(target_folder_id, item_id):
                        QMessageBox.warning(self, "Taşınamaz", "Klasör kendi alt klasörüne taşınamaz.")
                        return
                    for f in self._pending_doc_folders:
                        if f["id"] == item_id:
                            f["parent_id"] = target_folder_id
                            break
            else:
                if kind == "file":
                    self.store.move_contract_file(item_id, target_folder_id)
                elif kind == "folder":
                    self.store.move_contract_file_folder(item_id, target_folder_id)
            self._mark_documents_changed()
            self.render_contract_files()
        except Exception as exc:
            QMessageBox.warning(self, "Taşıma hatası", str(exc))

    def _add_contract_files(self, file_paths, folder_id=None):
        if not self._ensure_share_can_edit("Belgeler"):
            return
        if not self._ensure_document_access(interactive=True):
            return
        paths = [str(path or "").strip() for path in (file_paths or []) if str(path or "").strip()]
        if not paths:
            return
        added = 0
        duplicates = 0
        failures = []
        ALLOWED = {"pdf", "doc", "docx", "xls", "xlsx", "xlsm", "ppt", "pptx", "png", "jpg", "jpeg", "txt"}
        BLOCKED = {"exe", "bat", "cmd", "ps1", "sh", "msi", "dll", "com", "scr", "vbs", "js"}
        for raw_path in paths:
            path = Path(raw_path)
            try:
                if not path.exists():
                    raise ValueError("Dosya seçilemedi veya bulunamadı.")
                if path.is_dir():
                    raise ValueError("Klasör yüklenemez, lütfen dosya seçin.")
                if not path.is_file():
                    raise ValueError("Lütfen geçerli bir dosya seçin.")
                if not os.access(path, os.R_OK):
                    raise ValueError("Dosya okunamıyor.")
                if self.is_new_contract:
                    # Pending modda: bytes olarak in-memory sakla
                    import mimetypes as _mt
                    ext = path.suffix.lower().lstrip(".")
                    if ext in BLOCKED or ext not in ALLOWED:
                        raise ValueError("Bu dosya türü desteklenmiyor.")
                    from src.config.app_config import MAX_CONTRACT_FILE_SIZE_BYTES
                    size = path.stat().st_size
                    if size > MAX_CONTRACT_FILE_SIZE_BYTES:
                        raise ValueError("Dosya boyutu 120 MB üstünde olamaz.")
                    # Duplicate kontrolü
                    is_dup = any(
                        f["filename"] == path.name and f.get("size_bytes") == size
                        for f in self._pending_doc_files
                    )
                    if is_dup:
                        duplicates += 1
                        continue
                    content_bytes = path.read_bytes()
                    mime = _mt.guess_type(path.name)[0] or "application/octet-stream"
                    new_id = self._pending_doc_next_id
                    self._pending_doc_next_id -= 1
                    self._pending_doc_files.append({
                        "id": new_id,
                        "folder_id": folder_id,
                        "filename": path.name,
                        "file_ext": ext,
                        "mime_type": mime,
                        "size_bytes": size,
                        "content_blob": content_bytes,
                        "note": "",
                        "created_at": "",
                        "updated_at": "",
                        "_source_path": str(path),
                    })
                    added += 1
                else:
                    self.store.add_contract_file(self.ci.platform, self.ci.no, path, self.ci.contract_type, folder_id=folder_id)
                    added += 1
            except Exception as exc:
                message = str(exc)
                if "zaten ekli" in message.lower():
                    duplicates += 1
                else:
                    failures.append(f"{path.name or raw_path}: {message}")
        if added:
            self._mark_documents_changed()
            self.render_contract_files()
        if failures:
            QMessageBox.warning(self, "Bazı dosyalar eklenemedi", "\n".join(failures[:8]))
        if added or duplicates:
            if added and duplicates:
                QMessageBox.information(self, "Belgeler güncellendi", f"{added} dosya eklendi, {duplicates} dosya zaten ekli olduğu için atlandı.")
            elif duplicates and not failures:
                QMessageBox.information(self, "Bu belge zaten ekli", "Seçilen belge zaten ekli.")
            elif added > 1:
                QMessageBox.information(self, "Belgeler eklendi", f"{added} dosya eklendi.")

    def _handle_contract_files_drop(self, file_paths):
        self._add_contract_files(file_paths)

    def _flush_pending_documents_to_db(self):
        """Yeni sözleşme kaydedildikten sonra pending belge ve klasörleri DB'ye yaz."""
        if not self._pending_doc_folders and not self._pending_doc_files:
            return
        # Klasör id eşleme: pending (negatif) id → gerçek DB id
        id_map: dict = {}  # pending_id → real_db_id

        def flush_folder_tree(pending_parent_id, real_parent_id):
            """Pending klasörleri hiyerarşiyle DB'ye yaz."""
            children = [f for f in self._pending_doc_folders if f.get("parent_id") == pending_parent_id]
            for folder in children:
                created = self.store.create_contract_file_folder(
                    self.ci.platform, self.ci.no, self.ci.contract_type,
                    parent_id=real_parent_id, name=folder["name"]
                )
                real_id = int(created.get("id") or 0)
                id_map[folder["id"]] = real_id
                flush_folder_tree(folder["id"], real_id)

        flush_folder_tree(None, None)

        # Dosyaları DB'ye yaz
        for f in self._pending_doc_files:
            pending_folder_id = f.get("folder_id")
            real_folder_id = id_map.get(pending_folder_id) if pending_folder_id else None
            try:
                src = f.get("_source_path")
                if src and Path(src).is_file():
                    self.store.add_contract_file(
                        self.ci.platform, self.ci.no, Path(src),
                        self.ci.contract_type, folder_id=real_folder_id, note=f.get("note", "")
                    )
                else:
                    # Dosya path'i yoksa bytes'tan geçici dosya oluştur
                    import tempfile
                    suffix = f".{f.get('file_ext', 'bin')}"
                    with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
                        tmp.write(f.get("content_blob", b""))
                        tmp_path = tmp.name
                    try:
                        # Orijinal dosya adını koruyarak ekle
                        original_name = f.get("filename", "dosya")
                        target_tmp = Path(tmp_path).parent / original_name
                        Path(tmp_path).rename(target_tmp)
                        self.store.add_contract_file(
                            self.ci.platform, self.ci.no, target_tmp,
                            self.ci.contract_type, folder_id=real_folder_id, note=f.get("note", "")
                        )
                    finally:
                        try:
                            import os
                            if Path(tmp_path).exists():
                                os.unlink(tmp_path)
                        except Exception:
                            pass
            except Exception:
                pass  # Tek dosya hatasını yutma ama devam et (duplicate vs)

        # Pending yapıyı temizle
        self._pending_doc_folders.clear()
        self._pending_doc_files.clear()

    def _import_contract_folders(self, folder_paths, parent_folder_id=None):
        if not self._ensure_share_can_edit("Belgeler"):
            return
        """Windows'tan sürüklenen klasörleri recursive olarak STS içine aktarır."""
        if not self._ensure_document_access(interactive=True):
            return
        ALLOWED_EXTS = {"pdf", "doc", "docx", "xls", "xlsx", "xlsm", "ppt", "pptx", "png", "jpg", "jpeg", "txt"}

        added_files = 0
        skipped_files = 0
        errors = []

        def import_folder(fs_path, db_parent_id):
            """Tek klasörü recursive içe aktar. Klasörü DB'de oluştur, dosyalarını ekle."""
            nonlocal added_files, skipped_files
            fs_path = Path(fs_path)
            folder_name = fs_path.name or "Klasör"
            # Klasörü DB'ye oluştur
            try:
                created = self.store.create_contract_file_folder(
                    self.ci.platform, self.ci.no, self.ci.contract_type,
                    parent_id=db_parent_id, name=folder_name
                )
                db_folder_id = int(created.get("id") or 0)
            except Exception as exc:
                errors.append(f"Klasör oluşturulamadı ({folder_name}): {exc}")
                return

            # Dosyaları ekle
            try:
                children = sorted(fs_path.iterdir(), key=lambda p: (p.is_dir(), p.name.casefold()))
            except Exception as exc:
                errors.append(f"Klasör okunamadı ({folder_name}): {exc}")
                return

            for child in children:
                QApplication.processEvents()
                if child.is_dir():
                    import_folder(child, db_folder_id)
                elif child.is_file():
                    ext = child.suffix.lower().lstrip(".")
                    if ext not in ALLOWED_EXTS:
                        skipped_files += 1
                        continue
                    try:
                        self.store.add_contract_file(
                            self.ci.platform, self.ci.no, child,
                            self.ci.contract_type, folder_id=db_folder_id
                        )
                        added_files += 1
                    except Exception as exc:
                        msg = str(exc)
                        if "zaten ekli" in msg.lower():
                            skipped_files += 1
                        else:
                            errors.append(f"{child.name}: {msg}")

        for fp in (folder_paths or []):
            import_folder(fp, parent_folder_id)
            QApplication.processEvents()

        self._mark_documents_changed()
        self.render_contract_files()

        # Özet mesajı
        parts = []
        if added_files:
            parts.append(f"{added_files} dosya eklendi")
        if skipped_files:
            parts.append(f"{skipped_files} desteklenmeyen/zaten ekli dosya atlandı")
        summary = ", ".join(parts) if parts else "Eklenecek dosya bulunamadı."
        if errors:
            summary += f"\n\nHatalar ({len(errors)}):\n" + "\n".join(errors[:5])
            QMessageBox.warning(self, "Klasör İçe Aktarma", summary)
        elif added_files or skipped_files:
            QMessageBox.information(self, "Klasör İçe Aktarıldı", summary)

    def add_contract_file(self):
        self._pick_contract_files()

    def open_contract_file(self, file_id: int):
        if not self._ensure_document_access(interactive=True):
            return
        try:
            filename, _mime, content = self.store.get_contract_file_bytes(file_id)
            suffix = Path(filename).suffix
            temp = tempfile.NamedTemporaryFile(prefix="sts_contract_", suffix=suffix, delete=False)
            try:
                temp.write(content)
            finally:
                temp.close()
            QDesktopServices.openUrl(QUrl.fromLocalFile(temp.name))
        except Exception as exc:
            QMessageBox.warning(self, "Belge açılamadı", str(exc))

    def export_contract_file(self, file_id: int):
        if not self.require_permission_ui("export_data", "Belge Dışa Aktar"):
            return
        if not self._ensure_document_access(interactive=True):
            return
        try:
            filename, _mime, _content = self.store.get_contract_file_bytes(file_id)
            target, _ = QFileDialog.getSaveFileName(self, "Belgeyi Dışa Aktar", filename)
            if not target:
                return
            self.store.export_contract_file(file_id, target)
            QMessageBox.information(self, "Belge dışa aktarıldı", "Belge başarıyla dışa aktarıldı.")
        except Exception as exc:
            QMessageBox.warning(self, "Belge dışa aktarılamadı", str(exc))

    def delete_contract_file(self, file_id: int):
        if not self._ensure_share_can_edit("Belgeler"):
            return
        if not self._ensure_document_access(interactive=True):
            return
        self._begin_side_meta_modal_action()
        try:
            if not ask_yes_no(self, "Belgeyi Sil", "Belge STS dosyasından silinsin mi? Orijinal dosyaya dokunulmaz."):
                return
            try:
                if self.is_new_contract:
                    self._pending_doc_files = [f for f in self._pending_doc_files if f["id"] != file_id]
                else:
                    self.store.delete_contract_file(file_id)
                self._mark_documents_changed()
                self.render_contract_files()
            except Exception as exc:
                QMessageBox.warning(self, "Belge silinemedi", str(exc))
        finally:
            self._end_side_meta_modal_action()

    def show_contract_file_button_menu(self, file_id: int, button):
        if not self._ensure_share_can_edit("Belgeler"):
            return
        if not self._ensure_document_access(interactive=True):
            return
        menu = QMenu(self)
        menu.addAction("Aç", lambda: self.open_contract_file(file_id))
        menu.addAction("Dışa Aktar", lambda: self.export_contract_file(file_id))
        menu.addSeparator()
        menu.addAction("Sil", lambda: self.delete_contract_file(file_id))
        menu.exec(button.mapToGlobal(QPoint(0, button.height())))

    def open_tag_assign_dialog(self):
        if not self._ensure_share_can_edit("Etiketler"):
            return
        dlg = TagAssignDialog(self.store, self.contract_tags, self)
        if not dlg.exec() or not dlg.result:
            return
        existing_map = {self._tag_key(str(t.get("name", ""))): dict(t) for t in self.contract_tags}
        for t in dlg.result:
            key = self._tag_key(str(t.get("name", "")))
            if not key:
                continue
            existing_map[key] = {
                "name": str(t.get("name", "")).strip(),
                "color": str(t.get("color", "#3B82F6")),
                "note": str(t.get("note", "")).strip(),
            }
        self.contract_tags = list(existing_map.values())
        self._set_dirty()
        self.render_contract_tags()

    def remove_contract_tag(self, tag_name: str):
        if not self._ensure_share_can_edit("Etiketler"):
            return
        key = self._tag_key(tag_name)
        self.contract_tags = [t for t in self.contract_tags if self._tag_key(str((t or {}).get("name", ""))) != key]
        self._set_dirty()
        self.render_contract_tags()

    def add_system(self):
        if not self._ensure_share_can_edit():
            return
        dlg = MultiSystemDialog(
            self.store,
            self.ci.platform,
            contract_t0_date=str(self.ci.t0_date or ""),
            existing_names=[s.name for s in self.systems],
            parent=self,
            events_provider=self.date_picker_events,
        )
        if dlg.exec() and dlg.result:
            first_name = ""
            for sys_info in dlg.result:
                if not first_name:
                    first_name = sys_info.name
                self.systems.append(sys_info)
                self._set_dirty()
                self.deliveries[sys_info.name] = []
            self.selected_system = first_name or self.selected_system
            self.expanded_delivery_index = None
            self.refresh()

    def edit_system(self):
        if not self._ensure_share_can_edit():
            return
        r = self.system_list.currentRow()
        if r < 0 or r >= len(self.systems):
            QMessageBox.warning(self, "Sistem yok", "Düzenlemek için bir sistem seçin.")
            return
        self.sync_summary_to_system()
        current = self.systems[r]
        old_name = current.name

        # pre_selected: gercekten kullanilan bilesenleri sec (qty>0 veya teslimatta gecen)
        pre_selected = self._component_display_keys(current)
        dlg = SystemDialog(
            self.store,
            self.ci.platform,
            default_name=current.name,
            parent=self,
            existing_system=current,
            edit_mode=True,
            pre_selected=pre_selected,
            default_t0_date=str(self.ci.t0_date or ""),
            events_provider=self.date_picker_events,
        )
        if not dlg.exec() or not dlg.result:
            return

        updated = dlg.result
        new_name = updated.name
        for i, s in enumerate(self.systems):
            if i != r and s.name.strip().lower() == new_name.strip().lower():
                QMessageBox.warning(self, "Çakışma", f"'{new_name}' adında başka bir sistem var.")
                return

        current.name = new_name
        removed_components = set(getattr(updated, "removed_components", set()) or set())
        current.components = {k: v for k, v in dict(updated.components).items() if k not in removed_components}
        current.component_notes = {k: v for k, v in (getattr(current, "component_notes", {}) or {}).items() if k in current.components}
        current.t0_date = updated.t0_date
        current.t0_months = updated.t0_months
        current.completion_date = updated.completion_date
        current.status = updated.status
        current.acceptance_date = updated.acceptance_date

        # Sistem adı değiştiyse teslimat anahtarını da taşı.
        if old_name != new_name:
            old_deliveries = self.deliveries.pop(old_name, [])
            self.deliveries[new_name] = old_deliveries

        # Bileşen seti değiştiyse teslimat satırlarını yeni sete hizala.
        # Kaldırılan bileşenler kabullerin planned/delivered sözlüklerinden tamamen çıkarılır;
        # böylece arayüzde görünmez ve Excel'e 0 yerine boş hücre olarak yazılır.
        comps = list(current.components.keys())
        comp_set = set(comps)
        self.deliveries.setdefault(new_name, [])
        for d in self.deliveries.get(new_name, []):
            d.planned = {k: max(as_number((d.planned or {}).get(k, 0)), 0) for k in comps}
            d.delivered = {k: max(as_number((d.delivered or {}).get(k, 0)), 0) for k in comps}
            for k in comps:
                if d.delivered[k] > d.planned[k]:
                    d.delivered[k] = d.planned[k]
                if d.planned[k] > current.components.get(k, 0):
                    d.planned[k] = current.components.get(k, 0)
                    d.delivered[k] = min(d.delivered[k], d.planned[k])
            for removed_key in set((d.planned or {}).keys()) - comp_set:
                d.planned.pop(removed_key, None)
            for removed_key in set((d.delivered or {}).keys()) - comp_set:
                d.delivered.pop(removed_key, None)

        self.selected_system = new_name
        self.expanded_delivery_index = None
        self.refresh()
        self.system_list.setCurrentRow(r)
        self.refresh_right()
        self._set_dirty()

    def delete_system(self):
        if not self._ensure_share_can_edit():
            return
        r = self.system_list.currentRow()
        if r >= 0:
            name = self.systems[r].name
            self.systems.pop(r)
            self._set_dirty()
            self.deliveries.pop(name, None)
            self.expanded_delivery_index = None
            self.refresh()

    def add_delivery(self):
        if not self._ensure_share_can_edit():
            return
        from src.ui.contract import work_window_deliveries as cw_deliveries
        cw_deliveries.add_delivery(self)

    def current_system(self) -> Optional[SystemInfo]:
        r = self.system_list.currentRow()
        if 0 <= r < len(self.systems):
            return self.systems[r]
        return None

    def select_system(self, idx):
        self.expanded_delivery_index = None

        if 0 <= idx < len(self.systems):
            self.selected_system = self.systems[idx].name

        self._populate_system_list(keep_row=idx)
        self.refresh_right()

    def _is_sd_type(self, contract_type: str) -> bool:
        return bool(re.match(r"^SD-\d+$", str(contract_type or "").strip().upper()))

    def _sd_sort_key(self, text: str):
        raw = str(text or "").strip().upper()
        m = re.match(r"^SD[-_ ]?(\d+)$", raw)
        if m:
            return (0, int(m.group(1)))
        parts = re.split(r"(\d+)", raw.lower())
        return (1, [int(p) if p.isdigit() else p for p in parts])

    def _context_key(self, ci: Optional[ContractInfo] = None) -> Tuple[str, str, str]:
        c = ci or self.ci
        return (
            str(getattr(c, "platform", "") or "").strip(),
            str(getattr(c, "no", "") or "").strip(),
            str(getattr(c, "contract_type", "") or "").strip(),
        )

    def _cache_current_context(self):
        if getattr(self, "_refreshing_contract_context", False):
            return
        key = self._context_key()
        if not all(key):
            return
        self.sync_summary_to_system()
        self._context_cache[key] = {
            "ci": copy.deepcopy(self.ci),
            "systems": copy.deepcopy(self.systems),
            "deliveries": copy.deepcopy(self.deliveries),
            "deleted_delivery_systems": set(self._deleted_delivery_systems),
            "tags": copy.deepcopy(self.contract_tags),
            "original_platform": str(self.original_platform or ""),
            "original_contract_no": str(self.original_contract_no or ""),
            "original_contract_type": str(self.original_contract_type or ""),
            "original_entry_start_row": int(self.original_entry_start_row or 0),
        }

    def _load_cached_context(self, key: Tuple[str, str, str]) -> bool:
        ctx = self._context_cache.get(key)
        if not ctx:
            return False
        self.ci = copy.deepcopy(ctx["ci"])
        self.systems = copy.deepcopy(ctx.get("systems") or [])
        self.deliveries = copy.deepcopy(ctx.get("deliveries") or {})
        self._deleted_delivery_systems = set(ctx.get("deleted_delivery_systems") or set())
        self.contract_tags = copy.deepcopy(ctx.get("tags") or [])
        self.original_platform = str(ctx.get("original_platform") or self.ci.platform or "")
        self.original_contract_no = str(ctx.get("original_contract_no") or self.ci.no or "")
        self.original_contract_type = str(ctx.get("original_contract_type") or self.ci.contract_type or "")
        self.original_entry_start_row = int(ctx.get("original_entry_start_row") or getattr(self.ci, "entry_start_row", 0) or 0)
        return True

    def _family_context_rows(self) -> List[dict]:
        platform = str(self.ci.platform or "").strip()
        no = str(self.ci.no or "").strip()
        rows = []
        try:
            rows = [
                dict(it) for it in self.store.list_main_contracts(platform)
                if str(it.get("no", "") or "").strip() == no
            ]
        except Exception:
            rows = []
        seen = {
            (str(it.get("platform", "") or "").strip(), str(it.get("no", "") or "").strip(), str(it.get("type", "") or "").strip())
            for it in rows
        }
        for key, ctx in self._context_cache.items():
            p, n, t = key
            if p != platform or n != no or key in seen:
                continue
            ci = ctx.get("ci")
            if not ci:
                continue
            rows.append({
                "platform": p,
                "no": n,
                "type": t,
                "type_display": t,
                "row": int(getattr(ci, "entry_start_row", 0) or ctx.get("original_entry_start_row") or 0),
                "is_main": not self._is_sd_type(t),
                "_cache_key": key,
            })
        return rows

    def _next_sd_code_for_family(self, platform: str, no: str) -> str:
        used: set[int] = set()
        try:
            for item in self.store.list_main_contracts(platform):
                if str(item.get("no", "") or "").strip() != no:
                    continue
                m = re.match(r"^SD-(\d+)$", str(item.get("type", "") or "").strip().upper())
                if m:
                    used.add(int(m.group(1)))
        except Exception:
            try:
                base_next = self.store.next_sd_code(platform, no)
                m = re.match(r"^SD-(\d+)$", str(base_next or "").strip().upper())
                if m and int(m.group(1)) > 1:
                    used.update(range(1, int(m.group(1))))
            except Exception:
                pass
        for key in self._context_cache.keys():
            p, n, t = key
            if p == platform and n == no:
                m = re.match(r"^SD-(\d+)$", str(t or "").strip().upper())
                if m:
                    used.add(int(m.group(1)))
        n = 1
        while n in used:
            n += 1
        return f"SD-{n}"

    def _drop_context_cache_key(self, key: Tuple[str, str, str]):
        if key and all(key):
            self._context_cache.pop(key, None)

    def _drop_deleted_context_cache(self, info: dict):
        deleted = list((info or {}).get("deleted_contracts") or [])
        if not deleted and info:
            deleted = [{
                "platform": info.get("platform"),
                "contract_no": info.get("contract_no"),
                "contract_type": info.get("contract_type") or self.original_contract_type,
            }]
        for item in deleted:
            key = (
                str(item.get("platform", "") or "").strip(),
                str(item.get("contract_no", "") or "").strip(),
                str(item.get("contract_type", "") or "").strip(),
            )
            self._drop_context_cache_key(key)

    def refresh_sd_sidebar(self):
        if not hasattr(self, "sd_list"):
            return
        if not getattr(self, "_refreshing_contract_context", False):
            self._cache_current_context()
        current_key = self._context_key()
        current_row = int(getattr(self.ci, "entry_start_row", 0) or 0)
        current_type = str(getattr(self.ci, "contract_type", "") or "").strip()
        self.sd_list.blockSignals(True)
        self.sd_list.clear()
        family = self._family_context_rows()
        main_rows = [it for it in family if bool(it.get("is_main"))]
        sd_rows = [it for it in family if not bool(it.get("is_main"))]
        sd_rows.sort(key=lambda it: self._sd_sort_key(str(it.get("type", "") or "")))

        def add_item(label: str, row_data: dict, active: bool = False):
            item = QListWidgetItem(label)
            item.setTextAlignment(Qt.AlignCenter)
            item.setData(Qt.UserRole, dict(row_data or {}))
            self.sd_list.addItem(item)
            if active:
                self.sd_list.setCurrentItem(item)

        if main_rows:
            main = dict(main_rows[0])
            main_key = tuple(main.get("_cache_key") or (main.get("platform"), main.get("no"), main.get("type")))
            add_item("SÖZ\nANA", main, main_key == current_key or (current_row and int(main.get("row") or 0) == current_row))
        else:
            platform, no, _t = current_key
            add_item("SÖZ\nANA", {
                "platform": platform, "no": no, "row": 0, "type": "Ana Sözleşme", "is_main": True,
                "_cache_key": (platform, no, "Ana Sözleşme"),
            }, not self._is_sd_type(current_type))
        for sd in sd_rows:
            label = str(sd.get("type", "SD") or "SD")
            sd_key = tuple(sd.get("_cache_key") or (sd.get("platform"), sd.get("no"), sd.get("type")))
            add_item(label, sd, sd_key == current_key or (current_row and int(sd.get("row") or 0) == current_row))
        self.sd_list.blockSignals(False)
        self.add_sd_btn.setEnabled(bool(current_key[0] and current_key[1]))

    def on_sd_nav_changed(self, row: int):
        if getattr(self, "_refreshing_contract_context", False):
            return
        if row < 0 or not hasattr(self, "sd_list"):
            return
        item = self.sd_list.item(row)
        data = dict(item.data(Qt.UserRole) or {}) if item else {}
        key = tuple(data.get("_cache_key") or (data.get("platform"), data.get("no"), data.get("type")))
        if key == self._context_key():
            return
        self.switch_contract_context(data)

    def switch_contract_context(self, item: dict):
        self._cache_current_context()
        key = tuple(item.get("_cache_key") or (item.get("platform"), item.get("no"), item.get("type")))
        start_row = int(item.get("row") or 0)
        self._refreshing_contract_context = True
        try:
            if key in self._context_cache and self._load_cached_context(key):
                pass
            else:
                platform = str(item.get("platform") or self.ci.platform or "").strip()
                no = str(item.get("no") or self.ci.no or "").strip()
                if not platform or not no or start_row <= 0:
                    return
                self.set_busy_overlay(True, "Sözleşme detayı yükleniyor...", 25)
                try:
                    ci, systems, deliveries = self.store.load_contract_structure(platform, no, start_row=start_row)
                finally:
                    self.set_busy_overlay(False)
                if not ci:
                    QMessageBox.warning(self, "Bulunamadı", "Seçilen sözleşme/SD detayı okunamadı.")
                    return
                self.ci = ci
                self.original_platform = str(ci.platform or "")
                self.original_contract_no = str(ci.no or "")
                self.original_contract_type = str(ci.contract_type or "")
                self.original_entry_start_row = int(getattr(ci, "entry_start_row", 0) or 0)
                self.systems = systems or []
                self.deliveries = deliveries or {}
                self.contract_tags = self.store.load_contract_tags(
                    self.original_platform, self.original_contract_no, self.original_contract_type
                )
            self.expanded_delivery_index = None
            self.refresh_contract_header()
            self.render_contract_tags()
            self.refresh()
        finally:
            self._refreshing_contract_context = False

    def add_sd_contract(self):
        self._cache_current_context()
        platform = str(self.ci.platform or "").strip()
        no = str(self.ci.no or "").strip()
        if not platform or not no:
            QMessageBox.warning(self, "Eksik", "SD eklemek için önce ana sözleşme bilgisi bulunmalı.")
            return
        main_info = None
        try:
            main_info = self.store.find_main_contract_info(platform, no)
        except Exception:
            main_info = None
        main_key = (platform, no, "Ana Sözleşme")
        main_ctx = self._context_cache.get(main_key)
        if not main_info and not main_ctx:
            QMessageBox.warning(self, "Ana sözleşme bulunamadı", "SD eklemek için önce ana sözleşme detayında kalıp sistem bilgilerini girin.")
            return
        sd_code = self._next_sd_code_for_family(platform, no)
        sd_key = (platform, no, sd_code)
        if sd_key in self._context_cache:
            self.switch_contract_context({"_cache_key": sd_key, "platform": platform, "no": no, "type": sd_code})
            return
        source_ci = main_ctx.get("ci") if main_ctx else self.ci
        sd_ci = ContractInfo(
            no=no,
            platform=platform,
            user=str((main_info or {}).get("user") or getattr(source_ci, "user", "") or ""),
            users=list(getattr(source_ci, "users", []) or []),
            yi_yd=str((main_info or {}).get("yi_yd") or getattr(source_ci, "yi_yd", "Yİ") or "Yİ"),
            contract_type=sd_code,
            signature_date="",
            t0_date="",
            t0_months=0,
            completion_date="",
            status="Başlanmadı",
            note="",
            entry_start_row=0,
            sd_anchor_start_row=int((main_info or {}).get("block_start") or (main_info or {}).get("row") or 0),
            sd_anchor_end_row=int((main_info or {}).get("block_end") or (main_info or {}).get("row") or 0),
            sd_anchor_platform=platform if main_info else "",
            sd_anchor_no=no if main_info else "",
            responsible_engineer_id=int(getattr(source_ci, "responsible_engineer_id", 0) or 0),
            responsible_engineer_name=str(getattr(source_ci, "responsible_engineer_name", "") or ""),
        )
        dlg = ContractEditDialog(
            self.store,
            sd_ci,
            self,
            title_text="SD Ekleme Tablosu",
            save_text="SD Ekle",
            info_text="SD temel bilgilerini girin. Platform ve sözleşme no ana sözleşmeye bağlıdır; değiştirilemez.",
        )
        if not dlg.exec() or not dlg.result:
            return
        sd_ci = dlg.result
        self._context_cache[sd_key] = {
            "ci": sd_ci,
            "systems": [],
            "deliveries": {},
            "deleted_delivery_systems": set(),
            "tags": [],
            "original_platform": platform,
            "original_contract_no": no,
            "original_contract_type": sd_code,
            "original_entry_start_row": 0,
        }
        self.switch_contract_context({"_cache_key": sd_key, "platform": platform, "no": no, "type": sd_code})
        QMessageBox.information(self, "SD hazır", f"{sd_code} oluşturuldu. Sistem ve teslimatları ekleyip Kaydet'e basın.")

    def _default_acceptance_for(self, sys_info: SystemInfo) -> Optional[DeliveryInfo]:
        planned = {comp: max(as_number(qty), 0) for comp, qty in (sys_info.components or {}).items()}
        planned = {comp: qty for comp, qty in planned.items() if qty > 0.0001}
        if not planned:
            return None
        return DeliveryInfo(
            name="Teslimat 1",
            status="Başlanmadı",
            acceptance_date="",
            note="Sistem kaydedilirken otomatik oluşturuldu.",
            planned_acceptance_date="",
            planned=planned,
            delivered={comp: 0 for comp in planned.keys()},
            t0_date=str(getattr(sys_info, "t0_date", "") or ""),
            t0_months=int(getattr(sys_info, "t0_months", 0) or 0),
            completion_date=str(getattr(sys_info, "completion_date", "") or ""),
        )

    def _norm_status_text(self, status: str) -> str:
        txt = str(status or "").strip().lower()
        repl = {"ı": "i", "İ": "i", "ş": "s", "Ş": "s", "ğ": "g", "Ğ": "g", "ü": "u", "Ü": "u", "ö": "o", "Ö": "o", "ç": "c", "Ç": "c"}
        for a, b in repl.items():
            txt = txt.replace(a, b)
        return re.sub(r"\s+", " ", txt)

    def _delivery_is_delivered(self, delivery: DeliveryInfo) -> bool:
        return self._norm_status_text(getattr(delivery, "status", "")) in {"teslim edildi", "tamamlandi"}

    def _delivery_is_preparing(self, delivery: DeliveryInfo) -> bool:
        return self._norm_status_text(getattr(delivery, "status", "")) == "teslimata hazirlaniyor"

    def _derive_system_status(self, sys_info: SystemInfo, deliveries: List[DeliveryInfo]) -> str:
        if not deliveries:
            return "Başlanmadı"
        component_qty = {name: max(as_number(qty), 0) for name, qty in (sys_info.components or {}).items()}
        delivered_qty = {name: sum(max(as_number((delivery.delivered or {}).get(name, 0)), 0) for delivery in deliveries) for name in component_qty}
        active_components = [name for name, qty in component_qty.items() if qty > 0.0001]
        if active_components and all(delivered_qty.get(name, 0) >= component_qty[name] - 0.0001 for name in active_components):
            return "Teslim Edildi"
        if any(qty > 0.0001 for qty in delivered_qty.values()):
            return "Parçalı Teslimat"
        if any(self._delivery_is_preparing(d) for d in deliveries):
            return "Teslimata Hazırlanıyor"
        return "Başlanmadı"

    def _latest_acceptance_date(self, items: List[object]) -> str:
        latest = None
        for item in items:
            parsed = parse_iso_date(str(getattr(item, "acceptance_date", "") or ""))
            if parsed and (latest is None or parsed > latest):
                latest = parsed
        return latest.isoformat() if latest else ""

    def _system_acceptance_date(self, sys_info: SystemInfo, sys_deliveries: List[DeliveryInfo]) -> str:
        if self._norm_status_text(getattr(sys_info, "status", "")) != "teslim edildi":
            return ""
        return self._latest_acceptance_date(sys_deliveries)

    def _contract_acceptance_date(self, systems: List[SystemInfo]) -> str:
        if not systems:
            return ""
        completed = [
            sys_info for sys_info in systems
            if self._norm_status_text(getattr(sys_info, "status", "")) in {"teslim edildi", "tamamlandi"}
        ]
        if len(completed) != len(systems):
            return ""
        return self._latest_acceptance_date(completed)

    def _apply_derived_statuses(self, ci: ContractInfo, systems: List[SystemInfo], deliveries: Dict[str, List[DeliveryInfo]]):
        for sys_info in systems:
            sys_deliveries = list(deliveries.get(sys_info.name, []) or [])
            sys_info.status = self._derive_system_status(sys_info, sys_deliveries)
            sys_info.acceptance_date = self._system_acceptance_date(sys_info, sys_deliveries)

        ci.acceptance_date = self._contract_acceptance_date(systems)

        system_statuses = [self._norm_status_text(getattr(sys_info, "status", "")) for sys_info in systems]
        if not system_statuses or all(st == "baslanmadi" for st in system_statuses):
            ci.status = "Başlanmadı"
        elif all(st in {"teslim edildi", "tamamlandi"} for st in system_statuses):
            ci.status = "Tamamlandı"
        else:
            ci.status = "Devam ediyor"

    def _prepare_context_for_save(self, ctx: dict) -> Tuple[bool, str]:
        ci = ctx.get("ci")
        systems = ctx.get("systems") or []
        deliveries = ctx.get("deliveries") or {}
        if not ci:
            return False, "Sözleşme bilgisi eksik."
        if not systems:
            self._apply_derived_statuses(ci, systems, deliveries)
            ctx["deliveries"] = deliveries
            ctx["systems"] = systems
            return True, ""
        created_defaults = []
        deleted_delivery_systems = set(ctx.get("deleted_delivery_systems") or set())
        for sys_info in systems:
            if self._system_has_component_quantity(sys_info) and not deliveries.get(sys_info.name) and sys_info.name not in deleted_delivery_systems:
                d = self._default_acceptance_for(sys_info)
                if d:
                    deliveries.setdefault(sys_info.name, []).append(d)
                    created_defaults.append(sys_info.name)
        if created_defaults:
            ctx["deliveries"] = deliveries
            ctx["systems"] = systems
            ctx["_created_default_systems"] = created_defaults
            return False, (
                f"{ci.contract_type}: otomatik Teslimat 1 ekranda oluşturuldu. "
                "Lütfen açılan teslimat ekranını kontrol edip onaylayın; ardından tekrar Kaydet'e basın."
            )
        issues = self._acceptance_coverage_issues(systems, deliveries)
        if issues:
            title, message = self._acceptance_validation_message(issues)
            ctx["_acceptance_validation_title"] = title
            ctx["_acceptance_validation_issues"] = issues
            return False, f"{ci.contract_type}: {message}"
        self._apply_derived_statuses(ci, systems, deliveries)
        ctx["deliveries"] = deliveries
        ctx["systems"] = systems
        return True, ""

    def _save_context_family(self) -> bool:
        self._cache_current_context()
        family = self._family_context_rows()
        keys = []
        for it in family:
            key = tuple(it.get("_cache_key") or (it.get("platform"), it.get("no"), it.get("type")))
            if key in self._context_cache and key not in keys:
                keys.append(key)
        if not keys:
            return False
        keys.sort(key=lambda k: (1 if self._is_sd_type(k[2]) else 0, self._sd_sort_key(k[2])))
        for key in keys:
            ok, msg = self._prepare_context_for_save(self._context_cache[key])
            if not ok:
                ctx = self._context_cache[key]
                created_defaults = list(ctx.pop("_created_default_systems", []) or [])
                validation_title = str(ctx.pop("_acceptance_validation_title", "") or "")
                validation_issues = list(ctx.pop("_acceptance_validation_issues", []) or [])
                self._load_cached_context(key)
                if created_defaults:
                    self.selected_system = created_defaults[0]
                elif validation_issues:
                    self._focus_acceptance_issue(validation_issues)
                self.expanded_delivery_index = None
                self.refresh_contract_header()
                self.render_contract_tags()
                self.refresh()
                QMessageBox.information(
                    self,
                    "Teslimat oluşturuldu" if created_defaults else (validation_title or "Eksik"),
                    msg,
                )
                if created_defaults:
                    QTimer.singleShot(0, lambda name=created_defaults[0]: self._open_first_delivery_for_system(name))
                return True
        main_start = 0
        main_end = 0
        actor = self.store.current_actor()
        with self.store.batch_save():
            for key in keys:
                ctx = self._context_cache[key]
                ci = ctx["ci"]
                if self._is_sd_type(ci.contract_type) and main_start:
                    ci.sd_anchor_start_row = main_start
                    ci.sd_anchor_end_row = main_end or main_start
                    ci.sd_anchor_platform = ci.platform
                    ci.sd_anchor_no = ci.no
                written_start = self.store.write_contract(
                    ci,
                    ctx.get("systems") or [],
                    ctx.get("deliveries") or {},
                    old_contract_no=ctx.get("original_contract_no") or ci.no,
                    old_start_row=ctx.get("original_entry_start_row") or 0,
                )
                ci.entry_start_row = int(written_start or 0)
                if not self._is_sd_type(ci.contract_type):
                    try:
                        info = self.store.find_main_contract_info(ci.platform, ci.no)
                        main_start = int((info or {}).get("block_start") or written_start or 0)
                        main_end = int((info or {}).get("block_end") or main_start or 0)
                    except Exception:
                        main_start = int(written_start or 0)
                        main_end = main_start
                old_key = (
                    str(ctx.get("original_platform") or "").strip(),
                    str(ctx.get("original_contract_no") or "").strip(),
                    str(ctx.get("original_contract_type") or "").strip(),
                )
                new_key = (str(ci.platform or "").strip(), str(ci.no or "").strip(), str(ci.contract_type or "").strip())
                if old_key != new_key and all(old_key):
                    self.store.save_contract_tags(old_key[0], old_key[1], old_key[2], [], actor=actor)
                self.store.save_contract_tags(new_key[0], new_key[1], new_key[2], ctx.get("tags") or [], actor=actor)
                ctx["original_platform"] = new_key[0]
                ctx["original_contract_no"] = new_key[1]
                ctx["original_contract_type"] = new_key[2]
                ctx["original_entry_start_row"] = int(written_start or 0)
        current_key = self._context_key()
        if current_key in self._context_cache:
            self._load_cached_context(current_key)
        QMessageBox.information(self, "Kaydedildi", "Ana sözleşme ve bağlı SD kayıtları Excel'e yazıldı.")
        self._is_dirty = False
        self.accept()
        return True

    def _system_status_kind(self, status: str) -> str:
        norm = self._norm_status_text(status)
        if norm in {"teslim edildi", "tamamlandi"}:
            return "done"
        if norm in {"teslimata hazirlaniyor", "parcali teslimat"}:
            return "progress"
        return "notstarted"

    def _make_system_item_widget(self, sys_info: SystemInfo, selected: bool = False) -> QWidget:
        status = str(getattr(sys_info, "status", "") or "Başlanmadı")
        kind = self._system_status_kind(status)

        card = QFrame()
        card.setObjectName("systemListCard")
        card.setAttribute(Qt.WA_TransparentForMouseEvents, True)

        lay = QVBoxLayout(card)
        lay.setContentsMargins(10, 7, 10, 7)
        lay.setSpacing(2)

        row = QHBoxLayout()
        row.setContentsMargins(0, 0, 0, 0)
        row.setSpacing(8)

        name = QLabel(str(sys_info.name or "Sistem"))
        name.setObjectName("systemItemName")
        name.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        name.setMinimumWidth(0)
        name.setWordWrap(False)

        # Seçili değilse siyah, seçiliyse beyaz.
        name.setStyleSheet(
            "background: transparent; color: #ffffff; font-weight: 900; font-size: 12px;"
            if selected else
            "background: transparent; color: #0f172a; font-weight: 900; font-size: 12px;"
        )

        pill = QLabel(status)
        pill.setObjectName("systemStatusPill")
        pill.setProperty("kind", kind)
        pill.setAlignment(Qt.AlignCenter)
        pill.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        pill.setMinimumWidth(108)
        pill.setMaximumWidth(138)
        pill.setFixedHeight(22)
        pill.setWordWrap(False)

        row.addWidget(name, 1)
        row.addWidget(pill, 0, Qt.AlignRight | Qt.AlignVCenter)
        lay.addLayout(row)

        return card

    def _populate_system_list(self, keep_row: Optional[int] = None):
        current = self.system_list.currentRow() if keep_row is None else keep_row

        target = 0
        if self.systems:
            target = max(0, min(current, len(self.systems) - 1))
            if self.selected_system:
                for idx, sys_info in enumerate(self.systems):
                    if sys_info.name == self.selected_system:
                        target = idx
                        break

        self.system_list.blockSignals(True)
        self.system_list.clear()

        for idx, sys_info in enumerate(self.systems):
            item = self._make_system_list_item(sys_info)
            item.setSizeHint(QSize(0, 58))
            self.system_list.addItem(item)

            is_selected = idx == target
            self.system_list.setItemWidget(
                item,
                self._make_system_item_widget(sys_info, selected=is_selected)
            )

        if self.systems:
            self.system_list.setCurrentRow(target)

        self.system_list.blockSignals(False)

    def refresh_live_statuses(self):
        self._apply_derived_statuses(self.ci, self.systems, self.deliveries)
        self.refresh_contract_header()
        self._populate_system_list()

    def _make_system_list_item(self, sys_info: SystemInfo) -> QListWidgetItem:
        status = str(getattr(sys_info, "status", "") or "Başlanmadı")
        item = QListWidgetItem()
        item.setData(Qt.UserRole, status)
        return item

    def refresh(self):
        self.refresh_sd_sidebar()
        self._apply_derived_statuses(self.ci, self.systems, self.deliveries)
        self.refresh_contract_header()
        self._populate_system_list()
        self.refresh_right()
        if hasattr(self, 'update_timeline_bar'):
            self.update_timeline_bar()

    def _component_display_keys(self, sys_info: Optional[SystemInfo]) -> List[str]:
        """Sistemin tum secili bilesenleri dondur.
        excel_store artik sadece qty > 0 bilesenleri yukledigi icin
        sys_info.components'ta ne varsa kullanici secmis veya deger girmistir.
        qty=0 olsa bile (yeni eklenmis bilesenler) goster ki kullanici deger girebilsin.
        """
        if not sys_info:
            return []
        return list(sys_info.components.keys())

    def sync_summary_to_system(self):
        sys_info = self.current_system()
        if not sys_info:
            return
        for r in range(self.summary.rowCount()):
            comp_item = self.summary.item(r, 0)
            qty_item = self.summary.item(r, 1)
            if not comp_item or not qty_item:
                continue
            comp = comp_item.text()
            sys_info.components[comp] = as_number(qty_item.text())
            note_item = self.summary.item(r, 4)
            note = note_item.text() if note_item else ""
            if not hasattr(sys_info, "component_notes"):
                sys_info.component_notes = {}
            if note:
                sys_info.component_notes[comp] = note
            else:
                sys_info.component_notes.pop(comp, None)

    def on_summary_changed(self, item):
        if self._updating_summary or item.column() not in (1, 4):
            return
        self._set_dirty()
        self.sync_summary_to_system()
        if item.column() == 1:
            self.refresh_system_card_text()
            self.refresh_summary_only()

    def refresh_system_card_text(self):
        r = self.system_list.currentRow()
        if 0 <= r < len(self.systems):
            s = self.systems[r]
            self._populate_system_list(keep_row=r)

    def _parse_iso_date(self, text: str):
        return parse_iso_date(text)

    def _fmt_num(self, v):
        return fmt_num(v)

    def _as_number(self, v):
        return as_number(v)

    def _configure_table(self, table, compact: bool = False):
        return configure_table(table, compact=compact)

    @property
    def _DeliveryDialog(self):
        return DeliveryDialog

    def _show_warning(self, title: str, text: str):
        return QMessageBox.warning(self, title, text)

    def refresh_summary_only(self):
        cw_view.refresh_summary_only(self)

    def update_system_metric_cards(self, sys_info: Optional[SystemInfo]):
        cw_view.update_system_metric_cards(self, sys_info)

    def refresh_right(self):
        cw_view.refresh_right(self)

    def refresh_delivery_table(self):
        from src.ui.contract import work_window_deliveries as cw_deliveries
        cw_deliveries.refresh_delivery_table(self)

    def update_pinned_delivery(self, d: Optional[DeliveryInfo], headers: List[str], comps: List[str]):
        self.pinned_delivery.clear()
        self.pinned_delivery.setColumnCount(len(headers))
        self.pinned_delivery.setHorizontalHeaderLabels(headers)
        if d:
            self.pinned_delivery.setRowCount(1)
            vals = ["▼", d.name, d.status, d.acceptance_date] + [d.planned.get(c, 0) for c in comps]
            for c, v in enumerate(vals):
                it = QTableWidgetItem(fmt_num(v) if isinstance(v, (int, float)) else str(v))
                it.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
                if c == 0:
                    it.setTextAlignment(Qt.AlignCenter)
                self.pinned_delivery.setItem(0, c, it)
        else:
            self.pinned_delivery.setRowCount(0)
        self.pinned_delivery.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        if self.pinned_delivery.columnCount() > 0:
            self.pinned_delivery.horizontalHeader().setSectionResizeMode(0, QHeaderView.Fixed)
            self.pinned_delivery.setColumnWidth(0, 44)

        header_h = self.pinned_delivery.horizontalHeader().height()
        row_h = self.pinned_delivery.rowHeight(0) if self.pinned_delivery.rowCount() else 0
        frame = 2 * self.pinned_delivery.frameWidth()
        total_h = header_h + row_h + frame + (2 if row_h else 0)
        self.pinned_delivery.setMinimumHeight(total_h)
        self.pinned_delivery.setMaximumHeight(total_h)

    def delivery_detail_widget(self, d: DeliveryInfo, comps: List[str], idx: int) -> QWidget:
        from src.ui.contract import work_window_deliveries as cw_deliveries
        return cw_deliveries.delivery_detail_widget(self, d, comps, idx)

    def edit_delivery(self, idx: int):
        from src.ui.contract import work_window_deliveries as cw_deliveries
        cw_deliveries.edit_delivery(self, idx)

    def on_delivery_clicked(self, row, col):
        if col != 0 or row not in self._delivery_row_map:
            return
        idx = self._delivery_row_map[row]
        self.edit_delivery(idx)

    def on_pinned_delivery_clicked(self, row, col):
        pass  # pinned panel kaldirildi

    def _system_has_component_quantity(self, sys_info: SystemInfo) -> bool:
        return any(as_number(v) > 0.0001 for v in (sys_info.components or {}).values())

    def _create_default_acceptance_for_system(self, sys_info: SystemInfo):
        d = self._default_acceptance_for(sys_info)
        if d:
            self.deliveries.setdefault(sys_info.name, []).append(d)
            self._deleted_delivery_systems.discard(sys_info.name)
            self._set_dirty()

    def _open_first_delivery_for_system(self, system_name: str):
        if not system_name:
            return
        self.selected_system = system_name
        self._populate_system_list()
        for idx, sys_info in enumerate(self.systems):
            if sys_info.name == system_name:
                self.system_list.setCurrentRow(idx)
                break
        self.refresh_right()
        if self.deliveries.get(system_name):
            self.edit_delivery(0)

    def _acceptance_coverage_issues(self, systems: List[SystemInfo], deliveries: Dict[str, List[DeliveryInfo]]) -> List[dict]:
        return acceptance_coverage_issues(systems, deliveries)

    def _acceptance_validation_message(self, issues: List[dict]) -> Tuple[str, str]:
        unassigned = [issue for issue in issues if issue.get("kind") == "unassigned"]
        over_delivered = [issue for issue in issues if issue.get("kind") == "over_delivered"]
        delivery_over_planned = [issue for issue in issues if issue.get("kind") == "delivery_over_planned"]
        over_assigned = [issue for issue in issues if issue.get("kind") == "over_assigned"]
        if over_delivered:
            title = "Teslim edilen miktar sözleşme adedini aşıyor"
            intro = "Bazı bileşenlerde teslim edilen miktar sözleşme adedini aşıyor. Kaydetmeden önce miktarları düzeltin."
            details = [f"• {issue['system']} / {issue['component']}: sözleşme {fmt_num(issue['contract_qty'])}, teslim edilen {fmt_num(issue['delivered_qty'])}" for issue in over_delivered]
        elif delivery_over_planned:
            title = "Teslim edilen miktar teslimat adedini aşıyor"
            intro = "Bazı teslimatlarda teslim edilen miktar teslimat adedini aşıyor. Kaydetmeden önce miktarları düzeltin."
            details = [f"• {issue['system']} / {issue['delivery']} / {issue['component']}: teslimat {fmt_num(issue['planned_qty'])}, teslim edilen {fmt_num(issue['delivered_qty'])}" for issue in delivery_over_planned]
        elif over_assigned:
            title = "Teslimat miktarı sistem adedini aşıyor"
            intro = "Bazı bileşenlerde teslimatlara atanan miktar sistem adedini aşıyor. Kaydetmeden önce miktarları düzeltin."
            details = [f"• {issue['system']} / {issue['component']}: sistem {fmt_num(issue['contract_qty'])}, teslimatlar {fmt_num(issue['planned_qty'])}" for issue in over_assigned]
        else:
            title = "Atanmamış bileşenler var"
            intro = "Bu sözleşmede teslimata atanmamış bileşenler bulunuyor. Kaydetmeden önce kalan bileşenleri bir teslimata atayın."
            details = [f"• {issue['system']} / {issue['component']}: {fmt_num(issue['qty'])} adet atanmadı" for issue in unassigned]
        shown = details[:10]
        if len(details) > len(shown):
            shown.append(f"... ve {len(details) - len(shown)} kalem daha")
        return title, intro + "\n\n" + "\n".join(shown)

    def _focus_acceptance_issue(self, issues: List[dict]):
        if not issues:
            return
        system_name = str(issues[0].get("system") or "")
        if not system_name:
            return
        self.selected_system = system_name
        self._populate_system_list()
        for index, sys_info in enumerate(self.systems):
            if sys_info.name == system_name:
                self.system_list.setCurrentRow(index)
                break
        self.refresh_right()

    def _validate_acceptance_totals(self, systems: List[SystemInfo], deliveries: Dict[str, List[DeliveryInfo]]) -> List[str]:
        errors = []
        for sys_info in systems:
            sys_deliveries = deliveries.get(sys_info.name, []) or []
            components = self._component_display_keys(sys_info)
            for comp in components:
                total = max(as_number((sys_info.components or {}).get(comp, 0)), 0)
                planned_sum = sum(max(as_number((d.planned or {}).get(comp, 0)), 0) for d in sys_deliveries)
                delivered_sum = sum(max(as_number((d.delivered or {}).get(comp, 0)), 0) for d in sys_deliveries)
                if planned_sum - total > 0.0001:
                    errors.append(f"{sys_info.name} / {comp}: sistem {fmt_num(total)}, teslimatlar {fmt_num(planned_sum)}")
                if delivered_sum - planned_sum > 0.0001:
                    errors.append(f"{sys_info.name} / {comp}: teslim edilen {fmt_num(delivered_sum)}, teslimat adedi {fmt_num(planned_sum)}")
            for delivery in sys_deliveries:
                for comp, qty in (delivery.delivered or {}).items():
                    planned_qty = max(as_number((delivery.planned or {}).get(comp, 0)), 0)
                    delivered_qty = max(as_number(qty), 0)
                    if delivered_qty - planned_qty > 0.0001:
                        errors.append(f"{sys_info.name} / {delivery.name} / {comp}: teslim edilen, teslimat adedini aşıyor")
        return errors

    def ensure_systems_have_acceptances(self) -> bool:
        missing_systems = [
            sys_info for sys_info in self.systems
            if self._system_has_component_quantity(sys_info)
            and not self.deliveries.get(sys_info.name)
            and sys_info.name not in self._deleted_delivery_systems
        ]
        if not missing_systems:
            return True

        names = "\n".join(f"• {sys_info.name}" for sys_info in missing_systems)
        if not ask_yes_no(
            self,
            "Teslimat eklenmemiş sistem var",
            "Aşağıdaki sistemlere teslimat eklemediniz:\n\n"
            f"{names}\n\n"
            "Onaylarsanız bu sistemlerin içine Teslimat 1 otomatik oluşturulacak ve "
            "sistemdeki tüm bileşen adetleri Teslimat 1'e atanacaktır.\n\n"
            "Onaylamazsanız teslimat eklemeden kaydetmeye izin verilmeyecektir.",
            default_yes=True,
        ):
            QMessageBox.warning(
                self,
                "Teslimat gerekli",
                "Teslimat eklenmemiş sistemler için teslimat oluşturmadan kaydetme yapılamaz.",
            )
            return False

        first_missing_name = missing_systems[0].name
        for sys_info in missing_systems:
            self._create_default_acceptance_for_system(sys_info)
        self.expanded_delivery_index = None
        if first_missing_name:
            self.selected_system = first_missing_name
        self._populate_system_list()
        self.refresh_right()
        QMessageBox.information(
            self,
            "Teslimat oluşturuldu",
            "Otomatik Teslimat 1 kayıtları ekranda oluşturuldu. "
            "Lütfen açılan teslimat ekranını kontrol edip onaylayın; ardından tekrar Kaydet'e basın.",
        )
        QTimer.singleShot(0, lambda name=first_missing_name: self._open_first_delivery_for_system(name))
        return False

    def reject(self) -> None:
        """Kapat butonuna basıldığında değişiklik varsa onay ister."""
        current_key = self._context_key() if hasattr(self, "_context_cache") else ("", "", "")
        if self._is_dirty:
            if not ask_yes_no(
                self,
                "Değişiklikler Kaydedilmedi",
                "Yaptığınız değişiklikler kaydedilmeyecektir.\n\n"
                "Onaylıyor musunuz?",
            ):
                return
        if hasattr(self, "_context_cache") and current_key in self._context_cache:
            ctx = self._context_cache.get(current_key) or {}
            ci = ctx.get("ci")
            if int(getattr(ci, "entry_start_row", 0) or ctx.get("original_entry_start_row") or 0) <= 0:
                self._context_cache.pop(current_key, None)
        super().reject()

    def save_all(self):
        if not self._ensure_share_can_edit("Sözleşme Kaydet"):
            return
        required_permission = "create_contracts" if self.is_new_contract else "edit_contracts"
        if not self.require_permission_ui(required_permission, "Sözleşme Kaydet"):
            return
        # Belgeler/klasörler STS veritabanına anında yazılır. Bu durumda Kaydet'e
        # basıldığında "Değişiklik Yok" uyarısı gösterme; pencereyi normal kapat.
        if not self._is_dirty and not self.is_new_contract:
            if self._documents_changed:
                self._finish_persisted_side_meta_only_save()
                return
            self.accept()
            return

        if self._save_context_family():
            return
        if not self.systems:
            QMessageBox.warning(self, "Eksik", "En az bir sistem ekleyin.")
            return
        self.sync_summary_to_system()
        if not self.ensure_systems_have_acceptances():
            return
        issues = self._acceptance_coverage_issues(self.systems, self.deliveries)
        if issues:
            title, message = self._acceptance_validation_message(issues)
            self._focus_acceptance_issue(issues)
            QMessageBox.warning(self, title, message)
            return
        self._apply_derived_statuses(self.ci, self.systems, self.deliveries)
        # ── Değişiklik tespiti ──────────────────────────────────────────────
        if not self.is_new_contract:
            if self._make_data_snapshot() == self._initial_snapshot:
                if self._documents_changed:
                    self._finish_persisted_side_meta_only_save()
                    return
                self.accept()
                return
        # ───────────────────────────────────────────────────────────────────
        old_key = (
            str(self.original_platform or "").strip(),
            str(self.original_contract_no or "").strip(),
            str(self.original_contract_type or "").strip(),
        )
        new_key = (
            str(self.ci.platform or "").strip(),
            str(self.ci.no or "").strip(),
            str(self.ci.contract_type or "").strip(),
        )
        actor = self.store.current_actor()
        self._pending_contract_save_context = {
            "old_key": old_key,
            "new_key": new_key,
            "actor": actor,
            "tags": [dict(t or {}) for t in self.contract_tags],
        }
        worker = ContractSaveWorker(
            self.store.path,
            "write",
            str(self.ci.platform or ""),
            str(self.ci.no or ""),
            ci=copy.deepcopy(self.ci),
            systems=copy.deepcopy(self.systems),
            deliveries=copy.deepcopy(self.deliveries),
            old_contract_no=self.original_contract_no,
            old_start_row=self.original_entry_start_row,
            actor=actor,
            store=self.store,
        )
        self._start_contract_save_worker(worker, "Sözleşme kaydediliyor...")





def _is_sts_store(store) -> bool:
    return bool(
        store is not None
        and hasattr(store, "db")
        and hasattr(store, "write_users")
        and hasattr(store, "write_components")
    )

def section_label(text):
    l = QLabel(text)
    l.setObjectName("sectionTitle")
    return l

def configure_table(table: QTableWidget, compact: bool = False):
    table.setItemDelegate(CenterTableDelegate(table))
    table.verticalHeader().setVisible(False)
    table.setAlternatingRowColors(True)
    table.setShowGrid(True)
    table.setSelectionBehavior(QTableWidget.SelectRows)
    table.setSelectionMode(QTableWidget.SingleSelection)
    table.horizontalHeader().setMinimumHeight(42 if not compact else 34)
    table.verticalHeader().setDefaultSectionSize(34 if not compact else 28)
    table.setWordWrap(False)

def fill_table(table, rows):
    table.setRowCount(len(rows)); table.setColumnCount(len(rows[0]) if rows else table.columnCount())
    for r,row in enumerate(rows):
        for c,v in enumerate(row): table.setItem(r,c,QTableWidgetItem(str(int(v) if isinstance(v,float) and v==int(v) else v)))
    table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)



from src.ui.dialogs.contract_summary_popup import ContractSummaryPopup

from src.ui.dialogs.workbook_start import WorkbookStartDialog
from src.ui.contract import work_window_view as cw_view

class MainWindow(QMainWindow):
    def __init__(self, store: Optional[ExcelStore] = None, contract_index: Optional[List[dict]] = None, initial_path: Optional[Path] = None, current_staff: Optional[dict] = None):
        super().__init__()
        self.path = Path(initial_path) if initial_path else (store.path if store else Path(DEFAULT_FILE))
        self.store = store
        self.current_staff = current_staff or auth.current_staff
        self.contract_index = contract_index if contract_index is not None else []
        self._tag_color_map_cache: Optional[Dict[str, str]] = None
        self._loading = False
        self._loader_thread: Optional[QThread] = None
        self._loader_worker: Optional[ExcelLoadWorker] = None
        self._sts_loader_thread: Optional[QThread] = None
        self._sts_loader_worker: Optional[STSLoadWorker] = None
        self._export_thread: Optional[QThread] = None
        self._export_worker = None
        self._streaming_index = False
        self._store_loading = False
        self._last_load_timings: Dict[str, float] = {}
        self._index_ready_for_use = False
        self._version_baseline_signature = None
        self.calendar_window: Optional[ContractCalendarWindow] = None
        self._pending_select_platform: Optional[str] = None
        self.selected_platforms: set[str] = set()
        self.multi_platform_mode: bool = False
        self._updating_platform_list = False
        self._platform_checkbox_changed: Optional[str] = None
        self.setWindowTitle(APP_TITLE)
        icon_path = app_icon_path()
        if icon_path.exists():
            self.setWindowIcon(QIcon(str(icon_path)))
        self.resize(1360, 820)
        self.setStyleSheet(STYLE)
        self.all_contract_rows = []
        self._filter_apply_timer = QTimer(self)
        self._filter_apply_timer.setSingleShot(True)
        self._filter_apply_timer.timeout.connect(self.apply_contract_filter)
        self.build()
        self._install_system_admin_shortcut()
        if self.store:
            if not self.contract_index:
                # UI thread'i bloklamamak için hazır store olsa bile indeksleme yükünü worker'a bırak.
                self.start_sts_load(self.store.path) if str(self.store.path).lower().endswith(".sts") else self.start_excel_load(self.store.path)
            else:
                self.refresh(rebuild_index=False)
                self._apply_version_to_ui()
                self._remember_version_baseline()
        else:
            self.set_empty_state()
            self.connection_label.setText("Excel bağlı değil")


    def export_sts_to_excel(self):
        if self._export_thread and self._export_thread.isRunning():
            QMessageBox.information(self, "Excel’e Aktar",
                                    "İşlem devam ediyor, lütfen bekleyin.")
            return
        if not self.require_permission_ui("export_data", "Excel’e Aktar"):
            return
        if not self.store:
            QMessageBox.information(self, "Veri dosyası gerekli", "Önce bir STS veri dosyası açın.")
            return
        if not hasattr(self.store, "export_to_excel"):
            QMessageBox.information(self, "Excel’e Aktar", "Excel’e aktarım yalnızca STS veri dosyalarında desteklenir.")
            return
        active_platform = ""
        selected = set(getattr(self, "selected_platforms", set()))
        if len(selected) == 1:
            active_platform = next(iter(selected))
        from src.ui.dialogs.excel_export_options import ExcelExportDialog
        dlg = ExcelExportDialog(self.store, self, active_platform=active_platform, contract_index=getattr(self, "contract_index", None))
        if not dlg.exec() or not dlg.result_options:
            return
        opts = dict(dlg.result_options)
        try:
            suggested_out = suggested_export_excel_path(getattr(self, "path", None), opts)
        except Exception:
            suggested_out = Path(self.path).with_suffix(".xlsx")
        try:
            out, _ = QFileDialog.getSaveFileName(self, "Excel’e Aktar", str(suggested_out), "Excel (*.xlsx)")
        except Exception as exc:
            QMessageBox.critical(self, "Excel’e Aktar", f"Kayıt penceresi açılamadı:\n{exc}")
            return
        if not out:
            return
        if not str(out).lower().endswith(".xlsx"):
            out = str(out) + ".xlsx"
        from src.workers.export_workers import ExcelExportWorker
        self._export_progress = QProgressDialog("Excel dosyası hazırlanıyor...", "", 0, 100, self)
        self._export_progress.setWindowTitle("Excel’e Aktar")
        self._export_progress.setLabelText("Excel dosyası hazırlanıyor...")
        self._export_progress.setCancelButton(None)
        self._export_progress.setMinimumDuration(0)
        self._export_progress.setAutoClose(False)
        self._export_progress.setAutoReset(False)
        self._export_progress.setValue(0)
        self._export_progress.show()
        QApplication.processEvents()

        db_path = getattr(getattr(self.store, "db", None), "path", None)
        if not db_path:
            self._export_progress.close()
            QMessageBox.critical(self, "Excel’e Aktar", "Veritabanı yolu belirlenemedi.")
            return

        self._export_thread = QThread(self)
        # Worker'a canlı store/bağlantı DEĞİL, dosya yolu verilir; worker kendi
        # salt-okunur sqlite bağlantısını kendi thread'inde açar (thread-safe).
        self._export_worker = ExcelExportWorker(db_path, out, opts)
        self._export_worker.moveToThread(self._export_thread)
        self._export_thread.started.connect(self._export_worker.run)
        # KRİTİK: Sinyaller lambda/yerel fonksiyona DEĞİL, ana penceredeki
        # gerçek metotlara bağlanır. Lambda bağlanırsa Qt onu sinyali yayan
        # thread'de (worker) çalıştırır ve GUI'ye worker thread'inden dokunmak
        # uygulamayı çökertir. Alıcı self (ana thread'de yaşayan QObject)
        # olduğunda Qt bağlantıyı otomatik QueuedConnection yapar; tüm GUI
        # işleri güvenle ana thread'de çalışır.
        self._export_out_path = str(out)
        self._export_opts = dict(opts)
        self._export_worker.progress.connect(self._on_export_progress)
        self._export_worker.finished.connect(self._on_export_finished)
        self._export_worker.failed.connect(self._on_export_failed)
        self._export_worker.finished.connect(self._export_thread.quit)
        self._export_worker.failed.connect(self._export_thread.quit)
        self._export_thread.finished.connect(self._export_worker.deleteLater)
        self._export_thread.finished.connect(self._export_thread.deleteLater)
        self._export_thread.finished.connect(self._clear_export_refs)
        self._export_thread.start()

    # --- Excel export slot'ları: sinyaller worker thread'inden gelir ama bu
    # metotlar ana pencerenin (ana thread) metodu olduğu için Qt bunları
    # QueuedConnection ile ana thread'de çalıştırır. GUI burada güvendedir. ---

    def _clear_export_refs(self):
        self._export_thread = None
        self._export_worker = None

    def _on_export_progress(self, p, m):
        dlg = getattr(self, "_export_progress", None)
        if dlg is not None:
            try:
                dlg.setLabelText(str(m))
                dlg.setValue(int(max(0, min(100, int(p)))))
            except Exception:
                pass

    def _on_export_finished(self, res):
        dlg = getattr(self, "_export_progress", None)
        if dlg is not None:
            try:
                dlg.setValue(100)
                dlg.close()
            except Exception:
                pass
        out = getattr(self, "_export_out_path", "")
        opts = getattr(self, "_export_opts", {})
        # Export işlemi artık uygulama işlem geçmişine ekstra log yazmaz.
        # Sadece kullanıcıya başarı mesajı gösterilir; teşhis/debug log dosyası
        # da worker tarafında kapatıldı.
        box = QMessageBox(self)
        box.setWindowTitle("Excel’e Aktar")
        box.setIcon(QMessageBox.Information)
        box.setText("Excel dosyası oluşturuldu.")
        open_btn = box.addButton("Dosyayı Aç", QMessageBox.AcceptRole)
        box.addButton("Kapat", QMessageBox.RejectRole)
        box.exec()
        if box.clickedButton() is open_btn:
            try:
                import os
                os.startfile(str(out))
            except Exception:
                QMessageBox.information(self, "Excel’e Aktar", f"Dosya konumu:\n{out}")

    def _on_export_failed(self, msg):
        dlg = getattr(self, "_export_progress", None)
        if dlg is not None:
            try:
                dlg.close()
            except Exception:
                pass
        # Export işlemi için debug/işlem geçmişi logu oluşturulmaz.
        QMessageBox.critical(self, "Excel’e Aktar", str(msg))

    def open_database_management(self):
        if not self.store:
            QMessageBox.information(self, "Veri dosyası gerekli", "Önce bir STS veri dosyası açın.")
            return
        if not hasattr(self.store, "database_stats"):
            QMessageBox.information(self, "Database Yönetimi", "Database yönetimi yalnızca STS veri dosyalarında desteklenir.")
            return
        from src.ui.dialogs.database_management import DatabaseManagementDialog
        dlg = DatabaseManagementDialog(self.store, self, current_staff=self.current_staff)
        dlg.exec()


    def open_performance_tracking(self):
        if not self.store:
            QMessageBox.information(self, "Veri dosyası gerekli", "Önce bir STS veri dosyası açın.")
            return

        if not hasattr(self.store, "performance_stats"):
            QMessageBox.information(
                self,
                "Performans Takip",
                "Performans takip ekranı yalnızca STS veri dosyalarında desteklenir."
            )
            return

        from src.ui.dialogs.performance_tracking import PerformanceTrackingDialog
        dlg = PerformanceTrackingDialog(self.store, self)
        dlg.exec()

    def open_activity_logs(self):
        if not self.require_permission_ui("view_action_history", "İşlem Geçmişi"):
            return
        if not self.store:
            QMessageBox.information(self, "Veri dosyası gerekli", "Önce bir STS veri dosyası açın.")
            return
        if not hasattr(self.store, "list_logs"):
            QMessageBox.information(self, "İşlem Geçmişi", "İşlem geçmişi yalnızca STS veri dosyalarında desteklenir.")
            return
        from src.ui.dialogs.activity_logs import ActivityLogDialog
        dlg = ActivityLogDialog(self.store, self)
        dlg.exec()

    def _permission_db(self):
        if self.store is not None and getattr(self.store, "db", None) is not None:
            return self.store.db.conn
        return self.path

    def has_permission(self, permission_code: str) -> bool:
        return auth.has_permission(self.current_staff, permission_code, self._permission_db())

    def require_permission_ui(self, permission_code: str, title: str = "Yetki gerekli") -> bool:
        if self.has_permission(permission_code):
            return True
        QMessageBox.warning(self, "Yetkisiz İşlem", "Bu işlemi yapmak için gerekli yetkiye sahip değilsiniz.")
        return False

    def _install_system_admin_shortcut(self) -> None:
        self.system_admin_shortcut = QShortcut(QKeySequence("Ctrl+Alt+Shift+A"), self)
        self.system_admin_shortcut.setContext(Qt.ApplicationShortcut)
        self.system_admin_shortcut.activated.connect(self.open_system_admin_login_from_shortcut)

    def open_system_admin_login_from_shortcut(self) -> None:
        if not self.store or not self.is_sts_mode():
            return
        if self.current_staff and bool((self.current_staff or {}).get("is_admin")):
            QMessageBox.information(self, "Sistem Yöneticisi", "Zaten sistem yöneticisi oturumundasınız.")
            return

        previous_staff = self.current_staff
        admin_staff = auth.show_system_admin_login_dialog(self._permission_db(), self)
        if not admin_staff:
            self.current_staff = previous_staff
            auth.current_staff = previous_staff
            return

        self.current_staff = admin_staff
        auth.current_staff = admin_staff
        actor_name = str(admin_staff.get("full_name") or admin_staff.get("admin_name") or "Sistem Yöneticisi")
        if self.store is not None and hasattr(self.store, "actor"):
            self.store.actor = actor_name
        self._propagate_current_staff_to_open_windows(admin_staff)
        self._refresh_permission_actions()
        QMessageBox.information(self, "Sistem Yöneticisi", "Sistem yöneticisi oturumu açıldı.")

    def _propagate_current_staff_to_open_windows(self, staff: dict) -> None:
        for widget in QApplication.topLevelWidgets():
            if widget is self or not isinstance(widget, ContractWorkWindow):
                continue
            try:
                widget.current_staff = staff
                if getattr(widget, "store", None) is self.store and hasattr(widget.store, "actor"):
                    widget.store.actor = str(staff.get("full_name") or "Sistem Yöneticisi")
            except Exception:
                pass

    def open_staff_permissions_dialog(self, initial_tab: str = "staffRoles"):
        if not self.store or not self.is_sts_mode():
            QMessageBox.information(self, "Veri dosyası gerekli", "Personel ve yetki yönetimi için önce bir STS veri dosyası açın.")
            return
        required_permission = "manage_roles" if initial_tab == "rolePermissions" else "manage_staff"
        if not self.require_permission_ui(required_permission, "Personel ve Yetki Yönetimi"):
            return
        from src.ui.dialogs.staff_permissions import StaffPermissionsDialog
        dlg = StaffPermissionsDialog(self._permission_db(), self.current_staff, self, initial_tab=initial_tab)
        dlg.permissions_saved.connect(self._refresh_permission_actions)
        dlg.exec()

    def open_user_management(self):
        if not self.store:
            QMessageBox.information(self, "Excel gerekli", "Önce bir Excel veya STS veri dosyası bağlayın.")
            return
        dlg = UserManagerDialog(self.store, self)
        dlg.exec()
        if dlg.changed:
            self.request_refresh(scope="users")

    def open_staff_management(self):
        self.open_staff_permissions_dialog("staffRoles")

    def open_personnel_permissions(self):
        initial_tab = "staffRoles" if self.has_permission("manage_staff") else "rolePermissions"
        self.open_staff_permissions_dialog(initial_tab)

    def open_role_permissions(self):
        self.open_staff_permissions_dialog("rolePermissions")

    def open_delivery_schedule_report(self):
        from src.ui.dialogs.delivery_schedule_report_dialog import DeliveryScheduleReportDialog

        dlg = DeliveryScheduleReportDialog(self, store=self.store)
        dlg.exec()

    def open_platform_delivery_report(self):
        if not self.store:
            QMessageBox.information(self, "Veri dosyası gerekli", "Raporu açmak için önce bir STS veri dosyası açın.")
            return
        dlg = PlatformTeslimatDurumuReportDialog(self, store=self.store)
        dlg.exec()

    def open_usage_guide(self):
        try:
            dlg = UsageGuideDialog(self)
            dlg.exec()
        except Exception as exc:
            traceback.print_exc()
            QMessageBox.warning(self, "Kullanım Kılavuzu", f"Kullanım kılavuzu açılamadı:\n{exc}")


    def build(self):
        root=QWidget(); self.setCentralWidget(root); main=QVBoxLayout(root)
        main.setContentsMargins(8, 8, 8, 8)
        main.setSpacing(8)

        top=QFrame(); top.setObjectName("topbar"); tl=QHBoxLayout(top); tl.setContentsMargins(12, 8, 12, 8); tl.setSpacing(10)
        # Logo: önce SVG, yoksa ico dene
        _svg_logo = Path(__file__).parent / "src" / "ui" / "assets" / "sts_logo.svg"
        _ico_logo = app_icon_path()
        _logo_src = _svg_logo if _svg_logo.exists() else (_ico_logo if _ico_logo.exists() else None)
        if _logo_src:
            logo = QLabel(); logo.setObjectName("appLogo")
            _pix = QPixmap(str(_logo_src))
            if not _pix.isNull():
                logo.setPixmap(_pix.scaled(40, 40, Qt.KeepAspectRatio, Qt.SmoothTransformation))
            logo.setFixedSize(48, 48)
            logo.setAlignment(Qt.AlignCenter)
            tl.addWidget(logo)
        title=QLabel(APP_TITLE); title.setObjectName("appTitle"); tl.addWidget(title)
        self.connection_label=QLabel("✓ Excel bağlı"); self.connection_label.setObjectName("okPill"); tl.addWidget(self.connection_label)
        tl.addStretch()
        self.top_actions_btn = QToolButton()
        self.top_actions_btn.setObjectName("topMenuBtn")
        self.top_actions_btn.setText("☰")
        self.top_actions_btn.setToolTip("Menü")
        self.top_actions_btn.setPopupMode(QToolButton.InstantPopup)
        self.top_actions_menu = QMenu(self.top_actions_btn)
        self.top_actions_menu.setObjectName("topActionsMenu")
        self.top_actions_menu.addAction("Veri Dosyası Değiştir", self.open_file)
        self.top_actions_menu.addAction("Excel’e Aktar", self.export_sts_to_excel)
        reports_menu = self.top_actions_menu.addMenu("Raporlar")
        reports_menu.addAction("Tahmini Teslimat Takvimi", self.open_delivery_schedule_report)
        reports_menu.addAction("Platform Teslimat Özeti", self.open_platform_delivery_report)
        self.top_actions_menu.addAction("Database Yönetimi", self.open_database_management)
        self.top_actions_menu.addAction("Performans Takip", self.open_performance_tracking)
        self.top_actions_menu.addAction("Platform ve Bileşen Yönetimi", self.manage_platforms)
        self.top_actions_menu.addSeparator()
        self.user_management_action = self.top_actions_menu.addAction("Kullanıcı Yönetimi", self.open_user_management)
        self.role_permissions_action = self.top_actions_menu.addAction("Personel ve Yetki Yönetimi", self.open_personnel_permissions)
        self.top_actions_menu.addAction("Etiket Yönetimi", self.manage_tags)
        self.activity_logs_action = self.top_actions_menu.addAction("İşlem Geçmişi", self.open_activity_logs)
        self.top_actions_menu.addSeparator()
        self.top_actions_menu.addAction("Kullanım Kılavuzu", self.open_usage_guide)
        self.top_actions_menu.aboutToShow.connect(self._refresh_permission_actions)
        self.top_actions_btn.setMenu(self.top_actions_menu)
        tl.addWidget(self.top_actions_btn)
        main.addWidget(top, 0)

        strip=QFrame(); strip.setObjectName("alertStrip"); sl=QHBoxLayout(strip); sl.setContentsMargins(12, 10, 12, 10); sl.setSpacing(10)
        today_box = QFrame(); today_box.setObjectName("todayBadge")
        today_l = QVBoxLayout(today_box); today_l.setContentsMargins(12, 8, 12, 8); today_l.setSpacing(1)
        self.today_num = QLabel(str(date.today().day)); self.today_num.setObjectName("todayDay"); self.today_num.setAlignment(Qt.AlignCenter)
        self.today_info = QLabel(""); self.today_info.setObjectName("todayInfo"); self.today_info.setAlignment(Qt.AlignCenter)
        today_l.addWidget(self.today_num); today_l.addWidget(self.today_info)
        sl.addWidget(today_box, 0)

        self.alert_divider1 = QFrame(); self.alert_divider1.setObjectName("stripDivider"); sl.addWidget(self.alert_divider1, 0)

        self.alert_overdue_group = QFrame(); self.alert_overdue_group.setObjectName("alertGroup")
        g1l = QHBoxLayout(self.alert_overdue_group); g1l.setContentsMargins(8, 6, 8, 6); g1l.setSpacing(8)
        i1 = QLabel("⚠"); i1.setObjectName("alertIconRed"); g1l.addWidget(i1)
        t1 = QVBoxLayout(); t1.setContentsMargins(0, 0, 0, 0); t1.setSpacing(0)
        self.overdue_count = QLabel("0"); self.overdue_count.setObjectName("alertCountRed")
        l1 = QLabel("GECİKEN"); l1.setObjectName("alertLabel")
        t1.addWidget(self.overdue_count); t1.addWidget(l1); g1l.addLayout(t1)
        sl.addWidget(self.alert_overdue_group, 0)

        self.alert_critical_group = QFrame(); self.alert_critical_group.setObjectName("alertGroup")
        g2l = QHBoxLayout(self.alert_critical_group); g2l.setContentsMargins(8, 6, 8, 6); g2l.setSpacing(8)
        i2 = QLabel("⏳"); i2.setObjectName("alertIconAmber"); g2l.addWidget(i2)
        t2 = QVBoxLayout(); t2.setContentsMargins(0, 0, 0, 0); t2.setSpacing(0)
        self.critical_count = QLabel("0"); self.critical_count.setObjectName("alertCountAmber")
        l2 = QLabel("60 GÜN İÇİNDE"); l2.setObjectName("alertLabel")
        t2.addWidget(self.critical_count); t2.addWidget(l2); g2l.addLayout(t2)
        sl.addWidget(self.alert_critical_group, 0)

        self.alert_divider2 = QFrame(); self.alert_divider2.setObjectName("stripDivider"); sl.addWidget(self.alert_divider2, 0)
        self.upcoming_label = QLabel("Yaklaşan:"); self.upcoming_label.setObjectName("upcomingLabel")
        sl.addWidget(self.upcoming_label, 0)
        self.upcoming_scroll = QScrollArea(); self.upcoming_scroll.setObjectName("upcomingScroll")
        self.upcoming_scroll.setWidgetResizable(True)
        self.upcoming_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.upcoming_scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.upcoming_host = QWidget()
        self.upcoming_layout = QHBoxLayout(self.upcoming_host)
        self.upcoming_layout.setContentsMargins(0, 0, 0, 0)
        self.upcoming_layout.setSpacing(6)
        self.upcoming_layout.addStretch()
        self.upcoming_scroll.setWidget(self.upcoming_host)
        sl.addWidget(self.upcoming_scroll, 1)

        calb=QPushButton("🗓 Takvim Görünümü"); calb.clicked.connect(self.open_calendar_tracking); sl.addWidget(calb, 0)
        main.addWidget(strip, 0)

        body=QHBoxLayout(); body.setSpacing(8); main.addLayout(body,1)
        left=QFrame(); left.setObjectName("panel"); left.setFixedWidth(350); lv=QVBoxLayout(left); lv.setContentsMargins(0, 0, 0, 0); lv.setSpacing(0)
        platform_head = QWidget(); ph = QHBoxLayout(platform_head); ph.setContentsMargins(12, 8, 12, 8); ph.setSpacing(6)
        h=QLabel("Platformlar"); h.setObjectName("panelTitle"); ph.addWidget(h); ph.addStretch(1)
        self.platform_selection_badge = QLabel(""); self.platform_selection_badge.setObjectName("platformSelectionBadge")
        self.platform_selection_badge.setStyleSheet("QLabel{background:#dbeafe;color:#1d4ed8;border-radius:9px;padding:2px 7px;font-size:11px;font-weight:800;}")
        self.platform_selection_badge.hide(); ph.addWidget(self.platform_selection_badge)
        lv.addWidget(platform_head)
        self.platform_list=QListWidget(); self.platform_list.setObjectName("mainPlatformList"); self.platform_list.setContextMenuPolicy(Qt.CustomContextMenu)
        self.platform_list.itemClicked.connect(self.on_platform_clicked)
        self.platform_list.itemChanged.connect(self._on_platform_item_changed)
        self.platform_list.customContextMenuRequested.connect(self._on_platform_context_menu_requested)
        self._platform_list_delegate = PlatformListDelegate(self.platform_list)
        self.platform_list.setItemDelegate(self._platform_list_delegate)
        lv.addWidget(self.platform_list,1)
        self.platform_info_bar = QFrame(); self.platform_info_bar.setObjectName("platformInfoBar")
        self.platform_info_bar.setStyleSheet("QFrame#platformInfoBar{background:#f8fbff;border-top:1px solid #dbe7f5;} QLabel{color:#64748b;font-size:11px;} QPushButton{background:transparent;border:0;color:#1d4ed8;font-size:11px;font-weight:800;padding:2px 4px;}")
        pi = QHBoxLayout(self.platform_info_bar); pi.setContentsMargins(10, 4, 8, 4); pi.setSpacing(4)
        self.platform_info_label = QLabel(""); pi.addWidget(self.platform_info_label); pi.addStretch(1)
        clear_platforms = QPushButton("temizle"); clear_platforms.clicked.connect(self.clear_platform_selection); pi.addWidget(clear_platforms)
        self.platform_info_bar.hide(); lv.addWidget(self.platform_info_bar)
        new=QPushButton("+ Yeni Sözleşme"); new.setObjectName("newContractBtn"); new.clicked.connect(self.new_contract); new.setMinimumHeight(46); lv.addWidget(new)
        body.addWidget(left, 0)

        right=QFrame(); right.setObjectName("panel"); rv=QVBoxLayout(right); rv.setContentsMargins(12, 10, 12, 10); rv.setSpacing(8)
        self.right_panel = right
        head_row = QHBoxLayout()
        self.right_title=QLabel("Sözleşme Sorgulama"); self.right_title.setObjectName("queryTitle")
        head_row.addWidget(self.right_title)
        self.search_input = QLineEdit(); self.search_input.setPlaceholderText("Sözleşme no, kullanıcı veya durum ara..."); self.search_input.textChanged.connect(self.schedule_apply_contract_filter)
        head_row.addWidget(self.search_input, 1)
        self.clear_col_filters_btn = QPushButton("Filtreleri Temizle")
        self.clear_col_filters_btn.setObjectName("secondary")
        self.clear_col_filters_btn.clicked.connect(self.clear_query_filters)
        head_row.addWidget(self.clear_col_filters_btn, 0)
        hint = QLabel("Sözleşmeye çift tıklayınca detay ekranı açılır."); hint.setObjectName("muted"); head_row.addWidget(hint)
        rv.addLayout(head_row)

        self.filter_bar = QFrame()
        self.filter_bar.setObjectName("filterBar")
        fb = QHBoxLayout(self.filter_bar)
        fb.setContentsMargins(8, 6, 8, 6)
        fb.setSpacing(8)
        self.filter_type = QComboBox()
        self.filter_type.addItem("Tüm Türler", "")
        self.filter_type.currentIndexChanged.connect(self.schedule_apply_contract_filter)
        fb.addWidget(self.filter_type, 0)
        self._date_from_active = False
        self._date_to_active = False
        fb.addWidget(QLabel("Sıralama:"))
        self.sort_combo = QComboBox()
        self.sort_combo.addItem("Varsayılan", "default")
        self.sort_combo.addItem("Sözleşme No (Artan)", "no_asc")
        self.sort_combo.addItem("Sözleşme No (Azalan)", "no_desc")
        self.sort_combo.addItem("Termin Tarihi (Erken)", "date_asc")
        self.sort_combo.addItem("Termin Tarihi (Geç)", "date_desc")
        self.sort_combo.addItem("Kalan Gün (Artan)", "days_asc")
        self.sort_combo.addItem("Kalan Gün (Azalan)", "days_desc")
        self.sort_combo.addItem("Kullanıcı (A-Z)", "user_asc")
        self.sort_combo.addItem("Kullanıcı (Z-A)", "user_desc")
        self.sort_combo.currentIndexChanged.connect(self.schedule_apply_contract_filter)
        fb.addWidget(self.sort_combo, 0)
        self.clear_filters_btn = QPushButton("Temizle")
        self.clear_filters_btn.setObjectName("secondary")
        self.clear_filters_btn.clicked.connect(self.clear_query_filters)
        fb.addWidget(self.clear_filters_btn, 0)
        fb.addStretch()
        self.filter_bar.setVisible(False)
        self.contract_table=QTableWidget(0,9)
        self.contract_table.setObjectName("contractTable")
        # Yatay scroll yok — sütunlar her zaman tablo içinde kalır
        self.contract_table.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        # Excel gibi sutun filtresi
        self._filter_header = FilterableHeaderView(Qt.Horizontal, self.contract_table)
        self._filter_header.filterChanged.connect(self.schedule_apply_contract_filter)
        self.contract_table.setHorizontalHeader(self._filter_header)
        self.contract_table.setHorizontalHeaderLabels(["Platform", "Sözleşme Türü", "Sözleşme No", "Kullanıcı", "Durum", "Termin Tarihi", "Kalan Gün", "Etiketler", "Özet"])
        self.contract_table._filter_col_keys = ["platform", "type", "no", "user", "status", "date", "days", "tags", "summary"]
        self.contract_table._sort_mode = 'default'  # siralama modu
        # Tüm sütunlar Stretch — her zaman ekranı doldurur, yatay kaymaz.
        # Özet sabit 72px sağ kenarda.
        _hh = self.contract_table.horizontalHeader()
        _hh.setSectionResizeMode(QHeaderView.Stretch)    # hepsi orantılı dolar
        _hh.setSectionResizeMode(COL_SUMMARY, QHeaderView.Fixed)   # Özet sabit
        self.contract_table.setColumnWidth(COL_SUMMARY, 72)
        self.contract_table.verticalHeader().setVisible(False)
        self.contract_table.cellDoubleClicked.connect(self.open_selected_contract)
        self.contract_table.cellClicked.connect(self._on_contract_cell_clicked)
        self.contract_table.viewport().installEventFilter(self)
        rv.addWidget(self.contract_table,1)
        self.query_logo_bg = QLabel(self.contract_table)
        self.query_logo_bg.setObjectName("logoWatermark")
        self.query_logo_bg.setStyleSheet("background: transparent;")
        self.query_logo_bg.setAlignment(Qt.AlignCenter)
        # Dekoratif logo sadece arka plan davranışı göstermeli; tablo etkileşimini
        # ve hücrelerin okunabilirliğini hiçbir durumda engellememeli.
        self.query_logo_bg.setAttribute(Qt.WA_TransparentForMouseEvents, True)
        self.query_logo_bg.hide()

        logo_opacity = QGraphicsOpacityEffect(self.query_logo_bg)
        logo_opacity.setOpacity(0.55)
        self.query_logo_bg.setGraphicsEffect(logo_opacity)
        self._query_logo_source: Optional[QPixmap] = None
        self.contract_table.verticalScrollBar().valueChanged.connect(lambda _v: self.position_query_logo_background())
        self.contract_table.horizontalScrollBar().valueChanged.connect(lambda _v: self.position_query_logo_background())
        body.addWidget(right,1)
        self.build_loading_overlay()
        self.index_progress_badge = QLabel(self.centralWidget())
        self.index_progress_badge.setObjectName("miniProgressPill")
        self.index_progress_badge.setAlignment(Qt.AlignCenter)
        self.index_progress_badge.setText("Excel %0")
        self.index_progress_badge.hide()
        self.index_progress_badge.raise_()
        self._send_query_logo_to_back()

    def update_connection_badge(self, mode: str):
        m = str(mode or "").strip().lower()
        if m == "ok":
            self.connection_label.setText("✓ Excel bağlı")
            self.connection_label.setProperty("status", "ok")
        elif m == "loading":
            self.connection_label.setText("Excel analiz ediliyor")
            self.connection_label.setProperty("status", "loading")
        else:
            self.connection_label.setText("Excel bağlı değil")
            self.connection_label.setProperty("status", "bad")
        st = self.connection_label.style()
        st.unpolish(self.connection_label)
        st.polish(self.connection_label)
        self.connection_label.update()

    def _apply_version_to_ui(self):
        try:
            from src.services.version_manager import read_version
            ver = read_version(self.store)
            if ver:
                workbook_name = Path(getattr(self.store, "path", self.path)).stem
                label_parts = [part for part in [workbook_name] if part]
                if ver.lower() not in workbook_name.lower():
                    label_parts.append(f"[{ver}]")
                self.setWindowTitle(f"{APP_TITLE}  [{ver}]")
                self.connection_label.setText(f"✓ Excel bağlı  {' '.join(label_parts) or ver}")
                self.connection_label.setProperty("status", "ok")
                st = self.connection_label.style()
                st.unpolish(self.connection_label)
                st.polish(self.connection_label)
                self.connection_label.update()
        except Exception:
            pass

    def _excel_file_signature(self):
        try:
            path = Path(getattr(self.store, "path", self.path))
            st = path.stat()
            return (str(path.resolve()), int(st.st_mtime_ns), int(st.st_size))
        except Exception:
            return None

    def _remember_version_baseline(self) -> None:
        self._version_baseline_signature = self._excel_file_signature()

    def _workbook_changed_since_load(self) -> bool:
        baseline = getattr(self, "_version_baseline_signature", None)
        current = self._excel_file_signature()
        return bool(baseline and current and current != baseline)

    def closeEvent(self, event: QCloseEvent) -> None:
        """Kapanışta değişiklik varsa dosya adını standart versiyon formatına taşır."""
        if getattr(self, "store", None) and self._workbook_changed_since_load():
            try:
                # STS dosyalarında tek aktif dosya korunur: mevcut dosya yeniden adlandırılır.
                if self.is_sts_mode():
                    new_path = rename_sts_file_to_next_version(self.store, getattr(self.store, "path", self.path))
                    if new_path:
                        self.path = Path(new_path)
                        self._remember_version_baseline()
                # Eski Excel modu için mevcut version_manager davranışı korunur.
                elif getattr(self.store, "wb", None):
                    from src.services.version_manager import bump_version, save_store_as_versioned_file
                    new_ver = bump_version(self.store)
                    self.path = save_store_as_versioned_file(self.store, new_ver)
                    self._remember_version_baseline()
            except Exception:
                pass
        super().closeEvent(event)

    def build_loading_overlay(self):
        self.loading_overlay = QFrame(self.centralWidget())
        self.loading_overlay.setStyleSheet("QFrame { background: rgba(248, 251, 255, 0.82); }")
        self.loading_overlay.hide()
        self.loading_overlay.raise_()

        self.loading_card = QFrame(self.loading_overlay)
        self.loading_card.setStyleSheet(
            "QFrame { background: rgba(255,255,255,0.96); border: 1px solid #d8e2ed; border-radius: 12px; }"
            "QLabel { background: transparent; }"
        )
        lay = QVBoxLayout(self.loading_card)
        lay.setContentsMargins(20, 20, 20, 20)
        lay.setSpacing(10)
        self.loading_label = QLabel("Analiz ediliyor...")
        self.loading_label.setObjectName("mainTitle")
        self.loading_label.setAlignment(Qt.AlignCenter)
        self.loading_progress = QProgressBar()
        self.loading_progress.setRange(0, 100)
        self.loading_progress.setValue(0)
        self.loading_progress.setTextVisible(True)
        self.loading_progress.setFormat("%p%")
        lay.addWidget(self.loading_label)
        lay.addWidget(self.loading_progress)

    def position_loading_overlay(self):
        if not hasattr(self, "loading_overlay"):
            return
        parent = self.centralWidget()
        if not parent:
            return
        self.loading_overlay.setGeometry(parent.rect())
        w, h = 460, 150
        x = max((self.loading_overlay.width() - w) // 2, 0)
        y = max((self.loading_overlay.height() - h) // 2, 0)
        self.loading_card.setGeometry(x, y, w, h)
        self.loading_overlay.raise_()

    def position_query_logo_background(self):
        """Sorgulama tablosunda sadece seçili platform logolarını watermark olarak konumlandırır.

        Platform seçimi yoksa varsayılan/Baykar logosu gösterilmez.
        """
        try:
            if not hasattr(self, "query_logo_bg") or not hasattr(self, "contract_table"):
                return

            if not getattr(self, "_query_logo_source", None) or self._query_logo_source.isNull():
                self.query_logo_bg.clear()
                self.query_logo_bg.hide()
                return

            vp = self.contract_table.viewport()
            rect = vp.geometry()

            if rect.width() <= 0 or rect.height() <= 0:
                self.query_logo_bg.hide()
                return

            max_w = int(rect.width() * 0.82)
            max_h = int(rect.height() * 0.62)
            source = self._query_logo_source

            # Logo tablo alanından büyükse küçült; küçükse olduğu gibi bırak.
            if source.width() > max_w or source.height() > max_h:
                scaled = source.scaled(
                    QSize(max_w, max_h),
                    Qt.KeepAspectRatio,
                    Qt.SmoothTransformation,
                )
            else:
                scaled = source

            x = max(rect.x() + (rect.width() - scaled.width()) // 2, 0)
            y = max(rect.y() + (rect.height() - scaled.height()) // 2, 0)

            self.query_logo_bg.setGeometry(x, y, scaled.width(), scaled.height())
            self.query_logo_bg.setPixmap(scaled)
            self.query_logo_bg.show()
            self._send_query_logo_to_back()
        except Exception:
            try:
                if hasattr(self, "query_logo_bg"):
                    self.query_logo_bg.clear()
                    self.query_logo_bg.hide()
            except Exception:
                pass

    def _scale_logo_smart(
        self,
        px: QPixmap,
        max_size: QSize,
        max_upscale: float = 3.0
    ) -> QPixmap:
        """
        Küçük logoları en fazla max_upscale kadar büyütür.
        Büyük logoları max_size içine sığacak şekilde küçültür.
        Oranı her zaman korur.
        """
        if not px or px.isNull():
            return px

        ow = max(px.width(), 1)
        oh = max(px.height(), 1)

        fit_scale = min(
            max_size.width() / ow,
            max_size.height() / oh
        )

        if fit_scale < 1:
            # Logo zaten büyükse küçült.
            scale = fit_scale
        else:
            # Logo küçükse en fazla 3 kat büyüt.
            scale = min(fit_scale, max_upscale)

        nw = max(1, int(ow * scale))
        nh = max(1, int(oh * scale))

        if nw == ow and nh == oh:
            return px

        return px.scaled(
            QSize(nw, nh),
            Qt.KeepAspectRatio,
            Qt.SmoothTransformation
        )

    def _has_permission_context(self) -> bool:
        return bool(self.current_staff and self.is_sts_mode())

    def _permission_action_visible(self, permission_code: str) -> bool:
        # Excel/başlangıç modunda personel oturumu yoktur; menü öğelerini saklamak
        # kullanıcıda "menüden silinmiş" hissi yaratır. Gerçek STS oturumu varsa
        # görünürlük merkezi has_permission kontrolüyle belirlenir.
        if not self._has_permission_context():
            return True
        return self.has_permission(permission_code)

    def _refresh_permission_actions(self):
        if hasattr(self, "user_management_action"):
            self.user_management_action.setVisible(True)
        if hasattr(self, "role_permissions_action"):
            self.role_permissions_action.setVisible(
                self._permission_action_visible("manage_staff")
                or self._permission_action_visible("manage_roles")
            )
        if hasattr(self, "activity_logs_action"):
            self.activity_logs_action.setVisible(self._permission_action_visible("view_action_history"))

    def resizeEvent(self, event):
        super().resizeEvent(event)
        self.position_loading_overlay()
        self.position_query_logo_background()
        self.position_index_progress_badge()

    def position_index_progress_badge(self):
        if not hasattr(self, "index_progress_badge"):
            return
        parent = self.centralWidget()
        if not parent:
            return
        w, h = 138, 30
        margin = 12
        x = max(parent.width() - w - margin, 0)
        y = max(parent.height() - h - margin, 0)
        self.index_progress_badge.setGeometry(x, y, w, h)
        self.index_progress_badge.raise_()

    def set_index_progress_badge(self, visible: bool, percent: int = 0):
        if not hasattr(self, "index_progress_badge"):
            return
        p = int(max(0, min(100, int(percent or 0))))
        self.index_progress_badge.setText(f"Excel okuma %{p}")
        self.position_index_progress_badge()
        self.index_progress_badge.setVisible(bool(visible))

    def eventFilter(self, obj, event):
        # PySide6: Qt'nin C++ tarafından çağırdığı Python override'larında
        # yakalanmamış hata uygulamayı doğrudan abort ettirebilir. Bu filtre
        # QApplication seviyesine kurulu olduğu için export sırasında açılan/kapanan
        # dialoglar dahil tüm objelerin event'leri buradan geçer. Bu yüzden burada
        # silinmiş C++ objelere ve tüm yardımcı çağrılara karşı savunmalı davranıyoruz.
        try:
            if not qt_obj_alive(obj) or event is None:
                return False

            try:
                etype = event.type()
            except Exception:
                return False

            contract_viewport = None
            try:
                table = getattr(self, "contract_table", None)
                if qt_obj_alive(table):
                    contract_viewport = table.viewport()
            except Exception:
                contract_viewport = None

            if obj is contract_viewport and etype in (QEvent.Resize, QEvent.Move, QEvent.Show):
                QTimer.singleShot(0, self.position_query_logo_background)

            side_host = getattr(self, "side_meta_host", None)
            if obj is side_host and etype in (QEvent.Resize, QEvent.Show):
                self.position_side_meta_popover()

            _editing = (
                getattr(self, "_tree_editing", False)
                or getattr(self, "_file_dialog_open", False)
                or getattr(self, "_side_meta_modal_open", False)
            )
            panel_open = bool(getattr(self, "_side_meta_open_panel", None))

            if etype in (QEvent.WindowDeactivate, QEvent.ApplicationDeactivate) and panel_open and not _editing:
                self.close_side_meta_popover()

            if etype == QEvent.MouseButtonPress and panel_open:
                if not _editing and not self._is_side_meta_inside_click(obj, event):
                    self.close_side_meta_popover()

            if etype == QEvent.MouseButtonDblClick:
                file_id = None
                try:
                    if hasattr(obj, "property"):
                        file_id = obj.property("contractFileId")
                except Exception:
                    file_id = None
                if file_id:
                    self.open_contract_file(int(file_id))
                    return True

        except Exception:
            # Event filter içinde hata dışarı kaçarsa PySide6 abort edebilir.
            # Log görünür kalsın ama uygulama kapanmasın.
            try:
                traceback.print_exc()
            except Exception:
                pass
            return False

        return False

    def _norm_tr(self, s: str) -> str:
        t = str(s or "").strip().lower()
        return t.replace("ı", "i").replace("İ", "i").replace("ş", "s").replace("ğ", "g").replace("ü", "u").replace("ö", "o").replace("ç", "c")

    def _contract_health(self, it: dict) -> tuple[str, str, str, str]:
        """(cls, status_label, days_text, date_txt)
        status_label = Excel'deki gercek durum.
        cls = renk siniflandirmasi icin.
        """
        status_txt = str(it.get("status", "") or "").strip()
        date_txt = str(it.get("_near_delivery_txt") or "-")
        days_text = str(it.get("_near_delivery_days") or "-")
        today_iso = date.today().isoformat()
        cache_key = (status_txt, date_txt, days_text, today_iso)
        if it.get("_health_cache_key") == cache_key and "_health_cache_value" in it:
            return it["_health_cache_value"]
        day_num = it.get("_day_num")
        timing_kind = ""
        # Siniflandirma (renk ve uyari listeleri icin). Tamamlanmis gec
        # teslimatlar kirmizi gorunur ancak aktif gecikme uyarilarina girmez.
        if timing_kind == "gecikmeli_teslim":
            cls = "gecikmeli_teslim"
        elif is_completed_status(status_txt):
            cls = "tamamlandi"
        elif day_num is not None:
            cls = "geciken" if day_num < 0 else ("kritik" if day_num <= 60 else "normal")
        else:
            cls = "normal"
        # Gosterilecek etiket: Excel'deki gercek durum degeri
        st_label = status_txt if status_txt else "—"
        result = (cls, st_label, days_text, date_txt)
        it["_health_cache_key"] = cache_key
        it["_health_cache_value"] = result
        return result

    def _platform_logo_pixmap(self, platform: str, size: Optional[QSize] = None) -> Optional[QPixmap]:
        if not self.store:
            return None

        raw = self.store.get_platform_logo_bytes(platform)
        if not raw:
            return None

        px = QPixmap()
        if not px.loadFromData(raw):
            return None

        if size:
            return self._scale_logo_smart(px, size, max_upscale=3.0)

        return px

    def _send_query_logo_to_back(self) -> None:
        """Keep the decorative platform logo behind the contract table viewport."""
        try:
            if not hasattr(self, "query_logo_bg") or not hasattr(self, "contract_table"):
                return
            self.query_logo_bg.lower()
            viewport = self.contract_table.viewport()
            if viewport is not None:
                self.query_logo_bg.stackUnder(viewport)
            self.query_logo_bg.setAttribute(Qt.WA_TransparentForMouseEvents, True)
        except Exception:
            pass

    def _build_query_logo_strip(self, platforms: List[str]) -> Optional[QPixmap]:
        logos: List[QPixmap] = []

        for p in platforms:
            raw = self._platform_logo_pixmap(p)

            if not raw or raw.isNull():
                continue

            # Küçük logo en fazla 3 kat büyür.
            # Büyük logo bu kutuya sığacak kadar küçülür.
            px = self._scale_logo_smart(
                raw,
                QSize(900, 300),
                max_upscale=3.0
            )

            if px and not px.isNull():
                logos.append(px)

        if not logos:
            return None

        spacing = 36
        padding_x = 30
        padding_y = 20

        width = sum(px.width() for px in logos) + spacing * (len(logos) - 1) + padding_x * 2
        height = max(px.height() for px in logos) + padding_y * 2

        canvas = QPixmap(width, height)
        canvas.fill(Qt.transparent)

        painter = QPainter(canvas)
        painter.setRenderHint(QPainter.SmoothPixmapTransform, True)

        x = padding_x
        for px in logos:
            y = (height - px.height()) // 2
            painter.drawPixmap(x, y, px)
            x += px.width() + spacing

        painter.end()
        return canvas

    def update_query_logo_background(self, selected_platform: Optional[str] = None):
        """Sorgulama tablosu watermark kontrolü.

        - Platform seçiliyse: o platformun / seçili platformların gömülü logosu gösterilir.
        - Platform seçimi yoksa: varsayılan Baykar/STS logosu gösterilmez.
        """
        try:
            if not hasattr(self, "query_logo_bg"):
                return

            if not self.store:
                self._query_logo_source = None
                self.query_logo_bg.hide()
                self.query_logo_bg.clear()
                return

            targets: List[str] = []

            if isinstance(selected_platform, (list, tuple, set)):
                targets = [str(p).strip() for p in selected_platform if str(p or "").strip()]
            elif selected_platform:
                targets = [str(selected_platform).strip()]
            else:
                # _apply_platform_selection çoklu seçimde None gönderebilir.
                # Bu durumda seçili platform seti varsa onların logolarını göster;
                # gerçekten hiç seçim yoksa varsayılan logo gösterme.
                current_selected = list(getattr(self, "selected_platforms", set()) or [])
                targets = [str(p).strip() for p in current_selected if str(p or "").strip()]

            # Dekoratif logo yalnızca tek platform seçiliyken gösterilir. Çoklu seçim,
            # tüm platformlar veya genel görünümde tabloyu kapatabilecek watermark yoktur.
            if len(set(targets)) != 1:
                self._query_logo_source = None
                self.query_logo_bg.hide()
                self.query_logo_bg.clear()
                return

            # Sıralama sabit kalsın; soldaki platform listesi sırası varsa onu kullan.
            try:
                order = {name: i for i, name in enumerate(self._all_platform_names())}
                targets = sorted(set(targets), key=lambda x: order.get(x, 9999))
            except Exception:
                targets = sorted(set(targets))

            strip = self._build_query_logo_strip(targets)
            if not strip or strip.isNull():
                self._query_logo_source = None
                self.query_logo_bg.hide()
                self.query_logo_bg.clear()
                return

            self._query_logo_source = strip
            self.position_query_logo_background()
        except Exception:
            try:
                self._query_logo_source = None
                if hasattr(self, "query_logo_bg"):
                    self.query_logo_bg.hide()
                    self.query_logo_bg.clear()
            except Exception:
                pass

    def _set_platform_items(self, platforms: List[str]):
        available = {str(p) for p in platforms}
        self.selected_platforms.intersection_update(available)
        if not self.selected_platforms:
            self.multi_platform_mode = False
        self._updating_platform_list = True
        try:
            self.platform_list.clear()
            counts: Dict[str, int] = {}
            for it in self.contract_index:
                p = str(it.get("platform", ""))
                counts[p] = counts.get(p, 0) + 1
            for i, p in enumerate(platforms):
                platform = str(p)
                row = QListWidgetItem(platform)
                row.setData(Qt.UserRole, platform)
                row.setSizeHint(QSize(0, 46))
                self.platform_list.addItem(row)
                if hasattr(self, "_platform_list_delegate"):
                    self._platform_list_delegate.set_count(i, counts.get(platform, 0))
        finally:
            self._updating_platform_list = False
        self.refresh_platform_list_ui()

    def _all_platform_names(self) -> List[str]:
        return [
            str(self.platform_list.item(i).data(Qt.UserRole) or "")
            for i in range(self.platform_list.count())
            if str(self.platform_list.item(i).data(Qt.UserRole) or "")
        ]

    def normalize_platform_selection_state(self):
        if not self.selected_platforms:
            self.multi_platform_mode = False

    def refresh_platform_list_ui(self):
        self.normalize_platform_selection_state()
        self._updating_platform_list = True
        try:
            for i in range(self.platform_list.count()):
                item = self.platform_list.item(i)
                platform = str(item.data(Qt.UserRole) or "")
                flags = item.flags()
                if self.multi_platform_mode:
                    item.setFlags(flags | Qt.ItemIsUserCheckable)
                    item.setCheckState(Qt.Checked if platform in self.selected_platforms else Qt.Unchecked)
                else:
                    item.setData(Qt.CheckStateRole, None)
                    item.setFlags(flags & ~Qt.ItemIsUserCheckable)
                is_selected = platform in self.selected_platforms
                item.setData(PLATFORM_SELECTED_ROLE, is_selected)
                item.setSelected(is_selected)
        finally:
            self._updating_platform_list = False
        count = len(self.selected_platforms)
        self.platform_list.viewport().update()
        self.platform_selection_badge.setText(f"{count} seçili")
        self.platform_selection_badge.setVisible(count > 0)
        self.platform_info_label.setText(f"{count} platform · sağ tık ile düzenle")
        self.platform_info_bar.setVisible(count > 0)

    def _apply_platform_selection(self):
        self.normalize_platform_selection_state()
        selected = set(self.selected_platforms)
        self.all_contract_rows = [
            dict(it) for it in self.contract_index
            if not selected or str(it.get("platform", "")) in selected
        ]
        self._prepare_contract_row_cache(self.all_contract_rows)
        self._refresh_query_filters()
        if not selected:
            self.right_title.setText("Sözleşme Sorgulama")
            self.update_query_logo_background(None)
        elif len(selected) == 1:
            platform = next(iter(selected))
            self.right_title.setText(f"{platform} - Sözleşmeler")
            self.update_query_logo_background(platform)
        else:
            self.right_title.setText(f"{len(selected)} Platform - Sözleşmeler")
            self.update_query_logo_background(None)
        self.refresh_platform_list_ui()
        self.schedule_apply_contract_filter()

    def on_platform_clicked(self, item: QListWidgetItem):
        platform = str(item.data(Qt.UserRole) or "")
        if not platform:
            return
        if self.multi_platform_mode:
            if self._platform_checkbox_changed == platform:
                self._platform_checkbox_changed = None
            else:
                self.toggle_platform_multi(platform)
        elif self.selected_platforms == {platform}:
            self.clear_platform_selection()
        else:
            self.selected_platforms = {platform}
            self._apply_platform_selection()

    def _on_platform_item_changed(self, item: QListWidgetItem):
        if self._updating_platform_list or not self.multi_platform_mode:
            return
        platform = str(item.data(Qt.UserRole) or "")
        checked = item.checkState() == Qt.Checked
        if checked != (platform in self.selected_platforms):
            self._platform_checkbox_changed = platform
            QTimer.singleShot(0, lambda p=platform: self._clear_platform_checkbox_marker(p))
            self.toggle_platform_multi(platform)

    def _clear_platform_checkbox_marker(self, platform: str):
        if self._platform_checkbox_changed == platform:
            self._platform_checkbox_changed = None

    def _on_platform_context_menu_requested(self, pos: QPoint):
        item = self.platform_list.itemAt(pos)
        if item:
            self.show_platform_context_menu(str(item.data(Qt.UserRole) or ""), self.platform_list.viewport().mapToGlobal(pos))

    def show_platform_context_menu(self, platform: str, global_pos: QPoint):
        if not platform:
            return
        menu = QMenu(self)
        label = "− Seçimden çıkar" if platform in self.selected_platforms else "+ Çoklu seçime ekle"
        menu.addAction(label, lambda: self.toggle_platform_multi(platform))
        menu.addAction("☑ Tümünü seç", self.select_all_platforms)
        menu.addAction("× Seçimi temizle", self.clear_platform_selection)
        menu.exec(global_pos)

    def toggle_platform_multi(self, platform: str):
        self.multi_platform_mode = True
        if platform in self.selected_platforms:
            self.selected_platforms.remove(platform)
        else:
            self.selected_platforms.add(platform)
        self.normalize_platform_selection_state()
        self._apply_platform_selection()

    def select_all_platforms(self):
        self.selected_platforms = set(self._all_platform_names())
        self.multi_platform_mode = bool(self.selected_platforms)
        self._apply_platform_selection()

    def clear_platform_selection(self):
        self.selected_platforms.clear()
        self.multi_platform_mode = False
        self._updating_platform_list = True
        try:
            for i in range(self.platform_list.count()):
                item = self.platform_list.item(i)
                item.setCheckState(Qt.Unchecked)
                item.setData(PLATFORM_SELECTED_ROLE, False)
                item.setSelected(False)
        finally:
            self._updating_platform_list = False
        self._apply_platform_selection()

    def _clear_upcoming_layout(self):
        while self.upcoming_layout.count():
            item = self.upcoming_layout.takeAt(0)
            w = item.widget()
            if w:
                w.deleteLater()

    def update_alert_strip(self):
        today = date.today()
        self.today_num.setText(str(today.day))
        self.today_info.setText(f"{TR_MONTHS[today.month - 1].upper()}\n{today.year}")

        overdue = []
        critical = []
        for it in self.contract_index:
            ctype = self._norm_tr(str(it.get("type", "") or ""))
            if ctype != self._norm_tr("Ana Sözleşme"):
                continue
            cls, _st, kgun, _dt = self._contract_health(it)
            if cls == "geciken":
                overdue.append((it, kgun))
            elif cls == "kritik":
                critical.append((it, kgun))
        self.overdue_count.setText(str(len(overdue)))
        self.critical_count.setText(str(len(critical)))

        show_overdue = len(overdue) > 0
        show_critical = len(critical) > 0
        self.alert_overdue_group.setVisible(show_overdue)
        self.alert_critical_group.setVisible(show_critical)
        self.alert_divider1.setVisible(show_overdue or show_critical)
        self.alert_divider2.setVisible(show_overdue or show_critical)

        self._clear_upcoming_layout()
        upcoming = overdue[:5] + critical[:10]
        for it, kgun in upcoming:
            p = str(it.get("platform", "") or "")
            no = str(it.get("no", "") or "")
            cls, _st, _kg, _dt = self._contract_health(it)
            b = QPushButton(f"• #{no} · {p} · {kgun}")
            b.setCursor(Qt.PointingHandCursor)
            b.setProperty("kind", "red" if cls == "geciken" else "amber")
            b.setObjectName("upcomingPill")
            b.clicked.connect(lambda _=False, item=dict(it): self.open_contract_item(item))
            self.upcoming_layout.addWidget(b)
        self.upcoming_layout.addStretch()

    def set_empty_state(self):
        self.platform_list.clear()
        self.selected_platforms.clear()
        self.multi_platform_mode = False
        if hasattr(self, "platform_selection_badge"):
            self.refresh_platform_list_ui()
        self.all_contract_rows = []
        self.contract_table.setRowCount(0)
        self.set_index_progress_badge(False, 0)
        self.overdue_count.setText("0")
        self.critical_count.setText("0")
        self.alert_overdue_group.hide()
        self.alert_critical_group.hide()
        self.alert_divider1.hide()
        self.alert_divider2.hide()
        self._clear_upcoming_layout()
        self.upcoming_layout.addStretch()
        self.right_title.setText("Sözleşme Sorgulama")
        self.update_query_logo_background(None)
        self.update_connection_badge("bad")

    def set_loading_state(self, loading: bool, message: str = "Analiz ediliyor..."):
        self._loading = loading
        if loading:
            self.loading_label.setText(message)
            self.loading_progress.setRange(0, 100)
            self.loading_progress.setValue(0)
            self.position_loading_overlay()
            self.loading_overlay.show()
            self.update_connection_badge("loading")
        else:
            self.loading_progress.setValue(100)
            self.loading_overlay.hide()
            self.update_connection_badge("ok" if self.store else "bad")

    def set_busy_overlay(self, visible: bool, message: str = "İşlem yapılıyor...", percent: int = 0):
        # Excel worker yüklemesi yokken kısa/orta süreli senkron işlemlerde kullan.
        if not hasattr(self, "_busy_cursor_on"):
            self._busy_cursor_on = False
        if visible:
            self.loading_label.setText(message)
            self.loading_progress.setRange(0, 100)
            self.loading_progress.setValue(int(max(0, min(100, percent))))
            self.position_loading_overlay()
            self.loading_overlay.show()
            if not self._busy_cursor_on:
                QApplication.setOverrideCursor(Qt.WaitCursor)
                self._busy_cursor_on = True
            QApplication.processEvents()
        else:
            self.loading_overlay.hide()
            if self._busy_cursor_on:
                QApplication.restoreOverrideCursor()
                self._busy_cursor_on = False
            QApplication.processEvents()

    def start_excel_load(self, path: Path):
        if self._loader_thread and self._loader_thread.isRunning():
            return
        self.path = Path(path)
        self.store = None
        self.contract_index = []
        self.all_contract_rows = []
        self.selected_platforms.clear()
        self.multi_platform_mode = False
        self._store_loading = True
        self._index_ready_for_use = False
        self._last_load_timings = {}
        self._version_baseline_signature = None
        if hasattr(self, "platform_list"):
            self.platform_list.clear()
            self.refresh_platform_list_ui()
        if hasattr(self, "contract_table"):
            self.contract_table.setRowCount(0)
        self._streaming_index = False
        self.set_index_progress_badge(True, 0)
        self.set_loading_state(True, "Analiz ediliyor...")
        self._loader_thread = QThread(self)
        self._loader_worker = ExcelLoadWorker(self.path)
        self._loader_worker.moveToThread(self._loader_thread)
        self._loader_thread.started.connect(self._loader_worker.run)
        self._loader_worker.store_ready.connect(self.on_excel_store_ready)
        self._loader_worker.batch_ready.connect(self.on_excel_index_batch)
        self._loader_worker.index_ready.connect(self.on_excel_index_ready)
        self._loader_worker.finished.connect(self.on_excel_loaded)
        self._loader_worker.failed.connect(self.on_excel_load_failed)
        self._loader_worker.progress.connect(self.on_excel_load_progress)
        self._loader_worker.finished.connect(self._loader_thread.quit)
        self._loader_worker.failed.connect(self._loader_thread.quit)
        self._loader_thread.finished.connect(self._loader_worker.deleteLater)
        self._loader_thread.finished.connect(self._loader_thread.deleteLater)
        self._loader_thread.finished.connect(self._clear_loader_refs)
        self._loader_thread.start()


    def start_sts_load(self, path: Path):
        """STS dosyasını yükler.

        Adımlar:
        1. STSLoadWorker arka planda dosya doğrulaması yapar (magic bytes).
        2. finished() sinyali gelince _on_sts_load_finished() ana thread'de
           STSStore ve contract_index oluşturur.

        SQLite connection YALNIZCA ana thread'de (adım 2'de) açılır.
        Worker hiçbir zaman connection nesnesi taşımaz.
        """
        if self._sts_loader_thread and self._sts_loader_thread.isRunning():
            return
        self.path = Path(path)
        self.store = None
        self.contract_index = []
        self._tag_color_map_cache = None
        self._store_loading = True
        self.set_loading_state(True, "STS dosyası yükleniyor...")

        thread = QThread(self)
        worker = STSLoadWorker(self.path)
        worker.moveToThread(thread)

        thread.started.connect(worker.run)
        worker.progress.connect(self._on_sts_load_progress)
        worker.finished.connect(self._on_sts_load_finished)
        worker.failed.connect(self._on_sts_load_failed)
        worker.finished.connect(thread.quit)
        worker.failed.connect(thread.quit)
        thread.finished.connect(worker.deleteLater)
        thread.finished.connect(thread.deleteLater)
        thread.finished.connect(self._clear_sts_loader_refs)

        self._sts_loader_thread = thread
        self._sts_loader_worker = worker
        thread.start()

    def _clear_sts_loader_refs(self):
        self._sts_loader_thread = None
        self._sts_loader_worker = None

    def _on_sts_load_progress(self, percent: int, message: str):
        self.set_loading_state(True, f"{message}  %{percent}")

    def _on_sts_load_finished(self):
        """Worker doğrulamayı geçti; STSStore ve index ANA THREAD'de açılır.

        SQLite connection burada oluşur — hiçbir zaman worker thread'den
        taşınmaz.  build_contract_index() büyük dosyalarda birkaç saniye
        sürebilir; ilerleyen sürümlerde bu aşama da ayrı bir worker'a
        taşınabilir (connection o worker'da açılıp kapatılır, yalnızca
        list[dict] ana thread'e döner).
        """
        actor = str((self.current_staff or {}).get("full_name") or "Personel")
        try:
            self.store = STSStore(self.path, actor=actor)
            self.contract_index = self.store.build_contract_index()
        except Exception as exc:
            _log.exception("STSStore ana-thread açılış hatası")
            self._store_loading = False
            self.set_loading_state(False)
            self.set_empty_state()
            QMessageBox.critical(self, "STS yükleme hatası",
                                 f"STS dosyası açılamadı.\n\n{exc}")
            return
        self._tag_color_map_cache = None
        self._store_loading = False
        self.set_loading_state(False)
        self._set_platform_items(self.store.platform_names())
        self.update_alert_strip()
        self._apply_platform_selection()
        self.connection_label.setText("✓ STS veri dosyası bağlı")
        self._apply_version_to_ui()
        self._remember_version_baseline()

    def _on_sts_load_failed(self, error_text: str):
        self._store_loading = False
        self.set_loading_state(False)
        self.set_empty_state()
        _log.error("STS doğrulama hatası: %s", error_text)
        QMessageBox.critical(self, "STS yükleme hatası",
                             f"STS dosyası okunamadı.\n\n{error_text}")

    def is_sts_mode(self) -> bool:
        return (
            self.store is not None
            and self.path is not None
            and str(self.path).lower().endswith(".sts")
        )

    def _clear_loader_refs(self):
        self._loader_worker = None
        self._loader_thread = None

    def _refresh_index_tags_only(self):
        if not self.store:
            return
        tags_map = self.store.all_contract_tags_map()
        for it in self.contract_index:
            p = str(it.get("platform", "") or "").strip()
            no = str(it.get("no", "") or "").strip()
            ctype = str(it.get("type", "") or "").strip()
            tags = tags_map.get((p, no, ctype), [])
            it["tags"] = list(tags)
            search = " ".join(
                str(it.get(k, "") or "")
                for k in ["platform", "no", "user", "type", "link", "status", "completion_date", "content"]
            ).lower()
            if tags:
                search = (search + " " + " ".join(tags).lower()).strip()
            it["search"] = search

    def _rebuild_index_in_platform_order(self, platform_rows: Dict[str, List[dict]]):
        if not self.store:
            self.contract_index = []
            return
        ordered: List[dict] = []
        for p in self.store.platform_names():
            ordered.extend([dict(it) for it in platform_rows.get(p, [])])
        self.contract_index = ordered

    def _refresh_single_platform_index(self, platform: str, select_platform: Optional[str] = None):
        if not self.store:
            self.set_empty_state()
            return
        p = safe_sheet_name(str(platform or ""))
        if not p:
            self.request_refresh(select_platform=select_platform, scope="all")
            return
        current_rows: Dict[str, List[dict]] = {}
        for it in self.contract_index:
            pp = str(it.get("platform", "") or "")
            current_rows.setdefault(pp, []).append(dict(it))
        self.set_busy_overlay(True, f"{p} güncelleniyor...", 25)
        try:
            QApplication.processEvents()
            tags_map = self.store.all_contract_tags_map()
            self.set_busy_overlay(True, "Sözleşme satırları okunuyor...", 62)
            QApplication.processEvents()
            current_rows[p] = self.store.list_main_contracts(p, tags_map=tags_map)
            self._rebuild_index_in_platform_order(current_rows)
            self.set_busy_overlay(True, "Arayüz yenileniyor...", 92)
            QApplication.processEvents()
            platforms = self.store.platform_names()
            self._set_platform_items(platforms)
            self.update_alert_strip()
            self.refresh_open_calendar()
            target = str(select_platform or p)
            if target:
                self.select_platform(target)
            elif self.platform_list.count():
                self._apply_platform_selection()
            else:
                self.set_empty_state()
        finally:
            self.set_busy_overlay(False)


    def refresh_open_calendar(self):
        cal = getattr(self, "calendar_window", None)
        if not cal or not cal.isVisible():
            return
        try:
            cal.refresh_from_index(self.store, self.contract_index)
        except RuntimeError:
            self.calendar_window = None

    def request_refresh(self, select_platform: Optional[str] = None, scope: str = "all", platform: Optional[str] = None):
        if not self.store:
            self.set_empty_state()
            return
        kind = str(scope or "all").strip().lower()
        if kind == "all":
            if self.is_sts_mode():
                self.contract_index = self.store.build_contract_index()
                self._set_platform_items(self.store.platform_names())
                self.update_alert_strip()
                self.refresh_open_calendar()
                if select_platform:
                    self.select_platform(select_platform)
                else:
                    self._apply_platform_selection()
                self.connection_label.setText("✓ STS veri dosyası bağlı")
                return
            self._pending_select_platform = select_platform
            self.start_excel_load(self.path)
            return
        if kind == "tags":
            self.set_busy_overlay(True, "Etiketler güncelleniyor...", 35)
            try:
                self._refresh_index_tags_only()
                self.set_busy_overlay(True, "Arayüz yenileniyor...", 90)
                self._set_platform_items(self.store.platform_names())
                self.update_alert_strip()
                self.refresh_open_calendar()
                if select_platform:
                    self.select_platform(select_platform)
                else:
                    self._apply_platform_selection()
            finally:
                self.set_busy_overlay(False)
            return
        if kind == "platform":
            self._refresh_single_platform_index(str(platform or select_platform or ""), select_platform=select_platform or platform)
            return
        # UI-only yenileme: tekrar excel indeks okumadan görünümü tazeler.
        self._set_platform_items(self.store.platform_names())
        self.update_alert_strip()
        self.refresh_open_calendar()
        if self.platform_list.count():
            self._apply_platform_selection()
        else:
            self.set_empty_state()

    def on_excel_store_ready(self, store):
        self.store = store
        self.path = self.store.path
        self.contract_index = []
        self._tag_color_map_cache = None
        self._streaming_index = True
        self.loading_overlay.hide()
        self._loading = False
        self.update_connection_badge("loading")
        self.connection_label.setText("Excel indeksleniyor %0")
        self.set_index_progress_badge(True, 0)
        platforms = self.store.platform_names()
        self._set_platform_items(platforms)
        self.update_alert_strip()
        self._apply_platform_selection()

    def on_excel_index_batch(self, platform: str, rows, mapped_percent: int, message: str):
        new_rows = [dict(it) for it in list(rows or [])]
        self.set_index_progress_badge(True, int(mapped_percent or 0))
        if not new_rows:
            self.connection_label.setText(f"Excel indeksleniyor %{int(mapped_percent or 0)}")
            return
        self.contract_index.extend(new_rows)
        self.connection_label.setText(f"Excel indeksleniyor %{int(mapped_percent or 0)}")
        if self.store:
            self._apply_platform_selection()

    def on_excel_index_ready(self, platforms, index, timings):
        self.contract_index = list(index or [])
        self._last_load_timings = dict(timings or {})
        self._tag_color_map_cache = None
        self._streaming_index = False
        self._loading = False
        if hasattr(self, "loading_overlay"):
            self.loading_overlay.hide()
        self._index_ready_for_use = True
        self.update_connection_badge("ok")
        self.connection_label.setText("Liste hazır (detay düzenleme arka planda hazırlanıyor)")
        self.set_index_progress_badge(False, 100)
        platform_list = list(platforms or [])
        self._set_platform_items(platform_list)
        self.update_alert_strip()
        self.refresh_open_calendar()
        if self._pending_select_platform:
            self.select_platform(self._pending_select_platform)
        elif self.platform_list.count():
            self._apply_platform_selection()
        else:
            self.contract_table.setRowCount(0)

    def on_excel_loaded(self, store, index):
        selected_platform = next(iter(self.selected_platforms)) if len(self.selected_platforms) == 1 else ""
        self.store = store
        self.contract_index = list(index or [])
        self.path = self.store.path
        self._tag_color_map_cache = None
        self._store_loading = False
        self._index_ready_for_use = False
        self.set_loading_state(False)
        self.set_index_progress_badge(False, 100)
        self._streaming_index = False
        self.refresh(rebuild_index=False)
        self.refresh_open_calendar()
        if self._pending_select_platform:
            self.select_platform(self._pending_select_platform)
        elif selected_platform:
            self.select_platform(selected_platform)
        self._pending_select_platform = None
        self._apply_version_to_ui()
        self._remember_version_baseline()

    def on_excel_load_failed(self, error_text: str):
        self._store_loading = False
        self._index_ready_for_use = False
        self.set_loading_state(False)
        self.set_index_progress_badge(False, 0)
        self.set_empty_state()
        self._pending_select_platform = None
        QMessageBox.critical(self, "Excel yükleme hatası", f"Excel dosyası okunamadı.\n\n{error_text}")

    def on_excel_load_progress(self, percent: int, message: str):
        p = int(max(0, min(100, int(percent or 0))))
        msg = str(message or "Analiz ediliyor...")
        self.set_index_progress_badge(True, p)
        if self._loading and hasattr(self, "loading_progress"):
            self.loading_progress.setRange(0, 100)
            self.loading_progress.setValue(p)
        if self._loading and hasattr(self, "loading_label"):
            self.loading_label.setText(f"{msg}  %{p}")
        elif getattr(self, "_store_loading", False):
            if self._index_ready_for_use:
                self.update_connection_badge("ok")
                self.connection_label.setText("Liste hazır (detay düzenleme arka planda hazırlanıyor)")
            else:
                self.update_connection_badge("loading")
                self.connection_label.setText(f"{msg} %{p}")
        elif self.store:
            self.update_connection_badge("loading")
            self.connection_label.setText(f"Excel indeksleniyor %{p}")

    def on_contract_save_progress(self, percent: int, message: str):
        self.set_busy_overlay(True, message, percent)

    def open_file(self):
        dlg = WorkbookStartDialog(self)
        if dlg.exec() and dlg.selected_path:
            sel = Path(dlg.selected_path)
            if sel.suffix.lower() == ".sts":
                if _share_metadata_from_path(sel):
                    try:
                        win = open_share_contract_window(sel)
                        if win:
                            if not hasattr(self, "_share_windows"):
                                self._share_windows = []
                            self._share_windows.append(win)
                            win.destroyed.connect(lambda *_args, w=win: self._share_windows.remove(w) if hasattr(self, "_share_windows") and w in self._share_windows else None)
                            win.show()
                    except Exception as exc:
                        QMessageBox.critical(self, "Paylaşım açılamadı", f"Paylaşım dosyası açılamadı.\n\n{exc}")
                    return
                if not auth.ensure_system_admin_setup(sel, self):
                    return
                staff = auth.require_staff_login(sel, self)
                if not staff:
                    return
                self.current_staff = staff
                auth.current_staff = staff
                self.start_sts_load(sel)
            else:
                self.start_excel_load(sel)

    def show_contract_summary(self, row: int, item: dict):
        if not self.store:
            return
        old_popup = getattr(self, "_summary_popup", None)
        if old_popup and old_popup.isVisible():
            old_popup.close()

        dialog = ContractSummaryDialog(
            self.store,
            item,
            self,
            detail_handler=self.open_summary_event_detail,
        )
        self._summary_popup = dialog
        dialog.destroyed.connect(lambda *_: setattr(self, "_summary_popup", None))
        dialog.show()
        dialog.raise_()
        dialog.activateWindow()

    def open_summary_event_detail(self, item: dict) -> bool:
        self.open_contract_item(item)
        return True

    def _on_contract_cell_clicked(self, row: int, col: int):
        """Ozet hucresine tiklayinca popup ac.
        Col disi tiklama open_selected_contract ile cakismasin.
        """
        if col != COL_SUMMARY:
            return
        rows = getattr(self.contract_table, "_visible_rows", [])
        if row < 0 or row >= len(rows):
            return
        self.show_contract_summary(row, rows[row])

    def manage_platforms(self):
        if not self.require_permission_ui("manage_platforms", "Platform ve Bileşen Yönetimi"):
            return
        if not self.store:
            QMessageBox.information(self, "Excel gerekli", "Önce bir Excel dosyası bağlayın.")
            return
        dlg = PlatformComponentManagerDialog(self.store, self, initial_tab=0)
        saved_via_signal = False

        def refresh_after_platform_save():
            nonlocal saved_via_signal
            saved_via_signal = True
            current = self.platform_list.currentItem() if hasattr(self, "platform_list") else None
            current_platform = str(current.data(Qt.UserRole) or "") if current else None
            self.request_refresh(select_platform=current_platform, scope="all")

        dlg.settings_saved.connect(refresh_after_platform_save)
        dlg.exec()
        if dlg.changed and not saved_via_signal:
            refresh_after_platform_save()

    def create_platform(self):
        """Eski uyumluluk - manage_platforms'u cagirir."""
        self.manage_platforms()

    def manage_users(self):
        self.open_user_management()

    def manage_tags(self):
        if not self.require_permission_ui("manage_labels", "Etiket Yönetimi"):
            return
        if not self.store:
            QMessageBox.information(self, "Excel gerekli", "Önce bir Excel dosyası bağlayın.")
            return
        dlg = TagManagerDialog(self.store, self.contract_index, self)
        dlg.exec()
        if dlg.changed:
            self._tag_color_map_cache = None
            current_platform = ""
            cur = self.platform_list.currentItem()
            if cur:
                current_platform = str(cur.data(Qt.UserRole) or "")
            self.request_refresh(select_platform=current_platform, scope="tags")

    def manage_components(self):
        if not self.require_permission_ui("manage_components", "Bileşen Yönetimi"):
            return
        if not self.store:
            QMessageBox.information(self, "Excel gerekli", "Önce bir Excel dosyası bağlayın.")
            return
        dlg = PlatformComponentManagerDialog(self.store, self, initial_tab=1)
        if dlg.exec():
            self.request_refresh(scope="ui")


    def open_calendar_tracking(self):
        if not self.store:
            QMessageBox.information(self, "Excel gerekli", "Önce bir Excel dosyası bağlayın.")
            return
        self.set_busy_overlay(True, "Takvim hazırlanıyor...")
        try:
            self.calendar_window = ContractCalendarWindow(
                self.store, self.contract_index, self, detail_handler=self.open_calendar_event_detail
            )
        finally:
            self.set_busy_overlay(False)
        self.calendar_window.showMaximized()
        self.calendar_window.show()
        self.calendar_window.raise_()
        self.calendar_window.activateWindow()


    def open_calendar_event_detail(self, ev: dict) -> bool:
        platform = str(ev.get("platform", "") or "")
        no = str(ev.get("no", "") or "")
        row = int(ev.get("row", 0) or 0)
        ci, systems, deliveries = self.store.load_contract_structure(platform, no, start_row=row if row > 0 else None)
        if not ci:
            QMessageBox.warning(self, "Bulunamadı", "Sözleşme detayları okunamadı.")
            return False
        work = ContractWorkWindow(self.store, ci, self, systems=systems, deliveries=deliveries)
        return bool(work.exec())

    def new_contract(self):
        if not self.require_permission_ui("create_contracts", "Sözleşme Ekleme"):
            return
        if not self.store:
            QMessageBox.information(self, "Excel gerekli", "Önce bir Excel dosyası bağlayın.")
            return
        if not self.store.load_users():
            QMessageBox.information(self, "Kullanıcı gerekli", "Sözleşme girmeden önce kullanıcı tanımlayın.")
            udlg = UserManagerDialog(self.store, self)
            if not udlg.exec() or not self.store.load_users():
                return
        dlg=ContractDialog(self.store,self)
        if dlg.exec() and dlg.result:
            try:
                new_contract_id = self.store.write_contract(dlg.result, [], {})
                dlg.result.entry_start_row = int(new_contract_id or 0)
                setattr(dlg.result, "id", int(new_contract_id or 0))
                setattr(dlg.result, "contract_id", int(new_contract_id or 0))
                active_pid = int((getattr(dlg.result, "platform_ids", []) or [0])[0] or getattr(dlg.result, "platform_id", 0) or 0)
                ci, systems, deliveries = self.store.load_contract_structure(
                    dlg.result.platform,
                    dlg.result.no,
                    start_row=new_contract_id,
                    contract_type=dlg.result.contract_type,
                    platform_id=active_pid,
                )
            except Exception as exc:
                traceback.print_exc()
                QMessageBox.critical(self, "Sözleşme Kaydedilemedi", f"Yeni sözleşme kaydedilemedi:\n{exc}")
                return
            work=ContractWorkWindow(self.store,ci,self,systems=systems,deliveries=deliveries)
            if work.exec():
                self.request_refresh(
                    select_platform=ci.platform,
                    scope="platform",
                    platform=ci.platform,
                )

    def refresh(self, rebuild_index: bool = True):
        if not self.store:
            self.set_empty_state()
            return
        # Kayıttan sonra indeks tek seferde yenilenir; arama bu indeks üzerinden yapılır.
        if rebuild_index:
            self.contract_index = self.store.build_contract_index()
        if self.is_sts_mode():
            self.connection_label.setText("✓ STS veri dosyası bağlı")
        platforms = self.store.platform_names()
        self._set_platform_items(platforms)
        self.update_query_logo_background(None)
        self.update_alert_strip()
        self.refresh_open_calendar()
        self._apply_platform_selection()

    def select_platform(self, p):
        platform = str(p or "")
        if platform and platform in self._all_platform_names():
            self.selected_platforms = {platform}
            self.multi_platform_mode = False
            self._apply_platform_selection()

    def load_platform_contracts(self, platform):
        self.select_platform(platform)

    def toggle_filter_bar(self):
        visible = not self.filter_bar.isVisible()
        self.filter_bar.setVisible(visible)
        sender = self.sender()
        if visible:
            if sender is self.sort_btn:
                self.sort_combo.setFocus()
            else:
                self.filter_type.setFocus()

    def clear_query_filters(self):
        self.search_input.clear()
        if hasattr(self, "date_from_filter"):
            self._date_from_active = False
            self.date_from_filter.setDate(QDate.currentDate())
        if hasattr(self, "date_to_filter"):
            self._date_to_active = False
            self.date_to_filter.setDate(QDate.currentDate())
        if hasattr(self, "days_min_filter"):
            self.days_min_filter.clear()
        if hasattr(self, "days_max_filter"):
            self.days_max_filter.clear()
        if hasattr(self, 'filter_type'):
            self.filter_type.setCurrentIndex(0)
        if hasattr(self, 'filter_status'):
            self.filter_status.setCurrentIndex(0)
        if hasattr(self, 'sort_combo'):
            self.sort_combo.setCurrentIndex(0)
        # Sutun filtrelerini temizle
        if hasattr(self, '_filter_header'):
            self._filter_header.clear_all_filters()
        self.schedule_apply_contract_filter()

    def _on_date_filter_changed(self, which: str):
        if which == "from":
            self._date_from_active = True
        else:
            self._date_to_active = True
        self.schedule_apply_contract_filter()

    def _refresh_query_filters(self):
        type_vals = sorted(
            {str(it.get("type_display", it.get("type", "")) or "").strip() for it in self.all_contract_rows if str(it.get("type_display", it.get("type", "")) or "").strip()},
            key=self._norm_tr,
        )
        self.filter_type.blockSignals(True)
        self.filter_type.clear()
        self.filter_type.addItem("Tüm Türler", "")
        for tv in type_vals:
            self.filter_type.addItem(tv, tv)
        self.filter_type.blockSignals(False)

    def _contract_no_sort_key(self, no_text: str):
        txt = str(no_text or "").strip()
        if txt.isdigit():
            return (0, int(txt), txt.lower())
        parts = re.split(r"(\d+)", txt.lower())
        key = []
        for p in parts:
            if p.isdigit():
                key.append((0, int(p)))
            else:
                key.append((1, p))
        return (1, key)

    def _completion_date_sort_key(self, it: dict):
        d = it.get("_completion_obj")
        if d:
            return (0, d.toordinal())
        return (1, 99999999)

    def _days_sort_key(self, it: dict):
        day_num = it.get("_day_num")
        if day_num is not None:
            return (0, int(day_num))
        return (1, 99999999)

    def _contract_delivery_dates(self, it: dict) -> tuple[str, str, Optional[int]]:
        exact_plans = []
        has_flexible = False
        try:
            _ci, _systems, deliveries = self.store.load_contract_structure(
                it.get("platform", ""), it.get("no", ""), it.get("row", None) or it.get("entry_start_row", None), it.get("type", None)
            )
            for items in (deliveries or {}).values():
                for delivery in items or []:
                    raw = str(getattr(delivery, "planned_acceptance_date", "") or "").strip()
                    if not raw:
                        continue
                    parsed = parse_flexible_date(raw)
                    if parsed:
                        exact_plans.append(parsed)
                    else:
                        has_flexible = True
        except Exception:
            return "-", "-", None
        if exact_plans:
            near = min(exact_plans)
            diff = (near - date.today()).days
            return near.isoformat(), (f"{diff} gün" if diff >= 0 else f"{abs(diff)} gün gecikti"), diff
        if has_flexible:
            return "Belirsiz", "-", None
        return "-", "-", None

    def _delivery_summary_map(self, rows: List[dict]) -> Dict[int, tuple[str, str, Optional[int]]]:
        ids = [int(r.get("row") or r.get("entry_start_row") or 0) for r in rows if int(r.get("row") or r.get("entry_start_row") or 0)]
        if not ids or not getattr(getattr(self.store, "db", None), "conn", None):
            return {}
        placeholders = ",".join("?" for _ in ids)
        summary: Dict[int, tuple[list, bool]] = {cid: ([], False) for cid in ids}
        try:
            for cid, raw in self.store.db.conn.execute(
                f"SELECT contract_id, planned_acceptance_date FROM deliveries WHERE contract_id IN ({placeholders})",
                ids,
            ).fetchall():
                exacts, has_flexible = summary.setdefault(int(cid), ([], False))
                text = str(raw or "").strip()
                if not text:
                    continue
                parsed = parse_flexible_date(text)
                if parsed:
                    exacts.append(parsed)
                else:
                    has_flexible = True
                summary[int(cid)] = (exacts, has_flexible)
        except Exception:
            return {}
        out: Dict[int, tuple[str, str, Optional[int]]] = {}
        today = date.today()
        for cid, (exacts, has_flexible) in summary.items():
            if exacts:
                near = min(exacts)
                diff = (near - today).days
                out[cid] = (near.isoformat(), f"{diff} gün" if diff >= 0 else f"{abs(diff)} gün gecikti", diff)
            elif has_flexible:
                out[cid] = ("Belirsiz", "-", None)
            else:
                out[cid] = ("-", "-", None)
        return out

    def _prepare_contract_row_cache(self, rows: List[dict]):
        today = date.today()
        delivery_summary = self._delivery_summary_map(rows)
        for it in rows:
            cid = int(it.get("row") or it.get("entry_start_row") or 0)
            delivery_txt, delivery_days, day_num = delivery_summary.get(cid) or self._contract_delivery_dates(it)
            completion = parse_flexible_date(delivery_txt)
            it["_completion_obj"] = completion
            it["_completion_ord"] = completion.toordinal() if completion else None
            it["_near_delivery_txt"] = delivery_txt
            it["_near_delivery_days"] = delivery_days
            it["_day_num"] = day_num
            tags_list = list(it.get("tags", []) or [])
            it["_tags_str"] = ", ".join(tags_list) if tags_list else ""
            hay = it.get("search") or " ".join(str(it.get(k, "")) for k in ["platform", "no", "user", "status", "completion_date", "content"]).lower()
            it["_search_norm"] = str(hay or "").lower()

    def schedule_apply_contract_filter(self):
        if not hasattr(self, "_filter_apply_timer"):
            self.apply_contract_filter()
            return
        self._filter_apply_timer.start(180)

    def apply_contract_filter(self):
        q = (self.search_input.text() if hasattr(self, "search_input") else "").strip().lower()
        selected_type = str(self.filter_type.currentData() or "").strip() if hasattr(self, "filter_type") else ""
        selected_status = str(self.filter_status.currentData() or "").strip() if hasattr(self, "filter_status") else ""
        sort_mode = str(self.sort_combo.currentData() or "default")
        # Kalan Gun kolonundan gelen siralama sort_combo'yu ezer
        tbl_sort = getattr(getattr(self, 'contract_table', None), '_sort_mode', 'default')
        if tbl_sort != 'default':
            sort_mode = tbl_sort
        rows = []
        date_from = None
        date_to = None
        if hasattr(self, "date_from_filter") and getattr(self, "_date_from_active", False):
            date_from = self.date_from_filter.date().toPython()
        if hasattr(self, "date_to_filter") and getattr(self, "_date_to_active", False):
            date_to = self.date_to_filter.date().toPython()
        days_min = None
        days_max = None
        if hasattr(self, "days_min_filter"):
            txt = self.days_min_filter.text().strip()
            if txt:
                try:
                    days_min = int(txt)
                except ValueError:
                    days_min = None
        if hasattr(self, "days_max_filter"):
            txt = self.days_max_filter.text().strip()
            if txt:
                try:
                    days_max = int(txt)
                except ValueError:
                    days_max = None
        if hasattr(self, "contract_table"):
            self.contract_table._all_rows_for_filter = list(getattr(self, "all_contract_rows", []))
        # Sutun filtreleri
        col_filters = {}
        date_ranges = {}
        day_ranges = {}
        if hasattr(self, '_filter_header'):
            col_filters = dict(self._filter_header._col_filters)
            date_ranges = dict(getattr(self._filter_header, "_date_ranges", {}))
            day_ranges = dict(getattr(self._filter_header, "_day_ranges", {}))
        selected_platforms = set(getattr(self, "selected_platforms", set()))
        for it in getattr(self, "all_contract_rows", []):
            if selected_platforms and str(it.get("platform", "")) not in selected_platforms:
                continue
            hay = str(it.get("_search_norm") or "")
            if q and q not in hay:
                continue
            tval = str(it.get("type_display", it.get("type", "")) or "").strip()
            if selected_type and tval != selected_type:
                continue
            cls, st_label, _days, _dt = self._contract_health(it)
            if selected_status and cls != selected_status:
                continue
            completion = it.get("_completion_obj")
            if date_from and (not completion or completion < date_from):
                continue
            if date_to and (not completion or completion > date_to):
                continue
            day_num = it.get("_day_num")
            if days_min is not None and (day_num is None or day_num < days_min):
                continue
            if days_max is not None and (day_num is None or day_num > days_max):
                continue
            if COL_T_DATE in date_ranges:
                start_date, end_date = date_ranges.get(COL_T_DATE, (None, None))
                if start_date and (not completion or completion < start_date):
                    continue
                if end_date and (not completion or completion > end_date):
                    continue
            if COL_REMAINING in day_ranges:
                min_day, max_day = day_ranges.get(COL_REMAINING, (None, None))
                if min_day is not None and (day_num is None or day_num < min_day):
                    continue
                if max_day is not None and (day_num is None or day_num > max_day):
                    continue
            # Sutun bazli filtreler (0=Platform,1=Tur,2=No,3=User,4=Durum,5=Tarih,6=Gun,7=Etiketler,8=Ozet)
            tags_str = str(it.get("_tags_str") or "")
            col_vals = [
                str(it.get("platform", "") or ""),
                str(it.get("type_display", it.get("type", "")) or ""),
                str(it.get("no", "") or ""),
                str(it.get("user", "") or ""),
                st_label,
                _dt,
                _days,
                tags_str,
                "Özet",
            ]
            skip = False
            for ci, fset in col_filters.items():
                if fset is None:
                    continue

                # Etiketler kolonu için özel kontrol:
                # seçilen etiketlerden en az biri satırdaki etiketlerde varsa geçir.
                if ci == COL_TAGS:
                    row_tags = [str(t or "").strip() for t in list(it.get("tags", []) or []) if str(t or "").strip()]
                    if not any(tag in fset for tag in row_tags):
                        skip = True
                        break
                    continue

                if ci < len(col_vals):
                    if col_vals[ci] not in fset:
                        skip = True
                        break

            if skip:
                continue
            rows.append(it)

        if sort_mode == "no_asc":
            rows.sort(key=lambda x: self._contract_no_sort_key(x.get("no", "")))
        elif sort_mode == "no_desc":
            rows.sort(key=lambda x: self._contract_no_sort_key(x.get("no", "")), reverse=True)
        elif sort_mode == "date_asc":
            rows.sort(key=self._completion_date_sort_key)
        elif sort_mode == "date_desc":
            rows.sort(key=lambda x: (0, -x["_completion_obj"].toordinal()) if x.get("_completion_obj") else (1, 99999999))
        elif sort_mode == "days_asc":
            rows.sort(key=self._days_sort_key)
        elif sort_mode == "days_desc":
            rows.sort(key=self._days_sort_key, reverse=True)
        elif sort_mode == "user_asc":
            rows.sort(key=lambda x: self._norm_tr(str(x.get("user", ""))))
        elif sort_mode == "user_desc":
            rows.sort(key=lambda x: self._norm_tr(str(x.get("user", ""))), reverse=True)

        # Etiket renk haritası (performans için bir kez yükle)
        if self._tag_color_map_cache is None:
            tag_color_map: Dict[str, str] = {}
            if self.store:
                try:
                    for td in self.store.load_tag_defs():
                        tag_color_map[self.store._normalize_label(td.name)] = td.color or "#3B82F6"
                except Exception:
                    pass
            self._tag_color_map_cache = tag_color_map
        _tag_color_map = self._tag_color_map_cache

        self.contract_table.setUpdatesEnabled(False)
        self.contract_table.blockSignals(True)
        try:
            self.contract_table.clearContents()
            self.contract_table.setRowCount(len(rows))
            self.contract_table._visible_rows = rows
            for r,it in enumerate(rows):
                for c in range(self.contract_table.columnCount()):
                    self.contract_table.removeCellWidget(r, c)
                self.contract_table.setRowHeight(r, 36)
                cls, st_label, days_text, tdate = self._contract_health(it)
                payload = {
                    "platform": str(it.get("platform", "") or ""),
                    "contract_no": str(it.get("no", "") or ""),
                    "contract_type": str(it.get("type_display", it.get("type", "")) or ""),
                    "contract_item": it,
                }
                vals=[
                    it.get("platform", ""),
                    it.get("type_display", it.get("type", "")) or "",
                    it.get("no", ""),
                    it.get("user", ""),
                    st_label,
                    tdate,
                    days_text,
                    None,  # col 7: Etiketler widget
                    None,  # col 8: Ozet butonu
                ]
                for c,v in enumerate(vals):
                    if c == COL_TAGS:
                        # Etiketler: dikey sıralı renkli chip'ler
                        tags_list = list(it.get("tags", []) or [])
                        if not tags_list:
                            empty = QTableWidgetItem("")
                            empty.setFlags(empty.flags() & ~Qt.ItemIsEditable)
                            empty.setData(Qt.UserRole, payload)
                            self.contract_table.setItem(r, COL_TAGS, empty)
                            self.contract_table.setRowHeight(r, max(self.contract_table.rowHeight(r), 36))
                            continue
                        wrap = QWidget()
                        wrap.setStyleSheet("QWidget{background:transparent;border:0px;}")
                        wl = QVBoxLayout(wrap)
                        wl.setContentsMargins(5, 4, 5, 4)
                        wl.setSpacing(3)
                        wl.setAlignment(Qt.AlignTop | Qt.AlignLeft)
                        for tag_name in tags_list:
                            color = _tag_color_map.get(self.store._normalize_label(tag_name) if self.store else tag_name, "#3B82F6")
                            base_rgb = _hex_to_rgb(color)
                            bg = _mix_rgb(base_rgb, (255, 255, 255), 0.78)
                            border_c = _mix_rgb(base_rgb, (255, 255, 255), 0.28)
                            txt_c = _rgb_to_hex(_mix_rgb(base_rgb, (15, 23, 42), 0.22))
                            chip = QLabel(tag_name)
                            chip.setStyleSheet(
                                f"QLabel{{background:{_rgb_to_hex(bg)};color:{txt_c};"
                                f"border:1px solid {_rgb_to_hex(border_c)};border-radius:9px;"
                                f"padding:2px 8px;font-size:11px;font-weight:700;}}"
                            )
                            wl.addWidget(chip)
                        placeholder = QTableWidgetItem("")
                        placeholder.setData(Qt.UserRole, payload)
                        self.contract_table.setItem(r, COL_TAGS, placeholder)
                        self.contract_table.setCellWidget(r, COL_TAGS, wrap)
                        # Satır yüksekliğini etiket sayısına göre ayarla
                        CHIP_H, CHIP_SP, PAD = 22, 3, 8
                        n = len(tags_list)
                        row_h = max(36, n * CHIP_H + max(0, n - 1) * CHIP_SP + PAD) if n > 0 else 36
                        self.contract_table.setRowHeight(r, max(self.contract_table.rowHeight(r), row_h))
                        continue
                    if c == COL_SUMMARY:
                        lbl = QLabel("\U0001F50D")
                        lbl.setAlignment(Qt.AlignCenter)
                        lbl.setToolTip("Bileşen özetini gör")
                        lbl.setStyleSheet(
                            "QLabel{font-size:16px;color:#185FA5;background:transparent;border:0px;}"
                            "QLabel:hover{color:#042C53;}"
                        )
                        wrap = QWidget()
                        wrap.setStyleSheet("QWidget{background:transparent;border:0px;}")
                        wl = QHBoxLayout(wrap)
                        wl.setContentsMargins(0,0,0,0)
                        wl.setSpacing(0)
                        wl.setAlignment(Qt.AlignCenter)
                        wl.addWidget(lbl)
                        placeholder = QTableWidgetItem("")
                        placeholder.setData(Qt.UserRole, payload)
                        self.contract_table.setItem(r, COL_SUMMARY, placeholder)
                        self.contract_table.setCellWidget(r, COL_SUMMARY, wrap)
                        continue
                    cell = QTableWidgetItem(str(v or ""))
                    cell.setFlags(cell.flags() & ~Qt.ItemIsEditable)
                    cell.setData(Qt.UserRole, payload)
                    if c == COL_STATUS:
                        if cls in {"geciken", "gecikmeli_teslim"}:
                            cell.setForeground(QColor("#dc2626"))
                        elif cls == "kritik":
                            cell.setForeground(QColor("#b45309"))
                        elif cls == "tamamlandi":
                            cell.setForeground(QColor("#047857"))
                        else:
                            cell.setForeground(QColor("#1f5be3"))
                    if c == COL_REMAINING:
                        if cls in {"geciken", "gecikmeli_teslim"}:
                            cell.setForeground(QColor("#dc2626"))
                        elif "erken teslim edildi" in str(v):
                            cell.setForeground(QColor("#047857"))
                        elif str(v) in {"Termin gününde teslim edildi", "Teslim tarihi yok", "—"}:
                            cell.setForeground(QColor("#64748b"))
                        elif str(v).endswith("gün"):
                            days_num = as_number(str(v).replace(" gün", ""))
                            if days_num <= 60:
                                cell.setForeground(QColor("#b45309"))
                            else:
                                cell.setForeground(QColor("#1f5be3"))
                        else:
                            cell.setForeground(QColor("#047857"))
                    self.contract_table.setItem(r,c,cell)
        finally:
            self.contract_table.blockSignals(False)
            self.contract_table.setUpdatesEnabled(True)
        self.position_query_logo_background()

    def open_contract_item(self, item: dict):
        if not self.store:
            if getattr(self, "_store_loading", False):
                QMessageBox.information(
                    self,
                    "Excel hazırlanıyor",
                    "Liste hazır. Sözleşme detayı için Excel düzenleme modu birkaç saniye içinde hazır olacak.",
                )
            return
        platform = item.get("platform")
        no = item.get("no")
        start_row = item.get("row")
        self.set_busy_overlay(True, "Sözleşme detayı yükleniyor...")
        try:
            ci, systems, deliveries = self.store.load_contract_structure(platform, no, start_row=start_row)
        finally:
            self.set_busy_overlay(False)
        if not ci:
            QMessageBox.warning(self, "Bulunamadı", "Sözleşme detayları okunamadı.")
            return
        work = ContractWorkWindow(self.store, ci, self, systems=systems, deliveries=deliveries)
        if work.exec():
            deleted_info = getattr(work, "deleted_contract_info", None)
            if deleted_info:
                # Silme sonrası tam excel indeksini yeniden kurmak (binlerce satırda)
                # gereksiz uzun sürüyor. Mevcut indeksi yerinde güncelleriz.
                self.set_busy_overlay(True, "Liste güncelleniyor...", 82)
                try:
                    self._apply_deleted_contract_to_index(deleted_info)
                    self.set_busy_overlay(True, "Arayüz yenileniyor...", 96)
                    platforms = self.store.platform_names()
                    self._set_platform_items(platforms)
                    self.update_alert_strip()
                    if str(platform or "") in platforms:
                        self.select_platform(platform)
                    elif self.platform_list.count():
                        self._apply_platform_selection()
                    else:
                        self.set_empty_state()
                finally:
                    self.set_busy_overlay(False)
            else:
                old_platform = str(platform or "")
                work_ci = getattr(work, "ci", None)
                new_platform = str(getattr(work_ci, "platform", "") or old_platform)
                if old_platform and new_platform and old_platform != new_platform:
                    self.request_refresh(select_platform=old_platform, scope="platform", platform=old_platform)
                    self.request_refresh(select_platform=new_platform, scope="platform", platform=new_platform)
                    self.select_platform(new_platform)
                else:
                    target = new_platform or old_platform
                    self.request_refresh(select_platform=target, scope="platform", platform=target)

    def open_selected_contract(self, row, col):
        rows = getattr(self.contract_table, "_visible_rows", [])
        if row < 0 or row >= self.contract_table.rowCount():
            return
        if col == COL_SUMMARY:
            if row < len(rows):
                self.show_contract_summary(row, rows[row])
            return
        payload = None
        for column in range(self.contract_table.columnCount()):
            cell = self.contract_table.item(row, column)
            candidate = cell.data(Qt.UserRole) if cell else None
            if isinstance(candidate, dict) and candidate.get("contract_no"):
                payload = candidate
                break
        if payload and isinstance(payload.get("contract_item"), dict):
            self.open_contract_item(payload["contract_item"])
            return
        if row < len(rows):
            self.open_contract_item(rows[row])
            return
        platform_item = self.contract_table.item(row, COL_PLATFORM)
        type_item = self.contract_table.item(row, COL_TYPE)
        no_item = self.contract_table.item(row, COL_CONTRACT_NO)
        self.open_contract_item({
            "platform": platform_item.text() if platform_item else "",
            "type": type_item.text() if type_item else "",
            "no": no_item.text() if no_item else "",
        })

    def _apply_deleted_contract_to_index(self, deleted_info: dict):
        p = str((deleted_info or {}).get("platform") or "")
        no = str((deleted_info or {}).get("contract_no") or "").strip()
        start = int((deleted_info or {}).get("start_row") or 0)
        end = int((deleted_info or {}).get("end_row") or 0)
        deleted_rows = int((deleted_info or {}).get("deleted_rows") or max(0, end - start + 1))
        if not p or start <= 0 or deleted_rows <= 0:
            return

        updated = []
        removed = False
        for it in getattr(self, "contract_index", []):
            ip = str(it.get("platform", "") or "")
            irow = int(it.get("row", 0) or 0)
            ino = str(it.get("no", "") or "").strip()

            if ip == p and irow == start and ino == no and not removed:
                removed = True
                continue

            rec = dict(it)
            if ip == p and irow > end:
                rec["row"] = irow - deleted_rows
            updated.append(rec)

        # Güvenli fallback: row eşleşmesi bozulduysa ilk aynı no'yu düş.
        if not removed:
            for idx, it in enumerate(updated):
                if str(it.get("platform", "") or "") == p and str(it.get("no", "") or "").strip() == no:
                    del updated[idx]
                    removed = True
                    break

        self.contract_index = updated


def open_share_contract_window(path: Path | str) -> Optional[ContractWorkWindow]:
    """Open a share-mode STS directly in ContractWorkWindow without main list/login."""
    meta = _share_metadata_from_path(path)
    if not meta:
        return None
    store = STSStore(Path(path), actor="Paylaşım")
    rows = store.build_contract_index()
    if not rows:
        raise ValueError("Paylaşım dosyasında sözleşme bulunamadı.")
    source_no = str(meta.get("source_contract_no") or "").strip()
    selected = None
    if source_no:
        selected = next((r for r in rows if str(r.get("no") or "").strip() == source_no), None)
    if selected is None:
        contract_id = int(meta.get("contract_id") or 0)
        selected = next((r for r in rows if int(r.get("row") or 0) == contract_id), None) if contract_id else None
    selected = selected or rows[0]
    contract_id = int(selected.get("row") or meta.get("contract_id") or 0)
    if contract_id <= 0:
        raise ValueError("Paylaşım sözleşmesi bulunamadı.")
    row = store.db.conn.execute("SELECT c.contract_no,c.contract_type,p.name AS platform,c.platform_id FROM contracts c JOIN platforms p ON p.id=c.platform_id WHERE c.id=?", (contract_id,)).fetchone()
    if not row:
        raise ValueError("Paylaşım sözleşmesi bulunamadı.")
    ci, systems, deliveries = store.load_contract_structure(
        str(row["platform"] or ""),
        contract_no=str(row["contract_no"] or ""),
        start_row=contract_id,
        contract_type=str(row["contract_type"] or ""),
        platform_id=int(row["platform_id"] or 0),
    )
    auth.current_staff = {
        "id": 0,
        "full_name": "Paylaşım Kullanıcısı",
        "username": "share",
        "is_active": 1,
        "is_admin": True,
        "permissions": {"view_contracts", "edit_contracts", "export_data", "manage_labels"},
    }
    win = ContractWorkWindow(store, ci, systems=systems, deliveries=deliveries)
    win.set_share_mode(str(meta.get("permission_mode") or "view"))
    return win


if __name__ == "__main__":
    sys.excepthook = _global_exc_handler
    configure_windows_app_identity()
    app = QApplication(sys.argv)

    # ── Global yakalanmamış exception handler ────────────────────────────────
    # QApplication oluşturulduktan SONRA kurulur; böylece handler içinde
    # QApplication.instance() kontrolü güvenle yapılabilir.
    # NOT: sys.excepthook yalnızca ana thread ve threading.Thread için çalışır;
    # QThread içindeki hatalar worker'ların kendi except bloklarında yakalanır.
    def _global_exc_handler(exc_type, exc_val, exc_tb):
        if issubclass(exc_type, (KeyboardInterrupt, SystemExit)):
            sys.__excepthook__(exc_type, exc_val, exc_tb)
            return
        _log.critical("Yakalanmamış hata", exc_info=(exc_type, exc_val, exc_tb))
        # QApplication yoksa veya kapanıyorsa sadece logla, GUI gösterme
        q_app = QApplication.instance()
        if q_app is None:
            return
        try:
            msg = f"Beklenmeyen bir hata oluştu.\n\n{exc_val}"
            QMessageBox.critical(None, "Kritik Hata", msg)
        except Exception:
            pass

    sys.excepthook = _global_exc_handler
    app.setApplicationName("STS")
    app.setApplicationDisplayName("STS")
    app.setDesktopFileName(APP_ID)
    app.setFont(QFont("Segoe UI", 10))
    icon_path = app_icon_path()
    if icon_path.exists():
        app.setWindowIcon(QIcon(str(icon_path)))

    # Startup uses modal dialogs before the main window exists.  Keep Qt from
    # treating an accepted file/staff dialog as the last-window-close event;
    # otherwise the first successful registration/login can request app quit
    # before the real MainWindow event loop starts.
    app.setQuitOnLastWindowClosed(False)

    cli_path = Path(sys.argv[1]).expanduser() if len(sys.argv) > 1 and str(sys.argv[1]).strip() else None
    if cli_path and _share_metadata_from_path(cli_path):
        try:
            win = open_share_contract_window(cli_path)
            if win is None:
                raise ValueError("Paylaşım metadata bulunamadı.")
            app.setQuitOnLastWindowClosed(True)
            win.show()
            sys.exit(app.exec())
        except Exception as exc:
            _log.exception("Paylaşım STS açılış hatası")
            QMessageBox.critical(None, "Paylaşım açılamadı", f"Paylaşım dosyası açılamadı.\n\n{exc}")
            sys.exit(1)

    start_dialog = WorkbookStartDialog()
    if not start_dialog.exec() or not start_dialog.selected_path:
        sys.exit(0)

    selected_path = Path(start_dialog.selected_path)
    if _share_metadata_from_path(selected_path):
        try:
            win = open_share_contract_window(selected_path)
            if win is None:
                raise ValueError("Paylaşım metadata bulunamadı.")
            app.setQuitOnLastWindowClosed(True)
            win.show()
            sys.exit(app.exec())
        except Exception as exc:
            _log.exception("Paylaşım STS açılış hatası")
            QMessageBox.critical(None, "Paylaşım açılamadı", f"Paylaşım dosyası açılamadı.\n\n{exc}")
            sys.exit(1)

    staff = None
    if selected_path.suffix.lower() == ".sts":
        if not auth.ensure_system_admin_setup(selected_path):
            sys.exit(0)
        staff = auth.require_staff_login(selected_path)
        if not staff:
            sys.exit(0)

    win = MainWindow(initial_path=selected_path, current_staff=staff)
    win.show()

    def _start_initial_load():
        app.setQuitOnLastWindowClosed(True)
        try:
            if selected_path.suffix.lower() == ".sts":
                win.start_sts_load(selected_path)
            else:
                win.start_excel_load(selected_path)
        except Exception as exc:
            _log.exception("Başlangıç yükleme hatası")
            traceback.print_exc()
            QMessageBox.critical(win, "Açılış hatası", f"Uygulama başlatılırken hata oluştu.\n\n{exc}")
            app.quit()

    QTimer.singleShot(0, _start_initial_load)
    sys.exit(app.exec())
